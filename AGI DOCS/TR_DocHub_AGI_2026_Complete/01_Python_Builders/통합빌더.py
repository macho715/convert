#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TR_DocHub_AGI_2026.xlsm Builder
통합 빌더: 정규화 모델 + Dashboard + Calendar + VBA_Pasteboard

Usage:
    python 통합빌더.py --output TR_DocHub_AGI_2026_Template.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Dict, List, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference


# =========================
# Helpers
# =========================
def col_letter(n: int) -> str:
    return get_column_letter(n)


def set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[col_letter(col_idx)].width = w


def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{col_letter(start_col)}{start_row}:{col_letter(end_col)}{end_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def get_styles() -> Dict[str, Any]:
    """스타일 딕셔너리 반환"""
    return {
        'font_title': Font(bold=True, size=14, color="1E3A5F"),
        'font_header': Font(bold=True, size=11, color="FFFFFF"),
        'font_bold': Font(bold=True, size=10),
        'font_normal': Font(size=10),
        'header_fill': PatternFill("solid", fgColor="4472C4"),
        'subheader_fill': PatternFill("solid", fgColor="70AD47"),
        'gray_fill': PatternFill("solid", fgColor="D9D9D9"),
        'align_center': Alignment(horizontal="center", vertical="center"),
        'align_left': Alignment(horizontal="left", vertical="center"),
        'border_thin': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
    }


# =========================
# Sheet Builders
# =========================
def create_lists_sheet(wb, styles: Dict) -> None:
    """Lists sheet (hidden) - 드롭다운 소스"""
    ws_lists = wb.create_sheet("Lists")
    
    status_list = ["Not Started", "In Progress", "Submitted", "Accepted", "Rejected", "On Hold", "Waived"]
    priority_list = ["Mandatory", "Important", "Optional"]
    due_basis_list = ["AUTO", "Doc Deadline", "Land Permit By", "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival"]
    yn_list = ["Y", "N"]
    site_list = ["Site-1", "Site-2", "Site-3", "Site-4", "Site-5", "Site-6"]
    
    ws_lists["A1"] = "Status"
    ws_lists["B1"] = "Priority"
    ws_lists["C1"] = "Due_Basis"
    ws_lists["D1"] = "Y/N"
    ws_lists["E1"] = "Sites"
    
    for i, v in enumerate(status_list, start=2):
        ws_lists[f"A{i}"] = v
    for i, v in enumerate(priority_list, start=2):
        ws_lists[f"B{i}"] = v
    for i, v in enumerate(due_basis_list, start=2):
        ws_lists[f"C{i}"] = v
    for i, v in enumerate(yn_list, start=2):
        ws_lists[f"D{i}"] = v
    for i, v in enumerate(site_list, start=2):
        ws_lists[f"E{i}"] = v
    
    for c in range(1, 6):
        cell = ws_lists.cell(row=1, column=c)
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
    
    set_col_widths(ws_lists, {1: 18, 2: 12, 3: 18, 4: 6, 5: 12})
    ws_lists.sheet_state = "hidden"


def create_c_config_sheet(wb, styles: Dict) -> None:
    """C_Config sheet - 임계값/경로/공휴일"""
    ws_cfg = wb.create_sheet("C_Config")
    ws_cfg["A1"] = "TR Document Tracker - Config"
    ws_cfg["A1"].font = styles['font_title']
    
    default_year = dt.date.today().year
    cfg_rows = [
        ("DocCutoff_Days", 3, "Doc Deadline = MZP Departure - DocCutoff_Days"),
        ("LandPermitLead_Days", 7, "Land Permit By = AGI Arrival - LandPermitLead_Days"),
        ("DueSoon_Threshold_Days", 7, "Dashboard/CF: upcoming items within N days"),
        ("Amber_Threshold_Days", 7, "Dashboard KPI: D-7 Amber warning threshold"),
        ("Red_Threshold_Days", 3, "Dashboard KPI: D-3 Red warning threshold"),
        ("Critical_Threshold_Days", 1, "Dashboard KPI: D-1 Critical alert threshold"),
        ("Default_Year", default_year, "Year used for sample schedule MM-DD parsing"),
        ("ExportBasePath", "", "Base path for Export Pack (leave empty for workbook folder)"),
    ]
    
    ws_cfg["A3"] = "Key"
    ws_cfg["B3"] = "Value"
    ws_cfg["C3"] = "Description"
    for cell in (ws_cfg["A3"], ws_cfg["B3"], ws_cfg["C3"]):
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.alignment = styles['align_left']
        cell.border = styles['border_thin']
    
    for i, (k, v, dsc) in enumerate(cfg_rows, start=4):
        ws_cfg[f"A{i}"] = k
        ws_cfg[f"B{i}"] = v
        ws_cfg[f"C{i}"] = dsc
        for c in range(1, 4):
            cell = ws_cfg.cell(row=i, column=c)
            cell.border = styles['border_thin']
            cell.alignment = styles['align_left']
    
    set_col_widths(ws_cfg, {1: 26, 2: 18, 3: 60})


def create_s_voyages_sheet(wb, styles: Dict) -> None:
    """S_Voyages sheet - 항차 일정 (정규화: VoyageID 추가)"""
    ws_voy = wb.create_sheet("S_Voyages")
    ws_voy["A1"] = "Voyage Schedule (Source)"
    ws_voy["A1"].font = styles['font_title']
    ws_voy["A2"] = f"Last Updated: {now_str()}"
    ws_voy["A2"].font = Font(italic=True, size=10)
    
    headers_voy = [
        "VoyageID", "VoyageName", "TR Units", "Site",
        "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival",
        "Doc Deadline", "Land Permit By",
        "Project", "Lot", "Remarks"
    ]
    header_row = 4
    
    for col, h in enumerate(headers_voy, start=1):
        cell = ws_voy.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # AGI TR 1-6 샘플 데이터
    year = 2026
    sample_rows = [
        ("V01", "Voyage 1", "TR 1-2", "Site-1", dt.date(year, 1, 27), dt.date(year, 1, 29), dt.date(year, 2, 1), dt.date(year, 2, 2), None, None, "HVDC", "Lot-1", ""),
        ("V02", "Voyage 2", "TR 3-4", "Site-2", dt.date(year, 2, 6), dt.date(year, 2, 7), dt.date(year, 2, 10), dt.date(year, 2, 11), None, None, "HVDC", "Lot-1", ""),
        ("V03", "Voyage 3", "TR 5-6", "Site-3", dt.date(year, 2, 15), dt.date(year, 2, 16), dt.date(year, 2, 19), dt.date(year, 2, 20), None, None, "HVDC", "Lot-1", ""),
        ("V04", "Voyage 4", "TR 7", "Site-4", dt.date(year, 2, 24), dt.date(year, 2, 25), dt.date(year, 2, 27), dt.date(year, 2, 28), None, None, "HVDC", "Lot-1", ""),
    ]
    
    start_row = header_row + 1
    for i, row_data in enumerate(sample_rows):
        r = start_row + i
        for c, val in enumerate(row_data, start=1):
            ws_voy.cell(row=r, column=c, value=val)
            cell = ws_voy.cell(row=r, column=c)
            cell.font = styles['font_normal']
            cell.border = styles['border_thin']
            cell.alignment = styles['align_center'] if c <= 10 else styles['align_left']
            if c in (5, 6, 7, 8, 9, 10) and val:
                cell.number_format = "yyyy-mm-dd"
    
    # Formulas for Doc Deadline and Land Permit By
    for r in range(start_row, start_row + len(sample_rows)):
        ws_voy.cell(row=r, column=9, value=f"=IF(G{r}=\"\",\"\",G{r}-C_Config!$B$4)")  # Doc Deadline
        ws_voy.cell(row=r, column=10, value=f"=IF(H{r}=\"\",\"\",H{r}-C_Config!$B$5)")  # Land Permit By
        ws_voy.cell(row=r, column=9).number_format = "yyyy-mm-dd"
        ws_voy.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    
    ws_voy.freeze_panes = "A5"
    set_col_widths(ws_voy, {
        1: 12, 2: 16, 3: 12, 4: 10,
        5: 12, 6: 12, 7: 14, 8: 12,
        9: 12, 10: 14,
        11: 12, 12: 12, 13: 28
    })
    add_table(ws_voy, "tbl_Voyage", header_row, 1, start_row + len(sample_rows) - 1, len(headers_voy))
    
    # Data validation
    dv_site = DataValidation(type="list", formula1="=Lists!$E$2:$E$7", allow_blank=True)
    ws_voy.add_data_validation(dv_site)
    dv_site.add(f"D{start_row}:D5000")


def create_m_parties_sheet(wb, styles: Dict) -> None:
    """M_Parties sheet - Responsible Party 마스터"""
    ws_party = wb.create_sheet("M_Parties")
    ws_party["A1"] = "Responsible Party Master"
    ws_party["A1"].font = styles['font_title']
    
    headers_party = ["PartyID", "PartyName", "OwnerEmail", "Contact", "ActiveFlag"]
    header_row = 3
    
    for col, h in enumerate(headers_party, start=1):
        cell = ws_party.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # AGI 프로젝트 파티
    parties_data = [
        ("FF", "Freight Forwarder", "[MASK]@company.com", "", "Y"),
        ("CUSTBROKER", "Customs Broker", "[MASK]@company.com", "", "Y"),
        ("EPC", "EPC Contractor", "[MASK]@company.com", "", "Y"),
        ("TRCON", "Transport Contractor", "[MASK]@company.com", "", "Y"),
        ("PORT", "Port Authority", "[MASK]@authority.ae", "", "Y"),
        ("OFCO", "OFCO Agency", "[MASK]@ofco-int.com", "", "Y"),
        ("MMT", "Mammoet", "[MASK]@mammoet.com", "", "Y"),
        ("SCT", "Samsung C&T", "[MASK]@samsung.com", "", "Y"),
    ]
    
    start_row = header_row + 1
    for i, (pid, name, email, contact, active) in enumerate(parties_data):
        r = start_row + i
        ws_party.cell(r, 1).value = pid
        ws_party.cell(r, 2).value = name
        ws_party.cell(r, 3).value = email
        ws_party.cell(r, 4).value = contact
        ws_party.cell(r, 5).value = active
        
        for c in range(1, 6):
            cell = ws_party.cell(r, c)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            cell.alignment = styles['align_left'] if c in (2, 3, 4) else styles['align_center']
    
    add_table(ws_party, "tbl_Party", header_row, 1, start_row + len(parties_data) - 1, len(headers_party))
    
    # Data validation
    dv_active = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=False)
    ws_party.add_data_validation(dv_active)
    dv_active.add(f"E{start_row}:E5000")
    
    set_col_widths(ws_party, {1: 14, 2: 24, 3: 30, 4: 20, 5: 12})


def create_m_doccatalog_sheet(wb, styles: Dict) -> None:
    """M_DocCatalog sheet - 문서 카탈로그 (정규화)"""
    ws_doc = wb.create_sheet("M_DocCatalog")
    ws_doc["A1"] = "Document Catalog (Master)"
    ws_doc["A1"].font = styles['font_title']
    ws_doc["A2"] = "Edit document requirements here. Tracker refresh will follow this."
    ws_doc["A2"].font = Font(italic=True, size=10)
    
    headers_doc = [
        "DocCode", "DocName", "DocCategory",
        "DefaultResponsiblePartyID", "RequiredFlag", "EvidenceRequiredFlag", "ActiveFlag",
        "DocDescription"
    ]
    doc_header_row = 4
    
    for col, h in enumerate(headers_doc, start=1):
        cell = ws_doc.cell(row=doc_header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # AGI 프로젝트 기본 문서 목록
    default_docs = [
        ("GATEPASS", "Gate Pass Application", "Gate", "FF", "Y", "Y", "Y", "Port gate pass for MZP entry"),
        ("CUSTOMS", "Customs Declaration", "Customs", "CUSTBROKER", "Y", "Y", "Y", "Customs clearance documentation"),
        ("PERMIT", "Land Permit Application", "Permit", "EPC", "Y", "Y", "Y", "Road/land permit for SPMT operations"),
        ("BL", "Bill of Lading", "Transport", "FF", "Y", "Y", "Y", "BL draft and final"),
        ("STOWAGE", "Stowage Plan", "Transport", "TRCON", "Y", "Y", "Y", "Cargo stowage plan"),
        ("LASHING", "Lashing Plan", "Transport", "TRCON", "Y", "Y", "Y", "Cargo lashing plan"),
        ("MWS", "Marine Warranty Survey", "Transport", "MMT", "Y", "Y", "Y", "MWS certificate"),
        ("NOC", "AD Maritime NOC", "Permit", "OFCO", "Y", "Y", "Y", "AD Maritime No Objection Certificate"),
    ]
    
    doc_start_row = doc_header_row + 1
    for i, (code, name, cat, party, req, evid, active, desc) in enumerate(default_docs):
        r = doc_start_row + i
        ws_doc.cell(r, 1).value = code
        ws_doc.cell(r, 2).value = name
        ws_doc.cell(r, 3).value = cat
        ws_doc.cell(r, 4).value = party
        ws_doc.cell(r, 5).value = req
        ws_doc.cell(r, 6).value = evid
        ws_doc.cell(r, 7).value = active
        ws_doc.cell(r, 8).value = desc
        
        for c in range(1, 9):
            cell = ws_doc.cell(r, c)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            cell.alignment = styles['align_left'] if c in (2, 3, 8) else styles['align_center']
    
    add_table(ws_doc, "tbl_DocCatalog", doc_header_row, 1, doc_start_row + len(default_docs) - 1, len(headers_doc))
    
    # Data validation
    dv_party = DataValidation(type="list", formula1="=M_Parties[PartyID]", allow_blank=False)
    dv_yn = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=False)
    ws_doc.add_data_validation(dv_party)
    ws_doc.add_data_validation(dv_yn)
    dv_party.add(f"D{doc_start_row}:D5000")
    dv_yn.add(f"E{doc_start_row}:G5000")
    
    set_col_widths(ws_doc, {
        1: 14, 2: 28, 3: 18,
        4: 24, 5: 14, 6: 18, 7: 12,
        8: 40
    })
    ws_doc.freeze_panes = "A5"


def create_r_deadline_rules_sheet(wb, styles: Dict) -> None:
    """R_DeadlineRules sheet - DueDate 룰테이블"""
    ws_rules = wb.create_sheet("R_DeadlineRules")
    ws_rules["A1"] = "Deadline Rules (DocCode → AnchorField + OffsetDays)"
    ws_rules["A1"].font = styles['font_title']
    ws_rules["A2"] = "Priority: Lower number = Higher priority. Multiple rules for same DocCode: lowest Priority wins."
    ws_rules["A2"].font = Font(italic=True, size=10)
    
    headers_rules = [
        "RuleID", "DocCode", "AnchorField", "OffsetDays",
        "CalendarType", "Priority", "ActiveFlag", "AppliesIf"
    ]
    header_row = 4
    
    for col, h in enumerate(headers_rules, start=1):
        cell = ws_rules.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # AGI 프로젝트 룰 예시
    rules_data = [
        ("R001", "GATEPASS", "Load-out", -1, "CAL", 1, "Y", ""),
        ("R002", "CUSTOMS", "Doc Deadline", -2, "WD", 1, "Y", ""),
        ("R003", "PERMIT", "MZP Arrival", 0, "CAL", 1, "Y", ""),
        ("R004", "BL", "MZP Departure", -3, "WD", 1, "Y", ""),
        ("R005", "STOWAGE", "Load-out", -2, "CAL", 1, "Y", ""),
        ("R006", "LASHING", "Load-out", -2, "CAL", 1, "Y", ""),
        ("R007", "MWS", "MZP Departure", -5, "WD", 1, "Y", ""),
        ("R008", "NOC", "AGI Arrival", -7, "WD", 1, "Y", ""),
    ]
    
    start_row = header_row + 1
    for i, (rid, doc, anchor, offset, cal, prio, active, cond) in enumerate(rules_data):
        r = start_row + i
        ws_rules.cell(r, 1).value = rid
        ws_rules.cell(r, 2).value = doc
        ws_rules.cell(r, 3).value = anchor
        ws_rules.cell(r, 4).value = offset
        ws_rules.cell(r, 5).value = cal
        ws_rules.cell(r, 6).value = prio
        ws_rules.cell(r, 7).value = active
        ws_rules.cell(r, 8).value = cond
        
        for c in range(1, 9):
            cell = ws_rules.cell(r, c)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            cell.alignment = styles['align_center']
    
    add_table(ws_rules, "tbl_RuleDeadline", header_row, 1, start_row + len(rules_data) - 1, len(headers_rules))
    
    # Data validation
    dv_doccode = DataValidation(type="list", formula1="=M_DocCatalog[DocCode]", allow_blank=False)
    dv_anchor = DataValidation(type="list", formula1='"MZP Arrival,Load-out,MZP Departure,AGI Arrival,Doc Deadline,Land Permit By"', allow_blank=False)
    dv_cal = DataValidation(type="list", formula1='"CAL,WD"', allow_blank=False)
    dv_yn = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=False)
    
    ws_rules.add_data_validation(dv_doccode)
    ws_rules.add_data_validation(dv_anchor)
    ws_rules.add_data_validation(dv_cal)
    ws_rules.add_data_validation(dv_yn)
    
    dv_doccode.add(f"B{start_row}:B5000")
    dv_anchor.add(f"C{start_row}:C5000")
    dv_cal.add(f"E{start_row}:E5000")
    dv_yn.add(f"G{start_row}:G5000")
    
    set_col_widths(ws_rules, {
        1: 10, 2: 14, 3: 18, 4: 12,
        5: 14, 6: 10, 7: 12, 8: 20
    })
    ws_rules.freeze_panes = "A5"


def create_t_tracker_sheet(wb, styles: Dict) -> None:
    """T_Tracker sheet - 제출 통제 원장 (정규화)"""
    ws_tr = wb.create_sheet("T_Tracker")
    ws_tr["A1"] = "TR Document Tracker (Transaction - Main)"
    ws_tr["A1"].font = styles['font_title']
    ws_tr["A2"] = f"Last Refreshed: {now_str()}  |  목적: Responsible Party별 서류 제출일자/데드라인/현황 공유 → 누락 방지"
    ws_tr["A2"].font = Font(italic=True, size=10)
    
    tracker_headers = [
        "VoyageID", "DocCode",  # Composite PK
        "ResponsiblePartyID", "AnchorField", "AnchorDate", "OffsetDays", "DueDate",
        "Status", "SubmittedDate", "AcceptedDate",
        "EvidenceLink", "EvidenceNote",
        "LastUpdatedBy", "LastUpdatedAt", "RAG"
    ]
    
    tr_header_row = 4
    for col, h in enumerate(tracker_headers, start=1):
        cell = ws_tr.cell(row=tr_header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # 초기 행은 GenerateTrackerRows VBA로 생성 (여기서는 헤더만)
    tr_start_row = tr_header_row + 1
    
    # 샘플 행 1개 (참고용)
    ws_tr.cell(tr_start_row, 1).value = "V01"
    ws_tr.cell(tr_start_row, 2).value = "GATEPASS"
    ws_tr.cell(tr_start_row, 3).value = "FF"
    ws_tr.cell(tr_start_row, 4).value = "Load-out"
    # AnchorDate 수식 (개념):
    ws_tr.cell(tr_start_row, 5).value = f"=IFERROR(INDEX(tbl_Voyage, MATCH(A{tr_start_row}, tbl_Voyage[VoyageID], 0), MATCH(D{tr_start_row}, tbl_Voyage[#Headers], 0)), \"\")"
    ws_tr.cell(tr_start_row, 6).value = -1
    # DueDate 수식 (B안 - VBA로 채움, 여기서는 수식 예시):
    ws_tr.cell(tr_start_row, 7).value = f"=IF(E{tr_start_row}=\"\",\"\",IF(INDEX(tbl_RuleDeadline,MATCH(B{tr_start_row},tbl_RuleDeadline[DocCode],0),5)=\"WD\",WORKDAY.INTL(E{tr_start_row},F{tr_start_row},\"0000011\",Holidays!$A$2:$A$100),E{tr_start_row}+F{tr_start_row}))"
    ws_tr.cell(tr_start_row, 8).value = "Not Started"
    # RAG 수식:
    ws_tr.cell(tr_start_row, 15).value = f"=IF(G{tr_start_row}=\"\",\"\",IF(G{tr_start_row}<TODAY(),\"Overdue\",IF(G{tr_start_row}<=TODAY()+7,\"DueSoon\",\"OK\")))"
    
    for c in range(1, 16):
        cell = ws_tr.cell(tr_start_row, c)
        cell.border = styles['border_thin']
        cell.font = styles['font_normal']
        cell.alignment = styles['align_left'] if c in (11, 12) else styles['align_center']
    
    for c in (5, 7, 9, 10, 14):
        ws_tr.cell(tr_start_row, c).number_format = "yyyy-mm-dd"
    
    ws_tr.freeze_panes = "A5"
    set_col_widths(ws_tr, {
        1: 12, 2: 14, 3: 20, 4: 18, 5: 12, 6: 12, 7: 12,
        8: 14, 9: 14, 10: 14,
        11: 30, 12: 30,
        13: 14, 14: 18, 15: 12
    })
    
    # Table will be created after rows are generated
    add_table(ws_tr, "tbl_Tracker", tr_header_row, 1, tr_start_row, len(tracker_headers))
    
    # Data validation
    dv_voyage = DataValidation(type="list", formula1="=S_Voyages[VoyageID]", allow_blank=False)
    dv_doccode = DataValidation(type="list", formula1="=M_DocCatalog[DocCode]", allow_blank=False)
    dv_party = DataValidation(type="list", formula1="=M_Parties[PartyID]", allow_blank=False)
    dv_status = DataValidation(type="list", formula1="=Lists!$A$2:$A$8", allow_blank=True)
    
    ws_tr.add_data_validation(dv_voyage)
    ws_tr.add_data_validation(dv_doccode)
    ws_tr.add_data_validation(dv_party)
    ws_tr.add_data_validation(dv_status)
    
    dv_voyage.add(f"A{tr_start_row}:A5000")
    dv_doccode.add(f"B{tr_start_row}:B5000")
    dv_party.add(f"C{tr_start_row}:C5000")
    dv_status.add(f"H{tr_start_row}:H5000")
    
    # Conditional formatting
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:O5000",
        FormulaRule(formula=[f"=$O{tr_start_row}=\"Overdue\""], fill=red_fill)
    )
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:O5000",
        FormulaRule(formula=[f"=$O{tr_start_row}=\"DueSoon\""], fill=yellow_fill)
    )
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:O5000",
        FormulaRule(formula=[f"=$O{tr_start_row}=\"OK\""], fill=green_fill)
    )


def create_d_dashboard_sheet(wb, styles: Dict) -> None:
    """D_Dashboard sheet - Responsible Party 중심 KPI"""
    ws_dash = wb.create_sheet("D_Dashboard", 0)
    ws_dash["A1"] = "TR Document Preparation / Submission Dashboard"
    ws_dash["A1"].font = Font(bold=True, size=18, color="1E3A5F")
    ws_dash["A3"] = f"Last Updated: {now_str()}"
    ws_dash["A3"].font = Font(italic=True, size=10)
    
    # Party 선택
    ws_dash["A5"] = "Responsible Party:"
    ws_dash["A5"].font = styles['font_bold']
    ws_dash["B5"] = "ALL"
    dv_party = DataValidation(type="list", formula1="=M_Parties[PartyName]", allow_blank=True)
    ws_dash.add_data_validation(dv_party)
    dv_party.add("B5")
    
    # KPI 섹션
    ws_dash["A7"] = "KPI"
    ws_dash["A7"].font = styles['font_bold']
    ws_dash["A8"] = "Overdue Count"
    ws_dash["A9"] = "Due in 7 days Count"
    ws_dash["A10"] = "Due in 14 days Count"
    ws_dash["A11"] = "Submitted (Pending) Count"
    ws_dash["A12"] = "Rejected Count"
    ws_dash["A13"] = "Completion %"
    
    # KPI 수식 (Party 필터 적용)
    ws_dash["B8"] = "=SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[RAG]=\"Overdue\"))"
    ws_dash["B9"] = "=SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[RAG]=\"DueSoon\")*(tbl_Tracker[DueDate]<=TODAY()+7))"
    ws_dash["B10"] = "=SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[DueDate]<=TODAY()+14)*(tbl_Tracker[DueDate]>TODAY()+7))"
    ws_dash["B11"] = "=SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[Status]=\"Submitted\")*(tbl_Tracker[AcceptedDate]=\"\"))"
    ws_dash["B12"] = "=SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[Status]=\"Rejected\"))"
    ws_dash["B13"] = "=IFERROR(SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*((tbl_Tracker[Status]=\"Accepted\")+(tbl_Tracker[Status]=\"Waived\")))/SUMPRODUCT((tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*(tbl_Tracker[VoyageID]<>\"\")),0)"
    ws_dash["B13"].number_format = "0%"
    
    for rr in range(8, 14):
        ws_dash[f"A{rr}"].border = styles['border_thin']
        ws_dash[f"B{rr}"].border = styles['border_thin']
        ws_dash[f"B{rr}"].font = Font(bold=True, size=11)
    
    # 상세 리스트 (FILTER 함수 사용)
    row = 15
    ws_dash.merge_cells(f"A{row}:G{row}")
    ws_dash[f"A{row}"] = "📋 Detail List (Filtered by Party)"
    ws_dash[f"A{row}"].font = styles['font_header']
    ws_dash[f"A{row}"].fill = styles['header_fill']
    ws_dash[f"A{row}"].alignment = styles['align_left']
    
    row += 1
    detail_headers = ["VoyageID", "DocCode", "DocName", "DueDate", "Status", "EvidenceLink", "RAG"]
    for col, h in enumerate(detail_headers, start=1):
        cell = ws_dash.cell(row=row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['subheader_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # FILTER 수식 (Party 필터 적용)
    row += 1
    ws_dash.cell(row, 1).value = (
        "=IFERROR("
        "FILTER("
        "CHOOSE({1,2,3,7,8,11,15},tbl_Tracker[VoyageID],tbl_Tracker[DocCode],"
        "INDEX(M_DocCatalog[DocName],MATCH(tbl_Tracker[DocCode],M_DocCatalog[DocCode],0)),"
        "tbl_Tracker[DueDate],tbl_Tracker[Status],tbl_Tracker[EvidenceLink],tbl_Tracker[RAG]),"
        "(tbl_Tracker[ResponsiblePartyID]=IF($B$5=\"ALL\",\"*\",INDEX(M_Parties[PartyID],MATCH($B$5,M_Parties[PartyName],0))))*"
        "(tbl_Tracker[VoyageID]<>\"\")"
        "),"
        "\"No rows\""
        ")"
    )
    
    set_col_widths(ws_dash, {1: 26, 2: 18, 3: 2, 4: 2, 5: 26})


def create_calendar_view_sheet(wb, styles: Dict) -> None:
    """Calendar_View sheet - 주간/월간 마감 캘린더"""
    ws_cal = wb.create_sheet("Calendar_View")
    ws_cal["A1"] = "Calendar View - Submission Deadlines"
    ws_cal["A1"].font = styles['font_title']
    
    # Filter row
    ws_cal["A2"] = "Voyage:"
    ws_cal["B2"] = "All"
    ws_cal["D2"] = "Party:"
    ws_cal["E2"] = "All"
    ws_cal["G2"] = "Period:"
    ws_cal["H2"] = "This Week"
    
    dv_voyage = DataValidation(type="list", formula1="=S_Voyages[VoyageID]", allow_blank=True)
    dv_party = DataValidation(type="list", formula1="=M_Parties[PartyName]", allow_blank=True)
    dv_period = DataValidation(type="list", formula1='"This Week,Next Week,This Month"', allow_blank=False)
    ws_cal.add_data_validation(dv_voyage)
    ws_cal.add_data_validation(dv_party)
    ws_cal.add_data_validation(dv_period)
    dv_voyage.add("B2")
    dv_party.add("E2")
    dv_period.add("H2")
    
    # Weekly view
    ws_cal["A4"] = "Weekly View"
    ws_cal["A4"].font = Font(bold=True, size=12)
    
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for col, day in enumerate(weekdays, start=1):
        cell = ws_cal.cell(5, col)
        cell.value = day
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.alignment = styles['align_center']
    
    # Monthly view
    ws_cal["A20"] = "Monthly View"
    ws_cal["A20"].font = Font(bold=True, size=12)
    
    set_col_widths(ws_cal, {1: 16, 2: 14, 3: 2, 4: 16, 5: 14, 6: 2, 7: 16, 8: 14})


def create_holidays_sheet(wb, styles: Dict) -> None:
    """Holidays sheet - UAE 휴일/프로젝트 휴무일 (WORKDAY.INTL 지원)"""
    ws_hol = wb.create_sheet("Holidays")
    ws_hol["A1"] = "Holidays Calendar (for WORKDAY.INTL)"
    ws_hol["A1"].font = styles['font_title']
    
    headers_hol = ["Date", "Holiday Name", "Type"]
    header_row = 3
    
    for col, h in enumerate(headers_hol, start=1):
        cell = ws_hol.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    # UAE National Holidays 2026
    uae_holidays = [
        (dt.date(2026, 1, 1), "New Year's Day", "UAE National"),
        (dt.date(2026, 5, 1), "Labour Day", "UAE National"),
        # Add more as needed
    ]
    
    start_row = header_row + 1
    for i, (hol_date, name, htype) in enumerate(uae_holidays):
        r = start_row + i
        ws_hol.cell(r, 1).value = hol_date
        ws_hol.cell(r, 2).value = name
        ws_hol.cell(r, 3).value = htype
        ws_hol.cell(r, 1).number_format = "yyyy-mm-dd"
        
        for c in range(1, 4):
            cell = ws_hol.cell(r, c)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            cell.alignment = styles['align_left'] if c == 2 else styles['align_center']
    
    set_col_widths(ws_hol, {1: 14, 2: 30, 3: 18})
    ws_hol.freeze_panes = "A4"


def create_vba_pasteboard_sheet(wb, styles: Dict) -> None:
    """VBA_Pasteboard sheet - VBA 설치 보드"""
    ws_vba = wb.create_sheet("VBA_Pasteboard")
    ws_vba["A1"] = "VBA Installation Pasteboard"
    ws_vba["A1"].font = styles['font_title']
    ws_vba["A2"] = "=" * 50
    ws_vba["A2"].font = Font(bold=True)
    
    # Installation Checklist
    row = 3
    ws_vba.cell(row, 1).value = "[Installation Checklist]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    
    checklist_headers = ["StepNo", "Action", "Done", "Notes"]
    for col, h in enumerate(checklist_headers, start=1):
        cell = ws_vba.cell(row, col, value=h)
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.border = styles['border_thin']
    
    checklist_data = [
        (1, "VBE에서 모듈 Import", "", "modOperations.bas, modDashboard.bas"),
        (2, "참조 라이브러리 확인", "", "Microsoft Scripting Runtime (선택, Late Binding 권장)"),
        (3, "버튼 연결 확인", "", "Dashboard 버튼 매핑"),
        (4, "샘플 Voyage 생성", "", "S_Voyages에 V01 추가"),
        (5, "GenerateTrackerRows 실행", "", "T_Tracker 행 생성 확인"),
    ]
    
    row += 1
    for step, action, done, notes in checklist_data:
        ws_vba.cell(row, 1).value = step
        ws_vba.cell(row, 2).value = action
        ws_vba.cell(row, 3).value = done
        ws_vba.cell(row, 4).value = notes
        for c in range(1, 5):
            cell = ws_vba.cell(row, c)
            cell.border = styles['border_thin']
        row += 1
    
    # Module Inventory
    row += 2
    ws_vba.cell(row, 1).value = "=" * 50
    ws_vba.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_vba.cell(row, 1).value = "[Module Inventory]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    
    module_headers = ["ModuleName", "Type", "Purpose", "Dependencies", "InstallOrder", "Buttons"]
    for col, h in enumerate(module_headers, start=1):
        cell = ws_vba.cell(row, col, value=h)
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.border = styles['border_thin']
    
    module_data = [
        ("modOperations", "Standard", "Init/Generate/Recalc/Validate/Export", "None", 1, "btn_Init, btn_Generate, btn_Recalc, btn_Validate, btn_Export"),
        ("modDashboard", "Standard", "Dashboard refresh/filter", "modOperations", 2, "btn_RefreshDashboard"),
        ("T_Tracker Sheet", "Sheet", "Worksheet_Change event", "None", 3, "(자동)"),
    ]
    
    row += 1
    for mod_name, mod_type, purpose, deps, order, buttons in module_data:
        ws_vba.cell(row, 1).value = mod_name
        ws_vba.cell(row, 2).value = mod_type
        ws_vba.cell(row, 3).value = purpose
        ws_vba.cell(row, 4).value = deps
        ws_vba.cell(row, 5).value = order
        ws_vba.cell(row, 6).value = buttons
        for c in range(1, 7):
            cell = ws_vba.cell(row, c)
            cell.border = styles['border_thin']
        row += 1
    
    # Buttons Mapping
    row += 2
    ws_vba.cell(row, 1).value = "=" * 50
    ws_vba.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_vba.cell(row, 1).value = "[Buttons Mapping]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    
    button_headers = ["Button Name", "Macro", "Location"]
    for col, h in enumerate(button_headers, start=1):
        cell = ws_vba.cell(row, col, value=h)
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.border = styles['border_thin']
    
    button_data = [
        ("btn_Init", "InitializeWorkbook", "Dashboard"),
        ("btn_Generate", "GenerateTrackerRows", "Dashboard"),
        ("btn_Recalc", "RecalcDeadlines", "Dashboard"),
        ("btn_Validate", "ValidateBeforeExport", "Dashboard"),
        ("btn_Export", "ExportVoyagePack", "Dashboard"),
        ("btn_RefreshDashboard", "RefreshDashboard", "Dashboard"),
    ]
    
    row += 1
    for btn_name, macro, location in button_data:
        ws_vba.cell(row, 1).value = btn_name
        ws_vba.cell(row, 2).value = macro
        ws_vba.cell(row, 3).value = location
        for c in range(1, 4):
            cell = ws_vba.cell(row, c)
            cell.border = styles['border_thin']
        row += 1
    
    set_col_widths(ws_vba, {1: 20, 2: 30, 3: 20, 4: 40, 5: 12, 6: 40})


def create_instructions_sheet(wb, styles: Dict) -> None:
    """Instructions sheet - 사용 가이드"""
    ws_ins = wb.create_sheet("Instructions")
    ws_ins["A1"] = "How to use TR_DocHub_AGI_2026"
    ws_ins["A1"].font = styles['font_title']
    
    instructions = [
        "1) S_Voyages에 항차 일정 입력 (VoyageID, MZP Arrival, Load-out 등)",
        "2) M_DocCatalog에서 문서 요구사항 관리 (DocCode, RequiredFlag, ActiveFlag)",
        "3) R_DeadlineRules에서 DueDate 룰 설정 (DocCode → AnchorField + OffsetDays)",
        "4) Dashboard에서 [Generate] 버튼 클릭 → T_Tracker 행 자동 생성",
        "5) Dashboard에서 [Recalc] 버튼 클릭 → DueDate 자동 계산",
        "6) T_Tracker에서 Status/SubmittedDate/AcceptedDate 업데이트",
        "7) Dashboard에서 Party 선택 → KPI 및 상세 리스트 확인",
        "8) [Export] 버튼으로 항차별 PDF/CSV 생성",
        "",
        "DueDate 계산 로직:",
        "- R_DeadlineRules에서 DocCode별 룰 조회 (Priority 최소값 우선)",
        "- AnchorField로 S_Voyages에서 해당 날짜 조회",
        "- DueDate = AnchorDate + OffsetDays (CalendarType=WD면 WORKDAY.INTL 적용)",
        "",
        "⚠️ 중요:",
        "- VoyageID와 DocCode는 복합키 (중복 불가)",
        "- GenerateTrackerRows 실행 전 S_Voyages와 M_DocCatalog 입력 완료 필요",
        "- EvidenceLink는 파일 경로 또는 하이퍼링크 형식",
    ]
    
    for i, line in enumerate(instructions, start=3):
        ws_ins[f"A{i}"] = line
        ws_ins[f"A{i}"].alignment = styles['align_left']
        ws_ins[f"A{i}"].font = styles['font_normal']
    
    ws_ins.column_dimensions["A"].width = 120


# =========================
# Main Builder
# =========================
def build_tr_dochub_agi_template(output_path: Path) -> Path:
    """Build complete TR_DocHub_AGI template"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    styles = get_styles()
    
    # Create sheets in order
    create_lists_sheet(wb, styles)
    create_c_config_sheet(wb, styles)
    create_s_voyages_sheet(wb, styles)
    create_m_parties_sheet(wb, styles)
    create_m_doccatalog_sheet(wb, styles)
    create_r_deadline_rules_sheet(wb, styles)
    create_t_tracker_sheet(wb, styles)
    create_d_dashboard_sheet(wb, styles)
    create_calendar_view_sheet(wb, styles)
    create_holidays_sheet(wb, styles)
    create_vba_pasteboard_sheet(wb, styles)
    create_instructions_sheet(wb, styles)
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    ap = argparse.ArgumentParser()
    # 기본값에 날짜 자동 추가 (중복 방지)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"TR_DocHub_AGI_2026_Template_{timestamp}.xlsx"
    ap.add_argument("--output", type=str, default=default_filename, help="Output xlsx path (default: auto-generated with timestamp)")
    args = ap.parse_args()
    
    output_path = build_tr_dochub_agi_template(Path(args.output))
    print(f"[OK] Created: {output_path}")
    print(f"Next steps:")
    print(f"   1. Open in Excel -> Save as .xlsm")
    print(f"   2. Import VBA modules from VBA_Pasteboard")
    print(f"   3. Run InitializeWorkbook() -> GenerateTrackerRows()")


if __name__ == "__main__":
    main()
