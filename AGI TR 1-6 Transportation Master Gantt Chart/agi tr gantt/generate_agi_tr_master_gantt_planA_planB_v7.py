#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI Site – HVDC Transformers (TR1–TR6) Master Gantt Generator (Plan-A/Plan-B)

Output
- Excel .xlsx with:
  1) Inputs: MS approval date + parameters + derived LO1 per plan
  2) Tide_Peaks_MZP: daily peak tide extracted from MAMMOET_AGI TR.pdf (Jan–Apr 2026)
  3) Plan_A_Reserve: Reserve/Standby scenario (Standby exposure starts from 2026-01-06)
  4) Plan_B_Release: Release/Re-mobilize scenario (No standby; reconfirm/remob lead time)
  5) Gantt_Chart_A / Gantt_Chart_B: stacked-bar Gantt (Offset+Duration)

Key idea
- Dates in Plan sheets are formula-driven by Inputs!MS_APPROVAL and lead days.
  Change Inputs!MS_APPROVAL to auto-shift the entire campaign.
- LO/LI must be aligned to high tide windows; planner should validate using Tide_Peaks_MZP.

Run
  pip install openpyxl pymupdf pandas
  python generate_agi_tr_master_gantt_planA_planB_v7.py --ms 2026-01-16 --out AGI_TR_Master_Gantt.xlsx

Notes
- Tide peaks are extracted from the uploaded PDF layout (hourly matrix). If the PDF layout changes, parsing may fail.
- Treat the tide table as planning support only; reconfirm with Port Control / Harbour Master for execution.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import List

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


PHASE_COLORS = {
    "DOC": "D9D2E9",
    "PORT": "9DC3E6",
    "MARINE": "BFBFBF",
    "AGI": "C6E0B4",
    "RETURN": "E7E6E6",
    "MILESTONE": "F8CBAD",
    "STANDBY": "F4CCCC",
    "REMOB": "FFE599",
    "BUFFER": "D0E0E3",
}
WEEKEND_SHADE = "F2F2F2"
TODAY_SHADE = "FFF2CC"
HEADER_BLUE = "1F4E79"
SUMMARY_GREY = "404040"

BORDER_THIN = Side(style="thin", color="A6A6A6")


def thin_border() -> Border:
    return Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)


def parse_datestr(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%d-%b-%Y").date()


def parse_tide_page(page):
    words = page.get_text("words")
    time_words = [w for w in words if re.match(r"^\d{1,2}:\d{2}$", w[4])]
    time_words = sorted(time_words, key=lambda w: w[0])
    if len(time_words) < 20:
        return None
    time_cols = [(w[4], (w[0] + w[2]) / 2) for w in time_words]

    date_words = [w for w in words if re.match(r"^\d{2}-[A-Za-z]{3}-2026$", w[4])]
    date_words = sorted(date_words, key=lambda w: w[1])
    if not date_words:
        return None

    num_words = [w for w in words if re.match(r"^\d\.\d{2}$", w[4]) or re.match(r"^\d\.\d{1}$", w[4])]

    rows = []
    for dw in date_words:
        date_str = dw[4]
        y_center = (dw[1] + dw[3]) / 2
        candidates = [w for w in num_words if abs(((w[1] + w[3]) / 2) - y_center) < 3.5]
        if not candidates:
            continue

        values = []
        for _, x in time_cols:
            near = [w for w in candidates if abs(((w[0] + w[2]) / 2) - x) < 8.0]
            w = min(near, key=lambda w: abs(((w[0] + w[2]) / 2) - x)) if near else min(
                candidates, key=lambda w: abs(((w[0] + w[2]) / 2) - x)
            )
            try:
                values.append(float(w[4]))
            except Exception:
                values.append(float("nan"))
        rows.append((date_str, values))

    if not rows:
        return None

    df = pd.DataFrame({"date_str": [r[0] for r in rows]})
    for i, (t, _) in enumerate(time_cols):
        df[t] = [r[1][i] for r in rows]
    return df


def daily_peaks(df: pd.DataFrame) -> pd.DataFrame:
    time_cols = [c for c in df.columns if c != "date_str"]
    out = []
    for _, row in df.iterrows():
        vals = row[time_cols].astype(float)
        out.append((row["date_str"], vals.idxmax(), float(vals.max())))
    return pd.DataFrame(out, columns=["date_str", "peak_time", "peak_m"])


def apply_cell_style(cell, *, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def auto_fit_metadata_columns(ws):
    widths = {"A": 7, "B": 9, "C": 56, "D": 11, "E": 16, "F": 22, "G": 12, "H": 12, "I": 8, "J": 24, "K": 22, "L": 46}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def set_gantt_day_column_width(ws, start_col, end_col, width=3.0):
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def add_named_cell(wb: Workbook, name: str, sheet: str, addr: str):
    wb.defined_names.add(DefinedName(name=name, attr_text=f"{sheet}!${addr[0]}${addr[1:]}"))


@dataclass(frozen=True)
class TaskRow:
    id: str
    wbs: str
    task: str
    phase: str
    location: str
    owner: str
    start_formula: str
    finish_formula: str
    tide_gate: str = ""
    weather_gate: str = ""
    notes: str = ""


def build_plan_tasks(plan: str, n_tr: int = 6) -> List[TaskRow]:
    assert plan in ("A", "B")
    LO1 = "LO1_PLANA" if plan == "A" else "LO1_PLANB"
    PLAN_NAME = "Plan-A (Reserve/Standby)" if plan == "A" else "Plan-B (Release/Re-mobilize)"
    camp_start = "EQUIP_ARRIVAL" if plan == "A" else "MS_APPROVAL"
    last_lo = f"={LO1}+({n_tr}-1)*CYCLE_SPACING"
    last_return_end = f"=({last_lo})+1+SAIL_DAYS+LI_DAYS+TURN_DAYS+JACK_DAYS+BUF_DAYS+RETURN_DAYS"

    rows: List[TaskRow] = []
    rows.append(
        TaskRow(
            id="0",
            wbs="0",
            task=f"Campaign: TR1–TR{n_tr} (1TR/Trip) – Realistic + Buffer | {PLAN_NAME}",
            phase="SUMMARY",
            location="MZP / AGI",
            owner="SCT / Mammoet / LCT Operator",
            start_formula=f"={camp_start}",
            finish_formula=last_return_end,
            notes="Edit Inputs!MS_APPROVAL + lead-days; validate LO/LI against Tide_Peaks_MZP.",
        )
    )

    if plan == "A":
        rows.append(
            TaskRow(
                id="A001",
                wbs="A.0",
                task="Equipment arrival at MZP + Mobilization/Assembly/Deck prep (non-LO/LI work)",
                phase="PORT",
                location="Mina Zayed Port",
                owner="Mammoet + Port Ops",
                start_formula="=EQUIP_ARRIVAL",
                finish_formula="=EQUIP_ARRIVAL+2",
                notes="Based on original tentative sequence (arrival→assembly→deck prep).",
            )
        )
        rows.append(
            TaskRow(
                id="A002",
                wbs="A.1",
                task="Standby reservation (equipment + manpower) pending MS approval / tide window",
                phase="STANDBY",
                location="Mina Zayed Port",
                owner="Mammoet",
                start_formula="=EQUIP_ARRIVAL",
                finish_formula=f"={LO1}-1",
                notes="Standby charge exposure increases with LO1 delay.",
            )
        )
    else:
        rows.append(
            TaskRow(
                id="B001",
                wbs="B.0",
                task="Release (no standby) – availability subject to reconfirmation",
                phase="MILESTONE",
                location="Mina Zayed Port",
                owner="SCT",
                start_formula="=MS_APPROVAL",
                finish_formula="=MS_APPROVAL",
                notes="No reservation; reconfirm availability once commencement date is finalized.",
            )
        )
        rows.append(
            TaskRow(
                id="B002",
                wbs="B.1",
                task="Reconfirm availability + Re-mobilization/Assembly/Deck prep",
                phase="REMOB",
                location="Mina Zayed Port",
                owner="Mammoet + SCT",
                start_formula="=MS_APPROVAL",
                finish_formula=f"={LO1}-1",
                notes="LO1 may slip if equipment/manpower not immediately available.",
            )
        )

    for tr in range(1, n_tr + 1):
        base = tr * 100
        lo = f"={LO1}+({tr}-1)*CYCLE_SPACING"
        docs_start = f"={lo}-DOCS_DAYS"
        docs_finish = f"={lo}-1"
        sail_start = f"={lo}+1"
        sail_finish = f"={lo}+SAIL_DAYS"
        li_start = f"={sail_finish}+1"
        li_finish = f"={li_start}+LI_DAYS-1"
        turn_start = f"={li_finish}+1"
        turn_finish = f"={turn_start}+TURN_DAYS-1"
        jack_start = f"={turn_finish}+1"
        jack_finish = f"={jack_start}+JACK_DAYS-1"
        buf_start = f"={jack_finish}+1"
        buf_finish = f"={buf_start}+BUF_DAYS-1"
        ret_start = f"={buf_finish}+1"
        ret_finish = f"={ret_start}+RETURN_DAYS-1"

        rows.append(
            TaskRow(
                id=str(tr),
                wbs=f"{tr}.0",
                task=f"TR{tr} – Transport Cycle (MZP→AGI→MZP)",
                phase="SUMMARY",
                location="MZP / AGI",
                owner="SCT / Mammoet / LCT Operator",
                start_formula=docs_start,
                finish_formula=ret_finish,
                notes="Includes docs/prep + LO/LI + Turning/JD + buffer + return/reset.",
            )
        )

        rows.extend(
            [
                TaskRow(
                    id=str(base + 1),
                    wbs=f"{tr}.1",
                    task=f"TR{tr} – Docs/Permits/Marine pack (MS, ballast, PTW, Gate pass)",
                    phase="DOC",
                    location="Mina Zayed Port",
                    owner="SCT + TP Eng + Mammoet",
                    start_formula=docs_start,
                    finish_formula=docs_finish,
                ),
                TaskRow(
                    id=str(base + 2),
                    wbs=f"{tr}.2",
                    task=f"TR{tr} – Load-out (RoRo) to LCT deck (tide window required)",
                    phase="PORT",
                    location="Mina Zayed Port",
                    owner="Mammoet + LCT + Port",
                    start_formula=lo,
                    finish_formula=lo,
                    tide_gate="High tide window (see Tide_Peaks_MZP)",
                    weather_gate="Per approved MS",
                ),
                TaskRow(
                    id=str(base + 3),
                    wbs=f"{tr}.3",
                    task=f"TR{tr} – Seafastening + Marine transit to AGI",
                    phase="MARINE",
                    location="MZP↔AGI (Marine)",
                    owner="LCT Crew + Mammoet",
                    start_formula=sail_start,
                    finish_formula=sail_finish,
                    weather_gate="Per approved MS",
                ),
                TaskRow(
                    id=str(base + 4),
                    wbs=f"{tr}.4",
                    task=f"TR{tr} – Load-in (RoRo) + Store on jetty (tide window required)",
                    phase="AGI",
                    location="AGI Site",
                    owner="AGI Port + Mammoet",
                    start_formula=li_start,
                    finish_formula=li_finish,
                    tide_gate="High tide window (see Tide_Peaks_MZP)",
                    weather_gate="Per approved MS",
                ),
                TaskRow(
                    id=str(base + 5),
                    wbs=f"{tr}.5",
                    task=f"TR{tr} – Turning on foundation area",
                    phase="AGI",
                    location="AGI Site",
                    owner="Mammoet Jacking Crew",
                    start_formula=turn_start,
                    finish_formula=turn_finish,
                ),
                TaskRow(
                    id=str(base + 6),
                    wbs=f"{tr}.6",
                    task=f"TR{tr} – Jacking-down on temporary support",
                    phase="AGI",
                    location="AGI Site",
                    owner="Mammoet Jacking Crew",
                    start_formula=jack_start,
                    finish_formula=jack_finish,
                ),
                TaskRow(
                    id=str(base + 7),
                    wbs=f"{tr}.7",
                    task=f"TR{tr} – Buffer (Weather/Tide/Port congestion)",
                    phase="BUFFER",
                    location="MZP / AGI",
                    owner="SCT + Mammoet",
                    start_formula=buf_start,
                    finish_formula=buf_finish,
                ),
                TaskRow(
                    id=str(base + 8),
                    wbs=f"{tr}.8",
                    task=f"TR{tr} – Return to MZP + Reset / Demob-Mob",
                    phase="RETURN",
                    location="MZP↔AGI (Marine)",
                    owner="LCT Crew + Mammoet",
                    start_formula=ret_start,
                    finish_formula=ret_finish,
                ),
            ]
        )

    rows.append(
        TaskRow(
            id="M1",
            wbs="M",
            task=f"Milestone: TR{n_tr} returned to MZP (Campaign complete) | {PLAN_NAME}",
            phase="MILESTONE",
            location="Mina Zayed Port",
            owner="SCT",
            start_formula=last_return_end,
            finish_formula=last_return_end,
        )
    )
    return rows


def create_inputs_sheet(wb: Workbook, ms_approval: dt.date):
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "AGI TR Campaign Inputs (Plan-A / Plan-B)"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A4"], ws["B4"] = "Name", "Value"
    for c in ("A4", "B4"):
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].fill = PatternFill("solid", fgColor=HEADER_BLUE)
        ws[c].alignment = Alignment(horizontal="center")
        ws[c].border = thin_border()

    items = [
        ("MS_APPROVAL (Method Statement + ballast calc approved date)", ms_approval),
        ("EQUIP_ARRIVAL (MZP arrival for reserved resources)", dt.date(2026, 1, 6)),
        ("PlanA_LeadDays_to_LO1 (reserve: days from MS approval to LO1)", 2),
        ("PlanB_LeadDays_to_LO1 (release: days from MS approval to LO1)", 7),
        ("DOCS_DAYS (pre-LO docs/prep days)", 3),
        ("SAIL_DAYS (seafastening+transit days)", 2),
        ("LI_DAYS (load-in + store days)", 2),
        ("TURN_DAYS (turning days)", 3),
        ("JACK_DAYS (jacking-down days)", 1),
        ("BUF_DAYS (embedded weather/tide buffer per trip)", 2),
        ("RETURN_DAYS (return + reset days)", 4),
        ("CYCLE_SPACING (days between LO of consecutive TRs)", 15),
    ]
    r = 5
    for name, val in items:
        ws[f"A{r}"] = name
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{r}"] = val
        ws[f"A{r}"].border = thin_border()
        ws[f"B{r}"].border = thin_border()
        if isinstance(val, dt.date):
            ws[f"B{r}"].number_format = "yyyy-mm-dd"
        r += 1

    r += 1
    ws[f"A{r}"] = "Derived"
    ws[f"A{r}"].font = Font(bold=True)
    ws[f"A{r+1}"], ws[f"A{r+2}"] = "LO1_PLANA", "LO1_PLANB"
    ws[f"B{r+1}"] = "=MS_APPROVAL + PlanA_LeadDays_to_LO1"
    ws[f"B{r+2}"] = "=MS_APPROVAL + PlanB_LeadDays_to_LO1"
    ws[f"B{r+1}"].number_format = "yyyy-mm-dd"
    ws[f"B{r+2}"].number_format = "yyyy-mm-dd"
    for rr in (r + 1, r + 2):
        ws[f"A{rr}"].border = thin_border()
        ws[f"B{rr}"].border = thin_border()

    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 20

    mapping = {
        "MS_APPROVAL": "B5",
        "EQUIP_ARRIVAL": "B6",
        "PlanA_LeadDays_to_LO1": "B7",
        "PlanB_LeadDays_to_LO1": "B8",
        "DOCS_DAYS": "B9",
        "SAIL_DAYS": "B10",
        "LI_DAYS": "B11",
        "TURN_DAYS": "B12",
        "JACK_DAYS": "B13",
        "BUF_DAYS": "B14",
        "RETURN_DAYS": "B15",
        "CYCLE_SPACING": "B16",
        "LO1_PLANA": f"B{r+1}",
        "LO1_PLANB": f"B{r+2}",
    }
    for nm, addr in mapping.items():
        add_named_cell(wb, nm, "Inputs", addr)


def create_tide_peaks_sheet(wb: Workbook, peaks_df: pd.DataFrame):
    ws = wb.create_sheet("Tide_Peaks_MZP")
    ws["A1"] = "Mina Zayed – Daily Peak Tide (GMT+4, metres above Chart Datum) | Source: MAMMOET_AGI TR.pdf"
    ws["A1"].font = Font(bold=True, size=12)

    for c, h in enumerate(["Date", "Peak time", "Peak tide (m)"], start=1):
        cell = ws.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    row0 = 4
    for i, r in enumerate(peaks_df.itertuples(index=False), start=row0):
        ws.cell(i, 1, value=r.date).number_format = "yyyy-mm-dd"
        ws.cell(i, 2, value=r.peak_time)
        ws.cell(i, 3, value=float(r.peak_m)).number_format = "0.00"
        for c in range(1, 4):
            ws.cell(i, c).border = thin_border()
            ws.cell(i, c).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    dxf = DifferentialStyle(fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"))
    rule = Rule(type="expression", dxf=dxf, formula=["$C4>=2"])
    rule.priority = 1
    ws.conditional_formatting.add(f"A4:C{row0+len(peaks_df)-1}", rule)


def create_plan_sheet(wb: Workbook, sheet_name: str, tasks: List[TaskRow], horizon_start: dt.date, horizon_end: dt.date):
    ws = wb.create_sheet(sheet_name)

    # day axis
    days = []
    cur = horizon_start
    while cur <= horizon_end:
        days.append(cur)
        cur += dt.timedelta(days=1)

    META_COLS = ["ID", "WBS", "Task", "Phase", "Location", "Owner", "Start", "Finish", "Dur (d)", "Tide gate", "Weather gate", "Notes"]
    gantt_start_col = len(META_COLS) + 1
    header_row = 5
    first_data_row = 7
    last_data_row = first_data_row + len(tasks) - 1
    gantt_end_col = gantt_start_col + len(days) - 1

    # titles
    for r in (1, 2, 3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=gantt_end_col)
    ws["A1"] = f"{sheet_name} – AGI TR1–TR6 (1TR/Trip) Realistic Gantt (Plan-A/Plan-B)"
    ws["A2"] = "Change Inputs!MS_APPROVAL and lead-days; schedule auto-shifts. Verify LO/LI against Tide_Peaks_MZP."
    ws["A3"] = "Dates are formulas; Excel recalculates."

    title_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    for r in (1, 2, 3):
        for c in range(1, gantt_end_col + 1):
            cell = ws.cell(r, c)
            cell.fill = title_fill
            cell.font = Font(bold=True if r == 1 else False, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 20

    # header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, name in enumerate(META_COLS, start=1):
        cell = ws.cell(header_row, i, value=name)
        apply_cell_style(cell, font=header_font, fill=header_fill, alignment=header_align, border=thin_border())

    for j, d in enumerate(days):
        col = gantt_start_col + j
        cell = ws.cell(header_row, col, value=dt.datetime(d.year, d.month, d.day))
        apply_cell_style(cell, font=header_font, fill=header_fill, alignment=header_align, border=thin_border(), number_format="dd-mmm")
    ws.row_dimensions[header_row].height = 44

    # data
    date_fmt = "yyyy-mm-dd"
    for r, t in enumerate(tasks, start=first_data_row):
        meta = [t.id, t.wbs, t.task, t.phase, t.location, t.owner, None, None, None, t.tide_gate, t.weather_gate, t.notes]
        for c, v in enumerate(meta, start=1):
            cell = ws.cell(r, c, value=v)
            if c == 7:
                cell.value = t.start_formula
                cell.number_format = date_fmt
                align = Alignment(horizontal="center")
            elif c == 8:
                cell.value = t.finish_formula
                cell.number_format = date_fmt
                align = Alignment(horizontal="center")
            elif c == 9:
                cell.value = f"=H{r}-G{r}+1"
                cell.number_format = "0"
                align = Alignment(horizontal="center")
            elif c in (1, 2, 4):
                align = Alignment(horizontal="center")
            else:
                align = Alignment(horizontal="left", wrap_text=True)
            apply_cell_style(cell, alignment=align, border=thin_border())

        if t.phase == "SUMMARY":
            for c in range(1, len(META_COLS) + 1):
                apply_cell_style(
                    ws.cell(r, c),
                    font=Font(bold=True, color="FFFFFF"),
                    fill=PatternFill("solid", fgColor=SUMMARY_GREY),
                    border=thin_border(),
                )

        for c in range(gantt_start_col, gantt_end_col + 1):
            ws.cell(r, c).border = thin_border()

    auto_fit_metadata_columns(ws)
    set_gantt_day_column_width(ws, gantt_start_col, gantt_end_col, width=3.0)
    ws.freeze_panes = ws.cell(first_data_row, gantt_start_col)

    # conditional formatting
    cf_range = f"{get_column_letter(gantt_start_col)}{first_data_row}:{get_column_letter(gantt_end_col)}{last_data_row}"
    day_col_letter = get_column_letter(gantt_start_col)

    def add_cf(formula, fill_hex, pr):
        dxf = DifferentialStyle(fill=PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid"))
        rule = Rule(type="expression", dxf=dxf, formula=[formula])
        rule.priority = pr
        ws.conditional_formatting.add(cf_range, rule)

    add_cf(f"WEEKDAY({day_col_letter}${header_row},2)>5", WEEKEND_SHADE, 1)
    add_cf(f"{day_col_letter}${header_row}=TODAY()", TODAY_SHADE, 2)

    pr = 3
    for phase, color in PHASE_COLORS.items():
        add_cf(
            f'AND($D{first_data_row}="{phase}",{day_col_letter}${header_row}>=$G{first_data_row},{day_col_letter}${header_row}<=$H{first_data_row})',
            color,
            pr,
        )
        pr += 1


def create_gantt_chart_sheet(wb: Workbook, plan_sheet_name: str, chart_sheet_name: str):
    ws_plan = wb[plan_sheet_name]
    wc = wb.create_sheet(chart_sheet_name)

    wc["A1"] = f"{chart_sheet_name} – Stacked Bar (Offset+Duration) from {plan_sheet_name}"
    wc["A2"] = "Technique: Microsoft Gantt via stacked bar (Offset hidden)."
    wc["A1"].font = Font(bold=True, size=12)
    wc["A3"], wc["B3"], wc["C3"], wc["D3"] = "Task", "Start", "Offset (d)", "Duration (d)"

    for c in range(1, 5):
        cell = wc.cell(3, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    first_data_row = 7
    last_data_row = ws_plan.max_row
    detail_rows = [r for r in range(first_data_row, last_data_row + 1) if ws_plan.cell(r, 4).value != "SUMMARY"]

    start_row = 4
    for i, rr in enumerate(detail_rows):
        r = start_row + i
        wc.cell(r, 1).value = f"='{plan_sheet_name}'!C{rr}"
        wc.cell(r, 2).value = f"='{plan_sheet_name}'!G{rr}"
        wc.cell(r, 2).number_format = "yyyy-mm-dd"
        min_rng = f"$B${start_row}:$B${start_row+len(detail_rows)-1}"
        wc.cell(r, 3).value = f"=B{r}-MIN({min_rng})"
        wc.cell(r, 4).value = f"='{plan_sheet_name}'!I{rr}"
        wc.cell(r, 4).number_format = "0"
        for c in range(1, 5):
            wc.cell(r, c).border = thin_border()
            wc.cell(r, c).alignment = Alignment(horizontal="left" if c == 1 else "center", wrap_text=(c == 1))

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 150
    chart.title = f"{plan_sheet_name} – Tasks"

    data = Reference(wc, min_col=3, min_row=3, max_col=4, max_row=start_row + len(detail_rows) - 1)
    cats = Reference(wc, min_col=1, min_row=start_row, max_row=start_row + len(detail_rows) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    if chart.series:
        chart.series[0].graphicalProperties.noFill = True
        chart.series[0].graphicalProperties.line.noFill = True

    wc.add_chart(chart, "F3")


def build_workbook(ms_approval: dt.date, tide_pdf_path: Path, out_path: Path, n_tr: int):
    # tide peaks
    doc = fitz.open(str(tide_pdf_path))
    peaks_all = []
    for pno in range(len(doc)):
        dfp = parse_tide_page(doc[pno])
        if dfp is None:
            continue
        peaks_all.append(daily_peaks(dfp))
    peaks = pd.concat(peaks_all, ignore_index=True)
    peaks["date"] = peaks["date_str"].apply(parse_datestr)
    peaks = peaks.sort_values("date").reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)

    create_inputs_sheet(wb, ms_approval)
    create_tide_peaks_sheet(wb, peaks[["date", "peak_time", "peak_m"]])

    # horizon (includes standby start from 2026-01-06)
    horizon_start = dt.date(2026, 1, 6)
    lo1_a = ms_approval + dt.timedelta(days=2)
    horizon_end = lo1_a + dt.timedelta(days=(n_tr - 1) * 15 + 20)

    tasks_a = build_plan_tasks("A", n_tr)
    tasks_b = build_plan_tasks("B", n_tr)

    create_plan_sheet(wb, "Plan_A_Reserve", tasks_a, horizon_start, horizon_end)
    create_plan_sheet(wb, "Plan_B_Release", tasks_b, horizon_start, horizon_end)
    create_gantt_chart_sheet(wb, "Plan_A_Reserve", "Gantt_Chart_A")
    create_gantt_chart_sheet(wb, "Plan_B_Release", "Gantt_Chart_B")

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ms", default="2026-01-16", help="MS approval date (YYYY-MM-DD)")
    ap.add_argument("--pdf", default="MAMMOET_AGI TR.pdf", help="Tide PDF path (default: MAMMOET_AGI TR.pdf)")
    ap.add_argument("--out", default="AGI_TR_Master_Gantt_PlanA_PlanB_v7.xlsx", help="Output xlsx path")
    ap.add_argument("--n", type=int, default=6, help="Number of transformers (default 6)")
    args = ap.parse_args()

    ms = dt.datetime.strptime(args.ms, "%Y-%m-%d").date()
    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        raise FileNotFoundError(f"Tide PDF not found: {pdf_path}")

    out_path = Path(args.out)
    build_workbook(ms, pdf_path, out_path, args.n)
    print(f"[OK] Wrote: {out_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
