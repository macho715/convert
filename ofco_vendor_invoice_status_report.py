#!/usr/bin/env python3
"""
OFCO Sheet1 Vendor Invoice Status Report (READY / PENDING / MISMATCH / DONE)

What this does
--------------
1) Reads an OFCO workbook (default sheet: "Sheet1")
2) Extracts vendor invoice numbers from the "SUBJECT" column
3) Compares those invoice numbers against parsed vendor-invoice files found in vendor_dir
4) Generates an Excel report with sheets:
   - SUMMARY, ALL, READY, PENDING, MISMATCH, DONE

Status meaning
--------------
- DONE     : vendor file exists AND Sheet1 row matches vendor aggregated values (cost_item_fields + Total_Amount_AED)
- READY    : vendor file exists AND vendor total matches Sheet1 total (±tolerance) BUT Sheet1 values are not yet applied
- PENDING  : vendor file not found in vendor_dir
- MISMATCH : vendor file exists BUT totals mismatch, multiple candidate files, or vendor file parsing failure

Inputs
------
- ofco_excel_path : OFCO workbook path
- vendor_dir      : directory containing parsed vendor invoice xlsx (aggregated single-row format recommended)
- cost_fields_json: cost_item_fields.JSON path (88 fields)

Notes
-----
- Vendor invoice discovery is filename-based by default (extracts "INV...digits" from filenames).
  If your vendor files do not include invoice numbers in filenames, you can extend the discovery
  routine to read inside the file(s) and detect VENDOR_INVOICE_NO columns.

- Tolerances:
  - absolute: 0.01 AED
  - relative: 2% of Sheet1 Total_Amount_AED
"""

from __future__ import annotations

import argparse
import datetime as _dt
import glob
import json
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd


# Loose pattern that works with underscores in filenames
FILENAME_INV_PATTERN = re.compile(r"INV[- ]?\d{5,}", re.IGNORECASE)
SUBJECT_INV_PATTERN = re.compile(r"\bINV[- ]?\d{5,}\b", re.IGNORECASE)


def _num(x: Any) -> float:
    if x is None:
        return 0.0
    try:
        return float(x)
    except Exception:
        try:
            return float(str(x).replace(",", ""))
        except Exception:
            return 0.0


def normalize_inv(inv: str) -> str:
    return inv.replace(" ", "").upper()


def extract_invs_from_subject(text: str) -> List[str]:
    if not text:
        return []
    return [normalize_inv(m.group(0)) for m in SUBJECT_INV_PATTERN.finditer(text)]


def extract_invs_from_filename(name: str) -> List[str]:
    if not name:
        return []
    t = name.replace("_", " ").replace(".", " ")
    return [normalize_inv(m.group(0)) for m in FILENAME_INV_PATTERN.finditer(t)]


def load_cost_item_fields(cost_fields_json_path: str) -> List[str]:
    with open(cost_fields_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    fields = data.get("cost_item_fields") or []
    if not isinstance(fields, list) or not fields:
        raise ValueError("Invalid cost_item_fields.JSON: 'cost_item_fields' list missing/empty")
    return [str(x) for x in fields]


def discover_vendor_files(vendor_dir: str, exclude_prefixes: Optional[List[str]] = None) -> Dict[str, List[str]]:
    """
    Map vendor invoice number -> list of candidate xlsx paths found in vendor_dir.

    Discovery is filename-based. If you need content-based discovery, extend this function.
    """
    exclude_prefixes = [p.upper() for p in (exclude_prefixes or [])]

    mapping: Dict[str, List[str]] = {}
    patterns = [os.path.join(vendor_dir, "*.xlsx"), os.path.join(vendor_dir, "*.xlsm")]
    for pat in patterns:
        for fp in glob.glob(pat):
            base = os.path.basename(fp)
            if any(base.upper().startswith(pref) for pref in exclude_prefixes):
                continue
            invs = extract_invs_from_filename(base)
            for inv in invs:
                mapping.setdefault(inv, []).append(fp)
    return mapping


def _vendor_total(vseries: pd.Series, amount_fields: List[str]) -> Optional[float]:
    if "Total_Amount_AED" in vseries.index:
        return _num(vseries["Total_Amount_AED"])
    found = False
    s = 0.0
    for f in amount_fields:
        if f in vseries.index:
            found = True
            s += _num(vseries[f])
    return s if found else None


def compare_sheet_row_to_vendor(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_row: int,
    header_to_col: Dict[Any, int],
    vseries: pd.Series,
    cost_fields: List[str],
    tol_abs: float = 0.01,
) -> Tuple[bool, List[str]]:
    """
    Exact-ish match check for DONE status.
    Compares all 88 cost item fields + Total_Amount_AED (if present) within tol_abs.
    """
    mismatches: List[str] = []
    for f in cost_fields:
        col = header_to_col.get(f)
        if col is None:
            continue
        sheet_val = _num(ws.cell(row=sheet_row, column=col).value)
        vend_val = _num(vseries.get(f, 0))
        if abs(sheet_val - vend_val) > tol_abs:
            mismatches.append(f"{f}: sheet={sheet_val} vendor={vend_val}")

    if "Total_Amount_AED" in header_to_col and "Total_Amount_AED" in vseries.index:
        s_total = _num(ws.cell(row=sheet_row, column=header_to_col["Total_Amount_AED"]).value)
        v_total = _num(vseries["Total_Amount_AED"])
        if abs(s_total - v_total) > tol_abs:
            mismatches.append(f"Total_Amount_AED: sheet={s_total} vendor={v_total}")

    return (len(mismatches) == 0, mismatches)


def generate_status_report(
    ofco_excel_path: str,
    vendor_dir: str,
    cost_fields_json_path: str,
    out_path: str,
    sheet_name: str = "Sheet1",
    tol_pct: float = 0.02,
    tol_abs: float = 0.01,
    exclude_vendor_prefixes: Optional[List[str]] = None,
) -> str:
    cost_fields = load_cost_item_fields(cost_fields_json_path)
    amount_fields = [f for f in cost_fields if f.endswith("_AMOUNT")]

    # OFCO workbook
    wb = openpyxl.load_workbook(ofco_excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Found: {wb.sheetnames}")
    ws = wb[sheet_name]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    if "SUBJECT" not in header_to_col:
        raise ValueError("Column 'SUBJECT' not found in header row")
    if "Total_Amount_AED" not in header_to_col:
        raise ValueError("Column 'Total_Amount_AED' not found in header row")

    subject_col = header_to_col["SUBJECT"]
    total_col = header_to_col["Total_Amount_AED"]

    # cost item amount columns in sheet (for diagnostics)
    amount_cols = {f: header_to_col.get(f) for f in amount_fields}

    # Vendor file discovery
    vendor_map = discover_vendor_files(vendor_dir, exclude_prefixes=exclude_vendor_prefixes or ["OFCO-INV-"])

    records: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        subj = ws.cell(row=r, column=subject_col).value
        invs = extract_invs_from_subject(str(subj) if subj else "")
        if not invs:
            continue

        # If multiple invoice numbers in one row, we emit one record per invoice number.
        for inv in invs:
            sheet_total = _num(ws.cell(row=r, column=total_col).value)

            # diagnostics: how many amount fields are non-zero in the current Sheet1 row
            nz_cnt = 0
            nz_sum = 0.0
            nz_fields: List[str] = []
            for f, cidx in amount_cols.items():
                if cidx is None:
                    continue
                v = _num(ws.cell(row=r, column=cidx).value)
                if abs(v) > 1e-9:
                    nz_cnt += 1
                    nz_sum += v
                    nz_fields.append(f)

            rec: Dict[str, Any] = {
                "vendor_invoice_no": inv,
                "sheet_row": r,
                "subject": str(subj),
                "sheet_total_aed": sheet_total,
                "sheet_nonzero_amount_fields": nz_cnt,
                "sheet_nonzero_amount_sum": nz_sum,
                "sheet_nonzero_amount_top_fields": ", ".join(nz_fields[:4]) + ("..." if len(nz_fields) > 4 else ""),
                "vendor_file": None,
                "vendor_total_aed": None,
                "diff_aed": None,
                "diff_pct": None,
                "status": None,
                "reason": None,
            }

            vendor_files = vendor_map.get(inv, [])
            if not vendor_files:
                rec["status"] = "PENDING"
                rec["reason"] = "vendor_file_not_found"
                records.append(rec)
                continue

            if len(vendor_files) > 1:
                rec["status"] = "MISMATCH"
                rec["reason"] = f"multiple_vendor_files({len(vendor_files)})"
                rec["vendor_file"] = ";".join(os.path.basename(f) for f in vendor_files)
                records.append(rec)
                continue

            vf = vendor_files[0]
            rec["vendor_file"] = os.path.basename(vf)

            try:
                vdf = pd.read_excel(vf, sheet_name=0)
                if vdf.empty:
                    raise ValueError("vendor_xlsx_empty")
                vseries = vdf.iloc[0]

                vtotal = _vendor_total(vseries, amount_fields)
                rec["vendor_total_aed"] = vtotal

                if vtotal is None:
                    rec["status"] = "MISMATCH"
                    rec["reason"] = "vendor_total_missing"
                    records.append(rec)
                    continue

                diff = vtotal - sheet_total
                rec["diff_aed"] = diff
                rec["diff_pct"] = (diff / sheet_total) if sheet_total else None

                tol = max(tol_abs, abs(sheet_total) * tol_pct)
                if abs(diff) > tol:
                    rec["status"] = "MISMATCH"
                    rec["reason"] = f"total_diff_exceeds_tol({tol:.2f})"
                    records.append(rec)
                    continue

                match_ok, mism = compare_sheet_row_to_vendor(
                    ws=ws,
                    sheet_row=r,
                    header_to_col=header_to_col,
                    vseries=vseries,
                    cost_fields=cost_fields,
                    tol_abs=tol_abs,
                )

                if match_ok:
                    rec["status"] = "DONE"
                    rec["reason"] = "sheet_matches_vendor"
                else:
                    rec["status"] = "READY"
                    rec["reason"] = "vendor_file_ready_sheet_not_applied"
                    rec["mismatch_fields_count"] = len(mism)

                records.append(rec)

            except Exception as e:
                rec["status"] = "MISMATCH"
                rec["reason"] = f"vendor_parse_error:{type(e).__name__}:{e}"
                records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        raise RuntimeError("No vendor invoice numbers found in Sheet1 SUBJECT")

    # Report write
    ts = _dt.datetime.now().isoformat(timespec="seconds")
    summary = pd.DataFrame(
        [
            {"key": "ofco_excel", "value": os.path.basename(ofco_excel_path)},
            {"key": "sheet", "value": sheet_name},
            {"key": "vendor_dir", "value": vendor_dir},
            {"key": "timestamp", "value": ts},
            {"key": "tolerance_pct", "value": tol_pct},
            {"key": "tolerance_abs_aed", "value": tol_abs},
            {"key": "total_vendor_invoices_found_in_sheet", "value": int(len(df))},
            {"key": "DONE", "value": int((df["status"] == "DONE").sum())},
            {"key": "READY", "value": int((df["status"] == "READY").sum())},
            {"key": "PENDING", "value": int((df["status"] == "PENDING").sum())},
            {"key": "MISMATCH", "value": int((df["status"] == "MISMATCH").sum())},
        ]
    )

    df_sorted = df.sort_values(["status", "vendor_invoice_no", "sheet_row"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="SUMMARY", index=False)
        df_sorted.to_excel(writer, sheet_name="ALL", index=False)
        for status in ["READY", "PENDING", "MISMATCH", "DONE"]:
            df_sorted[df_sorted["status"] == status].to_excel(writer, sheet_name=status, index=False)

    # Basic formatting
    wb_rep = openpyxl.load_workbook(out_path)
    for wsname in wb_rep.sheetnames:
        w = wb_rep[wsname]
        w.freeze_panes = "A2"
    wb_rep.save(out_path)

    return out_path


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="OFCO Vendor Invoice READY/PENDING/MISMATCH report generator")
    p.add_argument("--ofco", required=True, help="Path to OFCO workbook (xlsx)")
    p.add_argument("--vendor_dir", required=True, help="Directory containing vendor invoice parsed xlsx files")
    p.add_argument("--cost_fields", required=True, help="Path to cost_item_fields.JSON")
    p.add_argument("--out", required=True, help="Output report xlsx path")
    p.add_argument("--sheet", default="Sheet1", help="OFCO sheet name (default: Sheet1)")
    p.add_argument("--tol_pct", type=float, default=0.02, help="Relative tolerance (default: 0.02 = 2%)")
    p.add_argument("--tol_abs", type=float, default=0.01, help="Absolute tolerance in AED (default: 0.01)")
    return p


def main() -> None:
    args = _build_arg_parser().parse_args()
    out = generate_status_report(
        ofco_excel_path=args.ofco,
        vendor_dir=args.vendor_dir,
        cost_fields_json_path=args.cost_fields,
        out_path=args.out,
        sheet_name=args.sheet,
        tol_pct=args.tol_pct,
        tol_abs=args.tol_abs,
        exclude_vendor_prefixes=["OFCO-INV-"],
    )
    print(f"Report written: {out}")


if __name__ == "__main__":
    main()
