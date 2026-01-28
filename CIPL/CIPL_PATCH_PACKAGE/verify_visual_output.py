#!/usr/bin/env python3
"""
시각적 출력 검증 스크립트
- CIPL_PATCH_PACKAGE 버전 (기준)과 최적화된 CIPL 버전 비교
- 동일한 입력 데이터로 두 버전 생성 후 비교
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 경로 설정
PACKAGE_DIR = Path(__file__).parent
ROOT_DIR = PACKAGE_DIR.parent.parent
CIPL_DIR = ROOT_DIR / "CIPL"

sys.path.insert(0, str(ROOT_DIR))
sys.path.insert(0, str(PACKAGE_DIR))
sys.path.insert(0, str(CIPL_DIR))

def load_test_data():
    """테스트 데이터 로드"""
    json_path = PACKAGE_DIR / "voyage_input_sample_full.json"
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)

def generate_package_version(data):
    """CIPL_PATCH_PACKAGE 버전 생성 (기준)"""
    print("[BASELINE] Generating CIPL_PATCH_PACKAGE version...")
    
    # make_cipl_set.py 사용
    import subprocess
    import importlib.util
    
    # make_cipl_set 모듈 로드
    make_cipl_path = PACKAGE_DIR / "make_cipl_set.py"
    spec = importlib.util.spec_from_file_location("make_cipl_set", make_cipl_path)
    make_cipl = importlib.util.module_from_spec(spec)
    sys.modules["make_cipl_set"] = make_cipl
    spec.loader.exec_module(make_cipl)
    
    # 임시 JSON 파일 생성
    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        temp_json = f.name
    
    try:
        # Excel 생성
        output_path = PACKAGE_DIR / "VERIFY_BASELINE.xlsx"
        result = subprocess.run(
            [sys.executable, str(make_cipl_path), "--in", temp_json, "--out", str(output_path)],
            cwd=str(PACKAGE_DIR),
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            print(f"[ERROR] Error generating baseline: {result.stderr}")
            return None
        
        print(f"[OK] Baseline generated: {output_path}")
        return output_path
    finally:
        Path(temp_json).unlink()

def generate_optimized_version(data):
    """최적화된 CIPL 버전 생성"""
    print("[OPTIMIZED] Generating optimized CIPL version...")
    
    # CIPL.py 매퍼 사용 (commons 모드)
    cipL_path = PACKAGE_DIR / "CIPL.py"
    if not cipL_path.exists():
        print(f"[ERROR] CIPL.py not found at {cipL_path}")
        return None
    
    import importlib.util
    spec = importlib.util.spec_from_file_location("cipl_mapper", cipL_path)
    cipL_mapper = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cipL_mapper)
    
    # 데이터 변환
    page_data = cipL_mapper.make_4page_data_dicts(data)
    
    # 각 페이지 모듈 로드
    ci_p1_path = CIPL_DIR / "COMMERCIAL INVOICE.PY"
    pl_p1_path = CIPL_DIR / "PACKING LIST.PY"
    
    if not ci_p1_path.exists() or not pl_p1_path.exists():
        print(f"[ERROR] Optimized modules not found")
        return None
    
    # 모듈 로드
    spec_ci = importlib.util.spec_from_file_location("ci_p1", ci_p1_path)
    mod_ci = importlib.util.module_from_spec(spec_ci)
    spec_ci.loader.exec_module(mod_ci)
    
    spec_pl = importlib.util.spec_from_file_location("pl_p1", pl_p1_path)
    mod_pl = importlib.util.module_from_spec(spec_pl)
    spec_pl.loader.exec_module(mod_pl)
    
    # Workbook 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # Commercial Invoice
    ws_ci = wb.create_sheet("Commercial_Invoice")
    if hasattr(mod_ci, "build_invoice_p1"):
        mod_ci.build_invoice_p1(ws_ci)
    else:
        print("[ERROR] build_invoice_p1 not found")
        return None
    
    # Packing List
    ws_pl = wb.create_sheet("Packing_List")
    if hasattr(mod_pl, "build_packing_list"):
        # 데이터 변환 필요
        pl_data = page_data.get("pl_p1", {})
        mod_pl.build_packing_list(ws_pl, pl_data)
    else:
        print("[ERROR] build_packing_list not found")
        return None
    
    output_path = PACKAGE_DIR / "VERIFY_OPTIMIZED.xlsx"
    wb.save(output_path)
    print(f"[OK] Optimized version generated: {output_path}")
    return output_path

def compare_cells(ws1, ws2, cell_refs):
    """셀 비교"""
    differences = []
    for cell_ref in cell_refs:
        try:
            c1 = ws1[cell_ref]
            c2 = ws2[cell_ref]
            
            diff = {
                "cell": cell_ref,
                "value": c1.value != c2.value,
                "font": str(c1.font) != str(c2.font),
                "alignment": str(c1.alignment) != str(c2.alignment),
                "border": str(c1.border) != str(c2.border),
                "fill": str(c1.fill) != str(c2.fill),
            }
            
            if any(diff.values()):
                differences.append(diff)
        except Exception as e:
            differences.append({"cell": cell_ref, "error": str(e)})
    
    return differences

def main():
    print("=" * 60)
    print("시각적 출력 검증 시작")
    print("=" * 60)
    
    # 테스트 데이터 로드
    data = load_test_data()
    print(f"[OK] Test data loaded")
    
    # 기준 버전 생성
    baseline_path = generate_package_version(data)
    if not baseline_path or not baseline_path.exists():
        print("[ERROR] Failed to generate baseline")
        return 1
    
    # 최적화 버전 생성
    optimized_path = generate_optimized_version(data)
    if not optimized_path or not optimized_path.exists():
        print("[ERROR] Failed to generate optimized version")
        return 1
    
    # 비교
    print("\n" + "=" * 60)
    print("파일 비교 중...")
    print("=" * 60)
    
    wb_baseline = load_workbook(baseline_path)
    wb_optimized = load_workbook(optimized_path)
    
    # Commercial Invoice 비교
    if "Commercial_Invoice" in wb_baseline.sheetnames and "Commercial_Invoice" in wb_optimized.sheetnames:
        ws_baseline = wb_baseline["Commercial_Invoice"]
        ws_optimized = wb_optimized["Commercial_Invoice"]
        
        # 주요 셀 비교
        key_cells = ["B3", "A5", "B5", "F5", "E24", "C29", "J4"]
        diffs = compare_cells(ws_baseline, ws_optimized, key_cells)
        
        if diffs:
            print(f"\n[WARNING] Commercial Invoice 차이점 발견: {len(diffs)}개")
            for diff in diffs:
                print(f"  {diff}")
        else:
            print("[OK] Commercial Invoice: 주요 셀 일치")
    
    # Packing List 비교
    if "Packing_List" in wb_baseline.sheetnames and "Packing_List" in wb_optimized.sheetnames:
        ws_baseline = wb_baseline["Packing_List"]
        ws_optimized = wb_optimized["Packing_List"]
        
        key_cells = ["B3", "A5", "B5", "F5", "E24", "C29"]
        diffs = compare_cells(ws_baseline, ws_optimized, key_cells)
        
        if diffs:
            print(f"\n[WARNING] Packing List 차이점 발견: {len(diffs)}개")
            for diff in diffs:
                print(f"  {diff}")
        else:
            print("[OK] Packing List: 주요 셀 일치")
    
    print("\n" + "=" * 60)
    print("검증 완료")
    print("=" * 60)
    print(f"기준 버전: {baseline_path}")
    print(f"최적화 버전: {optimized_path}")
    print("\n[TIP] Excel에서 두 파일을 열어 시각적으로 비교하세요.")
    
    return 0

if __name__ == "__main__":
    sys.exit(main())

