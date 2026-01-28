# Py3.11.8
"""
테스트: jpt71_refresh_export_final.py의 날짜 필터링 로직 검증

테스트 케이스:
1. 모든 날짜가 비어있을 때 (NaT/None) → "No valid dates found" 메시지
2. 일부 날짜만 있을 때 → 유효한 날짜만 사용
3. 정상적인 날짜 데이터 → 정상 동작
"""

import sys
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook

# 테스트할 함수들을 import
from jpt71_refresh_export_final import _to_date, build_cross_gantt


def test_to_date_function():
    """_to_date 함수의 다양한 입력 케이스 테스트"""
    print("=" * 60)
    print("테스트 1: _to_date 함수 검증")
    print("=" * 60)
    
    test_cases = [
        (None, None, "None 입력"),
        ("", None, "빈 문자열"),
        (datetime(2024, 1, 15), date(2024, 1, 15), "datetime 객체"),
        (date(2024, 1, 15), date(2024, 1, 15), "date 객체"),
        ("2024-01-15", date(2024, 1, 15), "문자열 날짜"),
        ("2024/01/15", date(2024, 1, 15), "슬래시 구분 날짜"),
        (pd.NaT, None, "pandas NaT"),
        ("invalid", None, "잘못된 형식"),
    ]
    
    passed = 0
    failed = 0
    
    for input_val, expected, description in test_cases:
        result = _to_date(input_val)
        if result == expected:
            print(f"✓ {description}: {input_val} → {result}")
            passed += 1
        else:
            print(f"✗ {description}: {input_val} → {result} (예상: {expected})")
            failed += 1
    
    print(f"\n결과: {passed}개 통과, {failed}개 실패\n")
    return failed == 0


def test_empty_dates():
    """모든 날짜가 비어있을 때 테스트"""
    print("=" * 60)
    print("테스트 2: 모든 날짜가 비어있는 경우")
    print("=" * 60)
    
    # 빈 날짜를 가진 DataFrame 생성
    df = pd.DataFrame({
        "Trip": ["Trip-1", "Trip-2"],
        "Type": ["Agg", "Deb"],
        "Material": ["Material1", "Material2"],
        "Plan_MW4_Depart_Agg": [pd.NaT, pd.NaT],
        "Plan_AGI_Offload_Agg": [None, None],
        "Plan_AGI_Debris_Load": [pd.NaT, None],
        "Plan_MW4_Debris_Offload": ["", ""],
    })
    
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    try:
        build_cross_gantt(wb, df_plan=df)
        
        # Cross_Gantt 시트 확인
        if "Cross_Gantt" in wb.sheetnames:
            ws = wb["Cross_Gantt"]
            cell_a1 = ws["A1"].value
            if cell_a1 == "No valid dates found in Plan sheet.":
                print("✓ 빈 날짜 처리 성공: A1에 메시지 출력")
                print(f"  메시지: {cell_a1}")
                return True
            else:
                print(f"✗ 예상 메시지 없음. A1 값: {cell_a1}")
                return False
        else:
            print("✗ Cross_Gantt 시트가 생성되지 않음")
            return False
    except Exception as e:
        print(f"✗ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_partial_dates():
    """일부 날짜만 있을 때 테스트"""
    print("=" * 60)
    print("테스트 3: 일부 날짜만 있는 경우")
    print("=" * 60)
    
    # 일부만 날짜가 있는 DataFrame
    df = pd.DataFrame({
        "Trip": ["Trip-1", "Trip-2", "Trip-3"],
        "Type": ["Agg", "Deb", "Agg"],
        "Material": ["Material1", "Material2", "Material3"],
        "Plan_MW4_Depart_Agg": [date(2024, 1, 15), pd.NaT, date(2024, 1, 20)],
        "Plan_AGI_Offload_Agg": [pd.NaT, date(2024, 1, 18), None],
        "Plan_AGI_Debris_Load": [None, date(2024, 1, 19), pd.NaT],
        "Plan_MW4_Debris_Offload": [pd.NaT, None, date(2024, 1, 25)],
    })
    
    wb = Workbook()
    wb.remove(wb.active)
    
    try:
        build_cross_gantt(wb, df_plan=df)
        
        if "Cross_Gantt" in wb.sheetnames:
            ws = wb["Cross_Gantt"]
            cell_a1 = ws["A1"].value
            
            # 날짜 헤더가 생성되었는지 확인 (날짜 컬럼은 대략 10번째 컬럼 이후)
            has_dates = False
            for col in range(10, 20):
                cell = ws.cell(1, col)
                if cell.value and isinstance(cell.value, date):
                    has_dates = True
                    break
            
            if has_dates and cell_a1 != "No valid dates found in Plan sheet.":
                print("✓ 부분 날짜 처리 성공: 유효한 날짜만 사용하여 Gantt 생성")
                print(f"  최소 날짜: 2024-01-15, 최대 날짜: 2024-01-25")
                return True
            else:
                print(f"✗ 날짜 헤더가 생성되지 않음. A1: {cell_a1}")
                return False
        else:
            print("✗ Cross_Gantt 시트가 생성되지 않음")
            return False
    except Exception as e:
        print(f"✗ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_normal_dates():
    """정상적인 날짜 데이터 테스트"""
    print("=" * 60)
    print("테스트 4: 정상적인 날짜 데이터")
    print("=" * 60)
    
    # 모든 날짜가 정상인 DataFrame
    df = pd.DataFrame({
        "Trip": ["Trip-1", "Trip-2"],
        "Type": ["Agg", "Deb"],
        "Material": ["Material1", "Material2"],
        "Plan_MW4_Depart_Agg": [date(2024, 1, 15), date(2024, 1, 20)],
        "Plan_AGI_Offload_Agg": [date(2024, 1, 18), date(2024, 1, 23)],
        "Plan_AGI_Debris_Load": [date(2024, 1, 19), date(2024, 1, 24)],
        "Plan_MW4_Debris_Offload": [date(2024, 1, 25), date(2024, 1, 28)],
    })
    
    wb = Workbook()
    wb.remove(wb.active)
    
    try:
        build_cross_gantt(wb, df_plan=df)
        
        if "Cross_Gantt" in wb.sheetnames:
            ws = wb["Cross_Gantt"]
            cell_a1 = ws["A1"].value
            
            # 첫 번째 날짜 헤더 찾기 (대략 10번째 컬럼 이후)
            first_date = None
            for col in range(10, 30):
                cell = ws.cell(1, col)
                if cell.value and isinstance(cell.value, date):
                    first_date = cell.value
                    break
            
            if first_date:
                print("✓ 정상 날짜 처리 성공: Gantt 생성 완료")
                print(f"  첫 번째 날짜 헤더: {first_date}")
                return True
            else:
                print(f"✗ 날짜 헤더가 생성되지 않음. A1: {cell_a1}")
                return False
        else:
            print("✗ Cross_Gantt 시트가 생성되지 않음")
            return False
    except Exception as e:
        print(f"✗ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_date_filtering_logic():
    """날짜 필터링 로직 직접 테스트"""
    print("=" * 60)
    print("테스트 5: 날짜 필터링 로직 직접 검증")
    print("=" * 60)
    
    # 다양한 날짜 타입 혼합
    test_data = [
        date(2024, 1, 15),
        pd.NaT,
        None,
        "",
        datetime(2024, 1, 20),
        "2024-01-25",
        pd.NaT,
        None,
    ]
    
    all_dates = []
    for d in test_data:
        if pd.isna(d) or d is None:
            continue
        converted_date = _to_date(d)
        if converted_date is not None:
            all_dates.append(converted_date)
    
    if not all_dates:
        print("✗ 날짜 필터링 실패: 유효한 날짜가 없음")
        return False
    
    expected_dates = [date(2024, 1, 15), date(2024, 1, 20), date(2024, 1, 25)]
    
    if len(all_dates) == len(expected_dates) and all_dates == expected_dates:
        print("✓ 날짜 필터링 로직 정상 작동")
        print(f"  입력: {len(test_data)}개 항목")
        print(f"  필터링 후: {len(all_dates)}개 유효한 날짜")
        print(f"  날짜 범위: {min(all_dates)} ~ {max(all_dates)}")
        return True
    else:
        print(f"✗ 날짜 필터링 결과 불일치")
        print(f"  예상: {expected_dates}")
        print(f"  결과: {all_dates}")
        return False


def main():
    """모든 테스트 실행"""
    print("\n" + "=" * 60)
    print("jpt71_refresh_export_final.py 날짜 필터링 테스트")
    print("=" * 60 + "\n")
    
    results = []
    
    # 테스트 실행
    results.append(("_to_date 함수", test_to_date_function()))
    results.append(("빈 날짜 처리", test_empty_dates()))
    results.append(("부분 날짜 처리", test_partial_dates()))
    results.append(("정상 날짜 처리", test_normal_dates()))
    results.append(("날짜 필터링 로직", test_date_filtering_logic()))
    
    # 결과 요약
    print("\n" + "=" * 60)
    print("테스트 결과 요약")
    print("=" * 60)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "✓ 통과" if result else "✗ 실패"
        print(f"{status}: {test_name}")
    
    print(f"\n총 {total}개 테스트 중 {passed}개 통과, {total - passed}개 실패")
    
    if passed == total:
        print("\n🎉 모든 테스트 통과!")
        return 0
    else:
        print("\n⚠️  일부 테스트 실패. 위의 오류를 확인하세요.")
        return 1


if __name__ == "__main__":
    sys.exit(main())

