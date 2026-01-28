#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OUTLOOK_HVDC_ALL_rev.xlsx 파일 구조 분석
"""

import sys
import io
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import json
from collections import defaultdict

# Windows 콘솔 인코딩 설정
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def analyze_excel_structure(file_path: str):
    """Excel 파일의 구조를 상세 분석"""
    file_path = Path(file_path)
    
    if not file_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return
    
    print("=" * 80)
    print(f"[Excel 파일 구조 분석] {file_path.name}")
    print("=" * 80)
    
    # openpyxl로 시트 정보 확인
    wb = load_workbook(file_path, data_only=False, read_only=True)
    
    analysis_result = {
        'file_name': file_path.name,
        'total_sheets': len(wb.sheetnames),
        'sheets': []
    }
    
    print(f"\n[총 시트 수] {len(wb.sheetnames)}\n")
    
    # 각 시트 분석
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"[시트] {sheet_name}")
        print(f"{'='*80}")
        
        ws = wb[sheet_name]
        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'dimensions': f"{ws.max_row}행 x {ws.max_column}열"
        }
        
        print(f"크기: {sheet_info['dimensions']}")
        
        # pandas로 데이터 읽기 (처음 20행)
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=20, header=None)
            
            # 헤더 추정 (첫 행이 헤더일 가능성)
            if len(df) > 0:
                print(f"\n[샘플 데이터] (처음 5행):")
                print(df.head().to_string())
                
                # 컬럼별 데이터 타입 및 샘플 값
                print(f"\n[컬럼 정보] (처음 20개):")
                for idx, col in enumerate(df.columns[:20]):
                    col_data = df[col].dropna()
                    if len(col_data) > 0:
                        sample_val = col_data.iloc[0]
                        data_type = type(sample_val).__name__
                        col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
                        print(f"  컬럼 {idx+1} ({col_letter}): {str(sample_val)[:50]} ... [{data_type}]")
                
                sheet_info['sample_data'] = df.head(10).to_dict('records')
                sheet_info['column_count'] = len(df.columns)
        except Exception as e:
            print(f"[경고] 데이터 읽기 오류: {e}")
            sheet_info['error'] = str(e)
        
        # 수식 개수 확인 (처음 1000개 셀만)
        formula_count = 0
        for row in ws.iter_rows(max_row=min(100, ws.max_row), max_col=min(50, ws.max_column)):
            for cell in row:
                if cell.data_type == 'f':
                    formula_count += 1
        
        if formula_count > 0:
            print(f"\n[발견된 수식] {formula_count}개 (샘플링)")
            sheet_info['formula_count'] = formula_count
        
        analysis_result['sheets'].append(sheet_info)
    
    wb.close()
    
    # JSON으로 저장
    output_json = file_path.parent / f"{file_path.stem}_analysis.json"
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(analysis_result, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n\n[분석 완료]")
    print(f"[결과 저장] {output_json}")
    
    return analysis_result

def get_detailed_column_info(file_path: str, sheet_name: str, max_rows: int = 100):
    """특정 시트의 상세 컬럼 정보"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows)
        
        print(f"\n{'='*80}")
        print(f"[상세 컬럼 정보] {sheet_name}")
        print(f"{'='*80}\n")
        
        column_info = []
        for col in df.columns:
            col_data = df[col].dropna()
            info = {
                'column_name': str(col),
                'non_null_count': len(col_data),
                'null_count': len(df) - len(col_data),
                'data_type': str(df[col].dtype),
                'sample_values': col_data.head(5).tolist() if len(col_data) > 0 else []
            }
            column_info.append(info)
            
            print(f"컬럼: {col}")
            print(f"  - 데이터 타입: {info['data_type']}")
            print(f"  - 비어있지 않은 값: {info['non_null_count']}/{len(df)}")
            if info['sample_values']:
                print(f"  - 샘플 값: {info['sample_values'][:3]}")
            print()
        
        return column_info
    except Exception as e:
        print(f"❌ 오류: {e}")
        return None

if __name__ == "__main__":
    file_path = r"C:\Users\SAMSUNG\Downloads\CONVERT\OUTLOOK_HVDC_ALL_rev.xlsx"
    
    # 기본 분석
    result = analyze_excel_structure(file_path)
    
    # 첫 번째 시트 상세 분석
    if result and result['sheets']:
        first_sheet = result['sheets'][0]['name']
        print(f"\n\n{'='*80}")
        print(f"[첫 번째 시트 상세 분석] {first_sheet}")
        print(f"{'='*80}")
        get_detailed_column_info(file_path, first_sheet, max_rows=50)
