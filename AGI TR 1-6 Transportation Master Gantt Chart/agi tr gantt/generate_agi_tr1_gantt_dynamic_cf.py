#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI TR#1 (TM-63) | LCT BUSHRA | Tide-based Dynamic Gantt Schedule Generator (Excel)

Output:
  1) Inputs_Assumptions
  2) Tide_Jan12_to_16 (hourly, GST, Chart Datum)
  3) Schedule_Gantt (task table + 2-hour slot Dynamic Gantt via Conditional Formatting)
  4) FailSafe_Log (for VBA logging)

Key point (Dynamic):
  - Gantt bars are produced by Excel Conditional Formatting formulas, not by "painting cells".
    => If you edit Start/End in Excel, Gantt updates immediately without rerunning Python.
  - Tide@Start / Tide@End are Excel formulas (VLOOKUP + FLOOR to hour).
  - Tide Status uses the editable TIDE_THRESHOLD named range.

Dependencies:
  pip install openpyxl

Run example:
  python generate_agi_tr1_gantt_dynamic_cf.py --out AGI_TR1_BUSHRA_Gantt_DynamicCF_20260112.xlsx --max_rows 200
"""

from __future__ import annotations

import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# Tide data (GST, Chart Datum) — Jan 12–16, 2026 (hourly)
# Source: provided tide sheet (image). Values rounded to 2 decimals.
# =============================================================================
TIDE_DATA: List[Tuple[str, float]] = [
    ("2026-01-12 00:00", 0.99),
    ("2026-01-12 01:00", 0.99),
    ("2026-01-12 02:00", 1.07),
    ("2026-01-12 03:00", 1.20),
    ("2026-01-12 04:00", 1.38),
    ("2026-01-12 05:00", 1.55),
    ("2026-01-12 06:00", 1.70),
    ("2026-01-12 07:00", 1.78),
    ("2026-01-12 08:00", 1.80),
    ("2026-01-12 09:00", 1.73),
    ("2026-01-12 10:00", 1.60),
    ("2026-01-12 11:00", 1.44),
    ("2026-01-12 12:00", 1.27),
    ("2026-01-12 13:00", 1.13),
    ("2026-01-12 14:00", 1.04),
    ("2026-01-12 15:00", 1.00),
    ("2026-01-12 16:00", 1.02),
    ("2026-01-12 17:00", 1.07),
    ("2026-01-12 18:00", 1.13),
    ("2026-01-12 19:00", 1.18),
    ("2026-01-12 20:00", 1.20),
    ("2026-01-12 21:00", 1.19),
    ("2026-01-12 22:00", 1.15),
    ("2026-01-12 23:00", 1.10),
    ("2026-01-13 00:00", 1.06),
    ("2026-01-13 01:00", 1.06),
    ("2026-01-13 02:00", 1.10),
    ("2026-01-13 03:00", 1.20),
    ("2026-01-13 04:00", 1.34),
    ("2026-01-13 05:00", 1.50),
    ("2026-01-13 06:00", 1.65),
    ("2026-01-13 07:00", 1.77),
    ("2026-01-13 08:00", 1.83),
    ("2026-01-13 09:00", 1.82),
    ("2026-01-13 10:00", 1.74),
    ("2026-01-13 11:00", 1.60),
    ("2026-01-13 12:00", 1.43),
    ("2026-01-13 13:00", 1.25),
    ("2026-01-13 14:00", 1.10),
    ("2026-01-13 15:00", 0.99),
    ("2026-01-13 16:00", 0.94),
    ("2026-01-13 17:00", 0.93),
    ("2026-01-13 18:00", 0.96),
    ("2026-01-13 19:00", 1.02),
    ("2026-01-13 20:00", 1.07),
    ("2026-01-13 21:00", 1.11),
    ("2026-01-13 22:00", 1.12),  # interpolated
    ("2026-01-13 23:00", 1.12),  # interpolated
    ("2026-01-14 00:00", 1.13),  # interpolated
    ("2026-01-14 01:00", 1.13),  # interpolated
    ("2026-01-14 02:00", 1.14),
    ("2026-01-14 03:00", 1.21),
    ("2026-01-14 04:00", 1.31),
    ("2026-01-14 05:00", 1.44),
    ("2026-01-14 06:00", 1.58),
    ("2026-01-14 07:00", 1.72),
    ("2026-01-14 08:00", 1.82),
    ("2026-01-14 09:00", 1.87),
    ("2026-01-14 10:00", 1.84),
    ("2026-01-14 11:00", 1.75),
    ("2026-01-14 12:00", 1.60),
    ("2026-01-14 13:00", 1.43),
    ("2026-01-14 14:00", 1.26),
    ("2026-01-14 15:00", 1.11),
    ("2026-01-14 16:00", 1.03),
    ("2026-01-14 17:00", 0.84),
    ("2026-01-14 18:00", 0.83),
    ("2026-01-14 19:00", 0.70),
    ("2026-01-14 20:00", 0.69),
    ("2026-01-14 21:00", 0.76),
    ("2026-01-14 22:00", 0.87),
    ("2026-01-14 23:00", 0.97),
    ("2026-01-15 00:00", 1.00),
    ("2026-01-15 01:00", 1.04),
    ("2026-01-15 02:00", 1.13),
    ("2026-01-15 03:00", 1.25),
    ("2026-01-15 04:00", 1.37),
    ("2026-01-15 05:00", 1.49),
    ("2026-01-15 06:00", 1.61),
    ("2026-01-15 07:00", 1.71),
    ("2026-01-15 08:00", 1.78),
    ("2026-01-15 09:00", 1.80),
    ("2026-01-15 10:00", 1.77),
    ("2026-01-15 11:00", 1.69),
    ("2026-01-15 12:00", 1.56),
    ("2026-01-15 13:00", 1.40),
    ("2026-01-15 14:00", 1.23),
    ("2026-01-15 15:00", 1.09),
    ("2026-01-15 16:00", 1.01),
    ("2026-01-15 17:00", 0.94),
    ("2026-01-15 18:00", 0.91),
    ("2026-01-15 19:00", 0.74),
    ("2026-01-15 20:00", 0.74),
    ("2026-01-15 21:00", 0.83),
    ("2026-01-15 22:00", 0.94),
    ("2026-01-15 23:00", 1.02),
    ("2026-01-16 00:00", 1.06),
    ("2026-01-16 01:00", 1.12),
    ("2026-01-16 02:00", 1.22),
    ("2026-01-16 03:00", 1.33),
    ("2026-01-16 04:00", 1.44),
    ("2026-01-16 05:00", 1.54),
    ("2026-01-16 06:00", 1.63),
    ("2026-01-16 07:00", 1.70),
    ("2026-01-16 08:00", 1.73),
    ("2026-01-16 09:00", 1.74),
    ("2026-01-16 10:00", 1.71),
    ("2026-01-16 11:00", 1.64),
    ("2026-01-16 12:00", 1.55),
    ("2026-01-16 13:00", 1.53),
    ("2026-01-16 14:00", 1.55),
    ("2026-01-16 15:00", 1.31),
    ("2026-01-16 16:00", 1.06),
    ("2026-01-16 17:00", 0.84),
    ("2026-01-16 18:00", 0.69),
    ("2026-01-16 19:00", 0.63),
    ("2026-01-16 20:00", 0.65),
    ("2026-01-16 21:00", 0.74),
    ("2026-01-16 22:00", 0.88),
    ("2026-01-16 23:00", 1.02),
]


# =============================================================================
# Styles
# =============================================================================
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_NORM = Font(name="Calibri", size=11)

FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HDR = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="D9D9D9")
FILL_INPUT = PatternFill("solid", fgColor="D9E1F2")
FILL_NOTE = PatternFill("solid", fgColor="FFF2CC")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_WEEKEND = PatternFill("solid", fgColor="F2F2F2")
FILL_LOW = PatternFill("solid", fgColor="F8CBAD")
FILL_OK = PatternFill("solid", fgColor="C6E0B4")

THIN = Side(style="thin", color="C0C0C0")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

PHASE_COLORS = {
    "PREP": "9DC3E6",
    "RORO": "F8CBAD",
    "SEAFAST": "D9D9D9",
    "SAIL": "C6E0B4",
    "PORT": "BDD7EE",
    "LAND": "E2EFDA",
    "BUFFER": "FFF2CC",
}


# =============================================================================
# Helpers
# =============================================================================
def hide_grid(ws) -> None:
    ws.sheet_view.showGridLines = False


def set_col_widths(ws, widths: Dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = float(w)


def write_header(ws, row: int, headers: List[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN


def parse_dt(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d %H:%M")


def build_slots(start: datetime, end: datetime, step_hours: float) -> List[datetime]:
    slots = []
    cur = start
    step = timedelta(hours=step_hours)
    while cur <= end:
        slots.append(cur)
        cur += step
    return slots


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="AGI_TR1_BUSHRA_Gantt_DynamicCF_20260112.xlsx", help="Output .xlsx path")
    ap.add_argument("--max_rows", default=200, type=int, help="Max task rows preformatted (template capacity)")
    args = ap.parse_args()

    # =============================================================================
    # Base inputs (editable later inside the Excel file)
    # =============================================================================
    timeline_start = parse_dt("2026-01-12 00:00")
    timeline_end = parse_dt("2026-01-16 23:00")
    slot_hours = 2.00
    tide_threshold = 1.60

    # =============================================================================
    # Task plan (TR#1 only) — edit here if needed
    # =============================================================================
    tasks: List[Dict[str, Any]] = [
        dict(ID="T01", Phase="PREP", Activity="SPMT mobilization / assembly & function test", Location="Mina Zayed Port", Owner="Mammoet/SCT", Start="2026-01-12 04:00", End="2026-01-12 06:00", TideCritical="N", Status="PLANNED", Notes="Equipment readiness check"),
        dict(ID="T02", Phase="PREP", Activity="Toolbox talk + PTW confirmation + comms check", Location="Mina Zayed Port", Owner="Mammoet/SCT/HSE", Start="2026-01-12 06:00", End="2026-01-12 06:30", TideCritical="N", Status="PLANNED", Notes="Daylight ops start"),
        dict(ID="T03", Phase="RORO", Activity="RoRo LOAD-OUT: TR1 SPMT → LCT BUSHRA deck", Location="Mina Zayed Port", Owner="Mammoet", Start="2026-01-12 06:30", End="2026-01-12 09:30", TideCritical="Y", Status="PLANNED", Notes="Tide window control"),
        dict(ID="T04", Phase="SEAFAST", Activity="Sea fastening / lashing / welding QA", Location="Mina Zayed Port", Owner="Mammoet/OFCO", Start="2026-01-12 09:30", End="2026-01-12 12:30", TideCritical="N", Status="PLANNED", Notes="MWS/QA hold point"),
        dict(ID="T05", Phase="PORT", Activity="Port clearance / pilotage / final docs", Location="Mina Zayed Port", Owner="OFCO/Port", Start="2026-01-12 12:30", End="2026-01-12 14:00", TideCritical="N", Status="PLANNED", Notes="Sail-away readiness"),
        dict(ID="T06", Phase="SAIL", Activity="Sail-away & marine transport (MZP → AGI)", Location="At sea", Owner="LCT Master", Start="2026-01-12 14:00", End="2026-01-13 04:00", TideCritical="N", Status="PLANNED", Notes="Transit time 14.00h (editable)"),
        dict(ID="T07", Phase="PORT", Activity="Arrival / berthing / ramp set-up", Location="Al Ghallan Island (AGI) Jetty", Owner="LCT Master/Mammoet", Start="2026-01-13 04:00", End="2026-01-13 06:00", TideCritical="N", Status="PLANNED", Notes="Mooring double lines as required"),
        dict(ID="T08", Phase="RORO", Activity="RoRo LOAD-IN: TR1 deck → AGI jetty", Location="AGI Jetty", Owner="Mammoet", Start="2026-01-13 06:00", End="2026-01-13 09:00", TideCritical="Y", Status="PLANNED", Notes="Target tide peak around 07-09"),
        dict(ID="T09", Phase="LAND", Activity="Land transport: jetty → installation area", Location="AGI Site", Owner="Mammoet/SCT", Start="2026-01-13 09:00", End="2026-01-13 12:00", TideCritical="N", Status="PLANNED", Notes="Route & ground bearing by client"),
        dict(ID="T10", Phase="LAND", Activity="Set-down / final positioning at laydown", Location="AGI Site", Owner="Mammoet/SCT", Start="2026-01-13 12:00", End="2026-01-13 15:00", TideCritical="N", Status="PLANNED", Notes="Final inspection"),
        dict(ID="T11", Phase="PREP", Activity="Close-out docs / demobilization prep", Location="AGI Site", Owner="SCT/DSV", Start="2026-01-13 15:00", End="2026-01-13 17:00", TideCritical="N", Status="PLANNED", Notes="Evidence pack & sign-off"),
        dict(ID="T12", Phase="BUFFER", Activity="Contingency buffer (tide/weather/port)", Location="AGI/MZP", Owner="SCT", Start="2026-01-13 17:00", End="2026-01-13 21:00", TideCritical="N", Status="PLANNED", Notes="Use if hold / delay"),
    ]

    max_rows = max(args.max_rows, 20)

    # =============================================================================
    # Workbook build
    # =============================================================================
    wb = Workbook()
    wb.remove(wb.active)

    # -------------------------------------------------------------------------
    # Sheet 1: Inputs_Assumptions
    # -------------------------------------------------------------------------
    ws_in = wb.create_sheet("Inputs_Assumptions")
    hide_grid(ws_in)

    ws_in.merge_cells("A1:D1")
    ws_in["A1"] = "AGI TR#1 (TM-63) | LCT BUSHRA | Tide-based Schedule (Dynamic Gantt)"
    ws_in["A1"].fill = FILL_TITLE
    ws_in["A1"].font = FONT_TITLE
    ws_in["A1"].alignment = ALIGN_LEFT
    ws_in.row_dimensions[1].height = 24

    ws_in["A2"] = "Blue cells are editable inputs. Gantt bars update automatically via Conditional Formatting."
    ws_in["A2"].fill = FILL_NOTE
    ws_in["A2"].alignment = ALIGN_LEFT

    write_header(ws_in, 3, ["Parameter", "Value", "Unit", "Notes"])
    ws_in.row_dimensions[3].height = 20

    inputs_rows = [
        ("Project", "HVDC – Al Ghallan Island (AGI)", "", "Ref: Mammoet MS for TM-63"),
        ("Vessel", "LCT BUSHRA", "", ""),
        ("Cargo", "Transformer TM-63 (TR#1)", "", ""),
        ("Cargo Weight (t)", 217.00, "t", "From MS (TM-63)"),
        ("Route", "Mina Zayed Port → Al Ghallan Island (AGI)", "", ""),
        ("Timezone", "GST (UTC+4)", "", ""),
        ("Daylight window", "06:00–18:00", "", "Assumption: RoRo/land ops daylight"),
        ("RoRo wind limit", 20.00, "kt", "Go/No-Go threshold"),
        ("RoRo Hs limit", 0.60, "m", "Significant wave height"),
        ("Tide datum", "CD", "", "Chart Datum"),
        ("Tide threshold (for RoRo)", tide_threshold, "m", "Used in Tide Status check"),
        ("Gantt slot hours", slot_hours, "h", "2.00h = 12 columns/day"),
        ("Gantt timeline start", timeline_start, "GST", "Controls Gantt header row"),
        ("Gantt timeline end", timeline_end, "GST", "Controls Gantt header row"),
        ("Transit time (MZP→AGI)", 14.00, "h", "Editable assumption"),
        ("Buffer policy", "Include 4.00h contingency", "", "Weather/tide/port hold"),
    ]

    param_row_map: Dict[str, int] = {}
    for r, (k, v, u, note) in enumerate(inputs_rows, start=4):
        param_row_map[k] = r

        c1 = ws_in.cell(r, 1, k)
        c1.fill = FILL_SECTION
        c1.font = FONT_BOLD
        c1.alignment = ALIGN_LEFT
        c1.border = BORDER_THIN

        c2 = ws_in.cell(r, 2, v)
        c2.fill = FILL_INPUT
        c2.alignment = ALIGN_LEFT
        c2.border = BORDER_THIN
        if isinstance(v, (int, float)):
            c2.number_format = "0.00"
        if isinstance(v, datetime):
            c2.number_format = "yyyy-mm-dd hh:mm"

        c3 = ws_in.cell(r, 3, u)
        c3.alignment = ALIGN_CENTER
        c3.border = BORDER_THIN

        c4 = ws_in.cell(r, 4, note)
        c4.alignment = ALIGN_LEFT
        c4.border = BORDER_THIN

    set_col_widths(ws_in, {"A": 30, "B": 40, "C": 10, "D": 55})
    ws_in.freeze_panes = "A4"

    row_tide = param_row_map["Tide threshold (for RoRo)"]
    row_slot = param_row_map["Gantt slot hours"]
    row_gstart = param_row_map["Gantt timeline start"]
    row_gend = param_row_map["Gantt timeline end"]

    wb.defined_names.add(DefinedName("TIDE_THRESHOLD", attr_text=f"'Inputs_Assumptions'!$B${row_tide}"))
    wb.defined_names.add(DefinedName("GANTT_SLOT_HOURS", attr_text=f"'Inputs_Assumptions'!$B${row_slot}"))
    wb.defined_names.add(DefinedName("GANTT_START", attr_text=f"'Inputs_Assumptions'!$B${row_gstart}"))
    wb.defined_names.add(DefinedName("GANTT_END", attr_text=f"'Inputs_Assumptions'!$B${row_gend}"))

    # -------------------------------------------------------------------------
    # Sheet 2: Tide_Jan12_to_16
    # -------------------------------------------------------------------------
    ws_t = wb.create_sheet("Tide_Jan12_to_16")
    hide_grid(ws_t)

    ws_t.merge_cells("A1:E1")
    ws_t["A1"] = "Water Tide (GST) | Mina Zayed (CD) | Jan 12–16, 2026"
    ws_t["A1"].fill = FILL_TITLE
    ws_t["A1"].font = FONT_TITLE
    ws_t["A1"].alignment = ALIGN_LEFT
    ws_t.row_dimensions[1].height = 22

    ws_t["A2"] = "Hourly tide values (rounded to 2 decimals). Update this table if official tide sheet changes."
    ws_t["A2"].fill = FILL_NOTE
    ws_t["A2"].alignment = ALIGN_LEFT

    write_header(ws_t, 4, ["DateTime (GST)", "Tide_m (CD)", "Source", "QC", "Notes"])
    ws_t.freeze_panes = "A5"
    set_col_widths(ws_t, {"A": 22, "B": 14, "C": 24, "D": 8, "E": 40})

    row_idx = 5
    for dt_str, val in TIDE_DATA:
        dtv = parse_dt(dt_str)
        ws_t.cell(row_idx, 1, dtv).number_format = "yyyy-mm-dd hh:mm"
        ws_t.cell(row_idx, 2, float(val)).number_format = "0.00"
        ws_t.cell(row_idx, 3, "Provided (image)")
        ws_t.cell(row_idx, 4, "OK")
        ws_t.cell(row_idx, 5, "")
        for col in range(1, 6):
            cell = ws_t.cell(row_idx, col)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER if col in (1, 2, 4) else ALIGN_LEFT
            cell.fill = FILL_INPUT if col in (1, 2) else FILL_WHITE
        row_idx += 1

    tide_last_row = row_idx - 1
    tide_lookup_range = f"'Tide_Jan12_to_16'!$A$5:$B${tide_last_row}"

    # -------------------------------------------------------------------------
    # Sheet 3: Schedule_Gantt
    # -------------------------------------------------------------------------
    ws_g = wb.create_sheet("Schedule_Gantt")
    hide_grid(ws_g)

    slots = build_slots(timeline_start, timeline_end, slot_hours)
    gantt_start_col = 15  # O
    task_cols = 14
    total_cols = task_cols + len(slots)

    ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws_g.cell(1, 1).value = "TR#1 Master Schedule (Mina Zayed → AGI) | Dynamic Gantt (2h slots)"
    ws_g.cell(1, 1).fill = FILL_TITLE
    ws_g.cell(1, 1).font = FONT_TITLE
    ws_g.cell(1, 1).alignment = ALIGN_LEFT
    ws_g.row_dimensions[1].height = 22

    # Legends
    ws_g["A2"] = "Legend:"
    ws_g["A2"].font = FONT_BOLD
    legend_items = [
        ("PREP", "Prep & Close-out"),
        ("RORO", "RoRo ops (tide-critical)"),
        ("SEAFAST", "Sea fastening"),
        ("PORT", "Port/berth"),
        ("SAIL", "Marine transit"),
        ("LAND", "Inland move"),
        ("BUFFER", "Buffer"),
    ]
    col = 2
    for code, label in legend_items:
        c = ws_g.cell(2, col, code)
        c.fill = PatternFill("solid", fgColor=PHASE_COLORS[code])
        c.border = BORDER_THIN
        c.alignment = ALIGN_CENTER
        c.font = FONT_BOLD
        ws_g.cell(3, col, label).alignment = ALIGN_CENTER
        ws_g.cell(3, col).font = Font(size=9)
        col += 1

    ws_g["I2"] = "Status:"
    ws_g["I2"].font = FONT_BOLD
    ws_g["I2"].alignment = ALIGN_LEFT
    status_items = [
        ("PLANNED", "Planned", "D9D9D9"),
        ("INPROG", "In Progress", "FFF2CC"),
        ("DONE", "Done", "C6E0B4"),
        ("HOLD", "Hold", "F8CBAD"),
    ]
    start_col_status = 10  # J
    for i, (code, label, color) in enumerate(status_items):
        ccol = start_col_status + i
        cell_code = ws_g.cell(2, ccol, code)
        cell_code.fill = PatternFill("solid", fgColor=color)
        cell_code.border = BORDER_THIN
        cell_code.font = FONT_BOLD
        cell_code.alignment = ALIGN_CENTER
        cell_lbl = ws_g.cell(3, ccol, label)
        cell_lbl.font = Font(size=9)
        cell_lbl.alignment = ALIGN_CENTER

    ws_g.row_dimensions[2].height = 18
    ws_g.row_dimensions[3].height = 18

    # Task header row 5
    task_headers = ["ID", "Phase", "Activity", "Location", "Owner",
                    "Start(GST)", "End(GST)", "Dur(h)",
                    "Tide@Start(m)", "Tide@End(m)",
                    "Tide-Critical", "Tide Status", "Status", "Notes"]
    write_header(ws_g, 5, task_headers)

    set_col_widths(ws_g, {
        "A": 6, "B": 10, "C": 44, "D": 22, "E": 16, "F": 18, "G": 18,
        "H": 8, "I": 12, "J": 12, "K": 12, "L": 10, "M": 10, "N": 28
    })

    # Timeline header row 5 formulas
    for j in range(len(slots)):
        col_idx = gantt_start_col + j
        cell = ws_g.cell(5, col_idx)
        if j == 0:
            cell.value = "=GANTT_START"
        else:
            prev = ws_g.cell(5, col_idx - 1).coordinate
            cell.value = f"={prev}+GANTT_SLOT_HOURS/24"
        cell.number_format = "mm-dd hh:mm"
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        ws_g.column_dimensions[get_column_letter(col_idx)].width = 6.0

    # Timeline day header row 4 merges
    day_header_row = 4
    current_date = None
    group_start_col = gantt_start_col
    for j, dtv in enumerate(slots):
        d = dtv.date()
        col_idx = gantt_start_col + j
        if current_date is None:
            current_date = d
            group_start_col = col_idx
        elif d != current_date:
            ws_g.merge_cells(start_row=day_header_row, start_column=group_start_col,
                             end_row=day_header_row, end_column=col_idx - 1)
            c = ws_g.cell(day_header_row, group_start_col, current_date.strftime("%Y-%m-%d"))
            c.fill = FILL_SECTION
            c.font = FONT_BOLD
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
            for cc in range(group_start_col, col_idx):
                ws_g.cell(day_header_row, cc).border = BORDER_THIN
                ws_g.cell(day_header_row, cc).fill = FILL_SECTION
            current_date = d
            group_start_col = col_idx

    last_col_idx = gantt_start_col + len(slots) - 1
    ws_g.merge_cells(start_row=day_header_row, start_column=group_start_col,
                     end_row=day_header_row, end_column=last_col_idx)
    c = ws_g.cell(day_header_row, group_start_col, current_date.strftime("%Y-%m-%d"))
    c.fill = FILL_SECTION
    c.font = FONT_BOLD
    c.alignment = ALIGN_CENTER
    c.border = BORDER_THIN
    for cc in range(group_start_col, last_col_idx + 1):
        ws_g.cell(day_header_row, cc).border = BORDER_THIN
        ws_g.cell(day_header_row, cc).fill = FILL_SECTION

    ws_g.row_dimensions[4].height = 18
    ws_g.row_dimensions[5].height = 20

    # Prefill task rows (A..N) up to max_rows with formulas and styles
    task_start_row = 6
    max_row = max_rows

    for i in range(max_row - task_start_row + 1):
        r = task_start_row + i
        t = tasks[i] if i < len(tasks) else None

        if t:
            ws_g.cell(r, 1, t["ID"]).fill = FILL_SECTION
            ws_g.cell(r, 1).font = FONT_BOLD
            ws_g.cell(r, 1).alignment = ALIGN_CENTER

            ws_g.cell(r, 2, t["Phase"]).fill = PatternFill("solid", fgColor=PHASE_COLORS.get(t["Phase"], "FFFFFF"))
            ws_g.cell(r, 2).alignment = ALIGN_CENTER

            ws_g.cell(r, 3, t["Activity"]).alignment = ALIGN_LEFT
            ws_g.cell(r, 4, t["Location"]).alignment = ALIGN_LEFT
            ws_g.cell(r, 5, t["Owner"]).alignment = ALIGN_LEFT

            ws_g.cell(r, 6, parse_dt(t["Start"]))
            ws_g.cell(r, 7, parse_dt(t["End"]))
            ws_g.cell(r, 11, t.get("TideCritical", "N"))
            ws_g.cell(r, 13, t.get("Status", "PLANNED"))
            ws_g.cell(r, 14, t.get("Notes", ""))

        # Date formats
        ws_g.cell(r, 6).number_format = "yyyy-mm-dd hh:mm"
        ws_g.cell(r, 7).number_format = "yyyy-mm-dd hh:mm"
        ws_g.cell(r, 6).alignment = ALIGN_CENTER
        ws_g.cell(r, 7).alignment = ALIGN_CENTER

        # Formulas (IF-wrapped)
        ws_g.cell(r, 8, f'=IF(OR($F{r}="", $G{r}=""),"",ROUND((G{r}-F{r})*24,2))').number_format = "0.00"
        ws_g.cell(r, 8).alignment = ALIGN_CENTER

        ws_g.cell(r, 9, f'=IF($F{r}="","",IFERROR(VLOOKUP(FLOOR(F{r},1/24),{tide_lookup_range},2,TRUE),""))').number_format = "0.00"
        ws_g.cell(r, 9).alignment = ALIGN_CENTER

        ws_g.cell(r, 10, f'=IF($G{r}="","",IFERROR(VLOOKUP(FLOOR(G{r},1/24),{tide_lookup_range},2,TRUE),""))').number_format = "0.00"
        ws_g.cell(r, 10).alignment = ALIGN_CENTER

        ws_g.cell(r, 12, f'=IF($K{r}="Y",IF(AND($I{r}>=TIDE_THRESHOLD,$J{r}>=TIDE_THRESHOLD),"OK","LOW"),"")').alignment = ALIGN_CENTER
        ws_g.cell(r, 11).alignment = ALIGN_CENTER
        ws_g.cell(r, 13).alignment = ALIGN_CENTER
        ws_g.cell(r, 14).alignment = ALIGN_LEFT

        # Borders + base fills
        for c in range(1, 15):
            cell = ws_g.cell(r, c)
            cell.border = BORDER_THIN
            if c not in (1, 2):
                cell.fill = FILL_WHITE

        # Highlight tide-critical prefilled rows only
        if t and t.get("TideCritical") == "Y":
            for c in range(1, 15):
                if c not in (1, 2):
                    ws_g.cell(r, c).fill = FILL_NOTE

        ws_g.row_dimensions[r].height = 26

    # Prefill Gantt grid area (O..end) for all template rows
    gantt_col_end = gantt_start_col + len(slots) - 1
    gantt_range = f"{get_column_letter(gantt_start_col)}{task_start_row}:{get_column_letter(gantt_col_end)}{max_row}"

    for r in range(task_start_row, max_row + 1):
        for c in range(gantt_start_col, gantt_col_end + 1):
            cell = ws_g.cell(r, c)
            cell.border = BORDER_THIN
            cell.fill = FILL_WHITE
            cell.alignment = ALIGN_CENTER

    # Conditional formatting: Gantt bars by Phase
    for phase, color in PHASE_COLORS.items():
        fill = PatternFill("solid", fgColor=color)
        # Overlap test: [slot_start, slot_end) overlaps [task_start, task_end)
        formula = f'=AND(O$5 + GANTT_SLOT_HOURS/24 > $F6, O$5 < $G6, $B6="{phase}")'
        ws_g.conditional_formatting.add(gantt_range, FormulaRule(formula=[formula], fill=fill))

    # Weekend shading on timeline header row
    header_range = f"{get_column_letter(gantt_start_col)}5:{get_column_letter(gantt_col_end)}5"
    ws_g.conditional_formatting.add(header_range, FormulaRule(formula=[f"=WEEKDAY(O$5,2)>5"], fill=FILL_WEEKEND))

    # Tide Status highlight
    ws_g.conditional_formatting.add(f"L{task_start_row}:L{max_row}", FormulaRule(formula=[f'=$L{task_start_row}="LOW"'], fill=FILL_LOW))
    ws_g.conditional_formatting.add(f"L{task_start_row}:L{max_row}", FormulaRule(formula=[f'=$L{task_start_row}="OK"'], fill=FILL_OK))

    # Status highlight
    ws_g.conditional_formatting.add(f"M{task_start_row}:M{max_row}", FormulaRule(formula=[f'=$M{task_start_row}="DONE"'], fill=FILL_OK))
    ws_g.conditional_formatting.add(f"M{task_start_row}:M{max_row}", FormulaRule(formula=[f'=$M{task_start_row}="HOLD"'], fill=FILL_LOW))
    ws_g.conditional_formatting.add(f"M{task_start_row}:M{max_row}", FormulaRule(formula=[f'=$M{task_start_row}="INPROG"'], fill=FILL_NOTE))

    # Data validation lists
    phase_list = ",".join(PHASE_COLORS.keys())
    dv_phase = DataValidation(type="list", formula1=f'"{phase_list}"', allow_blank=True)
    ws_g.add_data_validation(dv_phase)
    dv_phase.add(f"B{task_start_row}:B{max_row}")

    dv_tidecrit = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws_g.add_data_validation(dv_tidecrit)
    dv_tidecrit.add(f"K{task_start_row}:K{max_row}")

    dv_status = DataValidation(type="list", formula1='"PLANNED,INPROG,DONE,HOLD"', allow_blank=True)
    ws_g.add_data_validation(dv_status)
    dv_status.add(f"M{task_start_row}:M{max_row}")

    # Filters & freeze panes
    ws_g.auto_filter.ref = f"A5:N{max_row}"
    ws_g.freeze_panes = ws_g.cell(task_start_row, gantt_start_col).coordinate  # O6

    # Print settings
    ws_g.page_setup.orientation = "landscape"
    ws_g.page_setup.fitToWidth = 1
    ws_g.page_setup.fitToHeight = 0
    ws_g.print_options.horizontalCentered = True

    # -------------------------------------------------------------------------
    # Sheet 4: FailSafe_Log
    # -------------------------------------------------------------------------
    ws_log = wb.create_sheet("FailSafe_Log")
    hide_grid(ws_log)
    ws_log.merge_cells("A1:G1")
    ws_log["A1"] = "Fail-safe Log (auto-filled by VBA)"
    ws_log["A1"].fill = FILL_TITLE
    ws_log["A1"].font = FONT_TITLE
    ws_log["A1"].alignment = ALIGN_LEFT
    ws_log.row_dimensions[1].height = 22

    write_header(ws_log, 3, ["Timestamp", "Task ID", "Issue", "TideStart", "TideEnd", "Threshold", "Action"])
    set_col_widths(ws_log, {"A": 20, "B": 10, "C": 40, "D": 10, "E": 10, "F": 10, "G": 30})
    ws_log.freeze_panes = "A4"

    out_path = Path(args.out).resolve()
    wb.save(out_path)
    print(f"[OK] Written: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
