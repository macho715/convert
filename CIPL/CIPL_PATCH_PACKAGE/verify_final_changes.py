#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify all final changes in generated CIPL Excel file"""

from openpyxl import load_workbook

wb = load_workbook('CIPL_TEST_FINAL.xlsx')

print("=" * 60)
print("=== COMMERCIAL INVOICE ===")
print("=" * 60)
ws = wb['Commercial_Invoice_P1']

# 1. A25:A26과 B25:B26 사이 테두리 제거 확인
print("\n1. A25:A26과 B25:B26 사이 테두리 확인:")
a25_border = ws['A25'].border
b25_border = ws['B25'].border
print(f"   A25 right border: {a25_border.right.style if a25_border.right.style else 'None (제거됨)'}")
print(f"   B25 left border: {b25_border.left.style if b25_border.left.style else 'None (제거됨)'}")

# 2. B28:G42 박스 테두리 확인
print("\n2. B28:G42 박스 테두리 확인:")
b28_border = ws['B28'].border
print(f"   B28 top border style: {b28_border.top.style if b28_border.top.style else 'None'}")
if b28_border.top.color and hasattr(b28_border.top.color, 'rgb'):
    print(f"   B28 top border color: {b28_border.top.color.rgb}")

# 3. 외곽 테두리 확인
print("\n3. 외곽 테두리 확인:")
a1_border = ws['A1'].border
print(f"   A1 left border style: {a1_border.left.style if a1_border.left.style else 'None'}")
if a1_border.left.color and hasattr(a1_border.left.color, 'rgb'):
    print(f"   A1 left border color: {a1_border.left.color.rgb}")

print("\n" + "=" * 60)
print("=== PACKING LIST ===")
print("=" * 60)
ws2 = wb['Packing_List_P1']

# 1. A24:A25와 B24:B25 사이 테두리 제거 확인
print("\n1. A24:A25와 B24:B25 사이 테두리 확인:")
a24_border = ws2['A24'].border
b24_border = ws2['B24'].border
print(f"   A24 right border: {a24_border.right.style if a24_border.right.style else 'None (제거됨)'}")
print(f"   B24 left border: {b24_border.left.style if b24_border.left.style else 'None (제거됨)'}")

# 2. B28:G42 박스 테두리 확인
print("\n2. B28:G42 박스 테두리 확인:")
b28_border2 = ws2['B28'].border
print(f"   B28 top border style: {b28_border2.top.style if b28_border2.top.style else 'None'}")
if b28_border2.top.color and hasattr(b28_border2.top.color, 'rgb'):
    print(f"   B28 top border color: {b28_border2.top.color.rgb}")

# 3. J7-K7 사이 테두리 확인
print("\n3. J7-K7 사이 테두리 확인:")
j7_border = ws2['J7'].border
k7_border = ws2['K7'].border
print(f"   J7 right border: {j7_border.right.style if j7_border.right.style else 'None'}")
print(f"   K7 left border: {k7_border.left.style if k7_border.left.style else 'None'}")

# 4. 외곽 테두리 확인
print("\n4. 외곽 테두리 확인:")
a1_border2 = ws2['A1'].border
print(f"   A1 left border style: {a1_border2.left.style if a1_border2.left.style else 'None'}")
if a1_border2.left.color and hasattr(a1_border2.left.color, 'rgb'):
    print(f"   A1 left border color: {a1_border2.left.color.rgb}")

print("\n" + "=" * 60)
print("=== CI RIDER ===")
print("=" * 60)
ws3 = wb['CI_Rider_P2']

# 외곽 테두리 확인
print("\n1. 외곽 테두리 확인:")
a1_border3 = ws3['A1'].border
print(f"   A1 left border style: {a1_border3.left.style if a1_border3.left.style else 'None'}")
if a1_border3.left.color and hasattr(a1_border3.left.color, 'rgb'):
    print(f"   A1 left border color: {a1_border3.left.color.rgb}")

print("\n" + "=" * 60)
print("=== PL RIDER ===")
print("=" * 60)
ws4 = wb['PL_Rider_P2']

# 외곽 테두리 확인
print("\n1. 외곽 테두리 확인:")
a1_border4 = ws4['A1'].border
print(f"   A1 left border style: {a1_border4.left.style if a1_border4.left.style else 'None'}")
if a1_border4.left.color and hasattr(a1_border4.left.color, 'rgb'):
    print(f"   A1 left border color: {a1_border4.left.color.rgb}")

print("\n" + "=" * 60)
print("검증 완료!")
print("=" * 60)

