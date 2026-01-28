#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI TR Transportation Schedule Generator (Single Train)
- Scenario A: 7 voyages (1+1+1+1+1+1+1)
- Scenario B: 4 voyages (1+2+2+2) = TR1 / TR2-3 / TR4-5 / TR6-7

Outputs an Excel workbook with:
- Control_Panel
- Schedule_Data
- Gantt_Chart (static coloring)

Usage:
    python agi_tr_schedule_generator_scenarios_start20260115.py --scenario single
    python agi_tr_schedule_generator_scenarios_start20260115.py --scenario batch
"""

import argparse
import datetime as dt
import json
import math
from dataclasses import dataclass
from typing import Dict, List, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# Config (edit as needed)
# -----------------------------
START_DATE = dt.date(2026, 1, 15)
TIDE_THRESHOLD_M = 1.90

# Durations (days) for simplified "Schedule_Data + Gantt" planning
DUR = {
    "MOB": 1,     # Mobilization
    "DECK": 3,    # Deck Prep (one-time)
    "LO": 1,      # Load-out per TR
    "SF": 1,      # Sea Fastening (per voyage)
    "MWS": 1,     # MWS + MPI + final
    "SAIL": 1,    # Sail away
    "UL": 1,      # Unload per TR
    "JD": 1,      # Jack-up event (per batch)
    "RET": 1,     # Return sail
    "BUF": 1,     # Buffer/reset
}

# Tide JSON file (provided by user)
import pathlib
SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
TIDE_JSON_PATH = str(SCRIPT_DIR / "MINA ZAYED PORT WATER TIDE.json")

# Output names
OUT_SINGLE = "AGI_TR_SingleTrain_7x1_Start20260115.xlsx"
OUT_BATCH  = "AGI_TR_Batch_1_2_2_2_Start20260115.xlsx"

# Colors
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "INPUT": "FFFDE7",
}

BORDER_SIDE = Side(style="thin", color="A6A6A6")
def thin_border():
    return Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

def load_tide_df(path: str) -> pd.DataFrame:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.DataFrame(data["tide_records"])
    df["date"] = pd.to_datetime(df["date"]).dt.date
    return df

def next_tide_block(start_date: dt.date, days_needed: int, tide_ok: Set[dt.date], max_date: dt.date) -> dt.date:
    """Earliest date >= start_date with 'days_needed' consecutive days passing tide threshold.
    If beyond max_date, returns start_date (tide TBC)."""
    d = start_date
    if d > max_date:
        return d
    while d <= max_date:
        ok = True
        for i in range(days_needed):
            di = d + dt.timedelta(days=i)
            if di > max_date:
                continue
            if di not in tide_ok:
                ok = False
                break
        if ok:
            return d
        d += dt.timedelta(days=1)
    return start_date

def schedule_simple(start_date: dt.date,
                    voyage_groups: List[List[int]],
                    jackup_after_voyage: Dict[int, int],
                    tide_df: pd.DataFrame,
                    tide_threshold: float) -> pd.DataFrame:
    """Sequential schedule (single train) with tide gating for LO/UL."""
    tide_ok = set(tide_df[tide_df["max_height_m"] >= tide_threshold]["date"])
    max_date = tide_df["date"].max()

    rows = []
    cur = start_date

    # MOB
    rows.append(dict(ID="MOB", WBS="1.0", Task="Mobilization", Phase="MOBILIZATION", Owner="Mammoet",
                     Start=cur, End=cur + dt.timedelta(days=DUR["MOB"]), Duration=DUR["MOB"],
                     Notes="SPMT + Marine equipment"))
    cur = cur + dt.timedelta(days=DUR["MOB"])

    # DECK
    rows.append(dict(ID="PREP", WBS="1.1", Task="Deck Prep", Phase="DECK_PREP", Owner="Mammoet",
                     Start=cur, End=cur + dt.timedelta(days=DUR["DECK"]), Duration=DUR["DECK"],
                     Notes="One-time setup"))
    cur = cur + dt.timedelta(days=DUR["DECK"])

    for i, vg in enumerate(voyage_groups, start=1):
        n = len(vg)
        rows.append(dict(ID=f"V{i}", WBS=f"{i+1}.0",
                         Task=f"Voyage {i}: TR{vg[0] if n==1 else str(vg[0])+'-'+str(vg[-1])} ({n} unit{'s' if n>1 else ''})",
                         Phase="MILESTONE", Owner="All",
                         Start=cur, End=cur, Duration=0, Notes=None))

        # LO (gated)
        lo_d = DUR["LO"] * n
        lo_start = next_tide_block(cur, lo_d, tide_ok, max_date)
        lo_end = lo_start + dt.timedelta(days=lo_d)
        if lo_start <= max_date:
            rec = tide_df[tide_df["date"] == lo_start].iloc[0]
            lo_note = f"MZP High tide {rec['high_tide_window']} (max {rec['max_height_m']}m)"
        else:
            lo_note = "MZP tide TBC"
        rows.append(dict(ID=f"LO{i}", WBS=f"{i+1}.1", Task=f"Load-out TRs {', '.join('TR'+str(x) for x in vg)}",
                         Phase="LOADOUT", Owner="Mammoet",
                         Start=lo_start, End=lo_end, Duration=lo_d, Notes=lo_note))
        cur = lo_end

        # SF
        rows.append(dict(ID=f"SF{i}", WBS=f"{i+1}.2", Task="Sea Fastening",
                         Phase="SEAFAST", Owner="Mammoet",
                         Start=cur, End=cur + dt.timedelta(days=DUR["SF"]), Duration=DUR["SF"], Notes=None))
        cur = cur + dt.timedelta(days=DUR["SF"])

        # MWS
        rows.append(dict(ID=f"MWS{i}", WBS=f"{i+1}.3", Task="MWS + MPI + Final",
                         Phase="BUFFER", Owner="Aries/Captain",
                         Start=cur, End=cur + dt.timedelta(days=DUR["MWS"]), Duration=DUR["MWS"], Notes="Hold point"))
        cur = cur + dt.timedelta(days=DUR["MWS"])

        # SAIL
        rows.append(dict(ID=f"SA{i}", WBS=f"{i+1}.4", Task=f"Sail V{i} MZP→AGI",
                         Phase="SAIL", Owner="LCT",
                         Start=cur, End=cur + dt.timedelta(days=DUR["SAIL"]), Duration=DUR["SAIL"], Notes=None))
        cur = cur + dt.timedelta(days=DUR["SAIL"])

        # UL (gated; AGI tide proxy uses same JSON)
        ul_d = DUR["UL"] * n
        ul_start = next_tide_block(cur, ul_d, tide_ok, max_date)
        ul_end = ul_start + dt.timedelta(days=ul_d)
        if ul_start <= max_date:
            rec = tide_df[tide_df["date"] == ul_start].iloc[0]
            ul_note = f"AGI tide proxy: {rec['high_tide_window']} (max {rec['max_height_m']}m)"
        else:
            ul_note = "AGI tide TBC"
        rows.append(dict(ID=f"UL{i}", WBS=f"{i+1}.5", Task="Unload TRs to AGI",
                         Phase="AGI_UNLOAD", Owner="Mammoet",
                         Start=ul_start, End=ul_end, Duration=ul_d, Notes=ul_note))
        cur = ul_end

        # JD events (optional)
        if i in jackup_after_voyage:
            units = jackup_after_voyage[i]
            rows.append(dict(ID=f"JD{i}", WBS=f"{i+1}.9", Task=f"Jack-up Batch ({units} units)",
                             Phase="JACKDOWN", Owner="Mammoet",
                             Start=cur, End=cur + dt.timedelta(days=DUR["JD"]), Duration=DUR["JD"],
                             Notes="On-site installation"))
            cur = cur + dt.timedelta(days=DUR["JD"])

        # RETURN
        rows.append(dict(ID=f"RET{i}", WBS=f"{i+1}.6", Task="Return AGI→MZP",
                         Phase="RETURN", Owner="LCT",
                         Start=cur, End=cur + dt.timedelta(days=DUR["RET"]), Duration=DUR["RET"], Notes=None))
        cur = cur + dt.timedelta(days=DUR["RET"])

        # BUFFER
        rows.append(dict(ID=f"BUF{i}", WBS=f"{i+1}.7", Task="Buffer / Reset",
                         Phase="BUFFER", Owner="All",
                         Start=cur, End=cur + dt.timedelta(days=DUR["BUF"]), Duration=DUR["BUF"], Notes=None))
        cur = cur + dt.timedelta(days=DUR["BUF"])

    return pd.DataFrame(rows)

def make_workbook(schedule_df: pd.DataFrame, scenario_name: str, pattern_desc: str, start_date: dt.date) -> Workbook:
    wb = Workbook()
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"

    # Header
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = f"AGI TR Transportation - {scenario_name}"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ctrl.row_dimensions[1].height = 28

    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = f"Start date fixed to {start_date.strftime('%Y-%m-%d')} | Tide gate: ≥{TIDE_THRESHOLD_M:.2f}m"
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    ws_ctrl["A2"].alignment = Alignment(horizontal="center")

    ws_ctrl["A4"] = "Project Start:"
    ws_ctrl["A4"].font = Font(bold=True)
    ws_ctrl["B4"] = start_date
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = thin_border()

    ws_ctrl["A5"] = "Scenario:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = scenario_name
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = thin_border()

    ws_ctrl["A6"] = "Pattern:"
    ws_ctrl["A6"].font = Font(bold=True)
    ws_ctrl["B6"] = pattern_desc
    ws_ctrl["B6"].alignment = Alignment(wrap_text=True)
    ws_ctrl["B6"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B6"].border = thin_border()

    end_date = schedule_df["End"].max()
    ws_ctrl["A8"] = "Calculated Finish:"
    ws_ctrl["A8"].font = Font(bold=True)
    ws_ctrl["B8"] = end_date
    ws_ctrl["B8"].number_format = "YYYY-MM-DD"
    ws_ctrl["B8"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B8"].border = thin_border()

    ws_ctrl["A9"] = "Total Days (calendar):"
    ws_ctrl["A9"].font = Font(bold=True)
    ws_ctrl["B9"] = (end_date - start_date).days
    ws_ctrl["B9"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B9"].border = thin_border()

    for col in "ABCDEFGH":
        ws_ctrl.column_dimensions[col].width = 18

    # Schedule_Data
    ws_sched = wb.create_sheet("Schedule_Data")
    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = f"AGI HVDC Transformer Transportation Schedule ({scenario_name})"
    ws_sched["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_sched["A1"].alignment = Alignment(horizontal="center")

    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = pattern_desc
    ws_sched["A2"].font = Font(size=10, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    ws_sched["A2"].alignment = Alignment(horizontal="center")

    headers = ["ID","WBS","Task","Phase","Owner","Start","End","Duration","Notes"]
    header_row = 5
    for c,h in enumerate(headers,1):
        cell = ws_sched.cell(header_row,c,value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    start_row = 6
    for i,row in enumerate(schedule_df.itertuples(index=False), start=start_row):
        ws_sched.cell(i,1,value=row.ID)
        ws_sched.cell(i,2,value=row.WBS)
        ws_sched.cell(i,3,value=row.Task)
        ws_sched.cell(i,4,value=row.Phase)
        ws_sched.cell(i,5,value=row.Owner)
        ws_sched.cell(i,6,value=row.Start)
        ws_sched.cell(i,7,value=row.End)
        ws_sched.cell(i,8,value=row.Duration)
        ws_sched.cell(i,9,value=row.Notes)

        ws_sched.cell(i,6).number_format="YYYY-MM-DD"
        ws_sched.cell(i,7).number_format="YYYY-MM-DD"

        # Phase color
        pc = COLORS.get(row.Phase, "FFFFFF")
        ws_sched.cell(i,4).fill = PatternFill("solid", fgColor=pc)

        for c in range(1,10):
            ws_sched.cell(i,c).border = thin_border()
            ws_sched.cell(i,c).alignment = Alignment(vertical="center")

        if row.Phase == "MILESTONE":
            for c in range(1,10):
                ws_sched.cell(i,c).font = Font(bold=True)

    widths = {"A":10,"B":8,"C":40,"D":12,"E":12,"F":12,"G":12,"H":10,"I":40}
    for col,w in widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"

    # Gantt_Chart (static)
    ws_g = wb.create_sheet("Gantt_Chart")
    date_start = start_date
    date_end = end_date
    num_days = (date_end - date_start).days + 1
    last_col = 7 + num_days

    ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws_g.cell(1,1,value=f"Gantt Chart ({scenario_name})").font = Font(bold=True, size=14, color="FFFFFF")
    ws_g.cell(1,1).fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_g.cell(1,1).alignment = Alignment(horizontal="center")

    meta_headers=["ID","WBS","Task","Phase","Start","End","Dur"]
    for c,h in enumerate(meta_headers,1):
        cell = ws_g.cell(4,c,value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    date_col_start = 8
    for i in range(num_days):
        d = date_start + dt.timedelta(days=i)
        cell = ws_g.cell(4, date_col_start+i, value=d)
        cell.number_format = "D"
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()
        ws_g.column_dimensions[get_column_letter(date_col_start+i)].width = 2.5

    for idx,row in enumerate(schedule_df.itertuples(index=False), start=5):
        ws_g.cell(idx,1,value=row.ID)
        ws_g.cell(idx,2,value=row.WBS)
        ws_g.cell(idx,3,value=row.Task)
        ws_g.cell(idx,4,value=row.Phase)
        ws_g.cell(idx,5,value=row.Start)
        ws_g.cell(idx,6,value=row.End)
        ws_g.cell(idx,7,value=row.Duration)
        ws_g.cell(idx,5).number_format="MM/DD"
        ws_g.cell(idx,6).number_format="MM/DD"

        for c in range(1,8):
            ws_g.cell(idx,c).border = thin_border()
            ws_g.cell(idx,c).alignment = Alignment(vertical="center")

        pc = COLORS.get(row.Phase, "FFFFFF")
        ws_g.cell(idx,4).fill = PatternFill("solid", fgColor=pc)

        if row.Duration == 0:
            # milestone star
            offset = (row.Start - date_start).days
            if 0 <= offset < num_days:
                cell = ws_g.cell(idx, date_col_start+offset, value="★")
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, size=8, color="000000")
                cell.fill = PatternFill("solid", fgColor=COLORS["MILESTONE"])
                cell.border = thin_border()
        else:
            for i in range(num_days):
                d = date_start + dt.timedelta(days=i)
                if d >= row.Start and d < row.End:
                    ws_g.cell(idx, date_col_start+i).fill = PatternFill("solid", fgColor=pc)
                ws_g.cell(idx, date_col_start+i).border = thin_border()

    ws_g.column_dimensions["A"].width = 10
    ws_g.column_dimensions["B"].width = 6
    ws_g.column_dimensions["C"].width = 28
    ws_g.column_dimensions["D"].width = 12
    ws_g.column_dimensions["E"].width = 7
    ws_g.column_dimensions["F"].width = 7
    ws_g.column_dimensions["G"].width = 4
    ws_g.freeze_panes = ws_g.cell(5, date_col_start)

    return wb

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--scenario", choices=["single","batch","both"], default="both",
                    help="single = 7x1, batch = 1+2+2+2, both = generate both")
    args = ap.parse_args()

    tide_df = load_tide_df(TIDE_JSON_PATH)

    if args.scenario in ("single","both"):
        voyage_groups = [[1],[2],[3],[4],[5],[6],[7]]
        jack_map = {2:2, 4:2, 6:2, 7:1}  # notes only
        df = schedule_simple(START_DATE, voyage_groups, jack_map, tide_df, TIDE_THRESHOLD_M)
        wb = make_workbook(df, "Single Train 7x1 (1+1+1+1+1+1+1)",
                           "단일 트레인: 7항차 (TR1~TR7 각 1기) | Jack-up: after V2/V4/V6/V7",
                           START_DATE)
        wb.save(OUT_SINGLE)
        print(f"Saved: {OUT_SINGLE}")

    if args.scenario in ("batch","both"):
        voyage_groups = [[1],[2,3],[4,5],[6,7]]
        jack_map = {2:3, 3:2, 4:2}  # notes only
        df = schedule_simple(START_DATE, voyage_groups, jack_map, tide_df, TIDE_THRESHOLD_M)
        wb = make_workbook(df, "Batch 1+2+2+2 (TR1/2-3/4-5/6-7)",
                           "배치 패턴: 4항차 (TR1 / TR2-3 / TR4-5 / TR6-7) | Jack-up: after V2(3 units), V3(2), V4(2)",
                           START_DATE)
        wb.save(OUT_BATCH)
        print(f"Saved: {OUT_BATCH}")

if __name__ == "__main__":
    main()