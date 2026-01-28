# Py 3.11.8 / pandas 2.33 / openpyxl 3.1.5
"""
TR6 Python Pipeline
- Validates workbook structure
- Produces a simple report (phase summary, status summary, risk flags)
- Writes results into new sheet "PY_REPORT" (or overwrites)
Usage:
  py tr6_pipeline.py --in "AGI_TR6_VBA_Enhanced_TEMPLATE.xlsx" --out "C:\Temp" --log "C:\Temp\tr6_ops.log"
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
from pathlib import Path

import pandas as pd
import openpyxl


S_DATA = "Schedule_Data"
S_TIDE = "Tide_Data"
S_WEATHER = "Weather_Analysis"

DATA_HEADER_ROW = 5  # 1-based
DATA_START_ROW = 6


def log_jsonl(path: str, event: str, **kv) -> None:
    rec = {"ts": dt.datetime.now().isoformat(timespec="seconds"), "event": event, **kv}
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def read_schedule(xlsx: str) -> pd.DataFrame:
    df = pd.read_excel(xlsx, sheet_name=S_DATA, engine="openpyxl", header=DATA_HEADER_ROW - 1, dtype=str)
    # Keep only rows with ID
    if "ID" not in df.columns:
        raise ValueError("Schedule_Data header not found (expected 'ID' column)")
    df = df[df["ID"].notna() & (df["ID"].astype(str).str.strip() != "")]
    # Parse date cols if present
    for c in ("Start", "End"):
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
    for c in ("Offset", "Duration"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def read_tide_risk(xlsx: str) -> dict[dt.date, str]:
    df = pd.read_excel(xlsx, sheet_name=S_TIDE, engine="openpyxl", header=3)
    # header row 4 in sheet -> header=3 (0-based)
    if "Date" not in df.columns or "Risk Level" not in df.columns:
        return {}
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    df["Risk Level"] = df["Risk Level"].astype(str).str.upper()
    return {d: r for d, r in zip(df["Date"], df["Risk Level"]) if pd.notna(d)}


def shamal_window_from_weather_sheet(xlsx: str) -> tuple[dt.date, dt.date] | None:
    # This workbook note says Jan 14-18, 2026; try to detect from cell E4 (Notes) or fall back.
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if S_WEATHER not in wb.sheetnames:
        return None
    ws = wb[S_WEATHER]
    note = str(ws["F4"].value or "")
    # Example: "Peak Shamal: Jan 14-18"
    import re
    m = re.search(r"Jan\s+(\d{1,2})\s*-\s*(\d{1,2})", note)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return dt.date(2026, 1, a), dt.date(2026, 1, b)
    return dt.date(2026, 1, 14), dt.date(2026, 1, 18)


def build_report(df: pd.DataFrame, tide: dict[dt.date, str], shamal: tuple[dt.date, dt.date]) -> dict[str, pd.DataFrame]:
    # Status summary
    status = df.groupby(df.get("Status", pd.Series(["UNKNOWN"] * len(df))).fillna("UNKNOWN")).size().reset_index(name="Count")
    status.columns = ["Status", "Count"]

    # Phase summary
    phase_col = "Phase" if "Phase" in df.columns else None
    if phase_col:
        phase = (
            df.groupby(phase_col)
            .agg(Tasks=("ID", "count"), Start=("Start", "min"), End=("End", "max"))
            .reset_index()
            .rename(columns={phase_col: "Phase"})
        )
    else:
        phase = pd.DataFrame({"Phase": [], "Tasks": [], "Start": [], "End": []})

    # Risk flags
    def tide_flag(d: dt.date | None) -> str:
        if d is None:
            return "UNKNOWN"
        return tide.get(d, "UNKNOWN")

    sh_s, sh_e = shamal
    def shamal_flag(d: dt.date | None) -> str:
        if d is None:
            return "UNKNOWN"
        return "HIGH" if (sh_s <= d <= sh_e) else "LOW"

    df_risk = df.copy()
    df_risk["TideRisk(Start)"] = df_risk["Start"].apply(tide_flag) if "Start" in df_risk.columns else "UNKNOWN"
    df_risk["ShamalRisk(Start)"] = df_risk["Start"].apply(shamal_flag) if "Start" in df_risk.columns else "UNKNOWN"
    # Focus: LOADOUT / AGI_UNLOAD
    if "Phase" in df_risk.columns:
        focus = df_risk[df_risk["Phase"].isin(["LOADOUT", "AGI_UNLOAD"])].copy()
    else:
        focus = df_risk.copy()

    return {"Status": status, "Phase": phase, "Risk_All": df_risk, "Risk_Focus": focus}


def write_report(xlsx_in: str, xlsx_out: str, sheets: dict[str, pd.DataFrame]) -> None:
    # Write to a new workbook (copy) so we don't corrupt original
    Path(xlsx_out).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_in)
    # Remove existing report sheets
    for name in list(wb.sheetnames):
        if name.startswith("PY_"):
            del wb[name]
    # Create report sheets
    for k, df in sheets.items():
        ws = wb.create_sheet(f"PY_{k}")
        # header
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        # basic widths
        for col in ws.columns:
            mx = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(60, max(10, mx + 2))
    wb.save(xlsx_out)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="input xlsx")
    ap.add_argument("--out", dest="out_dir", required=True, help="output directory")
    ap.add_argument("--log", dest="log", required=True, help="log file path (jsonl)")
    args = ap.parse_args()

    inp = args.inp
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / (Path(inp).stem + "_PY_REPORT.xlsx")

    log_jsonl(args.log, "start", inp=inp, out=str(out_xlsx))

    try:
        df = read_schedule(inp)
        tide = read_tide_risk(inp)
        shamal = shamal_window_from_weather_sheet(inp) or (dt.date(2026, 1, 14), dt.date(2026, 1, 18))
        sheets = build_report(df, tide, shamal)
        write_report(inp, str(out_xlsx), sheets)
        log_jsonl(args.log, "done", rows=int(df.shape[0]), out=str(out_xlsx))
        print(f"OK: wrote {out_xlsx}")
        return 0
    except Exception as e:
        log_jsonl(args.log, "error", err=str(e))
        print(f"ERROR: {e}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
