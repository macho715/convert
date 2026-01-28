#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AGI HVDC Transformers (TR1..TR7) – Dynamic Gantt Template Builder (v13)

What you get
------------
Creates an Excel workbook (XLSX) that is usable immediately *without* macros:

  - Inputs                : Start date, trip pattern, installation triggers, durations/buffers.
  - Tide_Peaks_MZP         : Daily tide peak list parsed from provided "MAMMOET_AGI TR.pdf".
  - Weather_Forecast       : Optional sheet to be populated by VBA/PowerQuery (left blank).
  - Calc_PlanA / Calc_PlanB: Formula engine (hidden) to auto-shift schedule when start date changes.
  - Plan_A_Realistic       : Gantt grid (days) + task table (sequential, install inserted in timeline).
  - Plan_B_Fast            : Gantt grid (days) + task table (trip cycle independent from installs).
  - Assumptions_Refs       : Limits from MS + template references.

Key scenario (default)
----------------------
  - Total transformers: 7
  - Trips: 4 (units per trip = [1,2,2,2])
  - Installation triggers (cumulative arrivals): [3,5,7]  => batches [3,2,2]
  - Goal: complete before March (Plan B is designed to hit this with parallel install teams).

Notes
-----
  - This workbook uses formulas; Excel will recalculate dates when you change Inputs!B5.
  - VBA is delivered separately as a .bas module (see AGI_TR7_Automation_v4.bas).
  - Weather averages are handled as a *planning buffer* (Inputs!B22). Optional forecast linkage is
    implemented in VBA using Open‑Meteo.

Dependencies
------------
  pip install openpyxl pandas pymupdf

Run
---
  python agi_tr7_package_builder_v13.py \
    --tide_pdf "MAMMOET_AGI TR.pdf" \
    --out "AGI_TR7_Dynamic_Gantt_Template_v13.xlsx"
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ----------------------------
# Styling constants
# ----------------------------
HEADER_BLUE = "1F4E79"
SUMMARY_GREY = "404040"
WEEKEND_SHADE = "F2F2F2"
TODAY_SHADE = "FFF2CC"

PHASE_COLORS = {
    "MOB": "D9E1F2",
    "PREP": "FCE4D6",
    "PORT": "9DC3E6",
    "MARINE": "BFBFBF",
    "AGI": "C6E0B4",
    "INSTALL": "FFE699",
    "RETURN": "E7E6E6",
    "SUMMARY": SUMMARY_GREY,
    "MILESTONE": "F8CBAD",
}


def thin_border() -> Border:
    side = Side(style="thin", color="A6A6A6")
    return Border(left=side, right=side, top=side, bottom=side)


def excel_dt(d: dt.date) -> dt.datetime:
    return dt.datetime(d.year, d.month, d.day)


def daterange(d0: dt.date, days: int) -> List[dt.date]:
    return [d0 + dt.timedelta(days=i) for i in range(days)]


# ----------------------------
# Tide parsing (daily peaks) from the provided PDF layout
# ----------------------------


def parse_datestr(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%d-%b-%Y").date()


def parse_tide_page(page) -> Optional[pd.DataFrame]:
    """Parse one PDF page into a wide table: date_str x times -> tide_m.

    This parser is tuned for the provided Mammoet tide PDF layout.
    """

    words = page.get_text("words")
    time_words = [w for w in words if re.match(r"^\d{1,2}:\d{2}$", str(w[4]))]
    time_words = sorted(time_words, key=lambda w: w[0])
    if len(time_words) < 8:
        return None
    time_cols = [(w[4], (w[0] + w[2]) / 2) for w in time_words]

    date_words = [w for w in words if re.match(r"^\d{2}-[A-Za-z]{3}-20\d{2}$", str(w[4]))]
    date_words = sorted(date_words, key=lambda w: w[1])
    if not date_words:
        return None

    num_words = [
        w
        for w in words
        if re.match(r"^\d\.\d{2}$", str(w[4])) or re.match(r"^\d\.\d{1}$", str(w[4]))
    ]

    rows = []
    for dw in date_words:
        date_str = str(dw[4])
        y_center = (dw[1] + dw[3]) / 2
        candidates = [w for w in num_words if abs(((w[1] + w[3]) / 2) - y_center) < 3.8]
        if not candidates:
            continue
        values: List[float] = []
        for _, x in time_cols:
            near = [w for w in candidates if abs(((w[0] + w[2]) / 2) - x) < 8.5]
            w = min(near, key=lambda ww: abs(((ww[0] + ww[2]) / 2) - x)) if near else min(
                candidates, key=lambda ww: abs(((ww[0] + ww[2]) / 2) - x)
            )
            try:
                values.append(float(w[4]))
            except Exception:
                values.append(float("nan"))
        rows.append((date_str, values))

    if not rows:
        return None
    df = pd.DataFrame({"date_str": [r[0] for r in rows]})
    for i, (t, _) in enumerate(time_cols):
        df[str(t)] = [r[1][i] for r in rows]
    return df


def daily_peaks(df: pd.DataFrame) -> pd.DataFrame:
    """Convert wide tide table to daily peak (max tide)."""
    time_cols = [c for c in df.columns if c != "date_str"]
    out = []
    for _, row in df.iterrows():
        vals = row[time_cols].astype(float)
        # idxmax returns column label (time)
        out.append((row["date_str"], str(vals.idxmax()), float(vals.max())))
    return pd.DataFrame(out, columns=["date_str", "peak_time", "peak_m"])


def extract_tide_daily_peaks(pdf_path: Path) -> pd.DataFrame:
    """Parse PDF and return daily peak list (Date, Peak time, Peak tide)."""
    doc = fitz.open(str(pdf_path))
    frames: List[pd.DataFrame] = []
    for page in doc:
        df = parse_tide_page(page)
        if df is not None:
            frames.append(df)
    if not frames:
        raise RuntimeError("No tide tables could be parsed from the PDF. Check PDF layout/quality.")
    wide = pd.concat(frames, ignore_index=True)
    peaks = daily_peaks(wide)
    peaks["date"] = peaks["date_str"].map(parse_datestr)
    peaks = peaks[["date", "peak_time", "peak_m"]].sort_values("date")
    # drop duplicates (keep first)
    peaks = peaks.drop_duplicates(subset=["date"], keep="first")
    return peaks.reset_index(drop=True)


# ----------------------------
# Workbook structure + formulas
# ----------------------------


@dataclass(frozen=True)
class RowSpec:
    id: str
    wbs: str
    task: str
    phase: str
    location: str
    owner: str
    start_formula: str
    finish_formula: str
    tide_formula: str
    weather_formula: str
    notes: str


def add_inputs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Inputs")
    ws["A1"].value = "AGI HVDC Transformer Transport – Dynamic Gantt Inputs (TR1..TR7)"
    ws["A1"].font = Font(bold=True, size=14)

    # Scenario & basis
    ws["A3"].value = "Scenario"
    ws["B3"].value = "S1_4Trips_3Inst"
    ws["A4"].value = "Install parallel teams (Plan basis)"
    ws["B4"].value = 2
    ws["A5"].value = "LO Commencement (TR1 LO start @MZP)"
    ws["B5"].value = excel_dt(dt.date(2026, 1, 9))
    ws["A6"].value = "Target complete by"
    ws["B6"].value = excel_dt(dt.date(2026, 3, 1))

    for c in ("A3", "A4", "A5", "A6"):
        ws[c].font = Font(bold=True)
    for c in ("B5", "B6"):
        ws[c].number_format = "dd-mmm-yy"

    # Trip pattern table (supports up to 7 trips; default 4)
    ws["A8"].value = "Trip Pattern"
    ws["A8"].font = Font(bold=True)
    ws["A9"].value = "Trip#"
    ws["B9"].value = "Units"
    ws["D8"].value = "Installation Triggers (cumulative arrivals)"
    ws["D8"].font = Font(bold=True)
    ws["D9"].value = "Inst#"
    ws["E9"].value = "Cumulative arrivals"
    ws["F9"].value = "Batch size (auto)"

    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ("A9", "B9", "D9", "E9", "F9"):
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = thin_border()
    for r in range(10, 17):
        ws[f"A{r}"].value = r - 9
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws[f"A{r}"].border = thin_border()
        ws[f"B{r}"].border = thin_border()
    # Default trips: [1,2,2,2]
    ws["B10"].value = 1
    ws["B11"].value = 2
    ws["B12"].value = 2
    ws["B13"].value = 2

    # Install triggers: [3,5,7] -> batch sizes via formulas
    for i, trig in enumerate([3, 5, 7], start=10):
        inst_idx = i - 9
        ws[f"D{i}"].value = inst_idx
        ws[f"E{i}"].value = trig
        ws[f"D{i}"].alignment = Alignment(horizontal="center")
        ws[f"E{i}"].alignment = Alignment(horizontal="center")
        ws[f"D{i}"].border = thin_border()
        ws[f"E{i}"].border = thin_border()
    # Batch size auto
    ws["F10"].value = '=IF(E10="","",E10)'
    ws["F11"].value = '=IF(E11="","",E11-E10)'
    ws["F12"].value = '=IF(E12="","",E12-E11)'
    for r in range(10, 13):
        ws[f"F{r}"].alignment = Alignment(horizontal="center")
        ws[f"F{r}"].border = thin_border()

    # Durations / buffers
    ws["A18"].value = "Durations / Buffers (days)"
    ws["A18"].font = Font(bold=True)
    ws["A19"].value = "Parameter"
    ws["B19"].value = "Value"
    for cell in ("A19", "B19"):
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = thin_border()

    params = [
        ("Seafastening+MWS (post LO)", 1),
        ("Sail (one-way)", 1),
        ("Weather allowance (avg)", 1),
        ("Berth/unlash prep", 1),
        ("Backshift+Load SPMT", 1),
        ("Return sail", 1),
        ("Reset/readiness", 1),
        ("Install Move+Turn per unit", 4),
        ("Install Jack-down per unit", 1),
    ]
    base_row = 20
    for i, (k, v) in enumerate(params):
        r = base_row + i
        ws[f"A{r}"].value = k
        ws[f"B{r}"].value = v
        ws[f"A{r}"].border = thin_border()
        ws[f"B{r}"].border = thin_border()
        ws[f"B{r}"].alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    # Weather gate config (optional; used by VBA to pull forecast)
    ws["D18"].value = "Weather Gate (Open-Meteo) – optional"
    ws["D18"].font = Font(bold=True)
    ws["D19"].value = "Parameter"
    ws["E19"].value = "Value"
    for cell in ("D19", "E19"):
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = thin_border()

    w_params = [
        ("MZP_lat", 24.50),
        ("MZP_lon", 54.40),
        ("AGI_lat", 24.80),
        ("AGI_lon", 52.80),
        ("Wind max (kn)", 20.00),
        ("Wave max (m)", 0.60),
        ("Timezone", "Asia/Dubai"),
        ("Forecast horizon (d)", 16),
    ]
    r0 = 20
    for i, (k, v) in enumerate(w_params):
        rr = r0 + i
        ws[f"D{rr}"].value = k
        ws[f"E{rr}"].value = v
        ws[f"D{rr}"].border = thin_border()
        ws[f"E{rr}"].border = thin_border()
        ws[f"E{rr}"].alignment = Alignment(horizontal="center")

    # Data validation for scenario
    dv = DataValidation(type="list", formula1='"S1_4Trips_3Inst"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])


def add_tide_sheet(wb: Workbook, peaks: pd.DataFrame, source_label: str) -> None:
    ws = wb.create_sheet("Tide_Peaks_MZP")
    header = ["Date (GST)", "Peak time", "Peak tide (m)", "Source"]
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(1, c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60
    for i, row in peaks.iterrows():
        r = i + 2
        ws.cell(r, 1, value=excel_dt(row["date"])).number_format = "dd-mmm-yy"
        ws.cell(r, 2, value=row["peak_time"])
        ws.cell(r, 3, value=float(row["peak_m"]))
        ws.cell(r, 4, value=source_label)
        for c in range(1, 5):
            ws.cell(r, c).border = thin_border()
            ws.cell(r, c).alignment = Alignment(horizontal="center" if c < 4 else "left", wrap_text=True)


def add_weather_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Weather_Forecast")
    header = [
        "Date",
        "MZP wind max (kn)",
        "MZP wave max (m)",
        "AGI wind max (kn)",
        "AGI wave max (m)",
        "Gate (GO/NO-GO)",
        "Notes",
    ]
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(1, c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    widths = [12, 16, 14, 16, 14, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_calc_sheets(wb: Workbook) -> None:
    """Hidden calc sheets with formulas to compute trip timelines."""
    for name in ("Calc_PlanA", "Calc_PlanB"):
        ws = wb.create_sheet(name)
        ws.sheet_state = "hidden"
        headers = [
            "Trip#",
            "Units",
            "CumArrivals",
            "LO_Start",
            "LO_End",
            "Prep_End",
            "WX_End",
            "Sail_End",
            "Berth_End",
            "LI_End",
            "BatchUnits",
            "Install_End",
            "Backshift_End",
            "Return_End",
            "Reset_End",
            "Trip_Finish",
        ]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(1, c, value=h)
            cell.font = Font(bold=True)
            cell.border = thin_border()
        for r in range(2, 6):
            trip = r - 1
            ws.cell(r, 1, value=trip).border = thin_border()
            # Units from Inputs B10:B13
            ws.cell(r, 2, value=f"=Inputs!$B${9+trip}").border = thin_border()
            # Cum arrivals
            ws.cell(r, 3, value=f"=SUM($B$2:B{r})").border = thin_border()

        # Constants (Inputs addresses)
        # Durations: Inputs B20..B28
        D_SEAF = "Inputs!$B$20"
        D_SAIL = "Inputs!$B$21"
        D_WX = "Inputs!$B$22"
        D_BERTH = "Inputs!$B$23"
        D_BACK = "Inputs!$B$24"
        D_RET = "Inputs!$B$25"
        D_RESET = "Inputs!$B$26"
        D_INST_MT = "Inputs!$B$27"
        D_INST_JD = "Inputs!$B$28"
        INST_TEAMS = "Inputs!$B$4"

        # Trigger values & batch sizes
        TR1 = "Inputs!$E$10"
        TR2 = "Inputs!$E$11"
        TR3 = "Inputs!$E$12"
        B1 = "Inputs!$F$10"
        B2 = "Inputs!$F$11"
        B3 = "Inputs!$F$12"

        def batch_formula(cum_cell: str) -> str:
            # Returns 0 when no trigger matched or when cum is blank.
            return (
                f'=IF({cum_cell}="","",'
                f'IF({cum_cell}={TR1},{B1},'
                f'IF({cum_cell}={TR2},{B2},'
                f'IF({cum_cell}={TR3},{B3},0))))'
            )

        # Formulas per trip
        for r in range(2, 6):
            # LO_Start = first trip uses Inputs!B5; next trips use prior Trip_Finish + 1
            ws.cell(r, 4, value=f'=IF($B{r}="","",IF($A{r}=1,Inputs!$B$5,$P{r-1}+1))')
            ws.cell(r, 5, value=f'=IF($B{r}="","",$D{r}+$B{r}-1)')
            ws.cell(r, 6, value=f'=IF($B{r}="","",$E{r}+{D_SEAF})')
            ws.cell(r, 7, value=f'=IF($B{r}="","",$F{r}+{D_WX})')
            ws.cell(r, 8, value=f'=IF($B{r}="","",$G{r}+{D_SAIL})')
            ws.cell(r, 9, value=f'=IF($B{r}="","",$H{r}+{D_BERTH})')
            ws.cell(r, 10, value=f'=IF($B{r}="","",$I{r}+$B{r})')
            ws.cell(r, 11, value=batch_formula(f"$C{r}"))

        # Install and rest differs between plans
        for r in range(2, 6):
            batch = f"$K{r}"
            per_unit = f"({D_INST_MT}+{D_INST_JD})"
            groups = f"CEILING({batch}/{INST_TEAMS},1)"
            inst_duration = f"{groups}*{per_unit}"
            if name == "Calc_PlanA":
                # Install is inserted in trip timeline (sequential)
                ws.cell(r, 12, value=f"=IF({batch}>0,$J{r}+{inst_duration},$J{r})")
                ws.cell(r, 13, value=f'=IF($B{r}="","",$L{r}+{D_BACK})')
            else:
                # No install in trip; install scheduled separately in Plan_B sheet
                ws.cell(r, 12, value=f"=$J{r}")
                ws.cell(r, 13, value=f'=IF($B{r}="","",$J{r}+{D_BACK})')
            ws.cell(r, 14, value=f'=IF($B{r}="","",$M{r}+{D_RET})')
            ws.cell(r, 15, value=f'=IF($B{r}="","",$N{r}+{D_RESET})')
            ws.cell(r, 16, value=f"=$O{r}")

        # Add borders and formats for all calc cells
        for r in range(2, 6):
            for c in range(1, 17):
                ws.cell(r, c).border = thin_border()
                if c >= 4:
                    ws.cell(r, c).number_format = "dd-mmm-yy"

        # Plan B additional install timing block (decoupled from voyages)
        if name == "Calc_PlanB":
            base_r = 8
            install_headers = [
                "Ready1", "Dur1", "Inst1_Start", "Inst1_End",
                "Ready2", "Dur2", "Inst2_Start", "Inst2_End",
                "Ready3", "Dur3", "Inst3_Start", "Inst3_End",
            ]
            for c, h in enumerate(install_headers, start=1):
                ws.cell(base_r, c, value=h).font = Font(bold=True)
                ws.cell(base_r, c).border = thin_border()

            r = base_r + 1
            # Ready dates (trigger reached) = LI_End + 1 (LI_End is column J)
            ws.cell(r, 1, value='=IFERROR(INDEX($J$2:$J$5, MATCH(Inputs!$E$10, $C$2:$C$5, 0))+1,"")')
            ws.cell(r, 5, value='=IFERROR(INDEX($J$2:$J$5, MATCH(Inputs!$E$11, $C$2:$C$5, 0))+1,"")')
            ws.cell(r, 9, value='=IFERROR(INDEX($J$2:$J$5, MATCH(Inputs!$E$12, $C$2:$C$5, 0))+1,"")')

            # Durations (days)
            ws.cell(r, 2, value=f'=CEILING(Inputs!$F$10/{INST_TEAMS},1)*({D_INST_MT}+{D_INST_JD})')
            ws.cell(r, 6, value=f'=CEILING(Inputs!$F$11/{INST_TEAMS},1)*({D_INST_MT}+{D_INST_JD})')
            ws.cell(r, 10, value=f'=CEILING(Inputs!$F$12/{INST_TEAMS},1)*({D_INST_MT}+{D_INST_JD})')

            # Start/End chains
            ws.cell(r, 3, value='=A9')
            ws.cell(r, 4, value='=IF(C9="","",C9+B9-1)')

            ws.cell(r, 7, value='=IF(E9="","",MAX(E9,D9+1))')
            ws.cell(r, 8, value='=IF(G9="","",G9+F9-1)')

            ws.cell(r, 11, value='=IF(I9="","",MAX(I9,H9+1))')
            ws.cell(r, 12, value='=IF(K9="","",K9+J9-1)')

            for c in range(1, 13):
                ws.cell(r, c).border = thin_border()
                if c in (1, 3, 4, 5, 7, 8, 9, 11, 12):
                    ws.cell(r, c).number_format = 'dd-mmm-yy'


def tide_lookup_formula(date_cell: str) -> str:
    return f'=IFERROR(INDEX(Tide_Peaks_MZP!$C:$C, MATCH({date_cell}, Tide_Peaks_MZP!$A:$A, 0)), "")'


def weather_lookup_formula(date_cell: str) -> str:
    # Weather_Forecast col F = Gate
    return f'=IFERROR(INDEX(Weather_Forecast!$F:$F, MATCH({date_cell}, Weather_Forecast!$A:$A, 0)), "")'


def build_plan_rows(plan: str) -> List[RowSpec]:
    """Task table rows for Plan A or Plan B.

    Dates are formula-driven via Calc sheets.
    """

    calc = "Calc_PlanA" if plan == "A" else "Calc_PlanB"
    owner = "Samsung / Mammoet / LCT Operator"
    loc_mzp = "Mina Zayed Port"
    loc_agi = "AGI Site"
    loc_marine = "MZP↔AGI (Marine)"

    rows: List[RowSpec] = []

    # Campaign summary (computed later by MIN/MAX across plan)
    rows.append(
        RowSpec(
            id="0",
            wbs="0",
            task=f"Campaign – TR1..TR7 | Plan {plan} ({'Realistic' if plan=='A' else 'Fast'})",
            phase="SUMMARY",
            location=f"{loc_mzp}/{loc_agi}",
            owner=owner,
            start_formula="=MIN($G$8:$G$200)",
            finish_formula="=MAX($H$8:$H$200)",
            tide_formula="",
            weather_formula="",
            notes="Auto-summary (min start / max finish).",
        )
    )

    # Mobilization & one-time prep (relative to LO start)
    # MOB: start-3
    rows += [
        RowSpec(
            id="M1",
            wbs="MOB",
            task="Mobilization of SPMTs; Assembly + Function Test",
            phase="MOB",
            location=loc_mzp,
            owner="Mammoet",
            start_formula="=Inputs!$B$5-3",
            finish_formula="=Inputs!$B$5-3",
            tide_formula="",
            weather_formula="AUTO_WEATHER",
            notes="Baseline from Mammoet pattern: equip on-site before LO.",
        ),
        RowSpec(
            id="M2",
            wbs="MOB",
            task="Mobilization of Marine Equipment and Steelworks",
            phase="MOB",
            location=loc_mzp,
            owner="Mammoet",
            start_formula="=Inputs!$B$5-3",
            finish_formula="=Inputs!$B$5-3",
            tide_formula="",
            weather_formula="AUTO_WEATHER",
            notes="Steelworks & ramp-related prep.",
        ),
        RowSpec(
            id="P1",
            wbs="PREP",
            task="Beam Replacement + Deck Preparations (one-time)",
            phase="PREP",
            location=loc_mzp,
            owner="Mammoet + Port Ops",
            start_formula="=Inputs!$B$5-2",
            finish_formula="=Inputs!$B$5-1",
            tide_formula="",
            weather_formula="AUTO_WEATHER",
            notes="From Mammoet pattern (2 days).",
        ),
    ]

    # Trip blocks (4 trips max in calc)
    for trip in range(1, 5):
        r = 1 + trip  # row index in Calc (2..5)
        lo_s = f"={calc}!$D${r}"
        lo_e = f"={calc}!$E${r}"
        prep_e = f"={calc}!$F${r}"
        wx_e = f"={calc}!$G${r}"
        sail_e = f"={calc}!$H${r}"
        berth_e = f"={calc}!$I${r}"
        li_e = f"={calc}!$J${r}"
        inst_end = f"={calc}!$L${r}"
        back_e = f"={calc}!$M${r}"
        ret_e = f"={calc}!$N${r}"
        reset_e = f"={calc}!$O${r}"
        units = f"{calc}!$B${r}"
        batch_units = f"{calc}!$K${r}"

        rows.append(
            RowSpec(
                id=f"T{trip}",
                wbs=f"TRIP{trip}",
                task=f"Trip {trip} – Deliver {trip} (units per Inputs) + return",
                phase="SUMMARY",
                location=f"{loc_mzp}/{loc_agi}",
                owner=owner,
                start_formula=lo_s,
                finish_formula=reset_e,
                tide_formula="",
                weather_formula="",
                notes="Trip summary (computed).",
            )
        )

        # Load-out window (tide gated)
        rows.append(
            RowSpec(
                id=f"{trip}.LO",
                wbs=f"TRIP{trip}.LO",
                task=f"Trip {trip} – Load-out @MZP (RoRo) – {units} unit(s)",
                phase="PORT",
                location=loc_mzp,
                owner="Mammoet + Port Control",
                start_formula=lo_s,
                finish_formula=lo_e,
                tide_formula="AUTO_TIDE",
                weather_formula="AUTO_WEATHER",
                notes="Tide-assisted RoRo window (confirm official tide table).",
            )
        )

        # Seafastening/MWS
        rows.append(
            RowSpec(
                id=f"{trip}.SF",
                wbs=f"TRIP{trip}.SF",
                task=f"Trip {trip} – Seafastening + MWS/MPI + final prep",
                phase="MARINE",
                location=loc_mzp,
                owner="Mammoet + MWS + LCT Crew",
                start_formula=f"={calc}!$E${r}+1",
                finish_formula=prep_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="Pre-sail checks.",
            )
        )

        # Weather allowance
        rows.append(
            RowSpec(
                id=f"{trip}.WX",
                wbs=f"TRIP{trip}.WX",
                task=f"Trip {trip} – Metocean allowance (avg, planning)",
                phase="MARINE",
                location=loc_marine,
                owner=owner,
                start_formula=f"={calc}!$F${r}+1",
                finish_formula=wx_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="Planning buffer; replace with daily forecast windows.",
            )
        )

        # Sail
        rows.append(
            RowSpec(
                id=f"{trip}.SA",
                wbs=f"TRIP{trip}.SA",
                task=f"Trip {trip} – Sail MZP→AGI",
                phase="MARINE",
                location=loc_marine,
                owner="LCT Crew",
                start_formula=f"={calc}!$G${r}+1",
                finish_formula=sail_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="Transit (planning 1 day).",
            )
        )

        # Berth/unlash
        rows.append(
            RowSpec(
                id=f"{trip}.BR",
                wbs=f"TRIP{trip}.BR",
                task=f"Trip {trip} – AGI berthing + unlashing/cutting prep",
                phase="AGI",
                location=loc_agi,
                owner="AGI Port Control + Mammoet",
                start_formula=f"={calc}!$H${r}+1",
                finish_formula=berth_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="",
            )
        )

        # Load-in
        rows.append(
            RowSpec(
                id=f"{trip}.LI",
                wbs=f"TRIP{trip}.LI",
                task=f"Trip {trip} – Load-in @AGI (RoRo) – {units} unit(s)",
                phase="AGI",
                location=loc_agi,
                owner="Mammoet + Client",
                start_formula=f"={calc}!$I${r}+1",
                finish_formula=li_e,
                tide_formula="AUTO_TIDE",
                weather_formula="AUTO_WEATHER",
                notes="Tide-assisted RoRo window.",
            )
        )

        # Installation block indicator
        if plan == "A":
            rows.append(
                RowSpec(
                    id=f"{trip}.IN",
                    wbs=f"TRIP{trip}.INSTALL",
                    task=f"Trip {trip} – Installation batch (if triggered) – {batch_units} unit(s)",
                    phase="INSTALL",
                    location=loc_agi,
                    owner="Mammoet Jacking Crew",
                    start_formula=f"=IF({batch_units}>0,{calc}!$J${r}+1,\"\")",
                    finish_formula=f"=IF({batch_units}>0,{inst_end},\"\")",
                    tide_formula="",
                    weather_formula="",
                    notes="Inserted in timeline (Plan A).",
                )
            )
            after_install = inst_end
        else:
            after_install = li_e

        # Backshift + load on LCT
        rows.append(
            RowSpec(
                id=f"{trip}.BS",
                wbs=f"TRIP{trip}.BS",
                task=f"Trip {trip} – Backshift + Load SPMT on LCT (tide)",
                phase="RETURN",
                location=loc_agi,
                owner="Mammoet + LCT Crew",
                start_formula=f"={after_install}+1",
                finish_formula=back_e,
                tide_formula="AUTO_TIDE",
                weather_formula="AUTO_WEATHER",
                notes="",
            )
        )
        # Return sail
        rows.append(
            RowSpec(
                id=f"{trip}.RT",
                wbs=f"TRIP{trip}.RT",
                task=f"Trip {trip} – Return sail AGI→MZP",
                phase="MARINE",
                location=loc_marine,
                owner="LCT Crew",
                start_formula=f"={calc}!$M${r}+1",
                finish_formula=ret_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="",
            )
        )
        # Reset
        rows.append(
            RowSpec(
                id=f"{trip}.RS",
                wbs=f"TRIP{trip}.RS",
                task=f"Trip {trip} – Reset/readiness for next loading",
                phase="RETURN",
                location=loc_mzp,
                owner="Mammoet",
                start_formula=f"={calc}!$N${r}+1",
                finish_formula=reset_e,
                tide_formula="",
                weather_formula="AUTO_WEATHER",
                notes="",
            )
        )

    # Plan B: explicit installation blocks (independent)
    if plan == "B":
        # Installation timing is computed in hidden Calc_PlanB (row 9, cols C/D, G/H, K/L).
        inst1_s = f"={calc}!$C$9"
        inst1_e = f"={calc}!$D$9"
        inst2_s = f"={calc}!$G$9"
        inst2_e = f"={calc}!$H$9"
        inst3_s = f"={calc}!$K$9"
        inst3_e = f"={calc}!$L$9"

        rows += [
            RowSpec(
                id="I1",
                wbs="INSTALL.1",
                task="Installation Batch 1 (after 3 units arrived)",
                phase="INSTALL",
                location=loc_agi,
                owner="Mammoet Jacking Crew",
                start_formula=inst1_s,
                finish_formula=inst1_e,
                tide_formula="",
                weather_formula="",
                notes="Decoupled from voyages (Plan B).",
            ),
            RowSpec(
                id="I2",
                wbs="INSTALL.2",
                task="Installation Batch 2 (after 5 units arrived)",
                phase="INSTALL",
                location=loc_agi,
                owner="Mammoet Jacking Crew",
                start_formula=inst2_s,
                finish_formula=inst2_e,
                tide_formula="",
                weather_formula="",
                notes="",
            ),
            RowSpec(
                id="I3",
                wbs="INSTALL.3",
                task="Installation Batch 3 (after 7 units arrived)",
                phase="INSTALL",
                location=loc_agi,
                owner="Mammoet Jacking Crew",
                start_formula=inst3_s,
                finish_formula=inst3_e,
                tide_formula="",
                weather_formula="",
                notes="",
            ),
        ]

    # Milestone
    rows.append(
        RowSpec(
            id="MS1",
            wbs="M",
            task="Milestone: All transport + installation complete",
            phase="MILESTONE",
            location=f"{loc_mzp}/{loc_agi}",
            owner=owner,
            start_formula="=MAX($H$8:$H$200)",
            finish_formula="=MAX($H$8:$H$200)",
            tide_formula="",
            weather_formula="",
            notes="",
        )
    )

    return rows


def write_schedule_sheet(wb: Workbook, sheet_name: str, rows: List[RowSpec], gantt_days: int = 140) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A7"

    # Determine columns
    meta_cols = [
        "ID",
        "WBS",
        "Task",
        "Phase",
        "Location",
        "Owner",
        "Start",
        "Finish",
        "Dur (d)",
        "Tide peak (m)",
        "Weather gate",
        "Notes",
    ]
    gantt_start_col = len(meta_cols) + 1  # M
    header_row = 5
    first_row = 7

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=gantt_start_col + gantt_days - 1)
    ws["A1"].value = f"AGI TR1..TR7 – {sheet_name} (auto-shifts with Inputs!B5)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, h in enumerate(meta_cols, start=1):
        cell = ws.cell(header_row, i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border()

    # Day header formulas
    # Start at min of Start column in this sheet
    min_start_formula = f"=MIN($G${first_row}:$G$200)"
    for i in range(gantt_days):
        col = gantt_start_col + i
        cell = ws.cell(header_row, col)
        if i == 0:
            cell.value = min_start_formula
        else:
            prev = ws.cell(header_row, col - 1).coordinate
            cell.value = f"={prev}+1"
        cell.number_format = "dd-mmm"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 3.2

    ws.row_dimensions[header_row].height = 48

    # Column widths
    widths = {
        "A": 8,
        "B": 14,
        "C": 52,
        "D": 10,
        "E": 16,
        "F": 20,
        "G": 12,
        "H": 12,
        "I": 8,
        "J": 14,
        "K": 14,
        "L": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Data rows
    date_fmt = "dd-mmm-yy"
    base_border = thin_border()
    for idx, spec in enumerate(rows):
        r = first_row + idx
        ws.row_dimensions[r].height = 18
        tide_v = spec.tide_formula
        if tide_v == "AUTO_TIDE":
            tide_v = tide_lookup_formula(f"$G{r}")
        weather_v = spec.weather_formula
        if weather_v == "AUTO_WEATHER":
            weather_v = weather_lookup_formula(f"$G{r}")

        values = [
            spec.id,
            spec.wbs,
            spec.task,
            spec.phase,
            spec.location,
            spec.owner,
            spec.start_formula,
            spec.finish_formula,
            None,  # Dur
            tide_v,
            weather_v,
            spec.notes,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(r, c, value=v)
            if c in (1, 2, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in (7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = date_fmt
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(c in (3, 12)))
            cell.border = base_border

        # Duration formula (inclusive)
        ws.cell(r, 9).value = f'=IF(OR($G{r}="",$H{r}=""),"",$H{r}-$G{r}+1)'
        ws.cell(r, 9).alignment = Alignment(horizontal="center")

        # Summary styling
        if spec.phase == "SUMMARY":
            for c in range(1, len(meta_cols) + 1):
                cc = ws.cell(r, c)
                cc.fill = PatternFill("solid", fgColor=SUMMARY_GREY)
                cc.font = Font(bold=True, color="FFFFFF")

        # Milestone styling
        if spec.phase == "MILESTONE":
            for c in range(1, len(meta_cols) + 1):
                cc = ws.cell(r, c)
                cc.fill = PatternFill("solid", fgColor=PHASE_COLORS["MILESTONE"])
                cc.font = Font(bold=True)

        # Gantt grid cells (blank but bordered)
        for col in range(gantt_start_col, gantt_start_col + gantt_days):
            gc = ws.cell(r, col, value=None)
            gc.border = base_border

    last_row = first_row + len(rows) - 1
    gantt_end_col = gantt_start_col + gantt_days - 1
    gantt_end_letter = get_column_letter(gantt_end_col)
    first_day_letter = get_column_letter(gantt_start_col)
    cf_range = f"{first_day_letter}{first_row}:{gantt_end_letter}{last_row}"

    def add_cf(formula: str, fill_hex: str, priority: int):
        dxf = DifferentialStyle(fill=PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid"))
        rule = Rule(type="expression", dxf=dxf, formula=[formula])
        rule.priority = priority
        ws.conditional_formatting.add(cf_range, rule)

    # Weekend shading
    add_cf(f"WEEKDAY({first_day_letter}${header_row},2)>5", WEEKEND_SHADE, 1)
    # Today shading
    add_cf(f"{first_day_letter}${header_row}=TODAY()", TODAY_SHADE, 2)

    # Phase bars (use relative row references)
    pr = 3
    for phase, color in PHASE_COLORS.items():
        if phase in ("SUMMARY",):
            continue
        add_cf(
            f'AND($D{first_row}="{phase}",{first_day_letter}${header_row}>=$G{first_row},{first_day_letter}${header_row}<=$H{first_row})',
            color,
            pr,
        )
        pr += 1


def add_assumptions_sheet(wb: Workbook, template_refs: List[Tuple[str, str]]) -> None:
    ws = wb.create_sheet("Assumptions_Refs")
    ws["A1"].value = "Assumptions / References"
    ws["A1"].font = Font(bold=True, size=14)

    bullets = [
        "• Planning basis: 7 transformers (TM-63) moved via RoRo from Mina Zayed Port to AGI, using SPMTs + LCT.",
        "• Limits (Method Statement): Land wind ≤ 20 kn; RoRo wind ≤ 20 kn; Hs ≤ 0.6 m; visibility thresholds apply.",
        "• Tide: Load-out / Load-in to be aligned with high tide windows; Tide_Peaks_MZP is parsed from the provided Mammoet tide PDF (planning).",
        "• Weather: Excel uses a planning allowance day (Inputs!B22). Optional: VBA can fetch daily forecast and flag NO‑GO days.",
        "• Plan A inserts installation into voyage timeline (single critical path). Plan B decouples installation from voyages.",
        "• Installation parallel teams (Inputs!B4) is a planning lever to meet the pre‑March target.",
    ]
    r = 3
    for b in bullets:
        ws[f"A{r}"].value = b
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    ws[f"A{r}"].value = "Excel Gantt template technique references"
    ws[f"A{r}"].font = Font(bold=True)
    r += 1

    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(bold=True, color="FFFFFF")
    ws[f"A{r}"].value = "Source"
    ws[f"B{r}"].value = "Link"
    for c in (f"A{r}", f"B{r}"):
        ws[c].fill = header_fill
        ws[c].font = header_font
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
        ws[c].border = thin_border()
    r += 1
    for name, link in template_refs:
        ws[f"A{r}"].value = name
        ws[f"B{r}"].value = link
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{r}"].border = thin_border()
        ws[f"B{r}"].border = thin_border()
        r += 1

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 92


def add_gantt_chart_sheet(wb: Workbook, source_sheet: str, sheet_name: str) -> None:
    """Stacked-bar Gantt chart (offset + duration) based on a schedule sheet."""
    ws = wb.create_sheet(sheet_name)
    ws["A1"].value = f"Gantt Chart (Stacked Bar) – {source_sheet}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].value = "Uses Start offset (days) + Duration; Offset series is hidden."

    # Headers
    ws["A4"].value = "Task"
    ws["B4"].value = "Start"
    ws["C4"].value = "Offset (d)"
    ws["D4"].value = "Duration (d)"
    for c in range(1, 5):
        cell = ws.cell(4, c)
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    # Read tasks from source (exclude SUMMARY rows)
    src = wb[source_sheet]
    # Find last row
    last = 7
    while src.cell(last, 1).value is not None:
        last += 1
        if last > 300:
            break
    last -= 1
    out_r = 5
    # MinStart range
    min_range = f"$B$5:$B${5 + (last - 7)}"
    for r in range(7, last + 1):
        phase = src.cell(r, 4).value
        if phase == "SUMMARY":
            continue
        ws.cell(out_r, 1).value = f"={source_sheet}!C{r}"
        ws.cell(out_r, 2).value = f"={source_sheet}!G{r}"
        ws.cell(out_r, 2).number_format = "dd-mmm-yy"
        # Offset
        ws.cell(out_r, 3).value = f"=B{out_r}-MIN({min_range})"
        ws.cell(out_r, 4).value = f"={source_sheet}!I{r}"
        ws.cell(out_r, 4).number_format = "0"
        for c in range(1, 5):
            ws.cell(out_r, c).border = thin_border()
            ws.cell(out_r, c).alignment = Alignment(wrap_text=(c == 1), vertical="top")
        out_r += 1

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A5"

    # Chart
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 150
    chart.title = f"{source_sheet} – Transport Tasks"

    data = Reference(ws, min_col=3, min_row=4, max_col=4, max_row=out_r - 1)
    cats = Reference(ws, min_col=1, min_row=5, max_row=out_r - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.noFill = True
        chart.series[0].graphicalProperties.line.noFill = True
    ws.add_chart(chart, "F4")


def build_workbook(tide_pdf: Path) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    add_inputs_sheet(wb)
    add_weather_sheet(wb)
    add_calc_sheets(wb)

    peaks = extract_tide_daily_peaks(tide_pdf)
    add_tide_sheet(wb, peaks, source_label=f"{tide_pdf.name} – daily peak (planning; confirm official port table)")

    plan_a_rows = build_plan_rows("A")
    plan_b_rows = build_plan_rows("B")
    write_schedule_sheet(wb, "Plan_A_Realistic", plan_a_rows)
    write_schedule_sheet(wb, "Plan_B_Fast", plan_b_rows)

    # Chart sheets
    add_gantt_chart_sheet(wb, "Plan_A_Realistic", "Chart_PlanA")
    add_gantt_chart_sheet(wb, "Plan_B_Fast", "Chart_PlanB")

    # Template references (kept as plain links in-sheet)
    template_refs = [
        ("Microsoft Support – Create a Gantt chart in Excel (stacked bar)", "https://support.microsoft.com/en-us/office/present-your-data-in-a-gantt-chart-in-excel-f8910ab4-ceda-4521-8207-f0fb34d9e2b6"),
        ("Smartsheet – Gantt chart Excel templates", "https://www.smartsheet.com/gantt-chart-excel-templates"),
        ("Vertex42 – Gantt chart template (Excel)", "https://www.vertex42.com/ExcelTemplates/excel-gantt-chart.html"),
    ]
    add_assumptions_sheet(wb, template_refs)

    # Order sheets
    desired_order = [
        "Inputs",
        "Plan_A_Realistic",
        "Plan_B_Fast",
        "Chart_PlanA",
        "Chart_PlanB",
        "Tide_Peaks_MZP",
        "Weather_Forecast",
        "Assumptions_Refs",
        "Calc_PlanA",
        "Calc_PlanB",
    ]
    wb._sheets = [wb[s] for s in desired_order]
    return wb


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Build AGI TR7 dynamic Excel Gantt (Plan A/B).")
    p.add_argument("--tide_pdf", required=True, help="Path to MAMMOET_AGI TR.pdf")
    p.add_argument("--out", default="AGI_TR7_Dynamic_Gantt_Template_v13.xlsx", help="Output xlsx path")
    return p


def main() -> int:
    args = build_parser().parse_args()
    tide_pdf = Path(args.tide_pdf).expanduser().resolve()
    if not tide_pdf.exists():
        raise FileNotFoundError(f"tide_pdf not found: {tide_pdf}")
    out_path = Path(args.out).expanduser().resolve()
    wb = build_workbook(tide_pdf)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[OK] wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
