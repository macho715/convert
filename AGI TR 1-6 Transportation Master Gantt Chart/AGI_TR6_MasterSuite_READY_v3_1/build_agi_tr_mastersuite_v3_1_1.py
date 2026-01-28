#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGI TR MasterSuite Workbook Builder (v3.1.1)

Generates a ready-to-use Excel workbook (template + prefilled baseline schedule)
for AGI HVDC Transformer transportation planning, compatible with the provided VBA module.

Outputs:
  - AGI_TR6_MasterSuite_READY_v3_1_1.xlsx
  - AGI_TR6_MasterSuite_READY_v3_1_1.xlsm (macro-enabled container; import .bas in Excel)

Notes:
  - Tide_Data is a PLACEHOLDER. Replace with official tide tables when available.
  - Weather_Risk contains editable winter windows (incl. Shamal-like period).
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
from dataclasses import dataclass
from typing import Dict, List, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


PHASE_COLORS = {
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "BUFFER": "D9D9D9",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "RETURN": "999999",
    "BRIDGE": "CFE2F3",
    "TRANSPORT": "B4A7D6",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "MILESTONE": "FF0000",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def ceil_days(d: float) -> int:
    """Match VBA CalcEndDate: RoundUp(dur,0)-1."""
    if d <= 0:
        return 0
    return int(math.ceil(d)) - 1


def parse_long_list(s: str) -> List[int]:
    if not s:
        return []
    out = []
    for p in str(s).split(","):
        p = p.strip()
        if not p:
            continue
        out.append(int(float(p)))
    return out


@dataclass
class Task:
    id: str
    wbs: str
    task: str
    phase: str
    owner: str
    offset: float
    dur: float
    tr_list: str = ""
    voy: Any = ""
    batch: Any = ""
    tide_risk: str = ""
    weather_risk: str = ""
    critical: str = ""  # "Y"/""
    start: dt.date = dt.date(1900, 1, 1)
    end: dt.date = dt.date(1900, 1, 1)


def create_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for name in [
        "Control_Panel",
        "Scenario_Library",
        "Pattern_Tasks",
        "Schedule_Data",
        "Dependencies",
        "Resource_Calendar",
        "Tide_Data",
        "Weather_Risk",
        "Risk_Register",
        "Gantt_Chart",
        "Dashboard",
        "Docs_Checklist",
        "Evidence_Checklist",
        "Baseline",
        "Change_Log",
        "Reports",
        "Logs",
        "Exports",
    ]:
        wb.create_sheet(name)

    _build_control_panel(wb["Control_Panel"])
    _build_scenario_library(wb["Scenario_Library"])
    _build_pattern_tasks(wb["Pattern_Tasks"])
    _build_schedule_sheet(wb["Schedule_Data"])
    _build_tide_data(wb["Tide_Data"])
    _build_weather_risk(wb["Weather_Risk"])
    _build_gantt_sheet(wb["Gantt_Chart"])
    _build_simple_tables(wb["Dependencies"], ["Pred_ID", "Succ_ID", "Type(FS/SS/FF)", "Lag_Days", "Notes"])
    _build_simple_tables(wb["Resource_Calendar"], ["Resource", "WorkDays(1-7)", "ShiftHours", "Notes"])
    _build_simple_tables(wb["Risk_Register"], ["Risk_ID", "Category", "Description", "Likelihood", "Impact", "Owner", "Mitigation", "Status"])
    _build_simple_tables(wb["Docs_Checklist"], ["Doc", "Owner", "Due", "Submitted?", "Notes"])
    _build_simple_tables(wb["Evidence_Checklist"], ["Evidence Item", "Min Count", "Collected?", "Link/Path", "Notes"])
    _build_simple_tables(wb["Baseline"], ["Baseline_ID", "Frozen_On", "Notes"])
    _build_simple_tables(wb["Change_Log"], ["Timestamp", "User", "Change", "Impact", "Notes"])
    _build_simple_tables(wb["Reports"], ["Generated_On", "Report_Type", "Notes"])
    _build_simple_tables(wb["Logs"], ["Timestamp", "Level", "Module", "Message"])
    _build_simple_tables(wb["Exports"], ["Timestamp", "Type", "FilePath", "Notes"])

    return wb


def _build_control_panel(ws):
    ws["A4"] = "INPUTS"
    ws["A4"].font = Font(bold=True)

    labels = [
        ("B5", "D0 (V1 Load-out Start Date)"),
        ("B6", "Scenario"),
        ("B7", "Trip Plan (TR count per voyage)"),
        ("B8", "Install Batches (TR count per batch)"),
        ("B9", 'Jackdown Parallel Lines ("3대 잭다운")'),
        ("B10", "Weather Buffer (days per voyage)"),
        ("B11", "Tide Hold if HIGH (days)"),
        ("B12", "Target Completion Date"),
        ("B13", "Monte Carlo Runs (for P50/P80)"),
        ("B14", "Confidence Target (e.g., 0.8 for P80)"),
        ("B15", "Hard Deadline (must finish before)"),
    ]
    for cell, text in labels:
        ws[cell] = text

    ws["C5"] = dt.date(2026, 1, 6)
    ws["C6"] = "S2_4Voy_1-2-2-1"
    ws["C7"] = "1,2,2,1"
    ws["C8"] = "3,3"
    ws["C9"] = 3
    ws["C10"] = 1
    ws["C11"] = 1
    ws["C12"] = dt.date(2026, 2, 28)
    ws["C13"] = 500
    ws["C14"] = 0.8
    ws["C15"] = dt.date(2026, 3, 1)

    ws["A17"] = "OUTPUTS"
    ws["A17"].font = Font(bold=True)
    out_labels = [
        ("B18", "Planned Finish (Deterministic)", "C18"),
        ("B19", "P50 Finish (Monte Carlo)", "C19"),
        ("B20", "P80 Finish (Monte Carlo)", "C20"),
        ("B21", "Meets 3/1 Deadline?", "C21"),
        ("B22", "Tide/Weather Conflicts (count)", "C22"),
        ("B23", "Critical Path Length (days)", "C23"),
        ("B24", "Notes / Actions", "C24"),
    ]
    for b, t, c in out_labels:
        ws[b] = t
        ws[c] = "(run macro)"

    # Basic formatting
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 26
    for r in range(5, 16):
        ws[f"B{r}"].font = Font(bold=True)

    # Scenario dropdown validation (applied after Scenario_Library populated)
    dv = DataValidation(type="list", formula1="=Scenario_Library!$A$2:$A$50", allow_blank=False)
    dv.error = "Select a valid scenario from the list."
    dv.prompt = "Select scenario."
    ws.add_data_validation(dv)
    dv.add(ws["C6"])


def _build_scenario_library(ws):
    headers = ["Scenario_ID", "Description", "TripPlan", "InstallBatches", "ParallelJacks", "DefaultWeatherBuffer", "DefaultTideHold"]
    ws.append(headers)
    ws.append(["S1_6Voy_1TR", "6 voyages (1 TR each) + 3 install batches (2 TR sequential)", "1,1,1,1,1,1", "2,2,2", 1, 1, 1])
    ws.append(["S2_4Voy_1-2-2-1", "4 voyages (1-2-2-1) + 2 install batches (3 TR parallel)", "1,2,2,1", "3,3", 3, 1, 1])
    ws.append(["S3_Custom", "User-defined (reads Control_Panel plans)", "(from Control_Panel)", "(from Control_Panel)", 3, 1, 1])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22


def _build_pattern_tasks(ws):
    headers = ["Template", "Seq", "Code", "Task", "Phase", "Owner", "RelOffset_Days", "Duration_Days", "TideSensitive", "WeatherSensitive", "KeyTag"]
    ws.append(headers)

    rows = [
        ("VOYAGE_1TR", 10, "LO", "TR Load-out on SPMT + RoRo to LCT", "LOADOUT", "Mammoet", 0, 1, 1, 0, ""),
        ("VOYAGE_1TR", 20, "SF", "Sea fastening + lashing", "SEAFAST", "Mammoet", 1, 1, 0, 0, ""),
        ("VOYAGE_1TR", 30, "MWS", "MWS + MPI + final preparations", "BUFFER", "MWS/Captain", 2, 0.5, 0, 0, ""),
        ("VOYAGE_1TR", 40, "SAIL", "Sail-away MZP→AGI", "SAIL", "LCT Bushra", 2.5, 1, 0, 1, ""),
        ("VOYAGE_1TR", 50, "UNLD", "Arrival/Berthing + RoRo Load-in", "AGI_UNLOAD", "Mammoet", 3.5, 1, 1, 0, "DELIVERY_READY"),
        ("VOYAGE_1TR", 60, "RET", "LCT Return AGI→MZP", "RETURN", "LCT Bushra", 4.5, 1, 0, 1, "VOY_END"),
        ("VOYAGE_1TR", 70, "RST", "Buffer/Reset", "BUFFER", "All", 5.5, 0.5, 0, 0, ""),

        ("VOYAGE_2TR", 10, "LO1", "TR-A Load-out + RoRo", "LOADOUT", "Mammoet", 0, 1, 1, 0, ""),
        ("VOYAGE_2TR", 20, "SF1", "TR-A Sea fastening", "SEAFAST", "Mammoet", 1, 1, 0, 0, ""),
        ("VOYAGE_2TR", 30, "LO2", "TR-B Load-out + RoRo", "LOADOUT", "Mammoet", 2, 1, 1, 0, ""),
        ("VOYAGE_2TR", 40, "SF2", "TR-B Sea fastening", "SEAFAST", "Mammoet", 3, 1, 0, 0, ""),
        ("VOYAGE_2TR", 50, "MWS", "MWS + MPI + final preparations", "BUFFER", "MWS/Captain", 4, 0.5, 0, 0, ""),
        ("VOYAGE_2TR", 60, "SAIL", "Sail-away MZP→AGI", "SAIL", "LCT Bushra", 4.5, 1, 0, 1, ""),
        ("VOYAGE_2TR", 70, "UN2", "RoRo Load-in TR-B", "AGI_UNLOAD", "Mammoet", 5.5, 1, 1, 0, ""),
        ("VOYAGE_2TR", 80, "UN1", "RoRo Load-in TR-A", "AGI_UNLOAD", "Mammoet", 6.5, 1, 1, 0, "DELIVERY_READY"),
        ("VOYAGE_2TR", 90, "RET", "LCT Return AGI→MZP", "RETURN", "LCT Bushra", 7.5, 1, 0, 1, "VOY_END"),
        ("VOYAGE_2TR", 100, "RST", "Buffer/Reset", "BUFFER", "All", 8.5, 0.5, 0, 0, ""),

        ("INSTALL_BATCH", 10, "BR1", "Steel bridge installation / access prep", "BRIDGE", "Mammoet", 0, 0.5, 0, 0, ""),
        ("INSTALL_BATCH", 20, "TRNS", "Load on SPMT + transport to Bay", "TRANSPORT", "Mammoet", 0.5, 0.5, 0, 0, ""),
        ("INSTALL_BATCH", 30, "TURN", "Turning operation (90°)", "TURNING", "Mammoet", 1, 3, 0, 0, ""),
        ("INSTALL_BATCH", 40, "JD", "Jacking down on temporary support", "JACKDOWN", "Mammoet", 4, 1, 0, 0, "INSTALL_COMPLETE"),
        ("INSTALL_BATCH", 50, "BR2", "Steel bridge relocation / restore", "BRIDGE", "Mammoet", 5, 0.5, 0, 0, ""),
    ]
    for row in rows:
        ws.append(list(row))

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16


def _build_schedule_sheet(ws):
    # Title
    ws["A1"] = "Schedule_Data"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "ID", "WBS", "Task", "Phase", "Owner",
        "Offset_Days", "Start", "End", "Duration_Days",
        "TR_List", "Voyage", "Batch", "TideRisk", "WeatherRisk",
        "Critical?", "Baseline_Start", "Baseline_End",
        "Actual_Start", "Actual_End", "%Complete", "Status", "Notes"
    ]
    ws.append([""]*len(headers))  # row 2
    ws.append([""]*len(headers))  # row 3
    ws.append([""]*len(headers))  # row 4
    ws.append(headers)            # row 5

    for c in range(1, len(headers) + 1):
        cell = ws.cell(5, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["V"].width = 40


def _build_gantt_sheet(ws):
    ws["A1"] = "Gantt_Chart"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur", "Owner", "TR_List", "Voyage", "Batch", "Critical?", "Timeline (auto-built by VBA)"]
    ws.append(headers)  # row 1 (we keep a single header row at row 4 via padding)
    ws.delete_rows(1, 1)  # remove

    ws.append([""]*len(headers))  # row 2
    ws.append([""]*len(headers))  # row 3
    ws.append(headers)            # row 4

    for c in range(1, len(headers) + 1):
        cell = ws.cell(4, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "M5"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["M"].width = 3


def _build_tide_data(ws):
    ws.append(["Date", "TideWindow", "RiskLevel(HIGH/MED/LOW)", "Notes"])
    for c in range(1, 5):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Placeholder daily tide risk (edit/replace later)
    start = dt.date(2026, 1, 1)
    end = dt.date(2026, 3, 5)
    cur = start
    i = 0
    while cur <= end:
        # Simple pattern: every 7th day HIGH, every 4th day MED, else LOW
        risk = "LOW"
        if i % 7 == 0:
            risk = "HIGH"
        elif i % 4 == 0:
            risk = "MED"
        ws.append([cur, "AM/PM (placeholder)", risk, "Replace with official tide table when available."])
        cur += dt.timedelta(days=1)
        i += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 45


def _build_weather_risk(ws):
    headers = [
        "StartDate", "EndDate", "RiskLevel(LOW/MED/HIGH)",
        "DelayDays_Min", "DelayDays_Mode", "DelayDays_Max",
        "AppliesTo(Phase)", "Notes"
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.append([dt.date(2026, 1, 1), dt.date(2026, 1, 13), "LOW", 0, 0, 1, "SAIL,RETURN", "Average winter conditions."])
    ws.append([dt.date(2026, 1, 14), dt.date(2026, 1, 18), "HIGH", 1, 2, 3, "SAIL,RETURN,LOADOUT,AGI_UNLOAD", "High-risk window (editable)."])
    ws.append([dt.date(2026, 1, 19), dt.date(2026, 1, 21), "MED", 0, 1, 2, "SAIL,RETURN,LOADOUT,AGI_UNLOAD", "Medium-risk window (editable)."])
    ws.append([dt.date(2026, 1, 22), dt.date(2026, 2, 28), "LOW", 0, 0, 1, "SAIL,RETURN", "Average conditions; keep voyage buffer."])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 40


def _build_simple_tables(ws, headers: List[str]):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = max(12, min(40, len(headers[i-1]) + 2))


def compute_schedule(wb: Workbook) -> Tuple[List[Task], dt.date, int]:
    ws = wb["Control_Panel"]
    d0 = ws["C5"].value
    if isinstance(d0, dt.datetime):
        d0 = d0.date()

    scen = str(ws["C6"].value or "S2_4Voy_1-2-2-1")
    trip_plan = str(ws["C7"].value or "")
    batch_plan = str(ws["C8"].value or "")
    parallel = int(ws["C9"].value or 3)
    wbuf = int(ws["C10"].value or 1)

    # scenario defaults
    scen_ws = wb["Scenario_Library"]
    defaults: Dict[str, Dict[str, Any]] = {}
    for r in range(2, scen_ws.max_row + 1):
        sid = scen_ws.cell(r, 1).value
        if not sid:
            continue
        defaults[str(sid)] = {
            "trip": scen_ws.cell(r, 3).value,
            "batch": scen_ws.cell(r, 4).value,
            "parallel": scen_ws.cell(r, 5).value,
            "wbuf": scen_ws.cell(r, 6).value,
        }

    if scen in defaults and scen != "S3_Custom":
        trip_plan = trip_plan.strip() or str(defaults[scen]["trip"])
        batch_plan = batch_plan.strip() or str(defaults[scen]["batch"])
        parallel = int(defaults[scen]["parallel"] or parallel)
        wbuf = int(defaults[scen]["wbuf"] or wbuf)

    trips = parse_long_list(trip_plan)
    batches = parse_long_list(batch_plan)

    # pattern tasks
    pat_ws = wb["Pattern_Tasks"]
    pat_rows = []
    for r in range(2, pat_ws.max_row + 1):
        tpl = pat_ws.cell(r, 1).value
        if not tpl:
            continue
        pat_rows.append({
            "tpl": str(tpl),
            "code": str(pat_ws.cell(r, 3).value),
            "task": str(pat_ws.cell(r, 4).value),
            "phase": str(pat_ws.cell(r, 5).value),
            "owner": str(pat_ws.cell(r, 6).value),
            "dur": float(pat_ws.cell(r, 8).value or 0),
            "tag": str(pat_ws.cell(r, 11).value or ""),
        })

    def build_voy(voy_no: int, load_n: int, first_tr: int, voy_start: float) -> Tuple[List[Task], float, float]:
        tpl = "VOYAGE_1TR" if load_n == 1 else "VOYAGE_2TR"
        offset = voy_start
        order = 0
        tr_a = first_tr
        tr_b = first_tr + 1
        tr_list = f"TR{tr_a}" if load_n == 1 else f"TR{tr_a},TR{tr_b}"
        delivered = None
        out: List[Task] = []

        for row in pat_rows:
            if row["tpl"] != tpl:
                continue
            order += 1
            code = row["code"]
            name = row["task"]
            if load_n == 2:
                if code in ("LO1", "SF1"):
                    name = name.replace("TR-A", f"TR{tr_a}")
                elif code in ("LO2", "SF2", "UN2"):
                    name = name.replace("TR-B", f"TR{tr_b}")
                elif code == "UN1":
                    name = name.replace("TR-A", f"TR{tr_a}")

            out.append(Task(
                id=f"V{voy_no:02d}-{code}",
                wbs=f"{voy_no+1}.{order:02d}",
                task=name,
                phase=row["phase"],
                owner=row["owner"],
                offset=offset,
                dur=row["dur"],
                tr_list=tr_list,
                voy=voy_no,
                batch=""
            ))
            offset += row["dur"]
            if row["tag"].upper() == "DELIVERY_READY":
                delivered = offset

        # voyage buffer
        if wbuf > 0:
            order += 1
            out.append(Task(
                id=f"V{voy_no:02d}-WBUF",
                wbs=f"{voy_no+1}.{order:02d}",
                task="Weather/Operational Buffer (voyage)",
                phase="BUFFER",
                owner="All",
                offset=offset,
                dur=float(wbuf),
                tr_list=tr_list,
                voy=voy_no,
                batch=""
            ))
            offset += wbuf

        if delivered is None:
            delivered = voy_start + 4
        return out, delivered, offset

    def batch_tr_start(batch_no: int) -> int:
        return 1 + sum(batches[:batch_no-1])

    def build_batch(batch_no: int, batch_size: int, start_offset: float) -> Tuple[List[Task], float]:
        out: List[Task] = []
        order = 1
        out.append(Task(
            id=f"B{batch_no:02d}-BR1",
            wbs=f"I.{batch_no}.01",
            task=f"Batch {batch_no} — Steel bridge/access prep",
            phase="BRIDGE",
            owner="Mammoet",
            offset=start_offset,
            dur=0.5,
            tr_list="",
            batch=batch_no,
        ))
        lanes = [start_offset + 0.5 for _ in range(max(1, parallel))]
        base = batch_tr_start(batch_no)
        for i in range(batch_size):
            tr = base + i
            lane = i % max(1, parallel)

            order += 1
            out.append(Task(
                id=f"B{batch_no:02d}-TR{tr}-TRNS",
                wbs=f"I.{batch_no}.{order:02d}",
                task=f"TR{tr} — Load on SPMT + Transport to Bay",
                phase="TRANSPORT",
                owner="Mammoet",
                offset=lanes[lane],
                dur=0.5,
                tr_list=f"TR{tr}",
                batch=batch_no,
            ))
            lanes[lane] += 0.5

            order += 1
            out.append(Task(
                id=f"B{batch_no:02d}-TR{tr}-TURN",
                wbs=f"I.{batch_no}.{order:02d}",
                task=f"TR{tr} — Turning operation (90°)",
                phase="TURNING",
                owner="Mammoet",
                offset=lanes[lane],
                dur=3.0,
                tr_list=f"TR{tr}",
                batch=batch_no,
            ))
            lanes[lane] += 3.0

            order += 1
            out.append(Task(
                id=f"B{batch_no:02d}-TR{tr}-JD",
                wbs=f"I.{batch_no}.{order:02d}",
                task=f"TR{tr} — Jacking down on temporary support (Install Complete)",
                phase="JACKDOWN",
                owner="Mammoet",
                offset=lanes[lane],
                dur=1.0,
                tr_list=f"TR{tr}",
                batch=batch_no,
            ))
            lanes[lane] += 1.0

        bend = max(lanes)
        order += 1
        out.append(Task(
            id=f"B{batch_no:02d}-BR2",
            wbs=f"I.{batch_no}.{order:02d}",
            task=f"Batch {batch_no} — Bridge relocation/restore",
            phase="BRIDGE",
            owner="Mammoet",
            offset=bend,
            dur=0.5,
            tr_list=f"TR{base}–TR{base+batch_size-1}",
            batch=batch_no,
        ))
        return out, bend + 0.5

    tasks: List[Task] = []
    cur_voy = 0.0
    cur_ins = 0.0
    tr_counter = 1
    cum_delivered = 0
    batch_idx = 0

    tasks.append(Task("MOB-001", "1.0", "Mobilization (SPMT/Marine/Steelworks) + Function Test", "MOBILIZATION", "Mammoet", cur_voy, 1.0))
    cur_voy += 1.0
    tasks.append(Task("PREP-001", "1.1", "Deck Preparations (D-ring, markings, steel sets, welding)", "DECK_PREP", "Mammoet", cur_voy, 2.0))
    cur_voy += 2.0

    for voy_no, load_n in enumerate(trips, start=1):
        if load_n <= 0:
            continue
        voy_tasks, delivered, voy_end = build_voy(voy_no, load_n, tr_counter, cur_voy)
        tasks.extend(voy_tasks)
        cur_voy = voy_end

        cum_delivered += load_n
        tr_counter += load_n

        if batch_idx < len(batches):
            threshold = sum(batches[:batch_idx+1])
            if cum_delivered >= threshold:
                batch_idx += 1
                bsize = batches[batch_idx-1]
                bstart = max(cur_ins, delivered + 0.5)
                b_tasks, b_end = build_batch(batch_idx, bsize, bstart)
                tasks.extend(b_tasks)
                cur_ins = b_end

    finish_off = max(cur_voy, cur_ins)
    tasks.append(Task("COMP", "99.0", "PROJECT COMPLETE — All TRs installed", "MILESTONE", "All", finish_off, 0.0))

    # date conversion
    for t in tasks:
        s = d0 + dt.timedelta(days=int(math.floor(t.offset)))
        e = s + dt.timedelta(days=ceil_days(t.dur))
        t.start = s
        t.end = e

    # risks
    tide = _tide_dict(wb)
    weather_rules = _weather_rules(wb)
    conflicts = 0
    for t in tasks:
        if t.phase.upper() in ("LOADOUT", "AGI_UNLOAD"):
            t.tide_risk = tide.get(t.start, "LOW")
            if t.tide_risk == "HIGH":
                conflicts += 1
        if t.phase.upper() in ("SAIL", "RETURN"):
            t.weather_risk = _weather_risk(t.start, t.phase.upper(), weather_rules)
            if t.weather_risk == "HIGH":
                conflicts += 1

    return tasks, d0, conflicts


def _tide_dict(wb: Workbook) -> Dict[dt.date, str]:
    ws = wb["Tide_Data"]
    out: Dict[dt.date, str] = {}
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        risk = ws.cell(r, 3).value
        if not d or not risk:
            continue
        if isinstance(d, dt.datetime):
            d = d.date()
        out[d] = str(risk).strip().upper()
    return out


def _weather_rules(wb: Workbook) -> List[Dict[str, Any]]:
    ws = wb["Weather_Risk"]
    rules = []
    for r in range(2, ws.max_row + 1):
        sd, ed = ws.cell(r, 1).value, ws.cell(r, 2).value
        risk = ws.cell(r, 3).value
        applies = ws.cell(r, 7).value
        if not sd or not ed or not risk:
            continue
        if isinstance(sd, dt.datetime):
            sd = sd.date()
        if isinstance(ed, dt.datetime):
            ed = ed.date()
        appl = [p.strip().upper() for p in str(applies or "").split(",") if p.strip()]
        rules.append({"start": sd, "end": ed, "risk": str(risk).strip().upper(), "applies": appl})
    return rules


def _weather_risk(d: dt.date, phase: str, rules: List[Dict[str, Any]]) -> str:
    for rule in rules:
        if rule["start"] <= d <= rule["end"] and phase in rule["applies"]:
            return rule["risk"]
    return "LOW"


def write_schedule_and_gantt(wb: Workbook, tasks: List[Task], d0: dt.date, conflicts: int) -> None:
    ws = wb["Schedule_Data"]

    # Clear existing rows from row 6 down
    max_r = ws.max_row
    if max_r >= 6:
        for r in range(6, max_r + 1):
            for c in range(1, 23):
                ws.cell(r, c).value = None
                ws.cell(r, c).fill = PatternFill()

    row = 6
    for t in tasks:
        ws.cell(row, 1).value = t.id
        ws.cell(row, 2).value = t.wbs
        ws.cell(row, 3).value = t.task
        ws.cell(row, 4).value = t.phase
        ws.cell(row, 5).value = t.owner
        ws.cell(row, 6).value = float(t.offset)
        ws.cell(row, 7).value = t.start
        ws.cell(row, 8).value = t.end
        ws.cell(row, 9).value = float(t.dur)
        ws.cell(row, 10).value = t.tr_list
        ws.cell(row, 11).value = t.voy
        ws.cell(row, 12).value = t.batch
        ws.cell(row, 13).value = t.tide_risk
        ws.cell(row, 14).value = t.weather_risk
        ws.cell(row, 15).value = t.critical
        ws.cell(row, 20).value = 0
        ws.cell(row, 21).value = "Not Started"
        ws.cell(row, 22).value = ""

        ws.cell(row, 6).number_format = "0.0"
        ws.cell(row, 7).number_format = "yyyy-mm-dd"
        ws.cell(row, 8).number_format = "yyyy-mm-dd"
        ws.cell(row, 9).number_format = "0.0"

        color = PHASE_COLORS.get(t.phase.upper())
        if color:
            ws.cell(row, 3).fill = PatternFill("solid", fgColor=color)
        if t.phase.upper() == "MILESTONE":
            for c in range(1, 15):
                ws.cell(row, c).fill = PatternFill("solid", fgColor=PHASE_COLORS["MILESTONE"])
                ws.cell(row, c).font = Font(bold=True, color="FFFFFF")
        row += 1

    # Control panel outputs
    ctrl = wb["Control_Panel"]
    finish = max(t.end for t in tasks)
    ctrl["C18"] = finish
    ctrl["C18"].number_format = "yyyy-mm-dd"
    ctrl["C22"] = conflicts

    deadline = ctrl["C15"].value
    if isinstance(deadline, dt.datetime):
        deadline = deadline.date()
    ctrl["C21"] = "YES" if finish < deadline else "NO"
    ctrl["C21"].font = Font(bold=True, color="006100" if ctrl["C21"].value == "YES" else "9C0006")

    _build_static_gantt(wb, tasks)


def _build_static_gantt(wb: Workbook, tasks: List[Task]) -> None:
    ws = wb["Gantt_Chart"]
    # clear rows
    max_r = ws.max_row
    for r in range(5, max_r + 1):
        for c in range(1, 250):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill()

    starts = [t.start for t in tasks]
    ends = [t.end for t in tasks]
    min_s = min(starts)
    max_e = max(ends)
    start_t = min_s - dt.timedelta(days=2)
    total_days = (max_e + dt.timedelta(days=2) - start_t).days + 1
    total_days = max(30, min(140, total_days))
    col0 = 13  # M

    # timeline header row 4
    weather_rules = _weather_rules(wb)
    for i in range(total_days):
        d = start_t + dt.timedelta(days=i)
        cell = ws.cell(4, col0 + i)
        cell.value = d
        cell.number_format = "d"
        cell.alignment = Alignment(horizontal="center")
        if d.weekday() >= 5:
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        if _weather_risk(d, "SAIL", weather_rules) == "HIGH":
            cell.fill = PatternFill("solid", fgColor="FFE699")

    # rows
    for idx, t in enumerate(tasks):
        row = 5 + idx
        ws.cell(row, 1).value = t.id
        ws.cell(row, 2).value = t.wbs
        ws.cell(row, 3).value = t.task
        ws.cell(row, 4).value = t.phase
        ws.cell(row, 5).value = t.start
        ws.cell(row, 6).value = t.end
        ws.cell(row, 7).value = float(t.dur)
        ws.cell(row, 8).value = t.owner
        ws.cell(row, 9).value = t.tr_list
        ws.cell(row, 10).value = t.voy
        ws.cell(row, 11).value = t.batch
        ws.cell(row, 12).value = t.critical

        ws.cell(row, 5).number_format = "yyyy-mm-dd"
        ws.cell(row, 6).number_format = "yyyy-mm-dd"
        ws.cell(row, 7).number_format = "0.0"

        fill = PHASE_COLORS.get(t.phase.upper(), "D9D9D9")
        c_start = col0 + (t.start - start_t).days
        c_end = col0 + (t.end - start_t).days
        for c in range(c_start, c_end + 1):
            ws.cell(row, c).fill = PatternFill("solid", fgColor=fill)

        if t.phase.upper() == "MILESTONE":
            ws.cell(row, c_start).value = "★"
            ws.cell(row, c_start).font = Font(bold=True, color="FFFFFF")
            ws.cell(row, c_start).alignment = Alignment(horizontal="center")


def self_test(wb: Workbook) -> List[str]:
    required = [
        "Control_Panel", "Scenario_Library", "Pattern_Tasks", "Schedule_Data",
        "Gantt_Chart", "Tide_Data", "Weather_Risk", "Logs", "Exports"
    ]
    issues = []
    for sh in required:
        if sh not in wb.sheetnames:
            issues.append(f"Missing sheet: {sh}")

    pt = wb["Pattern_Tasks"]
    if str(pt.cell(1, 1).value).strip().upper() != "TEMPLATE":
        issues.append("Pattern_Tasks header mismatch at A1")
    if str(pt.cell(1, 8).value).strip().upper() != "DURATION_DAYS":
        issues.append("Pattern_Tasks header mismatch at H1")
    return issues


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="AGI_TR6_MasterSuite_READY_v3_1_1.xlsx", help="Output xlsx filename")
    args = ap.parse_args()

    wb = create_workbook()
    tasks, d0, conflicts = compute_schedule(wb)
    write_schedule_and_gantt(wb, tasks, d0, conflicts)

    issues = self_test(wb)
    if issues:
        raise SystemExit("SELFTEST FAILED:\n- " + "\n- ".join(issues))

    wb.save(args.out)

    # Also save an .xlsm container (macros imported via .bas in Excel)
    xlsm = args.out.rsplit(".", 1)[0] + ".xlsm"
    wb.save(xlsm)

    print("OK")
    print("Saved:", args.out)
    print("Saved:", xlsm)
    print("Planned finish:", max(t.end for t in tasks))
    print("Conflicts:", conflicts)


if __name__ == "__main__":
    main()
