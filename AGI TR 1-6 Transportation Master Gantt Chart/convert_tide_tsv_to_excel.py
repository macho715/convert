"""
TSV를 Excel로 변환하고 각 항차 load-out 날짜의 tide 상위 3개를 색상으로 표시
- 2026년 1월 27일 이후의 모든 날짜 포함
- 모든 시간대(0:00 ~ 23:00) 표시
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import sys
import io
import re

# === COLORS (Untitled-1.py와 동일) ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

# === Border 스타일 (Untitled-1.py와 동일) ===
BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

# Windows 콘솔 UTF-8 인코딩 설정
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Voyage별 load-out 날짜
VOYAGE_LOADOUT_DATES = {
    1: '29-Jan-2026',
    2: '05-Feb-2026',
    3: '12-Feb-2026',
    4: '19-Feb-2026',
    5: '26-Feb-2026',
    6: '05-Mar-2026',
    7: '12-Mar-2026'
}

# 필터링 기준 날짜 (2026년 1월 27일)
FILTER_DATE = datetime(2026, 1, 27)

def parse_date_str(date_str: str) -> datetime:
    """날짜 문자열을 datetime으로 변환"""
    # "27-Jan-2026" 형식
    try:
        return datetime.strptime(date_str, '%d-%b-%Y')
    except:
        return None

def parse_tsv_tide_data(tsv_path: Path) -> dict:
    """TSV 파일에서 tide 데이터 파싱 (2026-01-27 이후만)"""
    tide_data = {}
    
    with open(tsv_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 헤더 찾기
    header_idx = None
    for i, line in enumerate(lines):
        if 'DATE:' in line and 'TIME' in line:
            header_idx = i
            break
    
    if header_idx is None:
        raise ValueError("TSV 헤더를 찾을 수 없습니다")
    
    # 시간 컬럼 추출 (모든 탭 유지, 빈 값은 None으로 처리)
    header_line = lines[header_idx].strip().split('\t')
    time_cols = []
    for col in header_line[1:]:  # DATE 제외
        col_stripped = col.strip()
        time_cols.append(col_stripped if col_stripped else None)
    
    print(f"  헤더에서 {len(time_cols)}개 탭 발견 (빈 탭 포함)")
    
    # 데이터 파싱
    for line in lines[header_idx + 1:]:
        if not line.strip() or line.startswith('#'):
            continue
        
        parts = line.strip().split('\t')
        if len(parts) < 2:
            continue
        
        date_str = parts[0].strip()
        if not date_str:
            continue
        
        # 날짜 파싱 및 필터링
        date_obj = parse_date_str(date_str)
        if date_obj is None or date_obj < FILTER_DATE:
            continue
        
        # 모든 탭의 값을 유지 (빈 값은 None)
        tide_values = []
        for val in parts[1:]:
            val = val.strip()
            if not val:
                tide_values.append(None)
            else:
                try:
                    tide_val = float(val)
                    tide_values.append(tide_val)
                except (ValueError, TypeError):
                    tide_values.append(None)
        
        # 시간 컬럼과 값 매칭 (인덱스로 정확히 매칭)
        # time_cols와 tide_values의 길이를 맞춤
        max_len = max(len(time_cols), len(tide_values))
        times = []
        values = []
        
        for i in range(max_len):
            # 시간대 가져오기
            if i < len(time_cols):
                time_val = time_cols[i]
            else:
                time_val = None
            
            # 값 가져오기
            if i < len(tide_values):
                tide_val = tide_values[i]
            else:
                tide_val = None
            
            # 시간대가 있는 경우만 추가 (빈 탭은 제외하되 인덱스는 유지)
            if time_val:
                times.append(time_val)
                values.append(tide_val)
        
        tide_data[date_str] = {
            'times': times,
            'values': values
        }
    
    # 파싱 결과 확인
    if tide_data:
        first_date = list(tide_data.keys())[0]
        first_data = tide_data[first_date]
        all_times = first_data['times']
        print(f"  첫 번째 날짜 ({first_date}) 시간대 수: {len(all_times)}")
        print(f"  시간대 목록: {', '.join(all_times)}")
        # 18:00과 22:00이 포함되어 있는지 확인
        if '18:00' in all_times:
            print(f"  ✓ 18:00 포함됨")
        else:
            print(f"  ⚠️  18:00 누락됨")
        if '22:00' in all_times:
            print(f"  ✓ 22:00 포함됨")
        else:
            print(f"  ⚠️  22:00 누락됨")
    
    return tide_data

def find_top3_tides_for_date(tide_data: dict, date_str: str) -> list:
    """특정 날짜에서 tide 상위 3개 찾기"""
    if date_str not in tide_data:
        return []
    
    data = tide_data[date_str]
    times = data['times']
    values = data['values']
    
    # 시간과 값 쌍으로 만들고 정렬
    tide_pairs = []
    for time, value in zip(times, values):
        if value is not None:
            tide_pairs.append({
                'time': time,
                'value': value
            })
    
    # 값 기준 내림차순 정렬
    tide_pairs.sort(key=lambda x: x['value'], reverse=True)
    
    # 상위 3개 반환
    return tide_pairs[:3]

def convert_tsv_to_excel_with_highlight(tsv_path: Path, excel_template_path: Path, output_path: Path):
    """TSV를 Excel로 변환하고 tide 상위 3개를 색상으로 표시"""
    
    # TSV 데이터 파싱
    print("TSV 데이터 파싱 중 (2026-01-27 이후만)...")
    tide_data = parse_tsv_tide_data(tsv_path)
    print(f"✓ {len(tide_data)}일의 데이터 파싱 완료")
    
    # Excel 템플릿 로드
    print(f"\nExcel 템플릿 로드 중: {excel_template_path}")
    wb = load_workbook(excel_template_path)
    
    # 첫 번째 시트 사용
    ws = wb.active
    print(f"✓ 시트 '{ws.title}' 활성화")
    
    # TSV 데이터를 Excel에 쓰기
    print("\nTSV 데이터를 Excel에 쓰는 중...")
    
    # 먼저 모든 병합 셀 해제 (데이터 영역)
    print("  병합된 셀 해제 중...")
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        try:
            ws.unmerge_cells(str(merged_range))
        except Exception as e:
            print(f"    경고: 병합 해제 실패 {merged_range}: {e}")
    
    # TSV의 모든 시간대 확인
    all_times = set()
    if tide_data:
        for date_data in tide_data.values():
            all_times.update(date_data['times'])
    
    # 타이틀 행 추가 (Untitled-1.py 스타일)
    max_col_letter = get_column_letter(1 + len(all_times) if all_times else 25)
    ws.merge_cells(f"A1:{max_col_letter}1")
    ws["A1"] = "MINA ZAYED PORT - Water Tide Data (Jan to Mar 2026)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells(f"A2:{max_col_letter}2")
    ws["A2"] = "Tide ≥1.90m required for Load-out and AGI Arrival | Top 3 highest tides highlighted in yellow"
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # 헤더 행은 4번째 행으로 설정
    header_row = 4
    
    # 시간대별 컬럼 인덱스 매핑 생성 (모든 시간대를 정렬하여 배치)
    time_to_col = {}
    if all_times:
        sorted_times = sorted(all_times, key=lambda x: int(x.split(':')[0]) if ':' in x else 999)
        
        # DATE 컬럼
        ws.cell(row=header_row, column=1, value="DATE: \\ TIME")
        header_cell = ws.cell(row=header_row, column=1)
        header_cell.font = Font(bold=True, color="FFFFFF", size=9)
        header_cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.border = tb()
        
        # 시간대 컬럼들
        for col_idx, time in enumerate(sorted_times, 2):
            ws.cell(row=header_row, column=col_idx, value=time)
            header_cell = ws.cell(row=header_row, column=col_idx)
            header_cell.font = Font(bold=True, color="FFFFFF", size=9)
            header_cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.border = tb()
            time_to_col[time] = col_idx
    
    print(f"  총 {len(time_to_col)}개 시간대 매핑 완료 (0:00~23:00 모두 포함)")
    
    # 데이터 쓰기 (날짜순 정렬)
    data_row = header_row + 1
    sorted_dates = sorted(tide_data.keys(), key=lambda x: parse_date_str(x) or datetime.max)
    
    for date_str in sorted_dates:
        data = tide_data[date_str]
        
        # 날짜 셀 (스타일 적용)
        date_cell = ws.cell(row=data_row, column=1)
        try:
            date_obj = parse_date_str(date_str)
            if date_obj:
                date_cell.value = date_obj
                date_cell.number_format = "YYYY-MM-DD"
            else:
                date_cell.value = date_str
        except:
            date_cell.value = date_str
        date_cell.border = tb()
        
        # 시간대별로 정확한 컬럼에 값 쓰기 (스타일 적용)
        for time, value in zip(data['times'], data['values']):
            if time in time_to_col and value is not None:
                col_idx = time_to_col[time]
                value_cell = ws.cell(row=data_row, column=col_idx, value=value)
                value_cell.number_format = "0.00"
                value_cell.border = tb()
        
        data_row += 1
    
    print(f"✓ 데이터 쓰기 완료 (행 {header_row + 1}부터 {data_row - 1}까지, 총 {len(sorted_dates)}일)")
    
    # 각 Voyage의 load-out 날짜에서 tide 상위 3개 색상 표시
    print("\n각 항차별 tide 상위 3개 색상 표시 중...")
    
    # 색상 정의 (노란색 계열)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for voyage_num, date_str in VOYAGE_LOADOUT_DATES.items():
        print(f"  Voyage {voyage_num} ({date_str})...")
        top3 = find_top3_tides_for_date(tide_data, date_str)
        
        if not top3:
            print(f"    ⚠️  데이터를 찾을 수 없습니다")
            continue
        
        # 해당 날짜의 행 찾기
        date_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # 날짜 형식 비교
                try:
                    if isinstance(cell_value, datetime):
                        if cell_value.date() == parse_date_str(date_str).date():
                            date_row = row_idx
                            break
                    elif cell_str == date_str:
                        date_row = row_idx
                        break
                except:
                    if cell_str == date_str:
                        date_row = row_idx
                        break
        
        if date_row:
            print(f"    ✓ 행 {date_row}에서 상위 3개 tide 발견:")
            for tide_info in top3:
                # 시간대 이름으로 컬럼 찾기
                time = tide_info['time']
                if time in time_to_col:
                    col_idx = time_to_col[time]
                    cell = ws.cell(row=date_row, column=col_idx)
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)  # 강조를 위해 볼드 추가
                    print(f"      - {time}: {tide_info['value']}m (컬럼 {col_idx})")
        else:
            print(f"    ⚠️  날짜 '{date_str}'를 Excel에서 찾을 수 없습니다")
    
    # 컬럼 너비 설정 (Untitled-1.py 스타일)
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 6
    
    # Freeze panes 설정 (Untitled-1.py 스타일)
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    
    # 저장
    print(f"\nExcel 파일 저장 중: {output_path}")
    wb.save(output_path)
    print("✅ 완료!")

# 실행
if __name__ == "__main__":
    script_dir = Path(__file__).parent.absolute()
    base_dir = script_dir
    
    tsv_path = base_dir / "water tide(jan to march).tsv"
    excel_template_path = base_dir / "water tide_7trip.xlsx"
    
    # 타임스탬프를 포함한 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = base_dir / f"water tide_7trip_highlighted_{timestamp}.xlsx"
    
    if not tsv_path.exists():
        print(f"❌ TSV 파일을 찾을 수 없습니다: {tsv_path}")
        sys.exit(1)
    
    if not excel_template_path.exists():
        print(f"❌ Excel 템플릿 파일을 찾을 수 없습니다: {excel_template_path}")
        sys.exit(1)
    
    convert_tsv_to_excel_with_highlight(tsv_path, excel_template_path, output_path)
