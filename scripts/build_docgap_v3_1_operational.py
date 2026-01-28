"""Build/Patch: OFCO AGI TR1 DocGap Tracker v3.1 (Operational)

- Adds schedule scenario selection (Inputs!B9) with scenario table
- Adds manual date inputs (Inputs column C) + auto date formulas (Inputs column B)
- Adds Submit To -> default lead time + anchor mapping (Inputs rows 20..)
- Expands OFCO_Req_1_15 and NOC_Req_1_6 to support:
  Lead Time Default / Override / Effective / Anchor / Target Submit Date

Usage:
  python build_docgap_v3_1_operational.py \
    --in  /mnt/data/OFCO_AGI_TR1_DocGap_Tracker_v3_FULLOPTIONS.xlsx \
    --out /mnt/data/OFCO_AGI_TR1_DocGap_Tracker_v3_1_Operational.xlsx

Note:
  This script produces .xlsx. If you need .xlsm, rename the extension or copy.
"""

import argparse
from datetime import datetime
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment

# VBA 매크로 코드
VBA_MACROS = """Option Explicit

'=====================
' Doc Gap Tracker Macros (v3.1)
'=====================

Sub RefreshAll()
    'Recalculate
    Application.CalculateFull
End Sub

Sub FilterMissing()
    'Filter current sheet by Status = Missing (assumes headers on row 2)
    Dim ws As Worksheet
    Set ws = ActiveSheet
    If ws.AutoFilterMode = False Then Exit Sub
    ws.Range("A2").AutoFilter Field:=4, Criteria1:="Missing"
End Sub

Sub ClearFilters()
    Dim ws As Worksheet
    Set ws = ActiveSheet
    If ws.AutoFilterMode Then
        On Error Resume Next
        ws.ShowAllData
        On Error GoTo 0
    End If
End Sub

Sub StampUpdated()
    'Stamp last updated time in Executive_Summary!C1
    With ThisWorkbook.Sheets("Executive_Summary")
        .Range("C1").Value = "Last updated: " & Format(Now, "dd-mmm-yy hh:nn")
    End With
End Sub

Sub ApplyScenarioToManual()
    'Copy auto schedule dates (Column B) to manual (Column C) and set scenario to CUSTOM
    With ThisWorkbook.Sheets("Inputs")
        .Range("C3").Value = .Range("B3").Value
        .Range("C4").Value = .Range("B4").Value
        .Range("C5").Value = .Range("B5").Value
        .Range("C6").Value = .Range("B6").Value
        .Range("B9").Value = "CUSTOM"
    End With
    Application.CalculateFull
End Sub
"""



def apply_v3_1(src: str, out: str) -> None:
    wb = openpyxl.load_workbook(src)

    # ---------------- Inputs sheet ----------------
    ws = wb["Inputs"]

    # Insert Manual column C
    if ws.cell(2, 3).value != "Value (Manual)":
        ws.insert_cols(3)
        ws.cell(2, 2).value = "Value (Auto)"
        ws.cell(2, 3).value = "Value (Manual)"

        # Copy existing manual values
        for r in range(3, 7):
            ws.cell(r, 3).value = ws.cell(r, 2).value

    # Scenario selection row
    ws.cell(9, 1).value = "Schedule Scenario (Select)"
    ws.cell(9, 2).value = ws.cell(9, 2).value or "SCN-01"
    ws.cell(9, 3).value = "(Select SCN code; choose CUSTOM to use manual values in Column C for rows 3–6)"

    # Scenario table
    ws.cell(11, 1).value = "Scenario"
    ws.cell(11, 2).value = "Arrival"
    ws.cell(11, 3).value = "RoRo Start"
    ws.cell(11, 4).value = "RoRo End"
    ws.cell(11, 5).value = "Departure"

    # Default scenario
    ws.cell(12, 1).value = "SCN-01"
    ws.cell(12, 2).value = datetime(2026, 1, 27)
    ws.cell(12, 3).value = datetime(2026, 1, 29)
    ws.cell(12, 4).value = datetime(2026, 1, 30)
    ws.cell(12, 5).value = datetime(2026, 2, 1)

    # Empty placeholders
    for i, scn in enumerate(["SCN-02", "SCN-03", "SCN-04", "SCN-05", "CUSTOM"], start=13):
        ws.cell(i, 1).value = scn
        for c in range(2, 6):
            ws.cell(i, c).value = None

    # Weekend pattern for WORKDAY.INTL (editable)
    ws.cell(7, 1).value = "Weekend Pattern (Mon..Sun, 0=Work,1=Weekend)"
    ws.cell(7, 2).value = "0000011"  # default: Sat-Sun
    ws.cell(7, 3).value = "Example: Sat-Sun=0000011 / Fri-Sun=0000111 (edit as needed)"

    # Auto date formulas (B3..B6)
    ws.cell(3, 2).value = '=IF($B$9="CUSTOM",$C3,VLOOKUP($B$9,$A$12:$E$17,2,FALSE))'
    ws.cell(4, 2).value = '=IF($B$9="CUSTOM",$C4,VLOOKUP($B$9,$A$12:$E$17,3,FALSE))'
    ws.cell(5, 2).value = '=IF($B$9="CUSTOM",$C5,VLOOKUP($B$9,$A$12:$E$17,4,FALSE))'
    ws.cell(6, 2).value = '=IF($B$9="CUSTOM",$C6,VLOOKUP($B$9,$A$12:$E$17,5,FALSE))'

    # Scenario + Weekend dropdown validation
    ws.data_validations.dataValidation = []  # clear and re-add
    dv_scn = DataValidation(type="list", formula1="=$A$12:$A$17", allow_blank=False)
    dv_wkend = DataValidation(type="list", formula1='"0000011,0000111"', allow_blank=False)
    ws.add_data_validation(dv_scn)
    ws.add_data_validation(dv_wkend)
    dv_scn.add("B9")
    dv_wkend.add("B7")

    # Lead time mapping
    ws.cell(19, 1).value = "Lead Time & Anchor Mapping (Editable)"
    ws.cell(20, 1).value = "Submit To"
    ws.cell(20, 2).value = "Default Lead Time (WD)"
    ws.cell(20, 3).value = "Anchor Key (Arrival/RoRoStart/RoRoEnd/Departure)"

    mapping = [
        ("HSE", 2, "RoRoStart"),
        ("HM", 2, "RoRoStart"),
        ("Maqta Gateway", 3, "RoRoStart"),
        ("Port Ops", 1, "Arrival"),
        ("Customs", 2, "Arrival"),
        ("AD Maritime", 5, "Departure"),
    ]

    start_row = 21
    for i, (k, lt, anc) in enumerate(mapping):
        r = start_row + i
        ws.cell(r, 1).value = k
        ws.cell(r, 2).value = lt
        ws.cell(r, 3).value = anc

    # Extend range to allow future rows (add new Submit To mappings without rewriting formulas)
    map_range = "$A$21:$C$40"

    # ---------------- Sheets: OFCO_Req_1_15 / NOC_Req_1_6 ----------------
    def patch_req_sheet(sheet_name: str):
        wsx = wb[sheet_name]
        # Insert 3 columns at I (position 9)
        wsx.insert_cols(9, amount=3)

        # Update headers row 2
        wsx.cell(2, 9).value = "Lead Time Default (WD)"
        wsx.cell(2, 10).value = "Lead Time Override (WD)"
        wsx.cell(2, 11).value = "Lead Time Effective (WD)"
        wsx.cell(2, 12).value = "Anchor (Auto)"
        wsx.cell(2, 13).value = "Target Submit Date"

        # find last row with data in col A
        last = 2
        for r in range(3, 1000):
            if wsx.cell(r, 1).value is not None:
                last = r

        for r in range(3, last + 1):
            old_lead = wsx.cell(r, 12).value  # old I moved to L
            if old_lead is not None:
                wsx.cell(r, 10).value = old_lead  # override

            # Default lead time
            wsx.cell(r, 9).value = f'=IFERROR(VLOOKUP($H{r},Inputs!{map_range},2,FALSE),"")'
            # Effective
            wsx.cell(r, 11).value = f'=IF($J{r}="",$I{r},$J{r})'
            # Anchor
            wsx.cell(r, 12).value = f'=IFERROR(VLOOKUP($H{r},Inputs!{map_range},3,FALSE),"RoRoStart")'
            # Target
            wsx.cell(r, 13).value = (
                f'=IF($K{r}="","",WORKDAY.INTL('
                f'IF($L{r}="Arrival",Inputs!$B$3,IF($L{r}="Departure",Inputs!$B$6,IF($L{r}="RoRoEnd",Inputs!$B$5,Inputs!$B$4))),'
                f'-$K{r},Inputs!$B$7))'
            )

        # Basic data validations (Status/Risk/Owner/Submit To)
        wsx.data_validations.dataValidation = []
        dv_status = DataValidation(type="list", formula1='"Missing,Review,Submitted,Approved,Not Required"')
        dv_risk = DataValidation(type="list", formula1='"CRITICAL,HIGH,MED,LOW"')
        dv_owner = DataValidation(type="list", formula1='"OFCO,MMT,ALS,SCT,KFS,MWS,TBD"')
        dv_submit = DataValidation(type="list", formula1='"HSE,HM,Customs,AD Maritime,Maqta Gateway,Port Ops,TBD"')

        wsx.add_data_validation(dv_status)
        wsx.add_data_validation(dv_risk)
        wsx.add_data_validation(dv_owner)
        wsx.add_data_validation(dv_submit)

        dv_status.add("D3:D500")
        dv_risk.add("F3:F500")
        dv_owner.add("G3:G500")
        dv_submit.add("H3:H500")

    patch_req_sheet("OFCO_Req_1_15")
    patch_req_sheet("NOC_Req_1_6")

    # ---------------- VBA_Module sheet ----------------
    # Create or update VBA_Module sheet
    if "VBA_Module" in wb.sheetnames:
        ws_vba = wb["VBA_Module"]
        ws_vba.delete_rows(1, ws_vba.max_row)  # Clear existing content
    else:
        ws_vba = wb.create_sheet("VBA_Module")
    
    # Write change log header
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = 1
    ws_vba.cell(row, 1).value = f"v3.1 Operational - {timestamp}"
    row += 1
    ws_vba.cell(row, 1).value = "=" * 50
    row += 1
    ws_vba.cell(row, 1).value = ""
    row += 1
    
    # Write change log items
    change_items = [
        "주요 변경사항:",
        "",
        "1. Schedule Scenario 선택 기능 추가 (Inputs!B9)",
        "   - SCN-01 ~ SCN-05, CUSTOM 옵션",
        "   - 시나리오 테이블 (A12:E17)",
        "",
        "2. Manual Date 입력 컬럼 추가 (Inputs Column C)",
        "   - Auto 날짜 공식 (Column B) + Manual 입력 (Column C)",
        "   - CUSTOM 선택 시 Manual 값 사용",
        "",
        "3. Lead Time & Anchor 매핑 테이블 추가 (Inputs A21:C40)",
        "   - Submit To → Default Lead Time → Anchor Key 매핑",
        "   - HSE, HM, Maqta Gateway, Port Ops, Customs, AD Maritime",
        "",
        "4. OFCO_Req_1_15 / NOC_Req_1_6 시트 확장",
        "   - Lead Time Default (WD) - Column I",
        "   - Lead Time Override (WD) - Column J",
        "   - Lead Time Effective (WD) - Column K",
        "   - Anchor (Auto) - Column L",
        "   - Target Submit Date - Column M",
        "",
        "5. VBA 매크로 추가",
        "   - RefreshAll: 전체 재계산",
        "   - FilterMissing: Missing 상태 필터",
        "   - ClearFilters: 필터 제거",
        "   - StampUpdated: 업데이트 시간 기록",
        "   - ApplyScenarioToManual: 시나리오 → 수동 복사",
        "",
        "사용법:",
        "- VBA Editor에서 이 시트의 매크로 코드를 복사하여 모듈에 붙여넣기",
        "- 또는 별도 .bas 파일로 저장 후 Import",
    ]
    
    for item in change_items:
        ws_vba.cell(row, 1).value = item
        row += 1
    
    # Write VBA macros section
    row += 1
    ws_vba.cell(row, 1).value = "=" * 50
    row += 1
    ws_vba.cell(row, 1).value = "VBA MACRO CODE"
    row += 1
    ws_vba.cell(row, 1).value = "=" * 50
    row += 1
    ws_vba.cell(row, 1).value = "Copy the code below and paste into VBA Editor Module:"
    row += 1
    ws_vba.cell(row, 1).value = ""
    row += 1
    
    # Write VBA code line by line
    vba_lines = VBA_MACROS.split('\n')
    for line in vba_lines:
        ws_vba.cell(row, 1).value = line
        row += 1
    
    # Format: Make column A wider and wrap text
    ws_vba.column_dimensions['A'].width = 100
    for row_num in range(1, ws_vba.max_row + 1):
        cell = ws_vba.cell(row_num, 1)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Hide the sheet (optional - users can unhide if needed)
    # ws_vba.sheet_state = 'hidden'

    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", dest="out", required=True)
    args = ap.parse_args()
    apply_v3_1(args.src, args.out)


if __name__ == "__main__":
    main()
