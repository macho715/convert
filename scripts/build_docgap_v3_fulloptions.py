"""build_docgap_v3_fulloptions.py

Creates a 'full options' Doc Gap Tracker (v3) from v2 using openpyxl.

- Adds recommended Owner / Submit To values (unknowns as 'TBD')
- Adds dropdowns (Status/Owner/Submit To/Risk)
- Adds conditional formatting for Status and Risk
- Adds additional operational items (Pre-arrival meeting, Firewatcher, Ramp cert, Sea fastening cert)
- Adds an AD Maritime NOC summary row (already present in v2, but ensures Owner/Lead time)
- Adds Inputs schedule (Arrival, RoRo start/end, Departure)

Usage:
  python build_docgap_v3_fulloptions.py \
    --src OFCO_AGI_TR1_DocGap_Tracker_v2.xlsx \
    --out_xlsx OFCO_AGI_TR1_DocGap_Tracker_v3_FULLOPTIONS.xlsx \
    --out_xlsm OFCO_AGI_TR1_DocGap_Tracker_v3_FULLOPTIONS.xlsm

Note on VBA macros:
- This script does not embed VBA projects into the workbook.
- It writes an import-ready .bas file (DocGapMacros.bas) which you can import in Excel VBA Editor.
"""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


def style_table(ws, header_row: int, first_data_row: int, last_row: int, col_count: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for c in range(1, col_count + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Body
    for r in range(first_data_row, last_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if c in (2, 5, 11):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = ws.cell(first_data_row, 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{last_row}"


def add_dropdown(ws, col_letter: str, first_row: int, last_row: int, values: list[str]) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def add_conditional_formats(ws, last_row: int) -> None:
    # Status col D
    ws.conditional_formatting.add(
        f"D3:D{last_row}",
        FormulaRule(formula=["$D3=\"Missing\""], fill=PatternFill("solid", fgColor="F8CBAD")),
    )
    ws.conditional_formatting.add(
        f"D3:D{last_row}",
        FormulaRule(formula=["$D3=\"Review\""], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    ws.conditional_formatting.add(
        f"D3:D{last_row}",
        FormulaRule(formula=["$D3=\"Submitted\""], fill=PatternFill("solid", fgColor="C6E0B4")),
    )
    ws.conditional_formatting.add(
        f"D3:D{last_row}",
        FormulaRule(formula=["$D3=\"Approved\""], fill=PatternFill("solid", fgColor="A9D18E")),
    )

    # Risk col F
    ws.conditional_formatting.add(
        f"F3:F{last_row}",
        FormulaRule(formula=["$F3=\"CRITICAL\""], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"F3:F{last_row}",
        FormulaRule(formula=["$F3=\"HIGH\""], fill=PatternFill("solid", fgColor="F8CBAD")),
    )


def write_vba_assets(out_dir: Path) -> Path:
    bas = out_dir / "DocGapMacros.bas"
    bas.write_text(
        """Option Explicit\n\n'
' Doc Gap Tracker Macros (Import this module into Excel VBA Editor)\n'
\nSub RefreshAll()\n    Application.CalculateFull\nEnd Sub\n\nSub FilterMissing()\n    Dim ws As Worksheet\n    Set ws = ActiveSheet\n    If ws.AutoFilterMode = False Then Exit Sub\n    ws.Range(\"A2\").AutoFilter Field:=4, Criteria1:=\"Missing\"\nEnd Sub\n\nSub ClearFilters()\n    Dim ws As Worksheet\n    Set ws = ActiveSheet\n    If ws.AutoFilterMode Then\n        On Error Resume Next\n        ws.ShowAllData\n        On Error GoTo 0\n    End If\nEnd Sub\n\nSub StampUpdated()\n    With ThisWorkbook.Sheets(\"Executive_Summary\")\n        .Range(\"C1\").Value = \"Last updated: \" & Format(Now, \"dd-mmm-yy hh:nn\")\n    End With\nEnd Sub\n""",
        encoding="utf-8",
    )
    return bas


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    ap.add_argument("--out_xlsx", required=True)
    ap.add_argument("--out_xlsm", required=True)
    args = ap.parse_args()

    src = Path(args.src)
    out_xlsx = Path(args.out_xlsx)
    out_xlsm = Path(args.out_xlsm)

    wb = load_workbook(src)

    # ---- Inputs schedule
    inputs = wb["Inputs"]
    inputs["A3"].value, inputs["B3"].value = "Vessel Arrival to Mina Zayed (UTC+4)", date(2026, 1, 27)
    inputs["A4"].value, inputs["B4"].value = "RoRo / Load-out Start (UTC+4)", date(2026, 1, 29)
    inputs["A5"].value, inputs["B5"].value = "RoRo / Load-out End (UTC+4)", date(2026, 1, 30)
    inputs["A6"].value, inputs["B6"].value = "Vessel Departure from Mina Zayed (UTC+4)", date(2026, 2, 1)
    for r in range(3, 7):
        inputs[f"B{r}"].number_format = "dd-mmm-yy"

    # ---- Owners/SubmitTo recommendations
    owners = ["OFCO", "MMT", "ALS", "SCT", "KFS", "MWS", "TBD"]
    submit_to = ["HSE", "HM", "Customs", "AD Maritime", "Maqta Gateway", "Port Ops", "TBD"]
    status_vals = ["Missing", "Review", "Submitted", "Approved", "Not Required", "TBD"]
    risk_vals = ["LOW", "MED", "HIGH", "CRITICAL", "TBD"]

    ofco_map = {
        1: ("MMT", "HSE", 2),
        2: ("SCT", "HSE", 1),
        3: ("OFCO", "Maqta Gateway", 3),
        4: ("ALS", "HM", 1),
        5: ("MMT", "HSE", 2),
        6: ("MMT", "HSE", 2),
        7: ("MMT", "Port Ops", 1),
        8: ("SCT", "Port Ops", 1),
        9: ("MWS", "HM", 2),
        10: ("TBD", "HSE", 2),
        11: ("MWS", "HM", 2),
        12: ("MWS", "HM", 2),
        13: ("SCT", "Port Ops", 1),
        14: ("MMT", "HM", 1),
        15: ("SCT", "Port Ops", 1),
    }

    ws = wb["OFCO_Req_1_15"]
    header_row = 2
    first_data = 3

    # Add extra rows if missing
    extra_rows = [
        ("ADD-01", "Pre-arrival meeting (Port Authorities / operations planning)", "Yes", "Missing", None, "HIGH", "OFCO", "Port Ops", 2, None, "Meeting to be arranged prior to arrival."),
        ("ADD-02", "Firewatcher Certificate", "Yes", "Missing", None, "HIGH", "MMT", "HSE", 1, None, "OFCO requested Firewatcher certificate."),
        ("ADD-03", "Ramp Certificate / Linkspan Load Rating", "Yes", "Missing", None, "HIGH", "MMT", "HM", 2, None, "Often requested by port for ramp/linkspan operations."),
        ("ADD-04", "Sea Fastening Certificate (if separate from lashing plan)", "Yes", "Missing", None, "MED", "MMT", "HM", 2, None, "If separate certificate required by surveyor/port."),
    ]

    # Ensure extra rows exist (v2 ends at 15)
    if ws.max_row < first_data + 15 - 1 + len(extra_rows):
        start_row = ws.max_row + 1
        for i, row in enumerate(extra_rows):
            for c, v in enumerate(row, start=1):
                ws.cell(start_row + i, c).value = v

    # Apply mappings + formulas
    last_row = ws.max_row
    for r in range(first_data, last_row + 1):
        no = ws.cell(r, 1).value
        if isinstance(no, int) and no in ofco_map:
            owner, submit, lt = ofco_map[no]
            ws.cell(r, 7).value = owner
            ws.cell(r, 8).value = submit
            if ws.cell(r, 9).value in (None, ""):
                ws.cell(r, 9).value = lt

        # Target submit date = WORKDAY(RoRo Start, -LeadTime)
        ws.cell(r, 10).value = f'=IF($I{r}="","",WORKDAY(Inputs!$B$4,-$I{r}))'
        ws.cell(r, 10).number_format = "dd-mmm-yy"

    # Styling / validations
    style_table(ws, header_row, first_data, last_row, col_count=11)
    add_dropdown(ws, "D", first_data, last_row, status_vals)
    add_dropdown(ws, "F", first_data, last_row, risk_vals)
    add_dropdown(ws, "G", first_data, last_row, owners)
    add_dropdown(ws, "H", first_data, last_row, submit_to)
    add_conditional_formats(ws, last_row)

    # ---- NOC sheet
    wsn = wb["NOC_Req_1_6"]
    last_n = wsn.max_row
    for r in range(3, last_n + 1):
        wsn.cell(r, 7).value = "SCT" if wsn.cell(r, 1).value else None
        wsn.cell(r, 8).value = "AD Maritime" if wsn.cell(r, 1).value else None
        if wsn.cell(r, 9).value in (None, ""):
            wsn.cell(r, 9).value = 5
        # Target submit date = WORKDAY(Departure, -LeadTime)
        wsn.cell(r, 10).value = f'=IF($I{r}="","",WORKDAY(Inputs!$B$6,-$I{r}))'
        wsn.cell(r, 10).number_format = "dd-mmm-yy"

    style_table(wsn, 2, 3, last_n, col_count=11)
    add_dropdown(wsn, "D", 3, last_n, status_vals)
    add_dropdown(wsn, "F", 3, last_n, risk_vals)
    add_dropdown(wsn, "G", 3, last_n, owners)
    add_dropdown(wsn, "H", 3, last_n, submit_to)
    add_conditional_formats(wsn, last_n)

    # ---- VBA sheet (text only)
    if "VBA_Module" not in wb.sheetnames:
        vb = wb.create_sheet("VBA_Module")
    else:
        vb = wb["VBA_Module"]

    vb["A1"].value = "VBA code (import-ready). Save the attached DocGapMacros.bas and import it in Excel (Alt+F11 > File > Import File)."
    vb["A1"].font = Font(bold=True)

    # Save outputs
    wb.save(out_xlsx)
    wb.save(out_xlsm)

    bas_path = write_vba_assets(out_xlsx.parent)
    print(f"Saved: {out_xlsx}")
    print(f"Saved: {out_xlsm}")
    print(f"Saved VBA module: {bas_path}")


if __name__ == "__main__":
    main()
