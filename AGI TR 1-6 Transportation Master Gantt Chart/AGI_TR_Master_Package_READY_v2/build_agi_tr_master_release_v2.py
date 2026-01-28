#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_agi_tr_master_release_v2.py

Builds AGI TR1-TR6 master schedule workbook (Scenario 1-2-2-1 / JD x2 @ 3 units)
and optionally embeds VBA (Windows + Excel + pywin32 required).

Default files in this package:
  - Template: AGI_TR_Master_RELEASE_v2.xlsx
  - VBA:      AGI_TR_Master_PATCHED_v2.bas

Usage (xlsx only):
  python build_agi_tr_master_release_v2.py --d0 2026-01-09

Usage (create .xlsm with VBA embedded):
  python build_agi_tr_master_release_v2.py --d0 2026-01-09 --embed-vba

Notes
  - Embedding VBA requires:
      * Windows OS
      * Microsoft Excel installed
      * pywin32 installed: pip install pywin32
      * Excel Trust Center: enable "Trust access to the VBA project object model"
"""

from __future__ import annotations

import argparse
import datetime as dt
import platform
import shutil
from pathlib import Path

from openpyxl import load_workbook


def parse_date(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%Y-%m-%d").date()


def update_control_panel_dates(wb_path: Path, d0: dt.date) -> None:
    wb = load_workbook(wb_path)
    if "Control_Panel" not in wb.sheetnames:
        raise RuntimeError("Template missing Control_Panel sheet")
    ws = wb["Control_Panel"]
    ws["C5"].value = d0
    # default Shamal HIGH window (planner can edit)
    ws["C18"].value = dt.date(d0.year, 1, 14)
    ws["C19"].value = dt.date(d0.year, 1, 18)
    wb.save(wb_path)


def embed_vba_xlsm(xlsx_path: Path, bas_path: Path, xlsm_out: Path) -> None:
    if platform.system().lower() != "windows":
        raise RuntimeError("--embed-vba requires Windows + Microsoft Excel")

    try:
        import win32com.client  # type: ignore
    except Exception as e:
        raise RuntimeError("pywin32 is required (pip install pywin32)") from e

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))

        vbproj = wb.VBProject
        vbproj.VBComponents.Import(str(bas_path.resolve()))

        # Add Workbook_Open hook (keyboard shortcuts)
        thisworkbook = vbproj.VBComponents("ThisWorkbook")
        code_mod = thisworkbook.CodeModule
        existing = code_mod.Lines(1, code_mod.CountOfLines)

        if "Workbook_Open" not in existing:
            code_mod.AddFromString(
                "Private Sub Workbook_Open()\n"
                "    On Error Resume Next\n"
                "    AGI_TR_Master.SetupKeyboardShortcuts\n"
                "End Sub\n"
            )

        wb.SaveAs(str(xlsm_out.resolve()), FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
        wb.Close(SaveChanges=False)

    finally:
        excel.Quit()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--d0", default="2026-01-09", help="D0 date (YYYY-MM-DD)")
    ap.add_argument("--template", default="AGI_TR_Master_RELEASE_v2.xlsx", help="Template workbook (.xlsx)")
    ap.add_argument("--vba", default="AGI_TR_Master_PATCHED_v2.bas", help="VBA module (.bas)")
    ap.add_argument("--out-xlsx", default="AGI_TR_Master_RELEASE_OUT.xlsx", help="Output .xlsx path")
    ap.add_argument("--embed-vba", action="store_true", help="Create .xlsm by embedding VBA (Windows+Excel only)")
    ap.add_argument("--out-xlsm", default="AGI_TR_Master_RELEASE_OUT.xlsm", help="Output .xlsm path (when --embed-vba)")
    args = ap.parse_args()

    d0 = parse_date(args.d0)
    template = Path(args.template)
    vba = Path(args.vba)
    out_xlsx = Path(args.out_xlsx)

    if not template.exists():
        raise SystemExit(f"Template not found: {template}")
    if not vba.exists():
        raise SystemExit(f"VBA module not found: {vba}")

    shutil.copyfile(template, out_xlsx)
    update_control_panel_dates(out_xlsx, d0)
    print(f"Wrote: {out_xlsx}")

    if args.embed_vba:
        out_xlsm = Path(args.out_xlsm)
        embed_vba_xlsm(out_xlsx, vba, out_xlsm)
        print(f"Wrote: {out_xlsm}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
