# Py3.11.8 / pandas2.33 / openpyxl3.1.5
"""
JPT71 - Refresh All
- Input:  Plan sheet (tblPlan + INPUT block in column Q/R)
- Output: New workbook file (recommended) to avoid "file is open/locked" issues.

Usage:
  py jpt71_refresh_all.py "C:\Path\JPT71.xlsx" "C:\Path\JPT71_REFRESHED.xlsx"
"""

import sys
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


# -----------------------------
# Helpers
# -----------------------------
def _to_date(x):
    if x is None or x == "":
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def _thin_border(color="BFBFBF"):
    thin = Side(style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_cross_gantt(wb, df_plan, delay_ref="Plan!$R$4", inprog_trip="Debris-8", view_days_after=14):
    """
    Rebuilds Cross_Gantt sheet based on df_plan columns:
      Trip, Type, Material,
      Plan_MW4_Depart_Agg, Plan_AGI_Offload_Agg, Plan_AGI_Debris_Load, Plan_MW4_Debris_Offload
    """
    # Remove existing sheet
    if "Cross_Gantt" in wb.sheetnames:
        ws_old = wb["Cross_Gantt"]
        wb.remove(ws_old)

    ws = wb.create_sheet("Cross_Gantt")

    thin_border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    fill_row_agg = PatternFill("solid", fgColor="E2EFDA")
    fill_row_deb = PatternFill("solid", fgColor="FCE4D6")
    fill_inprog = PatternFill("solid", fgColor="D9D9D9")

    fill_mw4 = PatternFill("solid", fgColor="4F81BD")
    fill_agi_off = PatternFill("solid", fgColor="C6EFCE")
    fill_agi_debload = PatternFill("solid", fgColor="F4B084")
    fill_mw4_deboff = PatternFill("solid", fgColor="F8CBAD")

    dash = Side(style="dashed", color="7F7F7F")
    border_dashed = Border(left=dash, right=dash, top=dash, bottom=dash)

    headers = [
        "Seq","Trip","Type","Material",
        "MW4 Depart (Agg)","AGI Offload (Agg)",
        "AGI Debris Loading (Deb)","MW4 Debris Offloading (Deb)",
        "Status"
    ]
    hidden_headers = ["ShiftFlag","Plan_MW4_Depart","Plan_AGI_Offload","Plan_AGI_Deb_Load","Plan_MW4_Deb_Off"]

    # Build plan list & detect in-progress seq
    df = df_plan.copy()

    # Force columns exist
    need_cols = ["Trip","Type","Material",
                 "Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]
    for c in need_cols:
        if c not in df.columns:
            df[c] = None

    # Parse dates
    for c in ["Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]:
        df[c] = df[c].apply(_to_date)

    df = df[df["Trip"].notna() & (df["Trip"].astype(str).str.strip() != "")]
    df.reset_index(drop=True, inplace=True)

    if df.empty:
        ws["A1"] = "No plan rows in Plan sheet."
        return

    # Determine start/end for timeline
    all_dates = []
    for c in ["Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]:
        all_dates += [d for d in df[c].tolist() if d]
    plan_start = min(all_dates)
    plan_end = max(all_dates)
    view_end = plan_end + timedelta(days=view_days_after)

    dates = []
    d = plan_start
    while d <= view_end:
        dates.append(d)
        d += timedelta(days=1)

    # Find in-progress row index (1-based seq)
    inprog_idx = None
    for i, trip in enumerate(df["Trip"].astype(str).tolist(), start=1):
        if trip.strip() == inprog_trip:
            inprog_idx = i
            break
    if inprog_idx is None:
        inprog_idx = 999999  # no shifting applies

    # Header row
    col = 1
    for h in headers:
        c = ws.cell(1, col, h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        col += 1
    for h in hidden_headers:
        c = ws.cell(1, col, h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        col += 1

    date_start_col = col
    for i, dt in enumerate(dates):
        c = ws.cell(1, date_start_col + i, dt)
        c.number_format = "mm-dd"
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = thin_border
        ws.column_dimensions[get_column_letter(date_start_col + i)].width = 5

    # Column widths (visible)
    for c,w in {1:5,2:12,3:10,4:14,5:14,6:14,7:18,8:20,9:14}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Hide plan/shift cols
    shift_col = len(headers) + 1
    for c in range(shift_col, date_start_col):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    ws.freeze_panes = "J2"
    ws.row_dimensions[1].height = 22

    # Populate rows
    for idx, row in enumerate(df.itertuples(index=False), start=1):
        r = 1 + idx
        trip = str(row.Trip)
        typ = str(row.Type)
        mat = str(row.Material)

        p_mw4 = getattr(row, "Plan_MW4_Depart_Agg")
        p_agi_off = getattr(row, "Plan_AGI_Offload_Agg")
        p_debload = getattr(row, "Plan_AGI_Debris_Load")
        p_deboff = getattr(row, "Plan_MW4_Debris_Offload")

        row_fill = fill_row_deb if typ.strip().lower() == "debris" else fill_row_agg

        ws.cell(r,1,idx).alignment=center
        ws.cell(r,2,trip).alignment=center
        ws.cell(r,3,typ).alignment=center
        ws.cell(r,4,mat).alignment=left
        ws.cell(r,9,"IN PROGRESS" if trip.strip()==inprog_trip else "").alignment=center

        # shift flag (rows AFTER in-progress)
        ws.cell(r, shift_col, 1 if idx > inprog_idx else 0).alignment=center

        plan_start_col = shift_col + 1
        plan_vals = [p_mw4, p_agi_off, p_debload, p_deboff]
        for j,val in enumerate(plan_vals):
            c = ws.cell(r, plan_start_col + j, val)
            if val:
                c.number_format = "yyyy-mm-dd"
            c.alignment = center

        flag_ref = f"${get_column_letter(shift_col)}{r}"
        plan_ref = [
            f"{get_column_letter(plan_start_col)}{r}",
            f"{get_column_letter(plan_start_col+1)}{r}",
            f"{get_column_letter(plan_start_col+2)}{r}",
            f"{get_column_letter(plan_start_col+3)}{r}",
        ]

        def adj(plan_cell_ref: str):
            return f'=IF({plan_cell_ref}="", "", IF({flag_ref}=1, {plan_cell_ref}+{delay_ref}, {plan_cell_ref}))'

        ws.cell(r,5).value = adj(plan_ref[0])
        ws.cell(r,6).value = adj(plan_ref[1])
        ws.cell(r,7).value = adj(plan_ref[2])
        ws.cell(r,8).value = adj(plan_ref[3])
        for c in [5,6,7,8]:
            ws.cell(r,c).number_format="yyyy-mm-dd"
            ws.cell(r,c).alignment=center

        # style visible cols
        for c in range(1,10):
            cell = ws.cell(r,c)
            cell.border = thin_border
            cell.fill = fill_inprog if trip.strip()==inprog_trip else row_fill

        # dashed row for in-progress
        if trip.strip() == inprog_trip:
            for c in range(1,10):
                ws.cell(r,c).border = border_dashed

    # Conditional formatting on timeline (exclude in-progress row)
    first_data = 2
    last_data = 1 + len(df)
    first_date = date_start_col
    last_date = date_start_col + len(dates) - 1
    inprog_row = 1 + inprog_idx

    def add_cf(r1, r2, fill, col_letter):
        if r1 > r2:
            return
        rng = f"{get_column_letter(first_date)}{r1}:{get_column_letter(last_date)}{r2}"
        top_left = get_column_letter(first_date)
        # formula anchored to first row in the range; Excel will adjust row references down
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'={top_left}$1=${col_letter}{r1}'], fill=fill)
        )

    for (r1,r2) in [(first_data, inprog_row-1), (inprog_row+1, last_data)]:
        add_cf(r1,r2,fill_mw4,'E')
        add_cf(r1,r2,fill_agi_off,'F')
        add_cf(r1,r2,fill_agi_debload,'G')
        add_cf(r1,r2,fill_mw4_deboff,'H')

    # Grey timeline for in-progress row
    for c in range(first_date, last_date+1):
        cell = ws.cell(inprog_row, c)
        cell.fill = fill_inprog
        cell.border = border_dashed

    # Ensure borders in whole used range (preserve dashed)
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell=ws.cell(r,c)
            styles=[cell.border.left.style, cell.border.right.style, cell.border.top.style, cell.border.bottom.style]
            if "dashed" in styles:
                continue
            cell.border = thin_border


def main(in_path: str, out_path: str):
    # Read Plan table area (A:O) with pandas
    df = pd.read_excel(in_path, sheet_name="Plan", usecols="A:O", dtype=str)
    # Keep only relevant columns; if user renamed columns, adjust here
    # (We keep extra columns but Cross_Gantt uses only needed ones)
    wb = load_workbook(in_path)

    build_cross_gantt(wb, df_plan=df, delay_ref="Plan!$R$4", inprog_trip="Debris-8", view_days_after=14)

    wb.save(out_path)
    print("Saved:", out_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit('Usage: py jpt71_refresh_all.py "IN.xlsx" "OUT.xlsx"')
    main(sys.argv[1], sys.argv[2])
