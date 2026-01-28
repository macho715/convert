#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_agi_tr_master_release.py

Creates a ready-to-use AGI TR1-TR6 master schedule workbook for:
  - Scenario: 1-2-2-1 voyages
  - Jack-down batches: 2 (3 units per batch)

Outputs
  - .xlsx (always)
  - .xlsm (optional, Windows+Excel only) with VBA injected from a .bas module

Usage
  python build_agi_tr_master_release.py \
    --d0 2026-01-09 \
    --template AGI_TR_Master_READY_1-2-2-1_JD3.xlsx \
    --vba AGI_TR_Master.bas \
    --out AGI_TR_Master_RELEASE

Optional
  --embed-vba --xlsm-out AGI_TR_Master_RELEASE.xlsm

Notes
  - Embedding VBA requires Windows + Microsoft Excel + "Trust access to the VBA project object model".
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import platform
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_date(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%Y-%m-%d").date()


def ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def update_d0(wb_path: Path, d0: dt.date) -> None:
    wb = load_workbook(wb_path)
    if "Control_Panel" not in wb.sheetnames:
        raise RuntimeError("Template missing Control_Panel sheet")
    ws = wb["Control_Panel"]
    ws["C5"].value = d0
    # default shamal window (planner can edit)
    ws["C18"].value = dt.date(d0.year, 1, 14)
    ws["C19"].value = dt.date(d0.year, 1, 18)
    wb.save(wb_path)


def try_embed_vba_xlsm(xlsx_path: Path, bas_path: Path, xlsm_out: Path) -> None:
    """Embed VBA using Excel COM automation (Windows-only)."""
    if platform.system().lower() != "windows":
        raise RuntimeError("--embed-vba requires Windows + Excel")

    try:
        import win32com.client  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "pywin32 is required for --embed-vba (pip install pywin32)"
        ) from e

    xlsx_path = xlsx_path.resolve()
    bas_path = bas_path.resolve()
    xlsm_out = xlsm_out.resolve()

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(str(xlsx_path))

        # Import .bas module
        vbproj = wb.VBProject
        vbproj.VBComponents.Import(str(bas_path))

        # Add Workbook_Open hook (optional): create if not present
        # NOTE: This writes into ThisWorkbook module.
        thisworkbook = vbproj.VBComponents("ThisWorkbook")
        code_mod = thisworkbook.CodeModule
        existing = code_mod.Lines(1, code_mod.CountOfLines)
        if "Workbook_Open" not in existing:
            code = (
                "Private Sub Workbook_Open()\n"
                "    On Error Resume Next\n"
                "    AGI_TR_Master.SetupKeyboardShortcuts\n"
                "End Sub\n"
            )
            code_mod.AddFromString(code)

        # Save as xlsm
        ensure_parent(xlsm_out)
        wb.SaveAs(str(xlsm_out), FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
        wb.Close(SaveChanges=False)

    finally:
        excel.Quit()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--d0", default="2026-01-09", help="D0 date (YYYY-MM-DD)")
    ap.add_argument("--template", default="AGI_TR_Master_READY_1-2-2-1_JD3.xlsx")
    ap.add_argument("--vba", default="AGI_TR_Master.bas")
    ap.add_argument("--out", default="AGI_TR_Master_RELEASE.xlsx", help="Output .xlsx path")
    ap.add_argument("--embed-vba", action="store_true", help="Create .xlsm by embedding VBA (Windows+Excel only)")
    ap.add_argument("--xlsm-out", default="AGI_TR_Master_RELEASE.xlsm", help="Output .xlsm path (used with --embed-vba)")
    args = ap.parse_args()

    d0 = parse_date(args.d0)
    template = Path(args.template)
    vba = Path(args.vba)
    out_xlsx = Path(args.out)

    if not template.exists():
        raise SystemExit(f"Template not found: {template}")
    if not vba.exists():
        raise SystemExit(f"VBA module not found: {vba}")

    ensure_parent(out_xlsx)
    shutil.copyfile(template, out_xlsx)
    update_d0(out_xlsx, d0)

    print(f"Wrote: {out_xlsx}")

    if args.embed_vba:
        xlsm_out = Path(args.xlsm_out)
        try_embed_vba_xlsm(out_xlsx, vba, xlsm_out)
        print(f"Wrote: {xlsm_out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
