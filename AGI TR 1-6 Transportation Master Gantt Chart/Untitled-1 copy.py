#!/usr/bin/env python3
"""
AGI TR 7-Voyage Master Gantt with VBA
기존 AGI_TR_7Voyage_Master_Gantt.xlsx와 동일한 레이아웃 + VBA 기능
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# === COLORS (기존과 동일) ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def load_tasks_from_tsv(tsv_path, project_start_date):
    """TSV file loader with flexible headers.

    Returns tuples in the form:
        (task_id, wbs, task_name, phase, owner, offset_days, duration_ref, notes, type_a, type_b, activity_id)

    Duration is only applied when a valid Activity ID exists.
    """

    import csv
    from datetime import datetime

    tasks: list[tuple] = []

    if hasattr(project_start_date, "date"):
        project_start_date = project_start_date.date()

    def infer_phase_from_task(task_name: str) -> str:
        task_upper = task_name.upper()
        if 'MOBILIZATION' in task_upper or 'DEMOBILIZATION' in task_upper or 'DEMOB' in task_upper:
            return 'MOBILIZATION'
        if 'DECK' in task_upper and ('PREP' in task_upper or 'PREPARATION' in task_upper):
            return 'DECK_PREP'
        if 'LOADOUT' in task_upper or 'LOAD-OUT' in task_upper or 'LOAD IN' in task_upper or 'LOAD-IN' in task_upper:
            return 'LOADOUT'
        if 'SEA FASTENING' in task_upper or 'SEAFAST' in task_upper or 'SEAFASTENING' in task_upper:
            return 'SEAFAST'
        if 'SAIL' in task_upper or 'SAIL-AWAY' in task_upper:
            return 'SAIL'
        if 'UNLOAD' in task_upper or 'ARRIVAL' in task_upper:
            return 'AGI_UNLOAD'
        if 'TURNING' in task_upper or 'TURN' in task_upper:
            return 'TURNING'
        if 'JACKDOWN' in task_upper or 'JACK-DOWN' in task_upper or 'JACKING DOWN' in task_upper:
            return 'JACKDOWN'
        if 'RETURN' in task_upper:
            return 'RETURN'
        if 'VOYAGE' in task_upper:
            return 'MILESTONE'
        if 'MILESTONE' in task_upper:
            return 'MILESTONE'
        return 'BUFFER'

    phase_mapping = {
        'Mobilization': 'MOBILIZATION',
        'Deck Prep': 'DECK_PREP',
        'MZP Loadout': 'LOADOUT',
        'Sea Fastening': 'SEAFAST',
        'Survey': 'BUFFER',
        'Sea Passage': 'SAIL',
        'AGI Arrival': 'AGI_UNLOAD',
        'AGI Laydown': 'BUFFER',
        'Onshore SPMT': 'TURNING',
        'AGI Gate Prep': 'TURNING',
        'Jackdown': 'JACKDOWN',
        'Return': 'RETURN',
        'Buffer': 'BUFFER',
        'Marine Transport': 'MILESTONE',
        'Demobilization': 'MOBILIZATION',
        'Handover': 'MILESTONE',
    }

    def get_duration_ref(duration_str: str, task_name: str, has_activity_id: bool = True):
        if not has_activity_id:
            return 0
        try:
            dur_val = float(duration_str)
            if dur_val == 0:
                return 0
            if dur_val == 0.5:
                return 'DUR_BUF'
            if dur_val == 1.0:
                task_upper = task_name.upper()
                if 'LOADOUT' in task_upper or 'LOAD-OUT' in task_upper:
                    return 'DUR_LO'
                if 'MOBILIZATION' in task_upper or 'DEMOBILIZATION' in task_upper:
                    return 'DUR_MOB'
                if 'SAIL' in task_upper:
                    return 'DUR_SAIL'
                if 'ARRIV' in task_upper or 'UNLOAD' in task_upper or 'LOAD-IN' in task_upper:
                    return 'DUR_UL'
                if 'RETURN' in task_upper:
                    return 'DUR_RET'
                if 'JACK' in task_upper:
                    return 'DUR_JD'
                return 'DUR_BUF'
            if dur_val == 2.0:
                return 2
            if dur_val == 3.0:
                task_upper = task_name.upper()
                if 'TURN' in task_upper:
                    return 'DUR_TURN'
                if 'DECK' in task_upper or 'PREP' in task_upper:
                    return 'DUR_DECK'
                return 3
            return dur_val
        except Exception:
            return 'DUR_BUF' if has_activity_id else 0

    def parse_date(date_str: str):
        s = (date_str or '').strip()
        if not s:
            return None

        for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d-%b-%y'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass

        try:
            d = datetime.strptime(s, '%d-%b').date()
            return d.replace(year=project_start_date.year)
        except Exception:
            return None

    def iter_rows_preserve_headers(path: str):
        with open(path, 'r', encoding='utf-8', newline='') as f:
            # Try tab delimiter first
            content = f.read()
            f.seek(0)
            
            # Check if file uses tabs or spaces
            first_line = f.readline()
            f.seek(0)
            
            # If first line contains "Samsung" or "HVDC", likely space-delimited option c format
            is_option_c_format = 'Samsung' in first_line or 'HVDC' in first_line
            
            if is_option_c_format:
                # Use space delimiter and split by multiple spaces
                reader = csv.reader(f, delimiter='\t')
                first_line_parts = first_line.strip().split()
                # Skip project info line
                next(reader)
                
                # No header, use default header for option c format
                headers = ['Activity ID', 'Activity Name', 'Original Duration', 'Planned Start', 'Planned Finish']
                
                # Process remaining rows
                for line in f:
                    if not line.strip():
                        continue
                    # Split by multiple spaces (but preserve single spaces in activity names)
                    parts = line.strip().split('\t')
                    if not parts or len(parts) < 2:
                        # Try space splitting
                        parts = line.strip().split()
                        # Reconstruct: ID, Name (may have spaces), Duration, Start, End
                        if len(parts) >= 5:
                            # Find where duration starts (first numeric value after ID)
                            id_part = parts[0]
                            # Find duration (first float value)
                            dur_idx = None
                            for i, p in enumerate(parts[1:], 1):
                                try:
                                    float(p)
                                    dur_idx = i
                                    break
                                except:
                                    pass
                            
                            if dur_idx:
                                activity_id = id_part
                                activity_name = ' '.join(parts[1:dur_idx])
                                duration = parts[dur_idx] if dur_idx < len(parts) else ''
                                start_date = parts[dur_idx + 1] if dur_idx + 1 < len(parts) else ''
                                end_date = parts[dur_idx + 2] if dur_idx + 2 < len(parts) else ''
                                row = [activity_id, activity_name, duration, start_date, end_date]
                            else:
                                row = parts[:5] if len(parts) >= 5 else parts + [''] * (5 - len(parts))
                        else:
                            row = parts + [''] * (5 - len(parts))
                    else:
                        row = parts
                    
                    if len(row) < len(headers):
                        row = row + [''] * (len(headers) - len(row))
                    yield headers, row
            else:
                # Standard tab-delimited format
                reader = csv.reader(f, delimiter='	')
                first_line = next(reader)
                
                # Check if first line looks like a header (contains "Activity ID" or "Activity Name")
                is_header = any('Activity ID' in str(cell) or 'Activity Name' in str(cell) for cell in first_line)
                
                if is_header:
                    headers = first_line
                else:
                    # No header, use default header
                    headers = ['Activity ID', 'Activity Name', 'Original Duration', 'Planned Start', 'Planned Finish']
                    # Process first line as data
                    row = first_line
                    if len(row) < len(headers):
                        row = row + [''] * (len(headers) - len(row))
                    yield headers, row
                
                # Continue with remaining rows
                for row in reader:
                    if len(row) < len(headers):
                        row = row + [''] * (len(headers) - len(row))
                    yield headers, row

    task_counter = 0

    for headers, row in iter_rows_preserve_headers(tsv_path):
        header_lc = [h.strip() for h in headers]

        def idx(name: str):
            try:
                return header_lc.index(name)
            except ValueError:
                return None

        id_cols: list[int] = []

        for k in ('Activity ID (1)', 'Activity ID (2)', 'Activity ID (3)'):
            i = idx(k)
            if i is not None:
                id_cols.append(i)

        if not id_cols:
            dup = [i for i, h in enumerate(header_lc) if h == 'Activity ID']
            if dup:
                id_cols = dup[:3]

        if not id_cols:
            i = idx('Activity ID')
            if i is not None:
                id_cols = [i]
            else:
                i = idx('ID')
                if i is not None:
                    id_cols = [i]

        # For option c format (no header), try to detect columns by position
        # Format: Activity ID, Activity Name, Duration, Start Date, End Date
        if not id_cols and len(row) >= 5:
            # Check if first column looks like Activity ID (A followed by digits or category name)
            first_col = row[0].strip() if len(row) > 0 else ''
            if first_col and (first_col.upper().startswith('A') or first_col in ['MOBILIZATION', 'DEMOBILIZATION', 'OPERATIONAL', 'SPMT', 'MARINE', 'JACKING EQUIPMENT, STEEL BRIDGE', 'Beam Replacement', 'Deck Preparations', 'AGI TR Unit']):
                id_cols = [0]
                # Map columns for option c format
                name_i = 1 if len(row) > 1 else None
                dur_i = 2 if len(row) > 2 else None
                start_i = 3 if len(row) > 3 else None
                end_i = 4 if len(row) > 4 else None
            else:
                name_i = None
                dur_i = None
                start_i = None
                end_i = None
        else:
            name_i = None
            dur_i = None
            start_i = None
            end_i = None

        id_parts = [row[i].strip() for i in id_cols if i is not None and row[i].strip()]

        activity_id = ''
        for part in reversed(id_parts):
            candidate = part.strip()
            if candidate.upper().startswith('A') and len(candidate) > 1 and candidate[1:].isdigit():
                activity_id = candidate.upper()
                break

        if activity_id:
            task_id = activity_id
        elif id_parts:
            task_id = '-'.join([p.replace(' ', '_') for p in id_parts])
        else:
            task_id = ''

        type_a = id_parts[0] if len(id_parts) > 0 else ""
        type_b = id_parts[1] if len(id_parts) > 1 else ""

        # Get task name
        if name_i is not None:
            task_name = row[name_i].strip() if name_i < len(row) else ''
        else:
            name_i = idx('Activity Name')
            if name_i is None:
                name_i = idx('Task')
            task_name = row[name_i].strip() if name_i is not None and name_i < len(row) else ''
        
        if not task_name:
            continue

        # Get start date
        if start_i is not None:
            start_str = row[start_i].strip() if start_i < len(row) else ''
        else:
            start_candidates = []
            for k in ('Planned Start', 'Actual Start', 'Start'):
                i = idx(k)
                if i is not None:
                    start_candidates.append(row[i].strip())
            start_str = next((x for x in start_candidates if x), '')
        
        start_date = parse_date(start_str)
        if not start_date:
            continue

        # Get end date
        if end_i is not None:
            end_str = row[end_i].strip() if end_i < len(row) else ''
        else:
            end_candidates = []
            for k in ('Planned Finish', 'Actual Finish', 'End'):
                i = idx(k)
                if i is not None:
                    end_candidates.append(row[i].strip())
            end_str = next((x for x in end_candidates if x), '')
        end_date = parse_date(end_str) if end_str else None

        # Get duration
        if dur_i is not None:
            duration_str = row[dur_i].strip() if dur_i < len(row) else ''
        else:
            dur_candidates = []
            for k in ('Original Duration', 'Duration_days', 'Dur'):
                i = idx(k)
                if i is not None:
                    dur_candidates.append(row[i].strip())
            duration_str = next((x for x in dur_candidates if x), '')
        if not duration_str and end_date:
            duration_str = str((end_date - start_date).days)

        has_activity_id = bool(activity_id and activity_id[1:].isdigit())
        dur_ref = get_duration_ref(duration_str or '0', task_name, has_activity_id)

        phase_i = idx('Phase')
        phase_raw = row[phase_i].strip() if phase_i is not None else ''
        if phase_raw:
            phase = phase_mapping.get(phase_raw, infer_phase_from_task(task_name))
        else:
            phase = infer_phase_from_task(task_name)

        owner_i = idx('Owner')
        owner = row[owner_i].strip() if owner_i is not None else 'All'
        if not owner:
            owner = 'All'

        notes_i = idx('Notes')
        notes = row[notes_i].strip() if notes_i is not None else ''

        if not task_id:
            task_counter += 1
            task_id = f'TASK-{task_counter:04d}'

        wbs_i = idx('WBS')
        if wbs_i is not None and row[wbs_i].strip():
            wbs = row[wbs_i].strip()
        else:
            wbs = f'{len(tasks)+1:.1f}'

        offset = (start_date - project_start_date).days

        tasks.append((
            task_id,
            wbs,
            task_name,
            phase,
            owner,
            offset,
            dur_ref,
            notes,
            type_a,
            type_b,
            activity_id,
        ))

    return tasks

def load_tide_data_json(json_path):
    """
    조석 데이터 JSON 파일 로드
    """
    import json
    
    tide_records = []
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if 'tide_records' in data:
                for record in data['tide_records']:
                    date_str = record.get('date', '')
                    if not date_str:
                        continue
                    tide_records.append((
                        date_str,
                        record.get('high_tide_window', '').strip(),
                        float(record.get('max_height_m', 0)) if record.get('max_height_m') else 0.0,
                        record.get('risk_level', 'LOW').strip()
                    ))
    except Exception as e:
        print(f"Warning: Could not load tide data from JSON: {e}")
    
    return tide_records

def load_tide_data(tsv_path=None, json_path=None):
    """
    조석 데이터 로드 (TSV 또는 JSON 지원)
    """
    import csv
    import os
    
    if json_path and os.path.exists(json_path):
        return load_tide_data_json(json_path)
    
    tide_records = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            with open(tsv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    if not row.get('Date'):
                        continue
                    try:
                        date_str = row['Date'].strip()
                        tide_records.append((
                            date_str,
                            row.get('High Tide Window', '').strip(),
                            float(row.get('Max Height (m)', '0').strip()) if row.get('Max Height (m)') else 0.0,
                            row.get('Risk Level', 'LOW').strip()
                        ))
                    except:
                        continue
        except Exception as e:
            print(f"Warning: Could not load tide data from TSV: {e}")
    
    return tide_records

def calculate_max_days(tasks, project_start, wb=None, min_days=120, buffer_days=30):
    """
    작업 목록에서 최대 프로젝트 기간 계산
    
    Args:
        tasks: 작업 리스트
        project_start: 프로젝트 시작일
    
    Returns:
        최대 일수 (정수)
    """
    max_offset = 0
    for task in tasks:
        if len(task) < 6:
            continue
        offset = task[5] if isinstance(task[5], (int, float)) else 0
        dur_ref = task[6] if len(task) > 6 else 0

        activity_id = task[10] if len(task) > 10 else None
        if activity_id is not None:
            activity_id = str(activity_id).strip()
            if not (activity_id.upper().startswith("A") and activity_id[1:].isdigit()):
                continue

        duration = 0.0
        if isinstance(dur_ref, (int, float)):
            duration = float(dur_ref)
        elif isinstance(dur_ref, str):
            duration = None
            if wb:
                try:
                    dn = wb.defined_names.get(dur_ref)
                    if dn:
                        for title, coord in dn.destinations:
                            val = wb[title][coord].value
                            if isinstance(val, (int, float)):
                                duration = float(val)
                                break
                except Exception:
                    duration = None
            if duration is None:
                duration = 1.0

        if duration < 0:
            duration = 0
        max_offset = max(max_offset, offset + duration)
    
    if wb:
        try:
            ws_ctrl = wb["Control_Panel"]
            min_days_val = ws_ctrl["H8"].value
            buffer_days_val = ws_ctrl["H9"].value
            if isinstance(min_days_val, (int, float)):
                min_days = int(min_days_val)
            if isinstance(buffer_days_val, (int, float)):
                buffer_days = int(buffer_days_val)
        except Exception:
            pass
    
    # Duration을 고려하여 여유 있게 계산
    return max(min_days, int(max_offset) + buffer_days)

def calculate_voyage_ranges(tasks):
    """
    tasks에서 각 Voyage의 Day 범위를 계산
    
    Returns:
        voyage_ranges: [(voyage_name, start_day, end_day), ...]
    """
    voyage_ranges = []
    current_voyage = None
    voyage_start = None
    voyage_end = None
    
    for task in tasks:
        if len(task) < 6:
            continue
        
        tid, wbs, task_name, phase, owner, offset = task[:6]
        
        if tid.startswith('V') and len(tid) == 2:
            if current_voyage is not None:
                voyage_ranges.append((current_voyage, voyage_start, voyage_end))
            current_voyage = task_name
            voyage_start = offset
            voyage_end = offset
        
        if current_voyage is not None:
            voyage_end = max(voyage_end, offset)
    
    if current_voyage is not None:
        voyage_ranges.append((current_voyage, voyage_start, voyage_end))
    
    if tasks and len(tasks[0]) >= 6:
        mob_start = tasks[0][5] if isinstance(tasks[0][5], (int, float)) else 0
        mob_end = mob_start + 1
        voyage_ranges.insert(0, ("Mobilization", mob_start, mob_end))
    
    return voyage_ranges

def generate_weather_periods(project_start, project_end):
    """
    프로젝트 기간을 기반으로 Weather Analysis 헤더 구간 생성
    """
    periods = []
    if project_end <= project_start:
        period_name = f"{project_start.strftime('%b %d')}"
        return [(period_name, project_start, project_start)]
    current = project_start
    
    while current < project_end:
        period_start = current
        period_end = min(period_start + dt.timedelta(days=9), project_end)
        
        if period_start.month == period_end.month:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%d')}"
        else:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%b %d')}"
        
        periods.append((period_name, period_start, period_end))
        current = period_end + dt.timedelta(days=1)
    
    return periods

def parse_voyage_pattern(pattern_str):
    """
    Parse voyage pattern like "1x1x1x1x1x1x1", "1-2-2-2", or "2-2-2-1" into TR groups.
    """
    if not pattern_str:
        return [[i] for i in range(1, 8)]

    voyage_groups = []
    tr_id = 1

    if "x" in pattern_str:
        parts = pattern_str.split("x")
    elif "-" in pattern_str:
        parts = pattern_str.split("-")
    else:
        parts = [pattern_str]

    for part in parts:
        part = part.strip()
        if not part:
            continue
        try:
            count = int(part)
        except ValueError:
            continue
        if count <= 0:
            continue
        voyage_groups.append(list(range(tr_id, tr_id + count)))
        tr_id += count

    if not voyage_groups:
        return [[i] for i in range(1, 8)]

    return voyage_groups

def generate_scenario_tasks(pattern_str, project_start, cycle_spacing=15, early_return=False):
    """
    Generate scenario tasks from voyage pattern.

    Returns:
        tasks: list of (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    """
    tasks = []
    voyage_groups = parse_voyage_pattern(pattern_str)

    # MOB-001: Mobilization
    offset = 0
    tasks.append((
        "MOB-001", "A0", "Mobilization (crew/equipment)", "MOBILIZATION",
        "Mammoet", offset, "DUR_MOB", "SPMT + grillage in MZP"
    ))
    offset += 1

    # PREP-001: Deck Prep
    tasks.append((
        "PREP-001", "A0", "LCT deck preparations + fenders + mooring", "DECK_PREP",
        "Mammoet/KFS", offset, "DUR_DECK", "MWS pre-check ready"
    ))
    offset += 3  # DUR_DECK = 3 days

    cycle_offset = offset

    for v_idx, tr_list in enumerate(voyage_groups, start=1):
        n_units = len(tr_list)
        tr_str = f"TR{tr_list[0]}" if n_units == 1 else f"TR{tr_list[0]}-TR{tr_list[-1]}"
        offset = cycle_offset
        voyage_start = offset

        # V{number}: Voyage milestone
        wbs = f"A{v_idx}"
        tasks.append((
            f"V{v_idx}", wbs, f"VOYAGE {v_idx}: {tr_str} Transport", "MILESTONE",
            "SCT/Mammoet/KFS", offset, 0, "TIDE>=1.90 required (Loadout start)"
        ))

        # LO-{number}: Loadout (n_units days for batch)
        lo_duration = n_units if n_units > 1 else 1
        tasks.append((
            f"LO-{v_idx:02d}", wbs, f"Loadout {tr_str} onto LCT", "LOADOUT",
            "Mammoet", offset, lo_duration, "OK 2.01m (tide window required)"
        ))
        offset += lo_duration

        # SF-{number}: Sea Fastening
        tasks.append((
            f"SF-{v_idx:02d}", wbs, f"Sea fastening + MWS checks ({tr_str})", "SEAFAST",
            "Mammoet/KFS/MWS", offset, "DUR_SF", "Lashing + survey"
        ))
        offset += 1  # DUR_SF = 0.5 days, treated as 1 here

        # SAIL-{number}: Sail-away
        tasks.append((
            f"SAIL-{v_idx:02d}", wbs, "Sail-away MZP->AGI", "SAIL",
            "LCT", offset, "DUR_SAIL", "WX gate"
        ))
        offset += 1  # DUR_SAIL = 1 day

        # UL-{number}-{TR}: Unload (per unit)
        for tr_num in tr_list:
            tasks.append((
                f"UL-{v_idx:02d}-{tr_num}", wbs, f"Unload TR{tr_num} at AGI (1 unit/day)", "AGI_UNLOAD",
                "Mammoet", offset, "DUR_UL", "RORO + ramp"
            ))
            offset += 1  # DUR_UL = 1 day/unit

        first_jd_offset = None

        # TURN + JD per unit (interleaved)
        for tr_num in tr_list:
            tasks.append((
                f"TURN-{v_idx:02d}-{tr_num}", wbs, f"Turning TR{tr_num} (90 deg)", "TURNING",
                "Mammoet", offset, "DUR_TURN", "3.0d/unit"
            ))
            offset += 3  # DUR_TURN = 3 days/unit

            tasks.append((
                f"JD-{v_idx:02d}-{tr_num}", wbs, f"Jackdown TR{tr_num}", "JACKDOWN",
                "Mammoet", offset, "DUR_JD", "1.0d/unit"
            ))
            if first_jd_offset is None:
                first_jd_offset = offset
            offset += 1  # DUR_JD = 1 day/unit

        return_offset = offset
        buffer_offset = offset + 1
        return_note = "After final JD"

        if early_return and n_units > 1 and first_jd_offset is not None:
            return_offset = first_jd_offset + 1
            buffer_offset = return_offset + 1
            return_note = "After first JD"

        # RET-{number}: Return
        tasks.append((
            f"RET-{v_idx:02d}", wbs, "LCT Return AGI->MZP", "RETURN",
            "LCT", return_offset, "DUR_RET", return_note
        ))

        # BUF-{number}: Buffer
        tasks.append((
            f"BUF-{v_idx:02d}", wbs, "Buffer / reset", "BUFFER",
            "All", buffer_offset, "DUR_BUF", "contingency"
        ))

        cycle_days = buffer_offset - voyage_start + 1
        cycle_offset = buffer_offset + 1
        if v_idx < len(voyage_groups) and cycle_spacing > cycle_days:
            cycle_offset += (cycle_spacing - cycle_days)

    return tasks

def get_scenario_sheet_names(scenario_name):
    sheet_name_map = {
        "Option_A": "Option_A",
        "Option_B": "Option_B",
        "Option_C": "Option_C",
    }
    short_name = sheet_name_map.get(scenario_name, scenario_name)
    sched_name = f"Schedule_Data_{short_name}"
    gantt_name = f"Gantt_Chart_{short_name}"
    return short_name, sched_name[:31], gantt_name[:31]

def create_scenario_sheets(wb, scenario_name, tsv_path, project_start, pattern_str=None, early_return=False):
    """
    시나리오별 Schedule_Data와 Gantt_Chart 시트 생성
    """
    import os
    
    short_name, sched_name, gantt_name = get_scenario_sheet_names(scenario_name)
    ws_sched = wb.create_sheet(sched_name)
    
    ws_sched.merge_cells("A1:J1")
    ws_sched["A1"] = f"AGI TR Transportation - {scenario_name} Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_sched["A1"].alignment = Alignment(horizontal="center")
    
    ws_sched.merge_cells("A2:J2")
    ws_sched["A2"] = f"Start = {project_start.isoformat()} | Auto-Updates from Control_Panel"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    
    try:
        ws_ctrl = wb["Control_Panel"]
        shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
        shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
        if hasattr(shamal_start, "date"):
            shamal_start = shamal_start.date()
        if hasattr(shamal_end, "date"):
            shamal_end = shamal_end.date()
        try:
            tide_threshold = float(ws_ctrl["H7"].value)
        except Exception:
            tide_threshold = 1.90
    except Exception:
        shamal_start = dt.date(2026, 1, 15)
        shamal_end = dt.date(2026, 4, 30)
        tide_threshold = 1.90
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"
    
    ws_sched.merge_cells("A3:J3")
    ws_sched["A3"] = f"⚠️ Winter Shamal Risk Period: {shamal_text_full} | Tide ≥{tide_threshold:.2f}m + Weather Gate"
    ws_sched["A3"].font = Font(size=10, italic=True)
    ws_sched["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    headers = [
        "Type A",
        "Type B",
        "Activity ID",
        "Activity Name",
        "Phase",
        "Owner",
        "Start",
        "End",
        "Duration",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    tasks = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
            if tasks:
                print(f"Loaded {len(tasks)} tasks for {scenario_name} from TSV file")
            else:
                print(f"Warning: No tasks loaded from {scenario_name} TSV")
        except Exception as e:
            print(f"Error loading {scenario_name} TSV: {e}")

    if not tasks and pattern_str:
        try:
            tasks = generate_scenario_tasks(
                pattern_str,
                project_start,
                early_return=early_return,
            )
            print(f"Generated {scenario_name} tasks from pattern: {pattern_str} (early_return={early_return})")
        except Exception as e:
            print(f"Error generating {scenario_name} tasks: {e}")

    if not tasks:
        print(f"Warning: No tasks for {scenario_name}")
        return False
    
    for r, t in enumerate(tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t[:8]
        type_a = t[8] if len(t) > 8 else ""
        type_b = t[9] if len(t) > 9 else ""
        activity_id = t[10] if len(t) > 10 else ""
        has_activity = bool(activity_id and str(activity_id).strip())
        if not has_activity and len(t) < 11:
            has_activity = True

        ws_sched.cell(r, 1, value=type_a)
        ws_sched.cell(r, 2, value=type_b)
        ws_sched.cell(r, 3, value=activity_id)
        ws_sched.cell(r, 4, value=task)
        ws_sched.cell(r, 5, value=phase)
        ws_sched.cell(r, 6, value=owner)

        ws_sched.cell(r, 7, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"

        if isinstance(dur_ref, str):
            ws_sched.cell(r, 9, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 9, value=dur_ref)

        ws_sched.cell(r, 8, value=f"=G{r}+I{r}")
        ws_sched.cell(r, 8).number_format = "YYYY-MM-DD"
        ws_sched.cell(r, 10, value=notes)

        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 11):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 5).fill = PatternFill("solid", fgColor=pc)

        if phase == "MILESTONE":
            for c in range(1, 11):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 11):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")
    
    col_widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 45,
        "E": 14,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 40,
    }
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    
    ws_gantt = wb.create_sheet(gantt_name)
    
    max_days = 70  # Fixed to 70 days for Gantt chart display
    date_col = 5
    last_col = date_col + max_days - 1
    
    ws_gantt.merge_cells(f"A1:{get_column_letter(last_col)}1")
    ws_gantt["A1"] = f"AGI TR Transportation - {scenario_name} Gantt Chart"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_gantt["A1"].alignment = Alignment(horizontal="center")
    
    ws_gantt.merge_cells(f"A2:{get_column_letter(last_col)}2")
    ws_gantt["A2"] = f"Orange Zone = Winter Shamal Risk Period ({shamal_text}) | Tide ≥{tide_threshold:.2f}m + Weather Gate | VBA: RefreshGanttChart_{scenario_name}로 색상 갱신"
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    month_str = project_start.strftime("%b %Y")
    ws_gantt["A3"] = month_str
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:D3")
    
    meta_headers = ["Type A", "Type B", "Activity ID", "Activity Name"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # date_col already set above
    for i in range(max_days):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "M/D"  # Format: 1/18, 2/15 - Shorter format
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5  # Narrower for M/D format
    
    for i in range(max_days):
        cell_date = project_start + dt.timedelta(days=i)
        if shamal_start <= cell_date <= shamal_end:
            ws_gantt.cell(4, date_col + i).fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    for r, t in enumerate(tasks, 5):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t[:8]
        activity_id = t[10] if len(t) > 10 else ""

        ws_gantt.cell(r, 1, value=f"='{sched_name}'!A{r+1}")
        ws_gantt.cell(r, 2, value=f"='{sched_name}'!B{r+1}")
        ws_gantt.cell(r, 3, value=f"='{sched_name}'!C{r+1}")
        ws_gantt.cell(r, 4, value=f"='{sched_name}'!D{r+1}")

        for c in range(1, 5):
            ws_gantt.cell(r, c).border = tb()

        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)

        if has_activity:
            start_date = project_start + dt.timedelta(days=offset)
            if isinstance(dur_ref, str):
                duration = 1
            else:
                try:
                    duration = float(dur_ref)
                except Exception:
                    duration = 1
            if duration < 0:
                duration = 0

            end_date = start_date + dt.timedelta(days=duration)

            for i in range(max_days):
                cell_date = project_start + dt.timedelta(days=i)
                cell = ws_gantt.cell(r, date_col + i)
                cell.border = tb()

                if start_date <= cell_date < end_date:
                    cell.fill = PatternFill("solid", fgColor=pc)
                elif cell_date == start_date and duration == 0:
                    cell.fill = PatternFill("solid", fgColor=pc)
                    cell.value = "*"
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=8)
        else:
            for i in range(max_days):
                cell = ws_gantt.cell(r, date_col + i)
                cell.border = tb()

    ws_gantt.column_dimensions["A"].width = 12
    ws_gantt.column_dimensions["B"].width = 12
    ws_gantt.column_dimensions["C"].width = 12
    ws_gantt.column_dimensions["D"].width = 45

    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)
    return True

def create_tide_data_sheet(wb, tide_tsv_path=None, tide_json_path=None):
    """
    조석 데이터 시트 생성
    """
    import os
    
    ws_tide = wb.create_sheet("Tide_Data")
    try:
        tide_threshold = float(wb["Control_Panel"]["H7"].value)
    except Exception:
        tide_threshold = 1.90
    
    ws_tide.merge_cells("A1:D1")
    ws_tide["A1"] = "MINA ZAYED PORT - High Tide Data"
    ws_tide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_tide["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_tide["A1"].alignment = Alignment(horizontal="center")
    
    ws_tide.merge_cells("A2:D2")
    ws_tide["A2"] = f"Tide ≥{tide_threshold:.2f}m required for Load-out and AGI Arrival | VBA: RefreshTideData로 업데이트"
    ws_tide["A2"].font = Font(size=10, italic=True)
    ws_tide["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    headers = ["Date", "High Tide Window", "Max Height (m)", "Risk Level"]
    for col, h in enumerate(headers, 1):
        cell = ws_tide.cell(4, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    tide_records = []
    if tide_json_path and os.path.exists(tide_json_path):
        tide_records = load_tide_data(json_path=tide_json_path)
        print(f"✅ Loaded {len(tide_records)} tide records from JSON")
    elif tide_tsv_path and os.path.exists(tide_tsv_path):
        tide_records = load_tide_data(tsv_path=tide_tsv_path)
        print(f"✅ Loaded {len(tide_records)} tide records from TSV")
    
    for r, (date_str, window, height, risk) in enumerate(tide_records, 5):
        ws_tide.cell(r, 1, value=date_str)
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            ws_tide.cell(r, 1, value=date_obj)
            ws_tide.cell(r, 1).number_format = "YYYY-MM-DD"
        except:
            ws_tide.cell(r, 1, value=date_str)
        
        ws_tide.cell(r, 2, value=window)
        ws_tide.cell(r, 3, value=height)
        ws_tide.cell(r, 3).number_format = "0.00"
        ws_tide.cell(r, 4, value=risk)
        
        for c in range(1, 5):
            ws_tide.cell(r, c).border = tb()
        
        if risk == "HIGH":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFCDD2")
        elif risk == "MEDIUM":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFE0B2")
        elif risk == "LOW":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="C8E6C9")
        
        if height >= tide_threshold:
            ws_tide.cell(r, 3).font = Font(bold=True, color="0066CC")
            ws_tide.cell(r, 1).fill = PatternFill("solid", fgColor="E3F2FD")
    
    ws_tide.column_dimensions["A"].width = 12
    ws_tide.column_dimensions["B"].width = 25
    ws_tide.column_dimensions["C"].width = 14
    ws_tide.column_dimensions["D"].width = 12
    
    ws_tide.freeze_panes = "A5"

def create_comparison_summary(
    wb,
    project_start,
    scenario_a="Option_A",
    scenario_b="Option_B",
    label_a=None,
    label_b=None,
    sheet_name="Scenario_Comparison",
):
    """
    두 시나리오 비교 Summary 시트 생성
    """
    short_a, sched_a, _ = get_scenario_sheet_names(scenario_a)
    short_b, sched_b, _ = get_scenario_sheet_names(scenario_b)
    if not label_a:
        label_a = short_a
    if not label_b:
        label_b = short_b

    ws_comp = wb.create_sheet(sheet_name)
    
    ws_comp.merge_cells("A1:F1")
    ws_comp["A1"] = "AGI TR Transportation - Scenario Comparison"
    ws_comp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_comp["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_comp["A1"].alignment = Alignment(horizontal="center")
    
    headers = ["Metric", label_a, label_b, "Difference", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_comp.cell(3, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    comparison_data = [
        (
            "Total Tasks",
            f"=COUNTIF('{sched_a}'!C6:C1000,\"A*\")",
            f"=COUNTIF('{sched_b}'!C6:C1000,\"A*\")",
            "=C4-B4",
            "Tasks with Activity ID starting with 'A'",
        ),
        (
            "Total Voyages",
            f"=SUMPRODUCT(('{sched_a}'!C6:C1000=\"\")*('{sched_a}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_a}'!D6:D1000)))",
            f"=SUMPRODUCT(('{sched_b}'!C6:C1000=\"\")*('{sched_b}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_b}'!D6:D1000)))",
            "=C5-B5",
            "Voyage groups",
        ),
        (
            "Project Duration (days)",
            f"=INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/('{sched_a}'!C6:C1000<>\"\"),1)-1E-9)-PROJECT_START+1",
            f"=INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/('{sched_b}'!C6:C1000<>\"\"),1)-1E-9)-PROJECT_START+1",
            "=C6-B6",
            "Days difference (inclusive)",
        ),
        (
            "Total Jack-down Events",
            f"=COUNTIF('{sched_a}'!E6:E1000,\"JACKDOWN\")",
            f"=COUNTIF('{sched_b}'!E6:E1000,\"JACKDOWN\")",
            "=C7-B7",
            "",
        ),
        (
            "Project End Date",
            f"=INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/('{sched_a}'!C6:C1000<>\"\"),1)-1E-9)",
            f"=INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/('{sched_b}'!C6:C1000<>\"\"),1)-1E-9)",
            "=C8-B8",
            "Date difference",
        ),
        (
            "Mobilization Duration (days)",
            f"=IFERROR(INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/(('{sched_a}'!E6:E1000=\"MOBILIZATION\")*('{sched_a}'!C6:C1000<>\"\")),1)-1E-9)-AGGREGATE(15,6,'{sched_a}'!G6:G1000/(('{sched_a}'!E6:E1000=\"MOBILIZATION\")*('{sched_a}'!C6:C1000<>\"\")),1)+1,\"\")",
            f"=IFERROR(INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/(('{sched_b}'!E6:E1000=\"MOBILIZATION\")*('{sched_b}'!C6:C1000<>\"\")),1)-1E-9)-AGGREGATE(15,6,'{sched_b}'!G6:G1000/(('{sched_b}'!E6:E1000=\"MOBILIZATION\")*('{sched_b}'!C6:C1000<>\"\")),1)+1,\"\")",
            "=C9-B9",
            "Mobilization period (first MOB start to last MOB end)",
        ),
        (
            "LCT Round Trips",
            f"=SUMPRODUCT(('{sched_a}'!C6:C1000=\"\")*('{sched_a}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_a}'!D6:D1000)))",
            f"=SUMPRODUCT(('{sched_b}'!C6:C1000=\"\")*('{sched_b}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_b}'!D6:D1000)))",
            "=C10-B10",
            "Derived from Voyage count (RETURN phase not in TSV)",
        ),
    ]
    
    for r, (metric, val_a, val_b, diff, notes) in enumerate(comparison_data, 4):
        ws_comp.cell(r, 1, value=metric)
        ws_comp.cell(r, 1).font = Font(bold=True)
        
        ws_comp.cell(r, 2, value=val_a)
        ws_comp.cell(r, 3, value=val_b)
        ws_comp.cell(r, 4, value=diff)
        ws_comp.cell(r, 5, value=notes)
        
        for c in range(1, 6):
            ws_comp.cell(r, c).border = tb()
            if c in [2, 3, 4] and "Date" in metric:
                ws_comp.cell(r, c).number_format = "YYYY-MM-DD"
            elif c == 4 and "Date" not in metric:
                ws_comp.cell(r, c).number_format = "0"
    
    for col in ["A", "B", "C", "D", "E"]:
        ws_comp.column_dimensions[col].width = 25





def create_option_schedule_comparison_sheet(
    wb,
    scenario_a="Option_A",
    scenario_b="Option_B",
    sheet_name="Option_Compare",
):
    """Option A/B schedule comparison by Activity Name."""

    short_a, sched_a, _ = get_scenario_sheet_names(scenario_a)
    short_b, sched_b, _ = get_scenario_sheet_names(scenario_b)

    if sched_a not in wb.sheetnames or sched_b not in wb.sheetnames:
        return False

    ws_a = wb[sched_a]
    ws_b = wb[sched_b]

    def collect_names(ws, col=4):
        names = []
        for r in range(6, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                names.append(s)
        return names

    name_match_col = "D"
    names_a = collect_names(ws_a, 4)
    names_b = collect_names(ws_b, 4)

    # Option A order first; then append names unique to Option B
    seen = set()
    all_names = []
    for x in names_a + names_b:
        if x in seen:
            continue
        seen.add(x)
        all_names.append(x)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_cmp = wb.create_sheet(sheet_name)

    ws_cmp.merge_cells("A1:N1")
    ws_cmp["A1"] = f"{short_a} vs {short_b} - Full Schedule Comparison (by Activity Name)"
    ws_cmp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_cmp["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_cmp["A1"].alignment = Alignment(horizontal="center")

    # Quick summary
    ws_cmp["A3"] = "Overall End (A)"
    ws_cmp["B3"] = f"=MAX('{sched_a}'!H:H)"
    ws_cmp["A4"] = "Overall End (B)"
    ws_cmp["B4"] = f"=MAX('{sched_b}'!H:H)"
    ws_cmp["A5"] = "Delta End (B-A)"
    ws_cmp["B5"] = "=B4-B3"
    for addr in ("A3", "A4", "A5"):
        ws_cmp[addr].font = Font(bold=True)
    for addr in ("B3", "B4"):
        ws_cmp[addr].number_format = "YYYY-MM-DD"
    ws_cmp["B5"].number_format = "0"

    headers = [
        "Activity Name",
        f"Activity ID ({short_a})",
        f"Phase ({short_a})",
        f"Start ({short_a})",
        f"End ({short_a})",
        f"Dur ({short_a})",
        f"Activity ID ({short_b})",
        f"Phase ({short_b})",
        f"Start ({short_b})",
        f"End ({short_b})",
        f"Dur ({short_b})",
        "Delta Start (B-A)",
        "Delta End (B-A)",
        "Delta Dur (B-A)",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws_cmp.cell(7, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = tb()

    def idxmatch_by_name(row_idx: int, sheet: str, col_letter: str) -> str:
        return f"""=IFERROR(INDEX('{sheet}'!{col_letter}:{col_letter},MATCH($A{row_idx},'{sheet}'!${name_match_col}:${name_match_col},0)),"")"""

    start_row = 8
    for n, task_name in enumerate(all_names):
        r = start_row + n
        ws_cmp.cell(r, 1, value=task_name)

        ws_cmp.cell(r, 2, value=idxmatch_by_name(r, sched_a, "C"))
        ws_cmp.cell(r, 3, value=idxmatch_by_name(r, sched_a, "E"))
        ws_cmp.cell(r, 4, value=idxmatch_by_name(r, sched_a, "G"))
        ws_cmp.cell(r, 5, value=idxmatch_by_name(r, sched_a, "H"))
        ws_cmp.cell(r, 6, value=idxmatch_by_name(r, sched_a, "I"))

        ws_cmp.cell(r, 7, value=idxmatch_by_name(r, sched_b, "C"))
        ws_cmp.cell(r, 8, value=idxmatch_by_name(r, sched_b, "E"))
        ws_cmp.cell(r, 9, value=idxmatch_by_name(r, sched_b, "G"))
        ws_cmp.cell(r, 10, value=idxmatch_by_name(r, sched_b, "H"))
        ws_cmp.cell(r, 11, value=idxmatch_by_name(r, sched_b, "I"))

        ws_cmp.cell(r, 12, value=f'=IF(AND(D{r}<>"",I{r}<>""),I{r}-D{r},"")')
        ws_cmp.cell(r, 13, value=f'=IF(AND(E{r}<>"",J{r}<>""),J{r}-E{r},"")')
        ws_cmp.cell(r, 14, value=f'=IF(AND(F{r}<>"",K{r}<>""),K{r}-F{r},"")')

        for c in range(1, 15):
            ws_cmp.cell(r, c).border = tb()

        for c in (4, 5, 9, 10):
            ws_cmp.cell(r, c).number_format = "YYYY-MM-DD"
        for c in (12, 13, 14):
            ws_cmp.cell(r, c).number_format = "0"

    widths = {
        "A": 50,
        "B": 14,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 10,
        "L": 12,
        "M": 12,
        "N": 12,
    }
    for col, w in widths.items():
        ws_cmp.column_dimensions[col].width = w

    # Conditional formatting: deltas (L:M:N)
    try:
        from openpyxl.formatting.rule import CellIsRule

        last_row = start_row + len(all_names) - 1
        if last_row >= start_row:
            red_fill = PatternFill("solid", fgColor="FFCDD2")
            green_fill = PatternFill("solid", fgColor="C8E6C9")
            delta_range = f"L{start_row}:N{last_row}"
            ws_cmp.conditional_formatting.add(
                delta_range,
                CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill),
            )
            ws_cmp.conditional_formatting.add(
                delta_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=green_fill),
            )
    except Exception:
        pass

    ws_cmp.freeze_panes = "A8"
    return True



def create_three_way_comparison(
    wb,
    project_start,
    scenario_a="Option_A",
    scenario_b="Option_B",
    scenario_c="Option_C",
    label_a="Option A",
    label_b="Option B",
    label_c="Option C",
    sheet_name="Three_Way_Comparison",
):
    """Create a three-way comparison sheet."""
    short_a, sched_a, _ = get_scenario_sheet_names(scenario_a)
    short_b, sched_b, _ = get_scenario_sheet_names(scenario_b)

    if sched_a not in wb.sheetnames or sched_b not in wb.sheetnames:
        return False

    sched_c = None
    short_c = None
    if scenario_c:
        short_c, sched_c, _ = get_scenario_sheet_names(scenario_c)
    has_c = bool(sched_c) and sched_c in wb.sheetnames

    if not label_a:
        label_a = short_a
    if not label_b:
        label_b = short_b
    if has_c and not label_c:
        label_c = short_c

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_comp = wb.create_sheet(sheet_name)

    header_cols = "A1:H1" if has_c else "A1:F1"
    ws_comp.merge_cells(header_cols)
    ws_comp["A1"] = "AGI TR Transportation - Three-Way Comparison (A/B/C)"
    ws_comp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_comp["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_comp["A1"].alignment = Alignment(horizontal="center")

    subtitle_cols = "A2:H2" if has_c else "A2:F2"
    ws_comp.merge_cells(subtitle_cols)
    ws_comp["A2"] = f"Project Start: {project_start.isoformat()} | All dates auto-update from Control_Panel"
    ws_comp["A2"].font = Font(size=10, italic=True)
    ws_comp["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    headers = ["Metric", label_a, label_b]
    if has_c:
        headers.extend([label_c, "Best", "Worst", "Range", "Notes"])
    else:
        headers.extend(["Difference", "Notes"])

    for col, h in enumerate(headers, 1):
        cell = ws_comp.cell(4, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = tb()

    if has_c:
        comparison_data = [
            (
                "Total Tasks",
                f"=COUNTIF('{sched_a}'!C6:C1000,\"A*\")",
                f"=COUNTIF('{sched_b}'!C6:C1000,\"A*\")",
                f"=COUNTIF('{sched_c}'!C6:C1000,\"A*\")",
                "Tasks with Activity ID starting with 'A'",
            ),
            (
                "Total Voyages",
                f"=SUMPRODUCT(('{sched_a}'!C6:C1000=\"\")*('{sched_a}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_a}'!D6:D1000)))",
                f"=SUMPRODUCT(('{sched_b}'!C6:C1000=\"\")*('{sched_b}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_b}'!D6:D1000)))",
                f"=SUMPRODUCT(('{sched_c}'!C6:C1000=\"\")*('{sched_c}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_c}'!D6:D1000)))",
                "Voyage groups",
            ),
            (
                "Project Duration (days)",
                f"=INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/('{sched_a}'!C6:C1000<>\"\"),1)-1E-9)-PROJECT_START+1",
                f"=INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/('{sched_b}'!C6:C1000<>\"\"),1)-1E-9)-PROJECT_START+1",
                f"=INT(AGGREGATE(14,6,'{sched_c}'!H6:H1000/('{sched_c}'!C6:C1000<>\"\"),1)-1E-9)-PROJECT_START+1",
                "Days difference (inclusive)",
            ),
            (
                "Total Jack-down Events",
                f"=COUNTIF('{sched_a}'!E6:E1000,\"JACKDOWN\")",
                f"=COUNTIF('{sched_b}'!E6:E1000,\"JACKDOWN\")",
                f"=COUNTIF('{sched_c}'!E6:E1000,\"JACKDOWN\")",
                "",
            ),
            (
                "Project End Date",
                f"=INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/('{sched_a}'!C6:C1000<>\"\"),1)-1E-9)",
                f"=INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/('{sched_b}'!C6:C1000<>\"\"),1)-1E-9)",
                f"=INT(AGGREGATE(14,6,'{sched_c}'!H6:H1000/('{sched_c}'!C6:C1000<>\"\"),1)-1E-9)",
                "Date difference",
            ),
            (
                "Mobilization Duration (days)",
                f"=IFERROR(INT(AGGREGATE(14,6,'{sched_a}'!H6:H1000/(('{sched_a}'!E6:E1000=\"MOBILIZATION\")*('{sched_a}'!C6:C1000<>\"\")),1)-1E-9)-AGGREGATE(15,6,'{sched_a}'!G6:G1000/(('{sched_a}'!E6:E1000=\"MOBILIZATION\")*('{sched_a}'!C6:C1000<>\"\")),1)+1,\"\")",
                f"=IFERROR(INT(AGGREGATE(14,6,'{sched_b}'!H6:H1000/(('{sched_b}'!E6:E1000=\"MOBILIZATION\")*('{sched_b}'!C6:C1000<>\"\")),1)-1E-9)-AGGREGATE(15,6,'{sched_b}'!G6:G1000/(('{sched_b}'!E6:E1000=\"MOBILIZATION\")*('{sched_b}'!C6:C1000<>\"\")),1)+1,\"\")",
                f"=IFERROR(INT(AGGREGATE(14,6,'{sched_c}'!H6:H1000/(('{sched_c}'!E6:E1000=\"MOBILIZATION\")*('{sched_c}'!C6:C1000<>\"\")),1)-1E-9)-AGGREGATE(15,6,'{sched_c}'!G6:G1000/(('{sched_c}'!E6:E1000=\"MOBILIZATION\")*('{sched_c}'!C6:C1000<>\"\")),1)+1,\"\")",
                "Mobilization period (first MOB start to last MOB end)",
            ),
            (
                "LCT Round Trips",
                f"=SUMPRODUCT(('{sched_a}'!C6:C1000=\"\")*('{sched_a}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_a}'!D6:D1000)))",
                f"=SUMPRODUCT(('{sched_b}'!C6:C1000=\"\")*('{sched_b}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_b}'!D6:D1000)))",
                f"=SUMPRODUCT(('{sched_c}'!C6:C1000=\"\")*('{sched_c}'!D6:D1000<>\"\")*ISNUMBER(SEARCH(\"AGI TR Unit\",'{sched_c}'!D6:D1000)))",
                "Derived from Voyage count (RETURN phase not in TSV)",
            ),
        ]

        for r, (metric, val_a, val_b, val_c, notes) in enumerate(comparison_data, 5):
            ws_comp.cell(r, 1, value=metric).font = Font(bold=True)
            ws_comp.cell(r, 2, value=val_a)
            ws_comp.cell(r, 3, value=val_b)
            ws_comp.cell(r, 4, value=val_c)
            ws_comp.cell(r, 5, value=f"=MIN(B{r}:D{r})")
            ws_comp.cell(r, 6, value=f"=MAX(B{r}:D{r})")
            ws_comp.cell(r, 7, value=f"=F{r}-E{r}")
            ws_comp.cell(r, 8, value=notes)

            for c in range(1, 9):
                ws_comp.cell(r, c).border = tb()

            if "Date" in metric:
                for c in (2, 3, 4, 5, 6):
                    ws_comp.cell(r, c).number_format = "YYYY-MM-DD"
                ws_comp.cell(r, 7).number_format = "0"
            else:
                for c in (2, 3, 4, 5, 6, 7):
                    ws_comp.cell(r, c).number_format = "0"

            if "Duration" in metric or "Round Trips" in metric:
                ws_comp.cell(r, 5).fill = PatternFill("solid", fgColor="C8E6C9")
                ws_comp.cell(r, 6).fill = PatternFill("solid", fgColor="FFCDD2")
    else:
        comparison_data = [
            (
                "Total Tasks",
                f"=COUNTIF('{sched_a}'!C6:C1000,\"A*\")",
                f"=COUNTIF('{sched_b}'!C6:C1000,\"A*\")",
                "=C5-B5",
                "Tasks with Activity ID starting with 'A'",
            ),
            (
                "Project Duration (days)",
                f"=DATEDIF(PROJECT_START,MAX('{sched_a}'!H6:H1000),\"d\")+1",
                f"=DATEDIF(PROJECT_START,MAX('{sched_b}'!H6:H1000),\"d\")+1",
                "=C6-B6",
                "Days difference",
            ),
            (
                "Project End Date",
                f"=MAX('{sched_a}'!H6:H1000)",
                f"=MAX('{sched_b}'!H6:H1000)",
                "=C7-B7",
                "Date difference",
            ),
        ]

        for r, (metric, val_a, val_b, diff, notes) in enumerate(comparison_data, 5):
            ws_comp.cell(r, 1, value=metric).font = Font(bold=True)
            ws_comp.cell(r, 2, value=val_a)
            ws_comp.cell(r, 3, value=val_b)
            ws_comp.cell(r, 4, value=diff)
            ws_comp.cell(r, 5, value=notes)

            for c in range(1, 6):
                ws_comp.cell(r, c).border = tb()
                if c in [2, 3, 4] and "Date" in metric:
                    ws_comp.cell(r, c).number_format = "YYYY-MM-DD"
                elif c == 4 and "Date" not in metric:
                    ws_comp.cell(r, c).number_format = "0"

    widths = {
        "A": 25,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 30,
    }
    for col, w in widths.items():
        ws_comp.column_dimensions[col].width = w

    ws_comp.freeze_panes = "A5"
    return True
def create_gantt_with_vba(
    tsv_path=None,
    option_a_tsv=None,
    option_b_tsv=None,
    option_c_tsv=None,
    tide_tsv=None,
    tide_json=None,
):
    import os
    
    wb = Workbook()
    

    # TSV 파일에서 프로젝트 시작일 자동 감지 (다양한 TSV 헤더 지원)
    project_start = dt.date(2026, 1, 18)
    start_tsv = None
    for candidate in [option_a_tsv, option_b_tsv, option_c_tsv, tsv_path]:
        if candidate and os.path.exists(candidate):
            start_tsv = candidate
            break

    def _parse_date_loose(s: str):
        s = (s or '').strip()
        if not s:
            return None
        for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d-%b-%y', '%d-%b'):
            try:
                d = dt.datetime.strptime(s, fmt).date()
                # dd-Mon (no year) -> assume same year as default project_start
                if fmt == '%d-%b':
                    d = d.replace(year=project_start.year)
                return d
            except Exception:
                pass
        return None

    if start_tsv:
        try:
            import csv
            earliest = None
            with open(start_tsv, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f, delimiter='	')
                headers = next(reader)
                header_lc = [h.strip() for h in headers]

                def _idx(name: str):
                    try:
                        return header_lc.index(name)
                    except ValueError:
                        return None

                start_cols = []
                for k in ('Planned Start', 'Actual Start', 'Start'):
                    i = _idx(k)
                    if i is not None:
                        start_cols.append(i)

                for row in reader:
                    for i in start_cols:
                        if i < len(row):
                            d = _parse_date_loose(row[i])
                            if d:
                                earliest = d if earliest is None else min(earliest, d)
            if earliest:
                project_start = earliest
        except Exception as e:
            print(f"Warning: Could not read TSV file for start date: {e}")
            print("Using default project start date: 2026-01-18")


    # === CONTROL PANEL (새로 추가) ===
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"
    
    # Title
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "🎛️ AGI TR Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30
    
    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "📌 Changing the start date (B4) will automatically update all schedules. VBA macros must be enabled."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    
    # Input Section
    ws_ctrl["A4"] = "📅 Project Start Date:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = project_start
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)
    
    ws_ctrl["A5"] = "🎯 Target Completion Date:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = dt.date(2026, 2, 28)
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()
    
    # Voyage Pattern Selection
    ws_ctrl["A6"] = "🚢 Voyage Pattern:"
    ws_ctrl["A6"].font = Font(bold=True)
    ws_ctrl["B6"] = "1-2-2-2"  # Default: 1 solo + 3 pairs
    ws_ctrl["B6"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B6"].border = tb()
    # Note: Valid patterns: "1x1x1x1x1x1x1", "1-2-2-2", "2-2-2-1"
    voyage_patterns = ["1x1x1x1x1x1x1", "1-2-2-2", "2-2-2-1"]
    dv_voyage = DataValidation(
        type="list",
        formula1=f'"{",".join(voyage_patterns)}"',
        allow_blank=True,
    )
    dv_voyage.error = "Invalid voyage pattern selected"
    dv_voyage.errorTitle = "Invalid Voyage Pattern"
    ws_ctrl.add_data_validation(dv_voyage)
    dv_voyage.add("B6")
    
    # Early Return Option
    ws_ctrl["A7"] = "🔄 Early Return (1st JD):"
    ws_ctrl["A7"].font = Font(bold=True)
    ws_ctrl["B7"] = "TRUE"  # TRUE = return after first JD, FALSE = return after batch JD
    ws_ctrl["B7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B7"].border = tb()
    
    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")
    wb.defined_names["VOYAGE_PATTERN"] = DefinedName("VOYAGE_PATTERN", attr_text="Control_Panel!$B$6")
    wb.defined_names["EARLY_RETURN"] = DefinedName("EARLY_RETURN", attr_text="Control_Panel!$B$7")
    
    # Duration Parameters
    ws_ctrl["D4"] = "⏱️ Task Duration (Days)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)
    
    durations = [
        ("D5", "Mobilization:", "E5", 1.0, "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", 3.0, "DUR_DECK"),
        ("D7", "Load-out:", "E7", 1.0, "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", 0.5, "DUR_SF"),
        ("D9", "MWS Approval:", "E9", 0.5, "DUR_MWS"),
        ("D10", "Sailing:", "E10", 1.0, "DUR_SAIL"),
        ("D11", "AGI Unload:", "E11", 1.0, "DUR_UL"),
        ("D12", "Turning:", "E12", 3.0, "DUR_TURN"),
        ("D13", "Jack-down:", "E13", 1.0, "DUR_JD"),
        ("D14", "Return:", "E14", 1.0, "DUR_RET"),
        ("D15", "Buffer:", "E15", 0.5, "DUR_BUF"),
    ]
    
    for lc, lt, vc, v, name in durations:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = v
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")
    
    # Weather Settings
    ws_ctrl["G4"] = "🌊 Weather Settings"
    ws_ctrl["G4"].font = Font(bold=True, size=12)
    ws_ctrl["G5"] = "Shamal Start:"
    ws_ctrl["H5"] = dt.date(2026, 1, 15)
    ws_ctrl["H5"].number_format = "YYYY-MM-DD"
    ws_ctrl["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G6"] = "Shamal End:"
    ws_ctrl["H6"] = dt.date(2026, 4, 30)
    ws_ctrl["H6"].number_format = "YYYY-MM-DD"
    ws_ctrl["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    ws_ctrl["G7"] = "Tide Threshold (m):"
    ws_ctrl["G7"].font = Font(bold=True)
    ws_ctrl["H7"] = 1.90
    ws_ctrl["H7"].number_format = "0.00"
    ws_ctrl["H7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H7"].border = tb()
    
    ws_ctrl["G8"] = "📊 Gantt Min Days:"
    ws_ctrl["G8"].font = Font(bold=True)
    ws_ctrl["H8"] = 120
    ws_ctrl["H8"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H8"].border = tb()
    ws_ctrl["H8"].number_format = "0"
    
    ws_ctrl["G9"] = "📊 Gantt Buffer Days:"
    ws_ctrl["G9"].font = Font(bold=True)
    ws_ctrl["H9"] = 30
    ws_ctrl["H9"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H9"].border = tb()
    ws_ctrl["H9"].number_format = "0"
    
    # LCT Maintenance Settings
    ws_ctrl["G10"] = "🔧 LCT Maint. Start:"
    ws_ctrl["G10"].font = Font(bold=True)
    ws_ctrl["H10"] = dt.date(2026, 2, 10)  # Default maintenance start
    ws_ctrl["H10"].number_format = "YYYY-MM-DD"
    ws_ctrl["H10"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H10"].border = tb()
    
    ws_ctrl["G11"] = "🔧 LCT Maint. End:"
    ws_ctrl["G11"].font = Font(bold=True)
    ws_ctrl["H11"] = dt.date(2026, 2, 14)  # Default maintenance end (4 days)
    ws_ctrl["H11"].number_format = "YYYY-MM-DD"
    ws_ctrl["H11"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H11"].border = tb()

    wb.defined_names["SHAMAL_START"] = DefinedName("SHAMAL_START", attr_text="Control_Panel!$H$5")
    wb.defined_names["SHAMAL_END"] = DefinedName("SHAMAL_END", attr_text="Control_Panel!$H$6")
    wb.defined_names["TIDE_THRESHOLD"] = DefinedName("TIDE_THRESHOLD", attr_text="Control_Panel!$H$7")
    wb.defined_names["GANTT_MIN_DAYS"] = DefinedName("GANTT_MIN_DAYS", attr_text="Control_Panel!$H$8")
    wb.defined_names["GANTT_BUFFER_DAYS"] = DefinedName("GANTT_BUFFER_DAYS", attr_text="Control_Panel!$H$9")
    wb.defined_names["LCT_MAINT_START"] = DefinedName("LCT_MAINT_START", attr_text="Control_Panel!$H$10")
    wb.defined_names["LCT_MAINT_END"] = DefinedName("LCT_MAINT_END", attr_text="Control_Panel!$H$11")

    
    # Summary Section
    ws_ctrl["A8"] = "📊 Auto Calculation Summary_Option A"
    ws_ctrl["A8"].font = Font(bold=True, size=12)

    summary_items_a = [
        ("A9", "Estimated Completion:", "B9", "=MAX(Schedule_Data_Option_A!H:H)"),
        ("A10", "Total Duration (Days):", "B10", '=IF(ISNUMBER(B9),B9-B4+1,"")'),
        ("A11", "Status vs Target:", "B11", '=IF(ISNUMBER(B9),IF(B9<=B5,"On Target","Delayed"),"")'),
        ("A12", "Remaining Days:", "B12", '=IF(ISNUMBER(B9),B5-B9,"")'),
    ]

    ws_ctrl["A14"] = "📊 Auto Calculation Summary_Option B"
    ws_ctrl["A14"].font = Font(bold=True, size=12)

    summary_items_b = [
        ("A15", "Estimated Completion:", "B15", "=MAX(Schedule_Data_Option_B!H:H)"),
        ("A16", "Total Duration (Days):", "B16", '=IF(ISNUMBER(B15),B15-B4+1,"")'),
        ("A17", "Status vs Target:", "B17", '=IF(ISNUMBER(B15),IF(B15<=B5,"On Target","Delayed"),"")'),
        ("A18", "Remaining Days:", "B18", '=IF(ISNUMBER(B15),B5-B15,"")'),
    ]

    for lc, lt, vc, formula in summary_items_a + summary_items_b:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if "MAX" in formula:
            ws_ctrl[vc].number_format = "YYYY-MM-DD"

    # VBA Button Info
    ws_ctrl["A20"] = "🔔 VBA Macros (Alt+F8)"
    ws_ctrl["A20"].font = Font(bold=True, size=12)

    buttons = [
        "▶ UpdateAllScenarios - Batch update (Option A/B/C + Compare)",
        "▶ RefreshAllGanttCharts - Refresh both Gantt charts",
        "▶ RefreshGanttChart_Option_A - Refresh Option A Gantt",
        "▶ RefreshGanttChart_Option_B - Refresh Option B Gantt",
        "▶ RefreshTideData - Highlight Tide Data",
        "▶ GenerateReport - Show End Dates (A/B)",
        "▶ ExportToPDF - Export key sheets to PDF",
        "▶ CheckShamalRisk - Check Shamal Risk (A/B)",
        "▶ ShowControlPanelSettings - Show current settings",
        "▶ ResetAppState - Restore Excel state (Events/Calc)",
    ]
    for i, btn in enumerate(buttons, 21):
        ws_ctrl[f"A{i}"] = btn
        ws_ctrl[f"A{i}"].font = Font(size=10)
    
    # Column widths
    ws_ctrl.column_dimensions["A"].width = 30
    ws_ctrl.column_dimensions["B"].width = 15
    ws_ctrl.column_dimensions["D"].width = 16
    ws_ctrl.column_dimensions["E"].width = 10
    ws_ctrl.column_dimensions["G"].width = 14
    ws_ctrl.column_dimensions["H"].width = 12
    ws_ctrl.column_dimensions["I"].width = 10
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    try:
        tide_threshold = float(ws_ctrl["H7"].value)
    except Exception:
        tide_threshold = 1.90
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"
    
    # === TASK DEFINITIONS (used for internal calculations) ===
    # Task definitions - TSV 파일에서 로드 또는 기본값 사용
    # (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    # 프로젝트 시작: 2026-01-18 (Day 0)
    default_tasks = [
        # Mobilization
        ("MOB-001", "1.0", "MOBILIZATION", "MOBILIZATION", "Mammoet", 0, "DUR_MOB", "SPMT Assembly + Marine Equipment Mobilization"),
        ("PREP-001", "1.1", "Deck Preparations", "DECK_PREP", "Mammoet", 1, "DUR_DECK", "One-time setup for all voyages"),
        
        # Voyage 1: LO 01-18, SAIL 01-20, ARR 01-22
        ("V1", "2.0", "VOYAGE 1: TR1 Transport", "MILESTONE", "All", 0, 0, "✅ Tide ≥1.90m (2.05m) | Good Weather Window"),
        ("LO-101", "2.1", "TR1 Load-out on LCT", "LOADOUT", "Mammoet", 0, "DUR_LO", "Tide ≥1.90m (2.05m) required"),
        ("SF-102", "2.2", "TR1 Sea Fastening", "SEAFAST", "Mammoet", 0, "DUR_SF", "12-point lashing"),
        ("MWS-103", "2.3", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 0, "DUR_MWS", "Marine Warranty Surveyor"),
        ("SAIL-104", "2.4", "V1 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 2, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-105", "2.5", "AGI Arrival + TR1 RORO Unload", "AGI_UNLOAD", "Mammoet", 4, "DUR_UL", "Tide ≥1.90m (1.91m) | AGI FWD Draft ≤ 2.70m"),
        ("STORE-106", "2.6", "TR1 Stored on AGI Laydown", "BUFFER", "Mammoet", 4, "DUR_BUF", "Awaiting pair TR2"),
        ("RET-107", "2.7", "V1 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 4, "DUR_RET", "Quick turnaround"),
        ("BUF-108", "2.99", "V1 Buffer / Equipment Reset", "BUFFER", "All", 5, "DUR_BUF", "Weather contingency"),
        
        # Voyage 2: LO 01-26, SAIL 01-27, ARR 01-29
        ("V2", "3.0", "VOYAGE 2: TR2 Transport + JD-1", "MILESTONE", "All", 8, 0, "✅ Tide ≥1.90m (1.91m) | Good Weather Window (before Shamal)"),
        ("LO-109", "3.1", "TR2 Load-out on LCT", "LOADOUT", "Mammoet", 8, "DUR_LO", "Tide ≥1.90m (1.91m) required"),
        ("SF-110", "3.2", "TR2 Sea Fastening", "SEAFAST", "Mammoet", 8, "DUR_SF", "12-point lashing"),
        ("MWS-110A", "3.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 8, "DUR_MWS", "Pre-sail verification"),
        ("SAIL-111", "3.3", "V2 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 9, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-112", "3.4", "AGI Arrival + TR2 RORO Unload", "AGI_UNLOAD", "Mammoet", 11, "DUR_UL", "Tide ≥1.90m (2.03m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-113", "3.5", "TR1 Transport to Bay-1", "TURNING", "Mammoet", 12, 1, "Steel bridge install"),
        ("TURN-114", "3.6", "TR1 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", "10t Forklift required"),
        ("TRN-116", "3.8", "TR2 Transport to Bay-2", "TURNING", "Mammoet", 12, 1, ""),
        ("TURN-117", "3.9", "TR2 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", ""),
        ("JD-120A", "3.95", "JD-1 Jack-Down TR1", "JACKDOWN", "Mammoet", 14, "DUR_JD", "MILESTONE: TR1 complete | 02-01"),
        ("RET-119", "3.11", "V2 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 15, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-120B", "3.96", "JD-1 Jack-Down TR2", "JACKDOWN", "Mammoet", 16, "DUR_JD", "MILESTONE: TR2 complete | 02-02"),
        ("BUF-120", "3.99", "V2 Buffer / Shamal Recovery", "BUFFER", "All", 17, "DUR_BUF", "Post-Shamal weather check"),
        
        # Voyage 3: LO 01-31, SAIL 02-02, ARR 02-03
        ("V3", "4.0", "VOYAGE 3: TR3 Transport", "MILESTONE", "All", 13, 0, "✅ Tide ≥1.90m (2.07m) | Post-Shamal Window"),
        ("LO-121", "4.1", "TR3 Load-out on LCT", "LOADOUT", "Mammoet", 13, "DUR_LO", "Tide ≥1.90m (2.07m)"),
        ("SF-122", "4.2", "TR3 Sea Fastening", "SEAFAST", "Mammoet", 13, "DUR_SF", ""),
        ("MWS-122A", "4.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 13, "DUR_MWS", ""),
        ("SAIL-123", "4.3", "V3 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 15, "DUR_SAIL", "Good weather"),
        ("ARR-124", "4.4", "AGI Arrival + TR3 RORO Unload", "AGI_UNLOAD", "Mammoet", 16, "DUR_UL", "Tide ≥1.90m (2.04m)"),
        ("STORE-125", "4.5", "TR3 Stored on AGI Laydown", "BUFFER", "Mammoet", 16, "DUR_BUF", "Awaiting pair TR4"),
        ("RET-126", "4.6", "V3 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 17, "DUR_RET", ""),
        ("BUF-127", "4.99", "V3 Buffer", "BUFFER", "All", 17, "DUR_BUF", ""),
        
        # Voyage 4: LO 02-15, SAIL 02-16, ARR 02-18
        ("V4", "5.0", "VOYAGE 4: TR4 Transport + JD-2", "MILESTONE", "All", 28, 0, "✅ Tide ≥1.90m (1.90m) | Shamal 종료 직후"),
        ("LO-128", "5.1", "TR4 Load-out on LCT", "LOADOUT", "Mammoet", 28, "DUR_LO", "Tide ≥1.90m (1.90m)"),
        ("SF-129", "5.2", "TR4 Sea Fastening", "SEAFAST", "Mammoet", 28, "DUR_SF", ""),
        ("MWS-129A", "5.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 28, "DUR_MWS", ""),
        ("SAIL-130", "5.3", "V4 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 29, "DUR_SAIL", ""),
        ("ARR-131", "5.4", "AGI Arrival + TR4 RORO Unload", "AGI_UNLOAD", "Mammoet", 31, "DUR_UL", "Tide ≥1.90m (1.96m)"),
        ("TRN-132", "5.5", "TR3 Transport to Bay-3", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-133", "5.6", "TR3 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("TRN-135", "5.8", "TR4 Transport to Bay-4", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-136", "5.9", "TR4 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("JD-139A", "5.95", "JD-2 Jack-Down TR3", "JACKDOWN", "Mammoet", 33, "DUR_JD", "MILESTONE: TR3 complete | 02-20"),
        ("RET-138", "5.11", "V4 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 34, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-139B", "5.96", "JD-2 Jack-Down TR4", "JACKDOWN", "Mammoet", 35, "DUR_JD", "MILESTONE: TR4 complete | 02-21"),
        ("BUF-140", "5.99", "V4 Buffer", "BUFFER", "All", 36, "DUR_BUF", ""),
        
        # Voyage 5: LO 02-23, SAIL 02-23, ARR 02-24 (Fast-turn)
        ("V5", "6.0", "VOYAGE 5: TR5 Transport", "MILESTONE", "All", 36, 0, "✅ Tide ≥1.90m (1.99m) | Fast-turn"),
        ("LO-140", "6.1", "TR5 Load-out on LCT", "LOADOUT", "Mammoet", 36, "DUR_LO", "Tide ≥1.90m (1.99m)"),
        ("SF-141", "6.2", "TR5 Sea Fastening", "SEAFAST", "Mammoet", 36, "DUR_SF", ""),
        ("MWS-141A", "6.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 36, "DUR_MWS", ""),
        ("SAIL-142", "6.3", "V5 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 36, "DUR_SAIL", "Fast-turn"),
        ("ARR-143", "6.4", "AGI Arrival + TR5 RORO Unload", "AGI_UNLOAD", "Mammoet", 37, "DUR_UL", "Tide ≥1.90m (2.01m)"),
        ("STORE-144", "6.5", "TR5 Stored on AGI Laydown", "BUFFER", "Mammoet", 37, "DUR_BUF", "Awaiting pair TR6"),
        ("RET-145", "6.6", "V5 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 37, "DUR_RET", ""),
        ("BUF-146", "6.99", "V5 Buffer", "BUFFER", "All", 37, "DUR_BUF", ""),
        
        # Voyage 6: LO 02-25, SAIL 02-25, ARR 02-26 (Fast-turn)
        ("V6", "7.0", "VOYAGE 6: TR6 Transport + JD-3", "MILESTONE", "All", 38, 0, "✅ Tide ≥1.90m (2.01m) | Fast-turn"),
        ("LO-147", "7.1", "TR6 Load-out on LCT", "LOADOUT", "Mammoet", 38, "DUR_LO", "Tide ≥1.90m (2.01m)"),
        ("SF-148", "7.2", "TR6 Sea Fastening", "SEAFAST", "Mammoet", 38, "DUR_SF", ""),
        ("MWS-148A", "7.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 38, "DUR_MWS", ""),
        ("SAIL-149", "7.3", "V6 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 38, "DUR_SAIL", "Fast-turn"),
        ("ARR-150", "7.4", "AGI Arrival + TR6 RORO Unload", "AGI_UNLOAD", "Mammoet", 39, "DUR_UL", "Tide ≥1.90m (1.98m)"),
        ("TRN-151", "7.5", "TR5 Transport to Bay-5", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-152", "7.6", "TR5 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("TRN-154", "7.8", "TR6 Transport to Bay-6", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-155", "7.9", "TR6 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("JD-157A", "7.95", "JD-3 Jack-Down TR5", "JACKDOWN", "Mammoet", 40, "DUR_JD", "MILESTONE: TR5 complete | 02-27"),
        ("RET-158", "7.11", "V6 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-157B", "7.96", "JD-3 Jack-Down TR6", "JACKDOWN", "Mammoet", 42, "DUR_JD", "MILESTONE: TR6 complete | 02-28"),
        ("BUF-159", "7.99", "V6 Buffer / Reset for V7", "BUFFER", "All", 43, "DUR_BUF", ""),
        
        # Voyage 7: LO 02-27, SAIL 02-27, ARR 02-28 (Final)
        ("V7", "8.0", "VOYAGE 7: TR7 Transport + JD-4", "MILESTONE", "All", 40, 0, "✅ Tide ≥1.90m (1.92m) | Final unit"),
        ("LO-201", "8.1", "TR7 Load-out on LCT", "LOADOUT", "Mammoet", 40, "DUR_LO", "Tide ≥1.90m (1.92m) required"),
        ("SF-202", "8.2", "TR7 Sea Fastening", "SEAFAST", "Mammoet", 40, "DUR_SF", "12-point lashing"),
        ("MWS-202A", "8.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 40, "DUR_MWS", ""),
        ("SAIL-203", "8.3", "V7 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 40, "DUR_SAIL", "Weather window required"),
        ("ARR-204", "8.4", "AGI Arrival + TR7 RORO Unload", "AGI_UNLOAD", "Mammoet", 41, "DUR_UL", "Tide ≥1.90m (1.93m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-205", "8.5", "TR7 Transport to Bay-7", "TURNING", "Mammoet", 41, 1, "Steel bridge install"),
        ("TURN-206", "8.6", "TR7 Turning (90° rotation)", "TURNING", "Mammoet", 41, "DUR_TURN", "10t Forklift required"),
        ("JD-207", "8.7", "★ JD-4 Jack-Down (TR7)", "JACKDOWN", "Mammoet", 41, "DUR_JD", "MILESTONE: TR7 Complete | 02-28"),
        ("RET-208", "8.8", "V7 LCT Final Return: AGI→MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Final return"),
        
        # Demobilization
        ("DEMOB", "9.0", "DEMOBILIZATION", "MOBILIZATION", "Mammoet", 42, "DUR_MOB", "Equipment return"),
        ("END", "99.0", "★★★ PROJECT COMPLETE ★★★", "MILESTONE", "All", 42, 0, "All 7 TRs Installed | Jan-Feb 2026 Complete"),
    ]

    tasks = default_tasks
    if tsv_path:
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
            print(f"✅ Loaded {len(tasks)} tasks from TSV file")
        except Exception as e:
            print(f"Error loading TSV: {e}")
            print("Using default tasks list")
    
    # Schedule_Data and Gantt_Chart sheets removed; tasks remain for summary/weather calculations.
    
    # === Option Sheets ===
    option_a_created = False
    option_b_created = False
    option_c_created = False

    if option_a_tsv and os.path.exists(option_a_tsv):
        option_a_created = create_scenario_sheets(
            wb,
            "Option_A",
            option_a_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if option_a_created:
            print(f"Created Option_A sheets from {os.path.basename(option_a_tsv)}")
        else:
            print("Warning: Option_A sheets not created")

    if option_b_tsv and os.path.exists(option_b_tsv):
        option_b_created = create_scenario_sheets(
            wb,
            "Option_B",
            option_b_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if option_b_created:
            print(f"Created Option_B sheets from {os.path.basename(option_b_tsv)}")
        else:
            print("Warning: Option_B sheets not created")

    if option_c_tsv and os.path.exists(option_c_tsv):
        option_c_created = create_scenario_sheets(
            wb,
            "Option_C",
            option_c_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if option_c_created:
            print(f"Created Option_C sheets from {os.path.basename(option_c_tsv)}")
        else:
            print("Warning: Option_C sheets not created")
    
    # === Tide Data Sheet ===
    if tide_json and os.path.exists(tide_json):
        create_tide_data_sheet(wb, tide_json_path=tide_json)
        print("✅ Created Tide_Data sheet (from JSON)")
    elif tide_tsv and os.path.exists(tide_tsv):
        create_tide_data_sheet(wb, tide_tsv_path=tide_tsv)
        print("✅ Created Tide_Data sheet (from TSV)")
    
    # === Scenario Comparison ===
    if option_a_created and option_b_created:
        create_comparison_summary(
            wb,
            project_start,
            scenario_a="Option_A",
            scenario_b="Option_B",
            label_a="Option A",
            label_b="Option B",
            sheet_name="Scenario_Comparison_AB",
        )
        create_option_schedule_comparison_sheet(
            wb,
            scenario_a="Option_A",
            scenario_b="Option_B",
            sheet_name="Option_Compare_AB",
        )

    if option_a_created and option_c_created:
        create_comparison_summary(
            wb,
            project_start,
            scenario_a="Option_A",
            scenario_b="Option_C",
            label_a="Option A",
            label_b="Option C",
            sheet_name="Scenario_Comparison_AC",
        )
        create_option_schedule_comparison_sheet(
            wb,
            scenario_a="Option_A",
            scenario_b="Option_C",
            sheet_name="Option_Compare_AC",
        )

    if option_b_created and option_c_created:
        create_comparison_summary(
            wb,
            project_start,
            scenario_a="Option_B",
            scenario_b="Option_C",
            label_a="Option B",
            label_b="Option C",
            sheet_name="Scenario_Comparison_BC",
        )
        create_option_schedule_comparison_sheet(
            wb,
            scenario_a="Option_B",
            scenario_b="Option_C",
            sheet_name="Option_Compare_BC",
        )

    if option_a_created and option_b_created:
        create_three_way_comparison(
            wb,
            project_start,
            scenario_a="Option_A",
            scenario_b="Option_B",
            scenario_c="Option_C" if option_c_created else None,
            label_a="Option A",
            label_b="Option B",
            label_c="Option C" if option_c_created else None,
            sheet_name="Three_Way_Comparison",
        )

    
    # === WEATHER ANALYSIS (동적화) ===
    ws_weather = wb.create_sheet("Weather_Analysis")
    
    try:
        project_end = ws_ctrl["B5"].value
        if hasattr(project_end, "date"):
            project_end = project_end.date()
        if not isinstance(project_end, dt.date):
            raise ValueError("Invalid target end date")
    except Exception:
        max_offset = 0
        for task in tasks:
            if len(task) >= 6:
                offset = task[5] if isinstance(task[5], (int, float)) else 0
                max_offset = max(max_offset, offset)
        project_end = project_start + dt.timedelta(days=int(max_offset) + 30)
    
    year_month = project_start.strftime("%b %Y")
    ws_weather["A1"] = f"UAE Winter Weather Analysis - {year_month}"
    ws_weather["A1"].font = Font(bold=True, size=14)
    
    weather_periods = generate_weather_periods(project_start, project_end)
    period_names = [p[0] for p in weather_periods]
    
    weather_headers = ["Parameter"] + period_names + ["Notes"]
    num_cols = len(weather_headers)
    ws_weather.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    
    for c, h in enumerate(weather_headers, 1):
        cell = ws_weather.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = tb()
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    
    def get_weather_for_period(period_start, period_end, shamal_start, shamal_end):
        is_shamal = (period_start <= shamal_end and period_end >= shamal_start)
        if is_shamal:
            return {
                "wind": "16-21",
                "gust": "25-30",
                "wave": "0.8-1.2",
                "visibility": "2-5",
                "risk": "HIGH",
                "recommendation": "NO-GO",
            }
        return {
            "wind": "11-13",
            "gust": "18-20",
            "wave": "0.4-0.6",
            "visibility": "8-10",
            "risk": "LOW",
            "recommendation": "GO",
        }
    
    weather_data_rows = [
        ["Avg Wind (kt)"] + [""] * len(weather_periods) + [f"Peak Shamal: {shamal_text}"],
        ["Max Gust (kt)"] + [""] * len(weather_periods) + ["NO-GO if >22kt gust"],
        ["Wave Height (m)"] + [""] * len(weather_periods) + ["HOLD if >0.8m"],
        ["Visibility (km)"] + [""] * len(weather_periods) + ["Reduced during Shamal"],
        ["Risk Level"] + [""] * len(weather_periods) + [""],
        ["Recommendation"] + [""] * len(weather_periods) + [""],
    ]
    
    for idx, (_name, period_start, period_end) in enumerate(weather_periods):
        weather = get_weather_for_period(period_start, period_end, shamal_start, shamal_end)
        col = idx + 1
        weather_data_rows[0][col] = weather["wind"]
        weather_data_rows[1][col] = weather["gust"]
        weather_data_rows[2][col] = weather["wave"]
        weather_data_rows[3][col] = weather["visibility"]
        weather_data_rows[4][col] = weather["risk"]
        weather_data_rows[5][col] = weather["recommendation"]
    
    for r, row in enumerate(weather_data_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_weather.cell(r, c, value=val)
            cell.border = tb()
            if val == "HIGH" or val == "NO-GO":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif val == "MEDIUM" or val == "CAUTION":
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
            elif val == "LOW" or val == "GO":
                cell.fill = PatternFill("solid", fgColor="C8E6C9")
    
    for col in range(1, num_cols + 1):
        ws_weather.column_dimensions[get_column_letter(col)].width = 15
    
    # === SUMMARY (기존과 동일) ===
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "AGI HVDC Transformer Transportation - Project Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")
    
    summary_data = [
        ("Key Parameters", ""),
        ("Total Transformers", "7 units (TR1-TR7)"),
        ("Weight per TR", "217-271 tons"),
        ("Total Voyages", "7 (1 TR per voyage)"),
        ("Jack-down Events", "4 (after V2, V4, V6, V7)"),
        ("Vessel", "LCT BUSHRA"),
        ("Route", "Mina Zayed Port ↔ AGI Site"),
        ("", ""),
        ("Schedule Summary", ""),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
    ]
    
    voyage_ranges = calculate_voyage_ranges(tasks)
    for voyage_name, start_day, end_day in voyage_ranges:
        if "Mobilization" in voyage_name:
            label = voyage_name
        else:
            if ":" in voyage_name:
                prefix, rest = voyage_name.split(":", 1)
                label = f"{prefix.title()} ({rest.strip()})"
            else:
                label = voyage_name.title()
        summary_data.append((label, f"Day {start_day}-{end_day}"))
    
    project_complete_formula = "=MAX(Schedule_Data_Option_A!H:H,Schedule_Data_Option_B!H:H)"
    if option_c_created:
        project_complete_formula = "=MAX(Schedule_Data_Option_A!H:H,Schedule_Data_Option_B!H:H,Schedule_Data_Option_C!H:H)"

    summary_data.extend([
        ("Project Complete", project_complete_formula),
        ("", ""),
        ("Weather Constraints", ""),
        ("Shamal Period", shamal_text_full),
        ("Tide Requirement", f"≥{tide_threshold:.2f}m for LO/ARR"),
        ("Wind Limit", "≤18kt sustained, ≤22kt gust"),
        ("Wave Limit", "≤0.8m (HOLD), ≤1.0m (NO-GO)"),
        ("AGI Draft Limit", "≤2.70m forward draft"),
    ])
    
    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30
    
    # === VBA CODE SHEET ===
    ws_vba = wb.create_sheet("VBA_Code")
    
    ws_vba["A1"] = "📋 VBA 코드 - Alt+F11 → Module에 붙여넣기 → .xlsm으로 저장"
    ws_vba["A1"].font = Font(bold=True, size=14)
    
    vba_code = '''
Option Explicit

' ============================================
' AGI TR Multi-Scenario Master Gantt - VBA Macros
' ============================================
' 사용법
' 1) Alt+F11 → Module 삽입 → 이 코드 붙여넣기
' 2) Control_Panel 시트 코드영역 + ThisWorkbook에도 아래 이벤트 코드 추가
' 3) .xlsm으로 저장
' ============================================

' ----------------------------
' App State (성능/안정성)
' ----------------------------
Private Type TAppState
    Calculation As XlCalculation
    ScreenUpdating As Boolean
    EnableEvents As Boolean
    DisplayStatusBar As Boolean
End Type

Private gState As TAppState

Private Sub BeginFastMode()
    With Application
        gState.Calculation = .Calculation
        gState.ScreenUpdating = .ScreenUpdating
        gState.EnableEvents = .EnableEvents
        gState.DisplayStatusBar = .DisplayStatusBar

        .ScreenUpdating = False
        .EnableEvents = False
        .DisplayStatusBar = True
        .Calculation = xlCalculationManual
    End With
End Sub

Private Sub EndFastMode()
    With Application
        .Calculation = gState.Calculation
        .ScreenUpdating = gState.ScreenUpdating
        .EnableEvents = gState.EnableEvents
        .DisplayStatusBar = gState.DisplayStatusBar
    End With
End Sub

' 안전장치: 이벤트/계산이 꺼진 채로 남았을 때 수동 복구
Public Sub ResetAppState()
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    Application.Calculation = xlCalculationAutomatic
End Sub

Private Function SheetExists(ByVal sheetName As String) As Boolean
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = Sheets(sheetName)
    SheetExists = Not ws Is Nothing
    Set ws = Nothing
    On Error GoTo 0
End Function

' ----------------------------
' Date Coercion (여러 타입의 날짜 대응)
' ----------------------------
Private Function CoerceDate(ByVal v As Variant, Optional ByVal baseDate As Date = 0) As Date
    On Error GoTo Fail

    If IsDate(v) Then
        CoerceDate = CDate(v)
        Exit Function
    End If

    ' Excel serial number (Double/Long)
    If IsNumeric(v) Then
        CoerceDate = DateSerial(1899, 12, 30) + CDbl(v)
        Exit Function
    End If

    Dim s As String
    s = Trim$(CStr(v))
    If s = "" Then Exit Function

    If baseDate = 0 Then baseDate = Date

    ' dd-mmm (no year) → baseDate 연도 적용
    If InStr(1, s, "-", vbTextCompare) > 0 And Len(s) <= 6 Then
        Dim dd As Integer, monText As String
        dd = CInt(Left$(s, InStr(1, s, "-") - 1))
        monText = Mid$(s, InStr(1, s, "-") + 1)
        CoerceDate = DateSerial(Year(baseDate), Month(DateValue("1-" & monText)), dd)
        Exit Function
    End If

    CoerceDate = DateValue(s)
    Exit Function

Fail:
    CoerceDate = 0
End Function

' ============================================
' 1) 통합 업데이트
' ============================================
Public Sub UpdateAllScenarios()
    On Error GoTo CleanUp
    BeginFastMode

    ' 계산 먼저
    Sheets("Schedule_Data_Option_A").Calculate
    Sheets("Schedule_Data_Option_B").Calculate
    Sheets("Gantt_Chart_Option_A").Calculate
    Sheets("Gantt_Chart_Option_B").Calculate

    On Error Resume Next
    Sheets("Schedule_Data_Option_C").Calculate
    Sheets("Gantt_Chart_Option_C").Calculate
    Sheets("Scenario_Comparison_AB").Calculate
    Sheets("Scenario_Comparison_AC").Calculate
    Sheets("Scenario_Comparison_BC").Calculate
    Sheets("Three_Way_Comparison").Calculate
    Sheets("Option_Compare_AB").Calculate
    Sheets("Option_Compare_AC").Calculate
    Sheets("Option_Compare_BC").Calculate
    Sheets("Tide_Data").Calculate
    Sheets("Weather_Analysis").Calculate
    Sheets("Summary").Calculate
    Sheets("Control_Panel").Calculate
    On Error GoTo CleanUp

    ' 시각 갱신
    RefreshAllGanttCharts
    RefreshTideData

CleanUp:
    EndFastMode

    If Err.Number <> 0 Then
        MsgBox "❌ 업데이트 실패: " & Err.Description, vbExclamation, "UpdateAllScenarios"
    Else
        MsgBox "✅ Option A/B/C + Compare 업데이트 완료", vbInformation, "Update Complete"
    End If
End Sub

' ============================================
' 2) Gantt 색상 갱신 (공통 코어)
' ============================================
Public Sub RefreshAllGanttCharts()
    On Error GoTo CleanUp
    BeginFastMode

    RefreshGanttChart_Option_A
    RefreshGanttChart_Option_B
    RefreshGanttChart_Option_C

CleanUp:
    EndFastMode
End Sub

Private Sub RefreshGanttCore(ByVal ganttSheetName As String, ByVal schedSheetName As String)
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long
    Dim startD As Date, endD As Date
    Dim projStart As Date, cellDate As Date
    Dim phase As String
    Dim dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date

    Set ws = Sheets(ganttSheetName)
    Set wsd = Sheets(schedSheetName)

    projStart = CoerceDate(Sheets("Control_Panel").Range("B4").Value)
    shamalStart = CoerceDate(Sheets("Control_Panel").Range("H5").Value, projStart)
    shamalEnd = CoerceDate(Sheets("Control_Panel").Range("H6").Value, projStart)

    dc = 5 ' Date columns start at E
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row

    ' Clear existing colors in date columns
    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone

    ' Reset header colors + Shamal highlight
    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If shamalStart <> 0 And shamalEnd <> 0 Then
            If cellDate >= shamalStart And cellDate <= shamalEnd Then
                ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
            End If
        End If
    Next j

    ' Apply Gantt bars
    For i = 6 To lastRow
        If Trim$(CStr(wsd.Cells(i, 3).Value)) = "" Then GoTo NextRow

        startD = CoerceDate(wsd.Cells(i, 7).Value, projStart)
        endD = CoerceDate(wsd.Cells(i, 8).Value, projStart)
        If startD = 0 Then GoTo NextRow
        If endD = 0 Then endD = startD

        phase = CStr(wsd.Cells(i, 5).Value)

        Dim ganttRow As Long
        ganttRow = i - 1

        For j = 0 To maxJ
            cellDate = projStart + j
            If cellDate >= startD And cellDate < endD Then
                ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
            ElseIf cellDate = startD And startD = endD Then
                With ws.Cells(ganttRow, dc + j)
                    .Interior.Color = GetPhaseColor(phase)
                    .Value = ChrW(9733)
                    .HorizontalAlignment = xlCenter
                    .Font.Size = 8
                End With
            End If
        Next j

NextRow:
    Next i
End Sub

Public Sub RefreshGanttChart_Option_A()
    RefreshGanttCore "Gantt_Chart_Option_A", "Schedule_Data_Option_A"
End Sub

Public Sub RefreshGanttChart_Option_B()
    RefreshGanttCore "Gantt_Chart_Option_B", "Schedule_Data_Option_B"
End Sub

Public Sub RefreshGanttChart_Option_C()
    RefreshGanttCore "Gantt_Chart_Option_C", "Schedule_Data_Option_C"
End Sub

' ============================================
' 3) 조석 데이터 강조 (매 실행 시 초기화 포함)
' ============================================
Public Sub RefreshTideData()
    Dim ws As Worksheet
    Dim i As Long
    Dim tideThreshold As Double

    Set ws = Sheets("Tide_Data")
    tideThreshold = 0
    On Error Resume Next
    tideThreshold = CDbl(Sheets("Control_Panel").Range("H7").Value)
    On Error GoTo 0
    If tideThreshold = 0 Then tideThreshold = 1.9

    ' Reset formatting (A:D, data rows)
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    If lastRow < 5 Then Exit Sub

    ws.Range(ws.Cells(5, 1), ws.Cells(lastRow, 4)).Font.Bold = False
    ws.Range(ws.Cells(5, 1), ws.Cells(lastRow, 4)).Font.Color = RGB(0, 0, 0)
    ws.Range(ws.Cells(5, 1), ws.Cells(lastRow, 1)).Interior.ColorIndex = xlNone

    For i = 5 To lastRow
        If IsNumeric(ws.Cells(i, 3).Value) Then
            If CDbl(ws.Cells(i, 3).Value) >= tideThreshold Then
                ws.Cells(i, 3).Font.Bold = True
                ws.Cells(i, 3).Font.Color = RGB(0, 102, 204)
                ws.Cells(i, 1).Interior.Color = RGB(227, 242, 253)
            End If
        End If
    Next i
End Sub

' ============================================
' 4) Phase Color Helper
' ============================================
Public Function GetPhaseColor(phase As String) As Long
    Select Case phase
        Case "MOBILIZATION": GetPhaseColor = RGB(142, 124, 195)
        Case "DECK_PREP": GetPhaseColor = RGB(111, 168, 220)
        Case "LOADOUT": GetPhaseColor = RGB(147, 196, 125)
        Case "SEAFAST": GetPhaseColor = RGB(118, 165, 175)
        Case "SAIL": GetPhaseColor = RGB(164, 194, 244)
        Case "AGI_UNLOAD": GetPhaseColor = RGB(246, 178, 107)
        Case "TURNING": GetPhaseColor = RGB(255, 217, 102)
        Case "JACKDOWN": GetPhaseColor = RGB(224, 102, 102)
        Case "RETURN": GetPhaseColor = RGB(153, 153, 153)
        Case "BUFFER": GetPhaseColor = RGB(217, 217, 217)
        Case "MILESTONE": GetPhaseColor = RGB(255, 0, 0)
        Case Else: GetPhaseColor = RGB(255, 255, 255)
    End Select
End Function

' ============================================
' 5) PDF 내보내기 (핵심 시트만)
' ============================================
Public Sub ExportToPDF()
    Dim fp As String
    Dim names As Variant
    Dim outNames() As String
    Dim i As Long
    Dim count As Long

    fp = ThisWorkbook.Path & "\AGI_TR_Gantt_" & Format(Date, "YYYYMMDD") & ".pdf"

    names = Array( _
        "Control_Panel", _
        "Scenario_Comparison_AB", _
        "Scenario_Comparison_AC", _
        "Scenario_Comparison_BC", _
        "Three_Way_Comparison", _
        "Option_Compare_AB", _
        "Option_Compare_AC", _
        "Option_Compare_BC", _
        "Gantt_Chart_Option_A", _
        "Gantt_Chart_Option_B", _
        "Gantt_Chart_Option_C" _
    )

    For i = LBound(names) To UBound(names)
        If SheetExists(CStr(names(i))) Then
            ReDim Preserve outNames(0 To count)
            outNames(count) = CStr(names(i))
            count = count + 1
        End If
    Next i

    If count = 0 Then
        MsgBox "No sheets to export.", vbExclamation, "Export Cancelled"
        Exit Sub
    End If

    Sheets(outNames).Select
    ActiveSheet.ExportAsFixedFormat xlTypePDF, fp, xlQualityStandard, True
    Sheets("Control_Panel").Select

    MsgBox "✅ PDF 저장 완료:" & vbCrLf & fp, vbInformation, "Export Complete"
End Sub

' ============================================
' 6) 간단 리포트 (Option A/B 끝나는 날짜)
' ============================================
Public Sub GenerateReport()
    Dim endA As Date, endB As Date
    Dim endC As Date
    Dim hasC As Boolean
    endA = WorksheetFunction.Max(Sheets("Schedule_Data_Option_A").Range("H:H"))
    endB = WorksheetFunction.Max(Sheets("Schedule_Data_Option_B").Range("H:H"))
    hasC = SheetExists("Schedule_Data_Option_C")
    If hasC Then
        endC = WorksheetFunction.Max(Sheets("Schedule_Data_Option_C").Range("H:H"))
    End If

    Dim msg As String
    msg = "Project Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
          "Option A End: " & Format(endA, "YYYY-MM-DD") & vbCrLf & _
          "Option B End: " & Format(endB, "YYYY-MM-DD") & vbCrLf & _
          "Delta (B-A, days): " & CStr(endB - endA)

    If hasC Then
        msg = msg & vbCrLf & "Option C End: " & Format(endC, "YYYY-MM-DD") & vbCrLf & _
              "Delta (C-A, days): " & CStr(endC - endA)
    End If

    MsgBox msg, vbInformation, "Project Report"
End Sub

' ============================================
' 7) Shamal 기간 작업 점검 (Option A/B)
' ============================================
Public Sub CheckShamalRisk()
    Dim shamalStart As Date, shamalEnd As Date
    shamalStart = CoerceDate(Sheets("Control_Panel").Range("H5").Value)
    shamalEnd = CoerceDate(Sheets("Control_Panel").Range("H6").Value)

    Dim riskTasks As String
    Dim cnt As Long

    cnt = cnt + CheckShamalRiskOne("Schedule_Data_Option_A", shamalStart, shamalEnd, riskTasks)
    cnt = cnt + CheckShamalRiskOne("Schedule_Data_Option_B", shamalStart, shamalEnd, riskTasks)
    If SheetExists("Schedule_Data_Option_C") Then
        cnt = cnt + CheckShamalRiskOne("Schedule_Data_Option_C", shamalStart, shamalEnd, riskTasks)
    End If

    If cnt > 0 Then
        MsgBox "⚠️ SHAMAL 위험 경고!" & vbCrLf & vbCrLf & _
               "Shamal 기간 (" & Format(shamalStart, "MM/DD") & "-" & Format(shamalEnd, "MM/DD") & ") 중 " & cnt & "개 기상 민감 작업 발견:" & vbCrLf & _
               riskTasks, vbExclamation, "Weather Risk Alert"
    Else
        MsgBox "✅ Shamal 기간 중 기상 민감 작업 없음", vbInformation, "Weather Check OK"
    End If
End Sub

Private Function CheckShamalRiskOne(ByVal schedName As String, ByVal shamalStart As Date, ByVal shamalEnd As Date, ByRef outList As String) As Long
    Dim wsd As Worksheet
    Dim i As Long, lastRow As Long
    Dim taskDate As Date

    Set wsd = Sheets(schedName)
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row

    For i = 6 To lastRow
        taskDate = CoerceDate(wsd.Cells(i, 7).Value)
        If taskDate <> 0 Then
            If taskDate >= shamalStart And taskDate <= shamalEnd Then
                If wsd.Cells(i, 5).Value = "SAIL" Or wsd.Cells(i, 5).Value = "LOADOUT" Then
                    CheckShamalRiskOne = CheckShamalRiskOne + 1
                    outList = outList & vbCrLf & "  - [" & schedName & "] " & wsd.Cells(i, 3).Value & ": " & wsd.Cells(i, 4).Value
                End If
            End If
        End If
    Next i
End Function

' ============================================
' 8) Control Panel 설정 보기
' ============================================
Public Sub ShowControlPanelSettings()
    Dim msg As String
    msg = "Project Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
          "Target End: " & Format(Sheets("Control_Panel").Range("B5").Value, "YYYY-MM-DD") & vbCrLf & _
          "Shamal: " & Format(Sheets("Control_Panel").Range("H5").Value, "YYYY-MM-DD") & " ~ " & Format(Sheets("Control_Panel").Range("H6").Value, "YYYY-MM-DD") & vbCrLf & _
          "Tide Threshold: " & Format(Sheets("Control_Panel").Range("H7").Value, "0.00") & "m"
    MsgBox msg, vbInformation, "Control Panel Settings"
End Sub

' ============================================
' 9) 이벤트 코드 (추가로 붙여넣기)
' ============================================
' [A] Control_Panel 시트 코드 영역에 추가:
' Private Sub Worksheet_Change(ByVal Target As Range)
'     On Error GoTo CleanUp
'     If Intersect(Target, Me.Range("B4:B5,H5:H7,H10:H11")) Is Nothing Then Exit Sub
'     Application.EnableEvents = False
'     UpdateAllScenarios
' CleanUp:
'     Application.EnableEvents = True
' End Sub

' [B] ThisWorkbook 코드 영역에 추가:
' Private Sub Workbook_Open()
'     Application.EnableEvents = True
' End Sub
'''
    for i, line in enumerate(vba_code.strip().split('\n'), 3):
        ws_vba.cell(i, 1, value=line)
        ws_vba.cell(i, 1).font = Font(name="Consolas", size=9)
    
    ws_vba.column_dimensions["A"].width = 100
    
    return wb

if __name__ == "__main__":
    import os
    import sys
    from datetime import datetime
    
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
    
    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    
    option_a_tsv_path = os.path.join(script_dir, "OPTION A.tsv")
    option_a_tsv = option_a_tsv_path if os.path.exists(option_a_tsv_path) else None

    option_b_tsv_path = os.path.join(script_dir, "option b.tsv")
    option_b_tsv = option_b_tsv_path if os.path.exists(option_b_tsv_path) else None

    option_c_tsv_path = os.path.join(script_dir, "option_c.tsv")
    if not os.path.exists(option_c_tsv_path):
        option_c_tsv_path = os.path.join(script_dir, "option c.tsv")
    option_c_tsv = option_c_tsv_path if os.path.exists(option_c_tsv_path) else None
    tide_tsv = os.path.join(script_dir, "Date High Tide Window Max Height (m) Ris.tsv")
    tide_json = os.path.join(script_dir, "MINA ZAYED PORT WATER TIDE_MERGED.json")
    
    default_tsv = os.path.join(script_dir, "ID WBS Task Phase Owner Start End Durati.tsv")
    tsv_path = default_tsv if os.path.exists(default_tsv) else None
    
    print("Generating AGI TR Multi-Scenario Master Gantt with VBA...")
    wb = create_gantt_with_vba(
        tsv_path=tsv_path,
        option_a_tsv=option_a_tsv,
        option_b_tsv=option_b_tsv,
        option_c_tsv=option_c_tsv,
        tide_tsv=tide_tsv if os.path.exists(tide_tsv) else None,
        tide_json=tide_json if os.path.exists(tide_json) else None
    )
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(
        os.getcwd(),
        f"AGI_TR_MultiScenario_Master_Gantt_{timestamp}.xlsx",
    )
    wb.save(output_path)
    print(f"[OK] Generated: {output_path}")
