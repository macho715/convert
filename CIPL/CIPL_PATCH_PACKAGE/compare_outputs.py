#!/usr/bin/env python3
"""
간단한 출력 비교 스크립트
기준 버전과 최적화 버전의 주요 셀을 비교
"""
from openpyxl import load_workbook
from pathlib import Path

def compare_sheets(ws1, ws2, sheet_name):
    """두 워크시트의 주요 셀 비교"""
    key_cells = [
        "B3",  # Title
        "A5", "B5", "F5",  # Header row
        "E24", "C29",  # Key content
        "J4",  # Page number
    ]
    
    differences = []
    matches = 0
    
    for cell_ref in key_cells:
        try:
            c1 = ws1[cell_ref]
            c2 = ws2[cell_ref]
            
            # Value 비교
            if c1.value != c2.value:
                differences.append(f"{cell_ref}: value mismatch ('{c1.value}' vs '{c2.value}')")
                continue
            
            # Font 비교 (간단히)
            if str(c1.font) != str(c2.font):
                differences.append(f"{cell_ref}: font mismatch")
                continue
            
            # Alignment 비교
            if str(c1.alignment) != str(c2.alignment):
                differences.append(f"{cell_ref}: alignment mismatch")
                continue
            
            matches += 1
        except Exception as e:
            differences.append(f"{cell_ref}: error - {e}")
    
    return matches, differences

def main():
    baseline_path = Path("VERIFY_BASELINE.xlsx")
    optimized_path = Path("VERIFY_OPTIMIZED_FINAL.xlsx")
    
    if not baseline_path.exists():
        print(f"[ERROR] Baseline file not found: {baseline_path}")
        return 1
    
    if not optimized_path.exists():
        print(f"[ERROR] Optimized file not found: {optimized_path}")
        return 1
    
    print("=" * 60)
    print("출력 비교 시작")
    print("=" * 60)
    
    wb_baseline = load_workbook(baseline_path)
    wb_optimized = load_workbook(optimized_path)
    
    # Commercial Invoice 비교 (시트 이름 확인)
    print(f"\nBaseline sheets: {wb_baseline.sheetnames}")
    print(f"Optimized sheets: {wb_optimized.sheetnames}")
    
    # 첫 번째 시트 비교
    if len(wb_baseline.sheetnames) > 0 and len(wb_optimized.sheetnames) > 0:
        sheet_name = wb_baseline.sheetnames[0]
        ws1 = wb_baseline[sheet_name]
        if sheet_name in wb_optimized.sheetnames:
            ws2 = wb_optimized[sheet_name]
        else:
            ws2 = wb_optimized[wb_optimized.sheetnames[0]]
        
        matches, diffs = compare_sheets(ws1, ws2, sheet_name)
        print(f"\n{sheet_name}:")
        print(f"  일치: {matches}/{8} 셀")
        if diffs:
            print(f"  차이점: {len(diffs)}개")
            for diff in diffs[:5]:  # 최대 5개만 표시
                print(f"    - {diff}")
            if len(diffs) > 5:
                print(f"    ... 외 {len(diffs) - 5}개 차이점")
        else:
            print("  [OK] 모든 주요 셀 일치")
    
    # 두 번째 시트 비교 (Packing List)
    if len(wb_baseline.sheetnames) > 1 and len(wb_optimized.sheetnames) > 1:
        sheet_name = wb_baseline.sheetnames[2]  # Packing_List_P1
        ws1 = wb_baseline[sheet_name]
        if sheet_name in wb_optimized.sheetnames:
            ws2 = wb_optimized[sheet_name]
        else:
            ws2 = wb_optimized[wb_optimized.sheetnames[2]]
        
        matches, diffs = compare_sheets(ws1, ws2, sheet_name)
        print(f"\n{sheet_name}:")
        print(f"  일치: {matches}/{8} 셀")
        if diffs:
            print(f"  차이점: {len(diffs)}개")
            for diff in diffs[:5]:
                print(f"    - {diff}")
            if len(diffs) > 5:
                print(f"    ... 외 {len(diffs) - 5}개 차이점")
        else:
            print("  [OK] 모든 주요 셀 일치")
    
    print("\n" + "=" * 60)
    print("비교 완료")
    print("=" * 60)
    print(f"\n파일 위치:")
    print(f"  기준: {baseline_path.absolute()}")
    print(f"  최적화: {optimized_path.absolute()}")
    print("\n[TIP] Excel에서 두 파일을 열어 시각적으로 확인하세요.")
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

