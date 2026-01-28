#!/usr/bin/env python3
"""
AGI Transformer Transportation Scheduler (openpyxl)
- Start date driven (Control_Panel!B4)
- Generates "Schedule_Data" + "Gantt_Chart"
- Single-train assumptions (1 LCT + 1 SPMT): tasks are serialized (no overlap)

Scenarios:
  1) Single-Train 1+1+1+1+1+1+1 (7 voyages, 1 TR each)
  2) Batch 1+2+2+2 (4 voyages: TR1 / TR2-3 / TR4-5 / TR6-7)

Usage:
  python agi_tr_scheduler_scenarios.py --scenario single --start 2026-01-15
  python agi_tr_scheduler_scenarios.py --scenario batch  --start 2026-01-15
"""
from __future__ import annotations

import argparse
import datetime as dt
import math
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "PREWORK": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "INPUT": "FFFDE7",
}

BORDER = Side(style="thin", color="A6A6A6")

def tb() -> Border:
    return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def build_workbook(start_date: dt.date, batch_pattern: list[int], title_suffix: str) -> Workbook:
    wb = Workbook()
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"

    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = f"AGI TR Transportation Schedule - {title_suffix}"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ctrl.row_dimensions[1].height = 28

    ws_ctrl["A3"] = "Scenario:"
    ws_ctrl["A3"].font = Font(bold=True)
    ws_ctrl["B3"] = title_suffix
    ws_ctrl["B3"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B3"].border = tb()

    dv = DataValidation(type="list",
                        formula1='"Single-Train 1+1+1+1+1+1+1,Batch 1+2+2+2"',
                        allow_blank=False)
    ws_ctrl.add_data_validation(dv)
    dv.add(ws_ctrl["B3"])

    ws_ctrl["A4"] = "Project Start Date:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = start_date
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)

    ws_ctrl["D4"] = "Durations (days)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)

    durations = [
        ("D5",  "Mobilization (SPMT+Marine):",        "E5",  1.0, "DUR_MOB"),
        ("D6",  "Pre-work: Beam replacement:",        "E6",  2.0, "DUR_BEAM"),
        ("D7",  "Pre-work: Deck preparations:",       "E7",  2.0, "DUR_DECK"),
        ("D8",  "Port: Load-out (per TR):",           "E8",  0.5, "DUR_PORT_LO"),
        ("D9",  "Port: Seafastening (per TR):",       "E9",  1.0, "DUR_PORT_SF"),
        ("D10", "MWS+MPI+Final (per voyage):",        "E10", 0.2, "DUR_MWS"),
        ("D11", "Sailing (one-way):",                 "E11", 1.0, "DUR_SAIL"),
        ("D12", "Arrival/Berthing:",                  "E12", 0.5, "DUR_ARR"),
        ("D13", "Unlashing/Cutting prep:",            "E13", 0.5, "DUR_UNLASH"),
        ("D14", "AGI Load-in/Unload (per TR):",       "E14", 1.0, "DUR_UL"),
        ("D15", "Steel bridge install (per TR):",     "E15", 0.3, "DUR_BRIDGE"),
        ("D16", "Transport to bay + jack-up:",        "E16", 0.7, "DUR_TRN"),
        ("D17", "Turning (per TR):",                  "E17", 3.0, "DUR_TURN"),
        ("D18", "Jack-down (per TR):",                "E18", 1.0, "DUR_JD"),
        ("D19", "Return prep (SPMT back-load):",      "E19", 1.0, "DUR_RET_PREP"),
        ("D20", "Buffer between voyages:",            "E20", 0.5, "DUR_BUF"),
    ]

    for lc, label, vc, val, name in durations:
        ws_ctrl[lc] = label
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = val
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")

    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")

    ws_ctrl.column_dimensions["A"].width = 26
    ws_ctrl.column_dimensions["B"].width = 24
    ws_ctrl.column_dimensions["D"].width = 30
    ws_ctrl.column_dimensions["E"].width = 10

    # Schedule_Data
    ws = wb.create_sheet("Schedule_Data")
    ws.merge_cells("A1:I1")
    ws["A1"] = "AGI Transformer Transportation - Schedule Data"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Start = {start_date.isoformat()} | Single Train (1 LCT + 1 SPMT) | Pattern: {'+'.join(map(str,batch_pattern))}"
    ws["A2"].font = Font(size=10, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()

    tasks: list[tuple] = []
    tasks.append(("MOB-001","1.0","Mobilization (SPMT + Marine Equipment)","MOBILIZATION","Mammoet","=DUR_MOB","SPMT assembly + marine equipment mobilization"))
    tasks.append(("PRE-001","1.1","Pre-work: Beam replacement (site bays)","PREWORK","Mammoet","=DUR_BEAM","Per Mammoet baseline (two bays)"))
    tasks.append(("PRE-002","1.2","Pre-work: Deck preparations (D-rings, steel sets, welding)","PREWORK","Mammoet","=DUR_DECK","One-time preparation"))

    tr_id = 1
    voyage_id = 1
    for batch_i, n_units in enumerate(batch_pattern, start=1):
        tr_list = list(range(tr_id, tr_id+n_units))
        tr_id += n_units

        tasks.append((f"V{voyage_id}","2.%d"%batch_i,f"VOYAGE {voyage_id}: TR{tr_list[0]}"+(f"-TR{tr_list[-1]}" if len(tr_list)>1 else ""),"MILESTONE","All",0,f"Batch size {n_units} | TRs: {', '.join('TR'+str(t) for t in tr_list)}"))

        for t in tr_list:
            tasks.append((f"LO-{voyage_id:02d}-{t}","3.%d"%t,f"Port Load-out (RoRo) - TR{t}","LOADOUT","Mammoet","=DUR_PORT_LO","Ramp/linkspan + roll-on; rising tide preferred"))
            tasks.append((f"SF-{voyage_id:02d}-{t}","3.%d"%t,f"Sea Fastening - TR{t}","SEAFAST","Mammoet","=DUR_PORT_SF","Sea fastening per calc"))

        tasks.append((f"MWS-{voyage_id:02d}","3.%d9"%batch_i,"MWS + MPI + Final Preparations","BUFFER","Aries/Captain","=DUR_MWS","MWS checklist + sail-away certificate"))
        tasks.append((f"SAIL-{voyage_id:02d}","3.%dA"%batch_i,"Sail-away: Mina Zayed → AGI","SAIL","LCT Bushra","=DUR_SAIL","Weather gate per MS (wind ≤20kt, Hs ≤0.6m)"))
        tasks.append((f"ARR-{voyage_id:02d}","3.%dB"%batch_i,"Arrival + Berthing at AGI RoRo Jetty","AGI_UNLOAD","LCT Bushra","=DUR_ARR","Falling tide preferred for load-in"))
        tasks.append((f"PREP-UL-{voyage_id:02d}","3.%dC"%batch_i,"Unlashing/Cutting + Steel-sets prep","AGI_UNLOAD","Mammoet","=DUR_UNLASH","Unlashing, cleat cutting, prep"))

        for t in reversed(tr_list):
            tasks.append((f"UL-{voyage_id:02d}-{t}","4.%d"%t,f"AGI Load-in (Roll-out) + Jetty Storage - TR{t}","AGI_UNLOAD","Mammoet","=DUR_UL","One unit/day baseline"))

        for t in tr_list:
            tasks.append((f"BR-{voyage_id:02d}-{t}","5.%d0"%t,f"Steel Bridge Installation - TR{t}","TURNING","Mammoet","=DUR_BRIDGE","Crane required"))
            tasks.append((f"TRN-{voyage_id:02d}-{t}","5.%d1"%t,f"Transport to Bay + Jack-up (temporary support) - TR{t}","TURNING","Mammoet","=DUR_TRN","Includes SPMT move + set on temp support"))
            tasks.append((f"TURN-{voyage_id:02d}-{t}","5.%d2"%t,f"Turning (90° rotation) - TR{t}","TURNING","Mammoet","=DUR_TURN","10t forklift required"))
            tasks.append((f"JD-{voyage_id:02d}-{t}","5.%d3"%t,f"Jack-down on temporary support - TR{t}","JACKDOWN","Mammoet","=DUR_JD","1 day per unit baseline"))

        tasks.append((f"RET-PREP-{voyage_id:02d}","6.%d"%batch_i,"SPMT shifting back to Jetty + Load on LCT (if tide allows)","RETURN","Mammoet","=DUR_RET_PREP","Tide gate"))
        tasks.append((f"RET-{voyage_id:02d}","6.%dA"%batch_i,"Return Sail: AGI → Mina Zayed","RETURN","LCT Bushra","=DUR_SAIL","Backhaul"))

        if batch_i != len(batch_pattern):
            tasks.append((f"BUF-{voyage_id:02d}","6.%dB"%batch_i,"Buffer / Reset / Permits","BUFFER","All","=DUR_BUF","Contingency + PTW synchronization"))

        voyage_id += 1

    tasks.append(("END","99.0","PROJECT COMPLETE","MILESTONE","All",0,"All batches completed"))

    start_row = 6
    for idx, t in enumerate(tasks):
        r = start_row + idx
        tid, wbs, task, phase, owner, dur, notes = t

        ws.cell(r,1,value=tid)
        ws.cell(r,2,value=wbs)
        ws.cell(r,3,value=task)
        ws.cell(r,4,value=phase)
        ws.cell(r,5,value=owner)

        if r == start_row:
            ws.cell(r,6,value="=PROJECT_START")
        else:
            ws.cell(r,6,value=f"=G{r-1}")
        ws.cell(r,6).number_format = "YYYY-MM-DD"

        ws.cell(r,8,value=dur if dur != 0 else 0)
        ws.cell(r,8).number_format = "0.0"

        ws.cell(r,7,value=f"=F{r}+H{r}")
        ws.cell(r,7).number_format = "YYYY-MM-DD"
        ws.cell(r,9,value=notes)

        ws.cell(r,4).fill = PatternFill("solid", fgColor=COLORS.get(phase,"FFFFFF"))
        for c in range(1,10):
            ws.cell(r,c).border = tb()
            ws.cell(r,c).alignment = Alignment(vertical="center", wrap_text=(c in [3,9]))

        if phase in ("MILESTONE","JACKDOWN"):
            for c in range(1,10):
                ws.cell(r,c).font = Font(bold=True, color=("B71C1C" if phase=="JACKDOWN" else "000000"))

    col_widths = {"A":12,"B":8,"C":52,"D":14,"E":14,"F":12,"G":12,"H":10,"I":44}
    for col,w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    # Gantt_Chart (static fill based on default durations)
    wg = wb.create_sheet("Gantt_Chart")

    # Compute default total days (evaluate formulas with defaults only)
    defaults = {name:val for _,_,_,val,name in durations}

    def dur_value(d) -> float:
        if isinstance(d,(int,float)):
            return float(d)
        if isinstance(d,str) and d.startswith("="):
            expr = d[1:]
            for k,v in defaults.items():
                expr = re.sub(rf"\\b{k}\\b", str(v), expr)
            return float(eval(expr))
        return float(d)

    total_days = int(math.ceil(sum(dur_value(t[5]) for t in tasks) + 2))

    wg.merge_cells("A1:AZ1")
    wg["A1"] = f"AGI TR Transportation - Gantt (Start {start_date.isoformat()})"
    wg["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    wg["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    wg["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wg.row_dimensions[1].height = 24

    meta_headers = ["ID","WBS","Task","Phase","Start","End","Dur"]
    for c,h in enumerate(meta_headers,1):
        cell = wg.cell(4,c,value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = tb()

    date_col0 = 8
    for i in range(total_days):
        cell = wg.cell(4, date_col0+i, value=start_date + dt.timedelta(days=i))
        cell.number_format = "d"
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
        wg.column_dimensions[get_column_letter(date_col0+i)].width = 2.4

    # Copy meta references
    for idx in range(len(tasks)):
        r = 5 + idx
        sched_r = start_row + idx
        wg.cell(r,1,value=f"=Schedule_Data!A{sched_r}")
        wg.cell(r,2,value=f"=Schedule_Data!B{sched_r}")
        wg.cell(r,3,value=f"=Schedule_Data!C{sched_r}")
        wg.cell(r,4,value=f"=Schedule_Data!D{sched_r}")
        sc = wg.cell(r,5,value=f"=Schedule_Data!F{sched_r}")
        sc.number_format="MM/DD"
        ec = wg.cell(r,6,value=f"=Schedule_Data!G{sched_r}")
        ec.number_format="MM/DD"
        wg.cell(r,7,value=f"=Schedule_Data!H{sched_r}")
        for c in range(1,8):
            wg.cell(r,c).border = tb()
            wg.cell(r,c).alignment = Alignment(vertical="center", wrap_text=(c==3))

    # Static bars using defaults
    cur = dt.datetime.combine(start_date, dt.time())
    computed = []
    for t in tasks:
        d = dur_value(t[5])
        s = cur
        e = cur + dt.timedelta(days=d)
        computed.append((s,e,t[3]))
        cur = e

    for idx,(s,e,phase) in enumerate(computed):
        r = 5 + idx
        wg.cell(r,4).fill = PatternFill("solid", fgColor=COLORS.get(phase,"FFFFFF"))
        for i in range(total_days):
            day_start = dt.datetime.combine(start_date + dt.timedelta(days=i), dt.time())
            if day_start >= s and day_start < e:
                wg.cell(r, date_col0+i).fill = PatternFill("solid", fgColor=COLORS.get(phase,"FFFFFF"))
            wg.cell(r, date_col0+i).border = tb()

    wg.column_dimensions["A"].width = 12
    wg.column_dimensions["B"].width = 7
    wg.column_dimensions["C"].width = 34
    wg.column_dimensions["D"].width = 12
    wg.column_dimensions["E"].width = 7
    wg.column_dimensions["F"].width = 7
    wg.column_dimensions["G"].width = 4
    wg.freeze_panes = wg.cell(5, date_col0)

    return wb

def parse_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--scenario", choices=["single","batch"], required=True)
    ap.add_argument("--start", type=parse_date, required=True)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if args.scenario == "single":
        pattern = [1,1,1,1,1,1,1]
        title = "Single-Train 1+1+1+1+1+1+1"
        default_out = f"AGI_TR_SingleTrain_1x7_Start{args.start:%Y%m%d}.xlsx"
    else:
        pattern = [1,2,2,2]
        title = "Batch 1+2+2+2"
        default_out = f"AGI_TR_Batch_1_2_2_2_Start{args.start:%Y%m%d}.xlsx"

    out = args.out or default_out
    wb = build_workbook(args.start, pattern, title)
    wb.save(out)
    print(f"[OK] Saved: {out}")

if __name__ == "__main__":
    main()
