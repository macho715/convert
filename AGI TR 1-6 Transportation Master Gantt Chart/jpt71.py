# pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import date, timedelta

# =========================
# 0) 입력 데이터 (필요시 여기만 수정)
# =========================
# 컬럼 의미:
# (Item, Type, Material, MW4_Agg_Load, AGI_Agg_Off1, AGI_Agg_Off2, AGI_Deb_Load, MW4_Deb_Off, Status)
#
# - Aggregate: MW4_Agg_Load / AGI_Agg_Off1 / (필요시) AGI_Agg_Off2 사용
# - Debris(Backload): AGI_Deb_Load / MW4_Deb_Off 사용
#
# 아래는 사용자 화이트보드/지시 기반 예시 세팅 (Dates are reference only)

rows_plan = [
    # --- 실적 예시(84 cycle) ---
    ("84",       "Aggregate", "5mm",       date(2025,12,5),  date(2025,12,6),  date(2025,12,7),  None,              None,              ""),
    ("Debris-5", "Debris",    "Debris",    None,             None,             None,             date(2025,12,7),   date(2025,12,8),   ""),

    # --- (선택) 다음 cycle 예시(85) ---
    ("85",       "Aggregate", "Dune Sand", date(2025,12,9),  date(2025,12,10), date(2025,12,11), None,              None,              ""),
    ("Debris-6", "Debris",    "Debris",    None,             None,             None,             date(2025,12,11),  date(2025,12,12),  ""),

    # --- Debris list (whiteboard) ---
    ("Debris-7", "Debris",    "Debris",    None, None, None, date(2025,12,25), date(2025,12,26), ""),
    ("Debris-8", "Debris",    "Debris",    None, None, None, date(2025,12,28), date(2025,12,29), "IN PROGRESS"),  # 현재 하역중(MW4)

    # --- Aggregate list / schedule board (예시) ---
    ("88",       "Aggregate", "Dune Sand", date(2025,12,30), None, None, None, None, ""),
    ("Debris-9", "Debris",    "Debris",    None, None, None, date(2026,1,7),  date(2026,1,8),  ""),

    ("89",       "Aggregate", "5mm",       date(2026,1,9),  date(2026,1,10), date(2026,1,11), None, None, ""),
    ("Debris-10","Debris",    "Debris",    None, None, None, date(2026,1,10), date(2026,1,11), ""),

    ("90",       "Aggregate", "20mm",      date(2026,1,12), date(2026,1,13), None, None, None, ""),
    ("Debris-11","Debris",    "Debris",    None, None, None, date(2026,1,13), date(2026,1,14), ""),

    ("91",       "Aggregate", "10mm",      date(2026,1,15), None, None, None, None, ""),
    ("Debris-12","Debris",    "Debris",    None, None, None, date(2026,1,19), date(2026,1,20), ""),

    ("92",       "Aggregate", "20mm",      date(2026,1,21), date(2026,1,24), None, None, None, ""),
    ("Debris-13","Debris",    "Debris",    None, None, None, date(2026,1,25), date(2026,1,26), ""),

    ("93",       "Aggregate", "5mm",       date(2026,1,27), None, None, None, None, ""),
]

# Debris-8 계획 MW4 하역일(Delay 기준일)
DEBRIS8_PLANNED_MW4_DEB_OFF = date(2025,12,29)

# =========================
# 1) 정렬/보조값
# =========================
def earliest_dt(r):
    ds = [x for x in r[3:8] if x is not None]
    return min(ds) if ds else date(2099,1,1)

rows_plan = sorted(rows_plan, key=earliest_dt)
deb8_seq = next((i+1 for i,r in enumerate(rows_plan) if r[0]=="Debris-8"), None)

all_dates=[]
for r in rows_plan:
    all_dates += [x for x in r[3:8] if x is not None]
start = min(all_dates) - timedelta(days=1)
end   = max(all_dates) + timedelta(days=14)

dates=[]
d=start
while d<=end:
    dates.append(d); d+=timedelta(days=1)

# =========================
# 2) 워크북 생성
# =========================
wb = Workbook()

# --- Control sheet ---
ctl = wb.active
ctl.title="Control"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ctl["A1"]="Debris-8(MW4 Debris Offloading) IN PROGRESS. B3에 실제 완료시각 입력 → 이후 일정 자동 Shift."
ctl["A1"].font=Font(bold=True); ctl["A1"].alignment=left
ctl.merge_cells("A1:D1")

ctl["A3"]="Debris-8 Actual Completion (date/time)"
ctl["B3"]=""  # 사용자 입력
ctl["A4"]="Debris-8 Planned MW4 Debris Offloading Date"
ctl["B4"]=DEBRIS8_PLANNED_MW4_DEB_OFF
ctl["A5"]="Delay Days (auto)"
ctl["B5"]='=IF(B3="",0,MAX(0,INT(B3)-B4))'

for r in range(3,6):
    for c in range(1,3):
        cell=ctl.cell(r,c); cell.border=border
        cell.alignment = left if c==1 else center

ctl.column_dimensions["A"].width=52
ctl.column_dimensions["B"].width=26

# --- Cross_Gantt sheet ---
ws = wb.create_sheet("Cross_Gantt")

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
fill_white  = PatternFill("solid", fgColor="FFFFFF")

fill_row_agg = PatternFill("solid", fgColor="E2EFDA")
fill_row_deb = PatternFill("solid", fgColor="FCE4D6")
fill_inprog  = PatternFill("solid", fgColor="D9D9D9")
fill_completed = PatternFill("solid", fgColor="C5E0B4")

fill_mw4_load     = PatternFill("solid", fgColor="4F81BD")  # MW4 Agg Loading
fill_agi_off1     = PatternFill("solid", fgColor="70AD47")  # AGI Offload Day1
fill_agi_off2     = PatternFill("solid", fgColor="92D050")  # AGI Offload+DebLoad Day2
fill_agi_deb_load = PatternFill("solid", fgColor="FFC000")  # AGI Debris Loading
fill_mw4_deb_off  = PatternFill("solid", fgColor="FF6600")  # MW4 Debris Offloading

fill_today = PatternFill("solid", fgColor="FFFF00")
today_border = Border(
    left=Side(style="thick", color="FF0000"),
    right=Side(style="thick", color="FF0000"),
    top=Side(style="thick", color="FF0000"),
    bottom=Side(style="thick", color="FF0000"),
)

dash = Side(style="dashed", color="7F7F7F")
border_dashed = Border(left=dash, right=dash, top=dash, bottom=dash)

visible_headers=[
    "Seq","Item","Type","Material",
    "MW4 Agg Loading/Depart",
    "AGI Agg Offload Day-1",
    "AGI Agg Offload+Deb Load (Mix) Day-2",
    "AGI Debris Loading (Deb)",
    "MW4 Debris Offloading (Deb)",
    "Status"
]
hidden_headers=["ShiftFlag","Plan_MW4_Agg","Plan_AGI_Off1","Plan_AGI_Off2","Plan_AGI_DebLoad","Plan_MW4_DebOff"]

col=1
for h in visible_headers:
    cell=ws.cell(1,col,h); cell.fill=header_fill; cell.font=header_font
    cell.alignment=center; cell.border=border
    col+=1
for h in hidden_headers:
    cell=ws.cell(1,col,h); cell.fill=header_fill; cell.font=header_font
    cell.alignment=center; cell.border=border
    col+=1

date_start_col=col
today = date.today()
today_col = None
for i,dt in enumerate(dates):
    c=date_start_col+i
    cell=ws.cell(1,c,dt); cell.number_format="mm-dd"
    cell.fill=header_fill; cell.font=header_font
    cell.alignment=center; cell.border=border
    ws.column_dimensions[get_column_letter(c)].width=5
    if dt == today:
        today_col = c
        cell.fill = fill_today
        cell.font = Font(color="000000", bold=True)
        cell.border = today_border

widths={1:5,2:12,3:10,4:16,5:18,6:18,7:26,8:18,9:20,10:12}
for c,w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width=w

shift_col=len(visible_headers)+1
for c in range(shift_col, date_start_col):
    ws.column_dimensions[get_column_letter(c)].hidden=True

ws.freeze_panes="K2"
ws.row_dimensions[1].height=28

delay_ref="Control!$B$5"

# 데이터 입력 + Debris-8 이후 Shift 적용
for idx,r in enumerate(rows_plan, start=1):
    item, typ, material, p_mw4, p_off1, p_off2, p_deb_load, p_deb_off, status = r
    row_num=1+idx
    row_fill = fill_row_deb if typ=="Debris" else fill_row_agg
    status_upper = str(status).upper() if status else ""
    if "COMPLETE" in status_upper:
        row_fill = fill_completed
    elif "IN PROGRESS" in status_upper or item=="Debris-8":
        row_fill = fill_inprog

    ws.cell(row_num,1,idx).alignment=center
    ws.cell(row_num,2,item).alignment=center
    ws.cell(row_num,3,typ).alignment=center
    ws.cell(row_num,4,material).alignment=left
    ws.cell(row_num,10,status).alignment=center

    shift_flag = 1 if (deb8_seq is not None and idx>deb8_seq) else 0
    ws.cell(row_num, shift_col, shift_flag).alignment=center

    plan_start_col=shift_col+1
    plan_vals=[p_mw4,p_off1,p_off2,p_deb_load,p_deb_off]
    for j,val in enumerate(plan_vals):
        cell=ws.cell(row_num, plan_start_col+j, val)
        if val: cell.number_format="yyyy-mm-dd"
        cell.alignment=center

    flag_ref=f"${get_column_letter(shift_col)}{row_num}"
    def adj(addr):
        return f'=IF({addr}="", "", IF({flag_ref}=1, {addr}+{delay_ref}, {addr}))'

    plan_refs=[
        f"{get_column_letter(plan_start_col)}{row_num}",
        f"{get_column_letter(plan_start_col+1)}{row_num}",
        f"{get_column_letter(plan_start_col+2)}{row_num}",
        f"{get_column_letter(plan_start_col+3)}{row_num}",
        f"{get_column_letter(plan_start_col+4)}{row_num}",
    ]
    for vis_col, plan_ref in zip([5,6,7,8,9], plan_refs):
        ws.cell(row_num, vis_col).value=adj(plan_ref)
        ws.cell(row_num, vis_col).number_format="yyyy-mm-dd"
        ws.cell(row_num, vis_col).alignment=center

    for c in range(1,11):
        cell=ws.cell(row_num,c); cell.border=border; cell.fill=row_fill
    if item=="Debris-8":
        for c in range(1,11):
            ws.cell(row_num,c).border=border_dashed

# 타임라인 기본 셋
first_data_row=2
last_data_row=1+len(rows_plan)
first_date_col=date_start_col
last_date_col=date_start_col+len(dates)-1
for r in range(first_data_row, last_data_row+1):
    for c in range(first_date_col, last_date_col+1):
        ws.cell(r,c).border=border
        ws.cell(r,c).fill=fill_white

deb8_row=(1+deb8_seq) if deb8_seq else None
top_left_col_letter=get_column_letter(first_date_col)

def apply_cf(r1,r2, col_letter, fill):
    if r1>r2: return
    rng=f"{get_column_letter(first_date_col)}{r1}:{get_column_letter(last_date_col)}{r2}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={top_left_col_letter}$1=${col_letter}{r1}'], fill=fill))

ranges=[]
if deb8_row:
    ranges.append((first_data_row, deb8_row-1))
    ranges.append((deb8_row+1, last_data_row))
else:
    ranges.append((first_data_row, last_data_row))

for r1,r2 in ranges:
    apply_cf(r1,r2,"E",fill_mw4_load)
    apply_cf(r1,r2,"F",fill_agi_off1)
    apply_cf(r1,r2,"G",fill_agi_off2)
    apply_cf(r1,r2,"H",fill_agi_deb_load)
    apply_cf(r1,r2,"I",fill_mw4_deb_off)

# Debris-8 타임라인: 회색/점선(조건부서식 제외)
if deb8_row:
    for c in range(first_date_col, last_date_col+1):
        cell=ws.cell(deb8_row,c)
        cell.fill=fill_inprog
        cell.border=border_dashed

# 오늘 날짜 세로선 강조
if today_col:
    for r in range(first_data_row, last_data_row+1):
        cell = ws.cell(r, today_col)
        cell.fill = fill_today
        cell.border = today_border

# Legend sheet
legend_ws = wb.create_sheet("Legend")
legend_ws.title = "Legend"

legend_data = [
    ("항목", "설명", "색상"),
    ("MW4 Agg Loading", "MW4 집계 적재/출발", "파란색 (4F81BD)"),
    ("AGI Agg Offload Day-1", "AGI 집계 하역 1일차", "녹색 (70AD47)"),
    ("AGI Agg Offload+Deb Load Day-2", "AGI 집계 하역+잔재 적재 혼합 2일차", "연한 녹색 (92D050)"),
    ("AGI Debris Loading", "AGI 잔재 적재", "주황색 (FFC000)"),
    ("MW4 Debris Offloading", "MW4 잔재 하역", "진한 주황색 (FF6600)"),
    ("", "", ""),
    ("행 배경색", "", ""),
    ("Aggregate 작업", "집계 작업 행", "연한 녹색 (E2EFDA)"),
    ("Debris 작업", "잔재 작업 행", "연한 주황색 (FCE4D6)"),
    ("IN PROGRESS", "진행 중 작업", "회색 (D9D9D9)"),
    ("COMPLETED", "완료된 작업", "완료 녹색 (C5E0B4)"),
    ("", "", ""),
    ("특수 표시", "", ""),
    ("오늘 날짜", "현재 날짜 강조", "노란색 배경 + 빨간 테두리"),
    ("Debris-8", "진행 중 작업 (점선 테두리)", "회색 배경 + 점선 테두리"),
]

for r_idx, (col1, col2, col3) in enumerate(legend_data, start=1):
    legend_ws.cell(r_idx, 1, col1).alignment = left
    legend_ws.cell(r_idx, 2, col2).alignment = left
    legend_ws.cell(r_idx, 3, col3).alignment = left
    if r_idx == 1:
        for c in range(1, 4):
            cell = legend_ws.cell(r_idx, c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center
    else:
        for c in range(1, 4):
            legend_ws.cell(r_idx, c).border = border

legend_ws.column_dimensions["A"].width = 30
legend_ws.column_dimensions["B"].width = 40
legend_ws.column_dimensions["C"].width = 30

out_path="JPT71_CrossGantt_Debris8_InProgress_AutoShift.xlsx"
wb.save(out_path)
print("Saved:", out_path)
