from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .markdown_to_docx import extract_json_ld_from_markdown
from .types import ConversionResult, ConversionError, EngineNotFoundError


def _extract_message_bodies(content: str) -> Dict[str, str]:
    """Extract message bodies from markdown content"""
    message_bodies = {}
    lines = content.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Check for message header: #### Msg X — Name @ Date {#msg-X}
        msg_match = re.match(r'^####\s+Msg\s+\d+\s+—\s+.+?\s+@\s+.+?\s+\{#(.+?)\}', line)
        if msg_match:
            msg_id = msg_match.group(1)  # e.g., "msg-13"
            
            # Find the code block with message body
            i += 1
            while i < len(lines) and not lines[i].strip().startswith('```'):
                i += 1
            
            # Extract text from code block
            if i < len(lines) and lines[i].strip().startswith('```'):
                i += 1  # Skip ``` marker
                body_lines = []
                while i < len(lines) and not lines[i].strip().startswith('```'):
                    body_lines.append(lines[i])
                    i += 1
                message_bodies[msg_id] = '\n'.join(body_lines).strip()
        
        i += 1
    
    return message_bodies


def _extract_timeline(content: str) -> list[Dict[str, str]]:
    """Extract timeline entries from markdown content"""
    timeline_entries = []
    lines = content.split('\n')
    i = 0
    in_timeline = False
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Check for Timeline section start
        if re.match(r'^###\s+Timeline', line, re.IGNORECASE):
            in_timeline = True
            i += 1
            continue
        
        # Check for section end (### or ---)
        if in_timeline and (line.startswith('###') or line.strip() == '---'):
            break
        
        # Parse timeline entries: - YYYY-MM-DD: Description <!-- data:ref=#msg-X -->
        if in_timeline and line.startswith('- '):
            # Extract date and description
            # Format: - 2025-09-19: Description <!-- data:ref=#msg-13 -->
            match = re.match(r'^-\s+(\d{4}-\d{2}-\d{2}):\s+(.+?)(?:\s*<!--\s*data:ref=#(.+?)\s*-->)?$', line)
            if match:
                date_str = match.group(1)
                description = match.group(2).strip()
                related_msg = match.group(3) if match.group(3) else ''
                
                timeline_entries.append({
                    'date': date_str,
                    'description': description,
                    'related_message': related_msg
                })
        
        i += 1
    
    return timeline_entries


def markdown_to_xlsx(
    src: Path,
    dst: Path,
    preserve_metadata: bool = True
) -> ConversionResult:
    """
    Convert markdown email file to Excel (.xlsx) format
    
    Args:
        src: Source markdown file path
        dst: Destination XLSX file path
        preserve_metadata: Whether to include JSON-LD metadata
        
    Returns:
        ConversionResult with conversion details
    """
    if not OPENPYXL_AVAILABLE:
        raise EngineNotFoundError(
            "openpyxl is required for XLSX conversion. "
            "Install it with: pip install openpyxl"
        )
    
    if not src.exists():
        raise FileNotFoundError(f"Source file not found: {src}")
    
    # Read markdown content
    content = src.read_text(encoding='utf-8')
    
    # Extract JSON-LD metadata
    metadata = None
    if preserve_metadata:
        metadata = extract_json_ld_from_markdown(content)
    
    if not metadata:
        raise ConversionError("No JSON-LD metadata found in markdown file")
    
    # Extract message bodies from markdown
    message_bodies = _extract_message_bodies(content)
    
    # Extract timeline entries from markdown
    timeline_entries = _extract_timeline(content)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary", 0)
    _add_summary_sheet(ws_summary, metadata, header_fill, header_font, border)
    
    # Sheet 2: Participants
    if 'participants' in metadata and metadata['participants']:
        ws_participants = wb.create_sheet("Participants")
        _add_participants_sheet(ws_participants, metadata['participants'], header_fill, header_font, border)
    
    # Sheet 3: Messages
    if 'messages' in metadata and metadata['messages']:
        ws_messages = wb.create_sheet("Messages")
        _add_messages_sheet(ws_messages, metadata['messages'], message_bodies, header_fill, header_font, border)
    
    # Sheet 4: Actions
    if 'actions' in metadata and metadata['actions']:
        ws_actions = wb.create_sheet("Actions")
        _add_actions_sheet(ws_actions, metadata['actions'], header_fill, header_font, border)
    
    # Sheet 5: Issues
    if 'issues' in metadata and metadata['issues']:
        ws_issues = wb.create_sheet("Issues")
        _add_issues_sheet(ws_issues, metadata['issues'], header_fill, header_font, border)
    
    # Sheet 6: Topics
    if 'topics' in metadata and metadata['topics']:
        ws_topics = wb.create_sheet("Topics")
        _add_topics_sheet(ws_topics, metadata['topics'], header_fill, header_font, border)
    
    # Sheet 7: Timeline
    if timeline_entries:
        ws_timeline = wb.create_sheet("Timeline")
        _add_timeline_sheet(ws_timeline, timeline_entries, header_fill, header_font, border)
    
    # Save workbook
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))
    
    if not dst.exists():
        raise ConversionError(f"Failed to create XLSX file: {dst}")
    
    return ConversionResult(src, dst, "markdown-to-xlsx")


def _add_summary_sheet(ws, metadata: Dict[str, Any], header_fill, header_font, border):
    """Add summary information sheet"""
    row = 1
    
    # Title
    ws['A1'] = "Email Thread Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    row += 2
    
    # Subject
    ws[f'A{row}'] = "Subject"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'] = metadata.get('subject', 'N/A')
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1
    
    # Date Range
    if 'dateRange' in metadata:
        ws[f'A{row}'] = "Date Range"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        date_range = metadata['dateRange']
        date_str = f"{date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"
        ws[f'B{row}'] = date_str
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Message Count
    if 'messages' in metadata:
        ws[f'A{row}'] = "Total Messages"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = len(metadata['messages'])
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Participant Count
    if 'participants' in metadata:
        ws[f'A{row}'] = "Total Participants"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = len(metadata['participants'])
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60


def _add_participants_sheet(ws, participants: list, header_fill, header_font, border):
    """Add participants sheet"""
    # Headers
    headers = ["Name", "Organization", "Email", "Role"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, participant in enumerate(participants, 2):
        ws.cell(row=row_idx, column=1, value=participant.get('name', 'N/A')).border = border
        ws.cell(row=row_idx, column=2, value=participant.get('org', 'N/A')).border = border
        ws.cell(row=row_idx, column=3, value=participant.get('email', 'N/A')).border = border
        ws.cell(row=row_idx, column=4, value=participant.get('role', 'N/A')).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20


def _add_messages_sheet(ws, messages: list, message_bodies: Dict[str, str], header_fill, header_font, border):
    """Add messages sheet with message bodies"""
    # Headers - 추가: Message Body 열
    headers = ["Order", "Message ID", "Date", "From", "To", "Subject", "Summary", "Message Body"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort messages by order if available, handling mixed types
    def get_sort_key(msg):
        order = msg.get('order')
        if order is not None:
            return (0, int(order) if isinstance(order, (int, str)) else 0)
        date = msg.get('isoDate', '')
        return (1, date)
    
    sorted_messages = sorted(messages, key=get_sort_key)
    
    # Data
    for row_idx, msg in enumerate(sorted_messages, 2):
        ws.cell(row=row_idx, column=1, value=msg.get('order', 'N/A')).border = border
        ws.cell(row=row_idx, column=2, value=msg.get('id', 'N/A')).border = border
        ws.cell(row=row_idx, column=3, value=msg.get('isoDate', 'N/A')).border = border
        
        # From field
        from_field = msg.get('from', 'N/A')
        if isinstance(from_field, str) and '<' in from_field:
            from_field = from_field.split('<')[0].strip()
        ws.cell(row=row_idx, column=4, value=from_field).border = border
        
        # To field
        to_field = msg.get('to', [])
        if isinstance(to_field, list):
            to_field = ', '.join(str(t) for t in to_field[:3])  # Limit to first 3
        ws.cell(row=row_idx, column=5, value=str(to_field)[:50]).border = border  # Limit length
        
        ws.cell(row=row_idx, column=6, value=msg.get('subject', 'N/A')).border = border
        ws.cell(row=row_idx, column=7, value=msg.get('summary', 'N/A')[:100]).border = border  # Limit length
        
        # Message Body - 마지막 열에 원문 추가
        msg_id = msg.get('id', '')
        # msg_id는 "msg-13" 형식, message_bodies의 키도 "msg-13" 형식
        # JSON-LD의 snippetRef는 "#msg-13" 형식이지만, 실제 추출된 키는 "msg-13"
        body_text = message_bodies.get(msg_id, 'N/A')
        
        # 메시지 본문이 없으면 snippetRef에서 시도
        if body_text == 'N/A' and 'snippetRef' in msg:
            snippet_ref = msg.get('snippetRef', '').lstrip('#')
            body_text = message_bodies.get(snippet_ref, 'N/A')
        
        body_cell = ws.cell(row=row_idx, column=8, value=body_text)
        body_cell.border = border
        body_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 80  # Message Body 열 너비 증가


def _add_actions_sheet(ws, actions: list, header_fill, header_font, border):
    """Add actions sheet"""
    # Headers
    headers = ["Owner", "Action", "Status", "Related Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, action in enumerate(actions, 2):
        ws.cell(row=row_idx, column=1, value=action.get('owner', 'N/A')).border = border
        ws.cell(row=row_idx, column=2, value=action.get('action', 'N/A')).border = border
        ws.cell(row=row_idx, column=3, value=action.get('status', 'Pending')).border = border
        ws.cell(row=row_idx, column=4, value=action.get('relatedMsg', 'N/A')).border = border
        
        # Color code status
        status_cell = ws.cell(row=row_idx, column=3)
        status = action.get('status', 'Pending').lower()
        if 'completed' in status or 'approved' in status:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif 'pending' in status:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20


def _add_issues_sheet(ws, issues: list, header_fill, header_font, border):
    """Add issues sheet"""
    # Headers
    headers = ["Type", "Description", "Related Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, issue in enumerate(issues, 2):
        ws.cell(row=row_idx, column=1, value=issue.get('type', 'N/A')).border = border
        ws.cell(row=row_idx, column=2, value=issue.get('description', 'N/A')).border = border
        ws.cell(row=row_idx, column=3, value=issue.get('relatedMsg', 'N/A')).border = border
        
        # Color code issue types
        type_cell = ws.cell(row=row_idx, column=1)
        issue_type = issue.get('type', '').lower()
        if 'critical' in issue_type or 'security' in issue_type:
            type_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif 'warning' in issue_type:
            type_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20


def _add_topics_sheet(ws, topics: list, header_fill, header_font, border):
    """Add topics sheet"""
    # Headers
    ws.cell(row=1, column=1, value="Topic").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = border
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row_idx, topic in enumerate(topics, 2):
        ws.cell(row=row_idx, column=1, value=topic).border = border
    
    # Adjust column width
    ws.column_dimensions['A'].width = 40


def _add_timeline_sheet(ws, timeline_entries: list, header_fill, header_font, border):
    """Add timeline sheet"""
    # Headers
    headers = ["Date", "Description", "Related Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort timeline entries by date
    sorted_entries = sorted(timeline_entries, key=lambda x: x.get('date', ''))
    
    # Data
    for row_idx, entry in enumerate(sorted_entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.get('date', 'N/A')).border = border
        desc_cell = ws.cell(row=row_idx, column=2, value=entry.get('description', 'N/A'))
        desc_cell.border = border
        desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=row_idx, column=3, value=entry.get('related_message', 'N/A')).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 100  # Description
    ws.column_dimensions['C'].width = 20  # Related Message

