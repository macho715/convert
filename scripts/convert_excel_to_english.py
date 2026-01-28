import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 한글-영문 번역 매핑
TRANSLATIONS = {
    # 시트명
    '1. 이고르 우려사항': '1. Igor Concerns',
    '2. Linkspan 체류 시간 분석': '2. Linkspan Dwell Time Analysis',
    '3. 조수 데이터 분석': '3. Tide Data Analysis',
    '4. Linkspan 체류 시간 시뮬레이션': '4. Linkspan Dwell Time Simulation',
    '5. 이고르 우려사항 대응 분석': '5. Igor Concerns Response Analysis',
    '6. 이고르 질문 답변': '6. Igor Questions & Answers',
    '7. 종합 리스크 매트릭스': '7. Comprehensive Risk Matrix',
    '8. 권장 조치': '8. Recommended Actions',
    '9. 결론': '9. Conclusion',
    
    # 일반 용어
    '이고르 우려사항 원문': 'Igor Concerns - Original Text',
    'Linkspan 체류 시간 분석': 'Linkspan Dwell Time Analysis',
    '조수 데이터 분석': 'Tide Data Analysis',
    'Linkspan 체류 시간 시뮬레이션': 'Linkspan Dwell Time Simulation',
    '이고르 우려사항 대응 분석': 'Igor Concerns Response Analysis',
    '이고르 질문에 대한 답변': 'Igor Questions & Answers',
    '종합 리스크 매트릭스': 'Comprehensive Risk Matrix',
    '권장 조치': 'Recommended Actions',
    '결론': 'Conclusion',
    
    # 표 헤더 및 내용
    '단계': 'Step',
    '작업 내용': 'Work Description',
    '소요 시간': 'Duration',
    '리스크': 'Risk',
    '시간': 'Time',
    '조수 (m)': 'Tide (m)',
    '≥1.8m': '≥1.8m',
    'Deck/Jetty 레벨 차': 'Deck/Jetty Level Difference',
    'Linkspan 상태': 'Linkspan Status',
    '조수 창 (≥1.8m)': 'Tide Window (≥1.8m)',
    '시나리오': 'Scenario',
    '펌프 구성': 'Pump Configuration',
    '발라스팅 시간': 'Ballasting Time',
    'Linkspan 체류': 'Linkspan Dwell',
    '조수 창 여유': 'Tide Window Buffer',
    '리스크 요소': 'Risk Factor',
    '이고르 우려': 'Igor Concern',
    '현재 대응': 'Current Response',
    '잔여 리스크': 'Remaining Risk',
    '항목': 'Item',
    '값': 'Value',
    '출처': 'Source',
    '현재 계획': 'Current Plan',
    '작업 탱크': 'Work Tank',
    'AFT 탱크 용량': 'AFT Tank Capacity',
    '시간': 'Time',
    '결론': 'Conclusion',
    '방안': 'Measure',
    '내용': 'Description',
    '효과': 'Effect',
    '체류 시간': 'Dwell Time',
    '리스크 등급': 'Risk Level',
    '대응': 'Response',
    'Option': 'Option',
    '권장': 'Recommendation',
    '#': 'No.',
    '조치': 'Action',
    '담당': 'Responsible',
    '시점': 'Timing',
    
    # 상태 및 평가
    'SPMT 이동': 'SPMT Movement',
    'Jetty → Linkspan': 'Jetty → Linkspan',
    'Linkspan 통과': 'Linkspan Passage',
    'Linkspan → Deck 진입': 'Linkspan → Deck Entry',
    '조수 대기': 'Tide Waiting',
    'Deck/Jetty 레벨 맞춤': 'Deck/Jetty Level Alignment',
    '발라스팅': 'Ballasting',
    'FWB2.P/S 방출': 'FWB2.P/S Discharge',
    '최종 진입': 'Final Entry',
    'SPMT Deck 완전 진입': 'SPMT Deck Full Entry',
    '낮음': 'Low',
    '높음': 'High',
    '매우 높음': 'Very High',
    '중': 'Medium',
    '부족': 'Insufficient',
    '안전': 'Safe',
    '대기': 'Waiting',
    '통과 가능': 'Passable',
    '최적': 'Optimal',
    '양호': 'Good',
    '위험': 'Dangerous',
    '불가': 'Not Possible',
    '만조': 'High Tide',
    '초과': 'Exceeded',
    '제한적': 'Limited',
    '해소': 'Resolved',
    '관리 필요': 'Management Required',
    '채택': 'Adopted',
    '미채택': 'Not Adopted',
    '권장': 'Recommended',
    '불가': 'Not Feasible',
    '조수 창 최대 활용': 'Maximum Tide Window Utilization',
    '시간 단축': 'Time Reduction',
    '지연 최소화': 'Delay Minimization',
    '정상 작업': 'Normal Operation',
    '모니터링 강화': 'Enhanced Monitoring',
    '작업 중단 검토': 'Work Suspension Review',
    '즉시 철수': 'Immediate Withdrawal',
    '조건부': 'Conditional',
    '외부 펌프 확보': 'Secure External Pump',
    '만조 직후 작업 시작': 'Start Work Immediately After High Tide',
    'Tug 지속 밀기': 'Continuous Tug Pushing',
    '기상 모니터링': 'Weather Monitoring',
    'Hold Point 설정': 'Set Hold Point',
    '해소 가능': 'Resolvable',
    '필수 조건': 'Required Conditions',
    
    # 날짜 및 단위
    '2026-01-29 (Stage 3 - TR Unit 1)': '2026-01-29 (Stage 3 - TR Unit 1)',
    '2026-01-30 (Stage 4 - TR Unit 2)': '2026-01-30 (Stage 4 - TR Unit 2)',
    '2026-01-29 타임라인:': '2026-01-29 Timeline:',
    '2026-01-30 타임라인:': '2026-01-30 Timeline:',
    
    # 특수 용어
    'Linkspan 길이': 'Linkspan Length',
    'Clearance 요구': 'Clearance Requirement',
    'Freeboard 목표': 'Freeboard Target',
    '최소 조수': 'Minimum Tide',
    'Pump Out': 'Pump Out',
    'Transfer FWD→AFT': 'Transfer FWD→AFT',
    'FWB2.P/S → Overboard': 'FWB2.P/S → Overboard',
    'FWD → AFT 탱크': 'FWD → AFT Tank',
    'FWB2.P/S 방출량': 'FWB2.P/S Discharge Volume',
    '만조 직후 시작': 'Start Immediately After High Tide',
    '외부 펌프 사용': 'Use External Pump',
    '사전 준비 완료': 'Pre-preparation Complete',
    '펌프 연결 완료 후 대기': 'Wait After Pump Connection Complete',
    '병렬 작업': 'Parallel Work',
    'FWB2.P/S 동시 방출': 'FWB2.P/S Simultaneous Discharge',
    '작업 3일 전': '3 Days Before Work',
    '당일': 'On the Day',
    '작업 중': 'During Work',
    '외부 펌프, 만조 직후 시작': 'External Pump, Start Immediately After High Tide',
    
    # 이모지 대체
    'OK': 'OK',
    'NG': 'NG',
    'WARN': 'WARN',
    'GREEN': 'GREEN',
    'YELLOW': 'YELLOW',
    'ORANGE': 'ORANGE',
    'RED': 'RED',
}

def translate_text(text):
    """텍스트를 영문으로 번역"""
    if not text or not isinstance(text, str):
        return text
    
    # 직접 매핑된 번역이 있으면 사용
    if text in TRANSLATIONS:
        return TRANSLATIONS[text]
    
    # 부분 매칭 시도
    translated = text
    for korean, english in TRANSLATIONS.items():
        if korean in translated:
            translated = translated.replace(korean, english)
    
    return translated

def translate_excel_to_english(input_file, output_file):
    """엑셀 파일의 한글 내용을 영문으로 변환"""
    wb = openpyxl.load_workbook(input_file)
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 새 워크북 생성
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        new_ws = new_wb.create_sheet(title=translate_text(sheet_name)[:31])
        
        # 모든 셀 복사 및 번역
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                
                # 값 번역
                if cell.value:
                    new_cell.value = translate_text(str(cell.value))
                else:
                    new_cell.value = cell.value
                
                # 스타일 복사
                if cell.has_style:
                    try:
                        if cell.font:
                            new_cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                color=cell.font.color
                            )
                        if cell.fill:
                            new_cell.fill = PatternFill(
                                fill_type=cell.fill.fill_type,
                                start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color
                            )
                        if cell.border:
                            new_cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                        if cell.alignment:
                            new_cell.alignment = Alignment(
                                horizontal=cell.alignment.horizontal,
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text
                            )
                        if cell.number_format:
                            new_cell.number_format = cell.number_format
                    except:
                        pass
        
        # 열 너비 복사
        for col in ws.column_dimensions:
            new_ws.column_dimensions[col].width = ws.column_dimensions[col].width
    
    new_wb.save(output_file)
    print(f"Excel file translated: {output_file}")

if __name__ == '__main__':
    input_file = '이고르_우려사항_Linkspan_체류시간_분석.xlsx'
    output_file = 'Igor_Concerns_Linkspan_Dwell_Time_Analysis.xlsx'
    
    translate_excel_to_english(input_file, output_file)
    try:
        print(f"\n✅ Translation completed!")
        print(f"📄 Input: {input_file}")
        print(f"📊 Output: {output_file}")
    except UnicodeEncodeError:
        print(f"\nTranslation completed!")
        print(f"Input: {input_file}")
        print(f"Output: {output_file}")
