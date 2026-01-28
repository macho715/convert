#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sanity tests for AGI TR MasterSuite workbook (v3.1.0).
These tests validate structure and baseline schedule generation (static).
"""

import datetime as dt
from openpyxl import load_workbook

REQUIRED_SHEETS = [
    "Control_Panel","Scenario_Library","Pattern_Tasks","Schedule_Data",
    "Gantt_Chart","Tide_Data","Weather_Risk","Logs","Exports"
]

def main(path: str):
    wb = load_workbook(path)
    issues=[]

    for sh in REQUIRED_SHEETS:
        if sh not in wb.sheetnames:
            issues.append(f"Missing sheet: {sh}")

    cp=wb["Control_Panel"]
    if cp["C6"].value is None:
        issues.append("Control_Panel!C6 (Scenario) is empty")

    sd=wb["Schedule_Data"]
    # header row
    if str(sd["A5"].value).strip()!="ID" or str(sd["G5"].value).strip()!="Start":
        issues.append("Schedule_Data header row mismatch at row 5")

    # must have at least 20 tasks in baseline
    count=0
    for r in range(6, 500):
        if sd.cell(r,1).value:
            count += 1
    if count < 20:
        issues.append(f"Schedule_Data has too few tasks: {count}")

    # planned finish should be before March 1, 2026 in baseline scenario
    fin=cp["C18"].value
    if isinstance(fin, dt.datetime):
        fin=fin.date()
    if not isinstance(fin, dt.date):
        issues.append("Control_Panel!C18 (finish) not a date")
    else:
        if fin >= dt.date(2026,3,1):
            issues.append(f"Finish not before 2026-03-01: {fin}")

    if issues:
        print("FAIL")
        for i in issues:
            print("-", i)
        raise SystemExit(1)

    print("PASS")
    print("Tasks:", count)
    print("Finish:", cp["C18"].value)
    print("Conflicts:", cp["C22"].value)

if __name__=="__main__":
    import sys
    if len(sys.argv)<2:
        raise SystemExit("Usage: test_mastersuite_v3_1.py <xlsx/xlsm>")
    main(sys.argv[1])
