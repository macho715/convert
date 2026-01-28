#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
TR Document Submission Tracker (Samsung C&T / HVDC TR Transportation)
Py 3.11.8 / openpyxl 3.1.5

What this script does
- CREATE mode: generates a ready-to-use Excel tracker template (Dashboard + Voyage_Schedule + Doc_Matrix + Document_Tracker + etc.)
- REFRESH mode: rebuilds Document_Tracker rows from Voyage_Schedule + Doc_Matrix while preserving user inputs (Status/Submitted/Remarks...)

Notes
- Office LTSC 2021 friendly (no 365-only "Python in Excel" dependency).
- Sensitive data should be kept out of the file. Use [MASK] for emails/IDs if needed.
'''
from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference


# =========================
# Helpers
# =========================
def parse_lead_days(text: Any) -> int:
    """Convert lead time strings into integer day count (safe: range uses MAX)."""
    if text is None:
        return 0
    s = str(text).strip().lower()
    if s in ("n/a", "na", "-", "", "none"):
        return 0
    nums = [int(x) for x in re.findall(r"\d+", s)]
    return max(nums) if nums else 0


def mmdd_to_date(mmdd: str, year: int) -> Optional[dt.date]:
    """Parse 'MM-DD' into date."""
    if not mmdd:
        return None
    s = str(mmdd).strip()
    m = re.match(r"^\s*(\d{1,2})[-/](\d{1,2})\s*$", s)
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def extract_start_mmdd(range_text: Any) -> Optional[str]:
    """From '01-29~30' return '01-29'."""
    if range_text is None:
        return None
    s = str(range_text).strip()
    if "~" in s:
        return s.split("~", 1)[0].strip()
    return s


def auto_due_basis(category: str) -> str:
    """Default Due Based On mapping (users can edit Doc_Matrix)."""
    c = (category or "").lower()
    if "gate pass" in c:
        return "MZP Arrival"
    if "traffic" in c or "site" in c or "permit" in c:
        return "Land Permit By"
    if "customs" in c or "port" in c or "mws" in c:
        return "Doc Deadline"
    return "AUTO"


def col_letter(n: int) -> str:
    return get_column_letter(n)


def set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[col_letter(col_idx)].width = w


def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{col_letter(start_col)}{start_row}:{col_letter(end_col)}{end_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


# =========================
# Workbook Builder
# =========================
def build_template(output_path: Path) -> Path:
    wb = Workbook()

    # ---- Styles
    header_fill = PatternFill("solid", fgColor="1E3A5F")     # Dark blue
    subheader_fill = PatternFill("solid", fgColor="2D5A8A")  # Medium blue
    gray_fill = PatternFill("solid", fgColor="F2F2F2")

    font_title = Font(bold=True, size=16, color="1E3A5F")
    font_header = Font(bold=True, size=11, color="FFFFFF")
    font_normal = Font(size=10, color="000000")
    font_bold = Font(bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =========================
    # Sheet: Lists (DV sources)
    # =========================
    ws_lists = wb.active
    ws_lists.title = "Lists"

    status_list = ["Not Started", "In Progress", "Submitted", "Approved", "Rejected", "On Hold", "Not Required"]
    priority_list = ["Mandatory", "Important", "Optional"]
    due_basis_list = ["AUTO", "Doc Deadline", "Land Permit By", "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival"]
    yn_list = ["Y", "N"]
    site_list = ["Site-1", "Site-2", "Site-3", "Site-4", "Site-5", "Site-6"]

    ws_lists["A1"] = "Status"
    ws_lists["B1"] = "Priority"
    ws_lists["C1"] = "Due_Basis"
    ws_lists["D1"] = "Y/N"
    ws_lists["E1"] = "Sites"

    for i, v in enumerate(status_list, start=2):
        ws_lists[f"A{i}"] = v
    for i, v in enumerate(priority_list, start=2):
        ws_lists[f"B{i}"] = v
    for i, v in enumerate(due_basis_list, start=2):
        ws_lists[f"C{i}"] = v
    for i, v in enumerate(yn_list, start=2):
        ws_lists[f"D{i}"] = v
    for i, v in enumerate(site_list, start=2):
        ws_lists[f"E{i}"] = v

    for c in range(1, 6):
        ws_lists.cell(row=1, column=c).font = font_bold
        ws_lists.cell(row=1, column=c).fill = gray_fill

    set_col_widths(ws_lists, {1: 18, 2: 12, 3: 18, 4: 6, 5: 12})
    ws_lists.sheet_state = "hidden"  # hide helper sheet in production

    # =========================
    # Sheet: Config
    # =========================
    ws_cfg = wb.create_sheet("Config")
    ws_cfg["A1"] = "TR Document Tracker - Config"
    ws_cfg["A1"].font = font_title

    default_year = dt.date.today().year
    cfg_rows = [
        ("DocCutoff_Days", 3, "Doc Deadline = MZP Departure - DocCutoff_Days"),
        ("LandPermitLead_Days", 7, "Land Permit By = AGI Arrival - LandPermitLead_Days"),
        ("DueSoon_Threshold_Days", 7, "Dashboard/CF: upcoming items within N days"),
        ("Amber_Threshold_Days", 7, "Dashboard KPI: D-7 Amber warning threshold"),
        ("Red_Threshold_Days", 3, "Dashboard KPI: D-3 Red warning threshold"),
        ("Critical_Threshold_Days", 1, "Dashboard KPI: D-1 Critical alert threshold"),
        ("Default_Year", default_year, "Year used for sample schedule MM-DD parsing"),
    ]

    ws_cfg["A3"] = "Key"
    ws_cfg["B3"] = "Value"
    ws_cfg["C3"] = "Description"
    for cell in (ws_cfg["A3"], ws_cfg["B3"], ws_cfg["C3"]):
        cell.font = font_bold
        cell.fill = gray_fill
        cell.alignment = align_left
        cell.border = border_thin

    for i, (k, v, dsc) in enumerate(cfg_rows, start=4):
        ws_cfg[f"A{i}"] = k
        ws_cfg[f"B{i}"] = v
        ws_cfg[f"C{i}"] = dsc
        for c in range(1, 4):
            ws_cfg.cell(row=i, column=c).border = border_thin
            ws_cfg.cell(row=i, column=c).alignment = align_left

    set_col_widths(ws_cfg, {1: 26, 2: 18, 3: 60})

    # =========================
    # Sheet: Inputs (Scenario)
    # =========================
    ws_inputs = wb.create_sheet("Inputs")
    ws_inputs["A1"] = "Schedule Inputs (Scenario)"
    ws_inputs["A1"].font = font_title

    ws_inputs["A2"] = "Item"
    ws_inputs["B2"] = "Value (Auto)"
    ws_inputs["C2"] = "Value (Manual)"
    for c in range(1, 4):
        cell = ws_inputs.cell(row=2, column=c)
        cell.font = font_bold
        cell.fill = gray_fill
        cell.border = border_thin
        cell.alignment = align_left if c == 1 else align_center

    labels = ["Arrival", "RoRo Start", "RoRo End", "Departure"]
    manual_defaults = [
        dt.date(default_year, 1, 27),
        dt.date(default_year, 1, 29),
        dt.date(default_year, 1, 30),
        dt.date(default_year, 2, 1),
    ]
    for i, label in enumerate(labels):
        r = 3 + i
        ws_inputs.cell(row=r, column=1, value=label)
        ws_inputs.cell(row=r, column=2, value=f'=IF($B$9="CUSTOM",$C{r},VLOOKUP($B$9,$A$12:$E$17,{i + 2},FALSE))')
        ws_inputs.cell(row=r, column=3, value=manual_defaults[i])
        for c in range(1, 4):
            cell = ws_inputs.cell(row=r, column=c)
            cell.border = border_thin
            cell.alignment = align_left if c == 1 else align_center
            if c > 1:
                cell.number_format = "yyyy-mm-dd"

    ws_inputs["A7"] = "Weekend Pattern (Mon..Sun, 0=Work,1=Weekend)"
    ws_inputs["B7"] = "0000011"
    ws_inputs["C7"] = "Example: Sat-Sun=0000011 / Fri-Sun=0000111 (edit as needed)"

    ws_inputs["A9"] = "Schedule Scenario (Select)"
    ws_inputs["B9"] = "SCN-01"
    ws_inputs["C9"] = "(Select SCN code; choose CUSTOM to use manual values in Column C for rows 3-6)"

    ws_inputs["A11"] = "Scenario"
    ws_inputs["B11"] = "Arrival"
    ws_inputs["C11"] = "RoRo Start"
    ws_inputs["D11"] = "RoRo End"
    ws_inputs["E11"] = "Departure"
    for c in range(1, 6):
        cell = ws_inputs.cell(row=11, column=c)
        cell.font = font_bold
        cell.fill = gray_fill
        cell.border = border_thin
        cell.alignment = align_center

    ws_inputs["A12"] = "SCN-01"
    ws_inputs["B12"] = dt.date(default_year, 1, 27)
    ws_inputs["C12"] = dt.date(default_year, 1, 29)
    ws_inputs["D12"] = dt.date(default_year, 1, 30)
    ws_inputs["E12"] = dt.date(default_year, 2, 1)
    for scn_row, scn in enumerate(["SCN-02", "SCN-03", "SCN-04", "SCN-05", "CUSTOM"], start=13):
        ws_inputs.cell(scn_row, 1).value = scn
        for c in range(2, 6):
            ws_inputs.cell(scn_row, c).value = None

    dv_scn = DataValidation(type="list", formula1="=$A$12:$A$17", allow_blank=False)
    dv_wkend = DataValidation(type="list", formula1='"0000011,0000111"', allow_blank=False)
    ws_inputs.add_data_validation(dv_scn)
    ws_inputs.add_data_validation(dv_wkend)
    dv_scn.add("B9")
    dv_wkend.add("B7")

    set_col_widths(ws_inputs, {1: 28, 2: 18, 3: 50, 4: 18, 5: 18})

    # =========================
    # Sheet: Voyage_Schedule (Input)
    # =========================
    ws_voy = wb.create_sheet("Voyage_Schedule")
    ws_voy["A1"] = "Voyage Schedule (Input)"
    ws_voy["A1"].font = font_title
    ws_voy["A2"] = f"Last Updated: {now_str()}"
    ws_voy["A2"].font = Font(italic=True, size=10)

    headers_voy = [
        "Voyage", "Site", "TR Units",
        "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival",
        "Doc Deadline", "Land Permit By", "Submission Date",
        "Status", "Remarks"
    ]
    header_row = 4
    for col, h in enumerate(headers_voy, start=1):
        cell = ws_voy.cell(row=header_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    year = int(ws_cfg["B10"].value)  # Default_Year (row 10)
    sample_rows = [
        ("Voyage 1", "Site-1", "TR 1-2", "01-27", "01-29~30", "02-01", "02-02"),
        ("Voyage 2", "Site-2", "TR 3-4", "02-06", "02-07~08", "02-10", "02-11"),
        ("Voyage 3", "Site-3", "TR 5-6", "02-15", "02-16~17", "02-19", "02-20"),
        ("Voyage 4", "Site-4", "TR 7", "02-24", "02-25", "02-27", "02-28"),
    ]

    start_row = header_row + 1
    for i, row_data in enumerate(sample_rows):
        r = start_row + i
        voyage, site, tr_units, mzp_arr, loadout, mzp_dep, agi_arr = row_data

        ws_voy.cell(row=r, column=1, value=voyage)
        ws_voy.cell(row=r, column=2, value=site)
        ws_voy.cell(row=r, column=3, value=tr_units)

        if i == 0:
            ws_voy.cell(row=r, column=4, value="=Inputs!$B$3")
            ws_voy.cell(row=r, column=5, value="=Inputs!$B$4")
            ws_voy.cell(row=r, column=6, value="=Inputs!$B$6")
            ws_voy.cell(row=r, column=7, value="=IF(Inputs!$B$6=\"\",\"\",Inputs!$B$6+1)")
        else:
            ws_voy.cell(row=r, column=4, value=mmdd_to_date(extract_start_mmdd(mzp_arr) or "", year))
            ws_voy.cell(row=r, column=5, value=mmdd_to_date(extract_start_mmdd(loadout) or "", year))
            ws_voy.cell(row=r, column=6, value=mmdd_to_date(extract_start_mmdd(mzp_dep) or "", year))
            ws_voy.cell(row=r, column=7, value=mmdd_to_date(extract_start_mmdd(agi_arr) or "", year))

        # Default formulas for deadlines
        ws_voy.cell(row=r, column=8, value=f"=IF(F{r}=\"\",\"\",F{r}-Config!$B$4)")
        ws_voy.cell(row=r, column=9, value=f"=IF(G{r}=\"\",\"\",G{r}-Config!$B$5)")
        ws_voy.cell(row=r, column=10, value=f"=IF(AND(H{r}<>\"\",I{r}<>\"\"),MIN(H{r},I{r}),IF(H{r}<>\"\",H{r},I{r}))")

        ws_voy.cell(row=r, column=11, value="Open")
        ws_voy.cell(row=r, column=12, value=f"Load-out plan: {loadout}")

        for c in range(1, 13):
            cell = ws_voy.cell(row=r, column=c)
            cell.font = font_normal
            cell.border = border_thin
            cell.alignment = align_center if c <= 11 else align_left

        for c in (4, 5, 6, 7, 8, 9, 10):
            ws_voy.cell(row=r, column=c).number_format = "yyyy-mm-dd"

    ws_voy.freeze_panes = "A5"
    set_col_widths(ws_voy, {
        1: 16, 2: 10, 3: 12,
        4: 12, 5: 12, 6: 14, 7: 12,
        8: 12, 9: 14, 10: 14,
        11: 10, 12: 28
    })
    add_table(ws_voy, "tblVoyage", header_row, 1, start_row + len(sample_rows) - 1, len(headers_voy))

    dv_site = DataValidation(type="list", formula1="=Lists!$E$2:$E$7", allow_blank=True)
    dv_status_voy = DataValidation(type="list", formula1="=Lists!$A$2:$A$8", allow_blank=True)
    ws_voy.add_data_validation(dv_site)
    ws_voy.add_data_validation(dv_status_voy)
    dv_site.add(f"B{start_row}:B5000")
    dv_status_voy.add(f"K{start_row}:K5000")

    # =========================
    # Sheet: Doc_Matrix (Input)
    # =========================
    ws_doc = wb.create_sheet("Doc_Matrix")
    ws_doc["A1"] = "Document Requirement Matrix (Input)"
    ws_doc["A1"].font = font_title
    ws_doc["A2"] = "Edit lead-times / due-basis here. Tracker refresh will follow this."
    ws_doc["A2"].font = Font(italic=True, size=10)

    headers_doc = [
        "Doc Code", "Category", "Document Name", "Description",
        "Priority", "Responsible Party",
        "Lead Days", "Due Based On", "Due Offset Days",
        "Required (Y/N)", "Approval Stage", "Default Recipient", "Notes"
    ]
    doc_header_row = 4
    for col, h in enumerate(headers_doc, start=1):
        cell = ws_doc.cell(row=doc_header_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    # Auto-pull doc list from prototype workbook if present
    default_docs: List[List[Any]] = []
    proto = output_path.parent / "TR_Document_Tracker_VBA.xlsx"
    if proto.exists():
        try:
            wb_proto = load_workbook(proto, data_only=True)
            ws_proto = wb_proto["Document_Tracker"]
            for rr in range(5, 2000):
                doc_no = ws_proto.cell(row=rr, column=1).value
                if doc_no is None:
                    continue
                default_docs.append([
                    ws_proto.cell(row=rr, column=1).value,
                    ws_proto.cell(row=rr, column=2).value,
                    ws_proto.cell(row=rr, column=3).value,
                    ws_proto.cell(row=rr, column=4).value,
                    ws_proto.cell(row=rr, column=5).value,
                    ws_proto.cell(row=rr, column=6).value,
                    ws_proto.cell(row=rr, column=7).value,
                ])
        except Exception:
            default_docs = []

    if not default_docs:
        default_docs = [
            ["1", "Customs & Port", "Bill of Lading Draft", "BL draft for review", "Mandatory", "Shipping Agent", "3 days"],
            ["2", "Permit", "Land Permit Application", "Road/land permit application", "Mandatory", "Agency", "5 days"],
        ]

    doc_start_row = doc_header_row + 1
    for i, d in enumerate(default_docs):
        r = doc_start_row + i
        doc_no, cat, name, desc, pri, resp, lead_str = d
        lead_days = parse_lead_days(lead_str)
        due_basis = auto_due_basis(str(cat))

        doc_code = f"D{int(doc_no):03d}" if str(doc_no).isdigit() else str(doc_no)
        ws_doc.cell(row=r, column=1, value=doc_code)
        ws_doc.cell(row=r, column=2, value=cat)
        ws_doc.cell(row=r, column=3, value=name)
        ws_doc.cell(row=r, column=4, value=desc)
        ws_doc.cell(row=r, column=5, value=pri)
        ws_doc.cell(row=r, column=6, value=resp)
        ws_doc.cell(row=r, column=7, value=lead_days)
        ws_doc.cell(row=r, column=8, value=due_basis)
        ws_doc.cell(row=r, column=9, value=0)
        ws_doc.cell(row=r, column=10, value="Y")
        ws_doc.cell(row=r, column=11, value="Submit → Review → Approve")
        ws_doc.cell(row=r, column=12, value="")
        ws_doc.cell(row=r, column=13, value="")

        for c in range(1, len(headers_doc) + 1):
            cell = ws_doc.cell(row=r, column=c)
            cell.font = font_normal
            cell.border = border_thin
            cell.alignment = align_left if c in (2, 3, 4, 6, 11, 12, 13) else align_center

    ws_doc.freeze_panes = "A5"
    set_col_widths(ws_doc, {
        1: 10, 2: 18, 3: 28, 4: 36,
        5: 12, 6: 18,
        7: 10, 8: 14, 9: 14,
        10: 14, 11: 22, 12: 20, 13: 20
    })
    add_table(ws_doc, "tblDoc", doc_header_row, 1, doc_start_row + len(default_docs) - 1, len(headers_doc))

    dv_priority = DataValidation(type="list", formula1="=Lists!$B$2:$B$4", allow_blank=True)
    dv_due_basis = DataValidation(type="list", formula1="=Lists!$C$2:$C$8", allow_blank=True)
    dv_yn = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=True)
    ws_doc.add_data_validation(dv_priority)
    ws_doc.add_data_validation(dv_due_basis)
    ws_doc.add_data_validation(dv_yn)
    dv_priority.add(f"E{doc_start_row}:E5000")
    dv_due_basis.add(f"H{doc_start_row}:H5000")
    dv_yn.add(f"J{doc_start_row}:J5000")

    # =========================
    # Sheet: Party_Contacts
    # =========================
    ws_ct = wb.create_sheet("Party_Contacts")
    ws_ct["A1"] = "Responsible Party Contacts (Optional)"
    ws_ct["A1"].font = font_title
    ws_ct["A2"] = "Fill email addresses for reminder automation (VBA)."
    ws_ct["A2"].font = Font(italic=True, size=10)

    headers_ct = ["Responsible Party", "Email", "CC (Optional)"]
    ct_header_row = 4
    for col, h in enumerate(headers_ct, start=1):
        cell = ws_ct.cell(row=ct_header_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    parties = sorted({str(ws_doc.cell(row=r, column=6).value).strip()
                      for r in range(doc_start_row, doc_start_row + len(default_docs))
                      if ws_doc.cell(row=r, column=6).value})

    ct_start_row = ct_header_row + 1
    for i, p in enumerate(parties):
        r = ct_start_row + i
        ws_ct.cell(row=r, column=1, value=p)
        ws_ct.cell(row=r, column=2, value="[MASK]@company.com")
        ws_ct.cell(row=r, column=3, value="")
        for c in range(1, 4):
            cell = ws_ct.cell(row=r, column=c)
            cell.font = font_normal
            cell.border = border_thin
            cell.alignment = align_left

    set_col_widths(ws_ct, {1: 22, 2: 30, 3: 30})
    ws_ct.freeze_panes = "A5"
    add_table(ws_ct, "tblContacts", ct_header_row, 1, ct_start_row + len(parties) - 1, 3)

    # =========================
    # Sheet: Document_Tracker (Main)
    # =========================
    ws_tr = wb.create_sheet("Document_Tracker")
    ws_tr["A1"] = "TR Document Tracker (Main)"
    ws_tr["A1"].font = font_title
    ws_tr["A2"] = f"Last Refreshed: {now_str()}  |  목적: Responsible Party별 서류 제출일자/데드라인/현황 공유 → 누락 방지"
    ws_tr["A2"].font = Font(italic=True, size=10)

    tracker_headers = [
        "Record ID", "Voyage", "Site", "TR Units",
        "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival",
        "Doc Deadline", "Land Permit By",
        "Doc Code", "Category", "Document Name", "Description",
        "Priority", "Responsible Party",
        "Lead Days", "Due Based On", "Due Offset Days",
        "SUBMISSION DATE", "Start By (Prep)",
        "Status", "Submitted Date", "Approved/Received Date",
        "Remarks", "File Link", "Revision", "Transmittal No",
        "Last Updated", "Updated By"
    ]

    tr_header_row = 4
    for col, h in enumerate(tracker_headers, start=1):
        cell = ws_tr.cell(row=tr_header_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    # Build rows: voyages x docs (Required=Y)
    # Read voyages directly from tblVoyage area
    voyages = []
    for rr in range(start_row, start_row + len(sample_rows)):
        v_name = ws_voy.cell(row=rr, column=1).value
        if v_name:
            voyages.append(str(v_name))

    # Collect docs where Required = Y
    docs_for_tracker = []
    for rr in range(doc_start_row, doc_start_row + len(default_docs)):
        doc_code = ws_doc.cell(row=rr, column=1).value
        if not doc_code:
            continue
        req = str(ws_doc.cell(row=rr, column=10).value or "").strip().upper()
        if req != "Y":
            continue
        docs_for_tracker.append({
            "Doc Code": ws_doc.cell(row=rr, column=1).value,
            "Category": ws_doc.cell(row=rr, column=2).value,
            "Document Name": ws_doc.cell(row=rr, column=3).value,
            "Description": ws_doc.cell(row=rr, column=4).value,
            "Priority": ws_doc.cell(row=rr, column=5).value,
            "Responsible Party": ws_doc.cell(row=rr, column=6).value,
            "Lead Days": ws_doc.cell(row=rr, column=7).value,
            "Due Based On": ws_doc.cell(row=rr, column=8).value,
            "Due Offset Days": ws_doc.cell(row=rr, column=9).value,
        })

    tr_start_row = tr_header_row + 1
    row_cursor = tr_start_row

    for v in voyages:
        for d in docs_for_tracker:
            key = f"{v}|{d['Doc Code']}"
            ws_tr.cell(row=row_cursor, column=1, value=key)
            ws_tr.cell(row=row_cursor, column=2, value=v)
            # Site/TR Units lookup from Voyage_Schedule
            ws_tr.cell(row=row_cursor, column=3, value=f"=IFERROR(INDEX(tblVoyage[Site],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=4, value=f"=IFERROR(INDEX(tblVoyage[TR Units],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")

            # Milestone lookups
            ws_tr.cell(row=row_cursor, column=5, value=f"=IFERROR(INDEX(tblVoyage[MZP Arrival],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=6, value=f"=IFERROR(INDEX(tblVoyage[Load-out],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=7, value=f"=IFERROR(INDEX(tblVoyage[MZP Departure],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=8, value=f"=IFERROR(INDEX(tblVoyage[AGI Arrival],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=9, value=f"=IFERROR(INDEX(tblVoyage[Doc Deadline],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=10, value=f"=IFERROR(INDEX(tblVoyage[Land Permit By],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")

            ws_tr.cell(row=row_cursor, column=11, value=d["Doc Code"])
            ws_tr.cell(row=row_cursor, column=12, value=d["Category"])
            ws_tr.cell(row=row_cursor, column=13, value=d["Document Name"])
            ws_tr.cell(row=row_cursor, column=14, value=d["Description"])
            ws_tr.cell(row=row_cursor, column=15, value=d["Priority"])
            ws_tr.cell(row=row_cursor, column=16, value=d["Responsible Party"])
            ws_tr.cell(row=row_cursor, column=17, value=d["Lead Days"])
            ws_tr.cell(row=row_cursor, column=18, value=d["Due Based On"])
            ws_tr.cell(row=row_cursor, column=19, value=d["Due Offset Days"])

            # SUBMISSION DATE formula (uses IF; if locale separator issue, use refresh mode to compute as values)
            ws_tr.cell(row=row_cursor, column=20, value=(
                f"=IF(R{row_cursor}=\"Doc Deadline\",I{row_cursor},"
                f"IF(R{row_cursor}=\"Land Permit By\",J{row_cursor},"
                f"IF(R{row_cursor}=\"MZP Arrival\",E{row_cursor},"
                f"IF(R{row_cursor}=\"Load-out\",F{row_cursor},"
                f"IF(R{row_cursor}=\"MZP Departure\",G{row_cursor},"
                f"IF(R{row_cursor}=\"AGI Arrival\",H{row_cursor},"
                f"IF(AND(I{row_cursor}<>\"\",J{row_cursor}<>\"\"),MIN(I{row_cursor},J{row_cursor}),IF(I{row_cursor}<>\"\",I{row_cursor},J{row_cursor}))"
                f"))))))"
                f"+S{row_cursor}"
            ))
            ws_tr.cell(row=row_cursor, column=21, value=f"=IF(T{row_cursor}=\"\",\"\",T{row_cursor}-Q{row_cursor})")

            ws_tr.cell(row=row_cursor, column=22, value="Not Started")
            # user fields blank
            for cc in range(23, 31):
                ws_tr.cell(row=row_cursor, column=cc, value="")

            for cc in range(1, len(tracker_headers) + 1):
                cell = ws_tr.cell(row=row_cursor, column=cc)
                cell.border = border_thin
                cell.font = font_normal
                cell.alignment = align_left if cc in (14, 25, 26) else align_center

            for cc in (5, 6, 7, 8, 9, 10, 20, 21, 23, 24, 29):
                ws_tr.cell(row=row_cursor, column=cc).number_format = "yyyy-mm-dd"

            row_cursor += 1

    ws_tr.freeze_panes = "A5"
    set_col_widths(ws_tr, {
        1: 28, 2: 14, 3: 10, 4: 12,
        5: 12, 6: 12, 7: 14, 8: 12,
        9: 12, 10: 14,
        11: 10, 12: 18, 13: 28, 14: 36,
        15: 12, 16: 18,
        17: 10, 18: 14, 19: 14,
        20: 14, 21: 14,
        22: 12, 23: 14, 24: 18,
        25: 24, 26: 18, 27: 10, 28: 14,
        29: 16, 30: 12
    })

    last_row = row_cursor - 1
    add_table(ws_tr, "tblTracker", tr_header_row, 1, last_row, len(tracker_headers))

    dv_status = DataValidation(type="list", formula1="=Lists!$A$2:$A$8", allow_blank=True)
    ws_tr.add_data_validation(dv_status)
    dv_status.add(f"V{tr_start_row}:V5000")

    # Conditional formatting (row-wise)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    gray2_fill = PatternFill("solid", fgColor="E7E6E6")

    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:AD5000",
        FormulaRule(formula=[f"=($T{tr_start_row}<TODAY())*($V{tr_start_row}<>\"Approved\")*($V{tr_start_row}<>\"Submitted\")*($V{tr_start_row}<>\"Not Required\")*($T{tr_start_row}<>\"\")"], fill=red_fill)
    )
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:AD5000",
        FormulaRule(formula=[f"=(($T{tr_start_row}-TODAY())<=Config!$B$6)*(($T{tr_start_row}-TODAY())>=0)*($V{tr_start_row}<>\"Approved\")*($V{tr_start_row}<>\"Submitted\")*($V{tr_start_row}<>\"Not Required\")*($T{tr_start_row}<>\"\")"], fill=yellow_fill)
    )
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:AD5000",
        FormulaRule(formula=[f"=($V{tr_start_row}=\"Approved\")+($V{tr_start_row}=\"Submitted\")"], fill=green_fill)
    )
    ws_tr.conditional_formatting.add(
        f"A{tr_start_row}:AD5000",
        FormulaRule(formula=[f"=($V{tr_start_row}=\"Not Required\")"], fill=gray2_fill)
    )

    # =========================
    # Dashboard (summary)
    # =========================
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash["A1"] = "TR Document Preparation / Submission Dashboard"
    ws_dash["A1"].font = Font(bold=True, size=18, color="1E3A5F")
    ws_dash["A3"] = f"Last Updated: {now_str()}"
    ws_dash["A3"].font = Font(italic=True, size=10)

    ws_dash["A5"] = "KPI"
    ws_dash["A5"].font = font_bold
    ws_dash["A6"] = "Total Docs"
    ws_dash["A7"] = "Completed"
    ws_dash["A8"] = "Overdue"
    ws_dash["A9"] = "Due Soon (<=N days)"
    ws_dash["A10"] = "Completion %"
    ws_dash["A11"] = "D-7 Count (Amber)"
    ws_dash["A12"] = "D-3 Count (Red)"
    ws_dash["A13"] = "D-1 Count (Critical)"

    ws_dash["B6"] = "=COUNTA(tblTracker[Record ID])"
    ws_dash["B7"] = "=SUMPRODUCT((tblTracker[Status]=\"Approved\")+(tblTracker[Status]=\"Submitted\"))"
    ws_dash["B8"] = "=SUMPRODUCT((tblTracker[SUBMISSION DATE]<TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))"
    ws_dash["B9"] = "=SUMPRODUCT((tblTracker[SUBMISSION DATE]>=TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$6)*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))"
    ws_dash["B10"] = "=IFERROR(B7/B6,0)"
    ws_dash["B10"].number_format = "0%"
    ws_dash["B11"] = "=SUMPRODUCT((tblTracker[SUBMISSION DATE]>=TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$7)*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))"
    ws_dash["B12"] = "=SUMPRODUCT((tblTracker[SUBMISSION DATE]>=TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$8)*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))"
    ws_dash["B13"] = "=SUMPRODUCT((tblTracker[SUBMISSION DATE]>=TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$9)*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))"

    for rr in range(6, 14):
        ws_dash[f"A{rr}"].border = border_thin
        ws_dash[f"B{rr}"].border = border_thin
        ws_dash[f"B{rr}"].font = Font(bold=True, size=11)

    # Voyage summary block
    row = 15
    ws_dash.merge_cells(f"A{row}:I{row}")
    ws_dash[f"A{row}"] = "📅 Voyage Schedule Summary"
    ws_dash[f"A{row}"].font = font_header
    ws_dash[f"A{row}"].fill = header_fill
    ws_dash[f"A{row}"].alignment = align_left

    row += 1
    voyage_sum_headers = ["Voyage", "TR Units", "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival", "Doc Deadline", "Land Permit By", "Submission Date"]
    for col, h in enumerate(voyage_sum_headers, start=1):
        cell = ws_dash.cell(row=row, column=col, value=h)
        cell.font = font_header
        cell.fill = subheader_fill
        cell.alignment = align_center
        cell.border = border_thin

    for i in range(len(sample_rows)):
        rr = row + 1 + i
        idx_excel = i + 1
        ws_dash.cell(row=rr, column=1, value=f"=INDEX(tblVoyage[Voyage],{idx_excel})")
        ws_dash.cell(row=rr, column=2, value=f"=INDEX(tblVoyage[TR Units],{idx_excel})")
        ws_dash.cell(row=rr, column=3, value=f"=INDEX(tblVoyage[MZP Arrival],{idx_excel})")
        ws_dash.cell(row=rr, column=4, value=f"=INDEX(tblVoyage[Load-out],{idx_excel})")
        ws_dash.cell(row=rr, column=5, value=f"=INDEX(tblVoyage[MZP Departure],{idx_excel})")
        ws_dash.cell(row=rr, column=6, value=f"=INDEX(tblVoyage[AGI Arrival],{idx_excel})")
        ws_dash.cell(row=rr, column=7, value=f"=INDEX(tblVoyage[Doc Deadline],{idx_excel})")
        ws_dash.cell(row=rr, column=8, value=f"=INDEX(tblVoyage[Land Permit By],{idx_excel})")
        ws_dash.cell(row=rr, column=9, value=f"=INDEX(tblVoyage[Submission Date],{idx_excel})")
        for c in range(1, 10):
            cell = ws_dash.cell(row=rr, column=c)
            cell.border = border_thin
            cell.alignment = align_center
            if c >= 3:
                cell.number_format = "yyyy-mm-dd"

    # Party progress table
    row = row + len(sample_rows) + 3
    ws_dash.merge_cells(f"A{row}:H{row}")
    ws_dash[f"A{row}"] = "📊 Responsible Party Progress"
    ws_dash[f"A{row}"].font = font_header
    ws_dash[f"A{row}"].fill = header_fill
    ws_dash[f"A{row}"].alignment = align_left

    row += 1
    party_headers = ["Responsible Party", "Total", "Completed", "In Progress", "Not Started", "Overdue", "Due Soon", "Progress %"]
    for col, h in enumerate(party_headers, start=1):
        cell = ws_dash.cell(row=row, column=col, value=h)
        cell.font = font_header
        cell.fill = subheader_fill
        cell.alignment = align_center
        cell.border = border_thin

    party_start = row + 1
    for i, p in enumerate(parties):
        rr = party_start + i
        ws_dash.cell(row=rr, column=1, value=p)
        ws_dash.cell(row=rr, column=2, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*(tblTracker[Record ID]<>\"\"))")
        ws_dash.cell(row=rr, column=3, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*((tblTracker[Status]=\"Approved\")+(tblTracker[Status]=\"Submitted\")))")
        ws_dash.cell(row=rr, column=4, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*(tblTracker[Status]=\"In Progress\"))")
        ws_dash.cell(row=rr, column=5, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*(tblTracker[Status]=\"Not Started\"))")
        ws_dash.cell(row=rr, column=6, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*(tblTracker[SUBMISSION DATE]<TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))")
        ws_dash.cell(row=rr, column=7, value=f"=SUMPRODUCT((tblTracker[Responsible Party]=$A{rr})*(tblTracker[SUBMISSION DATE]>=TODAY())*(tblTracker[SUBMISSION DATE]<>\"\")*(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$6)*(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\"))")
        ws_dash.cell(row=rr, column=8, value=f"=IFERROR(C{rr}/B{rr},0)")
        ws_dash.cell(row=rr, column=8).number_format = "0%"
        for c in range(1, 9):
            cell = ws_dash.cell(row=rr, column=c)
            cell.border = border_thin
            cell.alignment = align_left if c == 1 else align_center

    chart = BarChart()
    chart.type = "col"
    chart.title = "Completed Docs by Responsible Party"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Party"
    data = Reference(ws_dash, min_col=3, min_row=party_start-1, max_row=party_start + len(parties) - 1)
    cats = Reference(ws_dash, min_col=1, min_row=party_start, max_row=party_start + len(parties) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws_dash.add_chart(chart, f"J{party_start-1}")

    set_col_widths(ws_dash, {1: 26, 2: 18, 3: 2, 4: 2, 5: 26})

    # =========================
    # Party_View (dynamic filter)
    # =========================
    ws_view = wb.create_sheet("Party_View")
    ws_view["A1"] = "Responsible Party View (Pick 1 party)"
    ws_view["A1"].font = font_title
    ws_view["A2"] = "Select party in B3 → see tasks due soon/overdue."
    ws_view["A2"].font = Font(italic=True, size=10)
    ws_view["A3"] = "Responsible Party"
    ws_view["A3"].font = font_bold
    ws_view["B3"] = parties[0] if parties else ""

    dv_party = DataValidation(type="list", formula1=f"=Party_Contacts!$A$5:$A${4+len(parties)}", allow_blank=True)
    ws_view.add_data_validation(dv_party)
    dv_party.add("B3")

    view_headers = ["Record ID", "Voyage", "Doc Code", "Document Name", "SUBMISSION DATE", "Status", "Remarks"]
    base_row = 6
    for col, h in enumerate(view_headers, start=1):
        cell = ws_view.cell(row=base_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    ws_view["A7"] = (
        "=IFERROR("
        "FILTER("
        "tblTracker[[Record ID]:[Remarks]],"
        "(tblTracker[Responsible Party]=$B$3)*"
        "((tblTracker[SUBMISSION DATE]<TODAY())+(tblTracker[SUBMISSION DATE]<=TODAY()+Config!$B$6))*"
        "(tblTracker[SUBMISSION DATE]<>\"\")*"
        "(tblTracker[Status]<>\"Approved\")*(tblTracker[Status]<>\"Submitted\")*(tblTracker[Status]<>\"Not Required\")"
        "),"
        "\"No rows\""
        ")"
    )

    set_col_widths(ws_view, {1: 28, 2: 14, 3: 10, 4: 28, 5: 14, 6: 12, 7: 30})
    ws_view.freeze_panes = "A7"

    # =========================
    # LOG (Optional)
    # =========================
    ws_log = wb.create_sheet("LOG")
    ws_log["A1"] = "Activity / Error Log"
    ws_log["A1"].font = font_title
    ws_log["A2"] = "VBA/Python 실행 로그를 남기고 싶을 때 사용 (민감정보 기록 금지)."
    ws_log["A2"].font = Font(italic=True, size=10)

    log_headers = ["Timestamp", "Action", "User", "Message"]
    log_header_row = 4
    for col, h in enumerate(log_headers, start=1):
        cell = ws_log.cell(row=log_header_row, column=col, value=h)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    # Sample first row
    ws_log.cell(row=5, column=1, value=now_str())
    ws_log.cell(row=5, column=2, value="CREATE")
    ws_log.cell(row=5, column=3, value="")
    ws_log.cell(row=5, column=4, value="Template created")

    for c in range(1, 5):
        ws_log.cell(row=5, column=c).border = border_thin
        ws_log.cell(row=5, column=c).alignment = align_left if c == 4 else align_center
        ws_log.cell(row=5, column=c).font = font_normal

    ws_log.freeze_panes = "A5"
    set_col_widths(ws_log, {1: 20, 2: 14, 3: 18, 4: 80})

    # =========================
    # Instructions
    # =========================
    ws_ins = wb.create_sheet("Instructions")
    ws_ins["A1"] = "How to use"
    ws_ins["A1"].font = font_title
    instructions = [
        "1) Voyage_Schedule 입력/수정 (항차, Site, 날짜 등)",
        "2) Doc_Matrix에서 Responsible Party / Lead Days / Due Based On / Required(Y/N) 관리",
        "3) Document_Tracker에서 Status + Submitted/Approved Date 업데이트",
        "4) Dashboard에서 Overdue/Due Soon 및 Party별 진행률 확인",
        "5) (옵션) Party_Contacts에 이메일 입력 후 VBA로 리마인더 드래프트 생성",
        "",
        "SUBMISSION DATE 계산 로직:",
        "- Doc_Matrix[Due Based On] 값 기준으로 해당 마일스톤 날짜를 가져와 제출기한으로 사용",
        "- AUTO는 Doc Deadline vs Land Permit By 중 빠른 날짜를 사용",
        "- Due Offset Days로 +/- 조정 가능",
        "",
        "⚠️ Excel 로케일에서 수식 구분자가 ';'인 경우 일부 IF 수식이 오류일 수 있습니다.",
        "   (1) Windows '목록 구분 기호'를 ','로 설정하거나",
        "   (2) --refresh 모드에서 Python이 값을 다시 계산하도록 운영하세요.",
    ]
    for i, line in enumerate(instructions, start=3):
        ws_ins[f"A{i}"] = line
        ws_ins[f"A{i}"].alignment = align_left
        ws_ins[f"A{i}"].font = font_normal
    ws_ins.column_dimensions["A"].width = 120

    # =========================
    # VBA_Pasteboard (copy/paste)
    # =========================
    if "VBA_Pasteboard" in wb.sheetnames:
        ws_vba = wb["VBA_Pasteboard"]
        row = ws_vba.max_row + 2
    else:
        ws_vba = wb.create_sheet("VBA_Pasteboard")
        ws_vba["A1"] = "VBA Code Repository (Copy & Paste into VBA Editor)"
        ws_vba["A1"].font = font_title
        row = 3

    start_row = row
    ws_vba.cell(row, 1, value="=" * 50)
    row += 1
    ws_vba.cell(row, 1, value="TR Document Tracker Macros")
    row += 1
    ws_vba.cell(row, 1, value="=" * 50)
    row += 1
    ws_vba.cell(row, 1, value="")
    row += 1

    vba_text = '''Option Explicit

' ============================
' Module: modTRDocTracker
' ============================

Private Function GetPythonExe() As String
    ' Prefer: "py" launcher (Windows)
    GetPythonExe = "py"
End Function

Private Function GetScriptPath() As String
    Dim wbPath As String
    wbPath = ThisWorkbook.Path
    GetScriptPath = wbPath & Application.PathSeparator & ".." & Application.PathSeparator & _
                    "01_Python_Builders" & Application.PathSeparator & "create_tr_document_tracker_v2.py"
End Function

Public Sub TR_Refresh_Document_Tracker()
    On Error GoTo EH

    Dim py As String: py = GetPythonExe()
    Dim script As String: script = GetScriptPath()
    Dim wbFile As String: wbFile = ThisWorkbook.FullName

    Dim cmd As String
    cmd = py & " " & """" & script & """" & " --refresh " & """" & wbFile & """"

    Shell cmd, vbHide
    MsgBox "Refresh started. Re-open workbook after script completes.", vbInformation
    Exit Sub

EH:
    MsgBox "Refresh failed: " & Err.Description, vbCritical
End Sub
'''

    for line in vba_text.splitlines():
        ws_vba.cell(row, 1, value=line)
        row += 1

    ws_vba.column_dimensions["A"].width = 120
    for row_num in range(start_row, ws_vba.max_row + 1):
        cell = ws_vba.cell(row_num, 1)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name="Consolas", size=10)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# =========================
# Refresh Logic
# =========================
def refresh_workbook(path: Path) -> None:
    """Rebuild tracker rows from Voyage_Schedule x Doc_Matrix, preserving user fields.

    Preserved (Document_Tracker):
      - Status, Submitted Date, Approved/Received Date
      - Remarks, File Link, Revision, Transmittal No
      - Last Updated, Updated By
    """
    wb = load_workbook(path)
    ws_voy = wb["Voyage_Schedule"]
    ws_doc = wb["Doc_Matrix"]
    ws_tr = wb["Document_Tracker"]

    # Styles (same look as template)
    thin = Side(style="thin", color="BFBFBF")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_normal = Font(size=10, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    voy_header_row = 4
    doc_header_row = 4
    tr_header_row = 4

    # Update timestamp
    ws_tr["A2"] = f"Last Refreshed: {now_str()}  |  목적: Responsible Party별 서류 제출일자/데드라인/현황 공유 → 누락 방지"

    # ---- Read voyages
    voyages: List[str] = []
    r = voy_header_row + 1
    while True:
        v = ws_voy.cell(row=r, column=1).value
        if not v:
            break
        voyages.append(str(v))
        r += 1

    # ---- Read docs (Required=Y)
    docs: List[Dict[str, Any]] = []
    r = doc_header_row + 1
    while True:
        code = ws_doc.cell(row=r, column=1).value
        if not code:
            break
        req = str(ws_doc.cell(row=r, column=10).value or "").strip().upper()
        if req == "Y":
            docs.append({
                "Doc Code": ws_doc.cell(row=r, column=1).value,
                "Category": ws_doc.cell(row=r, column=2).value,
                "Document Name": ws_doc.cell(row=r, column=3).value,
                "Description": ws_doc.cell(row=r, column=4).value,
                "Priority": ws_doc.cell(row=r, column=5).value,
                "Responsible Party": ws_doc.cell(row=r, column=6).value,
                "Lead Days": ws_doc.cell(row=r, column=7).value,
                "Due Based On": ws_doc.cell(row=r, column=8).value,
                "Due Offset Days": ws_doc.cell(row=r, column=9).value,
            })
        r += 1

    # ---- Preserve user inputs by key
    preserve_cols = list(range(22, 31))
    preserved: Dict[str, List[Any]] = {}
    r = tr_header_row + 1
    while True:
        key = ws_tr.cell(row=r, column=1).value
        if not key:
            break
        preserved[str(key)] = [ws_tr.cell(row=r, column=c).value for c in preserve_cols]
        r += 1

    # ---- Clear old rows (keep header area)
    if ws_tr.max_row > tr_header_row:
        ws_tr.delete_rows(tr_header_row + 1, ws_tr.max_row - tr_header_row)

    # ---- Rebuild
    row_cursor = tr_header_row + 1
    for v in voyages:
        for d in docs:
            key = f"{v}|{d['Doc Code']}"
            ws_tr.cell(row=row_cursor, column=1, value=key)
            ws_tr.cell(row=row_cursor, column=2, value=v)

            # Lookup Site/TR Units from Voyage_Schedule (tblVoyage)
            ws_tr.cell(row=row_cursor, column=3, value=f"=IFERROR(INDEX(tblVoyage[Site],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=4, value=f"=IFERROR(INDEX(tblVoyage[TR Units],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")

            # Milestones/deadlines from tblVoyage
            ws_tr.cell(row=row_cursor, column=5, value=f"=IFERROR(INDEX(tblVoyage[MZP Arrival],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=6, value=f"=IFERROR(INDEX(tblVoyage[Load-out],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=7, value=f"=IFERROR(INDEX(tblVoyage[MZP Departure],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=8, value=f"=IFERROR(INDEX(tblVoyage[AGI Arrival],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=9, value=f"=IFERROR(INDEX(tblVoyage[Doc Deadline],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")
            ws_tr.cell(row=row_cursor, column=10, value=f"=IFERROR(INDEX(tblVoyage[Land Permit By],MATCH($B{row_cursor},tblVoyage[Voyage],0)),\"\")")

            # Doc attributes
            ws_tr.cell(row=row_cursor, column=11, value=d["Doc Code"])
            ws_tr.cell(row=row_cursor, column=12, value=d["Category"])
            ws_tr.cell(row=row_cursor, column=13, value=d["Document Name"])
            ws_tr.cell(row=row_cursor, column=14, value=d["Description"])
            ws_tr.cell(row=row_cursor, column=15, value=d["Priority"])
            ws_tr.cell(row=row_cursor, column=16, value=d["Responsible Party"])
            ws_tr.cell(row=row_cursor, column=17, value=d["Lead Days"])
            ws_tr.cell(row=row_cursor, column=18, value=d["Due Based On"])
            ws_tr.cell(row=row_cursor, column=19, value=d["Due Offset Days"])

            # SUBMISSION DATE + Start By
            ws_tr.cell(row=row_cursor, column=20, value=(
                f"=IF(R{row_cursor}=\"Doc Deadline\",I{row_cursor},"
                f"IF(R{row_cursor}=\"Land Permit By\",J{row_cursor},"
                f"IF(R{row_cursor}=\"MZP Arrival\",E{row_cursor},"
                f"IF(R{row_cursor}=\"Load-out\",F{row_cursor},"
                f"IF(R{row_cursor}=\"MZP Departure\",G{row_cursor},"
                f"IF(R{row_cursor}=\"AGI Arrival\",H{row_cursor},"
                f"IF(AND(I{row_cursor}<>\"\",J{row_cursor}<>\"\"),MIN(I{row_cursor},J{row_cursor}),IF(I{row_cursor}<>\"\",I{row_cursor},J{row_cursor}))"
                f"))))))"
                f"+S{row_cursor}"
            ))
            ws_tr.cell(row=row_cursor, column=21, value=f"=IF(T{row_cursor}=\"\",\"\",T{row_cursor}-Q{row_cursor})")

            # Restore preserved user inputs
            if key in preserved:
                vals = preserved[key]
                for idx, c in enumerate(preserve_cols):
                    ws_tr.cell(row=row_cursor, column=c, value=vals[idx])
            else:
                ws_tr.cell(row=row_cursor, column=22, value="Not Started")

            # Styling for the entire row
            for cc in range(1, 31):
                cell = ws_tr.cell(row=row_cursor, column=cc)
                cell.border = border_thin
                cell.font = font_normal
                cell.alignment = align_left if cc in (14, 25, 26) else align_center

            for cc in (5, 6, 7, 8, 9, 10, 20, 21, 23, 24, 29):
                ws_tr.cell(row=row_cursor, column=cc).number_format = "yyyy-mm-dd"

            row_cursor += 1

    # ---- Update table range
    tbl = ws_tr._tables.get("tblTracker")
    if tbl is not None:
        tbl.ref = f"A{tr_header_row}:AD{row_cursor-1}"

    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", type=str, default="TR_Document_Tracker_READY.xlsx", help="Output xlsx path (CREATE mode)")
    ap.add_argument("--refresh", type=str, default="", help="Existing workbook path to refresh (REFRESH mode)")
    args = ap.parse_args()

    if args.refresh:
        refresh_workbook(Path(args.refresh))
        return

    build_template(Path(args.output))


if __name__ == "__main__":
    main()
