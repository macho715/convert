# Py 3.11.8 / pandas 2.33 / openpyxl 3.1.5
"""
AGI TR Schedule Runner
- Update Schedule_Data dates using D0 and Offset/Duration
- Rebuild Gantt_Chart bars (fills)
- Validate schedule and write a simple report (LOG sheet + jsonl)

Usage:
  py agi_tr_runner.py --in "AGI_TR6_VBA_Enhanced_AUTOMATION.xlsx" --out "C:\Temp\AGI_TR_Output" --mode update
  py agi_tr_runner.py --in "..." --mode validate
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
from pathlib import Path
from typing import Dict, Tuple, List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

CTRL_SHEET = "Control_Panel"
DATA_SHEET = "Schedule_Data"
GANTT_SHEET = "Gantt_Chart"
LOG_SHEET = "LOG"
SETTINGS_SHEET = "SETTINGS"

D0_CELL = "C5"  # Control_Panel D0 input
DATA_START_ROW = 6
GANTT_DATE_START_COL = 9  # I
GANTT_HEADER_ROW = 4
GANTT_FIRST_DATA_ROW = 5

# Timeline base (matches current workbook layout: Jan 4 - Feb 28, 2026)
TIMELINE_START = dt.date(2026, 1, 4)
TIMELINE_DAYS = 56  # 28 + 28

PHASE_FILL = {
    "MOBILIZATION": "4472C4",
    "DECK_PREP": "00B0F0",
    "LOADOUT": "4472C4",
    "SAIL": "70AD47",
    "AGI_UNLOAD": "ED7D31",
    "RETURN": "92D050",
    "TURNING": "FFC000",
    "JACKDOWN": "7030A0",
    "MILESTONE": "C00000",
    # Others (SEAFAST, BUFFER, etc.) intentionally left blank
}

def find_input_file(in_path: str) -> Path:
    """Find input file, preferring READY.xlsm when available."""
    path = Path(in_path)
    if path.exists():
        if path.suffix.lower() in (".xlsx", ".xlsm") and not path.stem.upper().endswith("_READY"):
            ready_path = path.with_name(f"{path.stem}_READY.xlsm")
            if ready_path.exists():
                print(f"Using READY version: {ready_path}")
                return ready_path
        return path
    ready_fallback = path.parent / f"{path.stem}_READY.xlsm"
    if ready_fallback.exists():
        print(f"Using READY version: {ready_fallback}")
        return ready_fallback
    return path

def fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)

def to_date(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    # Excel might store date as string
    try:
        return dt.datetime.fromisoformat(str(v)).date()
    except Exception:
        return None

def now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")

def log_jsonl(path: Path, event: str, **kv) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rec = {"ts": now_iso(), "event": event, **kv}
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")

def append_log_sheet(wb, level: str, proc: str, msg: str, details: str = "") -> None:
    if LOG_SHEET not in wb.sheetnames:
        return
    ws = wb[LOG_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, now_iso())
    ws.cell(r, 2, level)
    ws.cell(r, 3, proc)
    ws.cell(r, 4, msg)
    ws.cell(r, 5, details)
    ws.cell(r, 6, os.getenv("USERNAME", ""))

def update_schedule_dates(wb, d0: dt.date) -> Tuple[int, List[str]]:
    ws = wb[DATA_SHEET]
    last_row = ws.max_row
    updated = 0
    issues: List[str] = []

    for r in range(DATA_START_ROW, last_row + 1):
        task_id = ws.cell(r, 1).value
        if task_id in (None, ""):
            continue
        try:
            offset = float(ws.cell(r, 6).value or 0)   # F
            duration = float(ws.cell(r, 9).value or 0) # I
        except Exception:
            issues.append(f"Row {r}: offset/duration parse fail")
            continue

        start = d0 + dt.timedelta(days=offset)
        if duration > 0:
            end = start + dt.timedelta(days=duration - 1)
        else:
            end = start

        ws.cell(r, 7, dt.datetime.combine(start, dt.time()))  # G
        ws.cell(r, 8, dt.datetime.combine(end, dt.time()))    # H
        updated += 1

    append_log_sheet(wb, "INFO", "update_schedule_dates", f"Updated {updated} tasks", "")
    return updated, issues

def clear_gantt_bars(ws) -> None:
    # clear fills and values only in the bar area (not header)
    for r in range(GANTT_FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(GANTT_DATE_START_COL, GANTT_DATE_START_COL + TIMELINE_DAYS):
            cell = ws.cell(r, c)
            # keep Shamal header row intact (we clear only data rows)
            cell.value = None
            cell.fill = PatternFill()  # reset

def date_to_col(d: dt.date) -> int:
    delta = (d - TIMELINE_START).days
    return GANTT_DATE_START_COL + delta

def rebuild_gantt(wb) -> Tuple[int, List[str]]:
    ws_data = wb[DATA_SHEET]
    ws_g = wb[GANTT_SHEET]
    last_row = ws_data.max_row
    filled = 0
    issues: List[str] = []

    clear_gantt_bars(ws_g)

    for r in range(DATA_START_ROW, last_row + 1):
        task_id = ws_data.cell(r, 1).value
        if task_id in (None, ""):
            continue

        # Copy meta columns A-H into gantt row (same row index)
        for c in range(1, 9):
            ws_g.cell(r-1, c, ws_data.cell(r, c).value)  # data starts row6, gantt starts row5

        start = to_date(ws_data.cell(r, 7).value)
        end = to_date(ws_data.cell(r, 8).value)
        phase = str(ws_data.cell(r, 4).value or "").strip()

        if not start:
            issues.append(f"{task_id}: missing start")
            continue
        if not end:
            end = start

        # Determine fill length: ceil(duration), minimum 1
        try:
            dur = float(ws_data.cell(r, 9).value or 0)
        except Exception:
            dur = 0.0
        days = max(1, int(math.ceil(dur))) if dur > 0 else 1

        # For milestones, fill only start day, and mark with symbol
        is_milestone = phase.upper() in ("MILESTONE", "JACKDOWN")

        fill_rgb = PHASE_FILL.get(phase.upper())
        if not fill_rgb:
            # no bars for BUFFER/SEAFAST/etc, but keep milestones
            fill_rgb = None

        for k in range(days):
            d = start + dt.timedelta(days=k)
            if d < TIMELINE_START or d >= TIMELINE_START + dt.timedelta(days=TIMELINE_DAYS):
                continue
            col = date_to_col(d)
            cell = ws_g.cell(r-1, col)
            if fill_rgb:
                cell.fill = fill(fill_rgb)
            if is_milestone and k == 0:
                cell.value = "★" if phase.upper() == "JACKDOWN" else "▶"
                cell.font = Font(bold=True, color="FFFFFF" if phase.upper() == "MILESTONE" else "000000")
            filled += 1

    append_log_sheet(wb, "INFO", "rebuild_gantt", f"Filled {filled} cells", "")
    return filled, issues

def validate_schedule(wb, d0: dt.date) -> List[Dict]:
    ws = wb[DATA_SHEET]
    last_row = ws.max_row
    problems: List[Dict] = []

    for r in range(DATA_START_ROW, last_row + 1):
        task_id = ws.cell(r, 1).value
        if task_id in (None, ""):
            continue
        start = to_date(ws.cell(r, 7).value)
        end = to_date(ws.cell(r, 8).value)
        if not start:
            problems.append({"row": r, "id": task_id, "issue": "Start missing"})
            continue
        if not end:
            problems.append({"row": r, "id": task_id, "issue": "End missing"})
            continue
        # end can be before start due to duration<1; accept if duration<1
        try:
            dur = float(ws.cell(r, 9).value or 0)
        except Exception:
            dur = None
        if dur is not None and dur >= 1 and end < start:
            problems.append({"row": r, "id": task_id, "issue": "End < Start (dur>=1)"})

    append_log_sheet(wb, "INFO", "validate_schedule", f"Problems {len(problems)}", "")
    return problems

def read_settings(wb) -> Dict[str, str]:
    if SETTINGS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SETTINGS_SHEET]
    out = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k:
            out[str(k).strip()] = "" if v is None else str(v)
    return out

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", required=True)
    ap.add_argument("--out", dest="out_dir", default="")
    ap.add_argument("--mode", choices=["update", "validate"], default="update")
    ap.add_argument("--d0", default="")  # YYYY-MM-DD override
    args = ap.parse_args()

    in_path = find_input_file(args.in_path)
    if not in_path.exists():
        raise FileNotFoundError(in_path)

    wb = load_workbook(in_path)
    settings = read_settings(wb)

    # Determine D0
    if args.d0:
        d0 = dt.date.fromisoformat(args.d0)
    else:
        d0 = to_date(wb[CTRL_SHEET][D0_CELL].value) or dt.date.today()

    # Log path
    log_file = Path(settings.get("LOG_FILE", str(in_path.with_suffix(".jsonl"))))
    log_jsonl(log_file, "start", mode=args.mode, d0=str(d0), in_path=str(in_path))

    try:
        if args.mode == "update":
            update_schedule_dates(wb, d0)
            rebuild_gantt(wb)
            probs = validate_schedule(wb, d0)
            log_jsonl(log_file, "validate", problems=len(probs))

            out_dir = Path(args.out_dir or settings.get("OUT_DIR", str(in_path.parent)))
            out_dir.mkdir(parents=True, exist_ok=True)
            out_path = out_dir / f"{in_path.stem}_OUT_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(out_path)
            log_jsonl(log_file, "done", out_path=str(out_path))
            print(f"OK: saved -> {out_path}")
        else:
            probs = validate_schedule(wb, d0)
            print(json.dumps(probs, ensure_ascii=False, indent=2))
            log_jsonl(log_file, "validate_only", problems=len(probs))

    except Exception as e:
        append_log_sheet(wb, "ERROR", "main", str(e), "")
        log_jsonl(log_file, "error", err=str(e))
        raise
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
