#!/usr/bin/env python3
"""
AGI TR 7-Voyage Master Gantt with VBA
기존 AGI_TR_7Voyage_Master_Gantt.xlsx와 동일한 레이아웃 + VBA 기능
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# === COLORS (기존과 동일) ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def load_tasks_from_tsv(tsv_path, project_start_date):
    """
    TSV 파일에서 작업 목록을 읽어서 tasks 리스트로 변환합니다.

    지원 포맷(헤더 기반 자동 인식):
      - 형식1: ID, WBS, Task, Phase, Owner, Start, End, Duration_days, Notes
      - 형식2: Task, Owner, Start, End, Dur, Notes (ID/WBS/Phase 자동 추론)
      - 형식3(P6): Activity ID, Activity Name, Original Duration, Planned Start, Planned Finish, Actual Start, Actual Finish

    주의:
      - 일부 P6 TSV는 그룹/요약 행에서 'Activity Name' 컬럼이 누락되어 열이 1칸씩 당겨지는 경우가 있습니다.
        간단한 휴리스틱으로 복구합니다(예: Activity Name=숫자, Original Duration=날짜 형태).
      - 프로젝트 시작일(project_start_date)은 offset(일수) 계산 기준이며, Control_Panel의 PROJECT_START와 연결됩니다.
    """
    import csv
    import math
    import re
    from datetime import datetime

    tasks = []

    if hasattr(project_start_date, "date"):
        project_start_date = project_start_date.date()

    # === Helpers ===
    def _is_number_like(value: str) -> bool:
        v = (value or "").strip()
        return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", v))

    def _parse_date(value: object):
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value
        s = str(value).strip()
        if not s:
            return None

        # 1) ISO: 2026-01-15
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass

        # 2) 12-Jan-2026 / 06-Jan-26
        for fmt in ("%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass

        # 3) 12-Jan  (year inferred from project_start_date)
        try:
            d = datetime.strptime(s, "%d-%b").date()
            return d.replace(year=project_start_date.year)
        except Exception:
            return None

    def _infer_phase_from_task(task_name: str) -> str:
        task_upper = (task_name or "").upper()
        if "MOBILIZATION" in task_upper or "DEMOBILIZATION" in task_upper or "DEMOB" in task_upper:
            return "MOBILIZATION"
        if "DECK" in task_upper and ("PREP" in task_upper or "PREPARATION" in task_upper):
            return "DECK_PREP"
        if "LOADOUT" in task_upper or "LOAD-OUT" in task_upper or "LOAD OUT" in task_upper or "LOAD-IN" in task_upper or "LOAD IN" in task_upper:
            return "LOADOUT"
        if "SEA FASTENING" in task_upper or "SEAFAST" in task_upper or "SEAFASTENING" in task_upper or ("SEA" in task_upper and "FAST" in task_upper):
            return "SEAFAST"
        if "SAIL" in task_upper or "SEA PASSAGE" in task_upper or "SAIL-AWAY" in task_upper:
            return "SAIL"
        if "UNLOAD" in task_upper or "ARRIVAL" in task_upper or "RORO" in task_upper:
            return "AGI_UNLOAD"
        if "TURNING" in task_upper or ("TURN" in task_upper and "RETURN" not in task_upper):
            return "TURNING"
        if "JACKDOWN" in task_upper or "JACK-DOWN" in task_upper or "JACKING DOWN" in task_upper:
            return "JACKDOWN"
        if "RETURN" in task_upper:
            return "RETURN"
        if "VOYAGE" in task_upper or "MILESTONE" in task_upper:
            return "MILESTONE"
        if "BUFFER" in task_upper or "RESET" in task_upper:
            return "BUFFER"
        return "BUFFER"

    phase_mapping = {
        "Mobilization": "MOBILIZATION",
        "Deck Prep": "DECK_PREP",
        "MZP Loadout": "LOADOUT",
        "Sea Fastening": "SEAFAST",
        "Survey": "BUFFER",
        "Sea Passage": "SAIL",
        "AGI Arrival": "AGI_UNLOAD",
        "AGI Laydown": "BUFFER",
        "Onshore SPMT": "TURNING",
        "AGI Gate Prep": "TURNING",
        "Jackdown": "JACKDOWN",
        "Return": "RETURN",
        "Buffer": "BUFFER",
        "Marine Transport": "MILESTONE",
        "Demobilization": "MOBILIZATION",
        "Handover": "MILESTONE",
    }

    def _duration_ref(duration_str: str, task_name: str):
        s = (duration_str or "").strip()
        if not s:
            return "DUR_BUF"
        try:
            dur_val = float(s)
        except Exception:
            return "DUR_BUF"

        if dur_val <= 0:
            return 0

        task_upper = (task_name or "").upper()

        # 0.5 day: Sea-fastening / MWS / Buffer
        if math.isclose(dur_val, 0.5, rel_tol=0.0, abs_tol=1e-9):
            if ("SEA" in task_upper and "FAST" in task_upper) or "SEAFAST" in task_upper:
                return "DUR_SF"
            if "MWS" in task_upper or "MPI" in task_upper or "WARRANTY" in task_upper or "APPROVAL" in task_upper:
                return "DUR_MWS"
            return "DUR_BUF"

        # 1.0 day: map by task keywords
        if math.isclose(dur_val, 1.0, rel_tol=0.0, abs_tol=1e-9):
            if "LOADOUT" in task_upper or "LOAD-OUT" in task_upper or "LOAD OUT" in task_upper or "LOAD-IN" in task_upper or "LOAD IN" in task_upper:
                return "DUR_LO"
            if "MOBILIZATION" in task_upper or "DEMOBILIZATION" in task_upper or "DEMOB" in task_upper:
                return "DUR_MOB"
            if "SAIL" in task_upper or "SEA PASSAGE" in task_upper:
                return "DUR_SAIL"
            if "ARRIV" in task_upper or "UNLOAD" in task_upper or "RORO" in task_upper:
                return "DUR_UL"
            if "RETURN" in task_upper:
                return "DUR_RET"
            if "JACKDOWN" in task_upper or "JACK-DOWN" in task_upper or "JACKING DOWN" in task_upper:
                return "DUR_JD"
            return "DUR_BUF"

        # 3.0 day: turning or deck prep
        if math.isclose(dur_val, 3.0, rel_tol=0.0, abs_tol=1e-9):
            if "TURN" in task_upper:
                return "DUR_TURN"
            if "DECK" in task_upper or "PREP" in task_upper:
                return "DUR_DECK"
            return 3

        # Default: keep numeric (including fractions like 0.2, 0.3)
        if dur_val.is_integer():
            return int(dur_val)
        return float(dur_val)

    # ID 자동 생성 카운터
    task_counter = 0
    voyage_num = 0

    expected_p6_headers = [
        "Activity ID",
        "Activity Name",
        "Original Duration",
        "Planned Start",
        "Planned Finish",
        "Actual Start",
        "Actual Finish",
    ]

    def _open_with_fallback(path):
        for enc in ("utf-8-sig", "utf-8"):
            try:
                return open(path, "r", encoding=enc, newline="")
            except UnicodeDecodeError:
                continue
        return open(path, "r", encoding="utf-8", newline="", errors="replace")

    with _open_with_fallback(tsv_path) as f:
        reader = csv.reader(f, delimiter="\t")
        headers = next(reader, None)
        if not headers:
            return []

        headers = [h.strip().lstrip("\ufeff") for h in headers]
        header_set = {h.lower() for h in headers}

        # column name accessor (case-insensitive)
        def _get(row, key, default=""):
            # exact match
            if key in row:
                return row.get(key, default)
            # case-insensitive lookup
            k_lower = key.lower()
            for k in row.keys():
                if k.lower() == k_lower:
                    return row.get(k, default)
            return default

        for values in reader:
            if not values or not any((v or "").strip() for v in values):
                continue

            # Repair: P6 TSV group/summary row with missing Activity Name column
            if headers[:7] == expected_p6_headers and len(values) == len(headers) - 1:
                # Example bad row:
                #   MOBILIZATION   1.00   06-Jan-26  06-Jan-26
                # should be:
                #   MOBILIZATION   (blank) 1.00       06-Jan-26  06-Jan-26
                if len(values) >= 4 and _is_number_like(values[1]) and _parse_date(values[2]) and _parse_date(values[3]):
                    values = [values[0], ""] + values[1:]

            if len(values) < len(headers):
                values = values + [""] * (len(headers) - len(values))
            if len(values) > len(headers):
                values = values[: len(headers)]

            row = {headers[i]: (values[i] or "").strip() for i in range(len(headers))}

            # Start 날짜 확인 (필수)
            start_str = (
                _get(row, "Planned Start").strip()
                or _get(row, "Actual Start").strip()
                or _get(row, "Start").strip()
            )
            if not start_str:
                continue

            start_date = _parse_date(start_str)
            if not start_date:
                continue

            # End(선택)
            end_str = (
                _get(row, "Planned Finish").strip()
                or _get(row, "Actual Finish").strip()
                or _get(row, "End").strip()
            )
            end_date = _parse_date(end_str) if end_str else None

            # Task 이름
            task_name = (_get(row, "Task").strip() or _get(row, "Activity Name").strip())
            # 휴리스틱: Activity Name이 숫자이고, Original Duration이 날짜처럼 보이면 'Activity Name 누락'으로 간주
            if task_name and _is_number_like(task_name):
                orig_dur_str = _get(row, "Original Duration").strip()
                if _parse_date(orig_dur_str):
                    task_name = ""

            if not task_name:
                task_name = (_get(row, "Activity ID").strip() or _get(row, "ID").strip())
            if not task_name:
                continue

            # ID 처리 (없으면 자동 생성)
            task_id = (_get(row, "Activity ID").strip() or _get(row, "ID").strip())
            if not task_id:
                if "VOYAGE" in task_name.upper():
                    voyage_num += 1
                    task_id = f"V{voyage_num}"
                elif "MOBILIZATION" in task_name.upper():
                    task_id = "MOB-001"
                elif "DECK" in task_name.upper() and "PREP" in task_name.upper():
                    task_id = "PREP-001"
                else:
                    task_counter += 1
                    words = task_name.split()
                    if words:
                        prefix = re.sub(r"[^A-Z0-9]", "", words[0][:3].upper())
                        task_id = f"{prefix or 'TSK'}-{task_counter:03d}"
                    else:
                        task_id = f"TASK-{task_counter:03d}"

            # WBS 처리 (없으면 자동 생성)
            wbs_raw = _get(row, "WBS").strip()
            if not wbs_raw:
                wbs = f"A{voyage_num}" if voyage_num > 0 else "A0"
            else:
                wbs = wbs_raw

            # Phase 처리 (없으면 Task 이름에서 추론)
            phase_raw = _get(row, "Phase").strip()
            if phase_raw:
                phase = phase_mapping.get(phase_raw, _infer_phase_from_task(task_name))
            else:
                phase = _infer_phase_from_task(task_name)

            # Duration 처리: End 날짜가 있으면 End - Start로 계산, 없으면 Duration 컬럼 사용
            if end_date:
                try:
                    duration_str = str((end_date - start_date).days)
                except Exception:
                    duration_str = ""
            else:
                duration_str = (
                    _get(row, "Original Duration").strip()
                    or _get(row, "Duration_days").strip()
                    or _get(row, "Dur").strip()
                    or _get(row, "Duration").strip()
                )

            dur_ref = _duration_ref(duration_str, task_name)

            # Owner
            owner = _get(row, "Owner", "All").strip() or "All"

            # Notes
            notes = _get(row, "Notes").strip()

            # Offset 계산 (Start - project_start_date)
            offset = (start_date - project_start_date).days

            tasks.append((task_id, wbs, task_name, phase, owner, offset, dur_ref, notes))

    return tasks

def extract_scenario_dates(tsv_path, default_year=None):
    """
    Extract earliest start and latest end dates from a TSV file.
    Returns (start_date, end_date) or (None, None).
    """
    import csv
    import os
    import re
    from datetime import datetime

    if not tsv_path or not os.path.exists(tsv_path):
        return None, None

    expected_p6_headers = [
        "Activity ID",
        "Activity Name",
        "Original Duration",
        "Planned Start",
        "Planned Finish",
        "Actual Start",
        "Actual Finish",
    ]

    def _try_parse(value: str):
        s = (value or "").strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y", "%d-%b"):
            try:
                d = datetime.strptime(s, fmt).date()
                if fmt == "%d-%b" and default_year:
                    d = d.replace(year=default_year)
                return d
            except Exception:
                continue
        return None

    try:
        with open(tsv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter="\t")
            headers = next(reader, None)
            if not headers:
                return None, None
            headers = [h.strip().lstrip("\ufeff") for h in headers]

            start_idx = [i for i, h in enumerate(headers) if h.lower() in {"start", "planned start", "actual start"}]
            end_idx = [i for i, h in enumerate(headers) if h.lower() in {"end", "finish", "planned finish", "actual finish"}]
            if not start_idx:
                start_idx = [i for i, h in enumerate(headers) if "start" in h.lower()]
            if not end_idx:
                end_idx = [i for i, h in enumerate(headers) if "finish" in h.lower() or "end" in h.lower()]

            earliest = None
            latest = None

            for values in reader:
                if not values or not any((v or "").strip() for v in values):
                    continue

                if headers[:7] == expected_p6_headers and len(values) == len(headers) - 1:
                    if len(values) >= 4 and re.fullmatch(r"-?\d+(?:\.\d+)?", (values[1] or "").strip() or ""):
                        values = [values[0], ""] + values[1:]

                if len(values) < len(headers):
                    values = values + [""] * (len(headers) - len(values))
                if len(values) > len(headers):
                    values = values[: len(headers)]

                for idx in start_idx:
                    if idx < len(values):
                        d = _try_parse(values[idx])
                        if d:
                            earliest = d if earliest is None else min(earliest, d)
                for idx in end_idx:
                    if idx < len(values):
                        d = _try_parse(values[idx])
                        if d:
                            latest = d if latest is None else max(latest, d)

            return earliest, latest
    except Exception as e:
        print(f"Error extracting dates from {tsv_path}: {e}")
        return None, None

def load_tide_data_json(json_path):
    """
    조석 데이터 JSON 파일 로드
    """
    import json
    
    tide_records = []
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if 'tide_records' in data:
                for record in data['tide_records']:
                    date_str = record.get('date', '')
                    if not date_str:
                        continue
                    tide_records.append((
                        date_str,
                        record.get('high_tide_window', '').strip(),
                        float(record.get('max_height_m', 0)) if record.get('max_height_m') else 0.0,
                        record.get('risk_level', 'LOW').strip()
                    ))
    except Exception as e:
        print(f"Warning: Could not load tide data from JSON: {e}")
    
    return tide_records

def load_tide_data(tsv_path=None, json_path=None):
    """
    조석 데이터 로드 (TSV 또는 JSON 지원)
    """
    import csv
    import os
    
    if json_path and os.path.exists(json_path):
        return load_tide_data_json(json_path)
    
    tide_records = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            with open(tsv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    if not row.get('Date'):
                        continue
                    try:
                        date_str = row['Date'].strip()
                        tide_records.append((
                            date_str,
                            row.get('High Tide Window', '').strip(),
                            float(row.get('Max Height (m)', '0').strip()) if row.get('Max Height (m)') else 0.0,
                            row.get('Risk Level', 'LOW').strip()
                        ))
                    except:
                        continue
        except Exception as e:
            print(f"Warning: Could not load tide data from TSV: {e}")
    
    return tide_records

def calculate_max_days(tasks, project_start, wb=None, min_days=120, buffer_days=30):
    """
    작업 목록에서 최대 프로젝트 기간(일)을 계산합니다.

    - offset + duration(소수 포함) 기반으로 최대 종료 지점을 계산합니다.
    - wb(Control_Panel)가 주어지면 GANTT_MIN_DAYS(H8), GANTT_BUFFER_DAYS(H9) 값을 우선 적용합니다.
    """
    import math

    if hasattr(project_start, "date"):
        project_start = project_start.date()

    # Default duration lookup (Control_Panel이 없을 때 사용)
    default_duration_values = {
        "DUR_MOB": 1.0,
        "DUR_DECK": 3.0,
        "DUR_LO": 1.0,
        "DUR_SF": 0.5,
        "DUR_MWS": 0.5,
        "DUR_SAIL": 1.0,
        "DUR_UL": 1.0,
        "DUR_TURN": 3.0,
        "DUR_JD": 1.0,
        "DUR_RET": 1.0,
        "DUR_BUF": 0.5,
    }

    # Control_Panel 값 반영
    if wb:
        try:
            ws_ctrl = wb["Control_Panel"]
            # Min/Buffer days
            min_days_val = ws_ctrl["H8"].value
            buffer_days_val = ws_ctrl["H9"].value
            if isinstance(min_days_val, (int, float)):
                min_days = int(min_days_val)
            if isinstance(buffer_days_val, (int, float)):
                buffer_days = int(buffer_days_val)

            # Duration named ranges (E5~E15)
            # (기존 레이아웃 유지: create_gantt_with_vba의 durations list 기반)
            # DUR_MOB=E5 ... DUR_BUF=E15
            duration_cells = {
                "DUR_MOB": "E5",
                "DUR_DECK": "E6",
                "DUR_LO": "E7",
                "DUR_SF": "E8",
                "DUR_MWS": "E9",
                "DUR_SAIL": "E10",
                "DUR_UL": "E11",
                "DUR_TURN": "E12",
                "DUR_JD": "E13",
                "DUR_RET": "E14",
                "DUR_BUF": "E15",
            }
            for k, addr in duration_cells.items():
                v = ws_ctrl[addr].value
                if isinstance(v, (int, float)):
                    default_duration_values[k] = float(v)
        except Exception:
            pass

    def _duration_to_float(dur_ref) -> float:
        if isinstance(dur_ref, (int, float)):
            try:
                return max(0.0, float(dur_ref))
            except Exception:
                return 0.0
        if isinstance(dur_ref, str):
            key = dur_ref.strip().lstrip("=")
            return float(default_duration_values.get(key, 1.0))
        return 0.0

    max_end = 0.0
    for task in tasks:
        if len(task) < 7:
            continue
        offset = task[5] if isinstance(task[5], (int, float)) else 0
        dur_ref = task[6]
        duration = _duration_to_float(dur_ref)
        # 소수 duration이 있어도 Excel의 일자 열은 Day 단위이므로 ceiling으로 여유 확보
        max_end = max(max_end, float(offset) + float(duration))

    # 최소치 보장 + 버퍼 반영
    return max(int(min_days), int(math.ceil(max_end)) + int(buffer_days))

def calculate_voyage_ranges(tasks):
    """
    tasks에서 각 Voyage의 Day 범위를 계산합니다.

    Returns:
        voyage_ranges: [(voyage_name, start_day, end_day), ...]
    """
    import re

    voyage_ranges = []
    current_voyage = None
    voyage_start = None
    voyage_end = None

    voyage_id_re = re.compile(r"^V\d+$", re.IGNORECASE)

    for task in tasks:
        if len(task) < 6:
            continue

        tid, wbs, task_name, phase, owner, offset = task[:6]

        if isinstance(tid, str) and voyage_id_re.match(tid.strip()):
            # close previous
            if current_voyage is not None:
                voyage_ranges.append((current_voyage, voyage_start, voyage_end))
            current_voyage = task_name
            voyage_start = offset
            voyage_end = offset
            continue

        if current_voyage is not None and isinstance(offset, (int, float)):
            voyage_end = max(voyage_end, offset)

    if current_voyage is not None:
        voyage_ranges.append((current_voyage, voyage_start, voyage_end))

    # Mobilization(선택) - 첫 번째 태스크가 실제 Mobilization일 때만 삽입
    if tasks and len(tasks[0]) >= 6:
        first_phase = str(tasks[0][3]) if len(tasks[0]) >= 4 else ""
        if first_phase == "MOBILIZATION":
            mob_start = tasks[0][5] if isinstance(tasks[0][5], (int, float)) else 0
            mob_end = mob_start + 1
            voyage_ranges.insert(0, ("Mobilization", mob_start, mob_end))

    return voyage_ranges

def generate_weather_periods(project_start, project_end):
    """
    프로젝트 기간을 기반으로 Weather Analysis 헤더 구간 생성
    """
    periods = []
    if project_end <= project_start:
        period_name = f"{project_start.strftime('%b %d')}"
        return [(period_name, project_start, project_start)]
    current = project_start
    
    while current < project_end:
        period_start = current
        period_end = min(period_start + dt.timedelta(days=9), project_end)
        
        if period_start.month == period_end.month:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%d')}"
        else:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%b %d')}"
        
        periods.append((period_name, period_start, period_end))
        current = period_end + dt.timedelta(days=1)
    
    return periods

def parse_voyage_pattern(pattern_str):
    """
    Parse voyage pattern like "1x1x1x1x1x1x1", "1-2-2-2", or "2-2-2-1" into TR groups.
    """
    if not pattern_str:
        return [[i] for i in range(1, 8)]

    voyage_groups = []
    tr_id = 1

    if "x" in pattern_str:
        parts = pattern_str.split("x")
    elif "-" in pattern_str:
        parts = pattern_str.split("-")
    else:
        parts = [pattern_str]

    for part in parts:
        part = part.strip()
        if not part:
            continue
        try:
            count = int(part)
        except ValueError:
            continue
        if count <= 0:
            continue
        voyage_groups.append(list(range(tr_id, tr_id + count)))
        tr_id += count

    if not voyage_groups:
        return [[i] for i in range(1, 8)]

    return voyage_groups

def generate_scenario_tasks(pattern_str, project_start, cycle_spacing=15, early_return=False):
    """
    Generate scenario tasks from voyage pattern.

    Returns:
        tasks: list of (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    """
    tasks = []
    voyage_groups = parse_voyage_pattern(pattern_str)

    # MOB-001: Mobilization
    offset = 0
    tasks.append((
        "MOB-001", "A0", "Mobilization (crew/equipment)", "MOBILIZATION",
        "Mammoet", offset, "DUR_MOB", "SPMT + grillage in MZP"
    ))
    offset += 1

    # PREP-001: Deck Prep
    tasks.append((
        "PREP-001", "A0", "LCT deck preparations + fenders + mooring", "DECK_PREP",
        "Mammoet/KFS", offset, "DUR_DECK", "MWS pre-check ready"
    ))
    offset += 3  # DUR_DECK = 3 days

    cycle_offset = offset

    for v_idx, tr_list in enumerate(voyage_groups, start=1):
        n_units = len(tr_list)
        tr_str = f"TR{tr_list[0]}" if n_units == 1 else f"TR{tr_list[0]}-TR{tr_list[-1]}"
        offset = cycle_offset
        voyage_start = offset

        # V{number}: Voyage milestone
        wbs = f"A{v_idx}"
        tasks.append((
            f"V{v_idx}", wbs, f"VOYAGE {v_idx}: {tr_str} Transport", "MILESTONE",
            "SCT/Mammoet/KFS", offset, 0, "TIDE>=1.90 required (Loadout start)"
        ))

        # LO-{number}: Loadout (n_units days for batch)
        lo_duration = n_units if n_units > 1 else 1
        tasks.append((
            f"LO-{v_idx:02d}", wbs, f"Loadout {tr_str} onto LCT", "LOADOUT",
            "Mammoet", offset, lo_duration, "OK 2.01m (tide window required)"
        ))
        offset += lo_duration

        # SF-{number}: Sea Fastening
        tasks.append((
            f"SF-{v_idx:02d}", wbs, f"Sea fastening + MWS checks ({tr_str})", "SEAFAST",
            "Mammoet/KFS/MWS", offset, "DUR_SF", "Lashing + survey"
        ))
        offset += 1  # DUR_SF = 0.5 days, treated as 1 here

        # SAIL-{number}: Sail-away
        tasks.append((
            f"SAIL-{v_idx:02d}", wbs, "Sail-away MZP->AGI", "SAIL",
            "LCT", offset, "DUR_SAIL", "WX gate"
        ))
        offset += 1  # DUR_SAIL = 1 day

        # UL-{number}-{TR}: Unload (per unit)
        for tr_num in tr_list:
            tasks.append((
                f"UL-{v_idx:02d}-{tr_num}", wbs, f"Unload TR{tr_num} at AGI (1 unit/day)", "AGI_UNLOAD",
                "Mammoet", offset, "DUR_UL", "RORO + ramp"
            ))
            offset += 1  # DUR_UL = 1 day/unit

        first_jd_offset = None

        # TURN + JD per unit (interleaved)
        for tr_num in tr_list:
            tasks.append((
                f"TURN-{v_idx:02d}-{tr_num}", wbs, f"Turning TR{tr_num} (90 deg)", "TURNING",
                "Mammoet", offset, "DUR_TURN", "3.0d/unit"
            ))
            offset += 3  # DUR_TURN = 3 days/unit

            tasks.append((
                f"JD-{v_idx:02d}-{tr_num}", wbs, f"Jackdown TR{tr_num}", "JACKDOWN",
                "Mammoet", offset, "DUR_JD", "1.0d/unit"
            ))
            if first_jd_offset is None:
                first_jd_offset = offset
            offset += 1  # DUR_JD = 1 day/unit

        return_offset = offset
        buffer_offset = offset + 1
        return_note = "After final JD"

        if early_return and n_units > 1 and first_jd_offset is not None:
            return_offset = first_jd_offset + 1
            buffer_offset = return_offset + 1
            return_note = "After first JD"

        # RET-{number}: Return
        tasks.append((
            f"RET-{v_idx:02d}", wbs, "LCT Return AGI->MZP", "RETURN",
            "LCT", return_offset, "DUR_RET", return_note
        ))

        # BUF-{number}: Buffer
        tasks.append((
            f"BUF-{v_idx:02d}", wbs, "Buffer / reset", "BUFFER",
            "All", buffer_offset, "DUR_BUF", "contingency"
        ))

        cycle_days = buffer_offset - voyage_start + 1
        cycle_offset = buffer_offset + 1
        if v_idx < len(voyage_groups) and cycle_spacing > cycle_days:
            cycle_offset += (cycle_spacing - cycle_days)

    return tasks

def get_scenario_sheet_names(scenario_name):
    sheet_name_map = {
        "Mammoet_Original": "Mammoet_Orig",
        "Mammoet_ScenarioA": "Mammoet_ScnA",
        "Mammoet_Alternative": "Mammoet_Alt",
    }
    short_name = sheet_name_map.get(scenario_name, scenario_name)
    sched_name = f"Schedule_Data_{short_name}"
    gantt_name = f"Gantt_Chart_{short_name}"
    return short_name, sched_name[:31], gantt_name[:31]

def create_scenario_sheets(wb, scenario_name, tsv_path, project_start, pattern_str=None, early_return=False):
    """
    시나리오별 Schedule_Data와 Gantt_Chart 시트 생성
    """
    import os
    import math

    default_year = project_start.year if hasattr(project_start, "year") else None
    scenario_start, _ = extract_scenario_dates(tsv_path, default_year=default_year)
    if scenario_start:
        project_start = scenario_start
        print(f"Using scenario start date from TSV for {scenario_name}: {scenario_start}")

    if hasattr(project_start, "date"):
        project_start = project_start.date()

    short_name, sched_name, gantt_name = get_scenario_sheet_names(scenario_name)
    ws_sched = wb.create_sheet(sched_name)

    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = f"AGI TR Transportation - {scenario_name} Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_sched["A1"].alignment = Alignment(horizontal="center")

    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = f"Start = {project_start.isoformat()} | Auto-Updates from Control_Panel"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    # Control_Panel settings (Shamal/Tide threshold)
    try:
        ws_ctrl = wb["Control_Panel"]
        shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
        shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
        if hasattr(shamal_start, "date"):
            shamal_start = shamal_start.date()
        if hasattr(shamal_end, "date"):
            shamal_end = shamal_end.date()
        try:
            tide_threshold = float(ws_ctrl["H7"].value)
        except Exception:
            tide_threshold = 1.90
    except Exception:
        shamal_start = dt.date(2026, 1, 15)
        shamal_end = dt.date(2026, 4, 30)
        tide_threshold = 1.90

    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"

    ws_sched.merge_cells("A3:I3")
    ws_sched["A3"] = f"⚠️ Winter Shamal Risk Period: {shamal_text_full} | Tide ≥{tide_threshold:.2f}m + Weather Gate"
    ws_sched["A3"].font = Font(size=10, italic=True)
    ws_sched["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()

    tasks = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
            if tasks:
                print(f"Loaded {len(tasks)} tasks for {scenario_name} from TSV: {os.path.basename(tsv_path)}")
            else:
                print(f"Warning: No tasks loaded for {scenario_name} from TSV: {os.path.basename(tsv_path)}")
        except Exception as e:
            print(f"Error loading {scenario_name} TSV ({tsv_path}): {e}")

    if not tasks and pattern_str:
        try:
            tasks = generate_scenario_tasks(
                pattern_str,
                project_start,
                early_return=early_return,
            )
            print(f"Generated {scenario_name} tasks from pattern: {pattern_str} (early_return={early_return})")
        except Exception as e:
            print(f"Error generating {scenario_name} tasks: {e}")

    if not tasks:
        print(f"Warning: No tasks for {scenario_name}")
        return False

    def _is_valid_activity_id(task_id: str) -> bool:
        """A1000 형식만 허용: A로 시작하고 두 번째 문자가 숫자여야 함"""
        if not task_id:
            return False
        tid_upper = task_id.strip().upper()
        return len(tid_upper) >= 2 and tid_upper[0] == "A" and tid_upper[1].isdigit()

    filtered_tasks = [t for t in tasks if _is_valid_activity_id(t[0])]
    if not filtered_tasks:
        print(f"Warning: No valid A1000-format tasks found for {scenario_name} after filtering")
        return False
    if len(filtered_tasks) != len(tasks):
        print(f"Filtered {len(filtered_tasks)} valid tasks from {len(tasks)} total tasks for {scenario_name}")

    # Duration lookup for named durations (DUR_*)
    duration_values = {
        "DUR_MOB": 1.0,
        "DUR_DECK": 3.0,
        "DUR_LO": 1.0,
        "DUR_SF": 0.5,
        "DUR_MWS": 0.5,
        "DUR_SAIL": 1.0,
        "DUR_UL": 1.0,
        "DUR_TURN": 3.0,
        "DUR_JD": 1.0,
        "DUR_RET": 1.0,
        "DUR_BUF": 0.5,
    }
    try:
        ws_ctrl = wb["Control_Panel"]
        duration_cells = {
            "DUR_MOB": "E5",
            "DUR_DECK": "E6",
            "DUR_LO": "E7",
            "DUR_SF": "E8",
            "DUR_MWS": "E9",
            "DUR_SAIL": "E10",
            "DUR_UL": "E11",
            "DUR_TURN": "E12",
            "DUR_JD": "E13",
            "DUR_RET": "E14",
            "DUR_BUF": "E15",
        }
        for k, addr in duration_cells.items():
            v = ws_ctrl[addr].value
            if isinstance(v, (int, float)):
                duration_values[k] = float(v)
    except Exception:
        pass

    def _duration_to_float(dur_ref) -> float:
        if isinstance(dur_ref, (int, float)):
            try:
                return max(0.0, float(dur_ref))
            except Exception:
                return 0.0
        if isinstance(dur_ref, str):
            key = dur_ref.strip().lstrip("=")
            return float(duration_values.get(key, 1.0))
        return 0.0

    # === Schedule rows ===
    for r, t in enumerate(filtered_tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t

        ws_sched.cell(r, 1, value=tid)
        ws_sched.cell(r, 2, value=wbs)
        ws_sched.cell(r, 3, value=task)
        ws_sched.cell(r, 4, value=phase)
        ws_sched.cell(r, 5, value=owner)

        ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"

        if isinstance(dur_ref, str):
            ws_sched.cell(r, 8, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 8, value=dur_ref)

        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"
        ws_sched.cell(r, 9, value=notes)

        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 10):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)

        if phase == "MILESTONE":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")

    col_widths = {"A": 10, "B": 6, "C": 38, "D": 14, "E": 14, "F": 12, "G": 12, "H": 10, "I": 40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    ws_sched.auto_filter.ref = "A5:I5"

    # === Gantt ===
    ws_gantt = wb.create_sheet(gantt_name)
    max_days = calculate_max_days(filtered_tasks, project_start, wb=wb)

    ws_gantt.merge_cells(f"A1:{get_column_letter(7 + max_days)}1")
    ws_gantt["A1"] = f"AGI TR Transportation - {scenario_name} Gantt Chart"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_gantt["A1"].alignment = Alignment(horizontal="center")

    ws_gantt.merge_cells(f"A2:{get_column_letter(7 + max_days)}2")
    ws_gantt["A2"] = (
        f"Orange Zone = Winter Shamal Risk Period ({shamal_text}) | Tide ≥{tide_threshold:.2f}m + Weather Gate "
        f"| VBA: RefreshGanttChart_{scenario_name}로 색상 갱신"
    )
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    month_str = project_start.strftime("%b %Y")
    ws_gantt["A3"] = month_str
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:G3")

    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()

    date_col = 8
    for i in range(max_days):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5

    # Shamal 기간 헤더 색(초기값). 이후 기간 변경은 VBA 갱신 필요.
    for i in range(max_days):
        cell_date = project_start + dt.timedelta(days=i)
        if shamal_start <= cell_date <= shamal_end:
            ws_gantt.cell(4, date_col + i).fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    base_dt = dt.datetime.combine(project_start, dt.time.min)

    for r, t in enumerate(filtered_tasks, 5):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t

        # link to Schedule_Data (rows start at 6)
        ws_gantt.cell(r, 1, value=f"='{sched_name}'!A{r+1}")
        ws_gantt.cell(r, 2, value=f"='{sched_name}'!B{r+1}")
        ws_gantt.cell(r, 3, value=f"='{sched_name}'!C{r+1}")
        ws_gantt.cell(r, 4, value=f"='{sched_name}'!D{r+1}")

        start_cell = ws_gantt.cell(r, 5, value=f"='{sched_name}'!F{r+1}")
        start_cell.number_format = "MM/DD"

        end_cell = ws_gantt.cell(r, 6, value=f"='{sched_name}'!G{r+1}")
        end_cell.number_format = "MM/DD"

        ws_gantt.cell(r, 7, value=f"='{sched_name}'!H{r+1}")

        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()

        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)

        # --- Gantt fill (초기 렌더) ---
        duration_days = _duration_to_float(dur_ref)
        if duration_days < 0:
            duration_days = 0.0

        start_dt = base_dt + dt.timedelta(days=float(offset))
        end_dt = start_dt + dt.timedelta(days=float(duration_days))

        for i in range(max_days):
            cell_dt = base_dt + dt.timedelta(days=i)
            cell = ws_gantt.cell(r, date_col + i)
            cell.border = tb()

            if duration_days == 0:
                # milestone: show marker on start date
                if cell_dt.date() == start_dt.date():
                    cell.fill = PatternFill("solid", fgColor=pc)
                    cell.value = "★"
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=8)
                continue

            if start_dt <= cell_dt < end_dt:
                cell.fill = PatternFill("solid", fgColor=pc)

    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 28
    ws_gantt.column_dimensions["D"].width = 12
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4

    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)
    return True

def create_tide_data_sheet(wb, tide_tsv_path=None, tide_json_path=None):
    """
    조석 데이터 시트 생성
    """
    import os
    
    ws_tide = wb.create_sheet("Tide_Data")
    try:
        tide_threshold = float(wb["Control_Panel"]["H7"].value)
    except Exception:
        tide_threshold = 1.90
    
    ws_tide.merge_cells("A1:D1")
    ws_tide["A1"] = "MINA ZAYED PORT - High Tide Data"
    ws_tide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_tide["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_tide["A1"].alignment = Alignment(horizontal="center")
    
    ws_tide.merge_cells("A2:D2")
    ws_tide["A2"] = f"Tide ≥{tide_threshold:.2f}m required for Load-out and AGI Arrival | VBA: RefreshTideData로 업데이트"
    ws_tide["A2"].font = Font(size=10, italic=True)
    ws_tide["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    headers = ["Date", "High Tide Window", "Max Height (m)", "Risk Level"]
    for col, h in enumerate(headers, 1):
        cell = ws_tide.cell(4, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    tide_records = []
    if tide_json_path and os.path.exists(tide_json_path):
        tide_records = load_tide_data(json_path=tide_json_path)
        print(f"✅ Loaded {len(tide_records)} tide records from JSON")
    elif tide_tsv_path and os.path.exists(tide_tsv_path):
        tide_records = load_tide_data(tsv_path=tide_tsv_path)
        print(f"✅ Loaded {len(tide_records)} tide records from TSV")
    
    for r, (date_str, window, height, risk) in enumerate(tide_records, 5):
        ws_tide.cell(r, 1, value=date_str)
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            ws_tide.cell(r, 1, value=date_obj)
            ws_tide.cell(r, 1).number_format = "YYYY-MM-DD"
        except:
            ws_tide.cell(r, 1, value=date_str)
        
        ws_tide.cell(r, 2, value=window)
        ws_tide.cell(r, 3, value=height)
        ws_tide.cell(r, 3).number_format = "0.00"
        ws_tide.cell(r, 4, value=risk)
        
        for c in range(1, 5):
            ws_tide.cell(r, c).border = tb()
        
        if risk == "HIGH":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFCDD2")
        elif risk == "MEDIUM":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFE0B2")
        elif risk == "LOW":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="C8E6C9")
        
        if height >= tide_threshold:
            ws_tide.cell(r, 3).font = Font(bold=True, color="0066CC")
            ws_tide.cell(r, 1).fill = PatternFill("solid", fgColor="E3F2FD")
    
    ws_tide.column_dimensions["A"].width = 12
    ws_tide.column_dimensions["B"].width = 25
    ws_tide.column_dimensions["C"].width = 14
    ws_tide.column_dimensions["D"].width = 12
    
    ws_tide.freeze_panes = "A5"

def create_comparison_summary(
    wb,
    project_start,
    scenario_a="ScenarioA",
    scenario_b="ScenarioB",
    label_a=None,
    label_b=None,
    sheet_name="Scenario_Comparison",
):
    """
    두 시나리오 비교 Summary 시트 생성
    """
    short_a, sched_a, _ = get_scenario_sheet_names(scenario_a)
    short_b, sched_b, _ = get_scenario_sheet_names(scenario_b)
    if not label_a:
        label_a = short_a
    if not label_b:
        label_b = short_b

    ws_comp = wb.create_sheet(sheet_name)
    
    ws_comp.merge_cells("A1:F1")
    ws_comp["A1"] = "AGI TR Transportation - Scenario Comparison"
    ws_comp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_comp["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_comp["A1"].alignment = Alignment(horizontal="center")
    
    headers = ["Metric", label_a, label_b, "Difference", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_comp.cell(3, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    comparison_data = [
        (f"Total Tasks", f"=COUNTA('{sched_a}'!A:A)-5", f"=COUNTA('{sched_b}'!A:A)-5", "=C4-B4", "Task count difference"),
        (f"Total Voyages", f"=COUNTIF('{sched_a}'!A:A,\"V*\")", f"=COUNTIF('{sched_b}'!A:A,\"V*\")", "=C5-B5", ""),
        (f"Project Duration (days)", f"=MAX('{sched_a}'!G:G)-PROJECT_START", f"=MAX('{sched_b}'!G:G)-PROJECT_START", "=C6-B6", "Days difference"),
        (f"Total Jack-down Events", f"=COUNTIF('{sched_a}'!D:D,\"JACKDOWN\")", f"=COUNTIF('{sched_b}'!D:D,\"JACKDOWN\")", "=C7-B7", ""),
        (f"Project End Date", f"=MAX('{sched_a}'!G:G)", f"=MAX('{sched_b}'!G:G)", "=C8-B8", "Date difference"),
    ]
    
    for r, (metric, val_a, val_b, diff, notes) in enumerate(comparison_data, 4):
        ws_comp.cell(r, 1, value=metric)
        ws_comp.cell(r, 1).font = Font(bold=True)
        
        ws_comp.cell(r, 2, value=val_a)
        ws_comp.cell(r, 3, value=val_b)
        ws_comp.cell(r, 4, value=diff)
        ws_comp.cell(r, 5, value=notes)
        
        for c in range(1, 6):
            ws_comp.cell(r, c).border = tb()
            if c in [2, 3, 4] and "Date" in metric:
                ws_comp.cell(r, c).number_format = "YYYY-MM-DD"
            elif c == 4 and "Date" not in metric:
                ws_comp.cell(r, c).number_format = "0"
    
    for col in ["A", "B", "C", "D", "E"]:
        ws_comp.column_dimensions[col].width = 25

def create_gantt_with_vba(
    tsv_path=None,
    mammoet_original_tsv=None,
    mammoet_scenarioa_tsv=None,
    mammoet_alternative_tsv=None,
    mobilization_tsv=None,
    tide_tsv=None,
    tide_json=None,
    project_start_override=None,
):
    import os
    
    wb = Workbook()
    # TSV 파일에서 프로젝트 시작일 자동 감지 (다양한 TSV 형식 지원)
    project_start_default = dt.date(2026, 1, 18)

    # 사용자가 project_start를 직접 넘겼다면 그 값을 우선합니다(하위 호환)
    project_start = project_start_default
    if project_start_override:
        project_start = project_start_override
    else:
        candidates = [
            mammoet_original_tsv,
            mammoet_scenarioa_tsv,
            mammoet_alternative_tsv,
            mobilization_tsv,
            tsv_path,
        ]

        def _detect_earliest_date(path):
            import csv
            import re
            from datetime import datetime

            expected_p6_headers = [
                "Activity ID",
                "Activity Name",
                "Original Duration",
                "Planned Start",
                "Planned Finish",
                "Actual Start",
                "Actual Finish",
            ]

            def _try_parse(s: str):
                s = (s or "").strip()
                if not s:
                    return None
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y", "%d-%b"):
                    try:
                        d = datetime.strptime(s, fmt).date()
                        if fmt == "%d-%b":
                            d = d.replace(year=project_start_default.year)
                        return d
                    except Exception:
                        continue
                return None

            try:
                with open(path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.reader(f, delimiter="\t")
                    headers = next(reader, None)
                    if not headers:
                        return None
                    headers = [h.strip().lstrip("\ufeff") for h in headers]

                    # probable start columns (prefer Start/Planned Start/Actual Start)
                    start_idx = [i for i, h in enumerate(headers) if h.lower() in {"start", "planned start", "actual start"}]
                    if not start_idx:
                        start_idx = [i for i, h in enumerate(headers) if "start" in h.lower()]
                    scan_all = not start_idx

                    earliest = None
                    for values in reader:
                        if not values or not any((v or "").strip() for v in values):
                            continue

                        # Repair common P6 export row with missing Activity Name column
                        if headers[:7] == expected_p6_headers and len(values) == len(headers) - 1:
                            if len(values) >= 4 and re.fullmatch(r"-?\d+(?:\.\d+)?", (values[1] or "").strip() or "") and _try_parse(values[2]) and _try_parse(values[3]):
                                values = [values[0], ""] + values[1:]

                        if len(values) < len(headers):
                            values = values + [""] * (len(headers) - len(values))
                        if len(values) > len(headers):
                            values = values[: len(headers)]

                        probe = values if scan_all else [values[i] for i in start_idx if i < len(values)]
                        for v in probe:
                            d = _try_parse(v)
                            if not d:
                                continue
                            earliest = d if earliest is None else min(earliest, d)

                    return earliest
            except Exception:
                return None

        earliest = None
        for candidate in candidates:
            if not candidate:
                continue
            if not os.path.exists(candidate):
                continue
            d = _detect_earliest_date(candidate)
            if d:
                earliest = d if earliest is None else min(earliest, d)

        project_start = earliest or project_start_default

    
    # === CONTROL PANEL (새로 추가) ===
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"
    
    # Title
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "🎛️ AGI TR Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30
    
    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "📌 Changing the start date (B4) will automatically update all schedules. VBA macros must be enabled."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    
    # Input Section
    ws_ctrl["A4"] = "📅 Project Start Date:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = project_start
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)
    
    ws_ctrl["A5"] = "🎯 Target Completion Date:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = dt.date(2026, 2, 28)
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()
    
    # Voyage Pattern Selection
    ws_ctrl["A6"] = "🚢 Voyage Pattern:"
    ws_ctrl["A6"].font = Font(bold=True)
    ws_ctrl["B6"] = "1-2-2-2"  # Default: 1 solo + 3 pairs
    ws_ctrl["B6"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B6"].border = tb()
    # Note: Valid patterns: "1x1x1x1x1x1x1", "1-2-2-2", "2-2-2-1"
    voyage_patterns = ["1x1x1x1x1x1x1", "1-2-2-2", "2-2-2-1"]
    dv_voyage = DataValidation(
        type="list",
        formula1=f'"{",".join(voyage_patterns)}"',
        allow_blank=True,
    )
    dv_voyage.error = "Invalid voyage pattern selected"
    dv_voyage.errorTitle = "Invalid Voyage Pattern"
    ws_ctrl.add_data_validation(dv_voyage)
    dv_voyage.add("B6")
    
    # Early Return Option
    ws_ctrl["A7"] = "🔄 Early Return (1st JD):"
    ws_ctrl["A7"].font = Font(bold=True)
    ws_ctrl["B7"] = "TRUE"  # TRUE = return after first JD, FALSE = return after batch JD
    ws_ctrl["B7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B7"].border = tb()
    
    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")
    wb.defined_names["VOYAGE_PATTERN"] = DefinedName("VOYAGE_PATTERN", attr_text="Control_Panel!$B$6")
    wb.defined_names["EARLY_RETURN"] = DefinedName("EARLY_RETURN", attr_text="Control_Panel!$B$7")
    
    # Duration Parameters
    ws_ctrl["D4"] = "⏱️ Task Duration (Days)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)
    
    durations = [
        ("D5", "Mobilization:", "E5", 1.0, "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", 3.0, "DUR_DECK"),
        ("D7", "Load-out:", "E7", 1.0, "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", 0.5, "DUR_SF"),
        ("D9", "MWS Approval:", "E9", 0.5, "DUR_MWS"),
        ("D10", "Sailing:", "E10", 1.0, "DUR_SAIL"),
        ("D11", "AGI Unload:", "E11", 1.0, "DUR_UL"),
        ("D12", "Turning:", "E12", 3.0, "DUR_TURN"),
        ("D13", "Jack-down:", "E13", 1.0, "DUR_JD"),
        ("D14", "Return:", "E14", 1.0, "DUR_RET"),
        ("D15", "Buffer:", "E15", 0.5, "DUR_BUF"),
    ]
    
    for lc, lt, vc, v, name in durations:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = v
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")
    
    # Weather Settings
    ws_ctrl["G4"] = "🌊 Weather Settings"
    ws_ctrl["G4"].font = Font(bold=True, size=12)
    ws_ctrl["G5"] = "Shamal Start:"
    ws_ctrl["H5"] = dt.date(2026, 1, 15)
    ws_ctrl["H5"].number_format = "YYYY-MM-DD"
    ws_ctrl["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G6"] = "Shamal End:"
    ws_ctrl["H6"] = dt.date(2026, 4, 30)
    ws_ctrl["H6"].number_format = "YYYY-MM-DD"
    ws_ctrl["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    ws_ctrl["G7"] = "Tide Threshold (m):"
    ws_ctrl["G7"].font = Font(bold=True)
    ws_ctrl["H7"] = 1.90
    ws_ctrl["H7"].number_format = "0.00"
    ws_ctrl["H7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H7"].border = tb()
    
    ws_ctrl["G8"] = "📊 Gantt Min Days:"
    ws_ctrl["G8"].font = Font(bold=True)
    ws_ctrl["H8"] = 120
    ws_ctrl["H8"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H8"].border = tb()
    ws_ctrl["H8"].number_format = "0"
    
    ws_ctrl["G9"] = "📊 Gantt Buffer Days:"
    ws_ctrl["G9"].font = Font(bold=True)
    ws_ctrl["H9"] = 30
    ws_ctrl["H9"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H9"].border = tb()
    ws_ctrl["H9"].number_format = "0"
    
    # LCT Maintenance Settings
    ws_ctrl["G10"] = "🔧 LCT Maint. Start:"
    ws_ctrl["G10"].font = Font(bold=True)
    ws_ctrl["H10"] = dt.date(2026, 2, 10)  # Default maintenance start
    ws_ctrl["H10"].number_format = "YYYY-MM-DD"
    ws_ctrl["H10"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H10"].border = tb()
    
    ws_ctrl["G11"] = "🔧 LCT Maint. End:"
    ws_ctrl["G11"].font = Font(bold=True)
    ws_ctrl["H11"] = dt.date(2026, 2, 14)  # Default maintenance end (4 days)
    ws_ctrl["H11"].number_format = "YYYY-MM-DD"
    ws_ctrl["H11"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H11"].border = tb()

    wb.defined_names["SHAMAL_START"] = DefinedName("SHAMAL_START", attr_text="Control_Panel!$H$5")
    wb.defined_names["SHAMAL_END"] = DefinedName("SHAMAL_END", attr_text="Control_Panel!$H$6")
    wb.defined_names["TIDE_THRESHOLD"] = DefinedName("TIDE_THRESHOLD", attr_text="Control_Panel!$H$7")
    wb.defined_names["GANTT_MIN_DAYS"] = DefinedName("GANTT_MIN_DAYS", attr_text="Control_Panel!$H$8")
    wb.defined_names["GANTT_BUFFER_DAYS"] = DefinedName("GANTT_BUFFER_DAYS", attr_text="Control_Panel!$H$9")
    wb.defined_names["LCT_MAINT_START"] = DefinedName("LCT_MAINT_START", attr_text="Control_Panel!$H$10")
    wb.defined_names["LCT_MAINT_END"] = DefinedName("LCT_MAINT_END", attr_text="Control_Panel!$H$11")

    
    # Summary Section
    ws_ctrl["A8"] = "📊 Auto Calculation Summary"
    ws_ctrl["A8"].font = Font(bold=True, size=12)
    
    summary_items = [
        ("A9", "Estimated Completion (Latest):", "B9", "N/A"),
        ("A10", "Total Duration (Days):", "B10", "N/A"),
        ("A11", "Status vs Target:", "B11", "N/A"),
        ("A12", "Remaining Days:", "B12", "N/A"),
        ("A13", "Estimated Completion (Orig):", "B13", "N/A"),
        ("A14", "Estimated Completion (ScnA):", "B14", "N/A"),
        ("A15", "Estimated Completion (Alt):", "B15", "N/A"),
    ]
    
    for lc, lt, vc, formula in summary_items:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if isinstance(formula, str) and formula.startswith("=") and "MAX" in formula:
            ws_ctrl[vc].number_format = "YYYY-MM-DD"
    
    # VBA Button Info
    ws_ctrl["A18"] = "🔘 VBA Macros (Alt+F8)"
    ws_ctrl["A18"].font = Font(bold=True, size=12)
    
    buttons = [
        "▶ UpdateAllSchedules - Recalculate All Schedules",
        "▶ UpdateAllScenarios - Batch Update All Scenarios",
        "▶ RefreshGanttChart - Refresh Gantt Chart Colors",
        "▶ RefreshAllGanttCharts - Refresh All Gantt Charts",
        "▶ RefreshTideData - Highlight Tide Data",
        "▶ GenerateReport - Generate Status Report",
        "▶ ExportToPDF - Export to PDF",
        "▶ SimulateDelay - Simulate Delay",
        "▶ HighlightCritical - Highlight Critical Path",
        "▶ HighlightToday - Show Today's Date",
        "▶ CheckShamalRisk - Check Shamal Risk",
        "▶ HighlightLCTMaintenance - LCT Maintenance Period",
        "▶ ShowControlPanelSettings - Show All Settings",
    ]
    for i, btn in enumerate(buttons, 19):
        ws_ctrl[f"A{i}"] = btn
        ws_ctrl[f"A{i}"].font = Font(size=10)
    
    # Column widths
    ws_ctrl.column_dimensions["A"].width = 20
    ws_ctrl.column_dimensions["B"].width = 15
    ws_ctrl.column_dimensions["D"].width = 16
    ws_ctrl.column_dimensions["E"].width = 10
    ws_ctrl.column_dimensions["G"].width = 14
    ws_ctrl.column_dimensions["H"].width = 12
    ws_ctrl.column_dimensions["I"].width = 10
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    try:
        tide_threshold = float(ws_ctrl["H7"].value)
    except Exception:
        tide_threshold = 1.90
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"
    
    # === TASK DEFINITIONS (used for internal calculations) ===
    # Task definitions - TSV 파일에서 로드 또는 기본값 사용
    # (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    # 프로젝트 시작: 2026-01-18 (Day 0)
    default_tasks = [
        # Mobilization
        ("MOB-001", "1.0", "MOBILIZATION", "MOBILIZATION", "Mammoet", 0, "DUR_MOB", "SPMT Assembly + Marine Equipment Mobilization"),
        ("PREP-001", "1.1", "Deck Preparations", "DECK_PREP", "Mammoet", 1, "DUR_DECK", "One-time setup for all voyages"),
        
        # Voyage 1: LO 01-18, SAIL 01-20, ARR 01-22
        ("V1", "2.0", "VOYAGE 1: TR1 Transport", "MILESTONE", "All", 0, 0, "✅ Tide ≥1.90m (2.05m) | Good Weather Window"),
        ("LO-101", "2.1", "TR1 Load-out on LCT", "LOADOUT", "Mammoet", 0, "DUR_LO", "Tide ≥1.90m (2.05m) required"),
        ("SF-102", "2.2", "TR1 Sea Fastening", "SEAFAST", "Mammoet", 0, "DUR_SF", "12-point lashing"),
        ("MWS-103", "2.3", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 0, "DUR_MWS", "Marine Warranty Surveyor"),
        ("SAIL-104", "2.4", "V1 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 2, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-105", "2.5", "AGI Arrival + TR1 RORO Unload", "AGI_UNLOAD", "Mammoet", 4, "DUR_UL", "Tide ≥1.90m (1.91m) | AGI FWD Draft ≤ 2.70m"),
        ("STORE-106", "2.6", "TR1 Stored on AGI Laydown", "BUFFER", "Mammoet", 4, "DUR_BUF", "Awaiting pair TR2"),
        ("RET-107", "2.7", "V1 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 4, "DUR_RET", "Quick turnaround"),
        ("BUF-108", "2.99", "V1 Buffer / Equipment Reset", "BUFFER", "All", 5, "DUR_BUF", "Weather contingency"),
        
        # Voyage 2: LO 01-26, SAIL 01-27, ARR 01-29
        ("V2", "3.0", "VOYAGE 2: TR2 Transport + JD-1", "MILESTONE", "All", 8, 0, "✅ Tide ≥1.90m (1.91m) | Good Weather Window (before Shamal)"),
        ("LO-109", "3.1", "TR2 Load-out on LCT", "LOADOUT", "Mammoet", 8, "DUR_LO", "Tide ≥1.90m (1.91m) required"),
        ("SF-110", "3.2", "TR2 Sea Fastening", "SEAFAST", "Mammoet", 8, "DUR_SF", "12-point lashing"),
        ("MWS-110A", "3.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 8, "DUR_MWS", "Pre-sail verification"),
        ("SAIL-111", "3.3", "V2 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 9, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-112", "3.4", "AGI Arrival + TR2 RORO Unload", "AGI_UNLOAD", "Mammoet", 11, "DUR_UL", "Tide ≥1.90m (2.03m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-113", "3.5", "TR1 Transport to Bay-1", "TURNING", "Mammoet", 12, 1, "Steel bridge install"),
        ("TURN-114", "3.6", "TR1 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", "10t Forklift required"),
        ("TRN-116", "3.8", "TR2 Transport to Bay-2", "TURNING", "Mammoet", 12, 1, ""),
        ("TURN-117", "3.9", "TR2 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", ""),
        ("JD-120A", "3.95", "JD-1 Jack-Down TR1", "JACKDOWN", "Mammoet", 14, "DUR_JD", "MILESTONE: TR1 complete | 02-01"),
        ("RET-119", "3.11", "V2 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 15, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-120B", "3.96", "JD-1 Jack-Down TR2", "JACKDOWN", "Mammoet", 16, "DUR_JD", "MILESTONE: TR2 complete | 02-02"),
        ("BUF-120", "3.99", "V2 Buffer / Shamal Recovery", "BUFFER", "All", 17, "DUR_BUF", "Post-Shamal weather check"),
        
        # Voyage 3: LO 01-31, SAIL 02-02, ARR 02-03
        ("V3", "4.0", "VOYAGE 3: TR3 Transport", "MILESTONE", "All", 13, 0, "✅ Tide ≥1.90m (2.07m) | Post-Shamal Window"),
        ("LO-121", "4.1", "TR3 Load-out on LCT", "LOADOUT", "Mammoet", 13, "DUR_LO", "Tide ≥1.90m (2.07m)"),
        ("SF-122", "4.2", "TR3 Sea Fastening", "SEAFAST", "Mammoet", 13, "DUR_SF", ""),
        ("MWS-122A", "4.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 13, "DUR_MWS", ""),
        ("SAIL-123", "4.3", "V3 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 15, "DUR_SAIL", "Good weather"),
        ("ARR-124", "4.4", "AGI Arrival + TR3 RORO Unload", "AGI_UNLOAD", "Mammoet", 16, "DUR_UL", "Tide ≥1.90m (2.04m)"),
        ("STORE-125", "4.5", "TR3 Stored on AGI Laydown", "BUFFER", "Mammoet", 16, "DUR_BUF", "Awaiting pair TR4"),
        ("RET-126", "4.6", "V3 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 17, "DUR_RET", ""),
        ("BUF-127", "4.99", "V3 Buffer", "BUFFER", "All", 17, "DUR_BUF", ""),
        
        # Voyage 4: LO 02-15, SAIL 02-16, ARR 02-18
        ("V4", "5.0", "VOYAGE 4: TR4 Transport + JD-2", "MILESTONE", "All", 28, 0, "✅ Tide ≥1.90m (1.90m) | Shamal 종료 직후"),
        ("LO-128", "5.1", "TR4 Load-out on LCT", "LOADOUT", "Mammoet", 28, "DUR_LO", "Tide ≥1.90m (1.90m)"),
        ("SF-129", "5.2", "TR4 Sea Fastening", "SEAFAST", "Mammoet", 28, "DUR_SF", ""),
        ("MWS-129A", "5.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 28, "DUR_MWS", ""),
        ("SAIL-130", "5.3", "V4 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 29, "DUR_SAIL", ""),
        ("ARR-131", "5.4", "AGI Arrival + TR4 RORO Unload", "AGI_UNLOAD", "Mammoet", 31, "DUR_UL", "Tide ≥1.90m (1.96m)"),
        ("TRN-132", "5.5", "TR3 Transport to Bay-3", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-133", "5.6", "TR3 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("TRN-135", "5.8", "TR4 Transport to Bay-4", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-136", "5.9", "TR4 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("JD-139A", "5.95", "JD-2 Jack-Down TR3", "JACKDOWN", "Mammoet", 33, "DUR_JD", "MILESTONE: TR3 complete | 02-20"),
        ("RET-138", "5.11", "V4 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 34, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-139B", "5.96", "JD-2 Jack-Down TR4", "JACKDOWN", "Mammoet", 35, "DUR_JD", "MILESTONE: TR4 complete | 02-21"),
        ("BUF-140", "5.99", "V4 Buffer", "BUFFER", "All", 36, "DUR_BUF", ""),
        
        # Voyage 5: LO 02-23, SAIL 02-23, ARR 02-24 (Fast-turn)
        ("V5", "6.0", "VOYAGE 5: TR5 Transport", "MILESTONE", "All", 36, 0, "✅ Tide ≥1.90m (1.99m) | Fast-turn"),
        ("LO-140", "6.1", "TR5 Load-out on LCT", "LOADOUT", "Mammoet", 36, "DUR_LO", "Tide ≥1.90m (1.99m)"),
        ("SF-141", "6.2", "TR5 Sea Fastening", "SEAFAST", "Mammoet", 36, "DUR_SF", ""),
        ("MWS-141A", "6.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 36, "DUR_MWS", ""),
        ("SAIL-142", "6.3", "V5 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 36, "DUR_SAIL", "Fast-turn"),
        ("ARR-143", "6.4", "AGI Arrival + TR5 RORO Unload", "AGI_UNLOAD", "Mammoet", 37, "DUR_UL", "Tide ≥1.90m (2.01m)"),
        ("STORE-144", "6.5", "TR5 Stored on AGI Laydown", "BUFFER", "Mammoet", 37, "DUR_BUF", "Awaiting pair TR6"),
        ("RET-145", "6.6", "V5 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 37, "DUR_RET", ""),
        ("BUF-146", "6.99", "V5 Buffer", "BUFFER", "All", 37, "DUR_BUF", ""),
        
        # Voyage 6: LO 02-25, SAIL 02-25, ARR 02-26 (Fast-turn)
        ("V6", "7.0", "VOYAGE 6: TR6 Transport + JD-3", "MILESTONE", "All", 38, 0, "✅ Tide ≥1.90m (2.01m) | Fast-turn"),
        ("LO-147", "7.1", "TR6 Load-out on LCT", "LOADOUT", "Mammoet", 38, "DUR_LO", "Tide ≥1.90m (2.01m)"),
        ("SF-148", "7.2", "TR6 Sea Fastening", "SEAFAST", "Mammoet", 38, "DUR_SF", ""),
        ("MWS-148A", "7.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 38, "DUR_MWS", ""),
        ("SAIL-149", "7.3", "V6 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 38, "DUR_SAIL", "Fast-turn"),
        ("ARR-150", "7.4", "AGI Arrival + TR6 RORO Unload", "AGI_UNLOAD", "Mammoet", 39, "DUR_UL", "Tide ≥1.90m (1.98m)"),
        ("TRN-151", "7.5", "TR5 Transport to Bay-5", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-152", "7.6", "TR5 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("TRN-154", "7.8", "TR6 Transport to Bay-6", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-155", "7.9", "TR6 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("JD-157A", "7.95", "JD-3 Jack-Down TR5", "JACKDOWN", "Mammoet", 40, "DUR_JD", "MILESTONE: TR5 complete | 02-27"),
        ("RET-158", "7.11", "V6 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-157B", "7.96", "JD-3 Jack-Down TR6", "JACKDOWN", "Mammoet", 42, "DUR_JD", "MILESTONE: TR6 complete | 02-28"),
        ("BUF-159", "7.99", "V6 Buffer / Reset for V7", "BUFFER", "All", 43, "DUR_BUF", ""),
        
        # Voyage 7: LO 02-27, SAIL 02-27, ARR 02-28 (Final)
        ("V7", "8.0", "VOYAGE 7: TR7 Transport + JD-4", "MILESTONE", "All", 40, 0, "✅ Tide ≥1.90m (1.92m) | Final unit"),
        ("LO-201", "8.1", "TR7 Load-out on LCT", "LOADOUT", "Mammoet", 40, "DUR_LO", "Tide ≥1.90m (1.92m) required"),
        ("SF-202", "8.2", "TR7 Sea Fastening", "SEAFAST", "Mammoet", 40, "DUR_SF", "12-point lashing"),
        ("MWS-202A", "8.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 40, "DUR_MWS", ""),
        ("SAIL-203", "8.3", "V7 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 40, "DUR_SAIL", "Weather window required"),
        ("ARR-204", "8.4", "AGI Arrival + TR7 RORO Unload", "AGI_UNLOAD", "Mammoet", 41, "DUR_UL", "Tide ≥1.90m (1.93m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-205", "8.5", "TR7 Transport to Bay-7", "TURNING", "Mammoet", 41, 1, "Steel bridge install"),
        ("TURN-206", "8.6", "TR7 Turning (90° rotation)", "TURNING", "Mammoet", 41, "DUR_TURN", "10t Forklift required"),
        ("JD-207", "8.7", "★ JD-4 Jack-Down (TR7)", "JACKDOWN", "Mammoet", 41, "DUR_JD", "MILESTONE: TR7 Complete | 02-28"),
        ("RET-208", "8.8", "V7 LCT Final Return: AGI→MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Final return"),
        
        # Demobilization
        ("DEMOB", "9.0", "DEMOBILIZATION", "MOBILIZATION", "Mammoet", 42, "DUR_MOB", "Equipment return"),
        ("END", "99.0", "★★★ PROJECT COMPLETE ★★★", "MILESTONE", "All", 42, 0, "All 7 TRs Installed | Jan-Feb 2026 Complete"),
    ]

    tasks = default_tasks
    if tsv_path:
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
            print(f"✅ Loaded {len(tasks)} tasks from TSV file")
        except Exception as e:
            print(f"Error loading TSV: {e}")
            print("Using default tasks list")
    
    # Schedule_Data and Gantt_Chart sheets removed; tasks remain for summary/weather calculations.
    
    # === Mammoet Sheets ===
    mammoet_original_created = False
    mammoet_scenarioa_created = False
    mammoet_alternative_created = False

    if mammoet_original_tsv and os.path.exists(mammoet_original_tsv):
        mammoet_original_created = create_scenario_sheets(
            wb,
            "Mammoet_Original",
            mammoet_original_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if mammoet_original_created:
            print(f"Created Mammoet_Original sheets from {os.path.basename(mammoet_original_tsv)}")
        else:
            print("Warning: Mammoet_Original sheets not created")
    elif mobilization_tsv and os.path.exists(mobilization_tsv):
        mammoet_original_created = create_scenario_sheets(
            wb,
            "Mammoet_Original",
            mobilization_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if mammoet_original_created:
            print(f"Created Mammoet_Original sheets from {os.path.basename(mobilization_tsv)}")
        else:
            print("Warning: Mammoet_Original sheets not created from MOBILIZATION.tsv")

    if mammoet_scenarioa_tsv and os.path.exists(mammoet_scenarioa_tsv):
        mammoet_scenarioa_created = create_scenario_sheets(
            wb,
            "Mammoet_ScenarioA",
            mammoet_scenarioa_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if mammoet_scenarioa_created:
            print(f"Created Mammoet_ScenarioA sheets from {os.path.basename(mammoet_scenarioa_tsv)}")
        else:
            print("Warning: Mammoet_ScenarioA sheets not created")

    if mammoet_alternative_tsv and os.path.exists(mammoet_alternative_tsv):
        mammoet_alternative_created = create_scenario_sheets(
            wb,
            "Mammoet_Alternative",
            mammoet_alternative_tsv,
            project_start,
            pattern_str=None,
            early_return=False,
        )
        if mammoet_alternative_created:
            print(f"Created Mammoet_Alternative sheets from {os.path.basename(mammoet_alternative_tsv)}")
        else:
            print("Warning: Mammoet_Alternative sheets not created")

    def _set_summary_date_cell(addr, value):
        ws_ctrl[addr] = value
        ws_ctrl[addr].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[addr].border = tb()
        if isinstance(value, str) and value.startswith("="):
            ws_ctrl[addr].number_format = "YYYY-MM-DD"

    end_sources = []
    if mammoet_original_created:
        end_sources.append("Schedule_Data_Mammoet_Orig!G:G")
        _set_summary_date_cell("B13", "=MAX(Schedule_Data_Mammoet_Orig!G:G)")
    else:
        _set_summary_date_cell("B13", "N/A")
    if mammoet_scenarioa_created:
        end_sources.append("Schedule_Data_Mammoet_ScnA!G:G")
        _set_summary_date_cell("B14", "=MAX(Schedule_Data_Mammoet_ScnA!G:G)")
    else:
        _set_summary_date_cell("B14", "N/A")
    if mammoet_alternative_created:
        end_sources.append("Schedule_Data_Mammoet_Alt!G:G")
        _set_summary_date_cell("B15", "=MAX(Schedule_Data_Mammoet_Alt!G:G)")
    else:
        _set_summary_date_cell("B15", "N/A")

    if end_sources:
        latest_formula = f"=MAX({','.join(end_sources)})"
        ws_ctrl["B5"] = latest_formula
        ws_ctrl["B5"].number_format = "YYYY-MM-DD"
        ws_ctrl["B9"] = latest_formula
        ws_ctrl["B9"].number_format = "YYYY-MM-DD"
        ws_ctrl["B10"] = "=B9-B4+1"
        ws_ctrl["B11"] = '=IF(B9<=B5,"✅ On Target","❌ Delayed")'
        ws_ctrl["B12"] = "=B5-B9"
        for addr in ("B10", "B11", "B12"):
            ws_ctrl[addr].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
            ws_ctrl[addr].border = tb()
    
    # === Tide Data Sheet ===
    if tide_json and os.path.exists(tide_json):
        create_tide_data_sheet(wb, tide_json_path=tide_json)
        print("✅ Created Tide_Data sheet (from JSON)")
    elif tide_tsv and os.path.exists(tide_tsv):
        create_tide_data_sheet(wb, tide_tsv_path=tide_tsv)
        print("✅ Created Tide_Data sheet (from TSV)")
    
    # === Scenario Comparison 제거 ===
    
    # === WEATHER ANALYSIS (동적화) ===
    ws_weather = wb.create_sheet("Weather_Analysis")
    
    try:
        project_end = ws_ctrl["B5"].value
        if hasattr(project_end, "date"):
            project_end = project_end.date()
        if not isinstance(project_end, dt.date):
            raise ValueError("Invalid target end date")
    except Exception:
        max_offset = 0
        for task in tasks:
            if len(task) >= 6:
                offset = task[5] if isinstance(task[5], (int, float)) else 0
                max_offset = max(max_offset, offset)
        project_end = project_start + dt.timedelta(days=int(max_offset) + 30)
    
    year_month = project_start.strftime("%b %Y")
    ws_weather["A1"] = f"UAE Winter Weather Analysis - {year_month}"
    ws_weather["A1"].font = Font(bold=True, size=14)
    
    weather_periods = generate_weather_periods(project_start, project_end)
    period_names = [p[0] for p in weather_periods]
    
    weather_headers = ["Parameter"] + period_names + ["Notes"]
    num_cols = len(weather_headers)
    ws_weather.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    
    for c, h in enumerate(weather_headers, 1):
        cell = ws_weather.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = tb()
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    
    def get_weather_for_period(period_start, period_end, shamal_start, shamal_end):
        is_shamal = (period_start <= shamal_end and period_end >= shamal_start)
        if is_shamal:
            return {
                "wind": "16-21",
                "gust": "25-30",
                "wave": "0.8-1.2",
                "visibility": "2-5",
                "risk": "HIGH",
                "recommendation": "NO-GO",
            }
        return {
            "wind": "11-13",
            "gust": "18-20",
            "wave": "0.4-0.6",
            "visibility": "8-10",
            "risk": "LOW",
            "recommendation": "GO",
        }
    
    weather_data_rows = [
        ["Avg Wind (kt)"] + [""] * len(weather_periods) + [f"Peak Shamal: {shamal_text}"],
        ["Max Gust (kt)"] + [""] * len(weather_periods) + ["NO-GO if >22kt gust"],
        ["Wave Height (m)"] + [""] * len(weather_periods) + ["HOLD if >0.8m"],
        ["Visibility (km)"] + [""] * len(weather_periods) + ["Reduced during Shamal"],
        ["Risk Level"] + [""] * len(weather_periods) + [""],
        ["Recommendation"] + [""] * len(weather_periods) + [""],
    ]
    
    for idx, (_name, period_start, period_end) in enumerate(weather_periods):
        weather = get_weather_for_period(period_start, period_end, shamal_start, shamal_end)
        col = idx + 1
        weather_data_rows[0][col] = weather["wind"]
        weather_data_rows[1][col] = weather["gust"]
        weather_data_rows[2][col] = weather["wave"]
        weather_data_rows[3][col] = weather["visibility"]
        weather_data_rows[4][col] = weather["risk"]
        weather_data_rows[5][col] = weather["recommendation"]
    
    for r, row in enumerate(weather_data_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_weather.cell(r, c, value=val)
            cell.border = tb()
            if val == "HIGH" or val == "NO-GO":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif val == "MEDIUM" or val == "CAUTION":
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
            elif val == "LOW" or val == "GO":
                cell.fill = PatternFill("solid", fgColor="C8E6C9")
    
    for col in range(1, num_cols + 1):
        ws_weather.column_dimensions[get_column_letter(col)].width = 15
    
    # === SUMMARY (기존과 동일) ===
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "AGI HVDC Transformer Transportation - Project Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")
    
    summary_data = [
        ("Key Parameters", ""),
        ("Total Transformers", "7 units (TR1-TR7)"),
        ("Weight per TR", "217-271 tons"),
        ("Total Voyages", "7 (1 TR per voyage)"),
        ("Jack-down Events", "4 (after V2, V4, V6, V7)"),
        ("Vessel", "LCT BUSHRA"),
        ("Route", "Mina Zayed Port ↔ AGI Site"),
        ("", ""),
        ("Schedule Summary", ""),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
    ]
    
    voyage_ranges = calculate_voyage_ranges(tasks)
    for voyage_name, start_day, end_day in voyage_ranges:
        if "Mobilization" in voyage_name:
            label = voyage_name
        else:
            if ":" in voyage_name:
                prefix, rest = voyage_name.split(":", 1)
                label = f"{prefix.title()} ({rest.strip()})"
            else:
                label = voyage_name.title()
        summary_data.append((label, f"Day {start_day}-{end_day}"))
    
    summary_data.extend([
        ("Project Complete", "=MAX(Schedule_Data_Mammoet_Orig!G:G)"),
        ("", ""),
        ("Weather Constraints", ""),
        ("Shamal Period", shamal_text_full),
        ("Tide Requirement", f"≥{tide_threshold:.2f}m for LO/ARR"),
        ("Wind Limit", "≤18kt sustained, ≤22kt gust"),
        ("Wave Limit", "≤0.8m (HOLD), ≤1.0m (NO-GO)"),
        ("AGI Draft Limit", "≤2.70m forward draft"),
    ])
    
    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30
    
    # === VBA CODE SHEET ===
    ws_vba = wb.create_sheet("VBA_Code")
    
    ws_vba["A1"] = "📋 VBA 코드 - Alt+F11 → Module에 붙여넣기 → .xlsm으로 저장"
    ws_vba["A1"].font = Font(bold=True, size=14)
    
    vba_code = '''
Option Explicit

' ============================================
' AGI TR Multi-Scenario Master Gantt - VBA Macros
' ============================================
' 사용법: Alt+F11 → Module 삽입 → 코드 붙여넣기
' ============================================

' === 통합 업데이트 함수 ===
Sub UpdateAllScenarios()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    On Error Resume Next
    Sheets("Schedule_Data_Mammoet_Orig").Calculate
    Sheets("Gantt_Chart_Mammoet_Orig").Calculate
    Sheets("Schedule_Data_Mammoet_ScnA").Calculate
    Sheets("Gantt_Chart_Mammoet_ScnA").Calculate
    Sheets("Schedule_Data_Mammoet_Alt").Calculate
    Sheets("Gantt_Chart_Mammoet_Alt").Calculate
    Sheets("Tide_Data").Calculate
    On Error GoTo 0
    
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    Sheets("Weather_Analysis").Calculate
    
    Call RefreshAllGanttCharts
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "✅ 모든 시나리오 업데이트 완료!", vbInformation, "Update Complete"
End Sub

' === 모든 Gantt 차트 색상 갱신 ===
Sub RefreshAllGanttCharts()
    On Error Resume Next
    Call RefreshGanttChart_Mammoet_Original
    Call RefreshGanttChart_Mammoet_ScenarioA
    Call RefreshGanttChart_Mammoet_Alternative
    On Error GoTo 0
End Sub

' === Mammoet_Original Gantt 갱신 ===
Sub RefreshGanttChart_Mammoet_Original()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date

    Set ws = Sheets("Gantt_Chart_Mammoet_Orig")
    Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8

    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row

    Application.ScreenUpdating = False

    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone

    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
        End If
    Next j

    For i = 6 To lastRow
        Dim activityId As String
        On Error Resume Next
        activityId = Trim(UCase(CStr(wsd.Cells(i, 1).Value)))
        On Error GoTo 0

        ' A1000 형식만 허용: A로 시작하고 두 번째 문자가 숫자여야 함
        ' (MOBILIZATION, SPMT, MARINE 등 텍스트 헤더는 제외)
        If activityId = "" Or Left(activityId, 1) <> "A" Then
            GoTo NextRow
        End If
        If Len(activityId) < 2 Or Not IsNumeric(Mid(activityId, 2, 1)) Then
            GoTo NextRow
        End If

        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value

            ganttRow = i - 1

            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733)
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
NextRow:
    Next i

    Application.ScreenUpdating = True
End Sub

' === Mammoet_ScenarioA Gantt 갱신 ===
Sub RefreshGanttChart_Mammoet_ScenarioA()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date

    Set ws = Sheets("Gantt_Chart_Mammoet_ScnA")
    Set wsd = Sheets("Schedule_Data_Mammoet_ScnA")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8

    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row

    Application.ScreenUpdating = False

    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone

    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
        End If
    Next j

    For i = 6 To lastRow
        Dim activityId As String
        On Error Resume Next
        activityId = Trim(UCase(CStr(wsd.Cells(i, 1).Value)))
        On Error GoTo 0

        ' A1000 형식만 허용: A로 시작하고 두 번째 문자가 숫자여야 함
        ' (MOBILIZATION, SPMT, MARINE 등 텍스트 헤더는 제외)
        If activityId = "" Or Left(activityId, 1) <> "A" Then
            GoTo NextRow
        End If
        If Len(activityId) < 2 Or Not IsNumeric(Mid(activityId, 2, 1)) Then
            GoTo NextRow
        End If

        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value

            ganttRow = i - 1

            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733)
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
NextRow:
    Next i

    Application.ScreenUpdating = True
End Sub

' === Mammoet_Alternative Gantt 갱신 ===
Sub RefreshGanttChart_Mammoet_Alternative()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date

    Set ws = Sheets("Gantt_Chart_Mammoet_Alt")
    Set wsd = Sheets("Schedule_Data_Mammoet_Alt")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8

    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row

    Application.ScreenUpdating = False

    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone

    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
        End If
    Next j

    For i = 6 To lastRow
        Dim activityId As String
        On Error Resume Next
        activityId = Trim(UCase(CStr(wsd.Cells(i, 1).Value)))
        On Error GoTo 0

        ' A1000 형식만 허용: A로 시작하고 두 번째 문자가 숫자여야 함
        ' (MOBILIZATION, SPMT, MARINE 등 텍스트 헤더는 제외)
        If activityId = "" Or Left(activityId, 1) <> "A" Then
            GoTo NextRow
        End If
        If Len(activityId) < 2 Or Not IsNumeric(Mid(activityId, 2, 1)) Then
            GoTo NextRow
        End If

        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value

            ganttRow = i - 1

            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733)
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
NextRow:
    Next i

    Application.ScreenUpdating = True
End Sub

' === 조석 데이터 갱신 ===
Sub RefreshTideData()
    Dim ws As Worksheet
    Dim i As Long
    Dim tideThreshold As Double
    
    Set ws = Sheets("Tide_Data")
    tideThreshold = Sheets("Control_Panel").Range("H7").Value
    If tideThreshold = 0 Then tideThreshold = 1.9
    
    For i = 5 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
        If IsNumeric(ws.Cells(i, 3).Value) Then
            If ws.Cells(i, 3).Value >= tideThreshold Then
                ws.Cells(i, 3).Font.Bold = True
                ws.Cells(i, 3).Font.Color = RGB(0, 102, 204)
                ws.Cells(i, 1).Interior.Color = RGB(227, 242, 253)
            End If
        End If
    Next i
    
    MsgBox "✅ 조석 데이터 강조 완료 (Tide ≥" & Format(tideThreshold, "0.00") & "m)", vbInformation
End Sub

' === 1. 전체 일정 업데이트 ===
Sub UpdateAllSchedules()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    Sheets("Schedule_Data_Mammoet_Orig").Calculate
    Sheets("Gantt_Chart_Mammoet_Orig").Calculate
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    
    Call RefreshGanttChart
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "✅ 일정 업데이트 완료!" & vbCrLf & vbCrLf & _
           "프로젝트 시작: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
           "예상 완료: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD"), _
           vbInformation, "Schedule Updated"
End Sub

' === 2. Gantt Chart 색상 갱신 ===
Sub RefreshGanttChart()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date
    
    Set ws = Sheets("Gantt_Chart_Mammoet_Orig")
    Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8 ' Date columns start at H
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    Application.ScreenUpdating = False
    
    ' Clear existing colors in date columns
    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone
    
    ' Reset header colors + Shamal highlight
    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121) ' HEADER color
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0) ' Orange
        End If
    Next j
    
    ' Apply Gantt bars
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value
            
            ganttRow = i - 1
            
            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733) ' Star
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
    Next i
    
    ' Highlight today
    For j = 0 To maxJ
        cellDate = projStart + j
        If cellDate = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Borders(xlEdgeLeft).Color = RGB(255, 0, 0)
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Borders(xlEdgeLeft).Weight = xlThick
            Exit For
        End If
    Next j
    
    Application.ScreenUpdating = True
End Sub

' === Phase Color Helper ===
Function GetPhaseColor(phase As String) As Long
    Select Case phase
        Case "MOBILIZATION": GetPhaseColor = RGB(142, 124, 195)
        Case "DECK_PREP": GetPhaseColor = RGB(111, 168, 220)
        Case "LOADOUT": GetPhaseColor = RGB(147, 196, 125)
        Case "SEAFAST": GetPhaseColor = RGB(118, 165, 175)
        Case "SAIL": GetPhaseColor = RGB(164, 194, 244)
        Case "AGI_UNLOAD": GetPhaseColor = RGB(246, 178, 107)
        Case "TURNING": GetPhaseColor = RGB(255, 217, 102)
        Case "JACKDOWN": GetPhaseColor = RGB(224, 102, 102)
        Case "RETURN": GetPhaseColor = RGB(153, 153, 153)
        Case "BUFFER": GetPhaseColor = RGB(217, 217, 217)
        Case "MILESTONE": GetPhaseColor = RGB(255, 0, 0)
        Case Else: GetPhaseColor = RGB(255, 255, 255)
    End Select
End Function

' === 3. 프로젝트 리포트 생성 ===
Sub GenerateReport()
    Dim wsd As Worksheet
    Dim i As Long, total As Long, jdCount As Long, lastRow As Long
    Dim voyages As Long, milestones As Long
    
    Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    For i = 6 To lastRow
        If wsd.Cells(i, 1).Value <> "" Then
            total = total + 1
            If wsd.Cells(i, 4).Value = "JACKDOWN" Then jdCount = jdCount + 1
            If wsd.Cells(i, 4).Value = "MILESTONE" Then milestones = milestones + 1
            If Left(wsd.Cells(i, 1).Value, 1) = "V" And Len(wsd.Cells(i, 1).Value) = 2 Then voyages = voyages + 1
        End If
    Next i
    
    Dim rpt As String
    rpt = "╔══════════════════════════════════════╗" & vbCrLf & _
          "║   AGI HVDC TR Transportation Report  ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ Report Date: " & Format(Now, "YYYY-MM-DD HH:MM") & "      ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ PROJECT STATUS                       ║" & vbCrLf & _
          "║  Total Tasks: " & total & "                      ║" & vbCrLf & _
          "║  Voyages: " & voyages & "                          ║" & vbCrLf & _
          "║  Jack-down Events: " & jdCount & "                 ║" & vbCrLf & _
          "║  Milestones: " & milestones & "                       ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ KEY DATES                            ║" & vbCrLf & _
          "║  Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & "              ║" & vbCrLf & _
          "║  Target: " & Format(Sheets("Control_Panel").Range("B5").Value, "YYYY-MM-DD") & "             ║" & vbCrLf & _
          "║  Est.End: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD") & "            ║" & vbCrLf & _
          "║  Status: " & Sheets("Control_Panel").Range("B11").Value & "               ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ WEATHER RISK                         ║" & vbCrLf & _
          "║  Shamal: " & Format(Sheets("Control_Panel").Range("H5").Value, "MM/DD") & " - " & Format(Sheets("Control_Panel").Range("H6").Value, "MM/DD") & "           ║" & vbCrLf & _
          "╚══════════════════════════════════════╝"
    
    MsgBox rpt, vbInformation, "Project Report"
End Sub

' === 4. PDF 내보내기 ===
Sub ExportToPDF()
    Dim fp As String
    fp = ThisWorkbook.Path & "\\AGI_TR_Gantt_" & Format(Date, "YYYYMMDD") & ".pdf"
    
    Sheets(Array("Schedule_Data_Mammoet_Orig", "Gantt_Chart_Mammoet_Orig", "Summary")).Select
    ActiveSheet.ExportAsFixedFormat xlTypePDF, fp, xlQualityStandard, True
    Sheets("Control_Panel").Select
    
    MsgBox "✅ PDF 저장 완료:" & vbCrLf & fp, vbInformation, "Export Complete"
End Sub

' === 5. 지연 시뮬레이션 ===
Sub SimulateDelay()
    Dim delayDays As Integer, origStart As Date
    Dim wsCtrl As Worksheet
    
    Set wsCtrl = Sheets("Control_Panel")
    origStart = wsCtrl.Range("B4").Value
    
    delayDays = InputBox("시뮬레이션할 지연 일수를 입력하세요:" & vbCrLf & _
                         "(현재 시작일: " & Format(origStart, "YYYY-MM-DD") & ")", _
                         "Delay Simulation", "7")
    
    If IsNumeric(delayDays) And delayDays <> 0 Then
        wsCtrl.Range("B4").Value = origStart + delayDays
        Call UpdateAllSchedules
        
        MsgBox "시뮬레이션 결과:" & vbCrLf & _
               "새 시작일: " & Format(wsCtrl.Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
               "새 완료일: " & Format(wsCtrl.Range("B9").Value, "YYYY-MM-DD") & vbCrLf & _
               "목표 대비: " & wsCtrl.Range("B11").Value, vbInformation, "Simulation Result"
        
        If MsgBox("원래 일정으로 복원하시겠습니까?", vbYesNo + vbQuestion, "Restore?") = vbYes Then
            wsCtrl.Range("B4").Value = origStart
            Call UpdateAllSchedules
        End If
    End If
End Sub

' === 6. Critical Path 강조 ===
Sub HighlightCritical()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    
    Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    ' Reset
    wsd.Range(wsd.Cells(6, 1), wsd.Cells(lastRow, 9)).Font.Bold = False
    wsd.Range(wsd.Cells(6, 1), wsd.Cells(lastRow, 9)).Font.Color = RGB(0, 0, 0)
    
    ' Highlight Jack-down and Milestones
    For i = 6 To lastRow
        If wsd.Cells(i, 4).Value = "JACKDOWN" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(183, 28, 28)
        ElseIf wsd.Cells(i, 4).Value = "MILESTONE" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(21, 101, 192)
        End If
    Next i
    
    MsgBox "✅ Critical Path 강조 완료" & vbCrLf & _
           "🔴 빨강 = Jack-down (Critical)" & vbCrLf & _
           "🔵 파랑 = Milestone", vbInformation, "Critical Path"
End Sub

' === 7. 오늘 날짜 하이라이트 ===
Sub HighlightToday()
    Dim ws As Worksheet, j As Long, lastCol As Long, maxJ As Long, lastRow As Long
    Dim projStart As Date, dc As Long
    
    Set ws = Sheets("Gantt_Chart_Mammoet_Orig")
    projStart = Sheets("Control_Panel").Range("B4").Value
    dc = 8
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    For j = 0 To maxJ
        If projStart + j = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Interior.Color = RGB(255, 255, 200)
            ws.Cells(3, dc + j).Value = "TODAY"
            ws.Cells(3, dc + j).Font.Bold = True
            ws.Cells(3, dc + j).Font.Color = RGB(255, 0, 0)
            MsgBox "오늘 날짜 (" & Format(Date, "MM/DD") & ") 컬럼이 강조되었습니다.", vbInformation
            Exit For
        End If
    Next j
End Sub

' === 8. 날짜 변경 자동 트리거 (Control_Panel 시트에 추가) ===
' 아래 코드를 Control_Panel 시트의 코드 영역에 붙여넣으세요:
'
' Private Sub Worksheet_Change(ByVal Target As Range)
'     If Target.Address = "$B$4" Then
'         Call UpdateAllSchedules
'     End If
' End Sub

' === 9. 진행률 일괄 업데이트 ===
Sub BulkProgressUpdate()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    Dim pctValue As Double
    
    pctValue = InputBox("일괄 적용할 진행률을 입력하세요 (0-100):", "Bulk Progress", "50")
    
    If IsNumeric(pctValue) Then
        pctValue = pctValue / 100
        Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
        lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
        
        ' Progress 컬럼이 없으면 추가
        If wsd.Cells(5, 10).Value <> "Progress" Then
            wsd.Cells(5, 10).Value = "Progress"
            wsd.Cells(5, 10).Font.Bold = True
            wsd.Cells(5, 10).Font.Color = RGB(255, 255, 255)
            wsd.Cells(5, 10).Fill.Color = RGB(31, 78, 121)
        End If
        
        For i = 6 To lastRow
            If wsd.Cells(i, 1).Value <> "" Then
                wsd.Cells(i, 10).Value = pctValue
                wsd.Cells(i, 10).NumberFormat = "0%"
            End If
        Next i
        
        MsgBox "진행률 " & Format(pctValue, "0%") & " 일괄 적용 완료", vbInformation
    End If
End Sub

' === 10. Shamal 위험 체크 ===
Sub CheckShamalRisk()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    Dim taskDate As Date, shamalStart As Date, shamalEnd As Date
    Dim riskTasks As String, cnt As Long
    
    Set wsd = Sheets("Schedule_Data_Mammoet_Orig")
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) Then
            taskDate = wsd.Cells(i, 6).Value
            If taskDate >= shamalStart And taskDate <= shamalEnd Then
                ' SAIL tasks are weather-critical
                If wsd.Cells(i, 4).Value = "SAIL" Or wsd.Cells(i, 4).Value = "LOADOUT" Then
                    cnt = cnt + 1
                    riskTasks = riskTasks & vbCrLf & "  ⚠️ " & wsd.Cells(i, 1).Value & ": " & wsd.Cells(i, 3).Value
                End If
            End If
        End If
    Next i
    
    If cnt > 0 Then
        MsgBox "⚠️ SHAMAL 위험 경고!" & vbCrLf & vbCrLf & _
               "Shamal 기간 (" & Format(shamalStart, "MM/DD") & "-" & Format(shamalEnd, "MM/DD") & ") 중 " & cnt & "개 기상 민감 작업 발견:" & vbCrLf & _
               riskTasks & vbCrLf & vbCrLf & _
               "일정 조정을 권장합니다.", vbExclamation, "Weather Risk Alert"
    Else
        MsgBox "✅ Shamal 기간 중 기상 민감 작업 없음" & vbCrLf & _
               "현재 일정은 안전합니다.", vbInformation, "Weather Check OK"
    End If
End Sub

' ============================================
' NEW: Control Panel Settings Reader Functions
' ============================================

' === Get Voyage Pattern from Control Panel ===
Function GetVoyagePattern() As String
    ' Returns: "1-2-2-2", "2-2-2-1", "2-2-2-1_TWO_SPMT", or "1x1x1x1x1x1x1"
    GetVoyagePattern = Sheets("Control_Panel").Range("B6").Value
    If GetVoyagePattern = "" Then GetVoyagePattern = "1-2-2-2"
End Function

' === Check if Early Return is enabled ===
Function IsEarlyReturn() As Boolean
    ' TRUE = LCT returns after first JD in a pair
    ' FALSE = LCT returns after batch JD (both TRs)
    Dim val As String
    val = UCase(Trim(Sheets("Control_Panel").Range("B7").Value))
    IsEarlyReturn = (val = "TRUE" Or val = "YES" Or val = "1")
End Function

' === Get LCT Maintenance Start Date ===
Function GetLCTMaintStart() As Date
    On Error Resume Next
    GetLCTMaintStart = Sheets("Control_Panel").Range("H10").Value
    If Err.Number <> 0 Then GetLCTMaintStart = #1/1/2099#
    On Error GoTo 0
End Function

' === Get LCT Maintenance End Date ===
Function GetLCTMaintEnd() As Date
    On Error Resume Next
    GetLCTMaintEnd = Sheets("Control_Panel").Range("H11").Value
    If Err.Number <> 0 Then GetLCTMaintEnd = #1/1/2099#
    On Error GoTo 0
End Function

' === Highlight LCT Maintenance Period in Gantt ===
Sub HighlightLCTMaintenance()
    Dim ws As Worksheet
    Dim j As Long, lastCol As Long, maxJ As Long, lastRow As Long
    Dim projStart As Date, cellDate As Date, dc As Long
    Dim maintStart As Date, maintEnd As Date
    
    Set ws = Sheets("Gantt_Chart_Mammoet_Orig")
    projStart = Sheets("Control_Panel").Range("B4").Value
    maintStart = GetLCTMaintStart()
    maintEnd = GetLCTMaintEnd()
    dc = 8
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    ' Highlight maintenance period with gray
    For j = 0 To maxJ
        cellDate = projStart + j
        If cellDate >= maintStart And cellDate <= maintEnd Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Interior.Color = RGB(200, 200, 200)
            ws.Cells(3, dc + j).Value = "MAINT"
            ws.Cells(3, dc + j).Font.Bold = True
            ws.Cells(3, dc + j).Font.Size = 7
            ws.Cells(3, dc + j).Font.Color = RGB(128, 0, 0)
        End If
    Next j
    
    MsgBox "🔧 LCT Maintenance 기간 강조 완료:" & vbCrLf & _
           Format(maintStart, "YYYY-MM-DD") & " ~ " & Format(maintEnd, "YYYY-MM-DD"), _
           vbInformation, "LCT Maintenance"
End Sub

' === Display Current Control Panel Settings ===
Sub ShowControlPanelSettings()
    Dim msg As String
    
    msg = "📋 현재 Control Panel 설정:" & vbCrLf & vbCrLf & _
          "📅 Project Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
          "🎯 Target End: " & Format(Sheets("Control_Panel").Range("B5").Value, "YYYY-MM-DD") & vbCrLf & _
          "🚢 Voyage Pattern: " & GetVoyagePattern() & vbCrLf & _
          "🔄 Early Return: " & IIf(IsEarlyReturn(), "YES", "NO") & vbCrLf & vbCrLf & _
          "🌊 Shamal Period: " & Format(Sheets("Control_Panel").Range("H5").Value, "MM/DD") & _
          " ~ " & Format(Sheets("Control_Panel").Range("H6").Value, "MM/DD") & vbCrLf & _
          "🌊 Tide Threshold: " & Format(Sheets("Control_Panel").Range("H7").Value, "0.00") & "m" & vbCrLf & vbCrLf & _
          "🔧 LCT Maintenance: " & Format(GetLCTMaintStart(), "MM/DD") & _
          " ~ " & Format(GetLCTMaintEnd(), "MM/DD")
    
    MsgBox msg, vbInformation, "Control Panel Settings"
End Sub
'''
    
    for i, line in enumerate(vba_code.strip().split('\n'), 3):
        ws_vba.cell(i, 1, value=line)
        ws_vba.cell(i, 1).font = Font(name="Consolas", size=9)
    
    ws_vba.column_dimensions["A"].width = 100
    
    return wb

if __name__ == "__main__":
    import argparse
    import sys
    from pathlib import Path
    from datetime import datetime

    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    def _parse_iso_date(value: str) -> dt.date:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError as e:
            raise argparse.ArgumentTypeError("Expected date format YYYY-MM-DD") from e

    def _first_existing(base: Path, *names: str):
        for name in names:
            p = base / name
            if p.exists():
                return p
        return None

    script_dir = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()

    defaults = {
        "mamm_orig": _first_existing(
            script_dir,
            "MOS_with_NOTES.tsv",  # 추가: _with_NOTES 우선
            "MOS.tsv",
            "Mammoet_original_schedule.tsv",
            "Mammoet original schedule.tsv",
        ),
        "mamm_a": _first_existing(
            script_dir,
            "MFC1_with_NOTES.tsv",  # 추가: _with_NOTES 우선
            "MFC1.tsv",
            "Mammoet format_ScenarioA_1x1x1x1x1x1x1.tsv",
        ),
        "mamm_alt": _first_existing(
            script_dir,
            "MFA2_with_NOTES.tsv",  # 추가: _with_NOTES 우선
            "MFA2.tsv",
            "Mammoet format_Alternative_2-2-2-1_Two SPMTs.tsv",
            "Mammoet formatAlternative_2-2-2-1_Two SPMTs.tsv",
        ),
        "mobilization": _first_existing(script_dir, "MOBILIZATION.tsv"),
        "tide_tsv": _first_existing(script_dir, "Date High Tide Window Max Height (m) Ris.tsv"),
        "tide_json": _first_existing(script_dir, "MINA ZAYED PORT WATER TIDE_MERGED.json"),
        "generic_tsv": _first_existing(script_dir, "ID WBS Task Phase Owner Start End Durati.tsv"),
    }

    p = argparse.ArgumentParser(
        description="Generate AGI TR Multi-Scenario Master Gantt (Excel) from TSV inputs."
    )

    p.add_argument("--original", default=str(defaults["mamm_orig"]) if defaults["mamm_orig"] else None, help="Mammoet Original TSV path")
    p.add_argument("--scenario-a", dest="scenario_a", default=str(defaults["mamm_a"]) if defaults["mamm_a"] else None, help="Mammoet ScenarioA TSV path")
    p.add_argument("--alternative", default=str(defaults["mamm_alt"]) if defaults["mamm_alt"] else None, help="Mammoet Alternative TSV path")
    p.add_argument("--mobilization", default=str(defaults["mobilization"]) if defaults["mobilization"] else None, help="Mobilization TSV path (fallback)")
    p.add_argument("--tide-tsv", dest="tide_tsv", default=str(defaults["tide_tsv"]) if defaults["tide_tsv"] else None, help="Tide TSV path")
    p.add_argument("--tide-json", dest="tide_json", default=str(defaults["tide_json"]) if defaults["tide_json"] else None, help="Tide JSON path")
    p.add_argument("--tsv", dest="tsv_path", default=str(defaults["generic_tsv"]) if defaults["generic_tsv"] else None, help="Generic TSV path (optional)")

    p.add_argument("--project-start", type=_parse_iso_date, default=None, help="Override project start date (YYYY-MM-DD). If omitted, auto-detect from TSVs.")

    p.add_argument("--output-dir", default=str(script_dir), help="Output directory (default: script directory)")
    p.add_argument("--output", default=None, help="Output .xlsx path (overrides --output-dir)")

    p.add_argument(
        "--export-vba-bas",
        action="store_true",
        help="Export VBA code to a .bas file next to the output workbook (for easy import in VBA editor).",
    )
    p.add_argument(
        "--create-xlsm",
        action="store_true",
        help="(Windows/Excel only) After generating .xlsx, import VBA and save as .xlsm using Excel automation.",
    )

    args = p.parse_args()

    print("Generating AGI TR Multi-Scenario Master Gantt...")

    wb = create_gantt_with_vba(
        tsv_path=args.tsv_path,
        mammoet_original_tsv=args.original,
        mammoet_scenarioa_tsv=args.scenario_a,
        mammoet_alternative_tsv=args.alternative,
        mobilization_tsv=args.mobilization,
        tide_tsv=args.tide_tsv,
        tide_json=args.tide_json,
        project_start_override=args.project_start,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        out_dir = Path(args.output_dir).expanduser().resolve()
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"AGI_TR_MultiScenario_Master_Gantt_{timestamp}.xlsx"

    wb.save(str(output_path))
    print(f"[OK] Generated: {output_path}")

    bas_path = None
    if args.export_vba_bas:
        try:
            ws_vba = wb["VBA_Code"]
            lines = []
            for row in ws_vba.iter_rows(min_row=3, max_col=1, values_only=True):
                if row and row[0] is not None:
                    lines.append(str(row[0]))
            bas_path = output_path.with_suffix(".bas")
            bas_path.write_text("\n".join(lines), encoding="utf-8")
            print(f"[OK] Exported VBA module: {bas_path}")
        except Exception as e:
            print(f"[WARN] Could not export VBA .bas: {e}")

    if args.create_xlsm:
        # Requires: Windows + Excel installed + Trust access to VBProject enabled.
        if sys.platform != "win32":
            print("[WARN] --create-xlsm is only supported on Windows with Excel installed.")
        else:
            try:
                import win32com.client  # type: ignore
            except Exception:
                print("[WARN] pywin32(win32com) is not installed. Install: pip install pywin32")
            else:
                try:
                    if bas_path is None:
                        ws_vba = wb["VBA_Code"]
                        lines = []
                        for row in ws_vba.iter_rows(min_row=3, max_col=1, values_only=True):
                            if row and row[0] is not None:
                                lines.append(str(row[0]))
                        bas_path = output_path.with_suffix(".bas")
                        bas_path.write_text("\n".join(lines), encoding="utf-8")

                    xlsm_path = output_path.with_suffix(".xlsm")
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb_xl = excel.Workbooks.Open(str(output_path))
                    wb_xl.VBProject.VBComponents.Import(str(bas_path))
                    wb_xl.SaveAs(str(xlsm_path), FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
                    wb_xl.Close(SaveChanges=True)
                    excel.Quit()
                    print(f"[OK] Generated .xlsm with VBA: {xlsm_path}")
                except Exception as e:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                    print(f"[WARN] Could not create .xlsm via Excel automation: {e}")
