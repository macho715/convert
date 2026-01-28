#!/usr/bin/env python3
"""
AGI HVDC Transformer Transportation – 4 Voyages Master Gantt (with VBA helper sheet)

Scope (AGI Site):
- Voyage pattern: V1=1 unit, V2=2 units, V3=2 units, V4=2 units (total 7 transformers)
- Onsite JACK-UP batches: 3 events (3 units, 2 units, 2 units)
- Load-out / Unload durations are scaled by number of units per voyage:
    * 1 unit  -> DUR_LO, DUR_UL
    * 2 units -> DUR_LO*2, DUR_UL*2

Note:
- This workbook is generated as .xlsx (macros are provided as text in VBA_Code sheet).
  To use macros: copy VBA code into Excel (Alt+F11) and save as .xlsm.
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# === COLORS ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

BORDER = Side(style="thin", color="A6A6A6")

def tb():
    return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def dur_scaled(base_name: str, units: int) -> str:
    """Return Excel formula string for duration scaling by units."""
    if units == 1:
        return f"={base_name}"
    return f"={base_name}*{units}"

def create_tasks_for_voyage(
    voyage_num,
    units,
    offset,
    tr_start,
    *,
    include_turning=True,
    early_return=True,
):
    """Create task list for a single voyage and return next offset."""
    tasks = []
    wbs_base = voyage_num + 1.0

    tr_list = "+".join([f"TR{i}" for i in range(tr_start, tr_start + units)])
    unit_label = f"{units} unit{'s' if units > 1 else ''}"
    tasks.append((
        f"V{voyage_num}",
        f"{wbs_base}.0",
        f"VOYAGE {voyage_num}: {tr_list} ({unit_label})",
        "MILESTONE",
        "All",
        offset,
        0,
        "Tide >=1.90m required (LO/ARR)" if voyage_num == 1 else "",
    ))

    lo_offset = offset
    tasks.append((
        f"LO-{voyage_num}01",
        f"{wbs_base}.1",
        f"Load-out: {tr_list} on LCT ({unit_label})",
        "LOADOUT",
        "Mammoet",
        lo_offset,
        None,
        f"Applied: DUR_LO ×{units}" if units > 1 else "Baseline (1 unit)",
    ))

    sf_offset = lo_offset + (1 if units == 1 else 2)
    tasks.append((
        f"SF-{voyage_num}02",
        f"{wbs_base}.2",
        f"Sea Fastening: {tr_list}",
        "SEAFAST",
        "Mammoet",
        sf_offset,
        "DUR_SF",
        "12-point lashing",
    ))

    mws_offset = sf_offset
    tasks.append((
        f"MWS-{voyage_num}03",
        f"{wbs_base}.3",
        f"MWS + MPI + Final Check (V{voyage_num})",
        "BUFFER",
        "Aries/Captain",
        mws_offset,
        "DUR_MWS",
        "Pre-sail verification",
    ))

    sail_offset = sf_offset + 1
    tasks.append((
        f"SAIL-{voyage_num}04",
        f"{wbs_base}.4",
        f"Sail-away V{voyage_num}: MZP->AGI",
        "SAIL",
        "LCT Bushra",
        sail_offset,
        "DUR_SAIL",
        "Weather window required",
    ))

    arr_offset = sail_offset + 1
    if units == 1:
        tasks.append((
            f"ARR-{voyage_num}05",
            f"{wbs_base}.5",
            f"Arrival + RORO Unload: TR{tr_start} (1 unit)",
            "AGI_UNLOAD",
            "Mammoet",
            arr_offset,
            "DUR_UL",
            "Baseline (1 unit)",
        ))
        unload_end_offset = arr_offset
    else:
        for i, tr_num in enumerate(range(tr_start, tr_start + units)):
            tasks.append((
                f"UL-{voyage_num}-{i+1}",
                f"{wbs_base}.5{i+1}",
                f"Unload TR{tr_num} at AGI (1 unit/day)",
                "AGI_UNLOAD",
                "Mammoet",
                arr_offset + i,
                "DUR_UL",
                "RORO + ramp",
            ))
        unload_end_offset = arr_offset + units - 1

    next_voyage_offset = unload_end_offset + 3
    return_offset = unload_end_offset + 1

    if include_turning:
        turn_offset = unload_end_offset + 1
        tasks.append((
            f"TURN-{voyage_num}-1",
            f"{wbs_base}.61",
            f"Turning TR{tr_start} (90 deg)",
            "TURNING",
            "Mammoet",
            turn_offset,
            3,
            "3.0d/unit",
        ))

        first_jd_offset = turn_offset + 3
        tasks.append((
            f"JD-{voyage_num}-1",
            f"{wbs_base}.71",
            f"Jackdown TR{tr_start}",
            "JACKDOWN",
            "Mammoet",
            first_jd_offset,
            "DUR_JD",
            "1.0d/unit",
        ))

        if units > 1:
            if early_return:
                return_offset = first_jd_offset + 1
                tasks.append((
                    f"RET-{voyage_num}06",
                    f"{wbs_base}.8",
                    f"LCT Return V{voyage_num}: AGI->MZP (After first JD)",
                    "RETURN",
                    "LCT Bushra",
                    return_offset,
                    "DUR_RET",
                    "SPMT reloaded for next voyage",
                ))

            turn_offset_2 = first_jd_offset + 1
            tasks.append((
                f"TURN-{voyage_num}-2",
                f"{wbs_base}.62",
                f"Turning TR{tr_start + 1} (90 deg)",
                "TURNING",
                "Mammoet",
                turn_offset_2,
                3,
                "3.0d/unit",
            ))

            second_jd_offset = turn_offset_2 + 3
            tasks.append((
                f"JD-{voyage_num}-2",
                f"{wbs_base}.72",
                f"Jackdown TR{tr_start + 1}",
                "JACKDOWN",
                "Mammoet",
                second_jd_offset,
                "DUR_JD",
                "1.0d/unit",
            ))

            if not early_return:
                return_offset = second_jd_offset + 1
                tasks.append((
                    f"RET-{voyage_num}06",
                    f"{wbs_base}.8",
                    f"LCT Return V{voyage_num}: AGI->MZP",
                    "RETURN",
                    "LCT Bushra",
                    return_offset,
                    "DUR_RET",
                    "After final JD",
                ))
        else:
            return_offset = first_jd_offset + 1
            tasks.append((
                f"RET-{voyage_num}06",
                f"{wbs_base}.8",
                f"LCT Return V{voyage_num}: AGI->MZP",
                "RETURN",
                "LCT Bushra",
                return_offset,
                "DUR_RET",
                "After JD",
            ))

        if return_offset is not None:
            next_voyage_offset = return_offset + 2
    else:
        return_offset = unload_end_offset + 1
        tasks.append((
            f"RET-{voyage_num}06",
            f"{wbs_base}.8",
            f"LCT Return V{voyage_num}: AGI->MZP",
            "RETURN",
            "LCT Bushra",
            return_offset,
            "DUR_RET",
            "Return after unload",
        ))
        next_voyage_offset = return_offset + 2

    tasks.append((
        f"BUF-{voyage_num}07",
        f"{wbs_base}.99",
        f"Buffer / Reset (V{voyage_num})",
        "BUFFER",
        "All",
        next_voyage_offset - 1,
        "DUR_BUF",
        "Contingency",
    ))

    return tasks, next_voyage_offset


def create_gantt_with_vba() -> Workbook:
    wb = Workbook()

    # ============================================================
    # CONTROL PANEL
    # ============================================================
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"

    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "AGI TR Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30

    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "Start date (B4) change will shift all schedules (VBA macro required for auto-refresh)."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")

    # Input dates
    ws_ctrl["A4"] = "Project Start:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)

    ws_ctrl["B4"] = dt.date(2026, 1, 18)
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)

    ws_ctrl["A5"] = "Target End:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = dt.date(2026, 2, 28)
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()

    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")

    # Duration parameters (per-unit baseline)
    ws_ctrl["D4"] = "Durations (days)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)

    durations = [
        ("D5", "Mobilization:", "E5", 1.0, "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", 3.0, "DUR_DECK"),
        ("D7", "Load-out (per unit):", "E7", 1.0, "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", 0.5, "DUR_SF"),
        ("D9", "MWS Approval:", "E9", 0.5, "DUR_MWS"),
        ("D10", "Sailing:", "E10", 1.0, "DUR_SAIL"),
        ("D11", "Unload (per unit):", "E11", 1.0, "DUR_UL"),
        ("D12", "Jack-up/down:", "E12", 1.0, "DUR_JD"),
        ("D13", "Return:", "E13", 1.0, "DUR_RET"),
        ("D14", "Buffer:", "E14", 0.5, "DUR_BUF"),
    ]

    for lc, lt, vc, v, name in durations:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = v
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")

    # Weather settings (Shamal window)
    ws_ctrl["G4"] = "Weather (Shamal)"
    ws_ctrl["G4"].font = Font(bold=True, size=12)
    ws_ctrl["G5"] = "Shamal Start:"
    ws_ctrl["H5"] = dt.date(2026, 2, 5)
    ws_ctrl["H5"].number_format = "YYYY-MM-DD"
    ws_ctrl["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G6"] = "Shamal End:"
    ws_ctrl["H6"] = dt.date(2026, 2, 14)
    ws_ctrl["H6"].number_format = "YYYY-MM-DD"
    ws_ctrl["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    wb.defined_names["SHAMAL_START"] = DefinedName("SHAMAL_START", attr_text="Control_Panel!$H$5")
    wb.defined_names["SHAMAL_END"] = DefinedName("SHAMAL_END", attr_text="Control_Panel!$H$6")

    # LCT Maintenance Information
    ws_ctrl["G8"] = "LCT Maintenance"
    ws_ctrl["G8"].font = Font(bold=True, size=12)
    ws_ctrl["G9"] = "Maintenance Start:"
    ws_ctrl["H9"] = dt.date(2026, 3, 1)
    ws_ctrl["H9"].number_format = "YYYY-MM-DD"
    ws_ctrl["H9"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G10"] = "Duration (days):"
    ws_ctrl["H10"] = 7
    ws_ctrl["H10"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["H10"].number_format = "0"
    ws_ctrl["G10"].font = Font(bold=True)

    # Summary formulas
    ws_ctrl["A8"] = "Auto Summary"
    ws_ctrl["A8"].font = Font(bold=True, size=12)

    summary_items = [
        ("A9", "Estimated Finish:", "B9", "=MAX(Schedule_Data!G:G)"),
        ("A10", "Total Duration (days):", "B10", "=B9-B4+1"),
        ("A11", "Against Target:", "B11", '=IF(B9<=B5,"ON TRACK","DELAY")'),
        ("A12", "Float (days):", "B12", "=B5-B9"),
    ]

    for lc, lt, vc, formula in summary_items:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if "MAX" in formula:
            ws_ctrl[vc].number_format = "YYYY-MM-DD"

    # Column widths
    ws_ctrl.column_dimensions["A"].width = 20
    ws_ctrl.column_dimensions["B"].width = 15
    ws_ctrl.column_dimensions["D"].width = 22
    ws_ctrl.column_dimensions["E"].width = 10
    ws_ctrl.column_dimensions["G"].width = 14
    ws_ctrl.column_dimensions["H"].width = 12

    # ============================================================
    # SCHEDULE DATA
    # ============================================================
    ws_sched = wb.create_sheet("Schedule_Data")

    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = "AGI HVDC Transformer Transportation Master Schedule (4 Voyages)"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])

    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = "V1=1 unit | V2=2 units | V3=2 units | V4=2 units | Onsite Turning/Jackdown per transformer"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    ws_sched.merge_cells("A3:I3")
    ws_sched["A3"] = "Shamal risk: Feb 5-14, 2026 | Tide gate >=1.90m | Load-out/Unload scaled by units"

    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()

    # ------------------------------------------------------------
    # TASK LIST (Offsets are relative to PROJECT_START)

    tasks = [
        # Mobilization / common deck prep
        ("MOB-001", "1.0", "MOBILIZATION", "MOBILIZATION", "Mammoet", 0, "DUR_MOB", "SPMT assembly + marine mobilization"),
        ("PREP-001", "1.1", "Deck Preparations", "DECK_PREP", "Mammoet", 1, "DUR_DECK", "One-time setup for all voyages"),
    ]

    offset = 0
    voyage_specs = [
        (1, 1, 1),
        (2, 2, 2),
        (3, 2, 4),
        (4, 2, 6),
    ]
    for voyage_num, units, tr_start in voyage_specs:
        voyage_tasks, offset = create_tasks_for_voyage(
            voyage_num,
            units,
            offset,
            tr_start,
            include_turning=True,
            early_return=True,
        )
        tasks.extend(voyage_tasks)

    # Demobilization / close-out
    tasks.extend([
        ("DEMOB-001", "9.0", "DEMOBILIZATION", "MOBILIZATION", "Mammoet", offset, "DUR_MOB", "Equipment return"),
        ("END", "99.0", "PROJECT COMPLETE", "MILESTONE", "All", offset, 0, "All transformers delivered and installed"),
    ])

    # LCT Maintenance (March 2026) - Added per MoM
    # Calculate offset for March 1, 2026 from PROJECT_START baseline (2026-01-18)
    project_start_baseline = dt.date(2026, 1, 18)
    lct_maintenance_start = dt.date(2026, 3, 1)
    lct_maintenance_offset = (lct_maintenance_start - project_start_baseline).days

    tasks.append((
        "LCT-MAINT-001",
        "10.0",
        "LCT Bushra Maintenance (March 2026)",
        "BUFFER",
        "ADNOC L&S",
        lct_maintenance_offset,
        7,  # Estimated 7 days maintenance period
        "Scheduled maintenance - may affect V4 if project delayed"
    ))

    # Write tasks
    for r, t in enumerate(tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t

        ws_sched.cell(r, 1, value=tid)
        ws_sched.cell(r, 2, value=wbs)
        ws_sched.cell(r, 3, value=task)
        ws_sched.cell(r, 4, value=phase)
        ws_sched.cell(r, 5, value=owner)

        # Start = PROJECT_START + offset
        ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"

        # Duration
        # If dur_ref is None, it means scaled (LO/UL) – decide units by task label
        if dur_ref is None:
            # Determine scaling by checking (2 units) in task name; default 2
            units = 2 if "(2 units)" in task else 1
            base = "DUR_LO" if "Load-out" in task else "DUR_UL"
            ws_sched.cell(r, 8, value=dur_scaled(base, units))
        elif isinstance(dur_ref, str):
            ws_sched.cell(r, 8, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 8, value=dur_ref)

        # End = Start + Duration
        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"

        ws_sched.cell(r, 9, value=notes)

        # Styling
        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 10):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)

        if phase == "MILESTONE":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")

    # Column widths
    col_widths = {"A": 10, "B": 6, "C": 44, "D": 14, "E": 14, "F": 12, "G": 12, "H": 10, "I": 40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"

    # ============================================================
    # GANTT CHART (42 days horizon)
    # ============================================================
    ws_gantt = wb.create_sheet("Gantt_Chart")

    ws_gantt.merge_cells("A1:CA1")
    ws_gantt["A1"] = "AGI HVDC Master Gantt Chart (Auto-Updated)"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])

    ws_gantt.merge_cells("A2:CA2")
    ws_gantt["A2"] = "Orange columns = Shamal risk window (Feb 5-14) | LCT Maintenance: Mar 1-7, 2026 | VBA macro RefreshGanttChart required"
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    ws_gantt["A3"] = "Jan/Feb 2026"
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:G3")

    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()

    # Date columns (42 days)
    date_col = 8  # H
    horizon = 42
    for i in range(horizon):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5

    # Task rows: link to Schedule_Data
    for r in range(5, 5 + len(tasks)):
        src_row = r + 1  # Schedule_Data starts at row 6
        ws_gantt.cell(r, 1, value=f"=Schedule_Data!A{src_row}")
        ws_gantt.cell(r, 2, value=f"=Schedule_Data!B{src_row}")
        ws_gantt.cell(r, 3, value=f"=Schedule_Data!C{src_row}")
        ws_gantt.cell(r, 4, value=f"=Schedule_Data!D{src_row}")

        start_cell = ws_gantt.cell(r, 5, value=f"=Schedule_Data!F{src_row}")
        start_cell.number_format = "MM/DD"

        end_cell = ws_gantt.cell(r, 6, value=f"=Schedule_Data!G{src_row}")
        end_cell.number_format = "MM/DD"

        ws_gantt.cell(r, 7, value=f"=Schedule_Data!H{src_row}")

        # Borders
        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()
        for i in range(horizon):
            ws_gantt.cell(r, date_col + i).border = tb()

        # Phase color
        phase = ws_gantt.cell(r, 4).value
        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)

    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 30
    ws_gantt.column_dimensions["D"].width = 12
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4
    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)

    # ============================================================
    # SUMMARY (simple)
    # ============================================================
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "AGI HVDC Transformer Transportation - Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")

    summary_data = [
        ("Key Parameters", ""),
        ("Total Transformers", "7 units (TR1-TR7)"),
        ("Voyage Pattern", "V1=1 unit, V2=2 units, V3=2 units, V4=2 units"),
        ("Total Voyages", "4"),
        ("Onsite Jack-up Events", "3 (3 units, 2 units, 2 units)"),
        ("Vessel", "LCT BUSHRA"),
        ("Route", "Mina Zayed Port ↔ AGI Site"),
        ("", ""),
        ("Schedule Summary", ""),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
        ("Voyage 1 (TR1)", "Offset 0-5"),
        ("Voyage 2 (TR2+TR3)", "Offset 8-15"),
        ("Jack-up #1 (TR1-TR3)", "Offset 17"),
        ("Voyage 3 (TR4+TR5)", "Offset 28-35"),
        ("Jack-up #2 (TR4-TR5)", "Offset 37"),
        ("Voyage 4 (TR6+TR7)", "Offset 36-41"),
        ("Jack-up #3 (TR6-TR7)", "Offset 41"),
        ("Project Complete", "=MAX(Schedule_Data!G:G)"),
        ("", ""),  # Added separator
        ("LCT Maintenance", ""),  # Added section
        ("LCT Bushra Maintenance", "March 1, 2026 (7 days)"),
        ("Note", "May affect V4 if project delayed beyond Feb 28"),
    ]

    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 40

    # ============================================================
    # WEATHER ANALYSIS (informative)
    # ============================================================
    ws_weather = wb.create_sheet("Weather_Analysis")
    ws_weather["A1"] = "UAE Winter Weather Analysis - Jan/Feb 2026"
    ws_weather["A1"].font = Font(bold=True, size=14)
    ws_weather.merge_cells("A1:F1")

    weather_headers = ["Parameter", "Jan 1-10", "Jan 11-20", "Jan 21-31", "Feb 1-15", "Notes"]
    for c, h in enumerate(weather_headers, 1):
        cell = ws_weather.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = tb()

    weather_data = [
        ("Avg Wind (kt)", "11-13", "16-21", "13-15", "12-14", "Peak Shamal: Feb 5-14"),
        ("Max Gust (kt)", "18-20", "25-30", "20-22", "18-20", "NO-GO if >22kt gust"),
        ("Wave Height (m)", "0.4-0.6", "0.8-1.2", "0.5-0.7", "0.4-0.6", "HOLD if >0.8m"),
        ("Visibility (km)", "8-10", "2-5", "6-8", "8-10", "Reduced during Shamal"),
        ("Risk Level", "LOW", "HIGH", "MEDIUM", "LOW", ""),
        ("Recommendation", "GO", "NO-GO", "CAUTION", "GO", ""),
    ]

    for r, row in enumerate(weather_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws_weather.cell(r, c, value=val)
            cell.border = tb()
            if val in ("HIGH", "NO-GO"):
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif val in ("MEDIUM", "CAUTION"):
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
            elif val in ("LOW", "GO"):
                cell.fill = PatternFill("solid", fgColor="C8E6C9")

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_weather.column_dimensions[col].width = 16

    # ============================================================
    # VBA CODE (text only)
    # ============================================================
    ws_vba = wb.create_sheet("VBA_Code")
    ws_vba["A1"] = "VBA Code: copy into Excel VBA Module (Alt+F11) and save as .xlsm"
    ws_vba["A1"].font = Font(bold=True, size=14)

    vba_code = r'''
Option Explicit

' ============================================
' AGI TR 4-Voyage Master Gantt - VBA Macros
' ============================================

Sub UpdateAllSchedules()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual

    Sheets("Schedule_Data").Calculate
    Sheets("Gantt_Chart").Calculate
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate

    Call RefreshGanttChart

    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True

    MsgBox "Schedule updated." & vbCrLf & _
           "Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
           "Finish: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD"), _
           vbInformation, "Schedule"
End Sub

Sub RefreshGanttChart()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long
    Dim shamalStart As Date, shamalEnd As Date

    Set ws = Sheets("Gantt_Chart")
    Set wsd = Sheets("Schedule_Data")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8 ' date columns start at H

    Application.ScreenUpdating = False

    ' Clear date area colors
    ws.Range(ws.Cells(5, dc), ws.Cells(200, dc + 41)).Interior.ColorIndex = xlNone

    ' Reset header colors and highlight Shamal in header
    For j = 0 To 41
        cellDate = projStart + j
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121) ' header blue
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0) ' orange
        End If
        ' Highlight LCT Maintenance period (March 1-7, 2026)
        If cellDate >= DateSerial(2026, 3, 1) And cellDate <= DateSerial(2026, 3, 7) Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 193, 7) ' amber/yellow for maintenance
        End If
    Next j

    ' Draw gantt bars
    For i = 5 To 5 + 60
        If IsDate(wsd.Cells(i + 1, 6).Value) Then
            startD = wsd.Cells(i + 1, 6).Value
            endD = wsd.Cells(i + 1, 7).Value
            phase = wsd.Cells(i + 1, 4).Value

            For j = 0 To 41
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(i, dc + j).Interior.Color = GetPhaseColor(phase)
                End If
            Next j
        End If
    Next i

    Application.ScreenUpdating = True
End Sub

Function GetPhaseColor(phase As String) As Long
    Select Case phase
        Case "MOBILIZATION": GetPhaseColor = RGB(142, 124, 195)
        Case "DECK_PREP": GetPhaseColor = RGB(111, 168, 220)
        Case "LOADOUT": GetPhaseColor = RGB(147, 196, 125)
        Case "SEAFAST": GetPhaseColor = RGB(118, 165, 175)
        Case "SAIL": GetPhaseColor = RGB(164, 194, 244)
        Case "AGI_UNLOAD": GetPhaseColor = RGB(246, 178, 107)
        Case "TURNING": GetPhaseColor = RGB(255, 217, 102)  ' Added for Turning phase
        Case "JACKDOWN": GetPhaseColor = RGB(224, 102, 102)
        Case "RETURN": GetPhaseColor = RGB(153, 153, 153)
        Case "BUFFER": GetPhaseColor = RGB(217, 217, 217)
        Case "MILESTONE": GetPhaseColor = RGB(255, 0, 0)
        Case Else: GetPhaseColor = RGB(255, 255, 255)
    End Select
End Function
'''
    for i, line in enumerate(vba_code.strip().splitlines(), 3):
        ws_vba.cell(i, 1, value=line)
        ws_vba.cell(i, 1).font = Font(name="Consolas", size=9)

    ws_vba.column_dimensions["A"].width = 120

    return wb


if __name__ == "__main__":
    import os
    print("Generating AGI TR 4-Voyage Master Gantt (LO/UL x2 for 2-unit voyages)...")
    wb = create_gantt_with_vba()
    out = os.path.join(os.getcwd(), "AGI_TR_4Voyage_Master_Gantt_VBA_LOLIx2.xlsx")
    wb.save(out)
    print(f"[OK] Saved: {out}")
