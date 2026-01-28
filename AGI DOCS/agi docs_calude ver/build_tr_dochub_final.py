#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MACHO-GPT AUTO-GENERATED CODE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Project: HVDC Logistics Automation
Module: TR_DocHub_AGI_2026 Final Builder
Created: 2026-01-19
Python: 3.11 | Excel: LTSC 2021
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Purpose: Build comprehensive TR Document Tracker with VBA modules
Dependencies: openpyxl, pandas
Input: None (creates from scratch)
Output: TR_DocHub_AGI_2026_Final.xlsm
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations
import datetime as dt
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ============================================================================
# VBA MODULES CONTENT
# ============================================================================

VBA_MOD_OPERATIONS = '''Attribute VB_Name = "modOperations"
Option Explicit

'===============================================================================
' OPERATIONS MODULE - Minimum Essential Functions for TR_DocHub_AGI
'===============================================================================

Public Sub InitializeWorkbook()
    On Error GoTo EH
    Application.ScreenUpdating = False
    
    Dim ws As Worksheet
    Dim required_sheets As Variant
    required_sheets = Array("S_Voyages", "M_DocCatalog", "M_Parties", "R_DeadlineRules", "T_Tracker", "D_Dashboard")
    
    Dim sh As Variant
    For Each sh In required_sheets
        On Error Resume Next
        Set ws = ThisWorkbook.Sheets(CStr(sh))
        On Error GoTo 0
        If ws Is Nothing Then
            MsgBox "Required sheet missing: " & sh, vbCritical
            Exit Sub
        End If
        Set ws = Nothing
    Next sh
    
    Application.ScreenUpdating = True
    MsgBox "Workbook initialized successfully.", vbInformation
    Exit Sub
    
EH:
    Application.ScreenUpdating = True
    MsgBox "Initialize failed: " & Err.Description, vbCritical
End Sub

Public Sub GenerateTrackerRows(Optional voyageID As String = "ALL")
    On Error GoTo EH
    
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    
    Dim ws_voy As Worksheet, ws_doc As Worksheet, ws_tracker As Worksheet
    Dim lo_voy As ListObject, lo_doc As ListObject, lo_tracker As ListObject
    
    Set ws_voy = ThisWorkbook.Sheets("S_Voyages")
    Set ws_doc = ThisWorkbook.Sheets("M_DocCatalog")
    Set ws_tracker = ThisWorkbook.Sheets("T_Tracker")
    
    Set lo_voy = ws_voy.ListObjects("tbl_Voyage")
    Set lo_doc = ws_doc.ListObjects("tbl_DocCatalog")
    Set lo_tracker = ws_tracker.ListObjects("tbl_Tracker")
    
    Dim existing_keys As Object
    Set existing_keys = CreateObject("Scripting.Dictionary")
    
    If Not lo_tracker.DataBodyRange Is Nothing Then
        Dim r As Long
        For r = 1 To lo_tracker.DataBodyRange.Rows.Count
            Dim v_id As String, d_code As String
            v_id = CStr(lo_tracker.DataBodyRange.Cells(r, 1).Value)
            d_code = CStr(lo_tracker.DataBodyRange.Cells(r, 2).Value)
            If v_id <> "" And d_code <> "" Then
                existing_keys(v_id & "|" & d_code) = True
            End If
        Next r
    End If
    
    Dim new_rows As Long: new_rows = 0
    Dim voy_row As ListRow, doc_row As ListRow
    
    For Each voy_row In lo_voy.ListRows
        Dim v_id_val As String
        v_id_val = CStr(voy_row.Range(1).Value)
        If voyageID <> "ALL" And v_id_val <> voyageID Then GoTo NextVoyage
        
        For Each doc_row In lo_doc.ListRows
            Dim d_code_val As String, req_flag As String, active_flag As String, default_party As String
            d_code_val = CStr(doc_row.Range(1).Value)
            req_flag = UCase(CStr(doc_row.Range(5).Value))
            active_flag = UCase(CStr(doc_row.Range(7).Value))
            default_party = CStr(doc_row.Range(4).Value)
            
            If req_flag <> "Y" Or active_flag <> "Y" Then GoTo NextDoc
            
            Dim key As String: key = v_id_val & "|" & d_code_val
            If existing_keys.Exists(key) Then GoTo NextDoc
            
            Dim new_row As ListRow
            Set new_row = lo_tracker.ListRows.Add
            new_row.Range(1).Value = v_id_val
            new_row.Range(2).Value = d_code_val
            new_row.Range(3).Value = default_party
            new_row.Range(8).Value = "Not Started"
            
            existing_keys(key) = True
            new_rows = new_rows + 1
NextDoc:
        Next doc_row
NextVoyage:
    Next voy_row
    
    Application.EnableEvents = True
    Application.ScreenUpdating = True
    MsgBox "Generated " & new_rows & " new tracker rows.", vbInformation
    Exit Sub
    
EH:
    Application.EnableEvents = True
    Application.ScreenUpdating = True
    MsgBox "GenerateTrackerRows failed: " & Err.Description, vbCritical
End Sub

Public Sub RecalcDeadlines()
    On Error GoTo EH
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    Dim ws_tracker As Worksheet, ws_voy As Worksheet, ws_rules As Worksheet
    Dim lo_tracker As ListObject, lo_voy As ListObject, lo_rules As ListObject
    
    Set ws_tracker = ThisWorkbook.Sheets("T_Tracker")
    Set ws_voy = ThisWorkbook.Sheets("S_Voyages")
    Set ws_rules = ThisWorkbook.Sheets("R_DeadlineRules")
    
    Set lo_tracker = ws_tracker.ListObjects("tbl_Tracker")
    Set lo_voy = ws_voy.ListObjects("tbl_Voyage")
    Set lo_rules = ws_rules.ListObjects("tbl_RuleDeadline")
    
    If lo_tracker.DataBodyRange Is Nothing Then
        MsgBox "No tracker data found. Run GenerateTrackerRows first.", vbExclamation
        GoTo Cleanup
    End If
    
    Dim updated_count As Long: updated_count = 0
    Dim tr_row As ListRow
    
    For Each tr_row In lo_tracker.ListRows
        Dim v_id As String, d_code As String
        v_id = CStr(tr_row.Range(1).Value)
        d_code = CStr(tr_row.Range(2).Value)
        
        If v_id = "" Or d_code = "" Then GoTo NextTrackerRow
        
        Dim best_rule As ListRow: Set best_rule = Nothing
        Dim best_priority As Long: best_priority = 9999
        Dim rule_row As ListRow
        
        For Each rule_row In lo_rules.ListRows
            Dim rule_doc As String, rule_active As String, rule_priority As Long
            rule_doc = CStr(rule_row.Range(2).Value)
            rule_active = UCase(CStr(rule_row.Range(7).Value))
            On Error Resume Next
            rule_priority = CLng(rule_row.Range(6).Value)
            On Error GoTo 0
            
            If rule_doc = d_code And rule_active = "Y" And rule_priority < best_priority Then
                Set best_rule = rule_row
                best_priority = rule_priority
            End If
        Next rule_row
        
        If best_rule Is Nothing Then GoTo NextTrackerRow
        
        Dim anchor_field As String, offset_days As Long, cal_type As String
        anchor_field = CStr(best_rule.Range(3).Value)
        offset_days = CLng(best_rule.Range(4).Value)
        cal_type = CStr(best_rule.Range(5).Value)
        
        tr_row.Range(4).Value = anchor_field
        tr_row.Range(6).Value = offset_days
        
        Dim voy_match As ListRow: Set voy_match = Nothing
        For Each voy_match In lo_voy.ListRows
            If CStr(voy_match.Range(1).Value) = v_id Then Exit For
        Next voy_match
        
        If Not voy_match Is Nothing Then
            Dim anchor_col_idx As Long: anchor_col_idx = 0
            Dim col_idx As Long
            For col_idx = 1 To lo_voy.ListColumns.Count
                If lo_voy.ListColumns(col_idx).Name = anchor_field Then
                    anchor_col_idx = col_idx
                    Exit For
                End If
            Next col_idx
            
            If anchor_col_idx > 0 Then
                Dim anchor_date As Variant
                anchor_date = voy_match.Range(anchor_col_idx).Value
                If IsDate(anchor_date) Then
                    tr_row.Range(5).Value = anchor_date
                    
                    Dim due_date As Date
                    If UCase(cal_type) = "WD" Then
                        due_date = WorksheetFunction.WorkDay(CDate(anchor_date), offset_days)
                    Else
                        due_date = CDate(anchor_date) + offset_days
                    End If
                    tr_row.Range(7).Value = due_date
                    updated_count = updated_count + 1
                End If
            End If
        End If
NextTrackerRow:
    Next tr_row
    
Cleanup:
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "Updated " & updated_count & " deadline(s).", vbInformation
    Exit Sub
    
EH:
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "RecalcDeadlines failed: " & Err.Description, vbCritical
End Sub

Public Sub ValidateBeforeExport()
    On Error GoTo EH
    
    Dim errors As Object: Set errors = CreateObject("Scripting.Dictionary")
    Dim ws_tracker As Worksheet, ws_voy As Worksheet, ws_doc As Worksheet
    Dim lo_tracker As ListObject, lo_voy As ListObject, lo_doc As ListObject
    
    Set ws_tracker = ThisWorkbook.Sheets("T_Tracker")
    Set ws_voy = ThisWorkbook.Sheets("S_Voyages")
    Set ws_doc = ThisWorkbook.Sheets("M_DocCatalog")
    
    Set lo_tracker = ws_tracker.ListObjects("tbl_Tracker")
    Set lo_voy = ws_voy.ListObjects("tbl_Voyage")
    Set lo_doc = ws_doc.ListObjects("tbl_DocCatalog")
    
    Dim voy_row As ListRow
    For Each voy_row In lo_voy.ListRows
        Dim vid As String: vid = CStr(voy_row.Range(1).Value)
        If vid = "" Then GoTo NextVoyCheck
        
        Dim required_anchors As Variant
        required_anchors = Array(5, 6, 7, 8)
        
        Dim a As Variant
        For Each a In required_anchors
            If IsEmpty(voy_row.Range(CLng(a)).Value) Then
                errors.Add errors.Count + 1, "Voyage " & vid & ": Missing anchor column " & lo_voy.ListColumns(CLng(a)).Name
            End If
        Next a
NextVoyCheck:
    Next voy_row
    
    If lo_tracker.DataBodyRange Is Nothing Then
        errors.Add errors.Count + 1, "T_Tracker has no data rows"
    Else
        Dim tr_row As ListRow
        For Each tr_row In lo_tracker.ListRows
            Dim status_val As String, evid_link As String, d_code_val As String
            status_val = CStr(tr_row.Range(8).Value)
            evid_link = CStr(tr_row.Range(11).Value)
            d_code_val = CStr(tr_row.Range(2).Value)
            
            If status_val = "Submitted" And evid_link = "" Then
                Dim doc_row As ListRow
                For Each doc_row In lo_doc.ListRows
                    If CStr(doc_row.Range(1).Value) = d_code_val Then
                        If UCase(CStr(doc_row.Range(6).Value)) = "Y" Then
                            errors.Add errors.Count + 1, "Tracker: " & tr_row.Range(1).Value & "|" & d_code_val & " - Submitted but EvidenceLink missing"
                        End If
                        Exit For
                    End If
                Next doc_row
            End If
            
            If IsEmpty(tr_row.Range(7).Value) Or tr_row.Range(7).Value = "" Then
                errors.Add errors.Count + 1, "Tracker: " & tr_row.Range(1).Value & "|" & tr_row.Range(2).Value & " - DueDate empty"
            End If
        Next tr_row
    End If
    
    If errors.Count > 0 Then
        Dim msg As String: msg = "Validation found " & errors.Count & " issue(s):" & vbCrLf & vbCrLf
        Dim k As Variant
        For Each k In errors.Keys
            msg = msg & errors(k) & vbCrLf
        Next k
        MsgBox msg, vbExclamation, "Validation Errors"
    Else
        MsgBox "Validation passed. No errors found.", vbInformation
    End If
    Exit Sub
    
EH:
    MsgBox "ValidateBeforeExport failed: " & Err.Description, vbCritical
End Sub

Public Sub ExportVoyagePack(Optional voyageID As String = "")
    On Error GoTo EH
    
    If voyageID = "" Then
        voyageID = InputBox("Enter VoyageID to export (e.g., V01):", "Export Voyage Pack", "V01")
        If voyageID = "" Then Exit Sub
    End If
    
    Application.ScreenUpdating = False
    
    Dim base_path As String: base_path = ThisWorkbook.Path
    If base_path = "" Then base_path = Environ("USERPROFILE") & "\\Documents"
    
    Dim export_folder As String
    export_folder = base_path & "\\TR_DocHub_Export_" & voyageID & "_" & Format(Now, "YYYYMMDD_HHMMSS")
    
    On Error Resume Next
    MkDir export_folder
    On Error GoTo EH
    
    Dim ws_dash As Worksheet: Set ws_dash = ThisWorkbook.Sheets("D_Dashboard")
    ws_dash.ExportAsFixedFormat Type:=xlTypePDF, _
                                 Filename:=export_folder & "\\Dashboard_" & voyageID & ".pdf", _
                                 Quality:=xlQualityStandard, _
                                 IncludeDocProperties:=True, _
                                 IgnorePrintAreas:=False, _
                                 OpenAfterPublish:=False
    
    Application.ScreenUpdating = True
    MsgBox "Export Pack created:" & vbCrLf & export_folder, vbInformation
    Exit Sub
    
EH:
    Application.ScreenUpdating = True
    MsgBox "ExportVoyagePack failed: " & Err.Description, vbCritical
End Sub
'''

VBA_MOD_CONTROLTOWER = '''Attribute VB_Name = "modControlTower"
Option Explicit

'===============================================================================
' CONTROL TOWER - Single Entry Point for All Refresh Operations
'===============================================================================

Private Sub AppStateGuard_Begin()
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    Application.Calculation = xlCalculationManual
    Application.DisplayAlerts = False
End Sub

Private Sub AppStateGuard_End()
    Application.Calculation = xlCalculationAutomatic
    Application.CalculateFull
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    Application.DisplayAlerts = True
End Sub

Public Sub RefreshAll_ControlTower()
    On Error GoTo EH
    
    Dim startTime As Double: startTime = Timer
    AppStateGuard_Begin
    
    Call DG_RefreshAll
    Call TR_ApplyStatusFormatting
    Call UpdateDashboardTimestamp
    
    AppStateGuard_End
    
    MsgBox "Control Tower Refresh completed" & vbCrLf & _
           "Elapsed time: " & Format(Timer - startTime, "0.00") & " sec", vbInformation
    Exit Sub
    
EH:
    AppStateGuard_End
    MsgBox "Error occurred: " & Err.Description, vbCritical
End Sub

Private Sub UpdateDashboardTimestamp()
    Dim wsDash As Worksheet
    On Error Resume Next
    Set wsDash = ThisWorkbook.Sheets("D_Dashboard")
    If Not wsDash Is Nothing Then
        wsDash.Cells(3, 1).Value = "Last Updated: " & Format(Now, "YYYY-MM-DD HH:MM:SS")
    End If
End Sub
'''

VBA_MOD_DOCGAP = '''Attribute VB_Name = "modDocGapMacros"
Option Explicit

'=====================
' Doc Gap Tracker Macros (v3.1)
'=====================

Public Sub DG_RefreshAll()
    Application.CalculateFull
End Sub

Public Sub DG_FilterMissing()
    Dim ws As Worksheet: Set ws = ActiveSheet
    If ws.AutoFilterMode = False Then Exit Sub
    ws.Range("A2").AutoFilter Field:=4, Criteria1:="Missing"
End Sub

Public Sub DG_ClearFilters()
    Dim ws As Worksheet: Set ws = ActiveSheet
    If ws.AutoFilterMode Then
        On Error Resume Next
        ws.ShowAllData
    End If
End Sub
'''

VBA_MOD_TR_TRACKER = '''Attribute VB_Name = "TR_DocTracker_Module"
Option Explicit

'===============================================================================
' HVDC TR TRANSPORTATION - DOCUMENT TRACKER VBA MODULE
'===============================================================================

Public Sub TR_ApplyStatusFormatting()
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = ThisWorkbook.Sheets("T_Tracker")
    If ws Is Nothing Then Exit Sub
    On Error GoTo 0
    
    Application.ScreenUpdating = False
    
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
    If lastRow < 5 Then lastRow = 5
    
    Dim cell As Range
    For Each cell In ws.Range(ws.Cells(5, 8), ws.Cells(lastRow, 8))
        Select Case cell.Value
            Case "Accepted", "Submitted"
                cell.Interior.Color = RGB(144, 238, 144)
                cell.Font.Color = RGB(0, 100, 0)
            Case "In Progress"
                cell.Interior.Color = RGB(255, 255, 153)
                cell.Font.Color = RGB(128, 128, 0)
            Case "Not Started"
                cell.Interior.Color = RGB(255, 204, 204)
                cell.Font.Color = RGB(139, 0, 0)
            Case "Rejected"
                cell.Interior.Color = RGB(255, 100, 100)
                cell.Font.Color = RGB(100, 0, 0)
            Case "On Hold", "Waived"
                cell.Interior.Color = RGB(211, 211, 211)
                cell.Font.Color = RGB(128, 128, 128)
            Case Else
                cell.Interior.ColorIndex = xlNone
                cell.Font.ColorIndex = xlAutomatic
        End Select
    Next cell
    
    Application.ScreenUpdating = True
End Sub

Public Sub TR_HighlightOverdue()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("T_Tracker")
    
    Application.ScreenUpdating = False
    
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
    
    Dim r As Long
    For r = 5 To lastRow
        Dim dueDate As Variant: dueDate = ws.Cells(r, 7).Value
        Dim status As String: status = CStr(ws.Cells(r, 8).Value)
        
        If IsDate(dueDate) And status <> "Accepted" And status <> "Waived" Then
            If CDate(dueDate) < Date Then
                ws.Range(ws.Cells(r, 1), ws.Cells(r, 15)).Interior.Color = RGB(255, 200, 200)
            ElseIf CDate(dueDate) <= Date + 7 Then
                ws.Range(ws.Cells(r, 1), ws.Cells(r, 15)).Interior.Color = RGB(255, 255, 200)
            End If
        End If
    Next r
    
    Application.ScreenUpdating = True
End Sub

Public Sub EXP_ExportToPDF()
    On Error GoTo EH
    
    Dim ws As Worksheet: Set ws = ActiveSheet
    Dim filePath As String
    filePath = ThisWorkbook.Path & "\\" & ws.Name & "_" & Format(Now, "YYYYMMDD_HHMMSS") & ".pdf"
    
    ws.ExportAsFixedFormat Type:=xlTypePDF, Filename:=filePath, _
                            Quality:=xlQualityStandard, IncludeDocProperties:=True
    
    MsgBox "PDF exported: " & filePath, vbInformation
    Exit Sub
    
EH:
    MsgBox "Export failed: " & Err.Description, vbCritical
End Sub

Public Sub TR_Draft_Reminder_Emails()
    MsgBox "Email draft feature requires Outlook integration." & vbCrLf & _
           "Please use the full VBA module from modTRDocTracker.bas", vbInformation
End Sub
'''

VBA_THISWORKBOOK = '''Attribute VB_Name = "ThisWorkbook"
Option Explicit

Private Sub Workbook_Open()
    On Error Resume Next
    Application.OnKey "^+R", "RefreshAll_ControlTower"
    Application.OnKey "^+P", "EXP_ExportToPDF"
    Application.OnKey "^+E", "TR_Draft_Reminder_Emails"
End Sub

Private Sub Workbook_BeforeClose(Cancel As Boolean)
    On Error Resume Next
    Application.OnKey "^+R"
    Application.OnKey "^+P"
    Application.OnKey "^+E"
End Sub
'''

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def col_letter(n: int) -> str:
    return get_column_letter(n)

def set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[col_letter(col_idx)].width = w

def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{col_letter(start_col)}{start_row}:{col_letter(end_col)}{end_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)

def safe_create_sheet(wb: Workbook, sheet_name: str, index: int = None):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    if index is not None:
        return wb.create_sheet(sheet_name, index)
    return wb.create_sheet(sheet_name)

def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")

def get_styles() -> Dict[str, Any]:
    return {
        'font_title': Font(bold=True, size=14, color="1E3A5F"),
        'font_header': Font(bold=True, size=11, color="FFFFFF"),
        'font_bold': Font(bold=True, size=10),
        'font_normal': Font(size=10),
        'header_fill': PatternFill("solid", fgColor="4472C4"),
        'subheader_fill': PatternFill("solid", fgColor="70AD47"),
        'gray_fill': PatternFill("solid", fgColor="D9D9D9"),
        'align_center': Alignment(horizontal="center", vertical="center"),
        'align_left': Alignment(horizontal="left", vertical="center"),
        'border_thin': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

# ============================================================================
# SHEET BUILDERS
# ============================================================================

def create_lists_sheet(wb, styles: Dict) -> None:
    ws_lists = safe_create_sheet(wb, "Lists")
    
    status_list = ["Not Started", "In Progress", "Submitted", "Accepted", "Rejected", "On Hold", "Waived"]
    priority_list = ["Mandatory", "Important", "Optional"]
    due_basis_list = ["AUTO", "Doc Deadline", "Land Permit By", "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival"]
    yn_list = ["Y", "N"]
    site_list = ["Site-1", "Site-2", "Site-3", "Site-4", "Site-5", "Site-6"]
    
    ws_lists["A1"], ws_lists["B1"], ws_lists["C1"], ws_lists["D1"], ws_lists["E1"] = "Status", "Priority", "Due_Basis", "Y/N", "Sites"
    
    for i, v in enumerate(status_list, start=2): ws_lists[f"A{i}"] = v
    for i, v in enumerate(priority_list, start=2): ws_lists[f"B{i}"] = v
    for i, v in enumerate(due_basis_list, start=2): ws_lists[f"C{i}"] = v
    for i, v in enumerate(yn_list, start=2): ws_lists[f"D{i}"] = v
    for i, v in enumerate(site_list, start=2): ws_lists[f"E{i}"] = v
    
    for c in range(1, 6):
        ws_lists.cell(row=1, column=c).font = styles['font_bold']
        ws_lists.cell(row=1, column=c).fill = styles['gray_fill']
    
    set_col_widths(ws_lists, {1: 18, 2: 12, 3: 18, 4: 6, 5: 12})
    ws_lists.sheet_state = "hidden"

def create_c_config_sheet(wb, styles: Dict) -> None:
    ws_cfg = safe_create_sheet(wb, "C_Config")
    ws_cfg["A1"] = "TR Document Tracker - Config"
    ws_cfg["A1"].font = styles['font_title']
    
    cfg_rows = [
        ("DocCutoff_Days", 3, "Doc Deadline = MZP Departure - DocCutoff_Days"),
        ("LandPermitLead_Days", 7, "Land Permit By = AGI Arrival - LandPermitLead_Days (2-3 working days + weekend buffer)"),
        ("DueSoon_Threshold_Days", 7, "Dashboard/CF: upcoming items within N days"),
        ("Amber_Threshold_Days", 7, "Dashboard KPI: D-7 Amber warning threshold"),
        ("Red_Threshold_Days", 3, "Dashboard KPI: D-3 Red warning threshold"),
        ("Critical_Threshold_Days", 1, "Dashboard KPI: D-1 Critical alert threshold"),
        ("Default_Year", 2026, "Year used for sample schedule MM-DD parsing"),
        ("ExportBasePath", "", "Base path for Export Pack"),
    ]
    
    ws_cfg["A3"], ws_cfg["B3"], ws_cfg["C3"] = "Key", "Value", "Description"
    for cell in (ws_cfg["A3"], ws_cfg["B3"], ws_cfg["C3"]):
        cell.font = styles['font_bold']
        cell.fill = styles['gray_fill']
        cell.border = styles['border_thin']
    
    for i, (k, v, dsc) in enumerate(cfg_rows, start=4):
        ws_cfg[f"A{i}"], ws_cfg[f"B{i}"], ws_cfg[f"C{i}"] = k, v, dsc
        for c in range(1, 4):
            ws_cfg.cell(row=i, column=c).border = styles['border_thin']
    
    set_col_widths(ws_cfg, {1: 26, 2: 18, 3: 60})

def create_s_voyages_sheet(wb, styles: Dict) -> None:
    ws_voy = safe_create_sheet(wb, "S_Voyages")
    ws_voy["A1"] = "Voyage Schedule (Source)"
    ws_voy["A1"].font = styles['font_title']
    ws_voy["A2"] = f"Last Updated: {now_str()}"
    ws_voy["A2"].font = Font(italic=True, size=10)
    
    headers_voy = ["VoyageID", "VoyageName", "TR Units", "Site", "MZP Arrival", "Load-out", 
                   "MZP Departure", "AGI Arrival", "Doc Deadline", "Land Permit By", "Project", "Lot", "Remarks"]
    header_row = 4
    
    for col, h in enumerate(headers_voy, start=1):
        cell = ws_voy.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    sample_rows = [
        ("V01", "Voyage 1", "TR 1-2", "Site-1", dt.date(2026, 1, 26), dt.date(2026, 1, 29), dt.date(2026, 2, 1), dt.date(2026, 2, 2), None, None, "HVDC", "Lot-1", "LCT Arrival: 26 Jan | Deck Prep: 27-28 Jan (D-ring, steel set) | TR1: 29 Jan 10:00-11:00 | TR2: 30 Jan 08:00-09:00 | Pre-arrival Meeting required"),
        ("V02", "Voyage 2", "TR 3-4", "Site-2", dt.date(2026, 2, 6), dt.date(2026, 2, 7), dt.date(2026, 2, 10), dt.date(2026, 2, 11), None, None, "HVDC", "Lot-1", ""),
        ("V03", "Voyage 3", "TR 5-6", "Site-3", dt.date(2026, 2, 15), dt.date(2026, 2, 16), dt.date(2026, 2, 19), dt.date(2026, 2, 20), None, None, "HVDC", "Lot-1", ""),
        ("V04", "Voyage 4", "TR 7", "Site-4", dt.date(2026, 2, 24), dt.date(2026, 2, 25), dt.date(2026, 2, 27), dt.date(2026, 2, 28), None, None, "HVDC", "Lot-1", ""),
    ]
    
    start_row = header_row + 1
    for i, row_data in enumerate(sample_rows):
        r = start_row + i
        for c, val in enumerate(row_data, start=1):
            ws_voy.cell(row=r, column=c, value=val)
            cell = ws_voy.cell(row=r, column=c)
            cell.border = styles['border_thin']
            cell.alignment = styles['align_center'] if c <= 10 else styles['align_left']
            if c in (5, 6, 7, 8, 9, 10) and val:
                cell.number_format = "yyyy-mm-dd"
    
    for r in range(start_row, start_row + len(sample_rows)):
        ws_voy.cell(row=r, column=9, value=f"=IF(G{r}=\"\",\"\",G{r}-C_Config!$B$4)")
        ws_voy.cell(row=r, column=10, value=f"=IF(H{r}=\"\",\"\",H{r}-C_Config!$B$5)")
        ws_voy.cell(row=r, column=9).number_format = "yyyy-mm-dd"
        ws_voy.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    
    ws_voy.freeze_panes = "A5"
    set_col_widths(ws_voy, {1: 12, 2: 16, 3: 12, 4: 10, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 14, 11: 12, 12: 12, 13: 28})
    add_table(ws_voy, "tbl_Voyage", header_row, 1, start_row + len(sample_rows) - 1, len(headers_voy))
    
    dv_site = DataValidation(type="list", formula1="=Lists!$E$2:$E$7", allow_blank=True)
    ws_voy.add_data_validation(dv_site)
    dv_site.add(f"D{start_row}:D5000")

def create_m_parties_sheet(wb, styles: Dict) -> None:
    ws_party = safe_create_sheet(wb, "M_Parties")
    ws_party["A1"] = "Responsible Party Master"
    ws_party["A1"].font = styles['font_title']
    
    headers_party = ["PartyID", "PartyName", "OwnerEmail", "Contact", "ActiveFlag"]
    header_row = 3
    
    for col, h in enumerate(headers_party, start=1):
        cell = ws_party.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    parties_data = [
        ("FF", "Freight Forwarder", "", "", "Y"),
        ("CUSTBROKER", "Customs Broker", "", "", "Y"),
        ("EPC", "EPC Contractor", "", "", "Y"),
        ("TRCON", "Transport Contractor", "", "", "Y"),
        ("PORT", "Port Authority", "", "", "Y"),
        ("OFCO", "OFCO Agency", "nkk@ofco-int.com", "Nanda Kumar / Das Gopal", "Y"),
        ("MMT", "Mammoet", "Yulia.Frolova@mammoet.com", "Yulia Frolova", "Y"),
        ("SCT", "Samsung C&T", "", "", "Y"),
        ("ADNOC", "ADNOC L&S", "moda@adnoc.ae", "Mahmoud Ouda", "Y"),
        ("KFS", "Khalid Faraj Shipping", "lct.bushra@khalidfarajshipping.com", "LCT Bushra Vessel Ops", "Y"),
        ("DSV", "DSV Solutions", "jay.manaloto@dsv.com", "Jay Manaloto", "Y"),
    ]
    
    start_row = header_row + 1
    for i, (pid, name, email, contact, active) in enumerate(parties_data):
        r = start_row + i
        ws_party.cell(r, 1).value, ws_party.cell(r, 2).value = pid, name
        ws_party.cell(r, 3).value, ws_party.cell(r, 4).value, ws_party.cell(r, 5).value = email, contact, active
        for c in range(1, 6):
            cell = ws_party.cell(r, c)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            if c in (2, 3, 4):
                cell.alignment = styles['align_left']
            else:
                cell.alignment = styles['align_center']
    
    add_table(ws_party, "tbl_Party", header_row, 1, start_row + len(parties_data) - 1, len(headers_party))
    set_col_widths(ws_party, {1: 14, 2: 24, 3: 30, 4: 20, 5: 12})

def create_m_doccatalog_sheet(wb, styles: Dict) -> None:
    ws_doc = safe_create_sheet(wb, "M_DocCatalog")
    ws_doc["A1"] = "Document Catalog (Master)"
    ws_doc["A1"].font = styles['font_title']
    ws_doc["A2"] = "Edit document requirements here. Tracker refresh will follow this."
    ws_doc["A2"].font = Font(italic=True, size=10)
    
    headers_doc = ["DocCode", "DocName", "DocCategory", "DefaultResponsiblePartyID", 
                   "RequiredFlag", "EvidenceRequiredFlag", "ActiveFlag", "DocDescription"]
    doc_header_row = 4
    
    for col, h in enumerate(headers_doc, start=1):
        cell = ws_doc.cell(row=doc_header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    default_docs = [
        # === Gate & Customs ===
        ("GATEPASS", "Gate Pass Application", "Gate", "FF", "Y", "Y", "Y", "Port gate pass for MZP entry"),
        ("CUSTOMS", "Customs Declaration", "Customs", "CUSTBROKER", "Y", "Y", "Y", "Customs clearance documentation"),

        # === PTW & Loading/Offloading Documents (15 items) ===
        ("RISKASSESS", "Risk Assessment", "PTW", "EPC", "Y", "Y", "Y", "AD Port form required"),
        ("PTWCONSENT", "PTW Applicant/Receiver Consent Form", "PTW", "EPC", "Y", "Y", "Y", "Permit-to-Work applicant/receiver consent form | AD Port format"),
        ("PTWAPP", "PTW Application (Land Oversized & Heavy Load)", "PTW", "EPC", "Y", "Y", "Y", "SPMT land operations permit | 2-3 working days for approval (weekend may add extra time)"),
        ("STOWAGE", "Stowage Plan", "Transport", "TRCON", "Y", "Y", "Y", "Cargo stowage plan"),
        ("METHODSTMT", "Method Statement (incl. Weather Criteria)", "PTW", "EPC", "Y", "Y", "Y", "Method statement | AD Port form | Includes weather criteria"),
        ("COUNTDOWN", "Countdown Plan", "PTW", "EPC", "Y", "Y", "Y", "Operation countdown schedule"),
        ("UNDERTAKING", "Undertaking Letter", "PTW", "EPC", "Y", "Y", "Y", "Undertaking letter"),
        ("STABILITY", "Stability Calculation", "Transport", "MMT", "Y", "Y", "Y", "Vessel stability calculation | AGI TR.PY system"),
        ("EQUIPCERT", "3rd Party Equipment Certificates", "PTW", "MMT", "Y", "Y", "Y", "Certificate of SPMT and SPMT Operator competency certificates etc."),
        ("MWS", "Marine Warranty Survey", "Transport", "MMT", "Y", "Y", "Y", "Marine warranty survey | Required before sailing from port for HM approval"),
        ("MOORING", "Mooring Plan", "Transport", "MMT", "N", "N", "Y", "Mooring plan"),
        ("LASHING", "Lashing Plan", "Transport", "TRCON", "Y", "Y", "Y", "Cargo lashing plan"),
        ("LIFTING", "Critical Lifting Plan", "PTW", "MMT", "N", "N", "N", "Heavy lift plan | Not in current OFCO list"),
        ("INDEMNITYLIFT", "Indemnity Letter - Lifting Plan", "PTW", "EPC", "N", "N", "N", "Lifting-related indemnity letter | Not in current OFCO list"),
        ("INDEMNITY", "Indemnity Letter", "PTW", "EPC", "N", "N", "N", "General indemnity letter | Not in current OFCO list"),

        # === Transport Documents ===
        ("BL", "Bill of Lading", "Transport", "FF", "Y", "Y", "Y", "BL draft and final"),

        # === AD Maritime NOC Documents (6 items) ===
        ("NOC", "AD Maritime NOC", "Permit", "OFCO", "Y", "Y", "Y", "AD Maritime No Objection Certificate | Required prior to AGI transit"),
        ("TRADELICENSE", "Local Trading License", "NOC", "OFCO", "Y", "Y", "Y", "Local trading license / business registration"),
        ("RISKEMERG", "Detailed Risk Assessment & Emergency Response Plan", "NOC", "OFCO", "Y", "Y", "Y", "Detailed risk assessment & emergency response plan"),
        ("NOCOBJECT", "No Objection from Relevant Authorities", "NOC", "OFCO", "Y", "Y", "Y", "No-objection letter from relevant authorities"),
        ("VOYAGEPLAN", "Voyage Plan", "NOC", "OFCO", "Y", "Y", "Y", "Voyage plan (MZP → AGI)"),
        ("ROUTEMAP", "Route Map", "NOC", "OFCO", "N", "N", "Y", "Route map"),
        ("CONTRACT", "Contract Award Letter Copy", "NOC", "OFCO", "N", "N", "Y", "Copy of contract award letter"),

        # === Permit Documents ===
        ("PERMIT", "Land Permit Application", "Permit", "EPC", "Y", "Y", "Y", "Road/land permit for SPMT operations | 2-3 working days approval (weekend may add extra time)"),
    ]
    
    doc_start_row = doc_header_row + 1
    for i, (code, name, cat, party, req, evid, active, desc) in enumerate(default_docs):
        r = doc_start_row + i
        for c, val in enumerate([code, name, cat, party, req, evid, active, desc], start=1):
            cell = ws_doc.cell(r, c, value=val)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            if c in (2, 3, 8):
                cell.alignment = styles['align_left']
            else:
                cell.alignment = styles['align_center']
    
    add_table(ws_doc, "tbl_DocCatalog", doc_header_row, 1, doc_start_row + len(default_docs) - 1, len(headers_doc))
    dv_party = DataValidation(type="list", formula1="=M_Parties[PartyID]", allow_blank=False)
    dv_yn = DataValidation(type="list", formula1="=Lists!$D$2:$D$3", allow_blank=False)
    ws_doc.add_data_validation(dv_party)
    ws_doc.add_data_validation(dv_yn)
    dv_party.add(f"D{doc_start_row}:D5000")
    dv_yn.add(f"E{doc_start_row}:G5000")
    set_col_widths(ws_doc, {1: 14, 2: 28, 3: 18, 4: 24, 5: 14, 6: 18, 7: 12, 8: 40})
    ws_doc.freeze_panes = "A5"

def create_r_deadline_rules_sheet(wb, styles: Dict) -> None:
    ws_rules = safe_create_sheet(wb, "R_DeadlineRules")
    ws_rules["A1"] = "Deadline Rules (DocCode -> AnchorField + OffsetDays)"
    ws_rules["A1"].font = styles['font_title']
    ws_rules["A2"] = "Priority: Lower number = Higher priority."
    ws_rules["A2"].font = Font(italic=True, size=10)
    
    headers_rules = ["RuleID", "DocCode", "AnchorField", "OffsetDays", "CalendarType", "Priority", "ActiveFlag", "AppliesIf"]
    header_row = 4
    
    for col, h in enumerate(headers_rules, start=1):
        cell = ws_rules.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    rules_data = [
        ("R001", "GATEPASS", "Load-out", -1, "CAL", 1, "Y", ""),
        ("R002", "CUSTOMS", "Doc Deadline", -2, "WD", 1, "Y", ""),
        ("R003", "RISKASSESS", "MZP Arrival", -3, "WD", 1, "Y", ""),
        ("R004", "PTWCONSENT", "MZP Arrival", -2, "WD", 1, "Y", ""),
        ("R005", "PTWAPP", "MZP Arrival", -7, "WD", 1, "Y", "SPMT land operations permit | 2-3 working days approval + weekend buffer"),
        ("R006", "STOWAGE", "Load-out", -2, "CAL", 1, "Y", ""),
        ("R008", "METHODSTMT", "MZP Arrival", -3, "WD", 1, "Y", ""),
        ("R009", "COUNTDOWN", "MZP Arrival", -1, "CAL", 1, "Y", ""),
        ("R010", "UNDERTAKING", "MZP Arrival", -2, "CAL", 1, "Y", ""),
        ("R011", "STABILITY", "MZP Departure", -5, "WD", 1, "Y", ""),
        ("R012", "EQUIPCERT", "MZP Arrival", -5, "WD", 1, "Y", "SPMT and SPMT Operator competency certificates"),
        ("R013", "MWS", "MZP Departure", -5, "WD", 1, "Y", "Required before sailing from port for HM approval"),
        ("R014", "MOORING", "MZP Departure", -2, "CAL", 1, "Y", ""),
        ("R016", "LASHING", "Load-out", -2, "CAL", 1, "Y", ""),
        ("R007", "LIFTING", "Load-out", -3, "CAL", 1, "N", "Not in current OFCO list"),
        ("R015", "INDEMNITYLIFT", "Load-out", -2, "CAL", 1, "N", "Not in current OFCO list"),
        ("R017", "INDEMNITY", "MZP Arrival", -2, "CAL", 1, "N", "Not in current OFCO list"),
        ("R018", "BL", "MZP Departure", -3, "WD", 1, "Y", ""),
        ("R019", "NOC", "AGI Arrival", -7, "WD", 1, "Y", ""),
        ("R020", "TRADELICENSE", "AGI Arrival", -10, "WD", 1, "Y", ""),
        ("R021", "RISKEMERG", "AGI Arrival", -10, "WD", 1, "Y", ""),
        ("R022", "NOCOBJECT", "AGI Arrival", -7, "WD", 1, "Y", ""),
        ("R023", "VOYAGEPLAN", "AGI Arrival", -5, "WD", 1, "Y", ""),
        ("R024", "ROUTEMAP", "AGI Arrival", -3, "CAL", 1, "Y", ""),
        ("R025", "CONTRACT", "AGI Arrival", -5, "WD", 1, "Y", ""),
        ("R026", "PERMIT", "MZP Arrival", -7, "WD", 1, "Y", "Land Permit: 2-3 working days approval + weekend buffer"),
    ]
    
    start_row = header_row + 1
    for i, row_data in enumerate(rules_data):
        r = start_row + i
        for c, val in enumerate(row_data, start=1):
            cell = ws_rules.cell(r, c, value=val)
            cell.border = styles['border_thin']
            cell.font = styles['font_normal']
            cell.alignment = styles['align_center']
    
    add_table(ws_rules, "tbl_RuleDeadline", header_row, 1, start_row + len(rules_data) - 1, len(headers_rules))
    set_col_widths(ws_rules, {1: 10, 2: 14, 3: 18, 4: 12, 5: 14, 6: 10, 7: 12, 8: 20})
    ws_rules.freeze_panes = "A5"

def create_t_tracker_sheet(wb, styles: Dict) -> None:
    ws_tr = safe_create_sheet(wb, "T_Tracker")
    ws_tr["A1"] = "TR Document Tracker (Transaction - Main)"
    ws_tr["A1"].font = styles['font_title']
    ws_tr["A2"] = f"Last Refreshed: {now_str()}"
    ws_tr["A2"].font = Font(italic=True, size=10)
    
    tracker_headers = ["VoyageID", "DocCode", "ResponsiblePartyID", "AnchorField", "AnchorDate", 
                       "OffsetDays", "DueDate", "Status", "SubmittedDate", "AcceptedDate",
                       "EvidenceLink", "EvidenceNote", "LastUpdatedBy", "LastUpdatedAt", "RAG"]
    
    tr_header_row = 4
    for col, h in enumerate(tracker_headers, start=1):
        cell = ws_tr.cell(row=tr_header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    tr_start_row = tr_header_row + 1
    ws_tr.cell(tr_start_row, 1).value = "V01"
    ws_tr.cell(tr_start_row, 2).value = "GATEPASS"
    ws_tr.cell(tr_start_row, 3).value = "FF"
    ws_tr.cell(tr_start_row, 4).value = "Load-out"
    ws_tr.cell(tr_start_row, 5).value = dt.date(2026, 1, 29)
    ws_tr.cell(tr_start_row, 6).value = -1
    ws_tr.cell(tr_start_row, 7).value = dt.date(2026, 1, 28)
    ws_tr.cell(tr_start_row, 8).value = "Not Started"
    ws_tr.cell(tr_start_row, 15).value = f'=IF(G{tr_start_row}="","",IF(G{tr_start_row}<TODAY(),"Overdue",IF(G{tr_start_row}<=TODAY()+7,"DueSoon","OK")))'
    
    for c in range(1, 16):
        ws_tr.cell(tr_start_row, c).border = styles['border_thin']
    
    for c in (5, 7, 9, 10, 14):
        ws_tr.cell(tr_start_row, c).number_format = "yyyy-mm-dd"
    
    ws_tr.freeze_panes = "A5"
    set_col_widths(ws_tr, {1: 12, 2: 14, 3: 20, 4: 18, 5: 12, 6: 12, 7: 12, 8: 14, 9: 14, 10: 14, 11: 30, 12: 30, 13: 14, 14: 18, 15: 12})
    add_table(ws_tr, "tbl_Tracker", tr_header_row, 1, tr_start_row, len(tracker_headers))
    
    dv_status = DataValidation(type="list", formula1="=Lists!$A$2:$A$8", allow_blank=True)
    ws_tr.add_data_validation(dv_status)
    dv_status.add(f"H{tr_start_row}:H5000")
    
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    
    ws_tr.conditional_formatting.add(f"A{tr_start_row}:O5000", FormulaRule(formula=[f'=$O{tr_start_row}="Overdue"'], fill=red_fill))
    ws_tr.conditional_formatting.add(f"A{tr_start_row}:O5000", FormulaRule(formula=[f'=$O{tr_start_row}="DueSoon"'], fill=yellow_fill))
    ws_tr.conditional_formatting.add(f"A{tr_start_row}:O5000", FormulaRule(formula=[f'=$O{tr_start_row}="OK"'], fill=green_fill))

def create_d_dashboard_sheet(wb, styles: Dict) -> None:
    ws_dash = safe_create_sheet(wb, "D_Dashboard", 0)
    ws_dash["A1"] = "TR Document Preparation / Submission Dashboard"
    ws_dash["A1"].font = Font(bold=True, size=18, color="1E3A5F")
    ws_dash["A3"] = f"Last Updated: {now_str()}"
    ws_dash["A3"].font = Font(italic=True, size=10)
    
    ws_dash["A5"] = "Responsible Party:"
    ws_dash["A5"].font = styles['font_bold']
    ws_dash["B5"] = "ALL"
    
    ws_dash["A7"] = "KPI Summary"
    ws_dash["A7"].font = Font(bold=True, size=12, color="1E3A5F")
    ws_dash["A8"], ws_dash["A9"], ws_dash["A10"] = "Overdue Count", "Due in 7 days", "Due in 14 days"
    ws_dash["A11"], ws_dash["A12"], ws_dash["A13"] = "Submitted (Pending)", "Rejected Count", "Completion %"
    
    ws_dash["B8"] = '=COUNTIF(tbl_Tracker[RAG],"Overdue")'
    ws_dash["B9"] = '=COUNTIF(tbl_Tracker[RAG],"DueSoon")'
    ws_dash["B10"] = '=COUNTIFS(tbl_Tracker[DueDate],">"&TODAY()+7,tbl_Tracker[DueDate],"<="&TODAY()+14)'
    ws_dash["B11"] = '=COUNTIF(tbl_Tracker[Status],"Submitted")'
    ws_dash["B12"] = '=COUNTIF(tbl_Tracker[Status],"Rejected")'
    ws_dash["B13"] = '=IFERROR(COUNTIFS(tbl_Tracker[Status],"Accepted")/COUNTA(tbl_Tracker[VoyageID]),0)'
    ws_dash["B13"].number_format = "0%"
    
    for rr in range(8, 14):
        ws_dash[f"A{rr}"].border = styles['border_thin']
        ws_dash[f"B{rr}"].border = styles['border_thin']
        ws_dash[f"B{rr}"].font = Font(bold=True, size=11)
    
    ws_dash["A15"] = "Quick Actions"
    ws_dash["A15"].font = Font(bold=True, size=12, color="1E3A5F")
    ws_dash["A16"] = "Ctrl+Shift+R: Refresh All"
    ws_dash["A17"] = "Ctrl+Shift+P: Export PDF"
    ws_dash["A18"] = "Ctrl+Shift+E: Draft Reminder Emails"
    
    set_col_widths(ws_dash, {1: 26, 2: 18})

def create_holidays_sheet(wb, styles: Dict) -> None:
    ws_hol = safe_create_sheet(wb, "Holidays")
    ws_hol["A1"] = "Holidays Calendar (for WORKDAY.INTL)"
    ws_hol["A1"].font = styles['font_title']
    
    headers_hol = ["Date", "Holiday Name", "Type"]
    header_row = 3
    
    for col, h in enumerate(headers_hol, start=1):
        cell = ws_hol.cell(row=header_row, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['border_thin']
    
    uae_holidays = [
        (dt.date(2026, 1, 1), "New Year's Day", "UAE National"),
        (dt.date(2026, 5, 1), "Labour Day", "UAE National"),
        (dt.date(2026, 12, 2), "National Day", "UAE National"),
        (dt.date(2026, 12, 3), "National Day", "UAE National"),
    ]
    
    start_row = header_row + 1
    for i, (hol_date, name, htype) in enumerate(uae_holidays):
        r = start_row + i
        ws_hol.cell(r, 1).value = hol_date
        ws_hol.cell(r, 2).value = name
        ws_hol.cell(r, 3).value = htype
        ws_hol.cell(r, 1).number_format = "yyyy-mm-dd"
        for c in range(1, 4):
            ws_hol.cell(r, c).border = styles['border_thin']
    
    set_col_widths(ws_hol, {1: 14, 2: 30, 3: 18})
    ws_hol.freeze_panes = "A4"

def create_vba_pasteboard_sheet(wb, styles: Dict) -> None:
    ws_vba = safe_create_sheet(wb, "VBA_Pasteboard")
    ws_vba["A1"] = "VBA Installation Pasteboard"
    ws_vba["A1"].font = styles['font_title']
    ws_vba["A2"] = "=" * 60
    
    row = 4
    ws_vba.cell(row, 1).value = "[Installation Checklist]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    
    checklist_headers = ["Step", "Action", "Done", "Notes"]
    for col, h in enumerate(checklist_headers, start=1):
        ws_vba.cell(row, col, value=h).font = styles['font_bold']
        ws_vba.cell(row, col).fill = styles['gray_fill']
        ws_vba.cell(row, col).border = styles['border_thin']
    
    checklist_data = [
        (1, "Save as .xlsm (Macro-Enabled Workbook)", "", "File > Save As > .xlsm"),
        (2, "Open VBA Editor (Alt+F11)", "", ""),
        (3, "Import modOperations.bas", "", "File > Import"),
        (4, "Import modControlTower.bas", "", "File > Import"),
        (5, "Import modDocGapMacros.bas", "", "File > Import"),
        (6, "Import TR_DocTracker_Module.bas", "", "File > Import"),
        (7, "Add ThisWorkbook code", "", "Double-click ThisWorkbook"),
        (8, "Test: Run InitializeWorkbook()", "", ""),
        (9, "Test: Run GenerateTrackerRows()", "", ""),
        (10, "Test: Ctrl+Shift+R (Refresh All)", "", ""),
    ]
    
    row += 1
    for step, action, done, notes in checklist_data:
        ws_vba.cell(row, 1).value = step
        ws_vba.cell(row, 2).value = action
        ws_vba.cell(row, 3).value = done
        ws_vba.cell(row, 4).value = notes
        for c in range(1, 5):
            ws_vba.cell(row, c).border = styles['border_thin']
        row += 1
    
    row += 2
    ws_vba.cell(row, 1).value = "=" * 60
    row += 1
    ws_vba.cell(row, 1).value = "[VBA Module: modOperations]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for line in VBA_MOD_OPERATIONS.split('\n'):
        ws_vba.cell(row, 1).value = line
        row += 1
    
    row += 2
    ws_vba.cell(row, 1).value = "=" * 60
    row += 1
    ws_vba.cell(row, 1).value = "[VBA Module: modControlTower]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for line in VBA_MOD_CONTROLTOWER.split('\n'):
        ws_vba.cell(row, 1).value = line
        row += 1
    
    row += 2
    ws_vba.cell(row, 1).value = "=" * 60
    row += 1
    ws_vba.cell(row, 1).value = "[VBA Module: modDocGapMacros]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for line in VBA_MOD_DOCGAP.split('\n'):
        ws_vba.cell(row, 1).value = line
        row += 1
    
    row += 2
    ws_vba.cell(row, 1).value = "=" * 60
    row += 1
    ws_vba.cell(row, 1).value = "[VBA Module: TR_DocTracker_Module]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for line in VBA_MOD_TR_TRACKER.split('\n'):
        ws_vba.cell(row, 1).value = line
        row += 1
    
    row += 2
    ws_vba.cell(row, 1).value = "=" * 60
    row += 1
    ws_vba.cell(row, 1).value = "[ThisWorkbook Code]"
    ws_vba.cell(row, 1).font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for line in VBA_THISWORKBOOK.split('\n'):
        ws_vba.cell(row, 1).value = line
        row += 1
    
    set_col_widths(ws_vba, {1: 80, 2: 50, 3: 10, 4: 40})

def create_instructions_sheet(wb, styles: Dict) -> None:
    ws_ins = safe_create_sheet(wb, "Instructions")
    ws_ins["A1"] = "How to use TR_DocHub_AGI_2026"
    ws_ins["A1"].font = styles['font_title']
    
    instructions = [
        "",
        "=== GETTING STARTED ===",
        "1) S_Voyages: Enter voyage schedule (VoyageID, MZP Arrival, Load-out, etc.)",
        "2) M_DocCatalog: Manage document requirements (DocCode, RequiredFlag, ActiveFlag)",
        "3) R_DeadlineRules: Set DueDate rules (DocCode -> AnchorField + OffsetDays)",
        "4) D_Dashboard: Click [Generate] button -> T_Tracker rows auto-generated",
        "5) D_Dashboard: Click [Recalc] button -> DueDate auto-calculated",
        "6) T_Tracker: Update Status/SubmittedDate/AcceptedDate as documents are processed",
        "7) D_Dashboard: Review KPIs and detail list",
        "8) [Export] button: Create voyage-specific PDF/CSV export pack",
        "",
        "=== KEYBOARD SHORTCUTS ===",
        "Ctrl+Shift+R: Refresh All (RefreshAll_ControlTower)",
        "Ctrl+Shift+P: Export current sheet to PDF",
        "Ctrl+Shift+E: Draft reminder emails (Outlook)",
        "",
        "=== DueDate CALCULATION LOGIC ===",
        "- R_DeadlineRules: DocCode -> AnchorField + OffsetDays (Priority ascending)",
        "- AnchorField: MZP Arrival, Load-out, MZP Departure, AGI Arrival, Doc Deadline, Land Permit By",
        "- CalendarType: CAL (calendar days) or WD (working days via WORKDAY.INTL)",
        "- DueDate = AnchorDate + OffsetDays",
        "",
        "=== IMPORTANT NOTES ===",
        "- VoyageID + DocCode = Composite Key (no duplicates)",
        "- Run GenerateTrackerRows after adding new voyages/documents",
        "- EvidenceLink: File path or hyperlink format",
        "- Save as .xlsm before importing VBA modules",
        "- Pre-arrival meeting required before vessel arrival",
        "- Land Permit: 2-3 working days approval (weekend may add extra time)",
        "- PTW is voyage-specific (do not apply monthly)",
        "",
        "=== VBA INSTALLATION ===",
        "1) Save workbook as .xlsm (Excel Macro-Enabled Workbook)",
        "2) Open VBA Editor: Alt+F11",
        "3) Import all .bas modules from VBA_Pasteboard sheet",
        "4) Copy ThisWorkbook code to ThisWorkbook module",
        "5) Save and reopen workbook to activate shortcuts",
    ]
    
    for i, line in enumerate(instructions, start=3):
        ws_ins[f"A{i}"] = line
        ws_ins[f"A{i}"].font = styles['font_normal']
    
    ws_ins.column_dimensions["A"].width = 100

def create_log_sheet(wb, styles: Dict) -> None:
    ws_log = safe_create_sheet(wb, "LOG")
    ws_log["A1"] = "Execution Log"
    ws_log["A1"].font = styles['font_title']
    
    headers = ["Timestamp", "Action", "User", "Message"]
    for col, h in enumerate(headers, start=1):
        cell = ws_log.cell(row=3, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.border = styles['border_thin']
    
    ws_log.cell(4, 1).value = dt.datetime.now()
    ws_log.cell(4, 2).value = "CREATED"
    ws_log.cell(4, 3).value = "System"
    ws_log.cell(4, 4).value = "Workbook created by Python builder"
    
    set_col_widths(ws_log, {1: 20, 2: 16, 3: 16, 4: 60})

def create_party_contacts_sheet(wb, styles: Dict) -> None:
    ws_contacts = safe_create_sheet(wb, "Party_Contacts")
    ws_contacts["A1"] = "Party Contact List"
    ws_contacts["A1"].font = styles['font_title']
    
    headers = ["PartyName", "Email", "Contact Person", "Phone"]
    for col, h in enumerate(headers, start=1):
        cell = ws_contacts.cell(row=3, column=col, value=h)
        cell.font = styles['font_header']
        cell.fill = styles['header_fill']
        cell.border = styles['border_thin']
    
    contacts = [
        ("Freight Forwarder", "", "", ""),
        ("Customs Broker", "", "", ""),
        ("EPC Contractor", "", "", ""),
        ("Transport Contractor", "", "", ""),
        ("Port Authority", "", "", ""),
        ("OFCO Agency", "", "", ""),
        ("Mammoet", "", "", ""),
        ("Samsung C&T", "", "", ""),
    ]
    
    start_row = 4
    for i, (party, email, contact, phone) in enumerate(contacts):
        r = start_row + i
        ws_contacts.cell(r, 1).value = party
        ws_contacts.cell(r, 2).value = email
        ws_contacts.cell(r, 3).value = contact
        ws_contacts.cell(r, 4).value = phone
        for c in range(1, 5):
            ws_contacts.cell(r, c).border = styles['border_thin']
    
    add_table(ws_contacts, "tbl_Contacts", 3, 1, start_row + len(contacts) - 1, 4)
    set_col_widths(ws_contacts, {1: 24, 2: 30, 3: 24, 4: 16})

# ============================================================================
# MAIN BUILDER
# ============================================================================

def build_tr_dochub_agi_final(output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    
    styles = get_styles()
    
    create_lists_sheet(wb, styles)
    create_c_config_sheet(wb, styles)
    create_s_voyages_sheet(wb, styles)
    create_m_parties_sheet(wb, styles)
    create_m_doccatalog_sheet(wb, styles)
    create_r_deadline_rules_sheet(wb, styles)
    create_t_tracker_sheet(wb, styles)
    create_d_dashboard_sheet(wb, styles)
    create_holidays_sheet(wb, styles)
    create_party_contacts_sheet(wb, styles)
    create_log_sheet(wb, styles)
    create_vba_pasteboard_sheet(wb, styles)
    create_instructions_sheet(wb, styles)

    if output_path.exists():
        output_path.unlink()
        print(f"[INFO] Removed existing file: {output_path.name}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path

def main():
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path.cwd() / f"TR_DocHub_AGI_2026_Final_{timestamp}.xlsx"

    result = build_tr_dochub_agi_final(output_path)
    print(f"[OK] Created: {result}")
    print(f"\nNext steps:")
    print(f"   1. Open in Excel -> Save as .xlsm (Macro-Enabled Workbook)")
    print(f"   2. Alt+F11 -> Import VBA modules from VBA_Pasteboard sheet")
    print(f"   3. Run InitializeWorkbook() -> GenerateTrackerRows()")
    print(f"   4. Test Ctrl+Shift+R (Refresh All)")

if __name__ == "__main__":
    main()
