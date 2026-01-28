import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def parse_markdown_table(table_text):
    """마크다운 표를 파싱하여 리스트로 변환"""
    lines = [line.strip() for line in table_text.strip().split('\n') if line.strip()]
    if not lines:
        return []
    
    # 헤더 추출
    header_line = lines[0]
    if '|' not in header_line:
        return []
    
    headers = [cell.strip() for cell in header_line.split('|')[1:-1]]
    
    # 구분선 제거
    data_lines = [line for line in lines[1:] if not re.match(r'^\|[\s\-:]+\|$', line)]
    
    rows = []
    for line in data_lines:
        if '|' in line:
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if len(cells) == len(headers):
                rows.append(cells)
    
    return [headers] + rows

def clean_text(text):
    """마크다운 포맷팅 제거"""
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)  # **bold**
    text = re.sub(r'\*(.*?)\*', r'\1', text)  # *italic*
    text = re.sub(r'`(.*?)`', r'\1', text)  # `code`
    text = re.sub(r'#+\s*', '', text)  # headers
    text = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', text)  # links
    text = text.replace('✅', 'OK').replace('❌', 'NG').replace('⚠️', 'WARN')
    text = text.replace('🟢', 'GREEN').replace('🟡', 'YELLOW').replace('🟠', 'ORANGE').replace('🔴', 'RED')
    return text.strip()

def create_excel_from_markdown(md_file_path, excel_file_path):
    """마크다운 파일을 엑셀로 변환"""
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 섹션별로 시트 생성
    sections = {
        '1. 이고르 우려사항': extract_section_1(content),
        '2. Linkspan 체류 시간 분석': extract_section_2(content),
        '3. 조수 데이터 분석': extract_section_3(content),
        '4. Linkspan 체류 시간 시뮬레이션': extract_section_4(content),
        '5. 이고르 우려사항 대응 분석': extract_section_5(content),
        '6. 이고르 질문 답변': extract_section_6(content),
        '7. 종합 리스크 매트릭스': extract_section_7(content),
        '8. 권장 조치': extract_section_8(content),
        '9. 결론': extract_section_9(content),
    }
    
    for sheet_name, data in sections.items():
        if not data:
            continue
        
        ws = wb.create_sheet(title=sheet_name[:31])  # 엑셀 시트명 제한
        
        row = 1
        for item in data:
            if isinstance(item, dict):
                if item.get('type') == 'title':
                    cell = ws.cell(row=row, column=1, value=item['text'])
                    cell.font = title_font
                    row += 2
                elif item.get('type') == 'text':
                    ws.cell(row=row, column=1, value=item['text'])
                    row += 1
                elif item.get('type') == 'table':
                    table_data = item['data']
                    if table_data:
                        # 헤더 작성
                        for col_idx, header in enumerate(table_data[0], 1):
                            cell = ws.cell(row=row, column=col_idx, value=clean_text(header))
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = border
                        
                        row += 1
                        
                        # 데이터 작성
                        for table_row in table_data[1:]:
                            for col_idx, cell_value in enumerate(table_row, 1):
                                cell = ws.cell(row=row, column=col_idx, value=clean_text(cell_value))
                                cell.border = border
                                cell.alignment = Alignment(vertical='top', wrap_text=True)
                            row += 1
                        row += 1
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    wb.save(excel_file_path)
    try:
        print(f"✅ 엑셀 파일 생성 완료: {excel_file_path}")
    except UnicodeEncodeError:
        print(f"Excel file created: {excel_file_path}")

def extract_section_1(content):
    """섹션 1 추출"""
    match = re.search(r'## 1\. 이고르 우려사항 원문(.*?)(?=## 2\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '이고르 우려사항 원문'}]
    
    # 인용문 추출
    quote_match = re.search(r'>\s*(.+?)(?=\n\n|$)', section, re.DOTALL)
    if quote_match:
        quote = clean_text(quote_match.group(1))
        data.append({'type': 'text', 'text': quote})
    
    return data

def extract_section_2(content):
    """섹션 2 추출"""
    match = re.search(r'## 2\. Linkspan 체류 시간 분석(.*?)(?=## 3\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': 'Linkspan 체류 시간 분석'}]
    
    # 표 추출
    table_match = re.search(r'\| 단계.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table_match:
        table_data = parse_markdown_table(table_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_3(content):
    """섹션 3 추출"""
    match = re.search(r'## 3\. 조수 데이터 분석(.*?)(?=## 4\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '조수 데이터 분석'}]
    
    # 2026-01-29 표
    table1_match = re.search(r'### 3\.1.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table1_match:
        table_data = parse_markdown_table(table1_match.group(1))
        if table_data:
            data.append({'type': 'text', 'text': '2026-01-29 (Stage 3 - TR Unit 1)'})
            data.append({'type': 'table', 'data': table_data})
    
    # 2026-01-30 표
    table2_match = re.search(r'### 3\.2.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table2_match:
        table_data = parse_markdown_table(table2_match.group(1))
        if table_data:
            data.append({'type': 'text', 'text': '2026-01-30 (Stage 4 - TR Unit 2)'})
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_4(content):
    """섹션 4 추출"""
    match = re.search(r'## 4\. Linkspan 체류 시간 시뮬레이션(.*?)(?=## 5\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': 'Linkspan 체류 시간 시뮬레이션'}]
    
    # 시나리오별 표
    table_match = re.search(r'\| 시나리오.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table_match:
        table_data = parse_markdown_table(table_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    # 타임라인 데이터 추출 (텍스트로)
    timeline_match = re.search(r'#### 2026-01-29.*?\n```(.*?)```', section, re.DOTALL)
    if timeline_match:
        data.append({'type': 'text', 'text': '2026-01-29 타임라인:'})
        data.append({'type': 'text', 'text': timeline_match.group(1).strip()})
    
    timeline_match = re.search(r'#### 2026-01-30.*?\n```(.*?)```', section, re.DOTALL)
    if timeline_match:
        data.append({'type': 'text', 'text': '2026-01-30 타임라인:'})
        data.append({'type': 'text', 'text': timeline_match.group(1).strip()})
    
    return data

def extract_section_5(content):
    """섹션 5 추출"""
    match = re.search(r'## 5\. 이고르 우려사항 대응 분석(.*?)(?=## 6\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '이고르 우려사항 대응 분석'}]
    
    # 리스크 요소별 표
    table1_match = re.search(r'\| 리스크 요소.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table1_match:
        table_data = parse_markdown_table(table1_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    # Clearance 표
    table2_match = re.search(r'### 5\.2.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table2_match:
        table_data = parse_markdown_table(table2_match.group(1))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_6(content):
    """섹션 6 추출"""
    match = re.search(r'## 6\. 이고르 질문에 대한 답변(.*?)(?=## 7\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '이고르 질문에 대한 답변'}]
    
    # Pump Out vs Transfer 표
    table1_match = re.search(r'\| 항목.*?Pump Out.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table1_match:
        table_data = parse_markdown_table(table1_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    # 조수 대기 시간 최소화 방안 표
    table2_match = re.search(r'\| 방안.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table2_match:
        table_data = parse_markdown_table(table2_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_7(content):
    """섹션 7 추출"""
    match = re.search(r'## 7\. 종합 리스크 매트릭스(.*?)(?=## 8\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '종합 리스크 매트릭스'}]
    
    # 리스크 등급 표
    table1_match = re.search(r'\| 체류 시간.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table1_match:
        table_data = parse_markdown_table(table1_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    # Option별 리스크 평가 표
    table2_match = re.search(r'\| Option.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table2_match:
        table_data = parse_markdown_table(table2_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_8(content):
    """섹션 8 추출"""
    match = re.search(r'## 8\. 권장 조치(.*?)(?=## 9\.|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '권장 조치'}]
    
    # 필수 조치 표
    table_match = re.search(r'\| #.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table_match:
        table_data = parse_markdown_table(table_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

def extract_section_9(content):
    """섹션 9 추출"""
    match = re.search(r'## 9\. 결론(.*?)(?=🔧|$)', content, re.DOTALL)
    if not match:
        return []
    
    section = match.group(1)
    data = [{'type': 'title', 'text': '결론'}]
    
    # 결론 표
    table_match = re.search(r'\| 항목.*?\n\|.*?\n((?:\|.*?\n)+)', section, re.DOTALL)
    if table_match:
        table_data = parse_markdown_table(table_match.group(0))
        if table_data:
            data.append({'type': 'table', 'data': table_data})
    
    return data

if __name__ == '__main__':
    md_file = '이고르 우려사항인 Linkspan 체류 시간 관련 분석을 위해 자료를 확.md'
    excel_file = '이고르_우려사항_Linkspan_체류시간_분석.xlsx'
    
    create_excel_from_markdown(md_file, excel_file)
    try:
        print(f"\n✅ 변환 완료!")
        print(f"📄 입력: {md_file}")
        print(f"📊 출력: {excel_file}")
    except UnicodeEncodeError:
        print(f"\nConversion completed!")
        print(f"Input: {md_file}")
        print(f"Output: {excel_file}")
