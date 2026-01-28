#!/usr/bin/env python3
"""
Shared Excel helpers with cached styles and faster border/merged-cell handling.
"""

from __future__ import annotations

from typing import Dict, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

_FONT_CACHE: Dict[Tuple[str, int, bool, Optional[str]], Font] = {}
_ALIGN_CACHE: Dict[Tuple[str, str, bool], Alignment] = {}
_BORDER_CACHE: Dict[Tuple[str, Optional[str]], Border] = {}
_MERGED_CACHE: Dict[int, Dict[str, str]] = {}
_MERGED_CACHE_LEN: Dict[int, int] = {}


def get_font(
    name: str = "Calibri",
    size: int = 10,
    bold: bool = False,
    color: Optional[str] = None,
) -> Font:
    key = (name, int(size), bool(bold), color)
    if key not in _FONT_CACHE:
        _FONT_CACHE[key] = Font(name=name, size=size, bold=bold, color=color)
    return _FONT_CACHE[key]


def get_alignment(
    horizontal: str = "left",
    vertical: str = "center",
    wrap_text: bool = False,
) -> Alignment:
    key = (horizontal, vertical, bool(wrap_text))
    if key not in _ALIGN_CACHE:
        _ALIGN_CACHE[key] = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=wrap_text
        )
    return _ALIGN_CACHE[key]


def _color_key(color: Optional[object]) -> Optional[str]:
    if color is None:
        return None
    if isinstance(color, str):
        return color
    return getattr(color, "rgb", None) or getattr(color, "value", None) or str(color)


def get_border_for_side(side: Side) -> Border:
    key = (side.style, _color_key(side.color))
    if key not in _BORDER_CACHE:
        _BORDER_CACHE[key] = Border(top=side, bottom=side, left=side, right=side)
    return _BORDER_CACHE[key]


def _build_merged_map(ws) -> Dict[str, str]:
    merged_map: Dict[str, str] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[f"{get_column_letter(col)}{row}"] = top_left
    _MERGED_CACHE[id(ws)] = merged_map
    _MERGED_CACHE_LEN[id(ws)] = len(ws.merged_cells.ranges)
    return merged_map


def resolve_merged_addr(ws, addr: str) -> str:
    ws_id = id(ws)
    if _MERGED_CACHE_LEN.get(ws_id) != len(ws.merged_cells.ranges):
        merged_map = _build_merged_map(ws)
    else:
        merged_map = _MERGED_CACHE.get(ws_id) or _build_merged_map(ws)
    return merged_map.get(addr, addr)


def _set_border_side(cell, *, left=None, right=None, top=None, bottom=None) -> None:
    current = cell.border if cell.border else Border()
    cell.border = Border(
        left=left if left is not None else current.left,
        right=right if right is not None else current.right,
        top=top if top is not None else current.top,
        bottom=bottom if bottom is not None else current.bottom,
        diagonal=current.diagonal,
        diagonal_direction=current.diagonal_direction,
        outline=current.outline,
        vertical=current.vertical,
        horizontal=current.horizontal,
    )


def apply_border_outline_fast(ws, r1: int, c1: int, r2: int, c2: int, side: Side) -> None:
    for c in range(c1, c2 + 1):
        _set_border_side(ws.cell(r1, c), top=side)
        _set_border_side(ws.cell(r2, c), bottom=side)
    for r in range(r1, r2 + 1):
        _set_border_side(ws.cell(r, c1), left=side)
        _set_border_side(ws.cell(r, c2), right=side)


def apply_inner_grid_fast(ws, r_min: int, c_min: int, r_max: int, c_max: int, side: Side) -> None:
    grid = get_border_for_side(side)
    for r in range(r_min, r_max + 1):
        for c in range(c_min, c_max + 1):
            ws.cell(row=r, column=c).border = grid
