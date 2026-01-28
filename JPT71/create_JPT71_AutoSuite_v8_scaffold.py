# Py3.11.8
"""
Create JPT71 AutoSuite v8 Scaffold (XLSX)
- This creates an .xlsx (no macros). Save as .xlsm and import VBA module separately.
- Sheets:
  README, Settings, Calendar_Data(hidden), Calendar_View, LOG, LOG_SNAPSHOT, Dashboard,
  Plan (tblPlan + INPUT block Q:R), Shipping_List, Mail_Draft

Usage:
  py create_JPT71_AutoSuite_v8_scaffold.py "JPT71_AutoSuite_v8_Scaffold.xlsx"
"""

from datetime import date
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def main(out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    beige = PatternFill("solid", fgColor="FFF2CC")
    sborder = thin_border()
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # README
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "JPT71 AutoSuite v8 (Scaffold)"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "1) Save as .xlsm"
    ws["A4"] = "2) Import mod_JPT71_Auto_COMPLETE.bas"
    ws["A5"] = "3) Paste Plan sheet event code + ThisWorkbook code (see .bas comments)"
    ws["A6"] = "4) Use Plan only: tblPlan(A:O) + INPUT(Q:R)."
    ws.column_dimensions["A"].width = 120

    # Settings
    ws = wb.create_sheet("Settings", 1)
    ws["A1"] = "Settings (Sites / Ports / Holidays / Options)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A3"] = "Sites"; ws["C3"] = "Ports"; ws["E3"] = "Holidays"; ws["G3"] = "Options"
    for a in ["A3","C3","E3","G3"]:
        ws[a].fill = hdr_fill; ws[a].font = hdr_font; ws[a].alignment = center; ws[a].border = sborder
    sites = ["MW4","MW5","AGI"]
    ports = ["MW4","Al Ghallan Islands"]
    hol = [date(2025,12,25), date(2026,1,1)]
    for i,v in enumerate(sites, start=4):
        ws[f"A{i}"]=v; ws[f"A{i}"].border=sborder
    for i,v in enumerate(ports, start=4):
        ws[f"C{i}"]=v; ws[f"C{i}"].border=sborder
    for i,v in enumerate(hol, start=4):
        ws[f"E{i}"]=v; ws[f"E{i}"].number_format="yyyy-mm-dd"; ws[f"E{i}"].border=sborder

    ws["G4"]="Max events per day"; ws["H4"]=6
    ws["G5"]="Calendar month (1-12)"; ws["H5"]=12
    ws["G6"]="Calendar year"; ws["H6"]=2025
    for r in range(4,7):
        ws[f"G{r}"].border=sborder; ws[f"H{r}"].border=sborder
        ws[f"G{r}"].alignment=left; ws[f"H{r}"].alignment=center

    for col,w in {"A":18,"C":22,"E":16,"G":24,"H":18}.items():
        ws.column_dimensions[col].width=w

    # Calendar_Data (hidden)
    ws = wb.create_sheet("Calendar_Data", 2)
    ws.append(["EventDate","Trip","Type","Action","Location","StartDate","EndDate","ConflictFlag"])
    for c in range(1,9):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.sheet_state="hidden"

    # Calendar_View
    ws = wb.create_sheet("Calendar_View", 3)
    ws.sheet_view.showGridLines = False
    ws["A1"]="Calendar View (Auto)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"]="Year"; ws["B2"]="=Settings!$H$6"
    ws["C2"]="Month"; ws["D2"]="=Settings!$H$5"
    for a in ["A2","C2"]:
        ws[a].font=Font(bold=True)
    for a in ["A2","B2","C2","D2"]:
        ws[a].border=sborder
    dows=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    for i,day in enumerate(dows, start=1):
        cell=ws.cell(4,i,day)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
        ws.column_dimensions[get_column_letter(i)].width=22
    for r in range(5,11):
        ws.row_dimensions[r].height=90
        for c in range(1,8):
            cell=ws.cell(r,c,"")
            cell.border=sborder
            cell.alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)

    # LOG
    ws = wb.create_sheet("LOG", 4)
    ws.append(["Timestamp","User","Sheet","Address","OldValue","NewValue/Msg"])
    for c in range(1,7):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.column_dimensions["A"].width=20
    ws.column_dimensions["B"].width=18
    ws.column_dimensions["C"].width=18
    ws.column_dimensions["D"].width=16
    ws.column_dimensions["E"].width=24
    ws.column_dimensions["F"].width=30

    # LOG_SNAPSHOT
    ws = wb.create_sheet("LOG_SNAPSHOT", 5)
    ws.append(["Version","Timestamp","User","Range","Hash","Note"])
    for c in range(1,7):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.column_dimensions["F"].width=50

    # Dashboard
    ws = wb.create_sheet("Dashboard", 6)
    ws["A1"]="Dashboard (Auto)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A3"]="Delay Days"; ws["B3"]="=Plan!$R$4"
    ws["A4"]="Next Event Date"; ws["B4"]=""
    ws["A5"]="Conflicts"; ws["B5"]=""
    for r in range(3,6):
        ws[f"A{r}"].font=Font(bold=True); ws[f"A{r}"].border=sborder; ws[f"B{r}"].border=sborder
    ws.column_dimensions["A"].width=22
    ws.column_dimensions["B"].width=18

    # Plan (tblPlan + INPUT)
    ws = wb.create_sheet("Plan", 7)
    headers = ["Trip","Type","Material","Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload",
               "Ref_No","Item_Qty","Qty_Unit","Trailers","Trailer_Capacity","Trailer_Unit","ETA_MW4","Loading_MW4_Time"]
    ws.append(headers)
    # sample row
    ws.append(["87","Aggregate","10mm",date(2025,12,24),date(2025,12,25),"","", "HVDC-AGI-GRM-J71-087","800","Tons","10","80","Tons",date(2025,12,25),"8:00 AM"])
    for c in range(1,len(headers)+1):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
        ws.column_dimensions[get_column_letter(c)].width=16
    ws.row_dimensions[1].height=22

    tab = Table(displayName="tblPlan", ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    # INPUT block at Q:R
    ws.column_dimensions["Q"].width=34
    ws.column_dimensions["R"].width=26
    ws["Q1"]="INPUT (edit values in column R)"
    ws["Q1"].fill=hdr_fill; ws["Q1"].font=hdr_font
    ws.merge_cells("Q1:R1")

    def put(r, label, val=None, fmt=None):
        ws[f"Q{r}"]=label; ws[f"Q{r}"].border=sborder; ws[f"Q{r}"].alignment=left; ws[f"Q{r}"].font=Font(bold=True)
        ws[f"R{r}"]=val; ws[f"R{r}"].border=sborder; ws[f"R{r}"].alignment=left
        ws[f"R{r}"].fill=beige
        if fmt: ws[f"R{r}"].number_format=fmt

    put(2,"Debris-8 Actual Completion (date/time)","")
    put(3,"Debris-8 Planned MW4 Offloading Date",date(2025,12,26),"yyyy-mm-dd")
    ws["R4"]="=IF(R2=\"\",0,MAX(0,INT(R2)-R3))"
    put(4,"Delay Days (auto)", ws["R4"].value)

    put(6,"Shipment Prefix","HVDC-AGI-GRM-J71-")
    put(7,"Shipment Seq",87)
    put(8,"Document Date",date(2025,12,23),"dd mmm yyyy")
    put(9,"Vessel Name","JOPETWIL 71")
    put(10,"POL","MW4")
    put(11,"POD","Al Ghallan Islands")
    put(12,"ETD",date(2025,11,26),"dd mmm yyyy")
    put(13,"ETA",date(2025,11,27),"dd mmm yyyy")
    put(14,"Receiver Name","Mr. Khemlal")
    put(15,"Receiver Mobile","(+971 54 586 1053)")
    put(16,"Vendor","Samsung - GRM")
    put(17,"Item Name","Aggregate")
    put(18,"Material","10 mm")
    put(19,"Item Suffix","- Trailer")
    put(20,"QTY",10)
    put(21,"N. W (Ton)",71.6,"0.0")
    put(22,"G. W (Ton)",716.0,"0.00")
    put(23,"Total Style","Nos")
    put(24,"Reference","HVDC-AGI-GRM-J71-086")
    put(25,"Approx Tons",800)
    put(26,"Trailer Capacity (Tons)",80)
    put(27,"Loading Time","8:00 AM")
    put(28,"Mail Row1 Date",date(2025,12,25),"yyyy-mm-dd")
    put(29,"Mail Row1 Action","Debris Offloading")
    put(30,"Mail Row1 Remark","MW4")
    put(31,"Mail Row2 Date",date(2025,12,26),"yyyy-mm-dd")
    put(32,"Mail Row2 Action","10mm Aggregate Loading")
    put(33,"Mail Row2 Remark","MW5")
    put(34,"Berth Location","MW4")
    ws["R35"]="=R28"; put(35,"Berth From (auto)", ws["R35"].value)
    ws["R36"]="=R31"; put(36,"Berth To (auto)", ws["R36"].value)

    # Shipping_List (simple but linked)
    ws = wb.create_sheet("Shipping_List", 8)
    ws["A1"]="SHIPPING LIST FOR LIGHTNING PROJECT"; ws["A1"].font=Font(bold=True, size=16)
    ws.merge_cells("A1:L1")
    ws["A3"]="Date"; ws["B3"]="=Plan!$R$8"
    ws["A4"]="Vessel"; ws["B4"]="=Plan!$R$9"
    ws["A5"]="Shipment No"; ws["B5"]="=Plan!$R$6&Plan!$R$7"
    ws["A6"]="ETD"; ws["B6"]="=Plan!$R$12"
    ws["A7"]="ETA"; ws["B7"]="=Plan!$R$13"
    ws["A9"]="Vendor"; ws["B9"]="=Plan!$R$16"
    ws["A10"]="Item"; ws["B10"]="=Plan!$R$17&\" (\"&Plan!$R$18&\")\""
    ws["A11"]="QTY"; ws["B11"]="=Plan!$R$20"
    ws["A12"]="N.W"; ws["B12"]="=Plan!$R$21"
    ws["A13"]="G.W"; ws["B13"]="=Plan!$R$22"
    for r in range(3,14):
        ws[f"A{r}"].border=sborder; ws[f"B{r}"].border=sborder
    ws.column_dimensions["A"].width=18
    ws.column_dimensions["B"].width=48

    # Mail_Draft (cell-based linked)
    ws = wb.create_sheet("Mail_Draft", 9)
    ws["A1"]="MAIL DRAFT (copy/paste)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A3"]="Please arrange the berth"
    ws["B3"]="=TEXT(Plan!$R$35,\"dd.mmm\")&\"~\"&TEXT(Plan!$R$36,\"dd.mmm\")"
    ws["C3"]="at"; ws["D3"]="=Plan!$R$34"; ws["E3"]="per the updated schedule below:"
    ws["A5"]="Date"; ws["B5"]="Action"; ws["C5"]="Remark"
    ws.merge_cells("C5:E5")
    ws["A6"]="=TEXT(Plan!$R$28,\"dd mmm\")"; ws["B6"]="=Plan!$R$29"; ws["C6"]="=Plan!$R$30"; ws.merge_cells("C6:E6")
    ws["A7"]="=TEXT(Plan!$R$31,\"dd mmm\")"; ws["B7"]="=Plan!$R$32"; ws["C7"]="=Plan!$R$33"; ws.merge_cells("C7:E7")
    ws["A9"]="• Reference"; ws["B9"]="=Plan!$R$24"; ws.merge_cells("B9:E9")
    ws["A10"]="• Item:"; ws["B10"]="=Plan!$R$18"; ws["C10"]="(Approx."; ws["D10"]="=Plan!$R$25"; ws["E10"]="Tons)"
    ws["A11"]="• Trailers:"; ws["B11"]="Total"; ws["C11"]="=Plan!$R$20"; ws["D11"]="trailers"; ws["E11"]="= \"(\"&Plan!$R$26&\" Tons per trailer)\""
    ws["A13"]="Schedule:"; ws.merge_cells("A13:E13")
    ws["A14"]="• ETA at MW4:"; ws["B14"]="=TEXT(Plan!$R$28,\"yyyy-mm-dd\")"; ws.merge_cells("B14:E14")
    ws["A15"]="• Loading at MW4:"; ws["B15"]="=TEXT(Plan!$R$31,\"yyyy-mm-dd\")"; ws["C15"]="=Plan!$R$27"; ws.merge_cells("C15:E15")
    ws.column_dimensions["A"].width=20
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width=32

    wb.save(out_path)
    print("Saved:", out_path)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "JPT71_AutoSuite_v8_Scaffold.xlsx"
    main(out)
