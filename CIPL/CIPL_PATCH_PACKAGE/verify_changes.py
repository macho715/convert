#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify changes in generated CIPL Excel file"""

from openpyxl import load_workbook

wb = load_workbook('CIPL_TEST_NEW.xlsx')

print("=== COMMERCIAL INVOICE ===")
ws = wb['Commercial_Invoice_P1']
print("Column widths:")
for col in ['A','B','C','D','E','F','G','H','I','J']:
    w = ws.column_dimensions[col].width
    print(f"  {col}: {w}")

print("\nBorder check (A1):")
border = ws['A1'].border
if border.left.style:
    print(f"  Left border style: {border.left.style}")
    if border.left.color and hasattr(border.left.color, 'rgb'):
        print(f"  Left border color: {border.left.color.rgb}")
    else:
        print(f"  Left border color: {border.left.color}")

print("\nA5-B5 border check:")
a5_border = ws['A5'].border
b5_border = ws['B5'].border
print(f"  A5 right border: {a5_border.right.style if a5_border.right.style else 'None'}")
print(f"  B5 left border: {b5_border.left.style if b5_border.left.style else 'None'}")

print("\n=== PACKING LIST ===")
ws2 = wb['Packing_List_P1']
print("Column widths:")
for col in ['A','B','C','D','E','F','G','H','I','J','K']:
    w = ws2.column_dimensions[col].width
    print(f"  {col}: {w}")

print("\nBorder check (A1):")
border2 = ws2['A1'].border
if border2.left.style:
    print(f"  Left border style: {border2.left.style}")
    if border2.left.color and hasattr(border2.left.color, 'rgb'):
        print(f"  Left border color: {border2.left.color.rgb}")
    else:
        print(f"  Left border color: {border2.left.color}")

print("\nA5-B5 border check:")
a5_border2 = ws2['A5'].border
b5_border2 = ws2['B5'].border
print(f"  A5 right border: {a5_border2.right.style if a5_border2.right.style else 'None'}")
print(f"  B5 left border: {b5_border2.left.style if b5_border2.left.style else 'None'}")

