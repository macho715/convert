#!/usr/bin/env python3
"""
AGI TR 7-Voyage Master Gantt with VBA
기존 AGI_TR_7Voyage_Master_Gantt.xlsx와 동일한 레이아웃 + VBA 기능
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# === COLORS (기존과 동일) ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def create_gantt_with_vba():
    wb = Workbook()
    
    # === CONTROL PANEL (새로 추가) ===
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"
    
    # Title
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "🎛️ AGI TR Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30
    
    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "📌 시작일(B4)을 변경하면 모든 일정이 자동 업데이트됩니다. VBA 매크로 활성화 필요."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    
    # Input Section
    ws_ctrl["A4"] = "📅 프로젝트 시작일:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = dt.date(2026, 1, 18)
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)
    
    ws_ctrl["A5"] = "🎯 목표 완료일:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = dt.date(2026, 2, 28)
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()
    
    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")
    
    # Duration Parameters
    ws_ctrl["D4"] = "⏱️ 작업 소요시간 (일)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)
    
    durations = [
        ("D5", "Mobilization:", "E5", 1.0, "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", 3.0, "DUR_DECK"),
        ("D7", "Load-out:", "E7", 1.0, "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", 0.5, "DUR_SF"),
        ("D9", "MWS Approval:", "E9", 0.5, "DUR_MWS"),
        ("D10", "Sailing:", "E10", 1.0, "DUR_SAIL"),
        ("D11", "AGI Unload:", "E11", 1.0, "DUR_UL"),
        ("D12", "Turning:", "E12", 3.0, "DUR_TURN"),
        ("D13", "Jack-down:", "E13", 1.0, "DUR_JD"),
        ("D14", "Return:", "E14", 1.0, "DUR_RET"),
        ("D15", "Buffer:", "E15", 0.5, "DUR_BUF"),
    ]
    
    for lc, lt, vc, v, name in durations:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = v
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")
    
    # Weather Settings
    ws_ctrl["G4"] = "🌊 기상 설정"
    ws_ctrl["G4"].font = Font(bold=True, size=12)
    ws_ctrl["G5"] = "Shamal 시작:"
    ws_ctrl["H5"] = dt.date(2026, 2, 5)
    ws_ctrl["H5"].number_format = "YYYY-MM-DD"
    ws_ctrl["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G6"] = "Shamal 종료:"
    ws_ctrl["H6"] = dt.date(2026, 2, 14)
    ws_ctrl["H6"].number_format = "YYYY-MM-DD"
    ws_ctrl["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    wb.defined_names["SHAMAL_START"] = DefinedName("SHAMAL_START", attr_text="Control_Panel!$H$5")
    wb.defined_names["SHAMAL_END"] = DefinedName("SHAMAL_END", attr_text="Control_Panel!$H$6")
    
    # Summary Section
    ws_ctrl["A8"] = "📊 자동 계산 요약"
    ws_ctrl["A8"].font = Font(bold=True, size=12)
    
    summary_items = [
        ("A9", "예상 완료일:", "B9", "=MAX(Schedule_Data!G:G)"),
        ("A10", "총 기간 (일):", "B10", "=B9-B4+1"),
        ("A11", "목표 대비:", "B11", '=IF(B9<=B5,"✅ 목표 달성","❌ 지연")'),
        ("A12", "잔여 일수:", "B12", "=B5-B9"),
    ]
    
    for lc, lt, vc, formula in summary_items:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if "MAX" in formula:
            ws_ctrl[vc].number_format = "YYYY-MM-DD"
    
    # VBA Button Info
    ws_ctrl["A15"] = "🔘 VBA 매크로 (Alt+F8)"
    ws_ctrl["A15"].font = Font(bold=True, size=12)
    
    buttons = [
        "▶ UpdateAllSchedules - 전체 일정 재계산",
        "▶ RefreshGanttChart - Gantt 색상 갱신",
        "▶ GenerateReport - 현황 리포트",
        "▶ ExportToPDF - PDF 내보내기",
        "▶ SimulateDelay - 지연 시뮬레이션",
        "▶ HighlightCritical - Critical Path 강조",
        "▶ HighlightToday - 오늘 날짜 표시",
    ]
    for i, btn in enumerate(buttons, 16):
        ws_ctrl[f"A{i}"] = btn
        ws_ctrl[f"A{i}"].font = Font(size=10)
    
    # Column widths
    ws_ctrl.column_dimensions["A"].width = 20
    ws_ctrl.column_dimensions["B"].width = 15
    ws_ctrl.column_dimensions["D"].width = 16
    ws_ctrl.column_dimensions["E"].width = 10
    ws_ctrl.column_dimensions["G"].width = 14
    ws_ctrl.column_dimensions["H"].width = 12
    
    # === SCHEDULE DATA (기존과 동일한 구조) ===
    ws_sched = wb.create_sheet("Schedule_Data")
    
    # Title rows (기존과 동일)
    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = "AGI HVDC TR 1-7 Transportation Master Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = "7 Voyages | 4 Jack-down Events | LCT BUSHRA | Mina Zayed ↔ AGI Site"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    
    ws_sched.merge_cells("A3:I3")
    ws_sched["A3"] = "⚠️ Winter Shamal Risk Period: Feb 5-14, 2026 | Tide ≥1.90m + Weather Gate | Schedule Auto-Updates from Control_Panel"
    ws_sched["A3"].font = Font(size=10, italic=True)
    ws_sched["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    # Headers (Row 5, 기존과 동일)
    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Task definitions (최종 실행 일정 반영)
    # (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    # 프로젝트 시작: 2026-01-18 (Day 0)
    tasks = [
        # Mobilization
        ("MOB-001", "1.0", "MOBILIZATION", "MOBILIZATION", "Mammoet", 0, "DUR_MOB", "SPMT Assembly + Marine Equipment Mobilization"),
        ("PREP-001", "1.1", "Deck Preparations", "DECK_PREP", "Mammoet", 1, "DUR_DECK", "One-time setup for all voyages"),
        
        # Voyage 1: LO 01-18, SAIL 01-20, ARR 01-22
        ("V1", "2.0", "VOYAGE 1: TR1 Transport", "MILESTONE", "All", 0, 0, "✅ Tide ≥1.90m (2.05m) | Good Weather Window"),
        ("LO-101", "2.1", "TR1 Load-out on LCT", "LOADOUT", "Mammoet", 0, "DUR_LO", "Tide ≥1.90m (2.05m) required"),
        ("SF-102", "2.2", "TR1 Sea Fastening", "SEAFAST", "Mammoet", 0, "DUR_SF", "12-point lashing"),
        ("MWS-103", "2.3", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 0, "DUR_MWS", "Marine Warranty Surveyor"),
        ("SAIL-104", "2.4", "V1 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 2, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-105", "2.5", "AGI Arrival + TR1 RORO Unload", "AGI_UNLOAD", "Mammoet", 4, "DUR_UL", "Tide ≥1.90m (1.91m) | AGI FWD Draft ≤ 2.70m"),
        ("STORE-106", "2.6", "TR1 Stored on AGI Laydown", "BUFFER", "Mammoet", 4, "DUR_BUF", "Awaiting pair TR2"),
        ("RET-107", "2.7", "V1 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 4, "DUR_RET", "Quick turnaround"),
        ("BUF-108", "2.99", "V1 Buffer / Equipment Reset", "BUFFER", "All", 5, "DUR_BUF", "Weather contingency"),
        
        # Voyage 2: LO 01-26, SAIL 01-27, ARR 01-29
        ("V2", "3.0", "VOYAGE 2: TR2 Transport + JD-1", "MILESTONE", "All", 8, 0, "✅ Tide ≥1.90m (1.91m) | Good Weather Window (before Shamal)"),
        ("LO-109", "3.1", "TR2 Load-out on LCT", "LOADOUT", "Mammoet", 8, "DUR_LO", "Tide ≥1.90m (1.91m) required"),
        ("SF-110", "3.2", "TR2 Sea Fastening", "SEAFAST", "Mammoet", 8, "DUR_SF", "12-point lashing"),
        ("MWS-110A", "3.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 8, "DUR_MWS", "Pre-sail verification"),
        ("SAIL-111", "3.3", "V2 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 9, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-112", "3.4", "AGI Arrival + TR2 RORO Unload", "AGI_UNLOAD", "Mammoet", 11, "DUR_UL", "Tide ≥1.90m (2.03m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-113", "3.5", "TR1 Transport to Bay-1", "TURNING", "Mammoet", 12, 1, "Steel bridge install"),
        ("TURN-114", "3.6", "TR1 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", "10t Forklift required"),
        ("TRN-116", "3.8", "TR2 Transport to Bay-2", "TURNING", "Mammoet", 12, 1, ""),
        ("TURN-117", "3.9", "TR2 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", ""),
        ("JD-120", "3.95", "★ JD-1 Jack-Down Batch (TR1+TR2)", "JACKDOWN", "Mammoet", 14, "DUR_JD", "MILESTONE: TR1+TR2 Complete (Batch) | 02-01 ~ 02-02"),
        ("RET-119", "3.11", "V2 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 15, "DUR_RET", "Return after JD-1"),
        ("BUF-120", "3.99", "V2 Buffer / Shamal Recovery", "BUFFER", "All", 15, "DUR_BUF", "Post-Shamal weather check"),
        
        # Voyage 3: LO 01-31, SAIL 02-02, ARR 02-03
        ("V3", "4.0", "VOYAGE 3: TR3 Transport", "MILESTONE", "All", 13, 0, "✅ Tide ≥1.90m (2.07m) | Post-Shamal Window"),
        ("LO-121", "4.1", "TR3 Load-out on LCT", "LOADOUT", "Mammoet", 13, "DUR_LO", "Tide ≥1.90m (2.07m)"),
        ("SF-122", "4.2", "TR3 Sea Fastening", "SEAFAST", "Mammoet", 13, "DUR_SF", ""),
        ("MWS-122A", "4.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 13, "DUR_MWS", ""),
        ("SAIL-123", "4.3", "V3 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 15, "DUR_SAIL", "Good weather"),
        ("ARR-124", "4.4", "AGI Arrival + TR3 RORO Unload", "AGI_UNLOAD", "Mammoet", 16, "DUR_UL", "Tide ≥1.90m (2.04m)"),
        ("STORE-125", "4.5", "TR3 Stored on AGI Laydown", "BUFFER", "Mammoet", 16, "DUR_BUF", "Awaiting pair TR4"),
        ("RET-126", "4.6", "V3 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 17, "DUR_RET", ""),
        ("BUF-127", "4.99", "V3 Buffer", "BUFFER", "All", 17, "DUR_BUF", ""),
        
        # Voyage 4: LO 02-15, SAIL 02-16, ARR 02-18
        ("V4", "5.0", "VOYAGE 4: TR4 Transport + JD-2", "MILESTONE", "All", 28, 0, "✅ Tide ≥1.90m (1.90m) | Shamal 종료 직후"),
        ("LO-128", "5.1", "TR4 Load-out on LCT", "LOADOUT", "Mammoet", 28, "DUR_LO", "Tide ≥1.90m (1.90m)"),
        ("SF-129", "5.2", "TR4 Sea Fastening", "SEAFAST", "Mammoet", 28, "DUR_SF", ""),
        ("MWS-129A", "5.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 28, "DUR_MWS", ""),
        ("SAIL-130", "5.3", "V4 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 29, "DUR_SAIL", ""),
        ("ARR-131", "5.4", "AGI Arrival + TR4 RORO Unload", "AGI_UNLOAD", "Mammoet", 31, "DUR_UL", "Tide ≥1.90m (1.96m)"),
        ("TRN-132", "5.5", "TR3 Transport to Bay-3", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-133", "5.6", "TR3 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("TRN-135", "5.8", "TR4 Transport to Bay-4", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-136", "5.9", "TR4 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("JD-139", "5.95", "★ JD-2 Jack-Down Batch (TR3+TR4)", "JACKDOWN", "Mammoet", 33, "DUR_JD", "MILESTONE: TR3+TR4 Complete (Batch) | 02-20"),
        ("RET-138", "5.11", "V4 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 33, "DUR_RET", "Return after JD-2"),
        ("BUF-140", "5.99", "V4 Buffer", "BUFFER", "All", 33, "DUR_BUF", ""),
        
        # Voyage 5: LO 02-23, SAIL 02-23, ARR 02-24 (Fast-turn)
        ("V5", "6.0", "VOYAGE 5: TR5 Transport", "MILESTONE", "All", 36, 0, "✅ Tide ≥1.90m (1.99m) | Fast-turn"),
        ("LO-140", "6.1", "TR5 Load-out on LCT", "LOADOUT", "Mammoet", 36, "DUR_LO", "Tide ≥1.90m (1.99m)"),
        ("SF-141", "6.2", "TR5 Sea Fastening", "SEAFAST", "Mammoet", 36, "DUR_SF", ""),
        ("MWS-141A", "6.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 36, "DUR_MWS", ""),
        ("SAIL-142", "6.3", "V5 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 36, "DUR_SAIL", "Fast-turn"),
        ("ARR-143", "6.4", "AGI Arrival + TR5 RORO Unload", "AGI_UNLOAD", "Mammoet", 37, "DUR_UL", "Tide ≥1.90m (2.01m)"),
        ("STORE-144", "6.5", "TR5 Stored on AGI Laydown", "BUFFER", "Mammoet", 37, "DUR_BUF", "Awaiting pair TR6"),
        ("RET-145", "6.6", "V5 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 37, "DUR_RET", ""),
        ("BUF-146", "6.99", "V5 Buffer", "BUFFER", "All", 37, "DUR_BUF", ""),
        
        # Voyage 6: LO 02-25, SAIL 02-25, ARR 02-26 (Fast-turn)
        ("V6", "7.0", "VOYAGE 6: TR6 Transport + JD-3", "MILESTONE", "All", 38, 0, "✅ Tide ≥1.90m (2.01m) | Fast-turn"),
        ("LO-147", "7.1", "TR6 Load-out on LCT", "LOADOUT", "Mammoet", 38, "DUR_LO", "Tide ≥1.90m (2.01m)"),
        ("SF-148", "7.2", "TR6 Sea Fastening", "SEAFAST", "Mammoet", 38, "DUR_SF", ""),
        ("MWS-148A", "7.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 38, "DUR_MWS", ""),
        ("SAIL-149", "7.3", "V6 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 38, "DUR_SAIL", "Fast-turn"),
        ("ARR-150", "7.4", "AGI Arrival + TR6 RORO Unload", "AGI_UNLOAD", "Mammoet", 39, "DUR_UL", "Tide ≥1.90m (1.98m)"),
        ("TRN-151", "7.5", "TR5 Transport to Bay-5", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-152", "7.6", "TR5 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("TRN-154", "7.8", "TR6 Transport to Bay-6", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-155", "7.9", "TR6 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("JD-157", "7.95", "★ JD-3 Jack-Down Batch (TR5+TR6)", "JACKDOWN", "Mammoet", 40, "DUR_JD", "MILESTONE: TR5+TR6 Complete (Batch) | 02-27"),
        ("RET-158", "7.11", "V6 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 40, "DUR_RET", "Return after JD-3"),
        ("BUF-159", "7.99", "V6 Buffer / Reset for V7", "BUFFER", "All", 40, "DUR_BUF", ""),
        
        # Voyage 7: LO 02-27, SAIL 02-27, ARR 02-28 (Final)
        ("V7", "8.0", "VOYAGE 7: TR7 Transport + JD-4", "MILESTONE", "All", 40, 0, "✅ Tide ≥1.90m (1.92m) | Final unit"),
        ("LO-201", "8.1", "TR7 Load-out on LCT", "LOADOUT", "Mammoet", 40, "DUR_LO", "Tide ≥1.90m (1.92m) required"),
        ("SF-202", "8.2", "TR7 Sea Fastening", "SEAFAST", "Mammoet", 40, "DUR_SF", "12-point lashing"),
        ("MWS-202A", "8.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 40, "DUR_MWS", ""),
        ("SAIL-203", "8.3", "V7 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 40, "DUR_SAIL", "Weather window required"),
        ("ARR-204", "8.4", "AGI Arrival + TR7 RORO Unload", "AGI_UNLOAD", "Mammoet", 41, "DUR_UL", "Tide ≥1.90m (1.93m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-205", "8.5", "TR7 Transport to Bay-7", "TURNING", "Mammoet", 41, 1, "Steel bridge install"),
        ("TURN-206", "8.6", "TR7 Turning (90° rotation)", "TURNING", "Mammoet", 41, "DUR_TURN", "10t Forklift required"),
        ("JD-207", "8.7", "★ JD-4 Jack-Down (TR7)", "JACKDOWN", "Mammoet", 41, "DUR_JD", "MILESTONE: TR7 Complete | 02-28"),
        ("RET-208", "8.8", "V7 LCT Final Return: AGI→MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Final return"),
        
        # Demobilization
        ("DEMOB", "9.0", "DEMOBILIZATION", "MOBILIZATION", "Mammoet", 42, "DUR_MOB", "Equipment return"),
        ("END", "99.0", "★★★ PROJECT COMPLETE ★★★", "MILESTONE", "All", 42, 0, "All 7 TRs Installed | Jan-Feb 2026 Complete"),
    ]
    
    for r, t in enumerate(tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_sched.cell(r, 1, value=tid)
        ws_sched.cell(r, 2, value=wbs)
        ws_sched.cell(r, 3, value=task)
        ws_sched.cell(r, 4, value=phase)
        ws_sched.cell(r, 5, value=owner)
        
        # Start = PROJECT_START + Offset
        ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"
        
        # Duration
        if isinstance(dur_ref, str):
            ws_sched.cell(r, 8, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 8, value=dur_ref)
        
        # End = Start + Duration
        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"
        
        ws_sched.cell(r, 9, value=notes)
        
        # Styling
        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 10):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        if phase == "MILESTONE":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")
    
    # Column widths
    col_widths = {"A":10, "B":6, "C":38, "D":14, "E":14, "F":12, "G":12, "H":10, "I":40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    
    # === GANTT CHART (기존과 동일한 구조) ===
    ws_gantt = wb.create_sheet("Gantt_Chart")
    
    # Title rows
    ws_gantt.merge_cells("A1:CA1")
    ws_gantt["A1"] = "AGI HVDC TR 1-7 Master Gantt Chart (Auto-Updated)"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    ws_gantt.merge_cells("A2:CA2")
    ws_gantt["A2"] = "Orange Zone = Winter Shamal Risk Period (Feb 5-14) | Tide ≥1.90m + Weather Gate | VBA: RefreshGanttChart로 색상 갱신"
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    # Month header
    ws_gantt["A3"] = "Jan 2026"
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:G3")
    
    # Column headers (Row 4)
    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Date columns (42 days from project start to cover all tasks)
    date_col = 8
    for i in range(42):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5
    
    # Task rows
    for r, t in enumerate(tasks, 5):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_gantt.cell(r, 1, value=f"=Schedule_Data!A{r+1}")
        ws_gantt.cell(r, 2, value=f"=Schedule_Data!B{r+1}")
        ws_gantt.cell(r, 3, value=f"=Schedule_Data!C{r+1}")
        ws_gantt.cell(r, 4, value=f"=Schedule_Data!D{r+1}")
        
        start_cell = ws_gantt.cell(r, 5, value=f"=Schedule_Data!F{r+1}")
        start_cell.number_format = "MM/DD"
        
        end_cell = ws_gantt.cell(r, 6, value=f"=Schedule_Data!G{r+1}")
        end_cell.number_format = "MM/DD"
        
        ws_gantt.cell(r, 7, value=f"=Schedule_Data!H{r+1}")
        
        # Meta columns borders
        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()
        
        # Phase color in column D
        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        # Date cells borders
        for i in range(42):
            ws_gantt.cell(r, date_col + i).border = tb()
    
    # Column widths for meta columns
    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 28
    ws_gantt.column_dimensions["D"].width = 12
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4
    
    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)
    
    # === WEATHER ANALYSIS (기존과 동일) ===
    ws_weather = wb.create_sheet("Weather_Analysis")
    
    ws_weather["A1"] = "UAE Winter Weather Analysis - Jan/Feb 2026"
    ws_weather["A1"].font = Font(bold=True, size=14)
    ws_weather.merge_cells("A1:F1")
    
    weather_headers = ["Parameter", "Jan 1-10", "Jan 11-20", "Jan 21-31", "Feb 1-15", "Notes"]
    for c, h in enumerate(weather_headers, 1):
        cell = ws_weather.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = tb()
    
    weather_data = [
        ("Avg Wind (kt)", "11-13", "16-21", "13-15", "12-14", "Peak Shamal: Feb 5-14"),
        ("Max Gust (kt)", "18-20", "25-30", "20-22", "18-20", "NO-GO if >22kt gust"),
        ("Wave Height (m)", "0.4-0.6", "0.8-1.2", "0.5-0.7", "0.4-0.6", "HOLD if >0.8m"),
        ("Visibility (km)", "8-10", "2-5", "6-8", "8-10", "Reduced during Shamal"),
        ("Risk Level", "LOW", "HIGH", "MEDIUM", "LOW", ""),
        ("Recommendation", "GO", "NO-GO", "CAUTION", "GO", ""),
    ]
    
    for r, row in enumerate(weather_data, 4):
        for c, val in enumerate(row, 1):
            cell = ws_weather.cell(r, c, value=val)
            cell.border = tb()
            if val == "HIGH" or val == "NO-GO":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif val == "MEDIUM" or val == "CAUTION":
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
            elif val == "LOW" or val == "GO":
                cell.fill = PatternFill("solid", fgColor="C8E6C9")
    
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_weather.column_dimensions[col].width = 15
    
    # === SUMMARY (기존과 동일) ===
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "AGI HVDC Transformer Transportation - Project Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")
    
    summary_data = [
        ("Key Parameters", ""),
        ("Total Transformers", "7 units (TR1-TR7)"),
        ("Weight per TR", "217-271 tons"),
        ("Total Voyages", "7 (1 TR per voyage)"),
        ("Jack-down Events", "4 (after V2, V4, V6, V7)"),
        ("Vessel", "LCT BUSHRA"),
        ("Route", "Mina Zayed Port ↔ AGI Site"),
        ("", ""),
        ("Schedule Summary", ""),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
        ("Mobilization", "Day 0-1"),
        ("Voyage 1 (TR1)", "Day 0-5"),
        ("Voyage 2 (TR2+JD)", "Day 8-15"),
        ("Voyage 3 (TR3)", "Day 13-17"),
        ("Voyage 4 (TR4+JD)", "Day 28-33"),
        ("Voyage 5 (TR5)", "Day 36-37"),
        ("Voyage 6 (TR6+JD)", "Day 38-40"),
        ("Voyage 7 (TR7+JD)", "Day 40-41"),
        ("Project Complete", "=MAX(Schedule_Data!G:G)"),
        ("", ""),
        ("Weather Constraints", ""),
        ("Shamal Period", "Feb 5-14, 2026"),
        ("Tide Requirement", "≥1.90m for LO/ARR"),
        ("Wind Limit", "≤18kt sustained, ≤22kt gust"),
        ("Wave Limit", "≤0.8m (HOLD), ≤1.0m (NO-GO)"),
        ("AGI Draft Limit", "≤2.70m forward draft"),
    ]
    
    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30
    
    # === VBA CODE SHEET ===
    ws_vba = wb.create_sheet("VBA_Code")
    
    ws_vba["A1"] = "📋 VBA 코드 - Alt+F11 → Module에 붙여넣기 → .xlsm으로 저장"
    ws_vba["A1"].font = Font(bold=True, size=14)
    
    vba_code = '''
Option Explicit

' ============================================
' AGI TR 7-Voyage Master Gantt - VBA Macros
' ============================================
' 사용법: Alt+F11 → Module 삽입 → 코드 붙여넣기
' ============================================

' === 1. 전체 일정 업데이트 ===
Sub UpdateAllSchedules()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    Sheets("Schedule_Data").Calculate
    Sheets("Gantt_Chart").Calculate
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    
    Call RefreshGanttChart
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "✅ 일정 업데이트 완료!" & vbCrLf & vbCrLf & _
           "프로젝트 시작: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
           "예상 완료: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD"), _
           vbInformation, "Schedule Updated"
End Sub

' === 2. Gantt Chart 색상 갱신 ===
Sub RefreshGanttChart()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long
    Dim shamalStart As Date, shamalEnd As Date
    
    Set ws = Sheets("Gantt_Chart")
    Set wsd = Sheets("Schedule_Data")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    dc = 8 ' Date columns start at H
    
    Application.ScreenUpdating = False
    
    ' Clear existing colors in date columns
    ws.Range(ws.Cells(5, dc), ws.Cells(85, dc + 41)).Interior.ColorIndex = xlNone
    
    ' Reset header colors
    For j = 0 To 41
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121) ' HEADER color
    Next j
    
    ' Highlight Shamal period in header
    For j = 0 To 41
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0) ' Orange
        End If
    Next j
    
    ' Apply Gantt bars
    For i = 5 To 85 ' Task rows
        If IsDate(wsd.Cells(i + 1, 6).Value) Then
            startD = wsd.Cells(i + 1, 6).Value
            endD = wsd.Cells(i + 1, 7).Value
            phase = wsd.Cells(i + 1, 4).Value
            
            For j = 0 To 41
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(i, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(i, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(i, dc + j).Value = Chr(9733) ' Star
                    ws.Cells(i, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(i, dc + j).Font.Size = 8
                End If
            Next j
        End If
    Next i
    
    ' Highlight today
    For j = 0 To 41
        cellDate = projStart + j
        If cellDate = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(85, dc + j)).Borders(xlEdgeLeft).Color = RGB(255, 0, 0)
            ws.Range(ws.Cells(4, dc + j), ws.Cells(85, dc + j)).Borders(xlEdgeLeft).Weight = xlThick
            Exit For
        End If
    Next j
    
    Application.ScreenUpdating = True
End Sub

' === Phase Color Helper ===
Function GetPhaseColor(phase As String) As Long
    Select Case phase
        Case "MOBILIZATION": GetPhaseColor = RGB(142, 124, 195)
        Case "DECK_PREP": GetPhaseColor = RGB(111, 168, 220)
        Case "LOADOUT": GetPhaseColor = RGB(147, 196, 125)
        Case "SEAFAST": GetPhaseColor = RGB(118, 165, 175)
        Case "SAIL": GetPhaseColor = RGB(164, 194, 244)
        Case "AGI_UNLOAD": GetPhaseColor = RGB(246, 178, 107)
        Case "TURNING": GetPhaseColor = RGB(255, 217, 102)
        Case "JACKDOWN": GetPhaseColor = RGB(224, 102, 102)
        Case "RETURN": GetPhaseColor = RGB(153, 153, 153)
        Case "BUFFER": GetPhaseColor = RGB(217, 217, 217)
        Case "MILESTONE": GetPhaseColor = RGB(255, 0, 0)
        Case Else: GetPhaseColor = RGB(255, 255, 255)
    End Select
End Function

' === 3. 프로젝트 리포트 생성 ===
Sub GenerateReport()
    Dim wsd As Worksheet
    Dim i As Long, total As Long, jdCount As Long
    Dim voyages As Long, milestones As Long
    
    Set wsd = Sheets("Schedule_Data")
    
    For i = 6 To 87
        If wsd.Cells(i, 1).Value <> "" Then
            total = total + 1
            If wsd.Cells(i, 4).Value = "JACKDOWN" Then jdCount = jdCount + 1
            If wsd.Cells(i, 4).Value = "MILESTONE" Then milestones = milestones + 1
            If Left(wsd.Cells(i, 1).Value, 1) = "V" And Len(wsd.Cells(i, 1).Value) = 2 Then voyages = voyages + 1
        End If
    Next i
    
    Dim rpt As String
    rpt = "╔══════════════════════════════════════╗" & vbCrLf & _
          "║   AGI HVDC TR Transportation Report  ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ Report Date: " & Format(Now, "YYYY-MM-DD HH:MM") & "      ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ PROJECT STATUS                       ║" & vbCrLf & _
          "║  Total Tasks: " & total & "                      ║" & vbCrLf & _
          "║  Voyages: " & voyages & "                          ║" & vbCrLf & _
          "║  Jack-down Events: " & jdCount & "                 ║" & vbCrLf & _
          "║  Milestones: " & milestones & "                       ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ KEY DATES                            ║" & vbCrLf & _
          "║  Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & "              ║" & vbCrLf & _
          "║  Target: " & Format(Sheets("Control_Panel").Range("B5").Value, "YYYY-MM-DD") & "             ║" & vbCrLf & _
          "║  Est.End: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD") & "            ║" & vbCrLf & _
          "║  Status: " & Sheets("Control_Panel").Range("B11").Value & "               ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ WEATHER RISK                         ║" & vbCrLf & _
          "║  Shamal: " & Format(Sheets("Control_Panel").Range("H5").Value, "MM/DD") & " - " & Format(Sheets("Control_Panel").Range("H6").Value, "MM/DD") & "           ║" & vbCrLf & _
          "╚══════════════════════════════════════╝"
    
    MsgBox rpt, vbInformation, "Project Report"
End Sub

' === 4. PDF 내보내기 ===
Sub ExportToPDF()
    Dim fp As String
    fp = ThisWorkbook.Path & "\\AGI_TR_Gantt_" & Format(Date, "YYYYMMDD") & ".pdf"
    
    Sheets(Array("Schedule_Data", "Gantt_Chart", "Summary")).Select
    ActiveSheet.ExportAsFixedFormat xlTypePDF, fp, xlQualityStandard, True
    Sheets("Control_Panel").Select
    
    MsgBox "✅ PDF 저장 완료:" & vbCrLf & fp, vbInformation, "Export Complete"
End Sub

' === 5. 지연 시뮬레이션 ===
Sub SimulateDelay()
    Dim delayDays As Integer, origStart As Date
    Dim wsCtrl As Worksheet
    
    Set wsCtrl = Sheets("Control_Panel")
    origStart = wsCtrl.Range("B4").Value
    
    delayDays = InputBox("시뮬레이션할 지연 일수를 입력하세요:" & vbCrLf & _
                         "(현재 시작일: " & Format(origStart, "YYYY-MM-DD") & ")", _
                         "Delay Simulation", "7")
    
    If IsNumeric(delayDays) And delayDays <> 0 Then
        wsCtrl.Range("B4").Value = origStart + delayDays
        Call UpdateAllSchedules
        
        MsgBox "시뮬레이션 결과:" & vbCrLf & _
               "새 시작일: " & Format(wsCtrl.Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
               "새 완료일: " & Format(wsCtrl.Range("B9").Value, "YYYY-MM-DD") & vbCrLf & _
               "목표 대비: " & wsCtrl.Range("B11").Value, vbInformation, "Simulation Result"
        
        If MsgBox("원래 일정으로 복원하시겠습니까?", vbYesNo + vbQuestion, "Restore?") = vbYes Then
            wsCtrl.Range("B4").Value = origStart
            Call UpdateAllSchedules
        End If
    End If
End Sub

' === 6. Critical Path 강조 ===
Sub HighlightCritical()
    Dim wsd As Worksheet, i As Long
    
    Set wsd = Sheets("Schedule_Data")
    
    ' Reset
    wsd.Range("A6:I90").Font.Bold = False
    wsd.Range("A6:I90").Font.Color = RGB(0, 0, 0)
    
    ' Highlight Jack-down and Milestones
    For i = 6 To 87
        If wsd.Cells(i, 4).Value = "JACKDOWN" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(183, 28, 28)
        ElseIf wsd.Cells(i, 4).Value = "MILESTONE" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(21, 101, 192)
        End If
    Next i
    
    MsgBox "✅ Critical Path 강조 완료" & vbCrLf & _
           "🔴 빨강 = Jack-down (Critical)" & vbCrLf & _
           "🔵 파랑 = Milestone", vbInformation, "Critical Path"
End Sub

' === 7. 오늘 날짜 하이라이트 ===
Sub HighlightToday()
    Dim ws As Worksheet, j As Long
    Dim projStart As Date, dc As Long
    
    Set ws = Sheets("Gantt_Chart")
    projStart = Sheets("Control_Panel").Range("B4").Value
    dc = 8
    
    For j = 0 To 41
        If projStart + j = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(85, dc + j)).Interior.Color = RGB(255, 255, 200)
            ws.Cells(3, dc + j).Value = "TODAY"
            ws.Cells(3, dc + j).Font.Bold = True
            ws.Cells(3, dc + j).Font.Color = RGB(255, 0, 0)
            MsgBox "오늘 날짜 (" & Format(Date, "MM/DD") & ") 컬럼이 강조되었습니다.", vbInformation
            Exit For
        End If
    Next j
End Sub

' === 8. 날짜 변경 자동 트리거 (Control_Panel 시트에 추가) ===
' 아래 코드를 Control_Panel 시트의 코드 영역에 붙여넣으세요:
'
' Private Sub Worksheet_Change(ByVal Target As Range)
'     If Target.Address = "$B$4" Then
'         Call UpdateAllSchedules
'     End If
' End Sub

' === 9. 진행률 일괄 업데이트 ===
Sub BulkProgressUpdate()
    Dim wsd As Worksheet, i As Long
    Dim pctValue As Double
    
    pctValue = InputBox("일괄 적용할 진행률을 입력하세요 (0-100):", "Bulk Progress", "50")
    
    If IsNumeric(pctValue) Then
        pctValue = pctValue / 100
        Set wsd = Sheets("Schedule_Data")
        
        ' Progress 컬럼이 없으면 추가
        If wsd.Cells(5, 10).Value <> "Progress" Then
            wsd.Cells(5, 10).Value = "Progress"
            wsd.Cells(5, 10).Font.Bold = True
            wsd.Cells(5, 10).Font.Color = RGB(255, 255, 255)
            wsd.Cells(5, 10).Fill.Color = RGB(31, 78, 121)
        End If
        
        For i = 6 To 87
            If wsd.Cells(i, 1).Value <> "" Then
                wsd.Cells(i, 10).Value = pctValue
                wsd.Cells(i, 10).NumberFormat = "0%"
            End If
        Next i
        
        MsgBox "진행률 " & Format(pctValue, "0%") & " 일괄 적용 완료", vbInformation
    End If
End Sub

' === 10. Shamal 위험 체크 ===
Sub CheckShamalRisk()
    Dim wsd As Worksheet, i As Long
    Dim taskDate As Date, shamalStart As Date, shamalEnd As Date
    Dim riskTasks As String, cnt As Long
    
    Set wsd = Sheets("Schedule_Data")
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    
    For i = 6 To 87
        If IsDate(wsd.Cells(i, 6).Value) Then
            taskDate = wsd.Cells(i, 6).Value
            If taskDate >= shamalStart And taskDate <= shamalEnd Then
                ' SAIL tasks are weather-critical
                If wsd.Cells(i, 4).Value = "SAIL" Or wsd.Cells(i, 4).Value = "LOADOUT" Then
                    cnt = cnt + 1
                    riskTasks = riskTasks & vbCrLf & "  ⚠️ " & wsd.Cells(i, 1).Value & ": " & wsd.Cells(i, 3).Value
                End If
            End If
        End If
    Next i
    
    If cnt > 0 Then
        MsgBox "⚠️ SHAMAL 위험 경고!" & vbCrLf & vbCrLf & _
               "Shamal 기간 (Feb 5-14) 중 " & cnt & "개 기상 민감 작업 발견:" & vbCrLf & _
               riskTasks & vbCrLf & vbCrLf & _
               "일정 조정을 권장합니다.", vbExclamation, "Weather Risk Alert"
    Else
        MsgBox "✅ Shamal 기간 중 기상 민감 작업 없음" & vbCrLf & _
               "현재 일정은 안전합니다.", vbInformation, "Weather Check OK"
    End If
End Sub
'''
    
    for i, line in enumerate(vba_code.strip().split('\n'), 3):
        ws_vba.cell(i, 1, value=line)
        ws_vba.cell(i, 1).font = Font(name="Consolas", size=9)
    
    ws_vba.column_dimensions["A"].width = 100
    
    return wb

if __name__ == "__main__":
    import os
    import sys
    # Windows 콘솔 UTF-8 인코딩 설정
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
    
    print("Generating AGI TR 7-Voyage Master Gantt with VBA...")
    wb = create_gantt_with_vba()
    # Windows 호환: 현재 작업 디렉토리에 저장
    output_path = os.path.join(os.getcwd(), "AGI_TR_7Voyage_Master_Gantt_VBA.xlsx")
    wb.save(output_path)
    print(f"[OK] Generated: {output_path}")
