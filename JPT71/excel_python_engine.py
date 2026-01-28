#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to Python Implementation Engine

Excel 파일을 Python으로 완전히 구현하기 위한 엔진
- 클래스 구조 설계
- Excel 함수를 Python 함수로 변환
- 셀 참조 처리 로직
- 계산 엔진
- 포맷/스타일 재현
"""

import sys
import io
import re
import json
from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional, Tuple, Set, Union
from datetime import datetime, date, timedelta
from collections import defaultdict, deque
from enum import Enum
import math

# UTF-8 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


class CellType(Enum):
    """셀 데이터 타입"""
    VALUE = "value"
    FORMULA = "formula"
    EMPTY = "empty"


@dataclass
class CellReference:
    """셀 참조 파싱 및 처리"""
    sheet: Optional[str] = None
    column: str = ""
    row: int = 0
    absolute_column: bool = False
    absolute_row: bool = False
    
    @classmethod
    def parse(cls, ref: str, current_sheet: str = None) -> 'CellReference':
        """
        셀 참조 파싱
        예: "A1", "$B$2", "Sheet1!A1", "Sheet1!$B$2", "$A1", "A$1"
        """
        ref = ref.strip()
        
        # 시트 이름 추출
        sheet = None
        if '!' in ref:
            sheet, cell_part = ref.split('!', 1)
            sheet = sheet.strip("'\"")
        else:
            cell_part = ref
        
        # 절대 참조 확인
        absolute_column = cell_part.startswith('$')
        if absolute_column:
            cell_part = cell_part[1:]
        
        # 열과 행 분리
        column = ""
        row_str = ""
        absolute_row = False
        
        # 열 부분 추출
        i = 0
        while i < len(cell_part) and cell_part[i].isalpha():
            column += cell_part[i]
            i += 1
        
        # 행 부분 추출 (앞에 $가 있을 수 있음)
        if i < len(cell_part) and cell_part[i] == '$':
            absolute_row = True
            i += 1
        
        while i < len(cell_part) and cell_part[i].isdigit():
            row_str += cell_part[i]
            i += 1
        
        if not row_str or not column:
            raise ValueError(f"Invalid cell reference: {ref}")
        
        row = int(row_str)
        
        return cls(
            sheet=sheet or current_sheet,
            column=column,
            row=row,
            absolute_column=absolute_column,
            absolute_row=absolute_row
        )
    
    def to_string(self) -> str:
        """셀 참조를 문자열로 변환"""
        col = f"${self.column}" if self.absolute_column else self.column
        row = f"${self.row}" if self.absolute_row else str(self.row)
        
        if self.sheet:
            return f"{self.sheet}!{col}{row}"
        return f"{col}{row}"
    
    def resolve(self, base_sheet: str, base_row: int, base_col: str) -> Tuple[str, str, int]:
        """상대 참조를 절대 좌표로 변환"""
        sheet = self.sheet or base_sheet
        
        if self.absolute_column:
            col = self.column
        else:
            # 상대 열 계산
            base_col_num = CellReference._column_to_number(base_col)
            col_num = CellReference._column_to_number(self.column)
            new_col_num = base_col_num + (col_num - 1)
            col = CellReference._number_to_column(new_col_num)
        
        if self.absolute_row:
            row = self.row
        else:
            row = base_row + (self.row - 1)
        
        return (sheet, col, row)
    
    @staticmethod
    def _column_to_number(col: str) -> int:
        """열 문자를 숫자로 변환 (A=1, B=2, ..., Z=26, AA=27, ...)"""
        result = 0
        for char in col:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result
    
    @staticmethod
    def _number_to_column(num: int) -> str:
        """숫자를 열 문자로 변환"""
        result = ""
        while num > 0:
            num -= 1
            result = chr(ord('A') + (num % 26)) + result
            num //= 26
        return result


@dataclass
class ExcelCell:
    """Excel 셀 데이터 및 포맷"""
    coordinate: str  # "A1"
    value: Any = None
    formula: Optional[str] = None
    data_type: CellType = CellType.EMPTY
    
    # 포맷 정보
    font: Optional[Dict] = None
    fill: Optional[Dict] = None
    alignment: Optional[Dict] = None
    border: Optional[Dict] = None
    number_format: Optional[str] = None
    
    # 계산된 값
    calculated_value: Any = None
    is_calculating: bool = False
    
    def get_value(self) -> Any:
        """셀의 최종 값 반환"""
        if self.calculated_value is not None:
            return self.calculated_value
        return self.value


@dataclass
class ExcelSheet:
    """Excel 시트"""
    name: str
    rows: int = 0
    cols: int = 0
    cells: Dict[str, ExcelCell] = field(default_factory=dict)
    column_widths: Dict[str, float] = field(default_factory=dict)
    row_heights: Dict[int, float] = field(default_factory=dict)
    merged_cells: List[str] = field(default_factory=list)
    
    def get_cell(self, coordinate: str) -> Optional[ExcelCell]:
        """셀 가져오기"""
        return self.cells.get(coordinate.upper())
    
    def set_cell(self, coordinate: str, cell: ExcelCell):
        """셀 설정"""
        self.cells[coordinate.upper()] = cell
    
    def get_cell_value(self, coordinate: str) -> Any:
        """셀 값 가져오기"""
        cell = self.get_cell(coordinate)
        if cell:
            return cell.get_value()
        return None


class FormulaEngine:
    """Excel 함수 계산 엔진"""
    
    def __init__(self, workbook: 'ExcelWorkbook'):
        self.workbook = workbook
        self.function_registry = self._register_functions()
    
    def _register_functions(self) -> Dict[str, callable]:
        """Excel 함수 등록"""
        return {
            'IF': self._excel_if,
            'IFERROR': self._excel_iferror,
            'INDEX': self._excel_index,
            'ROW': self._excel_row,
            'SMALL': self._excel_small,
            'VLOOKUP': self._excel_vlookup,
            'HYPERLINK': self._excel_hyperlink,
            'SUBSTITUTE': self._excel_substitute,
            'DATE': self._excel_date,
            'WEEKDAY': self._excel_weekday,
            'UPPER': self._excel_upper,
            'TEXT': self._excel_text,
            'COUNTIF': self._excel_countif,
            'TEXTJOIN': self._excel_textjoin,
            'OFFSET': self._excel_offset,
        }
    
    def evaluate(self, formula: str, sheet_name: str, cell_coord: str) -> Any:
        """
        함수 평가
        """
        if not formula or not formula.startswith('='):
            return formula
        
        formula = formula[1:].strip()  # '=' 제거
        
        # 셀 참조 추출 및 값으로 치환
        formula = self._replace_cell_references(formula, sheet_name, cell_coord)
        
        # 함수 호출 처리
        try:
            result = self._evaluate_expression(formula, sheet_name, cell_coord)
            return result
        except Exception as e:
            return f"#ERROR: {str(e)}"
    
    def _replace_cell_references(self, formula: str, sheet_name: str, cell_coord: str) -> str:
        """셀 참조를 값으로 치환"""
        # 셀 참조 패턴: Sheet!A1, A1, $A$1 등
        pattern = r"([A-Za-z0-9_]+!)?(\$?[A-Z]+\$?\d+)"
        
        def replace_ref(match):
            sheet_ref = match.group(1)
            cell_ref = match.group(2)
            
            if sheet_ref:
                ref_sheet = sheet_ref.rstrip('!')
            else:
                ref_sheet = sheet_name
            
            try:
                ref = CellReference.parse(f"{ref_sheet}!{cell_ref}" if sheet_ref else cell_ref, sheet_name)
                value = self._get_cell_value(ref, sheet_name, cell_coord)
                
                # 값이 문자열이면 따옴표로 감싸기
                if isinstance(value, str) and not value.replace('.', '').replace('-', '').isdigit():
                    return f'"{value}"'
                return str(value) if value is not None else "0"
            except:
                return match.group(0)
        
        return re.sub(pattern, replace_ref, formula)
    
    def _get_cell_value(self, ref: CellReference, current_sheet: str, current_cell: str) -> Any:
        """셀 참조로부터 값 가져오기"""
        sheet_name = ref.sheet or current_sheet
        sheet = self.workbook.get_sheet(sheet_name)
        
        if not sheet:
            return None
        
        coord = f"{ref.column}{ref.row}"
        cell = sheet.get_cell(coord)
        
        if cell and cell.formula:
            # 함수인 경우 재귀적으로 계산
            return self.evaluate(cell.formula, sheet_name, coord)
        
        if cell:
            return cell.get_value()
        
        return None
    
    def _evaluate_expression(self, expr: str, sheet_name: str, cell_coord: str) -> Any:
        """표현식 평가 (간단한 수식 및 함수 호출)"""
        expr = expr.strip()
        
        # 함수 호출 찾기
        func_pattern = r"([A-Z_]+)\s*\("
        match = re.search(func_pattern, expr)
        
        if match:
            func_name = match.group(1)
            if func_name in self.function_registry:
                # 함수 인자 추출
                args_str = self._extract_function_args(expr, match.end())
                args = self._parse_arguments(args_str, sheet_name, cell_coord)
                return self.function_registry[func_name](args, sheet_name, cell_coord)
        
        # 간단한 수식 평가
        try:
            # 안전한 수식 평가 (숫자 연산만)
            if re.match(r'^[\d\s\+\-\*\/\(\)\.]+$', expr):
                return eval(expr)
            return expr
        except:
            return expr
    
    def _extract_function_args(self, expr: str, start_pos: int) -> str:
        """함수 인자 문자열 추출"""
        depth = 0
        i = start_pos
        while i < len(expr):
            if expr[i] == '(':
                depth += 1
            elif expr[i] == ')':
                depth -= 1
                if depth == 0:
                    return expr[start_pos:i]
            i += 1
        return expr[start_pos:]
    
    def _parse_arguments(self, args_str: str, sheet_name: str, cell_coord: str) -> List[Any]:
        """함수 인자 파싱"""
        if not args_str.strip():
            return []
        
        args = []
        current = ""
        depth = 0
        in_quotes = False
        
        for char in args_str:
            if char == '"' and (not current or current[-1] != '\\'):
                in_quotes = not in_quotes
                current += char
            elif char == ',' and not in_quotes and depth == 0:
                args.append(self._parse_value(current.strip(), sheet_name, cell_coord))
                current = ""
            elif char == '(':
                depth += 1
                current += char
            elif char == ')':
                depth -= 1
                current += char
            else:
                current += char
        
        if current.strip():
            args.append(self._parse_value(current.strip(), sheet_name, cell_coord))
        
        return args
    
    def _parse_value(self, value: str, sheet_name: str, cell_coord: str) -> Any:
        """값 파싱 (숫자, 문자열, 불리언 등)"""
        value = value.strip()
        
        if value.startswith('"') and value.endswith('"'):
            return value[1:-1]
        elif value.upper() == 'TRUE':
            return True
        elif value.upper() == 'FALSE':
            return False
        elif value.replace('.', '').replace('-', '').isdigit():
            return float(value) if '.' in value else int(value)
        else:
            return value
    
    # Excel 함수 구현
    def _excel_if(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """IF(condition, true_value, false_value)"""
        if len(args) < 2:
            return None
        condition = args[0]
        true_value = args[1]
        false_value = args[2] if len(args) > 2 else None
        
        if condition:
            return true_value
        return false_value
    
    def _excel_iferror(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """IFERROR(value, error_value)"""
        if len(args) < 1:
            return None
        try:
            value = args[0]
            if isinstance(value, str) and value.startswith('#ERROR'):
                return args[1] if len(args) > 1 else None
            return value
        except:
            return args[1] if len(args) > 1 else None
    
    def _excel_index(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """INDEX(array, row_num, [col_num])"""
        if len(args) < 2:
            return None
        array = args[0]
        row_num = int(args[1]) if isinstance(args[1], (int, float)) else 1
        col_num = int(args[2]) if len(args) > 2 and isinstance(args[2], (int, float)) else 1
        
        if isinstance(array, list):
            if isinstance(array[0], list):
                # 2D 배열
                row_idx = row_num - 1
                col_idx = col_num - 1
                if 0 <= row_idx < len(array) and 0 <= col_idx < len(array[row_idx]):
                    return array[row_idx][col_idx]
            else:
                # 1D 배열
                idx = row_num - 1
                if 0 <= idx < len(array):
                    return array[idx]
        return None
    
    def _excel_row(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """ROW([reference])"""
        if len(args) > 0 and isinstance(args[0], str):
            # 셀 참조에서 행 번호 추출
            ref = CellReference.parse(args[0], sheet_name)
            return ref.row
        else:
            # 현재 셀의 행 번호
            match = re.match(r'([A-Z]+)(\d+)', cell_coord)
            if match:
                return int(match.group(2))
        return 1
    
    def _excel_small(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """SMALL(array, k)"""
        if len(args) < 2:
            return None
        array = args[0]
        k = int(args[1]) if isinstance(args[1], (int, float)) else 1
        
        if isinstance(array, list):
            numbers = [x for x in array if isinstance(x, (int, float))]
            if numbers:
                sorted_nums = sorted(numbers)
                if 1 <= k <= len(sorted_nums):
                    return sorted_nums[k - 1]
        return None
    
    def _excel_vlookup(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])"""
        if len(args) < 3:
            return None
        
        lookup_value = args[0]
        table_array = args[1]
        col_index = int(args[2]) if isinstance(args[2], (int, float)) else 1
        range_lookup = args[3] if len(args) > 3 else True
        
        if not isinstance(table_array, list):
            return None
        
        for row in table_array:
            if isinstance(row, list) and len(row) > 0:
                if row[0] == lookup_value:
                    if 1 <= col_index <= len(row):
                        return row[col_index - 1]
        
        return None
    
    def _excel_hyperlink(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """HYPERLINK(link_location, [friendly_name])"""
        if len(args) < 1:
            return None
        return args[1] if len(args) > 1 else args[0]
    
    def _excel_substitute(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """SUBSTITUTE(text, old_text, new_text, [instance_num])"""
        if len(args) < 3:
            return None
        
        text = str(args[0])
        old_text = str(args[1])
        new_text = str(args[2])
        instance_num = int(args[3]) if len(args) > 3 else None
        
        if instance_num:
            # 특정 인스턴스만 교체
            parts = text.split(old_text)
            if len(parts) > instance_num:
                result = old_text.join(parts[:instance_num]) + new_text + old_text.join(parts[instance_num:])
                return result
        return text.replace(old_text, new_text)
    
    def _excel_date(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """DATE(year, month, day)"""
        if len(args) < 3:
            return None
        try:
            year = int(args[0])
            month = int(args[1])
            day = int(args[2])
            return date(year, month, day)
        except:
            return None
    
    def _excel_weekday(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """WEEKDAY(serial_number, [return_type])"""
        if len(args) < 1:
            return None
        
        return_type = int(args[1]) if len(args) > 1 else 1
        
        date_value = args[0]
        if isinstance(date_value, date):
            weekday = date_value.weekday()
        elif isinstance(date_value, (int, float)):
            # Excel 날짜 시리얼 번호 (1900-01-01 기준)
            base_date = date(1900, 1, 1)
            target_date = base_date + timedelta(days=int(date_value) - 2)
            weekday = target_date.weekday()
        else:
            return None
        
        # return_type에 따른 변환
        if return_type == 1:  # 1 (일요일) ~ 7 (토요일)
            return weekday + 2 if weekday < 5 else weekday - 5
        elif return_type == 2:  # 1 (월요일) ~ 7 (일요일)
            return weekday + 1
        else:
            return weekday + 1
    
    def _excel_upper(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """UPPER(text)"""
        if len(args) < 1:
            return None
        return str(args[0]).upper()
    
    def _excel_text(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """TEXT(value, format_text)"""
        if len(args) < 2:
            return None
        
        value = args[0]
        format_text = str(args[1])
        
        if isinstance(value, date):
            # 날짜 포맷팅
            if 'yyyy' in format_text or 'YYYY' in format_text:
                return value.strftime('%Y')
            elif 'mm' in format_text or 'MM' in format_text:
                return value.strftime('%m')
            elif 'dd' in format_text or 'DD' in format_text:
                return value.strftime('%d')
        elif isinstance(value, (int, float)):
            # 숫자 포맷팅
            return format(value, format_text)
        
        return str(value)
    
    def _excel_countif(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """COUNTIF(range, criteria)"""
        if len(args) < 2:
            return 0
        
        range_data = args[0]
        criteria = args[1]
        
        if not isinstance(range_data, list):
            return 0
        
        count = 0
        for item in range_data:
            if isinstance(item, list):
                for subitem in item:
                    if self._match_criteria(subitem, criteria):
                        count += 1
            else:
                if self._match_criteria(item, criteria):
                    count += 1
        
        return count
    
    def _match_criteria(self, value: Any, criteria: Any) -> bool:
        """조건 매칭"""
        if isinstance(criteria, str):
            if criteria.startswith('='):
                return value == criteria[1:]
            elif criteria.startswith('>'):
                return value > float(criteria[1:]) if isinstance(value, (int, float)) else False
            elif criteria.startswith('<'):
                return value < float(criteria[1:]) if isinstance(value, (int, float)) else False
            else:
                return str(value) == criteria
        return value == criteria
    
    def _excel_textjoin(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """TEXTJOIN(delimiter, ignore_empty, text1, [text2], ...)"""
        if len(args) < 3:
            return None
        
        delimiter = str(args[0])
        ignore_empty = args[1]
        texts = args[2:]
        
        result = []
        for text in texts:
            if isinstance(text, list):
                for item in text:
                    if not ignore_empty or (item is not None and str(item).strip()):
                        result.append(str(item))
            else:
                if not ignore_empty or (text is not None and str(text).strip()):
                    result.append(str(text))
        
        return delimiter.join(result)
    
    def _excel_offset(self, args: List[Any], sheet_name: str, cell_coord: str) -> Any:
        """OFFSET(reference, rows, cols, [height], [width])"""
        if len(args) < 3:
            return None
        
        ref_str = str(args[0])
        rows_offset = int(args[1]) if isinstance(args[1], (int, float)) else 0
        cols_offset = int(args[2]) if isinstance(args[2], (int, float)) else 0
        
        try:
            ref = CellReference.parse(ref_str, sheet_name)
            new_row = ref.row + rows_offset
            new_col_num = CellReference._column_to_number(ref.column) + cols_offset
            new_col = CellReference._number_to_column(new_col_num)
            
            new_coord = f"{new_col}{new_row}"
            return self._get_cell_value(
                CellReference.parse(new_coord, sheet_name),
                sheet_name,
                new_coord
            )
        except:
            return None


class StyleEngine:
    """포맷/스타일 재현 엔진"""
    
    @staticmethod
    def apply_format_to_openpyxl(cell, excel_cell: ExcelCell):
        """openpyxl 셀에 포맷 적용"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 폰트
        if excel_cell.font:
            font_color = excel_cell.font.get('color', None)
            # 색상 문자열을 aRGB hex로 변환
            if font_color and isinstance(font_color, str):
                # "FF000000" 형식이면 그대로 사용, "000000" 형식이면 "FF" 추가
                if len(font_color) == 6:
                    font_color = "FF" + font_color
                elif len(font_color) != 8:
                    font_color = None
            
            cell.font = Font(
                name=excel_cell.font.get('name', 'Calibri'),
                size=excel_cell.font.get('size', 11),
                bold=excel_cell.font.get('bold', False),
                italic=excel_cell.font.get('italic', False),
                underline=excel_cell.font.get('underline', None),
                color=font_color
            )
        
        # 배경색
        if excel_cell.fill:
            fill_color = excel_cell.fill.get('fgColor', None)
            # 색상 문자열을 aRGB hex로 변환
            if fill_color and isinstance(fill_color, str):
                # "FF000000" 형식이면 그대로 사용, "000000" 형식이면 "FF" 추가
                if len(fill_color) == 6:
                    fill_color = "FF" + fill_color
                elif len(fill_color) != 8:
                    fill_color = None
            
            # fgColor가 None이면 fill을 설정하지 않음
            if fill_color:
                cell.fill = PatternFill(
                    patternType=excel_cell.fill.get('patternType', 'solid'),
                    fgColor=fill_color
                )
        
        # 정렬
        if excel_cell.alignment:
            cell.alignment = Alignment(
                horizontal=excel_cell.alignment.get('horizontal', 'general'),
                vertical=excel_cell.alignment.get('vertical', 'bottom'),
                wrap_text=excel_cell.alignment.get('wrapText', False)
            )
        
        # 테두리
        if excel_cell.border:
            border = Border()
            if excel_cell.border.get('left'):
                border.left = Side(style=excel_cell.border['left'])
            if excel_cell.border.get('right'):
                border.right = Side(style=excel_cell.border['right'])
            if excel_cell.border.get('top'):
                border.top = Side(style=excel_cell.border['top'])
            if excel_cell.border.get('bottom'):
                border.bottom = Side(style=excel_cell.border['bottom'])
            cell.border = border
        
        # 숫자 포맷
        if excel_cell.number_format:
            cell.number_format = excel_cell.number_format


@dataclass
class ExcelWorkbook:
    """Excel 워크북"""
    sheets: Dict[str, ExcelSheet] = field(default_factory=dict)
    formula_engine: Optional[FormulaEngine] = None
    
    def __post_init__(self):
        self.formula_engine = FormulaEngine(self)
    
    def get_sheet(self, name: str) -> Optional[ExcelSheet]:
        """시트 가져오기"""
        return self.sheets.get(name)
    
    def add_sheet(self, sheet: ExcelSheet):
        """시트 추가"""
        self.sheets[sheet.name] = sheet
    
    def calculate_all(self):
        """모든 함수 계산"""
        # 의존성 그래프 생성
        dependency_graph = self._build_dependency_graph()
        
        # 위상 정렬로 계산 순서 결정
        calculation_order = self._topological_sort(dependency_graph)
        
        # 순서대로 계산
        for cell_ref in calculation_order:
            sheet_name, coord = cell_ref
            sheet = self.get_sheet(sheet_name)
            if sheet:
                cell = sheet.get_cell(coord)
                if cell and cell.formula:
                    try:
                        cell.calculated_value = self.formula_engine.evaluate(
                            cell.formula, sheet_name, coord
                        )
                    except Exception as e:
                        cell.calculated_value = f"#ERROR: {str(e)}"
    
    def _build_dependency_graph(self) -> Dict[Tuple[str, str], Set[Tuple[str, str]]]:
        """의존성 그래프 생성"""
        graph = defaultdict(set)
        
        for sheet_name, sheet in self.sheets.items():
            for coord, cell in sheet.cells.items():
                if cell.formula:
                    # 함수에서 참조하는 셀 찾기
                    refs = self._extract_cell_references(cell.formula, sheet_name)
                    for ref_sheet, ref_coord in refs:
                        graph[(ref_sheet, ref_coord)].add((sheet_name, coord))
        
        return graph
    
    def _extract_cell_references(self, formula: str, current_sheet: str) -> List[Tuple[str, str]]:
        """함수에서 셀 참조 추출"""
        refs = []
        pattern = r"([A-Za-z0-9_]+!)?(\$?[A-Z]+\$?\d+)"
        
        for match in re.finditer(pattern, formula):
            sheet_ref = match.group(1)
            cell_ref = match.group(2)
            
            if sheet_ref:
                ref_sheet = sheet_ref.rstrip('!')
            else:
                ref_sheet = current_sheet
            
            refs.append((ref_sheet, cell_ref.upper()))
        
        return refs
    
    def _topological_sort(self, graph: Dict) -> List[Tuple[str, str]]:
        """위상 정렬 (순환 참조 감지)"""
        in_degree = defaultdict(int)
        all_nodes = set()
        
        # 모든 노드 수집
        for node in graph:
            all_nodes.add(node)
            for dep in graph[node]:
                all_nodes.add(dep)
                in_degree[dep] += 1
        
        # 진입 차수가 0인 노드 찾기
        queue = deque([node for node in all_nodes if in_degree[node] == 0])
        result = []
        visited = set()
        
        while queue:
            node = queue.popleft()
            if node in visited:
                continue
            visited.add(node)
            result.append(node)
            
            # 의존 노드의 진입 차수 감소
            for dep in graph.get(node, set()):
                in_degree[dep] -= 1
                if in_degree[dep] == 0:
                    queue.append(dep)
        
        # 순환 참조가 있는 경우
        remaining = all_nodes - visited
        if remaining:
            # 남은 노드를 결과에 추가 (순환 참조 경고)
            result.extend(remaining)
        
        return result
    
    @classmethod
    def load_from_excel(cls, excel_path: str) -> 'ExcelWorkbook':
        """Excel 파일에서 로드"""
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_path, data_only=False)
        workbook = cls()
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = ExcelSheet(
                name=sheet_name,
                rows=ws.max_row,
                cols=ws.max_column
            )
            
            # 셀 데이터 추출
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None or cell.data_type == 'f':
                        excel_cell = ExcelCell(
                            coordinate=cell.coordinate,
                            value=cell.value if cell.data_type != 'f' else None,
                            formula=str(cell.value) if cell.data_type == 'f' else None,
                            data_type=CellType.FORMULA if cell.data_type == 'f' else CellType.VALUE
                        )
                        
                        # 포맷 정보 추출
                        if cell.font:
                            excel_cell.font = {
                                'name': cell.font.name,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'underline': cell.font.underline,
                                'color': str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                            }
                        
                        if cell.fill and cell.fill.patternType:
                            excel_cell.fill = {
                                'patternType': cell.fill.patternType,
                                'fgColor': str(cell.fill.fgColor.rgb) if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                            }
                        
                        if cell.alignment:
                            excel_cell.alignment = {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'wrapText': cell.alignment.wrap_text
                            }
                        
                        if cell.number_format:
                            excel_cell.number_format = cell.number_format
                        
                        sheet.set_cell(cell.coordinate, excel_cell)
            
            # 열 너비
            for col_letter in ws.column_dimensions:
                if ws.column_dimensions[col_letter].width:
                    sheet.column_widths[col_letter] = ws.column_dimensions[col_letter].width
            
            # 행 높이
            for row_num in ws.row_dimensions:
                if ws.row_dimensions[row_num].height:
                    sheet.row_heights[row_num] = ws.row_dimensions[row_num].height
            
            # 병합된 셀
            for merged in ws.merged_cells.ranges:
                sheet.merged_cells.append(str(merged))
            
            workbook.add_sheet(sheet)
        
        return workbook
    
    def save_to_excel(self, output_path: str):
        """Excel 파일로 저장"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        for sheet_name, sheet in self.sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # 셀 데이터 및 포맷 적용
            for coord, excel_cell in sheet.cells.items():
                cell = ws[coord]
                
                # 값 설정
                if excel_cell.formula:
                    cell.value = excel_cell.formula
                else:
                    cell.value = excel_cell.get_value()
                
                # 포맷 적용
                StyleEngine.apply_format_to_openpyxl(cell, excel_cell)
            
            # 열 너비
            for col_letter, width in sheet.column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # 행 높이
            for row_num, height in sheet.row_heights.items():
                ws.row_dimensions[row_num].height = height
            
            # 병합된 셀
            for merged_range in sheet.merged_cells:
                ws.merge_cells(merged_range)
        
        wb.save(output_path)


if __name__ == "__main__":
    # 사용 예제
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_python_engine.py <excel_file> [output_file]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else excel_path.replace('.xlsx', '_python.xlsx')
    
    print(f"Loading Excel file: {excel_path}")
    workbook = ExcelWorkbook.load_from_excel(excel_path)
    
    print(f"Calculating formulas...")
    workbook.calculate_all()
    
    print(f"Saving to: {output_path}")
    workbook.save_to_excel(output_path)
    
    print("Done!")

