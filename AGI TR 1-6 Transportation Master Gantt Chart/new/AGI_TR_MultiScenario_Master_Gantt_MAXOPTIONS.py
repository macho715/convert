
#!/usr/bin/env python3
"""
AGI TR Multi-Scenario Master Gantt (Max Options Edition)

목표
- 여러 시나리오(무제한) TSV/CSV/XLSX 입력을 받아 Excel 기반 Master Gantt 워크북을 생성합니다.
- Control_Panel(설정), Scenario별 Schedule_Data + Gantt_Chart, Tide_Data, Summary/Weather/KPIs 등을 생성합니다.
- Gantt는 기본적으로 "조건부 서식(Conditional Formatting)" 기반으로 렌더링하여,
  PROJECT_START/기간 설정 변경 시 VBA 없이도 자동 갱신되도록 설계했습니다.
- (선택) VBA 코드(.bas) 내보내기, Windows+Excel 환경에서 .xlsm 자동 생성/ PDF 내보내기 옵션을 제공합니다.

의존성
- Python 3.10+
- openpyxl

권장 실행 예:
  python AGI_TR_MultiScenario_Master_Gantt_MAXOPTIONS.py \
    --original MOS.tsv \
    --scenario-a MFC1.tsv \
    --alternative MFA2.tsv \
    --tide-tsv Tide.tsv \
    --export-vba-bas

주의
- openpyxl은 VBA를 직접 삽입할 수 없습니다. 본 스크립트는 'VBA_Code' 시트에 코드를 생성하고
  --export-vba-bas 옵션으로 .bas 모듈 파일을 추출할 수 있습니다.
- --create-xlsm / --export-pdf 는 Windows + Excel + "Trust access to the VBA project object model" 설정이 필요합니다.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Union

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# ==========================
# Styling / Constants
# ==========================

COLORS: dict[str, str] = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
    "WARN": "FFCDD2",
    "CAUTION": "FFE0B2",
    "OK": "C8E6C9",
    "WEEKEND": "ECEFF1",
    "MAINT": "E0E0E0",
}

PHASE_ORDER = [
    "MOBILIZATION",
    "DECK_PREP",
    "LOADOUT",
    "SEAFAST",
    "SAIL",
    "AGI_UNLOAD",
    "TURNING",
    "JACKDOWN",
    "RETURN",
    "BUFFER",
    "MILESTONE",
]

PHASE_COLOR_KEYS = {p: p for p in PHASE_ORDER}

BORDER_SIDE = Side(style="thin", color="A6A6A6")


def thin_border() -> Border:
    return Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


DurRef = Union[str, int, float]


DEFAULT_DURATION_VALUES: dict[str, float] = {
    "DUR_MOB": 1.0,
    "DUR_DECK": 3.0,
    "DUR_LO": 1.0,
    "DUR_SF": 0.5,
    "DUR_MWS": 0.5,
    "DUR_SAIL": 1.0,
    "DUR_UL": 1.0,
    "DUR_TURN": 3.0,
    "DUR_JD": 1.0,
    "DUR_RET": 1.0,
    "DUR_BUF": 0.5,
}

# Common WORKDAY.INTL weekend patterns
# Excel expects a 7-char string (Mon..Sun) where '1' means weekend.
WEEKEND_PATTERNS: dict[str, str] = {
    "SAT_SUN": "0000011",
    "FRI_SAT": "0000110",
    "SUN_ONLY": "0000001",
    "FRI_ONLY": "0000100",
}


# ==========================
# Data Models
# ==========================

@dataclass(frozen=True)
class Scenario:
    name: str
    input_spec: Optional[str] = None  # path or "pattern:1-2-2-2"
    pattern: Optional[str] = None
    early_return: bool = False
    cycle_spacing: int = 15
    include_group_rows: bool = False  # include P6 group/summary rows
    label: Optional[str] = None       # display label


@dataclass(frozen=True)
class Task:
    task_id: str
    wbs: str
    name: str
    phase: str
    owner: str
    offset_days: int
    duration: DurRef
    notes: str = ""
    planned_start: Optional[dt.date] = None
    planned_finish: Optional[dt.date] = None
    actual_start: Optional[dt.date] = None
    actual_finish: Optional[dt.date] = None
    percent_complete: Optional[float] = None


# ==========================
# Parsing utilities
# ==========================

def parse_iso_date(value: str) -> dt.date:
    try:
        return dt.datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except Exception as e:
        raise argparse.ArgumentTypeError(f"Invalid date '{value}'. Expected YYYY-MM-DD") from e


def open_text_with_fallback(path: Path, encoding: Optional[str] = None):
    if encoding:
        return open(path, "r", encoding=encoding, newline="")
    for enc in ("utf-8-sig", "utf-8"):
        try:
            return open(path, "r", encoding=enc, newline="")
        except UnicodeDecodeError:
            continue
    return open(path, "r", encoding="utf-8", newline="", errors="replace")


def sniff_delimiter(sample: str, default: str = "\t") -> str:
    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return default


def is_number_like(value: str) -> bool:
    v = (value or "").strip()
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", v))


def parse_date(value: Any, project_year_hint: Optional[int] = None) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass

    for fmt in ("%d-%b-%Y", "%d-%b-%y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass

    if project_year_hint is not None:
        try:
            d = dt.datetime.strptime(s, "%d-%b").date()
            return d.replace(year=project_year_hint)
        except Exception:
            pass

    return None


def infer_phase(task_name: str) -> str:
    u = (task_name or "").upper()
    if "MOBILIZATION" in u or "DEMOBILIZATION" in u or "DEMOB" in u:
        return "MOBILIZATION"
    if "DECK" in u and ("PREP" in u or "PREPARATION" in u):
        return "DECK_PREP"
    if "LOADOUT" in u or "LOAD-OUT" in u or "LOAD OUT" in u or "LOAD-IN" in u or "LOAD IN" in u:
        return "LOADOUT"
    if "SEA FASTENING" in u or "SEAFAST" in u or "SEAFASTENING" in u or ("SEA" in u and "FAST" in u):
        return "SEAFAST"
    if "SAIL" in u or "SEA PASSAGE" in u or "SAIL-AWAY" in u:
        return "SAIL"
    if "UNLOAD" in u or "ARRIVAL" in u or "RORO" in u:
        return "AGI_UNLOAD"
    if "TURNING" in u or ("TURN" in u and "RETURN" not in u):
        return "TURNING"
    if "JACKDOWN" in u or "JACK-DOWN" in u or "JACKING DOWN" in u:
        return "JACKDOWN"
    if "RETURN" in u:
        return "RETURN"
    if "VOYAGE" in u or "MILESTONE" in u or (u.startswith("V") and len(u) <= 3):
        return "MILESTONE"
    if "BUFFER" in u or "RESET" in u:
        return "BUFFER"
    return "BUFFER"


PHASE_MAPPING = {
    "Mobilization": "MOBILIZATION",
    "Deck Prep": "DECK_PREP",
    "MZP Loadout": "LOADOUT",
    "Sea Fastening": "SEAFAST",
    "Survey": "BUFFER",
    "Sea Passage": "SAIL",
    "AGI Arrival": "AGI_UNLOAD",
    "AGI Laydown": "BUFFER",
    "Onshore SPMT": "TURNING",
    "AGI Gate Prep": "TURNING",
    "Jackdown": "JACKDOWN",
    "Return": "RETURN",
    "Buffer": "BUFFER",
    "Marine Transport": "MILESTONE",
    "Demobilization": "MOBILIZATION",
    "Handover": "MILESTONE",
}


def duration_ref(duration_str: str, task_name: str) -> DurRef:
    s = (duration_str or "").strip()
    if not s:
        return "DUR_BUF"
    try:
        dur_val = float(s)
    except Exception:
        return "DUR_BUF"

    if dur_val <= 0:
        return 0

    u = (task_name or "").upper()

    if math.isclose(dur_val, 0.5, rel_tol=0.0, abs_tol=1e-9):
        if ("SEA" in u and "FAST" in u) or "SEAFAST" in u:
            return "DUR_SF"
        if "MWS" in u or "MPI" in u or "APPROVAL" in u:
            return "DUR_MWS"
        return "DUR_BUF"

    if math.isclose(dur_val, 1.0, rel_tol=0.0, abs_tol=1e-9):
        if "LOAD" in u:
            return "DUR_LO"
        if "MOBILIZATION" in u or "DEMOB" in u:
            return "DUR_MOB"
        if "SAIL" in u or "SEA PASSAGE" in u:
            return "DUR_SAIL"
        if "ARRIV" in u or "UNLOAD" in u or "RORO" in u:
            return "DUR_UL"
        if "RETURN" in u:
            return "DUR_RET"
        if "JACK" in u:
            return "DUR_JD"
        return "DUR_BUF"

    if math.isclose(dur_val, 3.0, rel_tol=0.0, abs_tol=1e-9):
        if "TURN" in u:
            return "DUR_TURN"
        if "DECK" in u or "PREP" in u:
            return "DUR_DECK"
        return 3

    if dur_val.is_integer():
        return int(dur_val)
    return float(dur_val)


def normalize_sheet_name(name: str, max_len: int = 31) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]+", "_", name).strip()
    return cleaned if len(cleaned) <= max_len else cleaned[:max_len]


def scenario_sheet_names(scenario_name: str) -> tuple[str, str, str]:
    mapping = {
        "Mammoet_Original": "Mammoet_Orig",
        "Mammoet_ScenarioA": "Mammoet_ScnA",
        "Mammoet_Alternative": "Mammoet_Alt",
    }
    short = mapping.get(scenario_name, scenario_name)
    short = normalize_sheet_name(short, max_len=18)
    sched = normalize_sheet_name(f"Schedule_Data_{short}", max_len=31)
    gantt = normalize_sheet_name(f"Gantt_Chart_{short}", max_len=31)
    return short, sched, gantt


# ==========================
# Input loaders
# ==========================

def iter_rows_from_excel(path: Path, sheet_name: Optional[str] = None) -> Iterable[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    headers_clean = [str(h).strip() if h is not None else "" for h in headers]
    for values in rows:
        if not values:
            continue
        if not any(v is not None and str(v).strip() for v in values):
            continue
        values_list = list(values) + [None] * (len(headers_clean) - len(values))
        yield {headers_clean[i]: values_list[i] for i in range(len(headers_clean))}


def iter_rows_from_delimited(
    path: Path,
    delimiter: Optional[str] = None,
    encoding: Optional[str] = None,
) -> Iterable[dict[str, str]]:
    with open_text_with_fallback(path, encoding=encoding) as f:
        sample = f.read(4096)
        f.seek(0)
        delim = delimiter or sniff_delimiter(sample, default="\t" if path.suffix.lower() == ".tsv" else ",")
        reader = csv.reader(f, delimiter=delim)
        headers = next(reader, None)
        if not headers:
            return []
        headers = [h.strip().lstrip("\ufeff") for h in headers]

        for values in reader:
            if not values or not any((v or "").strip() for v in values):
                continue

            expected_p6 = [
                "Activity ID",
                "Activity Name",
                "Original Duration",
                "Planned Start",
                "Planned Finish",
                "Actual Start",
                "Actual Finish",
            ]
            if headers[:7] == expected_p6 and len(values) == len(headers) - 1:
                if len(values) >= 4 and is_number_like(values[1]) and parse_date(values[2]) and parse_date(values[3]):
                    values = [values[0], ""] + values[1:]

            if len(values) < len(headers):
                values = values + [""] * (len(headers) - len(values))
            if len(values) > len(headers):
                values = values[: len(headers)]

            yield {headers[i]: (values[i] or "").strip() for i in range(len(headers))}


def get_case_insensitive(row: dict[str, Any], key: str, default: Any = "") -> Any:
    if key in row:
        return row.get(key, default)
    k_lower = key.lower()
    for k in row.keys():
        if str(k).lower() == k_lower:
            return row.get(k, default)
    return default


def load_tasks(
    input_path: Path,
    project_start: dt.date,
    *,
    delimiter: Optional[str] = None,
    encoding: Optional[str] = None,
    sheet_name: Optional[str] = None,
    include_group_rows: bool = False,
) -> list[Task]:
    tasks: list[Task] = []
    voyage_num = 0
    task_counter = 0
    project_year = project_start.year

    ext = input_path.suffix.lower()
    rows_iter = iter_rows_from_excel(input_path, sheet_name=sheet_name) if ext in (".xlsx", ".xlsm") else iter_rows_from_delimited(input_path, delimiter=delimiter, encoding=encoding)

    for row in rows_iter:
        start_str = (
            str(get_case_insensitive(row, "Planned Start", "")).strip()
            or str(get_case_insensitive(row, "Actual Start", "")).strip()
            or str(get_case_insensitive(row, "Start", "")).strip()
        )
        if not start_str:
            continue
        planned_start = parse_date(start_str, project_year_hint=project_year)
        if not planned_start:
            continue

        finish_str = (
            str(get_case_insensitive(row, "Planned Finish", "")).strip()
            or str(get_case_insensitive(row, "Actual Finish", "")).strip()
            or str(get_case_insensitive(row, "End", "")).strip()
        )
        planned_finish = parse_date(finish_str, project_year_hint=project_year) if finish_str else None

        actual_start = parse_date(get_case_insensitive(row, "Actual Start", ""), project_year_hint=project_year)
        actual_finish = parse_date(get_case_insensitive(row, "Actual Finish", ""), project_year_hint=project_year)

        task_name = (
            str(get_case_insensitive(row, "Task", "")).strip()
            or str(get_case_insensitive(row, "Activity Name", "")).strip()
            or ""
        )
        if task_name and is_number_like(task_name):
            if parse_date(get_case_insensitive(row, "Original Duration", ""), project_year_hint=project_year):
                task_name = ""

        activity_id = str(get_case_insensitive(row, "Activity ID", "")).strip()
        if not task_name:
            task_name = activity_id or str(get_case_insensitive(row, "ID", "")).strip()
        if not task_name:
            continue

        if not include_group_rows:
            activity_name_raw = str(get_case_insensitive(row, "Activity Name", "")).strip()
            if (not activity_name_raw) and activity_id and re.fullmatch(r"[A-Za-z][A-Za-z\s\-_/]*", activity_id):
                continue

        task_id = activity_id or str(get_case_insensitive(row, "ID", "")).strip()
        if not task_id:
            if "VOYAGE" in task_name.upper():
                voyage_num += 1
                task_id = f"V{voyage_num}"
            else:
                task_counter += 1
                prefix = re.sub(r"[^A-Z0-9]", "", (task_name.split()[0][:3].upper() if task_name.split() else "TSK"))
                task_id = f"{prefix or 'TSK'}-{task_counter:03d}"

        wbs = str(get_case_insensitive(row, "WBS", "")).strip() or (f"A{voyage_num}" if voyage_num > 0 else "A0")

        phase_raw = str(get_case_insensitive(row, "Phase", "")).strip()
        phase = PHASE_MAPPING.get(phase_raw, infer_phase(task_name))
        owner = str(get_case_insensitive(row, "Owner", "")).strip() or "All"
        notes = str(get_case_insensitive(row, "Notes", "")).strip()

        if planned_finish:
            try:
                duration_str = str((planned_finish - planned_start).days)
            except Exception:
                duration_str = ""
        else:
            duration_str = (
                str(get_case_insensitive(row, "Original Duration", "")).strip()
                or str(get_case_insensitive(row, "Duration_days", "")).strip()
                or str(get_case_insensitive(row, "Dur", "")).strip()
                or str(get_case_insensitive(row, "Duration", "")).strip()
            )
        dur = duration_ref(duration_str, task_name)

        pc_raw = (
            str(get_case_insensitive(row, "Percent Complete", "")).strip()
            or str(get_case_insensitive(row, "% Complete", "")).strip()
            or str(get_case_insensitive(row, "Progress", "")).strip()
        )
        percent_complete: Optional[float] = None
        if pc_raw:
            try:
                pc_val = float(pc_raw.replace("%", ""))
                if pc_val > 1.0:
                    pc_val = pc_val / 100.0
                percent_complete = max(0.0, min(1.0, pc_val))
            except Exception:
                percent_complete = None

        offset = (planned_start - project_start).days

        tasks.append(
            Task(
                task_id=task_id,
                wbs=wbs,
                name=task_name,
                phase=phase,
                owner=owner,
                offset_days=int(offset),
                duration=dur,
                notes=notes,
                planned_start=planned_start,
                planned_finish=planned_finish,
                actual_start=actual_start,
                actual_finish=actual_finish,
                percent_complete=percent_complete,
            )
        )

    return tasks


# ==========================
# Scenario task generation (pattern-based)
# ==========================

def parse_voyage_pattern(pattern_str: str) -> list[list[int]]:
    if not pattern_str:
        return [[i] for i in range(1, 8)]

    voyage_groups: list[list[int]] = []
    tr_id = 1

    if "x" in pattern_str:
        parts = pattern_str.split("x")
    elif "-" in pattern_str:
        parts = pattern_str.split("-")
    else:
        parts = [pattern_str]

    for part in parts:
        part = part.strip()
        if not part:
            continue
        try:
            count = int(part)
        except ValueError:
            continue
        if count <= 0:
            continue
        voyage_groups.append(list(range(tr_id, tr_id + count)))
        tr_id += count

    return voyage_groups or [[i] for i in range(1, 8)]


def generate_scenario_tasks(
    pattern_str: str,
    project_start: dt.date,
    *,
    cycle_spacing: int = 15,
    early_return: bool = False,
) -> list[Task]:
    tasks: list[Task] = []
    groups = parse_voyage_pattern(pattern_str)

    def add(task_id: str, wbs: str, name: str, phase: str, owner: str, offset: int, dur: DurRef, notes: str = ""):
        tasks.append(
            Task(
                task_id=task_id,
                wbs=wbs,
                name=name,
                phase=phase,
                owner=owner,
                offset_days=offset,
                duration=dur,
                notes=notes,
                planned_start=project_start + dt.timedelta(days=offset),
            )
        )

    offset = 0
    add("MOB-001", "A0", "Mobilization (crew/equipment)", "MOBILIZATION", "Mammoet", offset, "DUR_MOB", "SPMT + grillage in MZP")
    offset += int(math.ceil(DEFAULT_DURATION_VALUES["DUR_MOB"]))
    add("PREP-001", "A0", "LCT deck preparations + fenders + mooring", "DECK_PREP", "Mammoet/KFS", offset, "DUR_DECK", "MWS pre-check ready")
    offset += int(math.ceil(DEFAULT_DURATION_VALUES["DUR_DECK"]))

    cycle_offset = offset

    for v_idx, tr_list in enumerate(groups, start=1):
        n = len(tr_list)
        tr_str = f"TR{tr_list[0]}" if n == 1 else f"TR{tr_list[0]}-TR{tr_list[-1]}"
        wbs = f"A{v_idx}"
        offset = cycle_offset
        voyage_start = offset

        add(f"V{v_idx}", wbs, f"VOYAGE {v_idx}: {tr_str} Transport", "MILESTONE", "SCT/Mammoet/KFS", offset, 0, "TIDE>=1.90 required (Loadout start)")

        lo_dur = n if n > 1 else 1
        add(f"LO-{v_idx:02d}", wbs, f"Loadout {tr_str} onto LCT", "LOADOUT", "Mammoet", offset, lo_dur, "Tide window required")
        offset += lo_dur

        add(f"SF-{v_idx:02d}", wbs, f"Sea fastening + MWS checks ({tr_str})", "SEAFAST", "Mammoet/KFS/MWS", offset, "DUR_SF", "Lashing + survey")
        offset += 1

        add(f"SAIL-{v_idx:02d}", wbs, "Sail-away MZP->AGI", "SAIL", "LCT", offset, "DUR_SAIL", "WX gate")
        offset += 1

        for tr in tr_list:
            add(f"UL-{v_idx:02d}-{tr}", wbs, f"Unload TR{tr} at AGI (1 unit/day)", "AGI_UNLOAD", "Mammoet", offset, "DUR_UL", "RORO + ramp")
            offset += 1

        first_jd_offset: Optional[int] = None
        for tr in tr_list:
            add(f"TURN-{v_idx:02d}-{tr}", wbs, f"Turning TR{tr} (90 deg)", "TURNING", "Mammoet", offset, "DUR_TURN", "3.0d/unit")
            offset += int(math.ceil(DEFAULT_DURATION_VALUES["DUR_TURN"]))
            add(f"JD-{v_idx:02d}-{tr}", wbs, f"Jackdown TR{tr}", "JACKDOWN", "Mammoet", offset, "DUR_JD", "1.0d/unit")
            if first_jd_offset is None:
                first_jd_offset = offset
            offset += int(math.ceil(DEFAULT_DURATION_VALUES["DUR_JD"]))

        return_offset = offset
        buffer_offset = offset + 1
        return_note = "After final JD"

        if early_return and n > 1 and first_jd_offset is not None:
            return_offset = first_jd_offset + 1
            buffer_offset = return_offset + 1
            return_note = "After first JD"

        add(f"RET-{v_idx:02d}", wbs, "LCT Return AGI->MZP", "RETURN", "LCT", return_offset, "DUR_RET", return_note)
        add(f"BUF-{v_idx:02d}", wbs, "Buffer / reset", "BUFFER", "All", buffer_offset, "DUR_BUF", "contingency")

        cycle_days = buffer_offset - voyage_start + 1
        cycle_offset = buffer_offset + 1
        if v_idx < len(groups) and cycle_spacing > cycle_days:
            cycle_offset += (cycle_spacing - cycle_days)

    return tasks


# ==========================
# Workbook builders
# ==========================

def define_name(wb: Workbook, name: str, ref: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


def create_control_panel(
    wb: Workbook,
    *,
    project_start: dt.date,
    target_end: dt.date,
    shamal_start: dt.date,
    shamal_end: dt.date,
    tide_threshold: float,
    gantt_min_days: int,
    gantt_buffer_days: int,
    lct_maint_start: dt.date,
    lct_maint_end: dt.date,
    calendar_mode: str,
    weekend_pattern: str,
) -> None:
    ws = wb.active
    ws.title = "Control_Panel"

    ws.merge_cells("A1:H1")
    ws["A1"] = "AGI TR Transportation - Control Panel"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = "Changing PROJECT_START will update schedules (Gantt uses conditional formatting by default)."
    ws["A2"].fill = PatternFill("solid", fgColor="FFF9C4")

    ws["A4"] = "Project Start Date:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = project_start
    ws["B4"].number_format = "YYYY-MM-DD"
    ws["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["B4"].border = thin_border()

    ws["A5"] = "Target Completion Date:"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = target_end
    ws["B5"].number_format = "YYYY-MM-DD"
    ws["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["B5"].border = thin_border()

    ws["A6"] = "Voyage Pattern:"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = "1-2-2-2"
    ws["B6"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["B6"].border = thin_border()

    patterns = ["1x1x1x1x1x1x1", "1-2-2-2", "2-2-2-1"]
    dv = DataValidation(type="list", formula1=f'"{",".join(patterns)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B6")

    ws["A7"] = "Early Return (1st JD):"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = "TRUE"
    ws["B7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["B7"].border = thin_border()

    define_name(wb, "PROJECT_START", "Control_Panel!$B$4")
    define_name(wb, "TARGET_END", "Control_Panel!$B$5")
    define_name(wb, "VOYAGE_PATTERN", "Control_Panel!$B$6")
    define_name(wb, "EARLY_RETURN", "Control_Panel!$B$7")

    ws["D4"] = "Task Duration (days)"
    ws["D4"].font = Font(bold=True, size=12)

    durations = [
        ("D5", "Mobilization:", "E5", DEFAULT_DURATION_VALUES["DUR_MOB"], "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", DEFAULT_DURATION_VALUES["DUR_DECK"], "DUR_DECK"),
        ("D7", "Load-out:", "E7", DEFAULT_DURATION_VALUES["DUR_LO"], "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", DEFAULT_DURATION_VALUES["DUR_SF"], "DUR_SF"),
        ("D9", "MWS Approval:", "E9", DEFAULT_DURATION_VALUES["DUR_MWS"], "DUR_MWS"),
        ("D10", "Sailing:", "E10", DEFAULT_DURATION_VALUES["DUR_SAIL"], "DUR_SAIL"),
        ("D11", "AGI Unload:", "E11", DEFAULT_DURATION_VALUES["DUR_UL"], "DUR_UL"),
        ("D12", "Turning:", "E12", DEFAULT_DURATION_VALUES["DUR_TURN"], "DUR_TURN"),
        ("D13", "Jack-down:", "E13", DEFAULT_DURATION_VALUES["DUR_JD"], "DUR_JD"),
        ("D14", "Return:", "E14", DEFAULT_DURATION_VALUES["DUR_RET"], "DUR_RET"),
        ("D15", "Buffer:", "E15", DEFAULT_DURATION_VALUES["DUR_BUF"], "DUR_BUF"),
    ]
    for lc, lt, vc, v, name in durations:
        ws[lc] = lt
        ws[lc].font = Font(bold=True)
        ws[vc] = v
        ws[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws[vc].border = thin_border()
        ws[vc].number_format = "0.0"
        define_name(wb, name, f"Control_Panel!${vc}")

    ws["G4"] = "Weather & Constraints"
    ws["G4"].font = Font(bold=True, size=12)

    ws["G5"] = "Shamal Start:"
    ws["H5"] = shamal_start
    ws["H5"].number_format = "YYYY-MM-DD"
    ws["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws["H5"].border = thin_border()

    ws["G6"] = "Shamal End:"
    ws["H6"] = shamal_end
    ws["H6"].number_format = "YYYY-MM-DD"
    ws["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws["H6"].border = thin_border()

    ws["G7"] = "Tide Threshold (m):"
    ws["G7"].font = Font(bold=True)
    ws["H7"] = float(tide_threshold)
    ws["H7"].number_format = "0.00"
    ws["H7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H7"].border = thin_border()

    ws["G8"] = "Gantt Min Days:"
    ws["G8"].font = Font(bold=True)
    ws["H8"] = int(gantt_min_days)
    ws["H8"].number_format = "0"
    ws["H8"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H8"].border = thin_border()

    ws["G9"] = "Gantt Buffer Days:"
    ws["G9"].font = Font(bold=True)
    ws["H9"] = int(gantt_buffer_days)
    ws["H9"].number_format = "0"
    ws["H9"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H9"].border = thin_border()

    ws["G10"] = "LCT Maint. Start:"
    ws["G10"].font = Font(bold=True)
    ws["H10"] = lct_maint_start
    ws["H10"].number_format = "YYYY-MM-DD"
    ws["H10"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H10"].border = thin_border()

    ws["G11"] = "LCT Maint. End:"
    ws["G11"].font = Font(bold=True)
    ws["H11"] = lct_maint_end
    ws["H11"].number_format = "YYYY-MM-DD"
    ws["H11"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H11"].border = thin_border()

    ws["G12"] = "Calendar Mode:"
    ws["G12"].font = Font(bold=True)
    ws["H12"] = calendar_mode.upper()
    ws["H12"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H12"].border = thin_border()
    dv_mode = DataValidation(type="list", formula1='"CALENDAR,WORKDAY"', allow_blank=False)
    ws.add_data_validation(dv_mode)
    dv_mode.add("H12")

    ws["G13"] = "Weekend Pattern:"
    ws["G13"].font = Font(bold=True)
    ws["H13"] = weekend_pattern
    ws["H13"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws["H13"].border = thin_border()
    dv_week = DataValidation(type="list", formula1=f'"{",".join(WEEKEND_PATTERNS.values())}"', allow_blank=False)
    ws.add_data_validation(dv_week)
    dv_week.add("H13")

    define_name(wb, "SHAMAL_START", "Control_Panel!$H$5")
    define_name(wb, "SHAMAL_END", "Control_Panel!$H$6")
    define_name(wb, "TIDE_THRESHOLD", "Control_Panel!$H$7")
    define_name(wb, "GANTT_MIN_DAYS", "Control_Panel!$H$8")
    define_name(wb, "GANTT_BUFFER_DAYS", "Control_Panel!$H$9")
    define_name(wb, "LCT_MAINT_START", "Control_Panel!$H$10")
    define_name(wb, "LCT_MAINT_END", "Control_Panel!$H$11")
    define_name(wb, "CALENDAR_MODE", "Control_Panel!$H$12")
    define_name(wb, "WEEKEND_PATTERN", "Control_Panel!$H$13")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 16


def create_calendar_sheet(wb: Workbook, holidays: list[dt.date]) -> None:
    ws = wb.create_sheet("Calendar")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Calendar / Holidays (used by WORKDAY.INTL when CALENDAR_MODE=WORKDAY)"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Holiday Date"
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A3"].border = thin_border()
    ws.column_dimensions["A"].width = 16

    r = 4
    for d in sorted(set(holidays)):
        ws.cell(r, 1, value=d).number_format = "YYYY-MM-DD"
        ws.cell(r, 1).border = thin_border()
        r += 1

    if r > 4:
        define_name(wb, "HOLIDAYS", f"Calendar!$A$4:$A${r-1}")
    else:
        define_name(wb, "HOLIDAYS", "Calendar!$A$4:$A$4")


def resolve_duration_value(dur: DurRef, duration_values: dict[str, float]) -> float:
    if isinstance(dur, (int, float)):
        return float(dur)
    return float(duration_values.get(str(dur).strip(), 1.0))


def calculate_max_days_from_tasks(tasks: list[Task], *, wb: Workbook, min_days_default: int, buffer_days_default: int) -> int:
    min_days = min_days_default
    buffer_days = buffer_days_default
    try:
        ws_ctrl = wb["Control_Panel"]
        md = ws_ctrl["H8"].value
        bd = ws_ctrl["H9"].value
        if isinstance(md, (int, float)):
            min_days = int(md)
        if isinstance(bd, (int, float)):
            buffer_days = int(bd)
    except Exception:
        pass

    max_end_offset = 0
    for t in tasks:
        est_dur = resolve_duration_value(t.duration, DEFAULT_DURATION_VALUES)
        end_off = t.offset_days + int(math.ceil(max(est_dur, 0)))
        max_end_offset = max(max_end_offset, end_off)

    return max(min_days, max_end_offset + buffer_days)


def export_schedules_json(
    scenario_tasks: dict[str, list[Task]],
    project_start: dt.date,
    output_path: Path,
    duration_values: dict[str, float],
) -> None:
    """Export all scenario schedules to JSON format."""
    export_data = {
        "metadata": {
            "project_start": project_start.isoformat(),
            "export_date": datetime.now().isoformat(),
            "scenarios": list(scenario_tasks.keys()),
        },
        "scenarios": {},
    }

    for scenario_name, tasks in scenario_tasks.items():
        scenario_data = {
            "name": scenario_name,
            "task_count": len(tasks),
            "tasks": [],
        }

        for task in tasks:
            duration_days = resolve_duration_value(task.duration, duration_values)
            start_date = project_start + dt.timedelta(days=task.offset_days)
            end_date = start_date + dt.timedelta(days=int(math.ceil(duration_days)))

            task_data = {
                "task_id": task.task_id,
                "wbs": task.wbs,
                "name": task.name,
                "phase": task.phase,
                "owner": task.owner,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "duration_days": duration_days,
                "offset_days": task.offset_days,
                "notes": task.notes,
            }

            if task.planned_start:
                task_data["planned_start"] = task.planned_start.isoformat()
            if task.planned_finish:
                task_data["planned_finish"] = task.planned_finish.isoformat()
            if task.actual_start:
                task_data["actual_start"] = task.actual_start.isoformat()
            if task.actual_finish:
                task_data["actual_finish"] = task.actual_finish.isoformat()
            if task.percent_complete is not None:
                task_data["percent_complete"] = task.percent_complete

            scenario_data["tasks"].append(task_data)

        export_data["scenarios"][scenario_name] = scenario_data

    json_path = output_path.with_suffix(".schedules.json")
    json_path.write_text(json.dumps(export_data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[OK] Schedules exported to JSON: {json_path}")


def export_schedules_csv(
    scenario_tasks: dict[str, list[Task]],
    project_start: dt.date,
    output_path: Path,
    duration_values: dict[str, float],
) -> None:
    """Export all scenario schedules to CSV format."""
    csv_path = output_path.with_suffix(".schedules.csv")

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Scenario", "Task_ID", "WBS", "Task_Name", "Phase", "Owner",
            "Start_Date", "End_Date", "Duration_Days", "Offset_Days", "Notes",
            "Planned_Start", "Planned_Finish", "Actual_Start", "Actual_Finish", "Percent_Complete",
        ])

        for scenario_name, tasks in scenario_tasks.items():
            for task in tasks:
                duration_days = resolve_duration_value(task.duration, duration_values)
                start_date = project_start + dt.timedelta(days=task.offset_days)
                end_date = start_date + dt.timedelta(days=int(math.ceil(duration_days)))

                row = [
                    scenario_name,
                    task.task_id,
                    task.wbs,
                    task.name,
                    task.phase,
                    task.owner,
                    start_date.isoformat(),
                    end_date.isoformat(),
                    f"{duration_days:.2f}",
                    task.offset_days,
                    task.notes,
                    task.planned_start.isoformat() if task.planned_start else "",
                    task.planned_finish.isoformat() if task.planned_finish else "",
                    task.actual_start.isoformat() if task.actual_start else "",
                    task.actual_finish.isoformat() if task.actual_finish else "",
                    f"{task.percent_complete:.1f}" if task.percent_complete is not None else "",
                ]
                writer.writerow(row)

    print(f"[OK] Schedules exported to CSV: {csv_path}")


def export_report_json(
    scenario_tasks: dict[str, list[Task]],
    scenario_rows: list[tuple[str, str, str]],
    project_start: dt.date,
    target_end: Optional[dt.date],
    output_path: Path,
    duration_values: dict[str, float],
    wb: Optional[Workbook] = None,
) -> None:
    """Export comprehensive project report to JSON format."""
    report = {
        "metadata": {
            "project_start": project_start.isoformat(),
            "target_end": target_end.isoformat() if target_end else None,
            "export_date": datetime.now().isoformat(),
            "scenario_count": len(scenario_tasks),
        },
        "scenarios": {},
        "summary": {},
    }

    all_end_dates: list[dt.date] = []

    for scenario_name, tasks in scenario_tasks.items():
        if not tasks:
            continue

        scenario_end_dates = []
        total_duration = 0.0

        for task in tasks:
            duration_days = resolve_duration_value(task.duration, duration_values)
            start_date = project_start + dt.timedelta(days=task.offset_days)
            end_date = start_date + dt.timedelta(days=int(math.ceil(duration_days)))
            scenario_end_dates.append(end_date)
            total_duration += duration_days

        latest_end = max(scenario_end_dates) if scenario_end_dates else None
        if latest_end:
            all_end_dates.append(latest_end)

        report["scenarios"][scenario_name] = {
            "name": scenario_name,
            "task_count": len(tasks),
            "estimated_completion": latest_end.isoformat() if latest_end else None,
            "total_duration_days": total_duration,
            "status": "On Target" if (latest_end and target_end and latest_end <= target_end) else "Delayed" if latest_end and target_end else "Unknown",
        }

    if all_end_dates:
        report["summary"] = {
            "latest_completion": max(all_end_dates).isoformat(),
            "earliest_completion": min(all_end_dates).isoformat(),
            "on_target": all(target_end and d <= target_end for d in all_end_dates) if target_end else None,
        }

    if wb:
        try:
            wb["Scenario_KPIs"]
        except Exception:
            pass

    report_path = output_path.with_suffix(".report.json")
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[OK] Report exported to JSON: {report_path}")


def start_formula(offset: int) -> str:
    return (
        f'=IF(CALENDAR_MODE="WORKDAY",'
        f'WORKDAY.INTL(PROJECT_START-1,{offset + 1},WEEKEND_PATTERN,HOLIDAYS),'
        f'PROJECT_START+{offset})'
    )


def end_formula(start_cell: str, duration_cell: str) -> str:
    return (
        f'=IF(CALENDAR_MODE="WORKDAY",'
        f'WORKDAY.INTL({start_cell}-1,CEILING({duration_cell},1),WEEKEND_PATTERN,HOLIDAYS),'
        f'{start_cell}+{duration_cell})'
    )


def create_scenario_schedule_sheet(wb: Workbook, scenario: Scenario, tasks: list[Task], *, tide_sheet_exists: bool, extended: bool) -> tuple[str, int]:
    _short, sched_name, _gantt = scenario_sheet_names(scenario.name)
    ws = wb.create_sheet(sched_name)

    ws.merge_cells("A1:R1" if extended else "A1:I1")
    ws["A1"] = f"AGI TR Transportation - {scenario.name} Schedule"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:R2" if extended else "A2:I2")
    ws["A2"] = "Start/End auto-updates from Control_Panel (PROJECT_START)."
    ws["A2"].font = Font(size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    ws.merge_cells("A3:R3" if extended else "A3:I3")
    ws["A3"] = "Risk: Shamal period & Tide threshold are highlighted via helper columns / conditional formats."
    ws["A3"].font = Font(size=10, italic=True)
    ws["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    base_headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    extra_headers = ["Actual Start", "Actual Finish", "% Complete", "Start Var (d)", "Finish Var (d)", "Shamal Overlap", "Tide Height", "Tide OK", "Maint Overlap", "Critical"]
    headers = base_headers + (extra_headers if extended else [])
    header_row = 5

    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    first_task_row = header_row + 1
    for idx, t in enumerate(tasks):
        r = first_task_row + idx
        ws.cell(r, 1, value=t.task_id)
        ws.cell(r, 2, value=t.wbs)
        ws.cell(r, 3, value=t.name)
        ws.cell(r, 4, value=t.phase)
        ws.cell(r, 5, value=t.owner)

        ws.cell(r, 6, value=start_formula(t.offset_days)).number_format = "YYYY-MM-DD"

        if isinstance(t.duration, str):
            ws.cell(r, 8, value=f"={t.duration}")
        else:
            ws.cell(r, 8, value=float(t.duration))
        ws.cell(r, 8).number_format = "0.0"

        ws.cell(r, 7, value=end_formula(f"F{r}", f"H{r}")).number_format = "YYYY-MM-DD"
        ws.cell(r, 9, value=t.notes)

        ws.cell(r, 4).fill = PatternFill("solid", fgColor=COLORS.get(t.phase, "FFFFFF"))

        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = thin_border()

        if t.phase == "MILESTONE":
            for c in range(1, min(len(headers), 9) + 1):
                ws.cell(r, c).font = Font(bold=True)

        if not extended:
            continue

        if t.actual_start:
            ws.cell(r, 10, value=t.actual_start).number_format = "YYYY-MM-DD"
        if t.actual_finish:
            ws.cell(r, 11, value=t.actual_finish).number_format = "YYYY-MM-DD"
        if t.percent_complete is not None:
            ws.cell(r, 12, value=float(t.percent_complete)).number_format = "0%"

        ws.cell(r, 13, value=f'=IF(OR(J{r}="",F{r}=""),"",J{r}-F{r})').number_format = "0"
        ws.cell(r, 14, value=f'=IF(OR(K{r}="",G{r}=""),"",K{r}-G{r})').number_format = "0"
        ws.cell(r, 15, value=f'=IF(AND(F{r}<=SHAMAL_END,G{r}>=SHAMAL_START),"Y","")')

        if tide_sheet_exists:
            ws.cell(r, 16, value=f'=IFERROR(VLOOKUP(INT(F{r}),Tide_Data!$A:$C,3,FALSE),"")').number_format = "0.00"
            ws.cell(r, 17, value=(
                f'=IF(OR($D{r}="LOADOUT",$D{r}="AGI_UNLOAD",ISNUMBER(SEARCH("TIDE",$I{r}))),'
                f'IF(P{r}="", "NO_DATA", IF(P{r}>=TIDE_THRESHOLD,"OK","NO")),"")'
            ))
        ws.cell(r, 18, value=f'=IF(AND(F{r}<=LCT_MAINT_END,G{r}>=LCT_MAINT_START),"Y","")')
        ws.cell(r, 19, value=f'=IF(OR($D{r}="MILESTONE",$D{r}="LOADOUT",$D{r}="SAIL",$D{r}="AGI_UNLOAD",$D{r}="JACKDOWN"),"Y","")')

    ws.freeze_panes = f"A{first_task_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{first_task_row + len(tasks) - 1}"

    for i in range(1, len(headers) + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = {1:10,2:8,3:44,4:14,5:16,6:12,7:12,8:10,9:40}.get(i, 12)

    if extended:
        last_row = first_task_row + len(tasks) - 1
        if last_row >= first_task_row:
            rng = f"A{first_task_row}:{get_column_letter(len(headers))}{last_row}"
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$O{first_task_row}="Y"'], fill=PatternFill("solid", fgColor=COLORS["SHAMAL"])))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$Q{first_task_row}="NO"'], fill=PatternFill("solid", fgColor=COLORS["WARN"])))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$Q{first_task_row}="NO_DATA"'], fill=PatternFill("solid", fgColor=COLORS["CAUTION"])))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{first_task_row}="Y"'], fill=PatternFill("solid", fgColor=COLORS["MAINT"])))

    return sched_name, len(tasks)


def create_scenario_gantt_sheet(wb: Workbook, scenario: Scenario, *, schedule_sheet_name: str, num_tasks: int, max_days: int) -> str:
    _short, _sched, gantt_name = scenario_sheet_names(scenario.name)
    ws = wb.create_sheet(gantt_name)

    total_cols = 7 + max_days
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"AGI TR Transportation - {scenario.name} Gantt Chart"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws["A2"] = "Gantt rendered by Conditional Formatting (auto-updates)."
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])

    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    date_col = 8
    for i in range(max_days):
        cell = ws.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        cell.number_format = "D"
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(date_col + i)].width = 2.5

    first_gantt_row = 5
    for r in range(first_gantt_row, first_gantt_row + num_tasks):
        sched_row = r + 1
        ws.cell(r, 1, value=f"='{schedule_sheet_name}'!A{sched_row}")
        ws.cell(r, 2, value=f"='{schedule_sheet_name}'!B{sched_row}")
        ws.cell(r, 3, value=f"='{schedule_sheet_name}'!C{sched_row}")
        ws.cell(r, 4, value=f"='{schedule_sheet_name}'!D{sched_row}")
        ws.cell(r, 5, value=f"='{schedule_sheet_name}'!F{sched_row}").number_format = "MM/DD"
        ws.cell(r, 6, value=f"='{schedule_sheet_name}'!G{sched_row}").number_format = "MM/DD"
        ws.cell(r, 7, value=f"='{schedule_sheet_name}'!H{sched_row}")

        for c in range(1, 8):
            ws.cell(r, c).border = thin_border()

    last_row = first_gantt_row + num_tasks - 1
    last_col = date_col + max_days - 1

    for r in range(first_gantt_row, last_row + 1):
        for c in range(date_col, last_col + 1):
            ws.cell(r, c).border = thin_border()

    header_range = f"{get_column_letter(date_col)}4:{get_column_letter(last_col)}4"
    base_col = get_column_letter(date_col)

    ws.conditional_formatting.add(header_range, FormulaRule(formula=[f"AND({base_col}$4>=LCT_MAINT_START,{base_col}$4<=LCT_MAINT_END)"], fill=PatternFill("solid", fgColor=COLORS["MAINT"])))
    ws.conditional_formatting.add(header_range, FormulaRule(formula=[f"AND({base_col}$4>=SHAMAL_START,{base_col}$4<=SHAMAL_END)"], fill=PatternFill("solid", fgColor=COLORS["SHAMAL"])))

    weekend_formula = (
        f'IF(WEEKEND_PATTERN="0000011",WEEKDAY({base_col}$4,2)>5,'
        f'IF(WEEKEND_PATTERN="0000110",OR(WEEKDAY({base_col}$4,2)=5,WEEKDAY({base_col}$4,2)=6),FALSE))'
    )
    ws.conditional_formatting.add(header_range, FormulaRule(formula=[weekend_formula], fill=PatternFill("solid", fgColor=COLORS["WEEKEND"])))
    ws.conditional_formatting.add(header_range, FormulaRule(formula=[f"{base_col}$4=TODAY()"], fill=PatternFill("solid", fgColor="FFF59D")))

    grid_range = f"{get_column_letter(date_col)}{first_gantt_row}:{get_column_letter(last_col)}{last_row}"
    for phase in PHASE_ORDER:
        color = COLORS.get(PHASE_COLOR_KEYS.get(phase, "BUFFER"), "FFFFFF")
        if phase == "MILESTONE":
            formula = f'AND($D{first_gantt_row}="{phase}",{base_col}$4=$E{first_gantt_row})'
        else:
            formula = f'AND($D{first_gantt_row}="{phase}",{base_col}$4>=$E{first_gantt_row},{base_col}$4<$F{first_gantt_row})'
        ws.conditional_formatting.add(grid_range, FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)))

    phase_col_range = f"D{first_gantt_row}:D{last_row}"
    for phase in PHASE_ORDER:
        color = COLORS.get(PHASE_COLOR_KEYS.get(phase, "BUFFER"), "FFFFFF")
        ws.conditional_formatting.add(phase_col_range, FormulaRule(formula=[f'$D{first_gantt_row}="{phase}"'], fill=PatternFill("solid", fgColor=color)))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 4

    ws.freeze_panes = ws.cell(first_gantt_row, date_col)

    return gantt_name


def create_tide_data_sheet(wb: Workbook, *, tide_tsv: Optional[Path] = None, tide_json: Optional[Path] = None) -> bool:
    ws = wb.create_sheet("Tide_Data")
    try:
        tide_threshold = float(wb["Control_Panel"]["H7"].value)
    except Exception:
        tide_threshold = 1.90

    ws.merge_cells("A1:D1")
    ws["A1"] = "MINA ZAYED PORT - High Tide Data"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Tide ≥{tide_threshold:.2f}m required for Load-out and AGI Arrival"
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    headers = ["Date", "High Tide Window", "Max Height (m)", "Risk Level"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    records: list[tuple[dt.date, str, float, str]] = []

    if tide_json and tide_json.exists():
        try:
            data = json.loads(tide_json.read_text(encoding="utf-8"))
            for rec in data.get("tide_records", []):
                d = parse_date(rec.get("date"))
                if not d:
                    continue
                records.append((d, str(rec.get("high_tide_window", "")).strip(), float(rec.get("max_height_m", 0) or 0.0), str(rec.get("risk_level", "LOW")).strip().upper()))
        except Exception as e:
            print(f"[WARN] Could not read tide JSON: {e}")

    elif tide_tsv and tide_tsv.exists():
        try:
            for row in iter_rows_from_delimited(tide_tsv, delimiter="\t"):
                d = parse_date(row.get("Date") or row.get("date"))
                if not d:
                    continue
                window = str(row.get("High Tide Window", "")).strip()
                try:
                    height = float(str(row.get("Max Height (m)", "0")).strip() or 0.0)
                except Exception:
                    height = 0.0
                risk = str(row.get("Risk Level", "LOW")).strip().upper()
                records.append((d, window, height, risk))
        except Exception as e:
            print(f"[WARN] Could not read tide TSV: {e}")

    r0 = 5
    for idx, (d, window, height, risk) in enumerate(records):
        r = r0 + idx
        ws.cell(r, 1, value=d).number_format = "YYYY-MM-DD"
        ws.cell(r, 2, value=window)
        ws.cell(r, 3, value=float(height)).number_format = "0.00"
        ws.cell(r, 4, value=risk)
        for c in range(1, 5):
            ws.cell(r, c).border = thin_border()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A5"

    return len(records) > 0


def create_scenario_kpis_sheet(wb: Workbook, scenario_rows: list[tuple[str, str, str]], *, extended_schedule: bool) -> None:
    ws = wb.create_sheet("Scenario_KPIs")

    ws.merge_cells("A1:I1")
    ws["A1"] = "Scenario KPIs / Comparison"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Scenario", "Tasks", "End Date", "Duration (days)", "Shamal Tasks", "Tide NO", "Maint Tasks", "Critical Tasks", "% Complete (avg)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for i, (scn_name, sched_name, _gantt_name) in enumerate(scenario_rows, start=4):
        ws.cell(i, 1, value=scn_name)
        ws.cell(i, 2, value=f"=COUNTA('{sched_name}'!A:A)-5")
        ws.cell(i, 3, value=f"=MAX('{sched_name}'!G:G)")
        ws.cell(i, 3).number_format = "YYYY-MM-DD"
        ws.cell(i, 4, value=f"=C{i}-PROJECT_START")
        ws.cell(i, 4).number_format = "0"
        if extended_schedule:
            ws.cell(i, 5, value=f"=COUNTIF('{sched_name}'!O:O,\"Y\")")
            ws.cell(i, 6, value=f"=COUNTIF('{sched_name}'!Q:Q,\"NO\")")
            ws.cell(i, 7, value=f"=COUNTIF('{sched_name}'!R:R,\"Y\")")
            ws.cell(i, 8, value=f"=COUNTIF('{sched_name}'!S:S,\"Y\")")
            ws.cell(i, 9, value=f"=IFERROR(AVERAGE('{sched_name}'!L:L),\"\")")
            ws.cell(i, 9).number_format = "0%"

        for c in range(1, 10):
            ws.cell(i, c).border = thin_border()

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    try:
        chart = BarChart()
        chart.title = "Scenario Duration (days)"
        chart.y_axis.title = "Days"
        chart.x_axis.title = "Scenario"
        data = Reference(ws, min_col=4, min_row=3, max_row=3 + len(scenario_rows))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(scenario_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "K3")
    except Exception:
        pass


def create_summary_sheet(wb: Workbook, primary_schedule_name: str) -> None:
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:C1")
    ws["A1"] = "AGI TR Transportation - Summary"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
        ("Estimated End (Primary)", f"=MAX('{primary_schedule_name}'!G:G)"),
        ("Duration (days)", "=B5-PROJECT_START"),
        ("Status vs Target", '=IF(B5<=TARGET_END,"ON TARGET","DELAYED")'),
        ("Shamal Period", "=SHAMAL_START&\" to \"&SHAMAL_END"),
        ("Tide Threshold (m)", "=TIDE_THRESHOLD"),
        ("LCT Maintenance", "=LCT_MAINT_START&\" to \"&LCT_MAINT_END"),
    ]
    r0 = 3
    for i, (k, v) in enumerate(rows):
        r = r0 + i
        ws.cell(r, 1, value=k).font = Font(bold=True)
        ws.cell(r, 2, value=v)
        ws.cell(r, 1).border = thin_border()
        ws.cell(r, 2).border = thin_border()
        if "End" in k or "Start" in k:
            ws.cell(r, 2).number_format = "YYYY-MM-DD"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 44


def generate_vba_code(scenario_rows: list[tuple[str, str, str]]) -> str:
    arr_lines = []
    for _scn, sched, gantt in scenario_rows:
        arr_lines.append(f'        Array("{sched}", "{gantt}")')
    arr_body = ", _\n".join(arr_lines) if arr_lines else ""

    vba = f"""
Option Explicit

' ============================================
' AGI TR Multi-Scenario Master Gantt - VBA (Generated)
' Notes:
' - Gantt is rendered by Conditional Formatting (no repaint needed).
' - Macros below are convenience utilities (recalc/export).
' ============================================

Private Function ScenarioPairs() As Variant
    ScenarioPairs = Array( _
{arr_body}
    )
End Function

Sub UpdateAll()
    Dim pairs As Variant, i As Long
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual

    On Error Resume Next
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    Sheets("Scenario_KPIs").Calculate
    Sheets("Weather_Analysis").Calculate
    Sheets("Tide_Data").Calculate
    On Error GoTo 0

    pairs = ScenarioPairs()
    For i = LBound(pairs) To UBound(pairs)
        On Error Resume Next
        Sheets(CStr(pairs(i)(0))).Calculate
        Sheets(CStr(pairs(i)(1))).Calculate
        On Error GoTo 0
    Next i

    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True

    MsgBox "Update complete", vbInformation
End Sub

Sub ExportToPDF()
    Dim outPath As String
    outPath = ThisWorkbook.Path & "\\" & Replace(ThisWorkbook.Name, ".xlsm", "") & "_export.pdf"
    On Error GoTo ErrHandler
    ActiveWorkbook.ExportAsFixedFormat Type:=xlTypePDF, Filename:=outPath, Quality:=xlQualityStandard, _
        IncludeDocProperties:=True, IgnorePrintAreas:=False, OpenAfterPublish:=True
    MsgBox "PDF exported: " & outPath, vbInformation
    Exit Sub

ErrHandler:
    MsgBox "PDF export failed: " & Err.Description, vbCritical
End Sub

Sub ShowControlPanelSettings()
    Dim ws As Worksheet
    Set ws = Sheets("Control_Panel")
    MsgBox _
        "PROJECT_START: " & ws.Range("B4").Value & vbCrLf & _
        "TARGET_END: " & ws.Range("B5").Value & vbCrLf & _
        "SHAMAL: " & ws.Range("H5").Value & " ~ " & ws.Range("H6").Value & vbCrLf & _
        "TIDE_THRESHOLD: " & ws.Range("H7").Value & vbCrLf & _
        "CALENDAR_MODE: " & ws.Range("H12").Value & vbCrLf & _
        "WEEKEND_PATTERN: " & ws.Range("H13").Value, _
        vbInformation, "Control Panel Settings"
End Sub
"""
    return vba.strip() + "\n"


def create_vba_code_sheet(wb: Workbook, vba_code: str) -> None:
    ws = wb.create_sheet("VBA_Code")
    ws["A1"] = "VBA 코드 - Alt+F11 → Module 삽입 → 붙여넣기 (또는 --export-vba-bas 사용)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "※ 조건부 서식 기반 Gantt이므로, VBA는 Update/PDF Export 등 보조 기능입니다."
    ws["A2"].font = Font(size=10)
    for i, line in enumerate(vba_code.splitlines(), start=4):
        ws.cell(i, 1, value=line)
    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A4"


def create_weather_analysis_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Weather_Analysis")
    ws.merge_cells("A1:H1")
    ws["A1"] = "UAE Winter Weather Analysis (Simplified)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Parameter", "Value", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = thin_border()

    rows = [
        ("Shamal period", "=SHAMAL_START&\" ~ \"&SHAMAL_END", "Higher wind / wave risk"),
        ("Wind limit (sustained)", "18 kt", "Hold above threshold"),
        ("Wind limit (gust)", "22 kt", "No-go above threshold"),
        ("Wave limit", "0.8 m", "Hold above threshold"),
        ("Visibility", "2-5 km (Shamal)", "Dust / reduced visibility"),
    ]
    for i, (p, v, n) in enumerate(rows, start=4):
        ws.cell(i, 1, value=p).border = thin_border()
        ws.cell(i, 2, value=v).border = thin_border()
        ws.cell(i, 3, value=n).border = thin_border()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 46


# ==========================
# Config / CLI
# ==========================

def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(path)
    ext = path.suffix.lower()
    if ext == ".json":
        return json.loads(path.read_text(encoding="utf-8"))
    if ext == ".toml":
        import tomllib
        return tomllib.loads(path.read_text(encoding="utf-8"))
    if ext in (".yml", ".yaml"):
        try:
            import yaml  # type: ignore
        except Exception as e:
            raise RuntimeError("YAML config requires PyYAML. Install: pip install pyyaml") from e
        return yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    raise ValueError(f"Unsupported config extension: {ext}")


def parse_scenario_arg(s: str) -> Scenario:
    raw = s.strip()
    if not raw:
        raise argparse.ArgumentTypeError("Empty scenario spec")
    parts = [p.strip() for p in raw.split(";") if p.strip()]
    main = parts[0]
    if "=" not in main:
        raise argparse.ArgumentTypeError("Scenario must be NAME=INPUT")
    name, input_spec = [x.strip() for x in main.split("=", 1)]
    kwargs: dict[str, Any] = {"name": name, "input_spec": input_spec}
    for p in parts[1:]:
        if "=" not in p:
            continue
        k, v = [x.strip() for x in p.split("=", 1)]
        k = k.lower()
        if k in ("early_return", "earlyreturn"):
            kwargs["early_return"] = v.lower() in ("1", "true", "yes", "y")
        elif k in ("cycle_spacing", "cycle", "spacing"):
            try:
                kwargs["cycle_spacing"] = int(v)
            except Exception:
                pass
        elif k in ("include_group_rows", "include_groups"):
            kwargs["include_group_rows"] = v.lower() in ("1", "true", "yes", "y")
        elif k == "pattern":
            kwargs["pattern"] = v
    if input_spec.lower().startswith("pattern:"):
        kwargs["pattern"] = input_spec.split(":", 1)[1].strip()
        kwargs["input_spec"] = None
    return Scenario(**kwargs)


def first_existing(dir_path: Path, *names: str) -> Optional[Path]:
    for n in names:
        p = dir_path / n
        if p.exists():
            return p
    return None


def build_scenarios_from_legacy_args(args: argparse.Namespace) -> list[Scenario]:
    scenarios: list[Scenario] = []
    if args.original:
        scenarios.append(Scenario(name="Mammoet_Original", input_spec=args.original))
    if args.scenario_a:
        scenarios.append(Scenario(name="Mammoet_ScenarioA", input_spec=args.scenario_a))
    if args.alternative:
        scenarios.append(Scenario(name="Mammoet_Alternative", input_spec=args.alternative))
    if args.tsv_path:
        scenarios.append(Scenario(name="Generic", input_spec=args.tsv_path))
    if args.pattern and not any(s.input_spec for s in scenarios):
        scenarios.append(Scenario(name="Pattern_Generated", pattern=args.pattern, early_return=args.early_return, cycle_spacing=args.cycle_spacing))
    return scenarios


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    defaults = {
        "orig": first_existing(script_dir, "MOS.tsv", "Mammoet format_Original.tsv", "Mammoet_Original.tsv"),
        "a": first_existing(script_dir, "MFC1.tsv", "Mammoet format_ScenarioA.tsv"),
        "alt": first_existing(script_dir, "MFA2.tsv", "Mammoet format_Alternative.tsv"),
        "tide_tsv": first_existing(script_dir, "Tide.tsv"),
        "tide_json": first_existing(script_dir, "Tide.json", "MINA ZAYED PORT WATER TIDE_MERGED.json"),
    }

    p = argparse.ArgumentParser(description="Generate AGI TR Multi-Scenario Master Gantt workbook (max options edition).")
    p.add_argument("--config", type=str, default=None, help="Config file (.json/.toml/.yml). CLI overrides config.")
    p.add_argument("--original", default=str(defaults["orig"]) if defaults["orig"] else None)
    p.add_argument("--scenario-a", dest="scenario_a", default=str(defaults["a"]) if defaults["a"] else None)
    p.add_argument("--alternative", default=str(defaults["alt"]) if defaults["alt"] else None)
    p.add_argument("--tsv", dest="tsv_path", default=None)
    p.add_argument("--scenario", action="append", default=[], type=parse_scenario_arg)

    p.add_argument("--delimiter", default=None)
    p.add_argument("--encoding", default=None)
    p.add_argument("--sheet", default=None)
    p.add_argument("--include-group-rows", action="store_true")

    p.add_argument("--pattern", default=None)
    p.add_argument("--early-return", action="store_true")
    p.add_argument("--cycle-spacing", type=int, default=15)

    p.add_argument("--project-start", type=parse_iso_date, default=None)
    p.add_argument("--target-end", type=parse_iso_date, default=None)
    p.add_argument("--shamal-start", type=parse_iso_date, default=None)
    p.add_argument("--shamal-end", type=parse_iso_date, default=None)
    p.add_argument("--tide-threshold", type=float, default=None)
    p.add_argument("--gantt-min-days", type=int, default=120)
    p.add_argument("--gantt-buffer-days", type=int, default=30)
    p.add_argument("--lct-maint-start", type=parse_iso_date, default=None)
    p.add_argument("--lct-maint-end", type=parse_iso_date, default=None)
    p.add_argument("--calendar-mode", choices=["CALENDAR", "WORKDAY"], default="CALENDAR")
    p.add_argument("--weekend-pattern", default=None)
    p.add_argument("--holiday", action="append", default=[])
    p.add_argument("--holiday-file", default=None)

    p.add_argument("--tide-tsv", dest="tide_tsv", default=str(defaults["tide_tsv"]) if defaults["tide_tsv"] else None)
    p.add_argument("--tide-json", dest="tide_json", default=str(defaults["tide_json"]) if defaults["tide_json"] else None)

    p.add_argument("--simple-schedule", action="store_true")
    p.add_argument("--no-weather", action="store_true")
    p.add_argument("--no-vba", action="store_true")
    p.add_argument("--no-kpis", action="store_true")

    p.add_argument("--output-dir", default=str(script_dir))
    p.add_argument("--output", default=None)
    p.add_argument("--export-vba-bas", action="store_true")
    p.add_argument("--create-xlsm", action="store_true")
    p.add_argument("--export-pdf", action="store_true")
    p.add_argument("--export-schedules-json", action="store_true")
    p.add_argument("--export-schedules-csv", action="store_true")
    p.add_argument("--report-json", default=None)

    args = p.parse_args()

    config: dict[str, Any] = {}
    if args.config:
        try:
            config = load_config(Path(args.config).expanduser().resolve())
        except Exception as e:
            print(f"[WARN] Config load failed: {e}")

    project_start = args.project_start or parse_iso_date(str(config.get("project", {}).get("start", "2026-01-18")))
    target_end = args.target_end or parse_iso_date(str(config.get("project", {}).get("target_end", "2026-02-28")))
    shamal_start = args.shamal_start or parse_iso_date(str(config.get("project", {}).get("shamal_start", "2026-01-15")))
    shamal_end = args.shamal_end or parse_iso_date(str(config.get("project", {}).get("shamal_end", "2026-04-30")))
    tide_threshold = args.tide_threshold if args.tide_threshold is not None else float(config.get("project", {}).get("tide_threshold", 1.90))
    lct_maint_start = args.lct_maint_start or parse_iso_date(str(config.get("project", {}).get("lct_maint_start", "2026-02-10")))
    lct_maint_end = args.lct_maint_end or parse_iso_date(str(config.get("project", {}).get("lct_maint_end", "2026-02-14")))
    calendar_mode = (str(config.get("project", {}).get("calendar_mode", args.calendar_mode)) or args.calendar_mode).upper()
    weekend_pattern = args.weekend_pattern or str(config.get("project", {}).get("weekend_pattern", WEEKEND_PATTERNS["FRI_SAT"]))
    gantt_min_days = int(config.get("project", {}).get("gantt_min_days", args.gantt_min_days))
    gantt_buffer_days = int(config.get("project", {}).get("gantt_buffer_days", args.gantt_buffer_days))

    holidays: list[dt.date] = []
    for h in config.get("project", {}).get("holidays", []) or []:
        try:
            holidays.append(parse_iso_date(str(h)))
        except Exception:
            pass
    for h in args.holiday:
        try:
            holidays.append(parse_iso_date(h))
        except Exception:
            pass
    if args.holiday_file:
        try:
            for line in Path(args.holiday_file).read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line:
                    try:
                        holidays.append(parse_iso_date(line))
                    except Exception:
                        pass
        except Exception:
            pass

    scenarios: list[Scenario] = []
    for sc in (config.get("scenarios", []) or []):
        try:
            scenarios.append(
                Scenario(
                    name=str(sc.get("name")),
                    input_spec=str(sc.get("input")) if sc.get("input") else None,
                    pattern=str(sc.get("pattern")) if sc.get("pattern") else None,
                    early_return=bool(sc.get("early_return", False)),
                    cycle_spacing=int(sc.get("cycle_spacing", 15)),
                    include_group_rows=bool(sc.get("include_group_rows", False)),
                )
            )
        except Exception:
            continue

    scenarios.extend(args.scenario or [])
    scenarios.extend(build_scenarios_from_legacy_args(args))

    seen: set[str] = set()
    uniq: list[Scenario] = []
    for sc in scenarios:
        if sc.name in seen:
            continue
        seen.add(sc.name)
        uniq.append(sc)
    scenarios = uniq

    if not scenarios:
        scenarios = [Scenario(name="Pattern_Generated", pattern=args.pattern or "1-2-2-2", early_return=args.early_return, cycle_spacing=args.cycle_spacing)]

    wb = Workbook()
    create_control_panel(
        wb,
        project_start=project_start,
        target_end=target_end,
        shamal_start=shamal_start,
        shamal_end=shamal_end,
        tide_threshold=tide_threshold,
        gantt_min_days=gantt_min_days,
        gantt_buffer_days=gantt_buffer_days,
        lct_maint_start=lct_maint_start,
        lct_maint_end=lct_maint_end,
        calendar_mode=calendar_mode,
        weekend_pattern=weekend_pattern,
    )
    create_calendar_sheet(wb, holidays=holidays)

    tide_sheet_exists = False
    tide_tsv = Path(args.tide_tsv).expanduser().resolve() if args.tide_tsv else None
    tide_json = Path(args.tide_json).expanduser().resolve() if args.tide_json else None
    if tide_json and tide_json.exists():
        tide_sheet_exists = create_tide_data_sheet(wb, tide_json=tide_json)
    elif tide_tsv and tide_tsv.exists():
        tide_sheet_exists = create_tide_data_sheet(wb, tide_tsv=tide_tsv)

    scenario_rows: list[tuple[str, str, str]] = []
    scenario_tasks: dict[str, list[Task]] = {}
    extended_schedule = not args.simple_schedule

    for sc in scenarios:
        tasks: list[Task] = []
        if sc.input_spec:
            path = Path(sc.input_spec).expanduser().resolve()
            if path.exists():
                tasks = load_tasks(path, project_start, delimiter=args.delimiter, encoding=args.encoding, sheet_name=args.sheet, include_group_rows=args.include_group_rows or sc.include_group_rows)
                print(f"[OK] Loaded {len(tasks)} tasks for {sc.name} from {path.name}")
            else:
                print(f"[WARN] Scenario input not found: {path}")

        if not tasks and (sc.pattern or args.pattern):
            pattern = sc.pattern or args.pattern or "1-2-2-2"
            tasks = generate_scenario_tasks(pattern, project_start, cycle_spacing=sc.cycle_spacing or args.cycle_spacing, early_return=sc.early_return or args.early_return)
            print(f"[OK] Generated {len(tasks)} tasks for {sc.name} from pattern: {pattern}")

        if not tasks:
            print(f"[WARN] No tasks for scenario: {sc.name} (skipped)")
            continue

        scenario_tasks[sc.name] = tasks
        sched_name, num_tasks = create_scenario_schedule_sheet(wb, sc, tasks, tide_sheet_exists=tide_sheet_exists, extended=extended_schedule)
        max_days = calculate_max_days_from_tasks(tasks, wb=wb, min_days_default=gantt_min_days, buffer_days_default=gantt_buffer_days)
        gantt_name = create_scenario_gantt_sheet(wb, sc, schedule_sheet_name=sched_name, num_tasks=num_tasks, max_days=max_days)
        scenario_rows.append((sc.name, sched_name, gantt_name))

    if not scenario_rows:
        print("[ERROR] No scenario sheets generated. Check inputs.")
        return 2

    primary_schedule = scenario_rows[0][1]
    create_summary_sheet(wb, primary_schedule)

    if not args.no_kpis:
        create_scenario_kpis_sheet(wb, scenario_rows, extended_schedule=extended_schedule)

    if not args.no_weather:
        create_weather_analysis_sheet(wb)

    if not args.no_vba:
        vba_code = generate_vba_code(scenario_rows)
        create_vba_code_sheet(wb, vba_code)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        out_dir = Path(args.output_dir).expanduser().resolve()
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"AGI_TR_MultiScenario_Master_Gantt_MAX_{timestamp}.xlsx"

    wb.save(str(output_path))
    print(f"[OK] Workbook generated: {output_path}")

    duration_values = DEFAULT_DURATION_VALUES.copy()
    try:
        ws_ctrl = wb["Control_Panel"]
        duration_cells = {
            "DUR_MOB": "E5", "DUR_DECK": "E6", "DUR_LO": "E7", "DUR_SF": "E8",
            "DUR_MWS": "E9", "DUR_SAIL": "E10", "DUR_UL": "E11", "DUR_TURN": "E12",
            "DUR_JD": "E13", "DUR_RET": "E14", "DUR_BUF": "E15",
        }
        for k, addr in duration_cells.items():
            v = ws_ctrl[addr].value
            if isinstance(v, (int, float)):
                duration_values[k] = float(v)
    except Exception:
        pass

    if args.export_schedules_json:
        export_schedules_json(scenario_tasks, project_start, output_path, duration_values)

    if args.export_schedules_csv:
        export_schedules_csv(scenario_tasks, project_start, output_path, duration_values)

    if args.report_json:
        report_path = Path(args.report_json).expanduser().resolve() if args.report_json != "auto" else output_path
        export_report_json(scenario_tasks, scenario_rows, project_start, target_end, report_path, duration_values, wb=wb)

    bas_path: Optional[Path] = None
    if args.export_vba_bas and not args.no_vba:
        try:
            ws_vba = wb["VBA_Code"]
            lines = [str(r[0]) for r in ws_vba.iter_rows(min_row=4, max_col=1, values_only=True) if r and r[0] is not None]
            bas_path = output_path.with_suffix(".bas")
            bas_path.write_text("\n".join(lines), encoding="utf-8")
            print(f"[OK] VBA module exported: {bas_path}")
        except Exception as e:
            print(f"[WARN] VBA .bas export failed: {e}")

    # Windows automation
    if args.create_xlsm or args.export_pdf:
        if sys.platform != "win32":
            print("[WARN] --create-xlsm / --export-pdf are only supported on Windows with Excel installed.")
        else:
            try:
                import win32com.client  # type: ignore
            except Exception:
                print("[WARN] pywin32(win32com) is not installed. Install: pip install pywin32")
            else:
                excel = None
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb_xl = excel.Workbooks.Open(str(output_path))

                    if args.create_xlsm:
                        if bas_path is None and not args.no_vba:
                            bas_path = output_path.with_suffix(".bas")
                            bas_path.write_text(generate_vba_code(scenario_rows), encoding="utf-8")
                        if bas_path and bas_path.exists():
                            wb_xl.VBProject.VBComponents.Import(str(bas_path))
                        xlsm_path = output_path.with_suffix(".xlsm")
                        wb_xl.SaveAs(str(xlsm_path), FileFormat=52)
                        print(f"[OK] XLSM generated: {xlsm_path}")

                    if args.export_pdf:
                        pdf_path = output_path.with_suffix(".pdf")
                        wb_xl.ExportAsFixedFormat(0, str(pdf_path))
                        print(f"[OK] PDF exported: {pdf_path}")

                    wb_xl.Close(SaveChanges=True)
                except Exception as e:
                    print(f"[WARN] Excel automation failed: {e}")
                finally:
                    try:
                        if excel is not None:
                            excel.Quit()
                    except Exception:
                        pass

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
