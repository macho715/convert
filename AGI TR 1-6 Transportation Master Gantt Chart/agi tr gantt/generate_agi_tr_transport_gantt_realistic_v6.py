#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AGI Site – HVDC Transformers (TR1–TR6) Transport Gantt Generator (Excel)

Workbook outputs
- Schedule: task register + month/week/day timeline + conditional-format Gantt grid
- Gantt_Chart: stacked bar Gantt chart (start offset hidden)
- Tide_Peaks_MZP: tide-peak list (planning placeholder for RoRo windows)
- Assumptions_Refs: controls (for VBA/CF) + references

Notes
- Tide list is a planning placeholder. Replace with the latest official / port-approved tide table.

Dependencies
- Python 3.11+
- openpyxl 3.x
"""

from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import Rule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


# ----------------------------
# Data model
# ----------------------------

@dataclass(frozen=True)
class Task:
    id: str
    wbs: str
    task: str
    phase: str  # SUMMARY / DOC / PORT / MARINE / AGI / RETURN / MILESTONE
    location: str
    owner: str
    start: dt.date
    finish: dt.date
    progress_pct: Optional[float] = None
    baseline_start: Optional[dt.date] = None
    baseline_finish: Optional[dt.date] = None
    status: str = ""
    tide_gate: str = ""
    weather_gate: str = ""
    notes: str = ""


# ----------------------------
# Planning inputs
# ----------------------------

PHASE_COLORS = {
    "DOC": "D9D2E9",        # light purple
    "PORT": "9DC3E6",       # light blue
    "MARINE": "BFBFBF",     # mid grey
    "AGI": "C6E0B4",        # light green
    "RETURN": "E7E6E6",     # light grey
    "MILESTONE": "F8CBAD",  # light orange
}

HEADER_BLUE = "1F4E79"
SUMMARY_GREY = "404040"
GRID_BORDER = "A6A6A6"

WEEKEND_SHADE = "F2F2F2"
TODAY_SHADE = "FFF2CC"

OWNER_DEFAULT = "Samsung / Mammoet / LCT Operator"
LOC_MZP = "Mina Zayed Port"
LOC_AGI = "AGI Site"
LOC_MARINE = "MZP↔AGI (Marine)"

# Planning gate text (must be confirmed against latest MS + port control before execution)
WEATHER_GATE_RORO = "RoRo/MS: wind ≤20kt; Hs ≤0.6m; vis ≥55m"
WEATHER_GATE_LAND = "Land/MS: wind ≤20kt; vis ≥30m"
WEATHER_GATE_MARINE = "Marine/MS: wind ≤20kt; Hs ≤0.6m (RoRo window)"

TIDE_GATE_RORO = "High tide window (see Tide_Peaks_MZP; confirm latest port tide table)"

PHASE_LIST = ["SUMMARY", "DOC", "PORT", "MARINE", "AGI", "RETURN", "MILESTONE"]
STATUS_LIST = ["", "Not Started", "In Progress", "Complete", "Hold", "Risk"]


# ----------------------------
# Helper functions
# ----------------------------


def parse_date(s: str) -> dt.date:
    return dt.datetime.strptime(s.strip(), "%Y-%m-%d").date()


def daterange(d0: dt.date, d1: dt.date) -> List[dt.date]:
    if d1 < d0:
        raise ValueError("end date earlier than start date")
    out: List[dt.date] = []
    cur = d0
    while cur <= d1:
        out.append(cur)
        cur += dt.timedelta(days=1)
    return out


def excel_date(d: dt.date) -> dt.datetime:
    return dt.datetime(d.year, d.month, d.day)


def thin_border(style: str = "thin") -> Border:
    side = Side(style=style, color=GRID_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_cell_style(cell, *, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width


def set_gantt_day_column_width(ws, start_col: int, end_col: int, width: float = 3.2):
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ----------------------------
# Planning logic (Realistic)
# ----------------------------


def build_realistic_tasks(
    campaign_start: dt.date,
    n_transformers: int = 6,
    docs_days_tr1: int = 2,
    docs_days_next: int = 3,
    return_days_each: int = 2,
    final_trip_return_days: int = 4,
) -> Tuple[List[Task], List[Tuple[dt.date, str, float, str]]]:
    """Build tasks for TR1..TRn and a simple tide-peak list (placeholder)."""

    if n_transformers < 1:
        raise ValueError("n_transformers must be >= 1")

    tasks: List[Task] = []

    # Tide peaks placeholder list aligned to load-out/load-in days in this simplified pattern.
    tide_peaks: List[Tuple[dt.date, str, float, str]] = []
    tide_source = "MAMMOET_AGI TR.pdf (Mina Zayed tidal prediction) – planning placeholder"

    base_peak_date = campaign_start + dt.timedelta(days=2)
    peak_times = ["10:00", "12:00", "15:00", "16:00", "07:00", "09:00", "13:00", "14:00"]
    peak_levels = [1.91, 1.99, 2.01, 1.84, 1.97, 2.03, 2.08, 1.92]
    for i in range(n_transformers * 2):
        d = base_peak_date + dt.timedelta(days=2 * i)
        t = peak_times[i % len(peak_times)]
        h = peak_levels[i % len(peak_levels)]
        tide_peaks.append((d, t, h, tide_source))

    prev_return_start: Optional[dt.date] = None
    campaign_first_start = campaign_start
    campaign_last_finish: Optional[dt.date] = None

    for tr in range(1, n_transformers + 1):
        docs_start = campaign_start if tr == 1 else (prev_return_start or campaign_start)
        docs_days = docs_days_tr1 if tr == 1 else docs_days_next
        docs_finish = docs_start + dt.timedelta(days=docs_days - 1)

        preload_day = docs_finish
        loadout_day = preload_day + dt.timedelta(days=1)
        sail_day = loadout_day + dt.timedelta(days=1)
        loadin_day = sail_day + dt.timedelta(days=1)

        return_start = loadin_day + dt.timedelta(days=1)
        return_days = final_trip_return_days if tr == n_transformers else return_days_each
        return_finish = return_start + dt.timedelta(days=return_days - 1)

        prev_return_start = return_start
        campaign_last_finish = return_finish if (campaign_last_finish is None or return_finish > campaign_last_finish) else campaign_last_finish

        tr_summary_start = docs_start
        tr_summary_finish = return_finish

        base_id = tr * 100

        tasks.append(
            Task(
                id=str(tr),
                wbs=f"{tr}.0",
                task=f"TR{tr} – Transport Cycle (MZP→AGI→MZP)",
                phase="SUMMARY",
                location=f"{LOC_MZP}/{LOC_AGI}",
                owner=OWNER_DEFAULT,
                start=tr_summary_start,
                finish=tr_summary_finish,
                status="",
                notes="Realistic cycle incl. return/reset buffer (metocean/tide/HM approvals risk).",
            )
        )

        tasks.extend(
            [
                Task(
                    id=str(base_id + 1),
                    wbs=f"{tr}.1",
                    task=f"TR{tr} – Docs/Permits package (PTW, gate pass, port marine pack)",
                    phase="DOC",
                    location=LOC_MZP,
                    owner=OWNER_DEFAULT,
                    start=docs_start,
                    finish=docs_finish,
                    status="Not Started",
                    notes="Complete ≥48h ahead of load-out where possible.",
                ),
                Task(
                    id=str(base_id + 2),
                    wbs=f"{tr}.2",
                    task=f"TR{tr} – Pre-load on SPMT, lashing, steelworks prep",
                    phase="PORT",
                    location=LOC_MZP,
                    owner="Mammoet + Port Ops",
                    start=preload_day,
                    finish=preload_day,
                    status="Not Started",
                    weather_gate=WEATHER_GATE_LAND,
                    notes="Daylight ops. Verify SPMT readiness & functional checks.",
                ),
                Task(
                    id=str(base_id + 3),
                    wbs=f"{tr}.3",
                    task=f"TR{tr} – Load-out (RoRo) to LCT deck (tide-assisted)",
                    phase="PORT",
                    location=LOC_MZP,
                    owner="Mammoet + LCT Master + Port Control",
                    start=loadout_day,
                    finish=loadout_day,
                    status="Not Started",
                    tide_gate=TIDE_GATE_RORO,
                    weather_gate=WEATHER_GATE_RORO,
                    notes="Ramp angle/UKC per RoRo calc; hold if metocean exceeds limits.",
                ),
                Task(
                    id=str(base_id + 4),
                    wbs=f"{tr}.4",
                    task=f"TR{tr} – Seafastening + MWS/MPI + Sail-away / Marine transit",
                    phase="MARINE",
                    location=LOC_MARINE,
                    owner="Mammoet + MWS + LCT Crew",
                    start=sail_day,
                    finish=sail_day,
                    status="Not Started",
                    weather_gate=WEATHER_GATE_MARINE,
                    notes="Seafastening & inspections prior to sail (per MS/MWS).",
                ),
                Task(
                    id=str(base_id + 5),
                    wbs=f"{tr}.5",
                    task=f"TR{tr} – AGI arrival/berth + Load-in (tide-assisted) to foundation",
                    phase="AGI",
                    location=LOC_AGI,
                    owner="AGI Port Control + Mammoet + Client",
                    start=loadin_day,
                    finish=loadin_day,
                    status="Not Started",
                    tide_gate=TIDE_GATE_RORO,
                    weather_gate=WEATHER_GATE_RORO,
                    notes="Coordinate AGI Port Control; ensure route readiness and exclusion zone.",
                ),
                Task(
                    id=str(base_id + 6),
                    wbs=f"{tr}.6",
                    task=f"TR{tr} – Return to Mina Zayed + reset + buffer (weather/tide)",
                    phase="RETURN",
                    location=LOC_MARINE,
                    owner="LCT Crew + Port Ops",
                    start=return_start,
                    finish=return_finish,
                    status="Not Started",
                    weather_gate="Return/reset: metocean per MS; buffer included",
                    notes=(
                        "Transit back; prepare SPMT configuration for next transformer."
                        if tr < n_transformers
                        else "Transit back + demobilization/close-out. Includes final contingency buffer."
                    ),
                ),
            ]
        )

    if campaign_last_finish is None:
        campaign_last_finish = campaign_start

    tasks.insert(
        0,
        Task(
            id="0",
            wbs="0",
            task="Campaign: TR1–TR6 (1TR per voyage) – Realistic w/ buffers",
            phase="SUMMARY",
            location=f"{LOC_MZP}/{LOC_AGI}",
            owner=OWNER_DEFAULT,
            start=campaign_first_start,
            finish=campaign_last_finish,
            notes="Planning basis: sequential 1TR/Trip; built-in buffer + final close-out/contingency.",
        ),
    )

    tasks.append(
        Task(
            id="M1",
            wbs="M",
            task="Milestone: TR6 returned to Mina Zayed (campaign complete)",
            phase="MILESTONE",
            location=LOC_MZP,
            owner=OWNER_DEFAULT,
            start=campaign_last_finish,
            finish=campaign_last_finish,
            status="",
            notes="Operational transport cycles complete.",
        )
    )

    return tasks, tide_peaks


# ----------------------------
# Excel generation
# ----------------------------


def create_workbook(
    tasks: List[Task],
    tide_peaks: List[Tuple[dt.date, str, float, str]],
    out_path: Path,
    title: str,
    planning_basis: str,
    generated_on: dt.date,
    timeline_pad_days: int = 14,
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("Schedule")

    min_start = min(t.start for t in tasks)
    max_finish = max(t.finish for t in tasks)
    days = daterange(min_start, max_finish + dt.timedelta(days=timeline_pad_days))

    # Layout
    TITLE_ROW_1 = 1
    TITLE_ROW_2 = 2
    TITLE_ROW_3 = 3
    MONTH_ROW = 4
    WEEK_ROW = 5
    HEADER_ROW = 6
    FIRST_DATA_ROW = 7

    META_COLS = [
        "ID",
        "WBS",
        "Task",
        "Phase",
        "Location",
        "Owner",
        "Start",
        "Finish",
        "Dur (d)",
        "Progress %",
        "Baseline Start",
        "Baseline Finish",
        "Status",
        "Tide gate (planned)",
        "Weather gate",
        "Notes",
    ]

    meta_start_col = 1
    gantt_start_col = len(META_COLS) + 1
    gantt_end_col = gantt_start_col + len(days) - 1
    gantt_end_letter = get_column_letter(gantt_end_col)

    last_data_row = FIRST_DATA_ROW + len(tasks) - 1

    # Title block
    for r in (TITLE_ROW_1, TITLE_ROW_2, TITLE_ROW_3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=gantt_end_col)

    ws["A1"].value = title
    ws["A2"].value = planning_basis
    ws["A3"].value = f"Generated: {generated_on.isoformat()} (system date)"

    title_font = Font(bold=True, size=14, color="FFFFFF")
    subtitle_font = Font(size=11, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    title_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in (TITLE_ROW_1, TITLE_ROW_2, TITLE_ROW_3):
        for c in range(1, gantt_end_col + 1):
            cell = ws.cell(r, c)
            apply_cell_style(
                cell,
                font=title_font if r == TITLE_ROW_1 else subtitle_font,
                fill=title_fill,
                alignment=title_align,
            )
        ws.row_dimensions[r].height = 20 if r == TITLE_ROW_1 else 18

    # Month row (merged)
    month_fill = PatternFill("solid", fgColor="2F75B5")
    month_font = Font(bold=True, color="FFFFFF")
    month_align = Alignment(horizontal="center", vertical="center")

    ws.cell(MONTH_ROW, gantt_start_col).value = "Month"
    for c in range(1, gantt_start_col):
        ws.cell(MONTH_ROW, c).fill = PatternFill("solid", fgColor=HEADER_BLUE)
    for c in range(1, gantt_start_col):
        ws.cell(MONTH_ROW, c).border = thin_border()

    # Merge month segments across day columns
    seg_start = 0
    while seg_start < len(days):
        y, m = days[seg_start].year, days[seg_start].month
        seg_end = seg_start
        while seg_end < len(days) and days[seg_end].year == y and days[seg_end].month == m:
            seg_end += 1
        c1 = gantt_start_col + seg_start
        c2 = gantt_start_col + seg_end - 1
        ws.merge_cells(start_row=MONTH_ROW, start_column=c1, end_row=MONTH_ROW, end_column=c2)
        cell = ws.cell(MONTH_ROW, c1)
        cell.value = dt.date(y, m, 1).strftime("%b-%Y")
        apply_cell_style(cell, font=month_font, fill=month_fill, alignment=month_align, border=thin_border("thin"))
        for c in range(c1 + 1, c2 + 1):
            ws.cell(MONTH_ROW, c).fill = month_fill
            ws.cell(MONTH_ROW, c).border = thin_border()
        seg_start = seg_end

    ws.row_dimensions[MONTH_ROW].height = 18

    # Week row
    week_fill = PatternFill("solid", fgColor="BDD7EE")
    week_font = Font(bold=True, color="1F4E79")
    for c in range(1, gantt_start_col):
        cell = ws.cell(WEEK_ROW, c)
        cell.fill = week_fill
        cell.border = thin_border()

    ws.cell(WEEK_ROW, gantt_start_col).value = "Week"
    ws.cell(WEEK_ROW, gantt_start_col).font = week_font
    ws.cell(WEEK_ROW, gantt_start_col).alignment = Alignment(horizontal="center", vertical="center")

    for i, d in enumerate(days):
        col = gantt_start_col + i
        cell = ws.cell(WEEK_ROW, col)
        cell.value = f"W{d.isocalendar().week:02d}"
        apply_cell_style(cell, font=week_font, fill=week_fill, alignment=Alignment(horizontal="center", vertical="center"), border=thin_border())

    ws.row_dimensions[WEEK_ROW].height = 18

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, name in enumerate(META_COLS, start=meta_start_col):
        cell = ws.cell(HEADER_ROW, idx, value=name)
        apply_cell_style(cell, font=header_font, fill=header_fill, alignment=header_align, border=thin_border())

    for i, d in enumerate(days):
        col = gantt_start_col + i
        cell = ws.cell(HEADER_ROW, col, value=excel_date(d))
        apply_cell_style(cell, font=header_font, fill=header_fill, alignment=header_align, border=thin_border(), number_format="dd-mmm")

    ws.row_dimensions[HEADER_ROW].height = 36

    # Data rows
    date_fmt = "dd-mmm-yy"
    base_border = thin_border()

    for r, t in enumerate(tasks, start=FIRST_DATA_ROW):
        ws.row_dimensions[r].height = 18

        row_values = [
            t.id,
            t.wbs,
            t.task,
            t.phase,
            t.location,
            t.owner,
            excel_date(t.start),
            excel_date(t.finish),
            None,  # duration formula
            t.progress_pct,
            excel_date(t.baseline_start) if t.baseline_start else None,
            excel_date(t.baseline_finish) if t.baseline_finish else None,
            t.status,
            t.tide_gate,
            t.weather_gate,
            t.notes,
        ]

        for c, v in enumerate(row_values, start=1):
            cell = ws.cell(r, c, value=v)

            if c == 7 or c == 8 or c == 11 or c == 12:
                if v is not None:
                    cell.number_format = date_fmt
                align = Alignment(horizontal="center", vertical="center")
            elif c in (1, 2):
                align = Alignment(horizontal="center", vertical="center")
            elif c in (4, 9, 10, 13):
                align = Alignment(horizontal="center", vertical="center")
            else:
                align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if c == 9:
                cell.value = f"=H{r}-G{r}+1"
                cell.number_format = "0"

            apply_cell_style(cell, alignment=align, border=base_border)

        # Summary styling
        if t.phase == "SUMMARY":
            for c in range(1, len(META_COLS) + 1):
                cell = ws.cell(r, c)
                apply_cell_style(
                    cell,
                    font=Font(bold=True, color="FFFFFF"),
                    fill=PatternFill("solid", fgColor=SUMMARY_GREY),
                    border=base_border,
                )

        # Gantt grid cells (blank but bordered)
        for c in range(gantt_start_col, gantt_end_col + 1):
            cell = ws.cell(r, c)
            apply_cell_style(cell, alignment=Alignment(horizontal="center", vertical="center"), border=base_border)

    # Column widths (template-like)
    widths = {
        "A": 6,
        "B": 8,
        "C": 50,
        "D": 10,
        "E": 14,
        "F": 18,
        "G": 12,
        "H": 12,
        "I": 8,
        "J": 10,
        "K": 13,
        "L": 13,
        "M": 12,
        "N": 24,
        "O": 22,
        "P": 42,
    }
    for col, w in widths.items():
        set_col_width(ws, col, w)
    set_gantt_day_column_width(ws, gantt_start_col, gantt_end_col, width=3.2)

    # Freeze panes
    ws.freeze_panes = ws.cell(FIRST_DATA_ROW, gantt_start_col)

    # Excel Table for metadata (filterable)
    table_ref = f"A{HEADER_ROW}:{get_column_letter(len(META_COLS))}{last_data_row}"
    tbl = Table(displayName="TaskRegister", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

    # Data validation
    dv_phase = DataValidation(type="list", formula1='"' + ','.join(PHASE_LIST) + '"', allow_blank=True)
    ws.add_data_validation(dv_phase)
    dv_phase.add(f"D{FIRST_DATA_ROW}:D{last_data_row}")

    dv_status = DataValidation(type="list", formula1='"' + ','.join(STATUS_LIST) + '"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"M{FIRST_DATA_ROW}:M{last_data_row}")

    dv_progress = DataValidation(type="whole", operator="between", formula1="0", formula2="100", allow_blank=True)
    ws.add_data_validation(dv_progress)
    dv_progress.add(f"J{FIRST_DATA_ROW}:J{last_data_row}")

    # Progress DataBar (Smartsheet-like)
    ws.conditional_formatting.add(
        f"J{FIRST_DATA_ROW}:J{last_data_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="63C384", showValue=True),
    )

    # Conditional formatting for Gantt grid
    cf_range = f"{get_column_letter(gantt_start_col)}{FIRST_DATA_ROW}:{gantt_end_letter}{last_data_row}"
    day_col_letter = get_column_letter(gantt_start_col)

    def add_fill_rule(formula: str, fill_hex: str, priority: int):
        dxf = DifferentialStyle(fill=PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid"))
        rule = Rule(type="expression", dxf=dxf, formula=[formula])
        rule.priority = priority
        ws.conditional_formatting.add(cf_range, rule)

    # Weekend shading (low priority)
    add_fill_rule(f"AND(WeekendShadeOn, WEEKDAY({day_col_letter}${HEADER_ROW},2)>5)", WEEKEND_SHADE, priority=90)
    # Today shading
    add_fill_rule(f"AND(TodayShadeOn, {day_col_letter}${HEADER_ROW}=TODAY())", TODAY_SHADE, priority=80)

    # Phase bars (IMPORTANT: row is relative; no $ on row number)
    pr = 10
    for phase, color in PHASE_COLORS.items():
        add_fill_rule(
            f'AND($D{FIRST_DATA_ROW}="{phase}",{day_col_letter}${HEADER_ROW}>=$G{FIRST_DATA_ROW},{day_col_letter}${HEADER_ROW}<=$H{FIRST_DATA_ROW})',
            color,
            priority=pr,
        )
        pr += 1

    # Month boundary thicker borders (visual)
    med_side = Side(style="medium", color=GRID_BORDER)
    for i, d in enumerate(days):
        if d.day == 1:
            col = gantt_start_col + i
            for r in range(MONTH_ROW, last_data_row + 1):
                cell = ws.cell(r, col)
                cell.border = Border(
                    left=med_side,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{TITLE_ROW_1}:{HEADER_ROW}"

    # ---------------- Gantt chart sheet ----------------
    wc = wb.create_sheet("Gantt_Chart")
    wc["A1"].value = "AGI TR1–TR6 Gantt (Stacked Bar) – generated from Schedule"
    wc["A2"].value = "Technique: stacked bar (Start offset hidden)."
    wc["A3"].value = "Task"
    wc["B3"].value = "Start (date)"
    wc["C3"].value = "Start offset (d)"
    wc["D3"].value = "Duration (d)"

    wc["A1"].font = Font(bold=True, size=12)
    wc["A2"].font = Font(size=10)

    for c in range(1, 5):
        h = wc.cell(3, c)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = base_border

    wc.column_dimensions["A"].width = 60
    wc.column_dimensions["B"].width = 14
    wc.column_dimensions["C"].width = 16
    wc.column_dimensions["D"].width = 14

    detail_rows = [rr for rr in range(FIRST_DATA_ROW, last_data_row + 1) if ws.cell(rr, 4).value != "SUMMARY"]
    chart_table_first_row = 4

    for i, rr in enumerate(detail_rows):
        out_r = chart_table_first_row + i
        wc.cell(out_r, 1).value = f"=Schedule!C{rr}"
        wc.cell(out_r, 2).value = f"=Schedule!G{rr}"
        wc.cell(out_r, 2).number_format = date_fmt
        min_range = f"$B${chart_table_first_row}:$B${chart_table_first_row + len(detail_rows) - 1}"
        wc.cell(out_r, 3).value = f"=B{out_r}-MIN({min_range})"
        wc.cell(out_r, 4).value = f"=Schedule!I{rr}"
        wc.cell(out_r, 4).number_format = "0"

        for c in range(1, 5):
            wc.cell(out_r, c).border = base_border
            wc.cell(out_r, c).alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center", wrap_text=(c == 1))

    wc.freeze_panes = "A4"

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 150
    chart.title = "Transport Tasks (Offset + Duration)"

    data = Reference(wc, min_col=3, min_row=3, max_col=4, max_row=chart_table_first_row + len(detail_rows) - 1)
    cats = Reference(wc, min_col=1, min_row=chart_table_first_row, max_row=chart_table_first_row + len(detail_rows) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    if chart.series:
        chart.series[0].graphicalProperties.noFill = True
        chart.series[0].graphicalProperties.line.noFill = True
        chart.series[1].graphicalProperties.solidFill = "4F81BD"
        chart.series[1].graphicalProperties.line.solidFill = "4F81BD"

    wc.add_chart(chart, "F3")

    # ---------------- Tide peaks sheet ----------------
    wt = wb.create_sheet("Tide_Peaks_MZP")
    headers = ["Date (GST)", "Peak time (approx)", "Peak tide (m)", "Source"]
    for c, h in enumerate(headers, start=1):
        cell = wt.cell(1, c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = base_border

    wt.column_dimensions["A"].width = 14
    wt.column_dimensions["B"].width = 18
    wt.column_dimensions["C"].width = 14
    wt.column_dimensions["D"].width = 60

    for i, (d, t, h, src) in enumerate(tide_peaks, start=2):
        wt.cell(i, 1, value=excel_date(d)).number_format = date_fmt
        wt.cell(i, 2, value=t)
        wt.cell(i, 3, value=h)
        wt.cell(i, 4, value=src)
        for c in range(1, 5):
            wt.cell(i, c).border = base_border
            wt.cell(i, c).alignment = Alignment(horizontal="center" if c in (1, 2, 3) else "left", vertical="center", wrap_text=(c == 4))

    # ---------------- Assumptions sheet ----------------
    wa = wb.create_sheet("Assumptions_Refs")
    wa["A1"].value = "Assumptions / Controls / References (AGI Site)"
    wa["A1"].font = Font(bold=True, size=14)

    wa["A3"].value = "Controls (used by VBA macros / conditional formatting)"
    wa["A3"].font = Font(bold=True, size=12)

    wa["A4"].value = "CampaignStart"
    wa["B4"].value = excel_date(min_start)
    wa["B4"].number_format = date_fmt

    wa["A5"].value = "WeekendShadeOn"
    wa["B5"].value = True

    wa["A6"].value = "TodayShadeOn"
    wa["B6"].value = True

    wa["A8"].value = "Operational gates (planning text)"
    wa["A8"].font = Font(bold=True)

    bullets = [
        "• One transformer per voyage (1TR/Trip); sequential campaign assumed.",
        "• Docs/permits overlapped with vessel return/reset to minimize total duration.",
        "• Tide-assisted RoRo on load-out and load-in days; replace Tide_Peaks_MZP with official tide table.",
        "• Weather gates used as planning text; confirm latest MS + daily forecast before execution.",
        "• Buffers included: return/reset per trip + final close-out/contingency buffer on TR6.",
    ]

    rr = 9
    for b in bullets:
        wa[f"A{rr}"].value = b
        wa[f"A{rr}"].alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    rr += 1
    wa[f"A{rr}"].value = "Template reference links (Excel Gantt techniques)"
    wa[f"A{rr}"].font = Font(bold=True)
    rr += 1

    refs = [
        ("Smartsheet – Free Excel Gantt templates", "https://www.smartsheet.com/gantt-chart-excel-templates"),
        ("Vertex42 – Gantt chart template for Excel", "https://www.vertex42.com/ExcelTemplates/excel-gantt-chart.html"),
        ("Microsoft Support – simulate Gantt with stacked bar", "https://support.microsoft.com/en-us/office/present-your-data-in-a-gantt-chart-in-excel-f8910ab4-ceda-4521-8207-f0fb34d9e2b6"),
    ]

    wa[f"A{rr}"].value = "Source"
    wa[f"B{rr}"].value = "Link"
    for c in (f"A{rr}", f"B{rr}"):
        wa[c].font = header_font
        wa[c].fill = header_fill
        wa[c].alignment = header_align
        wa[c].border = base_border
    rr += 1

    for name, link in refs:
        wa[f"A{rr}"].value = name
        wa[f"B{rr}"].value = link
        wa[f"A{rr}"].alignment = Alignment(wrap_text=True, vertical="top")
        wa[f"B{rr}"].alignment = Alignment(wrap_text=True, vertical="top")
        wa[f"A{rr}"].border = base_border
        wa[f"B{rr}"].border = base_border
        rr += 1

    rr += 1
    wa[f"A{rr}"].value = "VBA"
    wa[f"A{rr}"].font = Font(bold=True)
    rr += 1
    wa[f"A{rr}"].value = "Import the provided .bas file (Alt+F11 → File → Import File) and (optionally) add buttons to run macros."

    wa.column_dimensions["A"].width = 60
    wa.column_dimensions["B"].width = 90

    # Named ranges for controls
    wb.defined_names.add(DefinedName("CampaignStart", attr_text="'Assumptions_Refs'!$B$4"))
    wb.defined_names.add(DefinedName("WeekendShadeOn", attr_text="'Assumptions_Refs'!$B$5"))
    wb.defined_names.add(DefinedName("TodayShadeOn", attr_text="'Assumptions_Refs'!$B$6"))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------
# CLI
# ----------------------------


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Generate AGI TR transport Gantt Excel (realistic w/ buffers).")
    p.add_argument("--start", default="2026-01-13", help="Campaign start date (YYYY-MM-DD). Default: 2026-01-13")
    p.add_argument("--n", type=int, default=6, help="Number of transformers (default 6).")
    p.add_argument("--out", default="", help="Output xlsx path. Default auto-named.")
    return p


def main() -> int:
    args = build_parser().parse_args()
    start = parse_date(args.start)
    n = int(args.n)

    tasks, tide_peaks = build_realistic_tasks(
        campaign_start=start,
        n_transformers=n,
        docs_days_tr1=2,
        docs_days_next=3,
        return_days_each=2,
        final_trip_return_days=4,
    )

    min_start = min(t.start for t in tasks)
    max_finish = max(t.finish for t in tasks)

    out = Path(args.out) if args.out else Path(
        f"AGI_TR_Transport_Gantt_{min_start.isoformat()}_to_{max_finish.isoformat()}_realistic_buffer_v6.xlsx"
    )

    title = "AGI Site – HVDC Transformer (TR1–TR6) – Realistic Transport Schedule (1TR/Trip + Buffers)"
    planning_basis = "Basis: 1TR/Trip sequential; tide-assisted RoRo; weather-gated; buffers included."

    create_workbook(
        tasks=tasks,
        tide_peaks=tide_peaks,
        out_path=out,
        title=title,
        planning_basis=planning_basis,
        generated_on=dt.date.today(),
        timeline_pad_days=14,
    )

    print(f"[OK] Wrote: {out.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
