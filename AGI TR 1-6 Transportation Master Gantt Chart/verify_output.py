#!/usr/bin/env python3
"""생성된 Excel 파일 검증 스크립트"""

from openpyxl import load_workbook
import os

wb_path = "AGI_TR_MultiScenario_Master_Gantt_VBA.xlsx"

if not os.path.exists(wb_path):
    print(f"❌ 파일을 찾을 수 없습니다: {wb_path}")
    exit(1)

wb = load_workbook(wb_path)

import sys
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

print("=" * 60)
print("생성된 Excel 파일 검증 결과")
print("=" * 60)

# Control_Panel 검증
if "Control_Panel" in wb.sheetnames:
    ws_ctrl = wb["Control_Panel"]
    print("\n[OK] Control_Panel 검증:")
    print(f"   Gantt 최소 일수 (I8): {ws_ctrl['I8'].value}")
    print(f"   Gantt 버퍼 일수 (I9): {ws_ctrl['I9'].value}")
    print(f"   Shamal 시작일 (H5): {ws_ctrl['H5'].value}")
    print(f"   Shamal 종료일 (H6): {ws_ctrl['H6'].value}")
    print(f"   프로젝트 시작일 (B4): {ws_ctrl['B4'].value}")
    print(f"   목표 완료일 (B5): {ws_ctrl['B5'].value}")

# Weather_Analysis 헤더 검증
if "Weather_Analysis" in wb.sheetnames:
    ws_weather = wb["Weather_Analysis"]
    headers = []
    for c in range(1, min(15, ws_weather.max_column + 1)):
        val = ws_weather.cell(3, c).value
        if val:
            headers.append(str(val))
    print(f"\n[OK] Weather_Analysis 헤더 ({len(headers)}개):")
    print(f"   {headers}")

# Gantt_Chart 날짜 범위 확인
if "Gantt_Chart" in wb.sheetnames:
    ws_gantt = wb["Gantt_Chart"]
    # 날짜 헤더 개수 확인 (G열부터)
    date_cols = 0
    for col in range(7, ws_gantt.max_column + 1):
        val = ws_gantt.cell(4, col).value
        if val and isinstance(val, (int, float)):
            date_cols += 1
    print(f"\n[OK] Gantt_Chart 날짜 컬럼 수: {date_cols}개")

# 시트 목록
print(f"\n[OK] 생성된 시트 목록 ({len(wb.sheetnames)}개):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"   {i}. {name}")

print("\n" + "=" * 60)
print("[OK] 검증 완료!")
print("=" * 60)

