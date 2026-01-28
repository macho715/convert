# pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import date, timedelta

# -----------------------------
# Reference plan (dates are reference only)
# Aggregate: MW4 Depart -> AGI Offload (next day)
# Debris(backload): AGI Debris Loading -> MW4 Debris Offloading (next day)
# -----------------------------
plan = [
    ("86", "Aggregate", "20mm",        date(2025,12,22), date(2025,12,23), None,              None),
    ("Debris-7", "Debris", "Debris",   None,            None,              date(2025,12,23),  date(2025,12,24)),
    ("87", "Aggregate", "10mm",        date(2025,12,24), date(2025,12,25), None,              None),
    ("Debris-8", "Debris", "Debris",   None,            None,              date(2025,12,25),  date(2025,12,26)),  # IN PROGRESS
    ("88", "Aggregate", "Dune Sand",   date(2025,12,26), date(2025,12,27), None,              None),
    ("Debris-9", "Debris", "Debris",   None,            None,              date(2025,12,27),  date(2025,12,28)),
    ("89", "Aggregate", "5mm",         date(2025,12,28), date(2025,12,29), None,              None),
    ("Debris-10","Debris", "Debris",   None,            None,              date(2025,12,29),  date(2025,12,30)),
    ("90", "Aggregate", "20mm",        date(2025,12,30), date(2025,12,31), None,              None),
    ("Debris-11","Debris", "Debris",   None,            None,              date(2025,12,31),  date(2026,1,1)),
    ("91", "Aggregate", "Dune Sand",   date(2026,1,1),   date(2026,1,2),   None,              None),
    ("Debris-12","Debris", "Debris",   None,            None,              date(2026,1,2),    date(2026,1,3)),
    ("92", "Aggregate", "5mm",         date(2026,1,3),   date(2026,1,4),   None,              None),
    ("Debris-13","Debris", "Debris",   None,            None,              date(2026,1,4),    date(2026,1,5)),
    ("93", "Aggregate", "20mm",        date(2026,1,5),   date(2026,1,6),   None,              None),
]
DEBRIS8_SEQ = 4  # 86=1, D7=2, 87=3, D8=4

plan_start = min([p[3] for p in plan if p[3]] + [p[5] for p in plan if p[5]])
plan_end   = max([p[4] for p in plan if p[4]] + [p[6] for p in plan if p[6]])
view_end   = plan_end + timedelta(days=14)

dates=[]
d=plan_start
while d<=view_end:
    dates.append(d); d+=timedelta(days=1)

wb = Workbook()

# -----------------------------
# Control sheet (user input)
# -----------------------------
ctl = wb.active
ctl.title = "Control"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ctl["A1"] = "Debris-8: IN PROGRESS. Enter actual completion (date/time) to shift subsequent schedule automatically."
ctl["A1"].font = Font(bold=True); ctl["A1"].alignment = left
ctl.merge_cells("A1:D1")

ctl["A3"]="Debris-8 Actual Completion (date/time)"; ctl["B3"]=""  # <-- INPUT
ctl["A4"]="Debris-8 Planned MW4 Offloading Date";    ctl["B4"]=date(2025,12,26)
ctl["A5"]="Delay Days (auto)";                       ctl["B5"]='=IF(B3="",0,MAX(0,INT(B3)-B4))'

for r in range(3,6):
    for c in range(1,3):
        cell=ctl.cell(r,c); cell.border=border
        cell.alignment = left if c==1 else center

ctl.column_dimensions["A"].width=44
ctl.column_dimensions["B"].width=24

# -----------------------------
# Cross_Gantt sheet
# -----------------------------
ws = wb.create_sheet("Cross_Gantt")

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)

fill_row_agg = PatternFill("solid", fgColor="E2EFDA")
fill_row_deb = PatternFill("solid", fgColor="FCE4D6")
fill_inprog  = PatternFill("solid", fgColor="D9D9D9")

fill_mw4         = PatternFill("solid", fgColor="4F81BD")  # MW4 Depart (Agg)
fill_agi_off     = PatternFill("solid", fgColor="C6EFCE")  # AGI Offload (Agg)
fill_agi_debload = PatternFill("solid", fgColor="F4B084")  # AGI Debris Loading (Deb)
fill_mw4_deboff  = PatternFill("solid", fgColor="F8CBAD")  # MW4 Debris Offloading (Deb)

dash = Side(style="dashed", color="7F7F7F")
border_dashed = Border(left=dash, right=dash, top=dash, bottom=dash)

headers = [
    "Seq","Trip","Type","Material",
    "MW4 Depart (Agg)","AGI Offload (Agg)",
    "AGI Debris Loading (Deb)","MW4 Debris Offloading (Deb)",
    "Status"
]
hidden_headers = ["ShiftFlag","Plan_MW4_Depart","Plan_AGI_Offload","Plan_AGI_Deb_Load","Plan_MW4_Deb_Off"]

thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Header row
col=1
for h in headers:
    c=ws.cell(1,col,h); c.fill=header_fill; c.font=header_font; c.alignment=center; c.border=thin_border
    col+=1
for h in hidden_headers:
    c=ws.cell(1,col,h); c.fill=header_fill; c.font=header_font; c.alignment=center; c.border=thin_border
    col+=1

date_start_col = col
for i,dt in enumerate(dates):
    c=ws.cell(1,date_start_col+i,dt)
    c.number_format="mm-dd"
    c.fill=header_fill; c.font=header_font; c.alignment=center; c.border=thin_border
    ws.column_dimensions[get_column_letter(date_start_col+i)].width=5

# widths visible
for c,w in {1:5,2:12,3:10,4:14,5:14,6:14,7:18,8:20,9:14}.items():
    ws.column_dimensions[get_column_letter(c)].width=w

# hide shift/plan cols
shift_col = len(headers)+1
for c in range(shift_col, date_start_col):
    ws.column_dimensions[get_column_letter(c)].hidden=True

ws.freeze_panes="J2"
ws.row_dimensions[1].height=22

# Populate rows with plan + adjusted formulas
for idx,(trip,typ,mat,p_mw4,p_agi_off,p_debload,p_deboff) in enumerate(plan, start=1):
    r=1+idx
    row_fill = fill_row_deb if typ=="Debris" else fill_row_agg

    ws.cell(r,1,idx).alignment=center
    ws.cell(r,2,trip).alignment=center
    ws.cell(r,3,typ).alignment=center
    ws.cell(r,4,mat).alignment=left
    ws.cell(r,9,"IN PROGRESS" if trip=="Debris-8" else "").alignment=center

    # shift flag (rows AFTER Debris-8)
    ws.cell(r,shift_col, 1 if idx>DEBRIS8_SEQ else 0).alignment=center

    plan_start_col = shift_col+1
    plan_vals=[p_mw4,p_agi_off,p_debload,p_deboff]
    for j,val in enumerate(plan_vals):
        c=ws.cell(r,plan_start_col+j,val)
        if val: c.number_format="yyyy-mm-dd"
        c.alignment=center

    delay_ref="Control!$B$5"
    flag_ref=f"${get_column_letter(shift_col)}{r}"
    plan_ref_E=f"{get_column_letter(plan_start_col)}{r}"
    plan_ref_F=f"{get_column_letter(plan_start_col+1)}{r}"
    plan_ref_G=f"{get_column_letter(plan_start_col+2)}{r}"
    plan_ref_H=f"{get_column_letter(plan_start_col+3)}{r}"
    def adj(plan_ref):
        return f'=IF({plan_ref}="", "", IF({flag_ref}=1, {plan_ref}+{delay_ref}, {plan_ref}))'

    ws.cell(r,5).value=adj(plan_ref_E)
    ws.cell(r,6).value=adj(plan_ref_F)
    ws.cell(r,7).value=adj(plan_ref_G)
    ws.cell(r,8).value=adj(plan_ref_H)
    for c in [5,6,7,8]:
        ws.cell(r,c).number_format="yyyy-mm-dd"; ws.cell(r,c).alignment=center

    # style visible cols
    for c in range(1,10):
        cell=ws.cell(r,c)
        cell.border=thin_border
        cell.fill = fill_inprog if trip=="Debris-8" else row_fill
    if trip=="Debris-8":
        for c in range(1,10):
            ws.cell(r,c).border=border_dashed

# Timeline conditional formatting (exclude Debris-8 row to keep grey)
first_data=2
last_data=1+len(plan)
first_date=date_start_col
last_date=date_start_col+len(dates)-1

deb8_row = 2 + (DEBRIS8_SEQ-1)  # seq4 row index in sheet

def add_cf(r1, r2, fill, col_letter):
    if r1>r2: 
        return
    rng=f"{get_column_letter(first_date)}{r1}:{get_column_letter(last_date)}{r2}"
    top_left=get_column_letter(first_date)
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={top_left}$1=${col_letter}{r1}'], fill=fill))

# Apply CF on two ranges: above Debris-8, and below Debris-8
# MW4 depart / AGI offload / AGI deb load / MW4 deb off
for (r1,r2) in [(first_data, deb8_row-1), (deb8_row+1, last_data)]:
    add_cf(r1,r2,fill_mw4,'E')
    add_cf(r1,r2,fill_agi_off,'F')
    add_cf(r1,r2,fill_agi_debload,'G')
    add_cf(r1,r2,fill_mw4_deboff,'H')

# Grey dashed timeline for Debris-8 row
for c in range(first_date, last_date+1):
    cell=ws.cell(deb8_row,c)
    cell.fill=fill_inprog
    cell.border=border_dashed

out_path="JPT71_CrossGantt_Debris8_InProgress_AutoShift.xlsx"
wb.save(out_path)
print("Saved:", out_path)
