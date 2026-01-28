import datetime as dt
import csv
import os
import sys
import importlib.util

# agi tr submission.py 모듈 import (공백이 있는 파일명 처리)
script_dir = os.path.dirname(os.path.abspath(__file__))
module_path = os.path.join(script_dir, "agi tr submission.py")

spec = importlib.util.spec_from_file_location("agi_tr_submission", module_path)
agi_tr_submission = importlib.util.module_from_spec(spec)
spec.loader.exec_module(agi_tr_submission)
create_gantt_from_tsv = agi_tr_submission.create_gantt_from_tsv

def convert_submission_tsv(input_path, output_path, project_start_date=None):
    """submission.tsv를 agi tr submission.py 형식으로 변환"""
    
    # 프로젝트 시작일 설정 (기본값: 오늘로부터 14일 전)
    if project_start_date is None:
        project_start_date = dt.date.today() - dt.timedelta(days=14)
    else:
        project_start_date = dt.datetime.strptime(project_start_date, '%Y-%m-%d').date()
    
    # Category를 Phase로 매핑
    category_to_phase = {
        "Port / Permits": "PTW",
        "Engineering / Marine": "Engineering",
        "Certificates / Equipment": "Mandatory Docs",
        "Letters": "Submission",
        "AD Maritime NOC": "Submission"
    }
    
    # Lead Time 파싱
    def parse_lead_time(lead_time_str):
        if not lead_time_str or lead_time_str.strip() == "TBD" or "TBD" in lead_time_str:
            return 3
        if "–" in lead_time_str or "-" in lead_time_str:
            parts = lead_time_str.replace("–", "-").split("-")
            try:
                return int(float(parts[0].strip()))
            except:
                return 3
        try:
            return int(float(lead_time_str.strip()))
        except:
            return 3
    
    def determine_risk(category, document):
        if "PTW" in document or "Certificate" in document or "NOC" in document:
            return "HIGH"
        if "Critical" in document:
            return "CRITICAL"
        return "MED"
    
    tasks = []
    
    with open(input_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for idx, row in enumerate(reader, 1):
            category = row.get('Category', '').strip()
            document = row.get('Document', '').strip()
            owner = row.get('Owner', '').strip()
            submit_to = row.get('Submit To', '').strip()
            lead_time_str = row.get('Lead Time (working days)', '').strip()
            purpose = row.get('Purpose / Notes', '').strip()
            remarks = row.get('Remarks', '').strip()
            
            if not document:
                continue
            
            phase = category_to_phase.get(category, "Submission")
            duration = parse_lead_time(lead_time_str)
            risk = determine_risk(category, document)
            wbs = f"S{idx:02d}"
            
            # 날짜 계산 (Target Due Date 기준)
            target_due_date_str = row.get('Target Due Date', '').strip()
            if target_due_date_str:
                try:
                    due_date = dt.datetime.strptime(target_due_date_str, '%Y-%m-%d').date()
                    # Start Date = Due Date - Duration (working days를 고려하여 역산)
                    # 간단히 duration만큼 빼기 (실제로는 working days 계산이 필요하지만 일단 단순화)
                    start_date = due_date - dt.timedelta(days=duration)
                    # 프로젝트 시작일 기준으로 D-XX 태그 계산
                    days_from_start = (start_date - project_start_date).days
                    days_to_due = (due_date - project_start_date).days
                except:
                    # Target Due Date 파싱 실패 시 기존 로직 사용
                    total_tasks = 35
                    days_from_start = (total_tasks - idx) * 1
                    start_date = project_start_date + dt.timedelta(days=days_from_start)
                    due_date = start_date + dt.timedelta(days=duration)
                    days_to_due = days_from_start + duration
            else:
                # Target Due Date가 없으면 기존 로직 사용
                total_tasks = 35
                days_from_start = (total_tasks - idx) * 1
                start_date = project_start_date + dt.timedelta(days=days_from_start)
                due_date = start_date + dt.timedelta(days=duration)
                days_to_due = days_from_start + duration
            
            evidence_parts = []
            if purpose:
                evidence_parts.append(purpose)
            if remarks:
                evidence_parts.append(f"Note: {remarks}")
            evidence = " | ".join(evidence_parts) if evidence_parts else f"Submission item {idx}"
            
            tasks.append({
                'WBS': wbs,
                'Phase': phase,
                'Site': 'Common',
                'Task': document,
                'Owner': owner,
                'Approver': submit_to,
                'Predecessor': '',
                'Start_Tag': f"D+{days_from_start}" if days_from_start >= 0 else f"D{days_from_start}",
                'Start_Date': start_date,
                'Due_Tag': f"D+{days_to_due}" if days_to_due >= 0 else f"D{days_to_due}",
                'Due_Date': due_date,
                'Duration_days': duration,
                'Risk': risk,
                'Evidence': evidence
            })
    
    # 변환된 TSV 저장
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        fieldnames = ['WBS', 'Phase', 'Site', 'Task/Document', 'Owner', 'Approver', 
                     'Predecessor', 'Start_Tag', 'Start_Date', 'Due_Tag', 'Due_Date', 
                     'Duration_days', 'Risk', 'Evidence']
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter='\t')
        writer.writeheader()
        for task in tasks:
            writer.writerow({
                'WBS': task['WBS'],
                'Phase': task['Phase'],
                'Site': task['Site'],
                'Task/Document': task['Task'],
                'Owner': task['Owner'],
                'Approver': task['Approver'],
                'Predecessor': task['Predecessor'],
                'Start_Tag': task['Start_Tag'],
                'Start_Date': task['Start_Date'].strftime('%Y-%m-%d'),
                'Due_Tag': task['Due_Tag'],
                'Due_Date': task['Due_Date'].strftime('%Y-%m-%d'),
                'Duration_days': task['Duration_days'],
                'Risk': task['Risk'],
                'Evidence': task['Evidence']
            })
    
    print(f"✅ Converted {len(tasks)} tasks")
    return output_path

if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, "submission.tsv")
    
    if not os.path.exists(input_path):
        print(f"❌ Error: {input_path} not found")
        sys.exit(1)
    
    converted_path = os.path.join(script_dir, "submission_converted.tsv")
    
    print(f"📖 Reading: {input_path}")
    converted_file = convert_submission_tsv(input_path, converted_path)
    
    print(f"📊 Generating Gantt chart...")
    wb = create_gantt_from_tsv(converted_path)
    
    # VBA 코드 시트 추가
    from openpyxl.styles import Font, PatternFill, Alignment
    
    ws_vba = wb.create_sheet("VBA_Code")
    
    vba_code = """' ============================================
' AGI TR 1-6 Transportation - VBA Macros
' ============================================
' 사용 방법:
' 1. Alt+F11을 눌러 VBA 에디터를 엽니다
' 2. Insert > Module을 선택합니다
' 3. 아래 코드를 복사하여 붙여넣습니다
' 4. F5를 눌러 실행하거나 매크로로 등록합니다
' ============================================

Sub UpdateProjectStartDate()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Control_Panel")
    Dim newDate As Date
    newDate = InputBox("새 프로젝트 시작일을 입력하세요 (YYYY-MM-DD):", "프로젝트 시작일 업데이트")
    If IsDate(newDate) Then
        ws.Range("B2").Value = newDate
        Application.Calculate
        MsgBox "프로젝트 시작일이 업데이트되었습니다.", vbInformation
    Else
        MsgBox "올바른 날짜 형식을 입력하세요.", vbExclamation
    End If
End Sub

Sub RefreshGanttChart()
    Application.Calculate
    ThisWorkbook.Worksheets("Gantt_Chart").Activate
    MsgBox "Gantt 차트가 새로고침되었습니다.", vbInformation
End Sub

Sub FilterByPhase()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Schedule_Data")
    Dim phaseFilter As String
    phaseFilter = InputBox("필터링할 Phase를 입력하세요:", "Phase 필터")
    If phaseFilter <> "" Then
        ws.Range("A1").CurrentRegion.AutoFilter Field:=2, Criteria1:=phaseFilter
        MsgBox phaseFilter & " Phase로 필터링되었습니다.", vbInformation
    End If
End Sub

Sub FilterByRisk()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Schedule_Data")
    Dim riskFilter As String
    riskFilter = InputBox("필터링할 Risk를 입력하세요 (CRITICAL/HIGH/MED):", "Risk 필터")
    If riskFilter <> "" Then
        ws.Range("A1").CurrentRegion.AutoFilter Field:=13, Criteria1:=riskFilter
        MsgBox riskFilter & " Risk로 필터링되었습니다.", vbInformation
    End If
End Sub

Sub ClearFilters()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Schedule_Data")
    On Error Resume Next
    ws.AutoFilterMode = False
    On Error GoTo 0
    MsgBox "모든 필터가 제거되었습니다.", vbInformation
End Sub

Sub ExportToPDF()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Gantt_Chart")
    Dim filePath As String
    filePath = ThisWorkbook.Path & "\AGI_TR_Gantt_Chart_" & Format(Now, "YYYYMMDD_HHMMSS") & ".pdf"
    ws.ExportAsFixedFormat Type:=xlTypePDF, Filename:=filePath, Quality:=xlQualityStandard
    MsgBox "PDF가 생성되었습니다: " & filePath, vbInformation
End Sub

Sub HighlightCriticalTasks()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Gantt_Chart")
    Dim cell As Range
    For Each cell In ws.Range("A1").CurrentRegion
        If cell.Value = "CRITICAL" Then
            cell.Interior.Color = RGB(255, 0, 0)
            cell.Font.Color = RGB(255, 255, 255)
        End If
    Next cell
    MsgBox "Critical 작업이 강조되었습니다.", vbInformation
End Sub

Sub ShowProjectSummary()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Summary")
    ws.Activate
    MsgBox "프로젝트 요약 정보를 확인하세요.", vbInformation
End Sub

Sub AutoFitColumns()
    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        ws.Columns.AutoFit
    Next ws
    MsgBox "모든 열 너비가 자동 조정되었습니다.", vbInformation
End Sub
"""
    
    # VBA 코드를 시트에 텍스트로 추가
    ws_vba["A1"] = "VBA 코드 사용 안내"
    ws_vba["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_vba["A1"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_vba["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_vba["A3"] = "아래 코드를 복사하여 VBA 에디터(Alt+F11)에서 사용하세요:"
    ws_vba["A3"].font = Font(bold=True, size=11)
    
    # VBA 코드를 여러 셀에 나누어 작성
    vba_lines = vba_code.split('\n')
    for i, line in enumerate(vba_lines, start=5):
        ws_vba[f"A{i}"].value = line
        ws_vba[f"A{i}"].font = Font(name="Courier New", size=9)
        ws_vba[f"A{i}"].alignment = Alignment(vertical="top", wrap_text=True)
    
    # 열 너비 조정
    ws_vba.column_dimensions["A"].width = 100
    
    print(f"📝 VBA 코드 시트 추가 완료")
    
    output_path = os.path.join(script_dir, "AGI_TR_Submission_Gantt.xlsx")
    wb.save(output_path)
    print(f"✅ Generated: {output_path}")