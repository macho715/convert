#!/usr/bin/env python3
"""
AGI TR 7-Voyage Master Gantt with VBA
기존 AGI_TR_7Voyage_Master_Gantt.xlsx와 동일한 레이아웃 + VBA 기능
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# === COLORS (기존과 동일) ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
}

DEFAULT_SCENARIO_PATTERN = "1-2-2-2"
DEFAULT_EARLY_RETURN = True
DEFAULT_LCT_MAINT_START = dt.date(2026, 3, 1)
DEFAULT_LCT_MAINT_DAYS = 7
EARLY_RETURN_OFFSETS = {
    "RET-119": (15, 17),
    "RET-138": (34, 36),
    "RET-158": (41, 43),
}

BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def load_tasks_from_tsv(tsv_path, project_start_date):
    """
    TSV 파일에서 작업 목록을 읽어서 tasks 리스트로 변환
    
    Args:
        tsv_path: TSV 파일 경로
        project_start_date: 프로젝트 시작일 (datetime.date)
    
    Returns:
        tasks 리스트: (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    """
    import csv
    from datetime import datetime
    
    tasks = []
    
    # Phase 매핑 (TSV Phase -> Python Phase)
    phase_mapping = {
        "Mobilization": "MOBILIZATION",
        "Deck Prep": "DECK_PREP",
        "MZP Loadout": "LOADOUT",
        "Sea Fastening": "SEAFAST",
        "Survey": "BUFFER",  # MWS는 BUFFER로 매핑
        "Sea Passage": "SAIL",
        "AGI Arrival": "AGI_UNLOAD",
        "AGI Laydown": "BUFFER",
        "Onshore SPMT": "TURNING",
        "AGI Gate Prep": "TURNING",
        "Jackdown": "JACKDOWN",
        "Return": "RETURN",
        "Buffer": "BUFFER",
        "Marine Transport": "MILESTONE",
        "Demobilization": "MOBILIZATION",
        "Handover": "MILESTONE",
    }
    
    # Duration 매핑 (Duration_days -> Duration_Ref)
    def get_duration_ref(duration_str, task_name):
        try:
            dur_val = float(duration_str)
            if dur_val == 0:
                return 0
            elif dur_val == 0.5:
                return "DUR_BUF"
            elif dur_val == 1.0:
                if "Loadout" in task_name or "Load-out" in task_name:
                    return "DUR_LO"
                elif "Mobilization" in task_name or "Demobilization" in task_name:
                    return "DUR_MOB"
                elif "Sail" in task_name or "Sail-away" in task_name:
                    return "DUR_SAIL"
                elif "Arrive" in task_name or "Unload" in task_name:
                    return "DUR_UL"
                elif "Return" in task_name:
                    return "DUR_RET"
                elif "Jackdown" in task_name or "JACKDOWN" in task_name:
                    return "DUR_JD"
                else:
                    return "DUR_BUF"
            elif dur_val == 2.0:
                return 2
            elif dur_val == 3.0:
                if "Turn" in task_name or "Turning" in task_name:
                    return "DUR_TURN"
                elif "Deck" in task_name or "Prep" in task_name:
                    return "DUR_DECK"
                else:
                    return 3
            else:
                return dur_val
        except:
            return "DUR_BUF"
    
    with open(tsv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            if not row.get('ID') or not row.get('Start'):
                continue
            
            # 날짜 파싱
            try:
                start_date = datetime.strptime(row['Start'].strip(), '%Y-%m-%d').date()
            except:
                continue
            
            # Offset 계산 (프로젝트 시작일로부터의 일수)
            offset = (start_date - project_start_date).days
            
            # Duration 처리
            duration_str = row.get('Duration_days', '0').strip()
            dur_ref = get_duration_ref(duration_str, row.get('Task', ''))
            
            # Phase 매핑
            phase_raw = row.get('Phase', '').strip()
            phase = phase_mapping.get(phase_raw, 'BUFFER')
            
            # WBS 정규화 (A0, A1 등 -> 1.0, 2.0 등)
            wbs_raw = row.get('WBS', '').strip()
            if wbs_raw.startswith('A'):
                try:
                    wbs_num = float(wbs_raw[1:]) if wbs_raw[1:] else 0.0
                    wbs = f"{wbs_num:.1f}"
                except:
                    wbs = wbs_raw
            else:
                wbs = wbs_raw
            
            # Owner 정리
            owner = row.get('Owner', 'All').strip()
            if not owner:
                owner = 'All'
            
            # Notes
            notes = row.get('Notes', '').strip()
            
            task = (
                row['ID'].strip(),
                wbs,
                row['Task'].strip(),
                phase,
                owner,
                offset,
                dur_ref,
                notes
            )
            tasks.append(task)
    
    return tasks

def load_tide_data_json(json_path):
    """
    조석 데이터 JSON 파일 로드
    """
    import json
    
    tide_records = []
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if 'tide_records' in data:
                for record in data['tide_records']:
                    date_str = record.get('date', '')
                    if not date_str:
                        continue
                    tide_records.append((
                        date_str,
                        record.get('high_tide_window', '').strip(),
                        float(record.get('max_height_m', 0)) if record.get('max_height_m') else 0.0,
                        record.get('risk_level', 'LOW').strip()
                    ))
    except Exception as e:
        print(f"Warning: Could not load tide data from JSON: {e}")
    
    return tide_records

def load_tide_data(tsv_path=None, json_path=None):
    """
    조석 데이터 로드 (TSV 또는 JSON 지원)
    """
    import csv
    import os
    
    if json_path and os.path.exists(json_path):
        return load_tide_data_json(json_path)
    
    tide_records = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            with open(tsv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    if not row.get('Date'):
                        continue
                    try:
                        date_str = row['Date'].strip()
                        tide_records.append((
                            date_str,
                            row.get('High Tide Window', '').strip(),
                            float(row.get('Max Height (m)', '0').strip()) if row.get('Max Height (m)') else 0.0,
                            row.get('Risk Level', 'LOW').strip()
                        ))
                    except:
                        continue
        except Exception as e:
            print(f"Warning: Could not load tide data from TSV: {e}")
    
    return tide_records

def calculate_max_days(tasks, project_start, wb=None, min_days=120, buffer_days=30):
    """
    작업 목록에서 최대 프로젝트 기간 계산
    
    Args:
        tasks: 작업 리스트
        project_start: 프로젝트 시작일
    
    Returns:
        최대 일수 (정수)
    """
    max_offset = 0
    for task in tasks:
        if len(task) >= 6:
            offset = task[5] if isinstance(task[5], (int, float)) else 0
            max_offset = max(max_offset, offset)
    
    if wb:
        try:
            ws_ctrl = wb["Control_Panel"]
            min_days_val = ws_ctrl["H8"].value
            buffer_days_val = ws_ctrl["H9"].value
            if isinstance(min_days_val, (int, float)):
                min_days = int(min_days_val)
            if isinstance(buffer_days_val, (int, float)):
                buffer_days = int(buffer_days_val)
            maint_start = ws_ctrl["H10"].value
            maint_days = ws_ctrl["H11"].value
            if hasattr(maint_start, "date"):
                maint_start = maint_start.date()
            if isinstance(maint_start, dt.date):
                maint_offset = (maint_start - project_start).days
                if isinstance(maint_days, (int, float)):
                    maint_offset += int(maint_days)
                max_offset = max(max_offset, maint_offset)
        except Exception:
            pass
    
    # Duration을 고려하여 여유 있게 계산
    return max(min_days, int(max_offset) + buffer_days)

def calculate_voyage_ranges(tasks):
    """
    tasks에서 각 Voyage의 Day 범위를 계산
    
    Returns:
        voyage_ranges: [(voyage_name, start_day, end_day), ...]
    """
    voyage_ranges = []
    current_voyage = None
    voyage_start = None
    voyage_end = None
    
    for task in tasks:
        if len(task) < 6:
            continue
        
        tid, wbs, task_name, phase, owner, offset = task[:6]
        
        if tid.startswith('V') and len(tid) == 2:
            if current_voyage is not None:
                voyage_ranges.append((current_voyage, voyage_start, voyage_end))
            current_voyage = task_name
            voyage_start = offset
            voyage_end = offset
        
        if current_voyage is not None:
            voyage_end = max(voyage_end, offset)
    
    if current_voyage is not None:
        voyage_ranges.append((current_voyage, voyage_start, voyage_end))
    
    if tasks and len(tasks[0]) >= 6:
        mob_start = tasks[0][5] if isinstance(tasks[0][5], (int, float)) else 0
        mob_end = mob_start + 1
        voyage_ranges.insert(0, ("Mobilization", mob_start, mob_end))
    
    return voyage_ranges

def generate_weather_periods(project_start, project_end):
    """
    프로젝트 기간을 기반으로 Weather Analysis 헤더 구간 생성
    """
    periods = []
    if project_end <= project_start:
        period_name = f"{project_start.strftime('%b %d')}"
        return [(period_name, project_start, project_start)]
    current = project_start
    
    while current < project_end:
        period_start = current
        period_end = min(period_start + dt.timedelta(days=9), project_end)
        
        if period_start.month == period_end.month:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%d')}"
        else:
            period_name = f"{period_start.strftime('%b %d')}-{period_end.strftime('%b %d')}"
        
        periods.append((period_name, period_start, period_end))
        current = period_end + dt.timedelta(days=1)
    
    return periods

def create_scenario_sheets(wb, scenario_name, tsv_path, project_start):
    """
    시나리오별 Schedule_Data와 Gantt_Chart 시트 생성
    """
    import os
    
    ws_sched = wb.create_sheet(f"Schedule_Data_{scenario_name}")
    
    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = f"AGI TR Transportation - {scenario_name} Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_sched["A1"].alignment = Alignment(horizontal="center")
    
    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = f"Start = {project_start.isoformat()} | Auto-Updates from Control_Panel"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    
    try:
        ws_ctrl = wb["Control_Panel"]
        shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
        shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
        if hasattr(shamal_start, "date"):
            shamal_start = shamal_start.date()
        if hasattr(shamal_end, "date"):
            shamal_end = shamal_end.date()
        try:
            tide_threshold = float(ws_ctrl["H7"].value)
        except Exception:
            tide_threshold = 1.90
    except Exception:
        shamal_start = dt.date(2026, 1, 15)
        shamal_end = dt.date(2026, 4, 30)
        tide_threshold = 1.90
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"
    
    ws_sched.merge_cells("A3:I3")
    ws_sched["A3"] = f"⚠️ Winter Shamal Risk Period: {shamal_text_full} | Tide ≥{tide_threshold:.2f}m + Weather Gate"
    ws_sched["A3"].font = Font(size=10, italic=True)
    ws_sched["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    tasks = []
    if tsv_path and os.path.exists(tsv_path):
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
        except Exception as e:
            print(f"Error loading {scenario_name} TSV: {e}")
    
    lct_maint_offset = (DEFAULT_LCT_MAINT_START - project_start).days
    if tasks and not any(str(t[0]).startswith("LCT-MAINT") for t in tasks):
        tasks.append((
            "LCT-MAINT-001",
            "10.0",
            "LCT Bushra Maintenance",
            "BUFFER",
            "ADNOC L&S",
            lct_maint_offset,
            "LCT_MAINT_DAYS",
            "Scheduled maintenance window",
        ))

    if not tasks:
        return
    
    if not any(str(t[0]).startswith("LCT-MAINT") for t in tasks):
        tasks.append((
            "LCT-MAINT-001",
            "10.0",
            "LCT Bushra Maintenance",
            "BUFFER",
            "ADNOC L&S",
            lct_maint_offset,
            "LCT_MAINT_DAYS",
            "Scheduled maintenance window",
        ))

    for r, t in enumerate(tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_sched.cell(r, 1, value=tid)
        ws_sched.cell(r, 2, value=wbs)
        ws_sched.cell(r, 3, value=task)
        ws_sched.cell(r, 4, value=phase)
        ws_sched.cell(r, 5, value=owner)
        
        if str(tid).startswith("LCT-MAINT"):
            ws_sched.cell(r, 6, value="=LCT_MAINT_START")
        elif str(tid) in EARLY_RETURN_OFFSETS:
            early_offset, late_offset = EARLY_RETURN_OFFSETS[str(tid)]
            if isinstance(offset, (int, float)) and float(offset) == float(early_offset):
                ws_sched.cell(
                    r,
                    6,
                    value=f"=PROJECT_START+IF(EARLY_RETURN,{early_offset},{late_offset})",
                )
            else:
                ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        else:
            ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"
        
        if isinstance(dur_ref, str):
            ws_sched.cell(r, 8, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 8, value=dur_ref)
        
        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"
        ws_sched.cell(r, 9, value=notes)
        
        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 10):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        if phase == "MILESTONE":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")
    
    col_widths = {"A":10, "B":6, "C":38, "D":14, "E":14, "F":12, "G":12, "H":10, "I":40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    
    ws_gantt = wb.create_sheet(f"Gantt_Chart_{scenario_name}")
    
    max_days = calculate_max_days(tasks, project_start, wb=wb)
    
    ws_gantt.merge_cells(f"A1:{get_column_letter(7 + max_days)}1")
    ws_gantt["A1"] = f"AGI TR Transportation - {scenario_name} Gantt Chart"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_gantt["A1"].alignment = Alignment(horizontal="center")
    
    ws_gantt.merge_cells(f"A2:{get_column_letter(7 + max_days)}2")
    ws_gantt["A2"] = f"Orange Zone = Winter Shamal Risk Period ({shamal_text}) | Tide ≥{tide_threshold:.2f}m + Weather Gate | VBA: RefreshGanttChart_{scenario_name}로 색상 갱신"
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    month_str = project_start.strftime("%b %Y")
    ws_gantt["A3"] = month_str
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:G3")
    
    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    date_col = 8
    for i in range(max_days):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5
    
    for i in range(max_days):
        cell_date = project_start + dt.timedelta(days=i)
        if shamal_start <= cell_date <= shamal_end:
            ws_gantt.cell(4, date_col + i).fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    for r, t in enumerate(tasks, 5):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_gantt.cell(r, 1, value=f"=Schedule_Data_{scenario_name}!A{r+1}")
        ws_gantt.cell(r, 2, value=f"=Schedule_Data_{scenario_name}!B{r+1}")
        ws_gantt.cell(r, 3, value=f"=Schedule_Data_{scenario_name}!C{r+1}")
        ws_gantt.cell(r, 4, value=f"=Schedule_Data_{scenario_name}!D{r+1}")
        
        start_cell = ws_gantt.cell(r, 5, value=f"=Schedule_Data_{scenario_name}!F{r+1}")
        start_cell.number_format = "MM/DD"
        
        end_cell = ws_gantt.cell(r, 6, value=f"=Schedule_Data_{scenario_name}!G{r+1}")
        end_cell.number_format = "MM/DD"
        
        ws_gantt.cell(r, 7, value=f"=Schedule_Data_{scenario_name}!H{r+1}")
        
        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()
        
        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        start_date = project_start + dt.timedelta(days=offset)
        if isinstance(dur_ref, str):
            duration = 1
        else:
            try:
                duration = float(dur_ref)
            except:
                duration = 1
        if duration < 0:
            duration = 1
        
        end_date = start_date + dt.timedelta(days=duration)
        
        for i in range(max_days):
            cell_date = project_start + dt.timedelta(days=i)
            cell = ws_gantt.cell(r, date_col + i)
            cell.border = tb()
            
            if start_date <= cell_date < end_date:
                cell.fill = PatternFill("solid", fgColor=pc)
            elif cell_date == start_date and duration == 0:
                cell.fill = PatternFill("solid", fgColor=pc)
                cell.value = "★"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=8)
    
    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 28
    ws_gantt.column_dimensions["D"].width = 12
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4
    
    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)

def create_tide_data_sheet(wb, tide_tsv_path=None, tide_json_path=None):
    """
    조석 데이터 시트 생성
    """
    import os
    
    ws_tide = wb.create_sheet("Tide_Data")
    try:
        tide_threshold = float(wb["Control_Panel"]["H7"].value)
    except Exception:
        tide_threshold = 1.90
    
    ws_tide.merge_cells("A1:D1")
    ws_tide["A1"] = "MINA ZAYED PORT - High Tide Data"
    ws_tide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_tide["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_tide["A1"].alignment = Alignment(horizontal="center")
    
    ws_tide.merge_cells("A2:D2")
    ws_tide["A2"] = f"Tide ≥{tide_threshold:.2f}m required for Load-out and AGI Arrival | VBA: RefreshTideData로 업데이트"
    ws_tide["A2"].font = Font(size=10, italic=True)
    ws_tide["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    headers = ["Date", "High Tide Window", "Max Height (m)", "Risk Level"]
    for col, h in enumerate(headers, 1):
        cell = ws_tide.cell(4, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    tide_records = []
    if tide_json_path and os.path.exists(tide_json_path):
        tide_records = load_tide_data(json_path=tide_json_path)
        print(f"✅ Loaded {len(tide_records)} tide records from JSON")
    elif tide_tsv_path and os.path.exists(tide_tsv_path):
        tide_records = load_tide_data(tsv_path=tide_tsv_path)
        print(f"✅ Loaded {len(tide_records)} tide records from TSV")
    
    for r, (date_str, window, height, risk) in enumerate(tide_records, 5):
        ws_tide.cell(r, 1, value=date_str)
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            ws_tide.cell(r, 1, value=date_obj)
            ws_tide.cell(r, 1).number_format = "YYYY-MM-DD"
        except:
            ws_tide.cell(r, 1, value=date_str)
        
        ws_tide.cell(r, 2, value=window)
        ws_tide.cell(r, 3, value=height)
        ws_tide.cell(r, 3).number_format = "0.00"
        ws_tide.cell(r, 4, value=risk)
        
        for c in range(1, 5):
            ws_tide.cell(r, c).border = tb()
        
        if risk == "HIGH":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFCDD2")
        elif risk == "MEDIUM":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="FFE0B2")
        elif risk == "LOW":
            ws_tide.cell(r, 4).fill = PatternFill("solid", fgColor="C8E6C9")
        
        if height >= tide_threshold:
            ws_tide.cell(r, 3).font = Font(bold=True, color="0066CC")
            ws_tide.cell(r, 1).fill = PatternFill("solid", fgColor="E3F2FD")
    
    ws_tide.column_dimensions["A"].width = 12
    ws_tide.column_dimensions["B"].width = 25
    ws_tide.column_dimensions["C"].width = 14
    ws_tide.column_dimensions["D"].width = 12
    
    ws_tide.freeze_panes = "A5"

def create_comparison_summary(wb, project_start):
    """
    두 시나리오 비교 Summary 시트 생성
    """
    ws_comp = wb.create_sheet("Scenario_Comparison")
    
    ws_comp.merge_cells("A1:F1")
    ws_comp["A1"] = "AGI TR Transportation - Scenario Comparison"
    ws_comp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_comp["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_comp["A1"].alignment = Alignment(horizontal="center")
    
    headers = ["Metric", "ScenarioA (1x1x1x1x1x1x1)", "ScenarioB (1-2-2-2)", "Difference", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_comp.cell(3, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    comparison_data = [
        ("Total Tasks", "=COUNTA(Schedule_Data_ScenarioA!A:A)-5", "=COUNTA(Schedule_Data_ScenarioB!A:A)-5", "=C4-B4", "Task count difference"),
        ("Total Voyages", "=COUNTIF(Schedule_Data_ScenarioA!A:A,\"V*\")", "=COUNTIF(Schedule_Data_ScenarioB!A:A,\"V*\")", "=C5-B5", ""),
        ("Project Duration (days)", "=MAX(Schedule_Data_ScenarioA!G:G)-PROJECT_START", "=MAX(Schedule_Data_ScenarioB!G:G)-PROJECT_START", "=C6-B6", "Days difference"),
        ("Total Jack-down Events", "=COUNTIF(Schedule_Data_ScenarioA!D:D,\"JACKDOWN\")", "=COUNTIF(Schedule_Data_ScenarioB!D:D,\"JACKDOWN\")", "=C7-B7", ""),
        ("Project End Date", "=MAX(Schedule_Data_ScenarioA!G:G)", "=MAX(Schedule_Data_ScenarioB!G:G)", "=C8-B8", "Date difference"),
    ]
    
    for r, (metric, val_a, val_b, diff, notes) in enumerate(comparison_data, 4):
        ws_comp.cell(r, 1, value=metric)
        ws_comp.cell(r, 1).font = Font(bold=True)
        
        ws_comp.cell(r, 2, value=val_a)
        ws_comp.cell(r, 3, value=val_b)
        ws_comp.cell(r, 4, value=diff)
        ws_comp.cell(r, 5, value=notes)
        
        for c in range(1, 6):
            ws_comp.cell(r, c).border = tb()
            if c in [2, 3, 4] and "Date" in metric:
                ws_comp.cell(r, c).number_format = "YYYY-MM-DD"
            elif c == 4 and "Date" not in metric:
                ws_comp.cell(r, c).number_format = "0"
    
    for col in ["A", "B", "C", "D", "E"]:
        ws_comp.column_dimensions[col].width = 25

def create_gantt_with_vba(tsv_path=None, scenario_a_tsv=None, scenario_b_tsv=None, tide_tsv=None, tide_json=None):
    import os
    
    wb = Workbook()
    
    # TSV 파일에서 프로젝트 시작일 자동 감지
    project_start = dt.date(2026, 1, 18)
    start_tsv = None
    if scenario_a_tsv and os.path.exists(scenario_a_tsv):
        start_tsv = scenario_a_tsv
    elif tsv_path and os.path.exists(tsv_path):
        start_tsv = tsv_path
    
    if start_tsv:
        try:
            import csv
            from datetime import datetime
            with open(start_tsv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    if row.get('ID') and row.get('Start'):
                        try:
                            first_date = datetime.strptime(row['Start'].strip(), '%Y-%m-%d').date()
                            # MOB-001 우선, 없으면 가장 이른 날짜 사용
                            if 'MOB' in row['ID']:
                                project_start = first_date
                                break
                            if first_date < project_start:
                                project_start = first_date
                        except:
                            pass
        except Exception as e:
            print(f"Warning: Could not read TSV file for start date: {e}")
            print("Using default project start date: 2026-01-18")

    lct_maint_start = DEFAULT_LCT_MAINT_START
    lct_maint_days = DEFAULT_LCT_MAINT_DAYS
    lct_maint_offset = (lct_maint_start - project_start).days
    
    # === CONTROL PANEL (새로 추가) ===
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"
    
    # Title
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "🎛️ AGI TR Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30
    
    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "📌 Changing the start date (B4) will automatically update all schedules. VBA macros must be enabled."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    
    # Input Section
    ws_ctrl["A4"] = "📅 Project Start Date:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = project_start
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)
    
    ws_ctrl["A5"] = "🎯 Target Completion Date:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = dt.date(2026, 2, 28)
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()

    ws_ctrl["A6"] = "Scenario Pattern:"
    ws_ctrl["A6"].font = Font(bold=True)
    ws_ctrl["B6"] = DEFAULT_SCENARIO_PATTERN
    ws_ctrl["B6"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B6"].border = tb()

    ws_ctrl["A7"] = "Early Return after 1st JD:"
    ws_ctrl["A7"].font = Font(bold=True)
    ws_ctrl["B7"] = DEFAULT_EARLY_RETURN
    ws_ctrl["B7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B7"].border = tb()
    
    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")
    wb.defined_names["SCENARIO_PATTERN"] = DefinedName("SCENARIO_PATTERN", attr_text="Control_Panel!$B$6")
    wb.defined_names["EARLY_RETURN"] = DefinedName("EARLY_RETURN", attr_text="Control_Panel!$B$7")
    
    # Duration Parameters
    ws_ctrl["D4"] = "⏱️ Task Duration (Days)"
    ws_ctrl["D4"].font = Font(bold=True, size=12)
    
    durations = [
        ("D5", "Mobilization:", "E5", 1.0, "DUR_MOB"),
        ("D6", "Deck Prep:", "E6", 3.0, "DUR_DECK"),
        ("D7", "Load-out:", "E7", 1.0, "DUR_LO"),
        ("D8", "Sea Fastening:", "E8", 0.5, "DUR_SF"),
        ("D9", "MWS Approval:", "E9", 0.5, "DUR_MWS"),
        ("D10", "Sailing:", "E10", 1.0, "DUR_SAIL"),
        ("D11", "AGI Unload:", "E11", 1.0, "DUR_UL"),
        ("D12", "Turning:", "E12", 3.0, "DUR_TURN"),
        ("D13", "Jack-down:", "E13", 1.0, "DUR_JD"),
        ("D14", "Return:", "E14", 1.0, "DUR_RET"),
        ("D15", "Buffer:", "E15", 0.5, "DUR_BUF"),
    ]
    
    for lc, lt, vc, v, name in durations:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = v
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
        ws_ctrl[vc].border = tb()
        ws_ctrl[vc].number_format = "0.0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Control_Panel!${vc}")
    
    # Weather Settings
    ws_ctrl["G4"] = "🌊 Weather Settings"
    ws_ctrl["G4"].font = Font(bold=True, size=12)
    ws_ctrl["G5"] = "Shamal Start:"
    ws_ctrl["H5"] = dt.date(2026, 1, 15)
    ws_ctrl["H5"].number_format = "YYYY-MM-DD"
    ws_ctrl["H5"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    ws_ctrl["G6"] = "Shamal End:"
    ws_ctrl["H6"] = dt.date(2026, 4, 30)
    ws_ctrl["H6"].number_format = "YYYY-MM-DD"
    ws_ctrl["H6"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])

    ws_ctrl["G7"] = "Tide Threshold (m):"
    ws_ctrl["G7"].font = Font(bold=True)
    ws_ctrl["H7"] = 1.90
    ws_ctrl["H7"].number_format = "0.00"
    ws_ctrl["H7"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H7"].border = tb()
    
    ws_ctrl["G8"] = "📊 Gantt Min Days:"
    ws_ctrl["G8"].font = Font(bold=True)
    ws_ctrl["H8"] = 120
    ws_ctrl["H8"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H8"].border = tb()
    ws_ctrl["H8"].number_format = "0"
    
    ws_ctrl["G9"] = "📊 Gantt Buffer Days:"
    ws_ctrl["G9"].font = Font(bold=True)
    ws_ctrl["H9"] = 30
    ws_ctrl["H9"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H9"].border = tb()
    ws_ctrl["H9"].number_format = "0"

    ws_ctrl["G10"] = "LCT Maintenance Start:"
    ws_ctrl["G10"].font = Font(bold=True)
    ws_ctrl["H10"] = DEFAULT_LCT_MAINT_START
    ws_ctrl["H10"].number_format = "YYYY-MM-DD"
    ws_ctrl["H10"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H10"].border = tb()

    ws_ctrl["G11"] = "LCT Maintenance Days:"
    ws_ctrl["G11"].font = Font(bold=True)
    ws_ctrl["H11"] = DEFAULT_LCT_MAINT_DAYS
    ws_ctrl["H11"].number_format = "0"
    ws_ctrl["H11"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["H11"].border = tb()

    wb.defined_names["SHAMAL_START"] = DefinedName("SHAMAL_START", attr_text="Control_Panel!$H$5")
    wb.defined_names["SHAMAL_END"] = DefinedName("SHAMAL_END", attr_text="Control_Panel!$H$6")
    wb.defined_names["TIDE_THRESHOLD"] = DefinedName("TIDE_THRESHOLD", attr_text="Control_Panel!$H$7")
    wb.defined_names["GANTT_MIN_DAYS"] = DefinedName("GANTT_MIN_DAYS", attr_text="Control_Panel!$H$8")
    wb.defined_names["GANTT_BUFFER_DAYS"] = DefinedName("GANTT_BUFFER_DAYS", attr_text="Control_Panel!$H$9")
    wb.defined_names["LCT_MAINT_START"] = DefinedName("LCT_MAINT_START", attr_text="Control_Panel!$H$10")
    wb.defined_names["LCT_MAINT_DAYS"] = DefinedName("LCT_MAINT_DAYS", attr_text="Control_Panel!$H$11")
    
    # Summary Section
    ws_ctrl["A8"] = "📊 Auto Calculation Summary"
    ws_ctrl["A8"].font = Font(bold=True, size=12)
    
    summary_items = [
        ("A9", "Estimated Completion:", "B9", "=MAX(Schedule_Data!G:G)"),
        ("A10", "Total Duration (Days):", "B10", "=B9-B4+1"),
        ("A11", "Status vs Target:", "B11", '=IF(B9<=B5,"✅ On Target","❌ Delayed")'),
        ("A12", "Remaining Days:", "B12", "=B5-B9"),
        ("A13", "Scenario Pattern:", "B13", "=SCENARIO_PATTERN"),
        ("A14", "Early Return (after 1st JD):", "B14", '=IF(EARLY_RETURN,"Yes","No")'),
    ]
    
    for lc, lt, vc, formula in summary_items:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if "MAX" in formula:
            ws_ctrl[vc].number_format = "YYYY-MM-DD"
    
    # VBA Button Info
    ws_ctrl["A15"] = "🔘 VBA Macros (Alt+F8)"
    ws_ctrl["A15"].font = Font(bold=True, size=12)
    
    buttons = [
        "▶ UpdateAllSchedules - Recalculate All Schedules",
        "??ApplyScenarioFromControlPanel - Apply Scenario Pattern",
        "▶ UpdateAllScenarios - Batch Update All Scenarios",
        "▶ RefreshGanttChart - Refresh Gantt Chart Colors",
        "▶ RefreshAllGanttCharts - Refresh All Gantt Charts",
        "▶ RefreshTideData - Highlight Tide Data",
        "▶ GenerateReport - Generate Status Report",
        "▶ ExportToPDF - Export to PDF",
        "▶ SimulateDelay - Simulate Delay",
        "▶ HighlightCritical - Highlight Critical Path",
        "▶ HighlightToday - Show Today's Date",
        "▶ CheckShamalRisk - Check Shamal Risk",
    ]
    for i, btn in enumerate(buttons, 16):
        ws_ctrl[f"A{i}"] = btn
        ws_ctrl[f"A{i}"].font = Font(size=10)
    
    # Column widths
    ws_ctrl.column_dimensions["A"].width = 20
    ws_ctrl.column_dimensions["B"].width = 15
    ws_ctrl.column_dimensions["D"].width = 16
    ws_ctrl.column_dimensions["E"].width = 10
    ws_ctrl.column_dimensions["G"].width = 14
    ws_ctrl.column_dimensions["H"].width = 12
    ws_ctrl.column_dimensions["I"].width = 10
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    try:
        tide_threshold = float(ws_ctrl["H7"].value)
    except Exception:
        tide_threshold = 1.90
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    shamal_text_full = f"{shamal_text}, {shamal_start.year}"
    
    # === SCHEDULE DATA (기존과 동일한 구조) ===
    ws_sched = wb.create_sheet("Schedule_Data")
    
    # Title rows (기존과 동일)
    ws_sched.merge_cells("A1:I1")
    ws_sched["A1"] = "AGI HVDC TR 1-7 Transportation Master Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    ws_sched.merge_cells("A2:I2")
    ws_sched["A2"] = "7 Voyages | 4 Jack-down Events | LCT BUSHRA | Mina Zayed ↔ AGI Site"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    
    ws_sched.merge_cells("A3:I3")
    ws_sched["A3"] = f"⚠️ Winter Shamal Risk Period: {shamal_text_full} | Tide ≥{tide_threshold:.2f}m + Weather Gate | Schedule Auto-Updates from Control_Panel"
    ws_sched["A3"].font = Font(size=10, italic=True)
    ws_sched["A3"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    # Headers (Row 5, 기존과 동일)
    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Task definitions - TSV 파일에서 로드 또는 기본값 사용
    # (ID, WBS, Task, Phase, Owner, Offset, Duration_Ref, Notes)
    # 프로젝트 시작: 2026-01-18 (Day 0)
    default_tasks = [
        # Mobilization
        ("MOB-001", "1.0", "MOBILIZATION", "MOBILIZATION", "Mammoet", 0, "DUR_MOB", "SPMT Assembly + Marine Equipment Mobilization"),
        ("PREP-001", "1.1", "Deck Preparations", "DECK_PREP", "Mammoet", 1, "DUR_DECK", "One-time setup for all voyages"),
        
        # Voyage 1: LO 01-18, SAIL 01-20, ARR 01-22
        ("V1", "2.0", "VOYAGE 1: TR1 Transport", "MILESTONE", "All", 0, 0, "✅ Tide ≥1.90m (2.05m) | Good Weather Window"),
        ("LO-101", "2.1", "TR1 Load-out on LCT", "LOADOUT", "Mammoet", 0, "DUR_LO", "Tide ≥1.90m (2.05m) required"),
        ("SF-102", "2.2", "TR1 Sea Fastening", "SEAFAST", "Mammoet", 0, "DUR_SF", "12-point lashing"),
        ("MWS-103", "2.3", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 0, "DUR_MWS", "Marine Warranty Surveyor"),
        ("SAIL-104", "2.4", "V1 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 2, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-105", "2.5", "AGI Arrival + TR1 RORO Unload", "AGI_UNLOAD", "Mammoet", 4, "DUR_UL", "Tide ≥1.90m (1.91m) | AGI FWD Draft ≤ 2.70m"),
        ("STORE-106", "2.6", "TR1 Stored on AGI Laydown", "BUFFER", "Mammoet", 4, "DUR_BUF", "Awaiting pair TR2"),
        ("RET-107", "2.7", "V1 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 4, "DUR_RET", "Quick turnaround"),
        ("BUF-108", "2.99", "V1 Buffer / Equipment Reset", "BUFFER", "All", 5, "DUR_BUF", "Weather contingency"),
        
        # Voyage 2: LO 01-26, SAIL 01-27, ARR 01-29
        ("V2", "3.0", "VOYAGE 2: TR2 Transport + JD-1", "MILESTONE", "All", 8, 0, "✅ Tide ≥1.90m (1.91m) | Good Weather Window (before Shamal)"),
        ("LO-109", "3.1", "TR2 Load-out on LCT", "LOADOUT", "Mammoet", 8, "DUR_LO", "Tide ≥1.90m (1.91m) required"),
        ("SF-110", "3.2", "TR2 Sea Fastening", "SEAFAST", "Mammoet", 8, "DUR_SF", "12-point lashing"),
        ("MWS-110A", "3.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 8, "DUR_MWS", "Pre-sail verification"),
        ("SAIL-111", "3.3", "V2 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 9, "DUR_SAIL", "✅ Good Weather Window"),
        ("ARR-112", "3.4", "AGI Arrival + TR2 RORO Unload", "AGI_UNLOAD", "Mammoet", 11, "DUR_UL", "Tide ≥1.90m (2.03m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-113", "3.5", "TR1 Transport to Bay-1", "TURNING", "Mammoet", 12, 1, "Steel bridge install"),
        ("TURN-114", "3.6", "TR1 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", "10t Forklift required"),
        ("TRN-116", "3.8", "TR2 Transport to Bay-2", "TURNING", "Mammoet", 12, 1, ""),
        ("TURN-117", "3.9", "TR2 Turning (90° rotation)", "TURNING", "Mammoet", 12, "DUR_TURN", ""),
        ("JD-120A", "3.95", "JD-1 Jack-Down TR1", "JACKDOWN", "Mammoet", 14, "DUR_JD", "MILESTONE: TR1 complete | 02-01"),
        ("RET-119", "3.11", "V2 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 15, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-120B", "3.96", "JD-1 Jack-Down TR2", "JACKDOWN", "Mammoet", 16, "DUR_JD", "MILESTONE: TR2 complete | 02-02"),
        ("BUF-120", "3.99", "V2 Buffer / Shamal Recovery", "BUFFER", "All", 17, "DUR_BUF", "Post-Shamal weather check"),
        
        # Voyage 3: LO 01-31, SAIL 02-02, ARR 02-03
        ("V3", "4.0", "VOYAGE 3: TR3 Transport", "MILESTONE", "All", 13, 0, "✅ Tide ≥1.90m (2.07m) | Post-Shamal Window"),
        ("LO-121", "4.1", "TR3 Load-out on LCT", "LOADOUT", "Mammoet", 13, "DUR_LO", "Tide ≥1.90m (2.07m)"),
        ("SF-122", "4.2", "TR3 Sea Fastening", "SEAFAST", "Mammoet", 13, "DUR_SF", ""),
        ("MWS-122A", "4.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 13, "DUR_MWS", ""),
        ("SAIL-123", "4.3", "V3 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 15, "DUR_SAIL", "Good weather"),
        ("ARR-124", "4.4", "AGI Arrival + TR3 RORO Unload", "AGI_UNLOAD", "Mammoet", 16, "DUR_UL", "Tide ≥1.90m (2.04m)"),
        ("STORE-125", "4.5", "TR3 Stored on AGI Laydown", "BUFFER", "Mammoet", 16, "DUR_BUF", "Awaiting pair TR4"),
        ("RET-126", "4.6", "V3 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 17, "DUR_RET", ""),
        ("BUF-127", "4.99", "V3 Buffer", "BUFFER", "All", 17, "DUR_BUF", ""),
        
        # Voyage 4: LO 02-15, SAIL 02-16, ARR 02-18
        ("V4", "5.0", "VOYAGE 4: TR4 Transport + JD-2", "MILESTONE", "All", 28, 0, "✅ Tide ≥1.90m (1.90m) | Shamal 종료 직후"),
        ("LO-128", "5.1", "TR4 Load-out on LCT", "LOADOUT", "Mammoet", 28, "DUR_LO", "Tide ≥1.90m (1.90m)"),
        ("SF-129", "5.2", "TR4 Sea Fastening", "SEAFAST", "Mammoet", 28, "DUR_SF", ""),
        ("MWS-129A", "5.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 28, "DUR_MWS", ""),
        ("SAIL-130", "5.3", "V4 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 29, "DUR_SAIL", ""),
        ("ARR-131", "5.4", "AGI Arrival + TR4 RORO Unload", "AGI_UNLOAD", "Mammoet", 31, "DUR_UL", "Tide ≥1.90m (1.96m)"),
        ("TRN-132", "5.5", "TR3 Transport to Bay-3", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-133", "5.6", "TR3 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("TRN-135", "5.8", "TR4 Transport to Bay-4", "TURNING", "Mammoet", 31, 1, ""),
        ("TURN-136", "5.9", "TR4 Turning (90° rotation)", "TURNING", "Mammoet", 31, "DUR_TURN", ""),
        ("JD-139A", "5.95", "JD-2 Jack-Down TR3", "JACKDOWN", "Mammoet", 33, "DUR_JD", "MILESTONE: TR3 complete | 02-20"),
        ("RET-138", "5.11", "V4 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 34, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-139B", "5.96", "JD-2 Jack-Down TR4", "JACKDOWN", "Mammoet", 35, "DUR_JD", "MILESTONE: TR4 complete | 02-21"),
        ("BUF-140", "5.99", "V4 Buffer", "BUFFER", "All", 36, "DUR_BUF", ""),
        
        # Voyage 5: LO 02-23, SAIL 02-23, ARR 02-24 (Fast-turn)
        ("V5", "6.0", "VOYAGE 5: TR5 Transport", "MILESTONE", "All", 36, 0, "✅ Tide ≥1.90m (1.99m) | Fast-turn"),
        ("LO-140", "6.1", "TR5 Load-out on LCT", "LOADOUT", "Mammoet", 36, "DUR_LO", "Tide ≥1.90m (1.99m)"),
        ("SF-141", "6.2", "TR5 Sea Fastening", "SEAFAST", "Mammoet", 36, "DUR_SF", ""),
        ("MWS-141A", "6.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 36, "DUR_MWS", ""),
        ("SAIL-142", "6.3", "V5 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 36, "DUR_SAIL", "Fast-turn"),
        ("ARR-143", "6.4", "AGI Arrival + TR5 RORO Unload", "AGI_UNLOAD", "Mammoet", 37, "DUR_UL", "Tide ≥1.90m (2.01m)"),
        ("STORE-144", "6.5", "TR5 Stored on AGI Laydown", "BUFFER", "Mammoet", 37, "DUR_BUF", "Awaiting pair TR6"),
        ("RET-145", "6.6", "V5 LCT Return: AGI→MZP", "RETURN", "LCT Bushra", 37, "DUR_RET", ""),
        ("BUF-146", "6.99", "V5 Buffer", "BUFFER", "All", 37, "DUR_BUF", ""),
        
        # Voyage 6: LO 02-25, SAIL 02-25, ARR 02-26 (Fast-turn)
        ("V6", "7.0", "VOYAGE 6: TR6 Transport + JD-3", "MILESTONE", "All", 38, 0, "✅ Tide ≥1.90m (2.01m) | Fast-turn"),
        ("LO-147", "7.1", "TR6 Load-out on LCT", "LOADOUT", "Mammoet", 38, "DUR_LO", "Tide ≥1.90m (2.01m)"),
        ("SF-148", "7.2", "TR6 Sea Fastening", "SEAFAST", "Mammoet", 38, "DUR_SF", ""),
        ("MWS-148A", "7.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 38, "DUR_MWS", ""),
        ("SAIL-149", "7.3", "V6 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 38, "DUR_SAIL", "Fast-turn"),
        ("ARR-150", "7.4", "AGI Arrival + TR6 RORO Unload", "AGI_UNLOAD", "Mammoet", 39, "DUR_UL", "Tide ≥1.90m (1.98m)"),
        ("TRN-151", "7.5", "TR5 Transport to Bay-5", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-152", "7.6", "TR5 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("TRN-154", "7.8", "TR6 Transport to Bay-6", "TURNING", "Mammoet", 39, 1, ""),
        ("TURN-155", "7.9", "TR6 Turning (90° rotation)", "TURNING", "Mammoet", 39, "DUR_TURN", ""),
        ("JD-157A", "7.95", "JD-3 Jack-Down TR5", "JACKDOWN", "Mammoet", 40, "DUR_JD", "MILESTONE: TR5 complete | 02-27"),
        ("RET-158", "7.11", "V6 LCT Return: AGI->MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Return after first JD (SPMT reuse)"),
        ("JD-157B", "7.96", "JD-3 Jack-Down TR6", "JACKDOWN", "Mammoet", 42, "DUR_JD", "MILESTONE: TR6 complete | 02-28"),
        ("BUF-159", "7.99", "V6 Buffer / Reset for V7", "BUFFER", "All", 43, "DUR_BUF", ""),
        
        # Voyage 7: LO 02-27, SAIL 02-27, ARR 02-28 (Final)
        ("V7", "8.0", "VOYAGE 7: TR7 Transport + JD-4", "MILESTONE", "All", 40, 0, "✅ Tide ≥1.90m (1.92m) | Final unit"),
        ("LO-201", "8.1", "TR7 Load-out on LCT", "LOADOUT", "Mammoet", 40, "DUR_LO", "Tide ≥1.90m (1.92m) required"),
        ("SF-202", "8.2", "TR7 Sea Fastening", "SEAFAST", "Mammoet", 40, "DUR_SF", "12-point lashing"),
        ("MWS-202A", "8.25", "MWS + MPI + Final Check", "BUFFER", "Aries/Captain", 40, "DUR_MWS", ""),
        ("SAIL-203", "8.3", "V7 Sail-away: MZP→AGI", "SAIL", "LCT Bushra", 40, "DUR_SAIL", "Weather window required"),
        ("ARR-204", "8.4", "AGI Arrival + TR7 RORO Unload", "AGI_UNLOAD", "Mammoet", 41, "DUR_UL", "Tide ≥1.90m (1.93m) | AGI FWD Draft ≤ 2.70m"),
        ("TRN-205", "8.5", "TR7 Transport to Bay-7", "TURNING", "Mammoet", 41, 1, "Steel bridge install"),
        ("TURN-206", "8.6", "TR7 Turning (90° rotation)", "TURNING", "Mammoet", 41, "DUR_TURN", "10t Forklift required"),
        ("JD-207", "8.7", "★ JD-4 Jack-Down (TR7)", "JACKDOWN", "Mammoet", 41, "DUR_JD", "MILESTONE: TR7 Complete | 02-28"),
        ("RET-208", "8.8", "V7 LCT Final Return: AGI→MZP", "RETURN", "LCT Bushra", 41, "DUR_RET", "Final return"),
        
        # Demobilization
        ("DEMOB", "9.0", "DEMOBILIZATION", "MOBILIZATION", "Mammoet", 42, "DUR_MOB", "Equipment return"),
        ("LCT-MAINT-001", "10.0", "LCT Bushra Maintenance", "BUFFER", "ADNOC L&S", lct_maint_offset, "LCT_MAINT_DAYS", "Scheduled maintenance window"),
        ("END", "99.0", "★★★ PROJECT COMPLETE ★★★", "MILESTONE", "All", 42, 0, "All 7 TRs Installed | Jan-Feb 2026 Complete"),
    ]

    tasks = default_tasks
    if tsv_path:
        try:
            tasks = load_tasks_from_tsv(tsv_path, project_start)
            print(f"✅ Loaded {len(tasks)} tasks from TSV file")
        except Exception as e:
            print(f"Error loading TSV: {e}")
            print("Using default tasks list")
    
    for r, t in enumerate(tasks, 6):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_sched.cell(r, 1, value=tid)
        ws_sched.cell(r, 2, value=wbs)
        ws_sched.cell(r, 3, value=task)
        ws_sched.cell(r, 4, value=phase)
        ws_sched.cell(r, 5, value=owner)
        
        # Start = PROJECT_START + Offset
        if str(tid).startswith("LCT-MAINT"):
            ws_sched.cell(r, 6, value="=LCT_MAINT_START")
        elif str(tid) in EARLY_RETURN_OFFSETS:
            early_offset, late_offset = EARLY_RETURN_OFFSETS[str(tid)]
            if isinstance(offset, (int, float)) and float(offset) == float(early_offset):
                ws_sched.cell(
                    r,
                    6,
                    value=f"=PROJECT_START+IF(EARLY_RETURN,{early_offset},{late_offset})",
                )
            else:
                ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        else:
            ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"
        
        # Duration
        if isinstance(dur_ref, str):
            ws_sched.cell(r, 8, value=f"={dur_ref}")
        else:
            ws_sched.cell(r, 8, value=dur_ref)
        
        # End = Start + Duration
        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"
        
        ws_sched.cell(r, 9, value=notes)
        
        # Styling
        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 10):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        if phase == "MILESTONE":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True)
        if phase == "JACKDOWN":
            for c in range(1, 10):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")
    
    # Column widths
    col_widths = {"A":10, "B":6, "C":38, "D":14, "E":14, "F":12, "G":12, "H":10, "I":40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    
    # === GANTT CHART (기존과 동일한 구조) ===
    ws_gantt = wb.create_sheet("Gantt_Chart")
    max_days = calculate_max_days(tasks, project_start, wb=wb)
    
    # Title rows
    ws_gantt.merge_cells(f"A1:{get_column_letter(7 + max_days)}1")
    ws_gantt["A1"] = "AGI HVDC TR 1-7 Master Gantt Chart (Auto-Updated)"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    ws_gantt.merge_cells(f"A2:{get_column_letter(7 + max_days)}2")
    ws_gantt["A2"] = f"Orange Zone = Winter Shamal Risk Period ({shamal_text}) | Tide ≥{tide_threshold:.2f}m + Weather Gate | VBA: RefreshGanttChart로 색상 갱신"
    ws_gantt["A2"].font = Font(size=10, italic=True)
    ws_gantt["A2"].fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    # Month header
    month_str = project_start.strftime("%b %Y")
    ws_gantt["A3"] = month_str
    ws_gantt["A3"].font = Font(bold=True)
    ws_gantt.merge_cells("A3:G3")
    
    # Column headers (Row 4)
    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Date columns (dynamic days from project start to cover all tasks)
    date_col = 8
    for i in range(max_days):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5
        
        cell_date = project_start + dt.timedelta(days=i)
        if shamal_start <= cell_date <= shamal_end:
            c.fill = PatternFill("solid", fgColor=COLORS["SHAMAL"])
    
    # Task rows
    for r, t in enumerate(tasks, 5):
        tid, wbs, task, phase, owner, offset, dur_ref, notes = t
        
        ws_gantt.cell(r, 1, value=f"=Schedule_Data!A{r+1}")
        ws_gantt.cell(r, 2, value=f"=Schedule_Data!B{r+1}")
        ws_gantt.cell(r, 3, value=f"=Schedule_Data!C{r+1}")
        ws_gantt.cell(r, 4, value=f"=Schedule_Data!D{r+1}")
        
        start_cell = ws_gantt.cell(r, 5, value=f"=Schedule_Data!F{r+1}")
        start_cell.number_format = "MM/DD"
        
        end_cell = ws_gantt.cell(r, 6, value=f"=Schedule_Data!G{r+1}")
        end_cell.number_format = "MM/DD"
        
        ws_gantt.cell(r, 7, value=f"=Schedule_Data!H{r+1}")
        
        # Meta columns borders
        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()
        
        # Phase color in column D
        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        start_date = project_start + dt.timedelta(days=offset)
        if isinstance(dur_ref, str):
            duration = 1
        else:
            try:
                duration = float(dur_ref)
            except:
                duration = 1
        if duration < 0:
            duration = 1
        
        end_date = start_date + dt.timedelta(days=duration)
        
        # Date cells borders and base fills
        for i in range(max_days):
            cell_date = project_start + dt.timedelta(days=i)
            cell = ws_gantt.cell(r, date_col + i)
            cell.border = tb()
            
            if start_date <= cell_date < end_date:
                cell.fill = PatternFill("solid", fgColor=pc)
            elif cell_date == start_date and duration == 0:
                cell.fill = PatternFill("solid", fgColor=pc)
                cell.value = "★"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=8)
    
    # Column widths for meta columns
    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 28
    ws_gantt.column_dimensions["D"].width = 12
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4
    
    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)
    
    # === Scenario Sheets ===
    if scenario_a_tsv and os.path.exists(scenario_a_tsv):
        create_scenario_sheets(wb, "ScenarioA", scenario_a_tsv, project_start)
        print("✅ Created ScenarioA sheets")
    
    if scenario_b_tsv and os.path.exists(scenario_b_tsv):
        create_scenario_sheets(wb, "ScenarioB", scenario_b_tsv, project_start)
        print("✅ Created ScenarioB sheets")
    
    # === Tide Data Sheet ===
    if tide_json and os.path.exists(tide_json):
        create_tide_data_sheet(wb, tide_json_path=tide_json)
        print("✅ Created Tide_Data sheet (from JSON)")
    elif tide_tsv and os.path.exists(tide_tsv):
        create_tide_data_sheet(wb, tide_tsv_path=tide_tsv)
        print("✅ Created Tide_Data sheet (from TSV)")
    
    # === Scenario Comparison ===
    if (scenario_a_tsv and os.path.exists(scenario_a_tsv) and
            scenario_b_tsv and os.path.exists(scenario_b_tsv)):
        create_comparison_summary(wb, project_start)
        print("✅ Created Scenario_Comparison sheet")
    
    # === WEATHER ANALYSIS (동적화) ===
    ws_weather = wb.create_sheet("Weather_Analysis")
    
    try:
        project_end = ws_ctrl["B5"].value
        if hasattr(project_end, "date"):
            project_end = project_end.date()
        if not isinstance(project_end, dt.date):
            raise ValueError("Invalid target end date")
    except Exception:
        max_offset = 0
        for task in tasks:
            if len(task) >= 6:
                offset = task[5] if isinstance(task[5], (int, float)) else 0
                max_offset = max(max_offset, offset)
        project_end = project_start + dt.timedelta(days=int(max_offset) + 30)
    
    year_month = project_start.strftime("%b %Y")
    ws_weather["A1"] = f"UAE Winter Weather Analysis - {year_month}"
    ws_weather["A1"].font = Font(bold=True, size=14)
    
    weather_periods = generate_weather_periods(project_start, project_end)
    period_names = [p[0] for p in weather_periods]
    
    weather_headers = ["Parameter"] + period_names + ["Notes"]
    num_cols = len(weather_headers)
    ws_weather.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    
    for c, h in enumerate(weather_headers, 1):
        cell = ws_weather.cell(3, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.border = tb()
    
    shamal_start = ws_ctrl["H5"].value or dt.date(2026, 1, 15)
    shamal_end = ws_ctrl["H6"].value or dt.date(2026, 4, 30)
    if hasattr(shamal_start, "date"):
        shamal_start = shamal_start.date()
    if hasattr(shamal_end, "date"):
        shamal_end = shamal_end.date()
    if shamal_start.month == shamal_end.month:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%d')}"
    else:
        shamal_text = f"{shamal_start.strftime('%b %d')}-{shamal_end.strftime('%b %d')}"
    
    def get_weather_for_period(period_start, period_end, shamal_start, shamal_end):
        is_shamal = (period_start <= shamal_end and period_end >= shamal_start)
        if is_shamal:
            return {
                "wind": "16-21",
                "gust": "25-30",
                "wave": "0.8-1.2",
                "visibility": "2-5",
                "risk": "HIGH",
                "recommendation": "NO-GO",
            }
        return {
            "wind": "11-13",
            "gust": "18-20",
            "wave": "0.4-0.6",
            "visibility": "8-10",
            "risk": "LOW",
            "recommendation": "GO",
        }
    
    weather_data_rows = [
        ["Avg Wind (kt)"] + [""] * len(weather_periods) + [f"Peak Shamal: {shamal_text}"],
        ["Max Gust (kt)"] + [""] * len(weather_periods) + ["NO-GO if >22kt gust"],
        ["Wave Height (m)"] + [""] * len(weather_periods) + ["HOLD if >0.8m"],
        ["Visibility (km)"] + [""] * len(weather_periods) + ["Reduced during Shamal"],
        ["Risk Level"] + [""] * len(weather_periods) + [""],
        ["Recommendation"] + [""] * len(weather_periods) + [""],
    ]
    
    for idx, (_name, period_start, period_end) in enumerate(weather_periods):
        weather = get_weather_for_period(period_start, period_end, shamal_start, shamal_end)
        col = idx + 1
        weather_data_rows[0][col] = weather["wind"]
        weather_data_rows[1][col] = weather["gust"]
        weather_data_rows[2][col] = weather["wave"]
        weather_data_rows[3][col] = weather["visibility"]
        weather_data_rows[4][col] = weather["risk"]
        weather_data_rows[5][col] = weather["recommendation"]
    
    for r, row in enumerate(weather_data_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_weather.cell(r, c, value=val)
            cell.border = tb()
            if val == "HIGH" or val == "NO-GO":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif val == "MEDIUM" or val == "CAUTION":
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
            elif val == "LOW" or val == "GO":
                cell.fill = PatternFill("solid", fgColor="C8E6C9")
    
    for col in range(1, num_cols + 1):
        ws_weather.column_dimensions[get_column_letter(col)].width = 15
    
    # === SUMMARY (기존과 동일) ===
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "AGI HVDC Transformer Transportation - Project Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")
    
    summary_data = [
        ("Key Parameters", ""),
        ("Total Transformers", "7 units (TR1-TR7)"),
        ("Weight per TR", "217-271 tons"),
        ("Total Voyages", "7 (1 TR per voyage)"),
        ("Jack-down Events", "4 (after V2, V4, V6, V7)"),
        ("Vessel", "LCT BUSHRA"),
        ("Route", "Mina Zayed Port ↔ AGI Site"),
        ("Scenario Pattern", "=SCENARIO_PATTERN"),
        ("Early Return After 1st JD", "=IF(EARLY_RETURN,\"Yes\",\"No\")"),
        ("LCT Maintenance Start", "=LCT_MAINT_START"),
        ("LCT Maintenance Days", "=LCT_MAINT_DAYS"),
        ("", ""),
        ("Schedule Summary", ""),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
    ]
    
    voyage_ranges = calculate_voyage_ranges(tasks)
    for voyage_name, start_day, end_day in voyage_ranges:
        if "Mobilization" in voyage_name:
            label = voyage_name
        else:
            if ":" in voyage_name:
                prefix, rest = voyage_name.split(":", 1)
                label = f"{prefix.title()} ({rest.strip()})"
            else:
                label = voyage_name.title()
        summary_data.append((label, f"Day {start_day}-{end_day}"))
    
    summary_data.extend([
        ("Project Complete", "=MAX(Schedule_Data!G:G)"),
        ("", ""),
        ("Weather Constraints", ""),
        ("Shamal Period", shamal_text_full),
        ("Tide Requirement", f"≥{tide_threshold:.2f}m for LO/ARR"),
        ("Wind Limit", "≤18kt sustained, ≤22kt gust"),
        ("Wave Limit", "≤0.8m (HOLD), ≤1.0m (NO-GO)"),
        ("AGI Draft Limit", "≤2.70m forward draft"),
    ])
    
    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"
        if label == "LCT Maintenance Days":
            ws_summary.cell(r, 2).number_format = "0"
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30
    
    # === VBA CODE SHEET ===
    ws_vba = wb.create_sheet("VBA_Code")
    
    ws_vba["A1"] = "📋 VBA 코드 - Alt+F11 → Module에 붙여넣기 → .xlsm으로 저장"
    ws_vba["A1"].font = Font(bold=True, size=14)
    
    vba_code = '''
Option Explicit

' ============================================
' AGI TR Multi-Scenario Master Gantt - VBA Macros
' ============================================
' 사용법: Alt+F11 → Module 삽입 → 코드 붙여넣기
' ============================================

' === 통합 업데이트 함수 ===
Sub UpdateAllScenarios()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    On Error Resume Next
    Sheets("Schedule_Data_ScenarioA").Calculate
    Sheets("Gantt_Chart_ScenarioA").Calculate
    Sheets("Schedule_Data_ScenarioB").Calculate
    Sheets("Gantt_Chart_ScenarioB").Calculate
    Sheets("Tide_Data").Calculate
    Sheets("Scenario_Comparison").Calculate
    On Error GoTo 0
    
    Sheets("Schedule_Data").Calculate
    Sheets("Gantt_Chart").Calculate
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    Sheets("Weather_Analysis").Calculate
    
    Call RefreshAllGanttCharts
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "✅ 모든 시나리오 업데이트 완료!", vbInformation, "Update Complete"
End Sub

' === 모든 Gantt 차트 색상 갱신 ===
Sub RefreshAllGanttCharts()
    On Error Resume Next
    Call RefreshGanttChart_ScenarioA
    Call RefreshGanttChart_ScenarioB
    Call RefreshGanttChart
    On Error GoTo 0
End Sub

' === Apply Scenario Pattern from Control Panel ===
Sub ApplyScenarioFromControlPanel()
    Dim pattern As String
    Dim srcName As String
    Dim wsSrc As Worksheet, wsDst As Worksheet
    Dim lastRow As Long, lastCol As Long
    
    pattern = UCase(Trim(Sheets("Control_Panel").Range("B6").Value))
    If pattern = "" Then Exit Sub
    
    Select Case pattern
        Case "SCENARIOA", "1X1X1X1X1X1X1", "1-1-1-1-1-1-1"
            srcName = "Schedule_Data_ScenarioA"
        Case "SCENARIOB", "1-2-2-2", "1 2 2 2"
            srcName = "Schedule_Data_ScenarioB"
        Case Else
            Exit Sub
    End Select
    
    On Error Resume Next
    Set wsSrc = Sheets(srcName)
    Set wsDst = Sheets("Schedule_Data")
    On Error GoTo 0
    
    If wsSrc Is Nothing Or wsDst Is Nothing Then Exit Sub
    
    lastRow = wsSrc.Cells(wsSrc.Rows.Count, 1).End(xlUp).Row
    lastCol = 9
    
    wsDst.Range(wsDst.Cells(6, 1), wsDst.Cells(wsDst.Rows.Count, lastCol)).ClearContents
    wsDst.Range(wsDst.Cells(6, 1), wsDst.Cells(lastRow, lastCol)).Formula = _
        wsSrc.Range(wsSrc.Cells(6, 1), wsSrc.Cells(lastRow, lastCol)).Formula
    wsDst.Range(wsDst.Cells(6, 1), wsDst.Cells(lastRow, lastCol)).NumberFormat = _
        wsSrc.Range(wsSrc.Cells(6, 1), wsSrc.Cells(lastRow, lastCol)).NumberFormat
End Sub

' === ScenarioA Gantt 갱신 ===
Sub RefreshGanttChart_ScenarioA()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date
    Dim maintStart As Variant, maintDays As Variant, maintEnd As Date
    Dim maintStart As Variant, maintDays As Variant, maintEnd As Date
    Dim maintStart As Variant, maintDays As Variant, maintEnd As Date
    
    Set ws = Sheets("Gantt_Chart_ScenarioA")
    Set wsd = Sheets("Schedule_Data_ScenarioA")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    maintStart = Sheets("Control_Panel").Range("H10").Value
    maintDays = Sheets("Control_Panel").Range("H11").Value
    If IsDate(maintStart) And IsNumeric(maintDays) And maintDays > 0 Then
        maintEnd = maintStart + CLng(maintDays) - 1
    Else
        maintDays = 0
    End If
    dc = 8
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    Application.ScreenUpdating = False
    
    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone
    
    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
        End If
        If IsNumeric(maintDays) And maintDays > 0 And cellDate >= maintStart And cellDate <= maintEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 193, 7)
        End If
    Next j
    
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value
            
            ganttRow = i - 1
            
            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733)
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
    Next i
    
    Application.ScreenUpdating = True
End Sub

' === ScenarioB Gantt 갱신 ===
Sub RefreshGanttChart_ScenarioB()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date
    
    Set ws = Sheets("Gantt_Chart_ScenarioB")
    Set wsd = Sheets("Schedule_Data_ScenarioB")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    maintStart = Sheets("Control_Panel").Range("H10").Value
    maintDays = Sheets("Control_Panel").Range("H11").Value
    If IsDate(maintStart) And IsNumeric(maintDays) And maintDays > 0 Then
        maintEnd = maintStart + CLng(maintDays) - 1
    Else
        maintDays = 0
    End If
    dc = 8
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    Application.ScreenUpdating = False
    
    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone
    
    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121)
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0)
        End If
        If IsNumeric(maintDays) And maintDays > 0 And cellDate >= maintStart And cellDate <= maintEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 193, 7)
        End If
    Next j
    
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value
            
            ganttRow = i - 1
            
            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733)
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
    Next i
    
    Application.ScreenUpdating = True
End Sub

' === 조석 데이터 갱신 ===
Sub RefreshTideData()
    Dim ws As Worksheet
    Dim i As Long
    Dim tideThreshold As Double
    
    Set ws = Sheets("Tide_Data")
    tideThreshold = Sheets("Control_Panel").Range("H7").Value
    If tideThreshold = 0 Then tideThreshold = 1.9
    
    For i = 5 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
        If IsNumeric(ws.Cells(i, 3).Value) Then
            If ws.Cells(i, 3).Value >= tideThreshold Then
                ws.Cells(i, 3).Font.Bold = True
                ws.Cells(i, 3).Font.Color = RGB(0, 102, 204)
                ws.Cells(i, 1).Interior.Color = RGB(227, 242, 253)
            End If
        End If
    Next i
    
    MsgBox "✅ 조석 데이터 강조 완료 (Tide ≥" & Format(tideThreshold, "0.00") & "m)", vbInformation
End Sub

' === 1. 전체 일정 업데이트 ===
Sub UpdateAllSchedules()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    Call ApplyScenarioFromControlPanel
    
    Sheets("Schedule_Data").Calculate
    Sheets("Gantt_Chart").Calculate
    Sheets("Control_Panel").Calculate
    Sheets("Summary").Calculate
    
    Call RefreshGanttChart
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "✅ 일정 업데이트 완료!" & vbCrLf & vbCrLf & _
           "프로젝트 시작: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
           "예상 완료: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD"), _
           vbInformation, "Schedule Updated"
End Sub

' === 2. Gantt Chart 색상 갱신 ===
Sub RefreshGanttChart()
    Dim ws As Worksheet, wsd As Worksheet
    Dim i As Long, j As Long, lastRow As Long, ganttRow As Long
    Dim startD As Date, endD As Date, projStart As Date, cellDate As Date
    Dim phase As String, dc As Long, lastCol As Long, maxJ As Long
    Dim shamalStart As Date, shamalEnd As Date
    
    Set ws = Sheets("Gantt_Chart")
    Set wsd = Sheets("Schedule_Data")
    projStart = Sheets("Control_Panel").Range("B4").Value
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    maintStart = Sheets("Control_Panel").Range("H10").Value
    maintDays = Sheets("Control_Panel").Range("H11").Value
    If IsDate(maintStart) And IsNumeric(maintDays) And maintDays > 0 Then
        maintEnd = maintStart + CLng(maintDays) - 1
    Else
        maintDays = 0
    End If
    dc = 8 ' Date columns start at H
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    Application.ScreenUpdating = False
    
    ' Clear existing colors in date columns
    ws.Range(ws.Cells(5, dc), ws.Cells(lastRow, lastCol)).Interior.ColorIndex = xlNone
    
    ' Reset header colors + Shamal highlight
    For j = 0 To maxJ
        ws.Cells(4, dc + j).Interior.Color = RGB(31, 78, 121) ' HEADER color
        cellDate = projStart + j
        If cellDate >= shamalStart And cellDate <= shamalEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 152, 0) ' Orange
        End If
        If IsNumeric(maintDays) And maintDays > 0 And cellDate >= maintStart And cellDate <= maintEnd Then
            ws.Cells(4, dc + j).Interior.Color = RGB(255, 193, 7) ' Maintenance
        End If
    Next j
    
    ' Apply Gantt bars
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) And wsd.Cells(i, 6).Value <> "" Then
            startD = wsd.Cells(i, 6).Value
            If IsDate(wsd.Cells(i, 7).Value) Then
                endD = wsd.Cells(i, 7).Value
            Else
                endD = startD
            End If
            phase = wsd.Cells(i, 4).Value
            
            ganttRow = i - 1
            
            For j = 0 To maxJ
                cellDate = projStart + j
                If cellDate >= startD And cellDate < endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                ElseIf cellDate = startD And startD = endD Then
                    ws.Cells(ganttRow, dc + j).Interior.Color = GetPhaseColor(phase)
                    ws.Cells(ganttRow, dc + j).Value = Chr(9733) ' Star
                    ws.Cells(ganttRow, dc + j).HorizontalAlignment = xlCenter
                    ws.Cells(ganttRow, dc + j).Font.Size = 8
                End If
            Next j
        End If
    Next i
    
    ' Highlight today
    For j = 0 To maxJ
        cellDate = projStart + j
        If cellDate = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Borders(xlEdgeLeft).Color = RGB(255, 0, 0)
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Borders(xlEdgeLeft).Weight = xlThick
            Exit For
        End If
    Next j
    
    Application.ScreenUpdating = True
End Sub

' === Phase Color Helper ===
Function GetPhaseColor(phase As String) As Long
    Select Case phase
        Case "MOBILIZATION": GetPhaseColor = RGB(142, 124, 195)
        Case "DECK_PREP": GetPhaseColor = RGB(111, 168, 220)
        Case "LOADOUT": GetPhaseColor = RGB(147, 196, 125)
        Case "SEAFAST": GetPhaseColor = RGB(118, 165, 175)
        Case "SAIL": GetPhaseColor = RGB(164, 194, 244)
        Case "AGI_UNLOAD": GetPhaseColor = RGB(246, 178, 107)
        Case "TURNING": GetPhaseColor = RGB(255, 217, 102)
        Case "JACKDOWN": GetPhaseColor = RGB(224, 102, 102)
        Case "RETURN": GetPhaseColor = RGB(153, 153, 153)
        Case "BUFFER": GetPhaseColor = RGB(217, 217, 217)
        Case "MILESTONE": GetPhaseColor = RGB(255, 0, 0)
        Case Else: GetPhaseColor = RGB(255, 255, 255)
    End Select
End Function

' === 3. 프로젝트 리포트 생성 ===
Sub GenerateReport()
    Dim wsd As Worksheet
    Dim i As Long, total As Long, jdCount As Long, lastRow As Long
    Dim voyages As Long, milestones As Long
    
    Set wsd = Sheets("Schedule_Data")
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    For i = 6 To lastRow
        If wsd.Cells(i, 1).Value <> "" Then
            total = total + 1
            If wsd.Cells(i, 4).Value = "JACKDOWN" Then jdCount = jdCount + 1
            If wsd.Cells(i, 4).Value = "MILESTONE" Then milestones = milestones + 1
            If Left(wsd.Cells(i, 1).Value, 1) = "V" And Len(wsd.Cells(i, 1).Value) = 2 Then voyages = voyages + 1
        End If
    Next i
    
    Dim rpt As String
    rpt = "╔══════════════════════════════════════╗" & vbCrLf & _
          "║   AGI HVDC TR Transportation Report  ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ Report Date: " & Format(Now, "YYYY-MM-DD HH:MM") & "      ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ PROJECT STATUS                       ║" & vbCrLf & _
          "║  Total Tasks: " & total & "                      ║" & vbCrLf & _
          "║  Voyages: " & voyages & "                          ║" & vbCrLf & _
          "║  Jack-down Events: " & jdCount & "                 ║" & vbCrLf & _
          "║  Milestones: " & milestones & "                       ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ KEY DATES                            ║" & vbCrLf & _
          "║  Start: " & Format(Sheets("Control_Panel").Range("B4").Value, "YYYY-MM-DD") & "              ║" & vbCrLf & _
          "║  Target: " & Format(Sheets("Control_Panel").Range("B5").Value, "YYYY-MM-DD") & "             ║" & vbCrLf & _
          "║  Est.End: " & Format(Sheets("Control_Panel").Range("B9").Value, "YYYY-MM-DD") & "            ║" & vbCrLf & _
          "║  Status: " & Sheets("Control_Panel").Range("B11").Value & "               ║" & vbCrLf & _
          "╠══════════════════════════════════════╣" & vbCrLf & _
          "║ WEATHER RISK                         ║" & vbCrLf & _
          "║  Shamal: " & Format(Sheets("Control_Panel").Range("H5").Value, "MM/DD") & " - " & Format(Sheets("Control_Panel").Range("H6").Value, "MM/DD") & "           ║" & vbCrLf & _
          "╚══════════════════════════════════════╝"
    
    MsgBox rpt, vbInformation, "Project Report"
End Sub

' === 4. PDF 내보내기 ===
Sub ExportToPDF()
    Dim fp As String
    fp = ThisWorkbook.Path & "\\AGI_TR_Gantt_" & Format(Date, "YYYYMMDD") & ".pdf"
    
    Sheets(Array("Schedule_Data", "Gantt_Chart", "Summary")).Select
    ActiveSheet.ExportAsFixedFormat xlTypePDF, fp, xlQualityStandard, True
    Sheets("Control_Panel").Select
    
    MsgBox "✅ PDF 저장 완료:" & vbCrLf & fp, vbInformation, "Export Complete"
End Sub

' === 5. 지연 시뮬레이션 ===
Sub SimulateDelay()
    Dim delayDays As Integer, origStart As Date
    Dim wsCtrl As Worksheet
    
    Set wsCtrl = Sheets("Control_Panel")
    origStart = wsCtrl.Range("B4").Value
    
    delayDays = InputBox("시뮬레이션할 지연 일수를 입력하세요:" & vbCrLf & _
                         "(현재 시작일: " & Format(origStart, "YYYY-MM-DD") & ")", _
                         "Delay Simulation", "7")
    
    If IsNumeric(delayDays) And delayDays <> 0 Then
        wsCtrl.Range("B4").Value = origStart + delayDays
        Call UpdateAllSchedules
        
        MsgBox "시뮬레이션 결과:" & vbCrLf & _
               "새 시작일: " & Format(wsCtrl.Range("B4").Value, "YYYY-MM-DD") & vbCrLf & _
               "새 완료일: " & Format(wsCtrl.Range("B9").Value, "YYYY-MM-DD") & vbCrLf & _
               "목표 대비: " & wsCtrl.Range("B11").Value, vbInformation, "Simulation Result"
        
        If MsgBox("원래 일정으로 복원하시겠습니까?", vbYesNo + vbQuestion, "Restore?") = vbYes Then
            wsCtrl.Range("B4").Value = origStart
            Call UpdateAllSchedules
        End If
    End If
End Sub

' === 6. Critical Path 강조 ===
Sub HighlightCritical()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    
    Set wsd = Sheets("Schedule_Data")
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    ' Reset
    wsd.Range(wsd.Cells(6, 1), wsd.Cells(lastRow, 9)).Font.Bold = False
    wsd.Range(wsd.Cells(6, 1), wsd.Cells(lastRow, 9)).Font.Color = RGB(0, 0, 0)
    
    ' Highlight Jack-down and Milestones
    For i = 6 To lastRow
        If wsd.Cells(i, 4).Value = "JACKDOWN" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(183, 28, 28)
        ElseIf wsd.Cells(i, 4).Value = "MILESTONE" Then
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Bold = True
            wsd.Range(wsd.Cells(i, 1), wsd.Cells(i, 9)).Font.Color = RGB(21, 101, 192)
        End If
    Next i
    
    MsgBox "✅ Critical Path 강조 완료" & vbCrLf & _
           "🔴 빨강 = Jack-down (Critical)" & vbCrLf & _
           "🔵 파랑 = Milestone", vbInformation, "Critical Path"
End Sub

' === 7. 오늘 날짜 하이라이트 ===
Sub HighlightToday()
    Dim ws As Worksheet, j As Long, lastCol As Long, maxJ As Long, lastRow As Long
    Dim projStart As Date, dc As Long
    
    Set ws = Sheets("Gantt_Chart")
    projStart = Sheets("Control_Panel").Range("B4").Value
    dc = 8
    
    lastCol = ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
    maxJ = lastCol - dc
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    For j = 0 To maxJ
        If projStart + j = Date Then
            ws.Range(ws.Cells(4, dc + j), ws.Cells(lastRow, dc + j)).Interior.Color = RGB(255, 255, 200)
            ws.Cells(3, dc + j).Value = "TODAY"
            ws.Cells(3, dc + j).Font.Bold = True
            ws.Cells(3, dc + j).Font.Color = RGB(255, 0, 0)
            MsgBox "오늘 날짜 (" & Format(Date, "MM/DD") & ") 컬럼이 강조되었습니다.", vbInformation
            Exit For
        End If
    Next j
End Sub

' === 8. 날짜 변경 자동 트리거 (Control_Panel 시트에 추가) ===
' 아래 코드를 Control_Panel 시트의 코드 영역에 붙여넣으세요:
'
' Private Sub Worksheet_Change(ByVal Target As Range)
'     If Target.Address = "$B$4" Then
'         Call UpdateAllSchedules
'     End If
' End Sub

' === 9. 진행률 일괄 업데이트 ===
Sub BulkProgressUpdate()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    Dim pctValue As Double
    
    pctValue = InputBox("일괄 적용할 진행률을 입력하세요 (0-100):", "Bulk Progress", "50")
    
    If IsNumeric(pctValue) Then
        pctValue = pctValue / 100
        Set wsd = Sheets("Schedule_Data")
        lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
        
        ' Progress 컬럼이 없으면 추가
        If wsd.Cells(5, 10).Value <> "Progress" Then
            wsd.Cells(5, 10).Value = "Progress"
            wsd.Cells(5, 10).Font.Bold = True
            wsd.Cells(5, 10).Font.Color = RGB(255, 255, 255)
            wsd.Cells(5, 10).Fill.Color = RGB(31, 78, 121)
        End If
        
        For i = 6 To lastRow
            If wsd.Cells(i, 1).Value <> "" Then
                wsd.Cells(i, 10).Value = pctValue
                wsd.Cells(i, 10).NumberFormat = "0%"
            End If
        Next i
        
        MsgBox "진행률 " & Format(pctValue, "0%") & " 일괄 적용 완료", vbInformation
    End If
End Sub

' === 10. Shamal 위험 체크 ===
Sub CheckShamalRisk()
    Dim wsd As Worksheet, i As Long, lastRow As Long
    Dim taskDate As Date, shamalStart As Date, shamalEnd As Date
    Dim riskTasks As String, cnt As Long
    
    Set wsd = Sheets("Schedule_Data")
    shamalStart = Sheets("Control_Panel").Range("H5").Value
    shamalEnd = Sheets("Control_Panel").Range("H6").Value
    lastRow = wsd.Cells(wsd.Rows.Count, 1).End(xlUp).Row
    
    For i = 6 To lastRow
        If IsDate(wsd.Cells(i, 6).Value) Then
            taskDate = wsd.Cells(i, 6).Value
            If taskDate >= shamalStart And taskDate <= shamalEnd Then
                ' SAIL tasks are weather-critical
                If wsd.Cells(i, 4).Value = "SAIL" Or wsd.Cells(i, 4).Value = "LOADOUT" Then
                    cnt = cnt + 1
                    riskTasks = riskTasks & vbCrLf & "  ⚠️ " & wsd.Cells(i, 1).Value & ": " & wsd.Cells(i, 3).Value
                End If
            End If
        End If
    Next i
    
    If cnt > 0 Then
        MsgBox "⚠️ SHAMAL 위험 경고!" & vbCrLf & vbCrLf & _
               "Shamal 기간 (" & Format(shamalStart, "MM/DD") & "-" & Format(shamalEnd, "MM/DD") & ") 중 " & cnt & "개 기상 민감 작업 발견:" & vbCrLf & _
               riskTasks & vbCrLf & vbCrLf & _
               "일정 조정을 권장합니다.", vbExclamation, "Weather Risk Alert"
    Else
        MsgBox "✅ Shamal 기간 중 기상 민감 작업 없음" & vbCrLf & _
               "현재 일정은 안전합니다.", vbInformation, "Weather Check OK"
    End If
End Sub
'''
    
    for i, line in enumerate(vba_code.strip().split('\n'), 3):
        ws_vba.cell(i, 1, value=line)
        ws_vba.cell(i, 1).font = Font(name="Consolas", size=9)
    
    ws_vba.column_dimensions["A"].width = 100
    
    return wb

if __name__ == "__main__":
    import os
    import sys
    
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
    
    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    
    scenario_a_tsv = os.path.join(script_dir, "ScenarioA_1x1x1x1x1x1x1_start2026-01-15.tsv")
    scenario_b_tsv = os.path.join(script_dir, "ScenarioB_1-2-2-2_start2026-01-15.tsv")
    tide_tsv = os.path.join(script_dir, "Date High Tide Window Max Height (m) Ris.tsv")
    tide_json = os.path.join(script_dir, "MINA ZAYED PORT WATER TIDE_MERGED.json")
    
    default_tsv = os.path.join(script_dir, "ID WBS Task Phase Owner Start End Durati.tsv")
    tsv_path = default_tsv if os.path.exists(default_tsv) else None
    
    print("Generating AGI TR Multi-Scenario Master Gantt with VBA...")
    wb = create_gantt_with_vba(
        tsv_path=tsv_path,
        scenario_a_tsv=scenario_a_tsv if os.path.exists(scenario_a_tsv) else None,
        scenario_b_tsv=scenario_b_tsv if os.path.exists(scenario_b_tsv) else None,
        tide_tsv=tide_tsv if os.path.exists(tide_tsv) else None,
        tide_json=tide_json if os.path.exists(tide_json) else None
    )
    
    output_path = os.path.join(os.getcwd(), "AGI_TR_MultiScenario_Master_Gantt_VBA.xlsx")
    wb.save(output_path)
    print(f"[OK] Generated: {output_path}")
