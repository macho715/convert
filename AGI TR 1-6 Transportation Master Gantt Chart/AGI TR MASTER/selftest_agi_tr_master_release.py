#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Self-test for AGI TR Master workbook structure.

Validates that:
  - Required sheets exist
  - Schedule_Data has required headers
  - Offsets/durations are numeric
  - Gantt_Chart alignment matches Schedule_Data row numbers

Run:
  python selftest_agi_tr_master_release.py AGI_TR_Master_READY_1-2-2-1_JD3.xlsx
"""

from __future__ import annotations

import sys
import math
from pathlib import Path

from openpyxl import load_workbook


REQ_SHEETS = [
    "Control_Panel",
    "Schedule_Data",
    "Gantt_Chart",
    "Tide_Data",
]

HEADERS = [
    "ID",
    "WBS",
    "Task",
    "Phase",
    "Owner",
    "Offset",
    "Start",
    "End",
    "Duration",
    "Notes",
    "Status",
]


def is_num(x) -> bool:
    if x is None:
        return False
    if isinstance(x, (int, float)):
        return not (isinstance(x, float) and math.isnan(x))
    return False


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python selftest_agi_tr_master_release.py <workbook.xlsx>")
        return 2

    path = Path(sys.argv[1])
    wb = load_workbook(path)

    missing = [s for s in REQ_SHEETS if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Missing required sheets: {missing}")

    ws = wb["Schedule_Data"]
    hdr = [ws.cell(5, c).value for c in range(1, 12)]
    if hdr != HEADERS:
        raise SystemExit(f"Schedule_Data headers mismatch. Got: {hdr}")

    # find last data row
    last = ws.max_row
    while last >= 6 and not ws.cell(last, 1).value:
        last -= 1

    # validate numeric columns
    bad = []
    for r in range(6, last + 1):
        if not ws.cell(r, 1).value:
            continue
        off = ws.cell(r, 6).value
        dur = ws.cell(r, 9).value
        if not is_num(off) or not is_num(dur):
            bad.append((r, off, dur))

    if bad:
        raise SystemExit(f"Non-numeric Offset/Duration rows: {bad[:10]}")

    wg = wb["Gantt_Chart"]
    # row alignment check for a few rows
    for r in range(6, min(last, 20) + 1):
        a = wg.cell(r, 1).value
        # formula strings are fine; but should reference Schedule_Data!A{r}
        if isinstance(a, str) and "Schedule_Data!A" in a:
            if f"Schedule_Data!A{r}" not in a:
                raise SystemExit(f"Gantt row alignment issue at row {r}: {a}")

    print("SELFTEST PASS:", path)
    print(f"  Sheets OK. Tasks: {last-5}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
