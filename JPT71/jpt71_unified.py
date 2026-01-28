# Py3.11.8
"""
JPT71 Unified Tool - All-in-One
- Mode 1: Create scaffold workbook
- Mode 2: Refresh Cross_Gantt + Create FINAL sheets (optional)

Usage:
  # Create scaffold
  py jpt71_unified.py scaffold "OUTPUT.xlsx"
  
  # Refresh + Export FINAL (default)
  py jpt71_unified.py refresh "IN.xlsx" "OUT_REFRESHED.xlsx"
  
  # Refresh only (no FINAL sheets)
  py jpt71_unified.py refresh "IN.xlsx" "OUT_REFRESHED.xlsx" --no-final

Requirements:
  pip install pandas==2.33 openpyxl==3.1.5 pywin32
"""

import sys
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# Common Utilities
# ============================================================

def _to_date(x):
    """Convert various date formats to date object"""
    if x is None or x == "":
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def _thin_border(color="BFBFBF"):
    """Create thin border style"""
    thin = Side(style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================
# Mode 1: Create Scaffold
# ============================================================

def create_scaffold(out_path: str):
    """Create JPT71 AutoSuite v8 scaffold workbook"""
    wb = Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    beige = PatternFill("solid", fgColor="FFF2CC")
    sborder = _thin_border()
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # README
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "JPT71 AutoSuite v8 (Scaffold)"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "1) Save as .xlsm"
    ws["A4"] = "2) Import mod_JPT71_Auto_COMPLETE.bas"
    ws["A5"] = "3) Paste Plan sheet event code + ThisWorkbook code (see .bas comments)"
    ws["A6"] = "4) Use Plan only: tblPlan(A:O) + INPUT(Q:R)."
    ws.column_dimensions["A"].width = 120

    # Settings
    ws = wb.create_sheet("Settings", 1)
    ws["A1"] = "Settings (Sites / Ports / Holidays / Options)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A3"] = "Sites"; ws["C3"] = "Ports"; ws["E3"] = "Holidays"; ws["G3"] = "Options"
    for a in ["A3","C3","E3","G3"]:
        ws[a].fill = hdr_fill; ws[a].font = hdr_font; ws[a].alignment = center; ws[a].border = sborder
    sites = ["MW4","MW5","AGI"]
    ports = ["MW4","Al Ghallan Islands"]
    hol = [date(2025,12,25), date(2026,1,1)]
    for i,v in enumerate(sites, start=4):
        ws[f"A{i}"]=v; ws[f"A{i}"].border=sborder
    for i,v in enumerate(ports, start=4):
        ws[f"C{i}"]=v; ws[f"C{i}"].border=sborder
    for i,v in enumerate(hol, start=4):
        ws[f"E{i}"]=v; ws[f"E{i}"].number_format="yyyy-mm-dd"; ws[f"E{i}"].border=sborder

    ws["G4"]="Max events per day"; ws["H4"]=6
    ws["G5"]="Calendar month (1-12)"; ws["H5"]=12
    ws["G6"]="Calendar year"; ws["H6"]=2025
    for r in range(4,7):
        ws[f"G{r}"].border=sborder; ws[f"H{r}"].border=sborder
        ws[f"G{r}"].alignment=left; ws[f"H{r}"].alignment=center

    for col,w in {"A":18,"C":22,"E":16,"G":24,"H":18}.items():
        ws.column_dimensions[col].width=w

    # Calendar_Data (hidden)
    ws = wb.create_sheet("Calendar_Data", 2)
    ws.append(["EventDate","Trip","Type","Action","Location","StartDate","EndDate","ConflictFlag"])
    for c in range(1,9):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.sheet_state="hidden"

    # Calendar_View
    ws = wb.create_sheet("Calendar_View", 3)
    ws.sheet_view.showGridLines = False
    ws["A1"]="Calendar View (Auto)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"]="Year"; ws["B2"]="=Settings!$H$6"
    ws["C2"]="Month"; ws["D2"]="=Settings!$H$5"
    for a in ["A2","C2"]:
        ws[a].font=Font(bold=True)
    for a in ["A2","B2","C2","D2"]:
        ws[a].border=sborder
    dows=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    for i,day in enumerate(dows, start=1):
        cell=ws.cell(4,i,day)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
        ws.column_dimensions[get_column_letter(i)].width=22
    for r in range(5,11):
        ws.row_dimensions[r].height=90
        for c in range(1,8):
            cell=ws.cell(r,c,"")
            cell.border=sborder
            cell.alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)

    # LOG
    ws = wb.create_sheet("LOG", 4)
    ws.append(["Timestamp","User","Sheet","Address","OldValue","NewValue/Msg"])
    for c in range(1,7):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.column_dimensions["A"].width=20
    ws.column_dimensions["B"].width=18
    ws.column_dimensions["C"].width=18
    ws.column_dimensions["D"].width=16
    ws.column_dimensions["E"].width=24
    ws.column_dimensions["F"].width=30

    # LOG_SNAPSHOT
    ws = wb.create_sheet("LOG_SNAPSHOT", 5)
    ws.append(["Version","Timestamp","User","Range","Hash","Note"])
    for c in range(1,7):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
    ws.column_dimensions["F"].width=50

    # Dashboard
    ws = wb.create_sheet("Dashboard", 6)
    ws["A1"]="Dashboard (Auto)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A3"]="Delay Days"; ws["B3"]="=Plan!$R$4"
    ws["A4"]="Next Event Date"; ws["B4"]=""
    ws["A5"]="Conflicts"; ws["B5"]=""
    for r in range(3,6):
        ws[f"A{r}"].font=Font(bold=True); ws[f"A{r}"].border=sborder; ws[f"B{r}"].border=sborder
    ws.column_dimensions["A"].width=22
    ws.column_dimensions["B"].width=18

    # Plan (tblPlan + INPUT)
    ws = wb.create_sheet("Plan", 7)
    headers = ["Trip","Type","Material","Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload",
               "Ref_No","Item_Qty","Qty_Unit","Trailers","Trailer_Capacity","Trailer_Unit","ETA_MW4","Loading_MW4_Time"]
    ws.append(headers)
    # sample row
    ws.append(["87","Aggregate","10mm",date(2025,12,24),date(2025,12,25),"","", "HVDC-AGI-GRM-J71-087","800","Tons","10","80","Tons",date(2025,12,25),"8:00 AM"])
    for c in range(1,len(headers)+1):
        cell=ws.cell(1,c)
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=sborder
        ws.column_dimensions[get_column_letter(c)].width=16
    ws.row_dimensions[1].height=22

    tab = Table(displayName="tblPlan", ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    # INPUT block at Q:R
    ws.column_dimensions["Q"].width=34
    ws.column_dimensions["R"].width=26
    ws["Q1"]="INPUT (edit values in column R)"
    ws["Q1"].fill=hdr_fill; ws["Q1"].font=hdr_font
    ws.merge_cells("Q1:R1")

    def put(r, label, val=None, fmt=None):
        ws[f"Q{r}"]=label; ws[f"Q{r}"].border=sborder; ws[f"Q{r}"].alignment=left; ws[f"Q{r}"].font=Font(bold=True)
        ws[f"R{r}"]=val; ws[f"R{r}"].border=sborder; ws[f"R{r}"].alignment=left
        ws[f"R{r}"].fill=beige
        if fmt: ws[f"R{r}"].number_format=fmt

    put(2,"Debris-8 Actual Completion (date/time)","")
    put(3,"Debris-8 Planned MW4 Offloading Date",date(2025,12,26),"yyyy-mm-dd")
    ws["R4"]="=IF(R2=\"\",0,MAX(0,INT(R2)-R3))"
    put(4,"Delay Days (auto)", ws["R4"].value)

    put(6,"Shipment Prefix","HVDC-AGI-GRM-J71-")
    put(7,"Shipment Seq",87)
    put(8,"Document Date",date(2025,12,23),"dd mmm yyyy")
    put(9,"Vessel Name","JOPETWIL 71")
    put(10,"POL","MW4")
    put(11,"POD","Al Ghallan Islands")
    put(12,"ETD",date(2025,11,26),"dd mmm yyyy")
    put(13,"ETA",date(2025,11,27),"dd mmm yyyy")
    put(14,"Receiver Name","Mr. Khemlal")
    put(15,"Receiver Mobile","(+971 54 586 1053)")
    put(16,"Vendor","Samsung - GRM")
    put(17,"Item Name","Aggregate")
    put(18,"Material","10 mm")
    put(19,"Item Suffix","- Trailer")
    put(20,"QTY",10)
    put(21,"N. W (Ton)",71.6,"0.0")
    put(22,"G. W (Ton)",716.0,"0.00")
    put(23,"Total Style","Nos")
    put(24,"Reference","HVDC-AGI-GRM-J71-086")
    put(25,"Approx Tons",800)
    put(26,"Trailer Capacity (Tons)",80)
    put(27,"Loading Time","8:00 AM")
    put(28,"Mail Row1 Date",date(2025,12,25),"yyyy-mm-dd")
    put(29,"Mail Row1 Action","Debris Offloading")
    put(30,"Mail Row1 Remark","MW4")
    put(31,"Mail Row2 Date",date(2025,12,26),"yyyy-mm-dd")
    put(32,"Mail Row2 Action","10mm Aggregate Loading")
    put(33,"Mail Row2 Remark","MW5")
    put(34,"Berth Location","MW4")
    ws["R35"]="=R28"; put(35,"Berth From (auto)", ws["R35"].value)
    ws["R36"]="=R31"; put(36,"Berth To (auto)", ws["R36"].value)

    # Shipping_List (simple but linked)
    ws = wb.create_sheet("Shipping_List", 8)
    ws["A1"]="SHIPPING LIST FOR LIGHTNING PROJECT"; ws["A1"].font=Font(bold=True, size=16)
    ws.merge_cells("A1:L1")
    ws["A3"]="Date"; ws["B3"]="=Plan!$R$8"
    ws["A4"]="Vessel"; ws["B4"]="=Plan!$R$9"
    ws["A5"]="Shipment No"; ws["B5"]="=Plan!$R$6&Plan!$R$7"
    ws["A6"]="ETD"; ws["B6"]="=Plan!$R$12"
    ws["A7"]="ETA"; ws["B7"]="=Plan!$R$13"
    ws["A9"]="Vendor"; ws["B9"]="=Plan!$R$16"
    ws["A10"]="Item"; ws["B10"]="=Plan!$R$17&\" (\"&Plan!$R$18&\")\""
    ws["A11"]="QTY"; ws["B11"]="=Plan!$R$20"
    ws["A12"]="N.W"; ws["B12"]="=Plan!$R$21"
    ws["A13"]="G.W"; ws["B13"]="=Plan!$R$22"
    for r in range(3,14):
        ws[f"A{r}"].border=sborder; ws[f"B{r}"].border=sborder
    ws.column_dimensions["A"].width=18
    ws.column_dimensions["B"].width=48

    # Mail_Draft (cell-based linked)
    ws = wb.create_sheet("Mail_Draft", 9)
    ws["A1"]="MAIL DRAFT (copy/paste)"; ws["A1"].font=Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A3"]="Please arrange the berth"
    ws["B3"]="=TEXT(Plan!$R$35,\"dd.mmm\")&\"~\"&TEXT(Plan!$R$36,\"dd.mmm\")"
    ws["C3"]="at"; ws["D3"]="=Plan!$R$34"; ws["E3"]="per the updated schedule below:"
    ws["A5"]="Date"; ws["B5"]="Action"; ws["C5"]="Remark"
    ws.merge_cells("C5:E5")
    ws["A6"]="=TEXT(Plan!$R$28,\"dd mmm\")"; ws["B6"]="=Plan!$R$29"; ws["C6"]="=Plan!$R$30"; ws.merge_cells("C6:E6")
    ws["A7"]="=TEXT(Plan!$R$31,\"dd mmm\")"; ws["B7"]="=Plan!$R$32"; ws["C7"]="=Plan!$R$33"; ws.merge_cells("C7:E7")
    ws["A9"]="• Reference"; ws["B9"]="=Plan!$R$24"; ws.merge_cells("B9:E9")
    ws["A10"]="• Item:"; ws["B10"]="=Plan!$R$18"; ws["C10"]="(Approx."; ws["D10"]="=Plan!$R$25"; ws["E10"]="Tons)"
    ws["A11"]="• Trailers:"; ws["B11"]="Total"; ws["C11"]="=Plan!$R$20"; ws["D11"]="trailers"; ws["E11"]="= \"(\"&Plan!$R$26&\" Tons per trailer)\""
    ws["A13"]="Schedule:"; ws.merge_cells("A13:E13")
    ws["A14"]="• ETA at MW4:"; ws["B14"]="=TEXT(Plan!$R$28,\"yyyy-mm-dd\")"; ws.merge_cells("B14:E14")
    ws["A15"]="• Loading at MW4:"; ws["B15"]="=TEXT(Plan!$R$31,\"yyyy-mm-dd\")"; ws["C15"]="=Plan!$R$27"; ws.merge_cells("C15:E15")
    ws.column_dimensions["A"].width=20
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width=32

    wb.save(out_path)
    print(f"✓ Scaffold created: {out_path}")


# ============================================================
# Mode 2: Build Cross Gantt
# ============================================================

def build_cross_gantt(wb, df_plan, delay_ref="Plan!$R$4", inprog_trip="Debris-8", view_days_after=14):
    """Rebuild Cross_Gantt sheet from Plan data"""
    if "Cross_Gantt" in wb.sheetnames:
        wb.remove(wb["Cross_Gantt"])
    ws = wb.create_sheet("Cross_Gantt")

    thin_border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    fill_row_agg = PatternFill("solid", fgColor="E2EFDA")
    fill_row_deb = PatternFill("solid", fgColor="FCE4D6")
    fill_inprog = PatternFill("solid", fgColor="D9D9D9")

    fill_mw4 = PatternFill("solid", fgColor="4F81BD")
    fill_agi_off = PatternFill("solid", fgColor="C6EFCE")
    fill_agi_debload = PatternFill("solid", fgColor="F4B084")
    fill_mw4_deboff = PatternFill("solid", fgColor="F8CBAD")

    dash = Side(style="dashed", color="7F7F7F")
    border_dashed = Border(left=dash, right=dash, top=dash, bottom=dash)

    headers = [
        "Seq","Trip","Type","Material",
        "MW4 Depart (Agg)","AGI Offload (Agg)",
        "AGI Debris Loading (Deb)","MW4 Debris Offloading (Deb)",
        "Status"
    ]
    hidden_headers = ["ShiftFlag","Plan_MW4_Depart","Plan_AGI_Offload","Plan_AGI_Deb_Load","Plan_MW4_Deb_Off"]

    df = df_plan.copy()
    need_cols = ["Trip","Type","Material",
                 "Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]
    for c in need_cols:
        if c not in df.columns:
            df[c] = None

    for c in ["Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]:
        df[c] = df[c].apply(_to_date)

    df = df[df["Trip"].notna() & (df["Trip"].astype(str).str.strip() != "")].reset_index(drop=True)
    if df.empty:
        ws["A1"] = "No plan rows in Plan sheet."
        return

    all_dates = []
    for c in ["Plan_MW4_Depart_Agg","Plan_AGI_Offload_Agg","Plan_AGI_Debris_Load","Plan_MW4_Debris_Offload"]:
        for d in df[c].tolist():
            if pd.isna(d) or d is None:
                continue
            converted_date = _to_date(d)
            if converted_date is not None:
                all_dates.append(converted_date)

    if not all_dates:
        ws["A1"] = "No valid dates found in Plan sheet."
        return
    plan_start = min(all_dates)
    plan_end = max(all_dates)
    view_end = plan_end + timedelta(days=view_days_after)

    dates = []
    d = plan_start
    while d <= view_end:
        dates.append(d); d += timedelta(days=1)

    inprog_idx = None
    for i, trip in enumerate(df["Trip"].astype(str).tolist(), start=1):
        if trip.strip() == inprog_trip:
            inprog_idx = i
            break
    if inprog_idx is None:
        inprog_idx = 999999

    col = 1
    for h in headers:
        c = ws.cell(1, col, h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = thin_border
        col += 1
    for h in hidden_headers:
        c = ws.cell(1, col, h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = thin_border
        col += 1

    date_start_col = col
    for i, dt in enumerate(dates):
        c = ws.cell(1, date_start_col + i, dt)
        c.number_format = "mm-dd"
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = thin_border
        ws.column_dimensions[get_column_letter(date_start_col + i)].width = 5

    for c,w in {1:5,2:12,3:10,4:14,5:14,6:14,7:18,8:20,9:14}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    shift_col = len(headers) + 1
    for c in range(shift_col, date_start_col):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    ws.freeze_panes = "J2"
    ws.row_dimensions[1].height = 22

    for idx, row in enumerate(df.itertuples(index=False), start=1):
        r = 1 + idx
        trip = str(row.Trip)
        typ = str(row.Type)
        mat = str(row.Material)

        p_mw4 = getattr(row, "Plan_MW4_Depart_Agg")
        p_agi_off = getattr(row, "Plan_AGI_Offload_Agg")
        p_debload = getattr(row, "Plan_AGI_Debris_Load")
        p_deboff = getattr(row, "Plan_MW4_Debris_Offload")

        row_fill = fill_row_deb if typ.strip().lower() == "debris" else fill_row_agg

        ws.cell(r,1,idx).alignment=center
        ws.cell(r,2,trip).alignment=center
        ws.cell(r,3,typ).alignment=center
        ws.cell(r,4,mat).alignment=left
        ws.cell(r,9,"IN PROGRESS" if trip.strip()==inprog_trip else "").alignment=center

        ws.cell(r, shift_col, 1 if idx > inprog_idx else 0).alignment=center

        plan_start_col = shift_col + 1
        plan_vals = [p_mw4, p_agi_off, p_debload, p_deboff]
        for j,val in enumerate(plan_vals):
            c = ws.cell(r, plan_start_col + j, val)
            if val:
                c.number_format = "yyyy-mm-dd"
            c.alignment = center

        flag_ref = f"${get_column_letter(shift_col)}{r}"
        plan_ref = [
            f"{get_column_letter(plan_start_col)}{r}",
            f"{get_column_letter(plan_start_col+1)}{r}",
            f"{get_column_letter(plan_start_col+2)}{r}",
            f"{get_column_letter(plan_start_col+3)}{r}",
        ]

        def adj(plan_cell_ref: str):
            return f'=IF({plan_cell_ref}="", "", IF({flag_ref}=1, {plan_cell_ref}+{delay_ref}, {plan_cell_ref}))'

        ws.cell(r,5).value = adj(plan_ref[0])
        ws.cell(r,6).value = adj(plan_ref[1])
        ws.cell(r,7).value = adj(plan_ref[2])
        ws.cell(r,8).value = adj(plan_ref[3])
        for c in [5,6,7,8]:
            ws.cell(r,c).number_format="yyyy-mm-dd"
            ws.cell(r,c).alignment=center

        for c in range(1,10):
            cell = ws.cell(r,c)
            cell.border = thin_border
            cell.fill = fill_inprog if trip.strip()==inprog_trip else row_fill

        if trip.strip() == inprog_trip:
            for c in range(1,10):
                ws.cell(r,c).border = border_dashed

    first_data = 2
    last_data = 1 + len(df)
    first_date = date_start_col
    last_date = date_start_col + len(dates) - 1
    inprog_row = 1 + inprog_idx

    def add_cf(r1, r2, fill, col_letter):
        if r1 > r2:
            return
        rng = f"{get_column_letter(first_date)}{r1}:{get_column_letter(last_date)}{r2}"
        top_left = get_column_letter(first_date)
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'={top_left}$1=${col_letter}{r1}'], fill=fill))

    for (r1,r2) in [(first_data, inprog_row-1), (inprog_row+1, last_data)]:
        add_cf(r1,r2,fill_mw4,'E')
        add_cf(r1,r2,fill_agi_off,'F')
        add_cf(r1,r2,fill_agi_debload,'G')
        add_cf(r1,r2,fill_mw4_deboff,'H')

    for c in range(first_date, last_date+1):
        cell = ws.cell(inprog_row, c)
        cell.fill = fill_inprog
        cell.border = border_dashed

    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell=ws.cell(r,c)
            styles=[cell.border.left.style, cell.border.right.style, cell.border.top.style, cell.border.bottom.style]
            if "dashed" in styles:
                continue
            cell.border = thin_border


# ============================================================
# Mode 2: Create FINAL Sheets
# ============================================================

FINAL_SUFFIX = "_FINAL"


def _delete_sheet_if_exists(wb, sheet_name: str) -> None:
    """Delete sheet if exists (Excel COM)"""
    for i in range(1, wb.Worksheets.Count + 1):
        if wb.Worksheets(i).Name == sheet_name:
            wb.Worksheets(i).Delete()
            return


def create_final_sheets_in_workbook(refreshed_path: str, suffix: str = FINAL_SUFFIX) -> None:
    """Create FINAL sheets (values only) using Excel COM"""
    try:
        import win32com.client as win32
        from win32com.client import constants as xl
    except Exception as e:
        raise RuntimeError("pywin32 is required. Install with: pip install pywin32") from e

    xlApp = win32.DispatchEx("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = False

    try:
        wb = xlApp.Workbooks.Open(refreshed_path)
        xlApp.CalculateFullRebuild()

        def copy_sheet_as_values(sheet_name: str):
            ws_src = wb.Worksheets(sheet_name)
            tgt_name = f"{sheet_name}{suffix}"
            _delete_sheet_if_exists(wb, tgt_name)
            ws_tgt = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
            ws_tgt.Name = tgt_name

            # Column widths
            ws_src.UsedRange.Copy()
            ws_tgt.Range("A1").PasteSpecial(Paste=xl.xlPasteColumnWidths)

            # Values + number formats
            ws_src.UsedRange.Copy()
            ws_tgt.Range("A1").PasteSpecial(Paste=xl.xlPasteValuesAndNumberFormats)

            # Formats
            ws_src.UsedRange.Copy()
            ws_tgt.Range("A1").PasteSpecial(Paste=xl.xlPasteFormats)

            # Row heights
            last_row = ws_src.UsedRange.Rows.Count
            for r in range(1, last_row+1):
                ws_tgt.Rows(r).RowHeight = ws_src.Rows(r).RowHeight

            # Page setup
            ws_tgt.PageSetup.Orientation = ws_src.PageSetup.Orientation
            ws_tgt.PageSetup.FitToPagesWide = ws_src.PageSetup.FitToPagesWide
            ws_tgt.PageSetup.FitToPagesTall = ws_src.PageSetup.FitToPagesTall

        copy_sheet_as_values("Shipping_List")
        copy_sheet_as_values("Mail_Draft")
        copy_sheet_as_values("Cross_Gantt")

        wb.Save()
        wb.Close(SaveChanges=True)

    finally:
        xlApp.CutCopyMode = False
        xlApp.Quit()


# ============================================================
# Mode 2: Refresh + Export FINAL
# ============================================================

def refresh_and_export(in_path: str, out_refreshed: str, create_final: bool = True):
    """Refresh Cross_Gantt and optionally create FINAL sheets"""
    df = pd.read_excel(in_path, sheet_name="Plan", usecols="A:O", dtype=str)
    wb = load_workbook(in_path)

    build_cross_gantt(wb, df_plan=df, delay_ref="Plan!$R$4", inprog_trip="Debris-8", view_days_after=14)
    wb.save(out_refreshed)

    if create_final:
        create_final_sheets_in_workbook(out_refreshed)
        print(f"✓ Saved: {out_refreshed}")
        print(f"✓ Added FINAL sheets in: {out_refreshed}")
    else:
        print(f"✓ Saved: {out_refreshed}")
        print(f"✓ Cross_Gantt refreshed (no FINAL sheets)")


# ============================================================
# Main Entry Point
# ============================================================

def main():
    """Main entry point with mode selection"""
    if len(sys.argv) < 2:
        print("JPT71 Unified Tool")
        print("\nUsage:")
        print('  # Create scaffold workbook')
        print('  py jpt71_unified.py scaffold "OUTPUT.xlsx"')
        print('\n  # Refresh + Export FINAL')
        print('  py jpt71_unified.py refresh "IN.xlsx" "OUT_REFRESHED.xlsx"')
        print('\n  # Refresh only (no FINAL)')
        print('  py jpt71_unified.py refresh "IN.xlsx" "OUT_REFRESHED.xlsx" --no-final')
        sys.exit(1)

    mode = sys.argv[1].lower()

    if mode == "scaffold":
        if len(sys.argv) < 3:
            out_path = "JPT71_AutoSuite_v8_Scaffold.xlsx"
        else:
            out_path = sys.argv[2]
        create_scaffold(out_path)

    elif mode == "refresh":
        if len(sys.argv) < 4:
            raise SystemExit('Usage: py jpt71_unified.py refresh "IN.xlsx" "OUT_REFRESHED.xlsx" [--no-final]')
        
        create_final = True
        if len(sys.argv) >= 5 and sys.argv[4] == "--no-final":
            create_final = False
        
        refresh_and_export(sys.argv[2], sys.argv[3], create_final=create_final)

    else:
        raise SystemExit(f'Unknown mode: {mode}\nUse "scaffold" or "refresh"')


if __name__ == "__main__":
    main()