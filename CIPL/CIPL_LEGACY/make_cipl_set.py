#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CIPL Set Builder (CI P1 + CI Rider P2 + PL P1 + PL Rider P2) — Unified Runner

- Input: one JSON (voyage_input.json)
- Output: one XLSX with 4 sheets (same formats as your existing scripts)

Dependencies:
  pip install openpyxl

Run:
  python make_cipl_set.py --in voyage_input.json --out CIPL_HVDC-ADOPT-SCT-0159.xlsx
"""

from __future__ import annotations

import argparse
import importlib.util
import importlib.machinery
import json
import sys
from pathlib import Path
from typing import Any, Dict, Callable, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# show, don't guess: load a python file as a module
def load_module(py_path: Path, module_name: str):
    py_path_resolved = py_path.resolve()
    if not py_path_resolved.exists():
        raise FileNotFoundError(f"Module file not found: {py_path_resolved}")
    
    # .PY 확장자도 인식하도록 loader를 명시적으로 지정
    loader = importlib.machinery.SourceFileLoader(module_name, str(py_path_resolved))
    spec = importlib.util.spec_from_loader(module_name, loader)
    
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load module: {py_path_resolved} (spec is None or loader is None)")
    
    # 모듈을 sys.modules에 등록 (dataclass 등이 제대로 작동하도록)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = mod
    spec.loader.exec_module(mod)
    return mod


def read_json(p: Path) -> Dict[str, Any]:
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_key(d: Dict[str, Any], key: str) -> Any:
    if key not in d:
        raise KeyError(f"Missing required key: {key}")
    return d[key]


def build_ci_p1(ws: Worksheet, mod_ci_p1, data: Dict[str, Any]) -> None:
    # expected functions in COMMERCIAL INVOICE.PY:
    # set_page_setup(ws), set_dimensions(ws), build_commercial_invoice(ws, data_dict)
    mod_ci_p1.set_page_setup(ws)
    mod_ci_p1.set_dimensions(ws)
    mod_ci_p1.build_commercial_invoice(ws, data)


def build_ci_rider_p2(ws: Worksheet, mod_ci_rider, data: Dict[str, Any]) -> None:
    # expected in CI RIDER.PY:
    # set_page_setup(ws), set_col_widths(ws), build_sheet(ws, payload)
    mod_ci_rider.set_page_setup(ws)
    mod_ci_rider.set_col_widths(ws)
    mod_ci_rider.build_sheet(ws, data)


def build_pl_p1(ws: Worksheet, mod_pl_p1, data: Dict[str, Any]) -> None:
    # expected in PACKING LIST.PY:
    # set_page_setup(ws), set_dimensions(ws), build_packing_list(ws, data_dict)
    mod_pl_p1.set_page_setup(ws)
    mod_pl_p1.set_dimensions(ws)
    mod_pl_p1.build_packing_list(ws, data)


def build_pl_rider_p2(ws: Worksheet, mod_pl_rider, data: Dict[str, Any]) -> None:
    # expected in PACKING LIST ATTACHED RIDER.PY:
    # set_page_setup(ws), then column widths already inside builder OR need to set here
    # build_rider(ws, payload) returns last_row
    mod_pl_rider.set_page_setup(ws)
    # set widths A:R exactly
    for col, w in mod_pl_rider.COL_WIDTHS.items():
        ws.column_dimensions[col].width = float(w)
    mod_pl_rider.build_rider(ws, data)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input JSON path (voyage_input.json)")
    ap.add_argument("--out", required=True, help="Output XLSX path")

    # optional explicit file paths
    ap.add_argument("--ci_p1", default="COMMERCIAL INVOICE.PY", help="Path to COMMERCIAL INVOICE.PY")
    ap.add_argument("--ci_rider", default="CI RIDER.PY", help="Path to CI RIDER.PY")
    ap.add_argument("--pl_p1", default="PACKING LIST.PY", help="Path to PACKING LIST.PY")
    ap.add_argument("--pl_rider", default="PACKING LIST ATTACHED RIDER.PY", help="Path to PACKING LIST ATTACHED RIDER.PY")

    args = ap.parse_args()

    # 스크립트가 있는 디렉토리를 기준으로 상대 경로 해결
    script_dir = Path(__file__).parent.resolve()

    # 입력/출력 파일 경로 해결 (절대 경로가 아니면 스크립트 디렉토리 기준)
    inp_path = Path(args.inp)
    out_path = Path(args.out)
    inp = inp_path.resolve() if inp_path.is_absolute() else (script_dir / inp_path).resolve()
    out = out_path.resolve() if out_path.is_absolute() else (script_dir / out_path).resolve()

    # load inputs
    payload = read_json(inp)

    # input contract (single JSON -> 4 payload blocks)
    ci_p1_data = ensure_key(payload, "ci_p1")
    ci_rider_data = ensure_key(payload, "ci_rider_p2")
    pl_p1_data = ensure_key(payload, "pl_p1")
    pl_rider_data = ensure_key(payload, "pl_rider_p2")

    # load modules - 스크립트 디렉토리 기준으로 상대 경로 해결
    ci_p1_path = Path(args.ci_p1)
    ci_rider_path = Path(args.ci_rider)
    pl_p1_path = Path(args.pl_p1)
    pl_rider_path = Path(args.pl_rider)
    
    mod_ci_p1 = load_module(ci_p1_path.resolve() if ci_p1_path.is_absolute() else (script_dir / ci_p1_path).resolve(), "ci_p1_mod")
    mod_ci_rider = load_module(ci_rider_path.resolve() if ci_rider_path.is_absolute() else (script_dir / ci_rider_path).resolve(), "ci_rider_mod")
    mod_pl_p1 = load_module(pl_p1_path.resolve() if pl_p1_path.is_absolute() else (script_dir / pl_p1_path).resolve(), "pl_p1_mod")
    mod_pl_rider = load_module(pl_rider_path.resolve() if pl_rider_path.is_absolute() else (script_dir / pl_rider_path).resolve(), "pl_rider_mod")

    # build workbook
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Commercial_Invoice_P1"
    build_ci_p1(ws1, mod_ci_p1, ci_p1_data)

    ws2 = wb.create_sheet("CI_Rider_P2")
    build_ci_rider_p2(ws2, mod_ci_rider, ci_rider_data)

    ws3 = wb.create_sheet("Packing_List_P1")
    build_pl_p1(ws3, mod_pl_p1, pl_p1_data)

    ws4 = wb.create_sheet("PL_Rider_P2")
    build_pl_rider_p2(ws4, mod_pl_rider, pl_rider_data)

    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
