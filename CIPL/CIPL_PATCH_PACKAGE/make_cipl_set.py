#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""make_cipl_set.py — Unified CIPL Builder

- Supports two input modes:
  1) Expanded mode: {ci_p1, ci_rider_p2, pl_p1, pl_rider_p2}
  2) Commons mode: {commons, static_parties, (optional) ci_rider_items, pl_rider_items}
     -> uses CIPL.make_4page_data_dicts()

- Builds one workbook with 4 sheets using your existing format scripts:
  - COMMERCIAL INVOICE.PY
  - CI RIDER.PY
  - PACKING LIST.PY
  - PACKING LIST ATTACHED RIDER.PY

Run:
  python make_cipl_set.py --in voyage_input.json --out CIPL.xlsx
"""

from __future__ import annotations

import argparse
import importlib.util
import json
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_module(py_path: Path, module_name: str):
    # Robust loader even when file extension is not '.py' (e.g., '.PY')
    from importlib.machinery import SourceFileLoader
    import sys

    spec = importlib.util.spec_from_loader(module_name, SourceFileLoader(module_name, str(py_path)))
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load module: {py_path}")
    mod = importlib.util.module_from_spec(spec)
    # dataclasses (and some libs) expect the module to be registered in sys.modules during execution
    sys.modules[module_name] = mod
    spec.loader.exec_module(mod)
    return mod



def read_json(p: Path) -> Dict[str, Any]:
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def call_page_setup(mod, ws: Worksheet):
    # Normalizes inconsistency: some files use set_page_setup(), CI Rider uses apply_page_setup()
    if hasattr(mod, "set_page_setup"):
        mod.set_page_setup(ws)
        return
    if hasattr(mod, "apply_page_setup"):
        mod.apply_page_setup(ws)
        return
    raise AttributeError(f"No page-setup function found in module: {mod.__file__}")


def build_ci_p1(ws: Worksheet, mod_ci_p1, data: Dict[str, Any]) -> None:
    call_page_setup(mod_ci_p1, ws)
    if hasattr(mod_ci_p1, "set_dimensions"):
        mod_ci_p1.set_dimensions(ws)
    mod_ci_p1.build_commercial_invoice(ws, data)


def build_ci_rider_p2(ws: Worksheet, mod_ci_rider, data: Dict[str, Any]) -> None:
    call_page_setup(mod_ci_rider, ws)  # apply_page_setup in CI RIDER
    mod_ci_rider.set_col_widths(ws)
    mod_ci_rider.build_sheet(ws, data)


def build_pl_p1(ws: Worksheet, mod_pl_p1, data: Dict[str, Any]) -> None:
    call_page_setup(mod_pl_p1, ws)
    if hasattr(mod_pl_p1, "set_dimensions"):
        mod_pl_p1.set_dimensions(ws)
    mod_pl_p1.build_packing_list(ws, data)


def build_pl_rider_p2(ws: Worksheet, mod_pl_rider, data: Dict[str, Any]) -> None:
    call_page_setup(mod_pl_rider, ws)
    # ensure widths A:R exactly
    if hasattr(mod_pl_rider, "COL_WIDTHS"):
        for col, w in mod_pl_rider.COL_WIDTHS.items():
            ws.column_dimensions[col].width = float(w)
    mod_pl_rider.build_rider(ws, data)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input JSON path")
    ap.add_argument("--out", required=False, help="Output XLSX path (optional)")

    ap.add_argument("--ci_p1", default="COMMERCIAL INVOICE.PY", help="Path to COMMERCIAL INVOICE.PY")
    ap.add_argument("--ci_rider", default="CI RIDER.PY", help="Path to CI RIDER.PY")
    ap.add_argument("--pl_p1", default="PACKING LIST.PY", help="Path to PACKING LIST.PY")
    ap.add_argument("--pl_rider", default="PACKING LIST ATTACHED RIDER.PY", help="Path to PACKING LIST ATTACHED RIDER.PY")
    ap.add_argument("--cipl_mapper", default="CIPL.py", help="Path to CIPL.py (commons mapper)")

    args = ap.parse_args()

    inp = Path(args.inp).resolve()
    payload = read_json(inp)

    # If commons mode, use mapper. If expanded, pass-through.
    mapper_path = Path(args.cipl_mapper).resolve()
    if ("commons" in payload and "static_parties" in payload):
        mod_cipl = load_module(mapper_path, "cipl_mapper_mod")
        pages = mod_cipl.make_4page_data_dicts(payload, cbm_decimals=3)
    else:
        # Expect expanded mode
        required = ("ci_p1", "ci_rider_p2", "pl_p1", "pl_rider_p2")
        missing = [k for k in required if k not in payload]
        if missing:
            raise KeyError(f"Input JSON missing keys: {missing}. Provide commons+static_parties or expanded 4-page dicts.")
        pages = {k: payload[k] for k in required}

    # Auto output name if not provided
    out_path: Path
    if args.out:
        out_path = Path(args.out).resolve()
    else:
        # Try invoice no from commons or CI P1
        inv = None
        if "commons" in payload and "invoice_no" in payload.get("commons", {}):
            inv = payload["commons"]["invoice_no"]
        else:
            inv = pages["ci_p1"].get("invoice_no", "CIPL")
        out_path = inp.parent / f"CIPL_{inv}.xlsx"

    # Load page modules
    mod_ci_p1 = load_module(Path(args.ci_p1).resolve(), "ci_p1_mod")
    mod_ci_rider = load_module(Path(args.ci_rider).resolve(), "ci_rider_mod")
    mod_pl_p1 = load_module(Path(args.pl_p1).resolve(), "pl_p1_mod")
    mod_pl_rider = load_module(Path(args.pl_rider).resolve(), "pl_rider_mod")

    # Build workbook
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Commercial_Invoice_P1"
    build_ci_p1(ws1, mod_ci_p1, pages["ci_p1"])

    ws2 = wb.create_sheet("CI_Rider_P2")
    build_ci_rider_p2(ws2, mod_ci_rider, pages["ci_rider_p2"])

    ws3 = wb.create_sheet("Packing_List_P1")
    build_pl_p1(ws3, mod_pl_p1, pages["pl_p1"])

    ws4 = wb.create_sheet("PL_Rider_P2")
    build_pl_rider_p2(ws4, mod_pl_rider, pages["pl_rider_p2"])

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
