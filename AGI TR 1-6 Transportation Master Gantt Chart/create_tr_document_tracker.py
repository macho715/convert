#!/usr/bin/env python3
"""
TR Transportation Document Tracker with VBA
HVDC Transformer Transportation - Document Preparation & Progress Tracking
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
import datetime
import os
import time

# Create workbook
wb = Workbook()

# Colors
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")  # Dark blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2D5A8A")  # Medium blue
VOYAGE_1_FILL = PatternFill("solid", fgColor="FFD700")  # Gold
VOYAGE_2_FILL = PatternFill("solid", fgColor="90EE90")  # Light green
VOYAGE_3_FILL = PatternFill("solid", fgColor="87CEEB")  # Sky blue
VOYAGE_4_FILL = PatternFill("solid", fgColor="DDA0DD")  # Plum
CRITICAL_FILL = PatternFill("solid", fgColor="FF6B6B")  # Red
IMPORTANT_FILL = PatternFill("solid", fgColor="FFB347")  # Orange
STANDARD_FILL = PatternFill("solid", fgColor="77DD77")  # Pastel green
GATE_PASS_FILL = PatternFill("solid", fgColor="B0E0E6")  # Powder blue
COMPLETE_FILL = PatternFill("solid", fgColor="90EE90")  # Light green
IN_PROGRESS_FILL = PatternFill("solid", fgColor="FFFF99")  # Light yellow
NOT_STARTED_FILL = PatternFill("solid", fgColor="FFCCCC")  # Light red
NA_FILL = PatternFill("solid", fgColor="D3D3D3")  # Light gray

WHITE_FONT = Font(color="FFFFFF", bold=True)
BLACK_FONT = Font(color="000000")
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ==================== SHEET 1: DASHBOARD ====================
ws_dash = wb.active
ws_dash.title = "Dashboard"

# Title
ws_dash.merge_cells('A1:N2')
ws_dash['A1'] = "HVDC TR Transportation - Document Preparation Progress Dashboard"
ws_dash['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_dash['A1'].alignment = center_align

# Date
ws_dash['A3'] = f"Last Updated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws_dash['A3'].font = Font(italic=True, size=10)

# Voyage Summary Table
row = 5
voyage_headers = ["Voyage", "TR Units", "MZP Arrival", "Load-out", "MZP Departure", "AGI Arrival", "Doc Deadline", "Land Permit By", "Status"]
ws_dash.merge_cells(f'A{row}:I{row}')
ws_dash[f'A{row}'] = "📅 Voyage Schedule Summary"
ws_dash[f'A{row}'].font = TITLE_FONT
ws_dash[f'A{row}'].fill = HEADER_FILL
ws_dash[f'A{row}'].font = WHITE_FONT

row += 1
for col, header in enumerate(voyage_headers, 1):
    cell = ws_dash.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

voyage_data = [
    ["Voyage 1", "TR 1-2", "01-27", "01-29~30", "02-01", "02-02", "01-23", "01-22", ""],
    ["Voyage 2 ⚡", "TR 3-4", "02-06", "02-07~08", "02-10", "02-11", "02-03", "02-02", ""],
    ["Voyage 3", "TR 5-6", "02-15", "02-16~17", "02-19", "02-20", "02-12", "02-11", ""],
    ["Voyage 4 (Final)", "TR 7", "02-24", "02-25", "02-27", "02-28", "02-20", "02-19", ""],
]

voyage_fills = [VOYAGE_1_FILL, VOYAGE_2_FILL, VOYAGE_3_FILL, VOYAGE_4_FILL]
for i, v_data in enumerate(voyage_data):
    row += 1
    for col, val in enumerate(v_data, 1):
        cell = ws_dash.cell(row=row, column=col, value=val)
        cell.alignment = center_align
        cell.border = thin_border
        if col == 1:
            cell.fill = voyage_fills[i]
            cell.font = BOLD_FONT

# Party Progress Summary
row += 3
ws_dash.merge_cells(f'A{row}:G{row}')
ws_dash[f'A{row}'] = "📊 Party-wise Document Submission Progress"
ws_dash[f'A{row}'].font = TITLE_FONT
ws_dash[f'A{row}'].fill = HEADER_FILL
ws_dash[f'A{row}'].font = WHITE_FONT

row += 1
party_headers = ["Responsible Party", "Total Docs", "Completed", "In Progress", "Not Started", "N/A", "Progress %"]
for col, header in enumerate(party_headers, 1):
    cell = ws_dash.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

parties = ["Samsung C&T", "Mammoet", "OFCO Agency", "ADNOC L&S", "Vessel Owner (KFS)", "DSV Solutions", "MWS (Sterling)"]
for party in parties:
    row += 1
    ws_dash.cell(row=row, column=1, value=party).font = BOLD_FONT
    ws_dash.cell(row=row, column=1).border = thin_border
    for col in range(2, 8):
        cell = ws_dash.cell(row=row, column=col, value="-")
        cell.alignment = center_align
        cell.border = thin_border

# Key Contacts
row += 3
ws_dash.merge_cells(f'A{row}:E{row}')
ws_dash[f'A{row}'] = "📞 Key Contacts"
ws_dash[f'A{row}'].font = TITLE_FONT
ws_dash[f'A{row}'].fill = HEADER_FILL
ws_dash[f'A{row}'].font = WHITE_FONT

row += 1
contact_headers = ["Party", "Name", "Email", "Phone", "Role"]
for col, header in enumerate(contact_headers, 1):
    cell = ws_dash.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

contacts = [
    ["OFCO Agency", "Nanda Kumar", "nkk@ofco-int.com", "", "PTW Coordinator"],
    ["OFCO Agency", "Das Gopal", "das@ofco-int.com", "", "Documentation"],
    ["ADNOC L&S", "Mahmoud Ouda", "moda@adnoc.ae", "", "Port Coordinator"],
    ["LCT Bushra", "Capt. Joey Vargas", "lct.bushra@khalidfarajshipping.com", "", "Vessel Master"],
    ["Mammoet", "Yulia Frolova", "Yulia.Frolova@mammoet.com", "", "Project Manager"],
    ["DSV Solutions", "Jay Manaloto", "jay.manaloto@dsv.com", "", "Logistics Coordinator"],
    ["Samsung C&T", "Cha", "", "", "Project Manager"],
]

for contact in contacts:
    row += 1
    for col, val in enumerate(contact, 1):
        cell = ws_dash.cell(row=row, column=col, value=val)
        cell.alignment = left_align
        cell.border = thin_border

# Column widths for Dashboard
col_widths = [18, 12, 12, 12, 14, 12, 14, 14, 12, 12, 10, 10, 10, 10]
for i, width in enumerate(col_widths, 1):
    ws_dash.column_dimensions[get_column_letter(i)].width = width

# ==================== SHEET 2: DOCUMENT TRACKER (MAIN) ====================
ws_main = wb.create_sheet("Document_Tracker")

# Title
ws_main.merge_cells('A1:Q2')
ws_main['A1'] = "HVDC TR Transportation - Document Preparation Status by Voyage & Party"
ws_main['A1'].font = Font(bold=True, size=16, color="1E3A5F")
ws_main['A1'].alignment = center_align

# Headers
row = 4
headers = [
    "No", "Category", "Document Name", "Description", "Priority", 
    "Responsible Party", "Lead Time",
    "V1 Status", "V1 Date", "V1 Remarks",
    "V2 Status", "V2 Date", "V2 Remarks",
    "V3 Status", "V3 Date", "V3 Remarks",
    "V4 Status", "V4 Date", "V4 Remarks"
]

for col, header in enumerate(headers, 1):
    cell = ws_main.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

# Document Data
documents = [
    # PTW & Loading Documents
    ["1", "PTW & Loading", "Risk Assessment", "AD Port form required", "Mandatory", "OFCO Agency", "3 days"],
    ["2", "PTW & Loading", "PTW Applicant/Receiver Consent", "Permit-to-Work consent form", "Mandatory", "OFCO Agency", "3 days"],
    ["3", "PTW & Loading", "PTW Land Oversized & Heavy Load", "SPMT land operations permit", "Mandatory", "OFCO Agency", "2-3 days"],
    ["4", "PTW & Loading", "Stowage Plan", "Cargo stowage plan", "Mandatory", "Mammoet", "5 days"],
    ["5", "PTW & Loading", "Critical Lifting Plan", "Heavy lift plan", "Mandatory", "Mammoet", "5 days"],
    ["6", "PTW & Loading", "Method Statement (incl. Weather)", "Method statement with weather criteria", "Mandatory", "Mammoet", "5 days"],
    ["7", "PTW & Loading", "Countdown Plan", "Operation countdown schedule", "Important", "Samsung C&T", "3 days"],
    ["8", "PTW & Loading", "Undertaking Letter", "Undertaking letter", "Important", "Samsung C&T", "2 days"],
    ["9", "PTW & Loading", "Stability Calculation", "Vessel stability calculation (AGI TR.PY)", "Mandatory", "Vessel Owner (KFS)", "5 days"],
    ["10", "PTW & Loading", "3rd Party Equipment Certificates", "SPMT, crane certificates", "Important", "Mammoet", "3 days"],
    ["11", "PTW & Loading", "Marine Warranty Survey (MWS)", "Marine warranty survey", "Mandatory", "MWS (Sterling)", "7 days"],
    ["12", "PTW & Loading", "Mooring Plan", "Mooring plan", "Standard", "Vessel Owner (KFS)", "3 days"],
    ["13", "PTW & Loading", "Indemnity Letter - Lifting", "Lifting-related indemnity", "Important", "Samsung C&T", "2 days"],
    ["14", "PTW & Loading", "Lashing Plan", "Cargo lashing plan", "Mandatory", "Mammoet", "3 days"],
    ["15", "PTW & Loading", "Indemnity Letter (General)", "General indemnity letter", "Important", "Samsung C&T", "2 days"],
    
    # AD Maritime NOC Documents
    ["16", "AD Maritime NOC", "Local Trading License", "Business registration", "Mandatory", "Samsung C&T", "N/A"],
    ["17", "AD Maritime NOC", "Detailed Risk Assessment & ERP", "Risk assessment & emergency response plan", "Mandatory", "Mammoet", "5 days"],
    ["18", "AD Maritime NOC", "No Objection Letter", "NOL from relevant authorities", "Mandatory", "ADNOC L&S", "5 days"],
    ["19", "AD Maritime NOC", "Voyage Plan", "Voyage plan (MZP → AGI)", "Mandatory", "Vessel Owner (KFS)", "3 days"],
    ["20", "AD Maritime NOC", "Route Map", "Route map", "Important", "Vessel Owner (KFS)", "2 days"],
    ["21", "AD Maritime NOC", "Contract Award Letter Copy", "Copy of contract award", "Important", "Samsung C&T", "N/A"],
    
    # Engineering Documents
    ["22", "Engineering", "Vessel Data Package", "GA drawings, stability booklet", "Mandatory", "Vessel Owner (KFS)", "5 days"],
    ["23", "Engineering", "Ballasting Analysis", "Ballasting calculations", "Mandatory", "Vessel Owner (KFS)", "5 days"],
    ["24", "Engineering", "Ramp Strength Verification", "Ramp FEA results", "Mandatory", "Mammoet", "7 days"],
    ["25", "Engineering", "Engineering Simulation Report", "Combined engineering report", "Mandatory", "Mammoet", "7 days"],
    ["26", "Engineering", "SPMT Load Distribution Plan", "SPMT load calculations", "Mandatory", "Mammoet", "5 days"],
    ["27", "Engineering", "Sea Fastening Design", "Sea fastening calculations", "Mandatory", "Mammoet", "5 days"],
    
    # Vessel Documents
    ["28", "Vessel", "Vessel Class Certificate", "Classification certificate", "Mandatory", "Vessel Owner (KFS)", "N/A"],
    ["29", "Vessel", "Load Line Certificate", "International load line cert", "Mandatory", "Vessel Owner (KFS)", "N/A"],
    ["30", "Vessel", "Safety Construction Certificate", "Safety construction cert", "Mandatory", "Vessel Owner (KFS)", "N/A"],
    ["31", "Vessel", "Insurance Certificate", "Marine insurance cert", "Mandatory", "Vessel Owner (KFS)", "N/A"],
    ["32", "Vessel", "Crew List", "Crew list with certifications", "Standard", "Vessel Owner (KFS)", "1 day"],
    
    # MWS Documents
    ["33", "MWS", "MWS Scope of Work", "MWS engagement letter", "Mandatory", "MWS (Sterling)", "3 days"],
    ["34", "MWS", "Document Appraisal Sheet (DAS)", "MWS approval certificate", "Mandatory", "MWS (Sterling)", "5 days"],
    ["35", "MWS", "MWS Attendance Confirmation", "On-site attendance confirmation", "Mandatory", "MWS (Sterling)", "1 day"],
    
    # Gate Pass - Mina Zayed Port (Equipment)
    ["36", "Gate Pass-Equipment", "SPMT Gate Pass Application", "SPMT entry/exit permit (MZP)", "Mandatory", "OFCO Agency", "2 days"],
    ["37", "Gate Pass-Equipment", "Trailer Gate Pass Application", "Trailer entry/exit permit (MZP)", "Mandatory", "OFCO Agency", "2 days"],
    ["38", "Gate Pass-Equipment", "Crane Gate Pass Application", "Crane entry/exit permit (MZP)", "Important", "OFCO Agency", "2 days"],
    ["39", "Gate Pass-Equipment", "Equipment Insurance Certificate", "Equipment insurance proof", "Mandatory", "Mammoet", "N/A"],
    ["40", "Gate Pass-Equipment", "Equipment Registration", "Equipment registration docs", "Mandatory", "Mammoet", "N/A"],
    
    # Gate Pass - Mina Zayed Port (Personnel)
    ["41", "Gate Pass-Personnel", "Personnel Gate Pass Application", "Staff entry permit (MZP)", "Mandatory", "OFCO Agency", "2 days"],
    ["42", "Gate Pass-Personnel", "Personnel ID Copies (Emirates ID)", "ID copies for all personnel", "Mandatory", "Mammoet", "1 day"],
    ["43", "Gate Pass-Personnel", "Personnel ID Copies (Emirates ID)", "ID copies for all personnel", "Mandatory", "Samsung C&T", "1 day"],
    ["44", "Gate Pass-Personnel", "Personnel ID Copies (Emirates ID)", "ID copies for all personnel", "Mandatory", "DSV Solutions", "1 day"],
    ["45", "Gate Pass-Personnel", "HSE Training Certificates", "Safety training certificates", "Mandatory", "Mammoet", "N/A"],
    ["46", "Gate Pass-Personnel", "Welder Certificates", "Welder qualifications", "Mandatory", "Mammoet", "N/A"],
    ["47", "Gate Pass-Personnel", "Vessel Crew ID List", "Crew list for port access", "Mandatory", "Vessel Owner (KFS)", "1 day"],
    
    # Customs & Port
    ["48", "Customs & Port", "Pre-Arrival Cargo Declaration", "Maqta Gateway submission", "Mandatory", "OFCO Agency", "2 days"],
    ["49", "Customs & Port", "Berth Booking Confirmation", "RoRo Jetty Berth allocation", "Mandatory", "ADNOC L&S", "3 days"],
    ["50", "Customs & Port", "Tide Table (Current Month)", "Tide data for load-out timing", "Mandatory", "ADNOC L&S", "N/A"],
]

priority_fills = {
    "Mandatory": CRITICAL_FILL,
    "Important": IMPORTANT_FILL,
    "Standard": STANDARD_FILL
}

category_fills = {
    "Gate Pass-Equipment": GATE_PASS_FILL,
    "Gate Pass-Personnel": GATE_PASS_FILL,
}

row = 5
for doc in documents:
    for col, val in enumerate(doc, 1):
        cell = ws_main.cell(row=row, column=col, value=val)
        cell.alignment = center_align if col in [1, 5, 7] else left_align
        cell.border = thin_border
        
        # Priority coloring
        if col == 5:
            cell.fill = priority_fills.get(val, PatternFill())
            cell.font = BOLD_FONT
        
        # Category coloring
        if col == 2 and val in category_fills:
            cell.fill = category_fills[val]
    
    # Add empty status columns for each voyage
    for status_col in [8, 11, 14, 17]:  # V1, V2, V3, V4 status columns
        cell = ws_main.cell(row=row, column=status_col, value="Not Started")
        cell.alignment = center_align
        cell.border = thin_border
    
    for date_col in [9, 12, 15, 18]:  # Date columns
        cell = ws_main.cell(row=row, column=date_col, value="")
        cell.alignment = center_align
        cell.border = thin_border
    
    for remarks_col in [10, 13, 16, 19]:  # Remarks columns
        cell = ws_main.cell(row=row, column=remarks_col, value="")
        cell.alignment = left_align
        cell.border = thin_border
    
    row += 1

# Add Data Validation (Status dropdown)
status_validation = DataValidation(
    type="list",
    formula1='"Complete,In Progress,Not Started,N/A,Pending Review"',
    allow_blank=True
)
status_validation.error = "Please select from list"
status_validation.errorTitle = "Invalid Status"
ws_main.add_data_validation(status_validation)

# Apply validation to status columns (V1, V2, V3, V4)
for col in [8, 11, 14, 17]:
    for r in range(5, row):
        status_validation.add(ws_main.cell(row=r, column=col))

# Column widths
col_widths_main = [4, 16, 32, 28, 10, 18, 10, 12, 12, 20, 12, 12, 20, 12, 12, 20, 12, 12, 20]
for i, width in enumerate(col_widths_main, 1):
    if i <= 19:
        ws_main.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws_main.freeze_panes = 'H5'

# ==================== SHEET 3: VOYAGE 1 DETAIL ====================
def create_voyage_sheet(wb, voyage_num, tr_units, dates):
    ws = wb.create_sheet(f"Voyage_{voyage_num}")
    
    ws.merge_cells('A1:J2')
    ws[f'A1'] = f"Voyage {voyage_num}: {tr_units}"
    ws[f'A1'].font = Font(bold=True, size=16, color="1E3A5F")
    ws[f'A1'].alignment = center_align
    
    # Schedule info
    row = 4
    ws[f'A{row}'] = "Schedule Summary:"
    ws[f'A{row}'].font = BOLD_FONT
    
    row += 1
    schedule_items = [
        ("MZP Arrival:", dates[0]),
        ("Load-out:", dates[1]),
        ("MZP Departure:", dates[2]),
        ("AGI Arrival:", dates[3]),
        ("Document Deadline:", dates[4]),
    ]
    
    for label, val in schedule_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = val
        ws[f'B{row}'].font = BOLD_FONT
        row += 1
    
    # Document checklist header
    row += 1
    headers = ["No", "Category", "Document", "Responsible", "Priority", "Status", "Submit Date", "Received Date", "Remarks"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center_align
        cell.border = thin_border
    
    # Add validation
    status_val = DataValidation(
        type="list",
        formula1='"Complete,In Progress,Not Started,N/A,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(status_val)
    
    # Column widths
    widths = [4, 16, 32, 18, 10, 14, 12, 12, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    return ws

voyage_info = [
    ("1", "TR Units 1-2", ["01-27", "01-29~30", "02-01", "02-02", "01-23"]),
    ("2", "TR Units 3-4", ["02-06", "02-07~08", "02-10", "02-11", "02-03"]),
    ("3", "TR Units 5-6", ["02-15", "02-16~17", "02-19", "02-20", "02-12"]),
    ("4 (Final)", "TR Unit 7", ["02-24", "02-25", "02-27", "02-28", "02-20"]),
]

for v_num, tr_units, dates in voyage_info:
    create_voyage_sheet(wb, v_num, tr_units, dates)

# ==================== SHEET: GATE PASS TRACKER ====================
ws_gate = wb.create_sheet("Gate_Pass_MZP")

ws_gate.merge_cells('A1:K2')
ws_gate['A1'] = "🚪 Mina Zayed Port - Gate Pass Tracker (Equipment & Personnel)"
ws_gate['A1'].font = Font(bold=True, size=16, color="1E3A5F")
ws_gate['A1'].alignment = center_align

# Equipment Section
row = 4
ws_gate.merge_cells(f'A{row}:K{row}')
ws_gate[f'A{row}'] = "🚛 EQUIPMENT GATE PASS"
ws_gate[f'A{row}'].font = WHITE_FONT
ws_gate[f'A{row}'].fill = PatternFill("solid", fgColor="4A90D9")
ws_gate[f'A{row}'].alignment = center_align

row += 1
equip_headers = ["No", "Equipment Type", "Qty", "Registration No", "Owner/Company", "V1", "V2", "V3", "V4", "Status", "Remarks"]
for col, header in enumerate(equip_headers, 1):
    cell = ws_gate.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

equipment = [
    ["1", "SPMT (24-axle)", "14", "", "Mammoet", "", "", "", "", "", ""],
    ["2", "PPU (Power Pack Unit)", "2", "", "Mammoet", "", "", "", "", "", ""],
    ["3", "Trailer (Transport Beam)", "1", "", "Mammoet", "", "", "", "", "", ""],
    ["4", "Forklift", "2", "", "Mammoet", "", "", "", "", "", ""],
    ["5", "Crane (Mobile)", "1", "", "Mammoet", "", "", "", "", "", ""],
    ["6", "Support Vehicle", "3", "", "Mammoet", "", "", "", "", "", ""],
    ["7", "Surveyor Vehicle", "1", "", "MWS (Sterling)", "", "", "", "", "", ""],
]

for eq in equipment:
    row += 1
    for col, val in enumerate(eq, 1):
        cell = ws_gate.cell(row=row, column=col, value=val)
        cell.alignment = center_align if col <= 5 else center_align
        cell.border = thin_border

# Personnel Section
row += 3
ws_gate.merge_cells(f'A{row}:K{row}')
ws_gate[f'A{row}'] = "👷 PERSONNEL GATE PASS"
ws_gate[f'A{row}'].font = WHITE_FONT
ws_gate[f'A{row}'].fill = PatternFill("solid", fgColor="5BA55B")
ws_gate[f'A{row}'].alignment = center_align

row += 1
person_headers = ["No", "Name", "Company", "Role", "Emirates ID", "V1", "V2", "V3", "V4", "Status", "Remarks"]
for col, header in enumerate(person_headers, 1):
    cell = ws_gate.cell(row=row, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor="3D8B3D")
    cell.alignment = center_align
    cell.border = thin_border

# Sample personnel rows (to be filled)
for i in range(1, 21):
    row += 1
    ws_gate.cell(row=row, column=1, value=str(i)).alignment = center_align
    ws_gate.cell(row=row, column=1).border = thin_border
    for col in range(2, 12):
        cell = ws_gate.cell(row=row, column=col, value="")
        cell.border = thin_border

# Column widths
gate_widths = [4, 24, 18, 16, 16, 8, 8, 8, 8, 12, 20]
for i, w in enumerate(gate_widths, 1):
    ws_gate.column_dimensions[get_column_letter(i)].width = w

# Add validations for Gate Pass sheet
gate_status_val = DataValidation(
    type="list",
    formula1='"Applied,Approved,Pending,Rejected,N/A"',
    allow_blank=True
)
ws_gate.add_data_validation(gate_status_val)
for r in range(6, 35):
    gate_status_val.add(ws_gate.cell(row=r, column=10))

# ==================== SHEET: INSTRUCTIONS ====================
ws_inst = wb.create_sheet("Instructions")

ws_inst.merge_cells('A1:F2')
ws_inst['A1'] = "📋 Instructions & Legend"
ws_inst['A1'].font = Font(bold=True, size=16, color="1E3A5F")
ws_inst['A1'].alignment = center_align

instructions = [
    "",
    "▶ Status Options:",
    "   • Complete - Document submitted and approved",
    "   • In Progress - Document being prepared",
    "   • Not Started - Work not yet begun",
    "   • Pending Review - Submitted, awaiting approval",
    "   • N/A - Not applicable for this voyage",
    "",
    "▶ Priority Legend:",
    "   • Mandatory (Red) - Must be completed before operation",
    "   • Important (Orange) - Strongly recommended",
    "   • Standard (Green) - Good to have",
    "",
    "▶ Key Deadlines:",
    "   • All documents should be submitted 3-4 days before MZP arrival",
    "   • Land Permit (SPMT): Allow 2-3 business days for approval",
    "   • Gate Pass: Apply 2 days before entry required",
    "",
    "▶ Responsible Party Abbreviations:",
    "   • KFS = Khalid Faraj Shipping (Vessel Owner)",
    "   • MWS = Marine Warranty Survey (Sterling Technical)",
    "   • OFCO = OFCO Agency",
    "   • MMT = Mammoet",
    "",
    "▶ Voyage Schedule (2026):",
    "   • Voyage 1: MZP 01-27 → AGI 02-02 (TR 1-2)",
    "   • Voyage 2: MZP 02-06 → AGI 02-11 (TR 3-4) ⚡ Parallel",
    "   • Voyage 3: MZP 02-15 → AGI 02-20 (TR 5-6)",
    "   • Voyage 4: MZP 02-24 → AGI 02-28 (TR 7) Final",
    "",
    "▶ Key Contacts:",
    "   OFCO Agency: nkk@ofco-int.com (Nanda Kumar)",
    "   ADNOC L&S: moda@adnoc.ae (Mahmoud Ouda)",
    "   Mammoet: Yulia.Frolova@mammoet.com",
    "   DSV: jay.manaloto@dsv.com",
    "",
    "??VBA Macro Installation (Manual Method if Auto-Embed Fails):",
    "   Step 1: Save this file as .xlsm (Excel Macro-Enabled Workbook)",
    "   Step 2: Press Alt+F11 to open VBA Editor",
    "   Step 3: Enable 'Trust access to the VBA project object model'",
    "          (File > Options > Trust Center > Trust Center Settings >",
    "           Macro Settings > Trust access to the VBA project object model)",
    "   Step 4: Go to Insert > Module in VBA Editor",
    "   Step 5: Open the 'VBA_Code' sheet in this workbook",
    "   Step 6: Select column A from row 3 onwards (Ctrl+Shift+End)",
    "   Step 7: Copy all VBA code (Ctrl+C)",
    "   Step 8: Paste into the VBA module window (Ctrl+V)",
    "   Step 9: Save the workbook (Ctrl+S)",
    "   Step 10: Run macros from Alt+F8 or Developer > Macros",
    "   Note: If automatic embedding succeeded, you already have .xlsm file",
]

for i, text in enumerate(instructions, 4):
    ws_inst[f'A{i}'] = text
    if text.startswith("▶"):
        ws_inst[f'A{i}'].font = BOLD_FONT

ws_inst.column_dimensions['A'].width = 60

# ==================== ADD CONDITIONAL FORMATTING ====================
from openpyxl.formatting.rule import CellIsRule

# Status conditional formatting rules for Document_Tracker sheet
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

# Apply to status columns (H, K, N, Q)
status_ranges = ['H5:H60', 'K5:K60', 'N5:N60', 'Q5:Q60']

for status_range in status_ranges:
    ws_main.conditional_formatting.add(status_range, 
        CellIsRule(operator='equal', formula=['"Complete"'], fill=green_fill))
    ws_main.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"In Progress"'], fill=yellow_fill))
    ws_main.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Not Started"'], fill=red_fill))
    ws_main.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Pending Review"'], fill=blue_fill))
    ws_main.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"N/A"'], fill=gray_fill))

# Add print area and page setup
ws_main.print_area = 'A1:S60'
ws_main.page_setup.orientation = 'landscape'
ws_main.page_setup.fitToWidth = 1
ws_main.page_setup.fitToHeight = 0
ws_main.print_title_rows = '1:4'

# Dashboard page setup
ws_dash.page_setup.orientation = 'landscape'
ws_dash.page_setup.fitToPage = True

# Gate Pass page setup
ws_gate.page_setup.orientation = 'landscape'

# ==================== SUMMARY STATISTICS SHEET ====================
ws_stats = wb.create_sheet("Progress_Stats")

ws_stats.merge_cells('A1:F2')
ws_stats['A1'] = "📊 Document Submission Progress Statistics"
ws_stats['A1'].font = Font(bold=True, size=16, color="1E3A5F")
ws_stats['A1'].alignment = center_align

# Add formulas for automatic calculation
row = 4
ws_stats['A4'] = "Category"
ws_stats['B4'] = "Total"
ws_stats['C4'] = "Complete"
ws_stats['D4'] = "In Progress"
ws_stats['E4'] = "Not Started"
ws_stats['F4'] = "Progress %"

for col in range(1, 7):
    cell = ws_stats.cell(row=4, column=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = center_align
    cell.border = thin_border

categories = ["PTW & Loading", "AD Maritime NOC", "Engineering", "Vessel", "MWS", 
              "Gate Pass-Equipment", "Gate Pass-Personnel", "Customs & Port"]

for i, cat in enumerate(categories):
    row = 5 + i
    ws_stats.cell(row=row, column=1, value=cat).border = thin_border
    ws_stats.cell(row=row, column=1).font = BOLD_FONT
    for col in range(2, 7):
        cell = ws_stats.cell(row=row, column=col, value="-")
        cell.alignment = center_align
        cell.border = thin_border

# Column widths
for i, w in enumerate([20, 10, 10, 12, 12, 12], 1):
    ws_stats.column_dimensions[get_column_letter(i)].width = w

# Save workbook
script_dir = os.path.dirname(os.path.abspath(__file__))

vba_module_candidates = [
    os.path.join(script_dir, "TR_DocTracker_VBA_Module.bas"),
    os.path.join(script_dir, "TR_DocTracker_VBA_Module_1.bas"),
    os.path.join(os.path.dirname(script_dir), "TR_DocTracker_VBA_Module.bas"),
    os.path.join(os.path.dirname(script_dir), "TR_DocTracker_VBA_Module_1.bas"),
    os.path.join(os.path.expanduser("~"), "Downloads", "TR_DocTracker_VBA_Module.bas"),
    os.path.join(os.path.expanduser("~"), "Downloads", "TR_DocTracker_VBA_Module_1.bas"),
    r"C:\Users\SAMSUNG\Downloads\TR_DocTracker_VBA_Module.bas",
    r"C:\Users\SAMSUNG\Downloads\TR_DocTracker_VBA_Module_1.bas",
]

# ==================== SHEET: VBA CODE (MANUAL COPY) ====================
ws_vba = wb.create_sheet("VBA_Code")
ws_vba.merge_cells('A1:B1')
ws_vba['A1'] = "VBA Module Code - Copy All Content Below for Manual Installation"
ws_vba['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_vba['A1'].alignment = center_align
ws_vba['A1'].fill = HEADER_FILL

ws_vba['A2'] = "Instructions: Select column A from row 3 onwards, copy all, and paste into a new VBA module (Insert > Module)"
ws_vba['A2'].font = Font(italic=True, size=10)
ws_vba['A2'].alignment = left_align

vba_file_path = None
for candidate in vba_module_candidates:
    if os.path.isfile(candidate):
        vba_file_path = candidate
        break

if vba_file_path:
    try:
        with open(vba_file_path, "r", encoding="utf-8", errors="ignore") as handle:
            vba_lines = handle.readlines()
        row = 3
        for line in vba_lines:
            cell = ws_vba.cell(row=row, column=1, value=line.rstrip("\r\n"))
            cell.alignment = left_align
            cell.font = Font(name="Courier New", size=9)
            row += 1
    except OSError as exc:
        ws_vba['A3'] = f"Error reading VBA file: {exc}"
        ws_vba['A3'].font = Font(color="FF0000")
else:
    ws_vba['A3'] = "VBA module file not found. Please locate TR_DocTracker_VBA_Module_1.bas"
    ws_vba['A3'].font = Font(color="FF0000")
    ws_vba['A4'] = "Expected locations searched:"
    row = 5
    for candidate in vba_module_candidates[:8]:
        cell = ws_vba.cell(row=row, column=1, value=f"  - {candidate}")
        cell.font = Font(size=9, italic=True)
        row += 1

ws_vba.column_dimensions['A'].width = 120
ws_vba.column_dimensions['B'].width = 10

output_xlsx = os.path.join(script_dir, "TR_Document_Tracker_VBA.xlsx")
output_xlsm = os.path.join(script_dir, "TR_Document_Tracker_VBA.xlsm")

wb.save(output_xlsx)
print(f"Excel file created: {output_xlsx}")

def get_vba_module_name(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as handle:
            for _ in range(5):
                line = handle.readline()
                if line.startswith("Attribute VB_Name"):
                    parts = line.split('"')
                    if len(parts) >= 2:
                        return parts[1].strip()
    except OSError:
        return None
    return None

vba_module_paths = []
module_names = set()
for candidate in vba_module_candidates:
    if not os.path.isfile(candidate):
        continue
    module_name = get_vba_module_name(candidate) or os.path.basename(candidate)
    if module_name in module_names:
        continue
    module_names.add(module_name)
    vba_module_paths.append(candidate)

def embed_vba_modules(xlsx_path, xlsm_path, module_paths):
    if not module_paths:
        print("No VBA module files found. Skipping VBA embed.")
        return False
    try:
        import win32com.client as win32
    except ImportError:
        print("pywin32 not installed. Skipping VBA embed.")
        return False

    xlsx_path = os.path.abspath(xlsx_path)
    xlsm_path = os.path.abspath(xlsm_path)
    if not os.path.isfile(xlsx_path):
        print(f"Workbook not found: {xlsx_path}")
        return False

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb_com = None
    try:
        for attempt in range(3):
            try:
                wb_com = excel.Workbooks.Open(Filename=xlsx_path, ReadOnly=False)
                break
            except Exception as exc:
                if attempt == 2:
                    raise
                time.sleep(1)
        vbproj = wb_com.VBProject
        for module_path in module_paths:
            vbproj.VBComponents.Import(module_path)
        wb_com.SaveAs(xlsm_path, FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
        return True
    except Exception as exc:
        print(f"VBA embed failed: {exc}")
        return False
    finally:
        if wb_com is not None:
            wb_com.Close(SaveChanges=False)
        excel.Quit()

if embed_vba_modules(output_xlsx, output_xlsm, vba_module_paths):
    print(f"Excel macro-enabled file created: {output_xlsm}")
