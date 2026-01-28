#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 파일 구조 분석 - Python 구현 가이드 문서 생성

Excel 파일의 모든 시트, 포맷, 스타일, 함수를 분석하여
Python 구현을 위한 가이드 문서를 생성합니다.
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border
from collections import defaultdict
import re
import io

# UTF-8 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


def analyze_cell_formatting(cell):
    """셀 포맷 정보 추출"""
    fmt = {
        'value': None,
        'data_type': None,
        'formula': None,
        'font': None,
        'fill': None,
        'alignment': None,
        'border': None,
        'number_format': None
    }
    
    if cell.value is not None:
        fmt['value'] = str(cell.value)
    
    fmt['data_type'] = cell.data_type
    
    # 함수 추출
    if cell.data_type == 'f':
        if hasattr(cell.value, 'text'):
            fmt['formula'] = cell.value.text
        else:
            fmt['formula'] = str(cell.value) if cell.value else None
    
    # 폰트 정보
    if cell.font:
        fmt['font'] = {
            'name': cell.font.name,
            'size': cell.font.size,
            'bold': cell.font.bold,
            'italic': cell.font.italic,
            'underline': cell.font.underline,
            'color': str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None
        }
    
    # 배경색
    if cell.fill and cell.fill.patternType:
        fmt['fill'] = {
            'patternType': cell.fill.patternType,
            'fgColor': str(cell.fill.fgColor.rgb) if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None,
            'bgColor': str(cell.fill.bgColor.rgb) if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor and hasattr(cell.fill.bgColor, 'rgb') else None
        }
    
    # 정렬
    if cell.alignment:
        fmt['alignment'] = {
            'horizontal': cell.alignment.horizontal,
            'vertical': cell.alignment.vertical,
            'wrap_text': cell.alignment.wrap_text,
            'text_rotation': cell.alignment.text_rotation
        }
    
    # 테두리
    if cell.border:
        fmt['border'] = {
            'left': str(cell.border.left.style) if cell.border.left else None,
            'right': str(cell.border.right.style) if cell.border.right else None,
            'top': str(cell.border.top.style) if cell.border.top else None,
            'bottom': str(cell.border.bottom.style) if cell.border.bottom else None
        }
    
    # 숫자 포맷
    if cell.number_format:
        fmt['number_format'] = cell.number_format
    
    return fmt


def extract_functions_from_formula(formula):
    """함수에서 사용된 Excel 함수 추출"""
    if not formula:
        return []
    
    # ArrayFormula 객체 처리
    if hasattr(formula, 'text'):
        formula_str = formula.text
    elif isinstance(formula, str):
        formula_str = formula
    else:
        formula_str = str(formula)
    
    if not formula_str.startswith('='):
        return []
    
    # Excel 함수 패턴: 함수명(인자)
    pattern = r'([A-Z][A-Z0-9_]*)\s*\('
    functions = re.findall(pattern, formula_str)
    return functions


def analyze_excel_file(excel_path: str):
    """Excel 파일 전체 분석"""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    print(f"Analyzing Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=False)
    
    analysis = {
        'file_name': excel_path.name,
        'sheets': [],
        'all_functions': defaultdict(list),
        'all_formats': defaultdict(int),
        'all_styles': defaultdict(int)
    }
    
    # 각 시트 분석
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nAnalyzing sheet: {sheet_name}")
        
        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'column_widths': {},
            'row_heights': {},
            'merged_cells': [],
            'formulas': [],
            'data_types': defaultdict(int),
            'number_formats': defaultdict(int),
            'sample_cells': []
        }
        
        # 열 너비
        for col_letter in ws.column_dimensions:
            if ws.column_dimensions[col_letter].width:
                sheet_info['column_widths'][col_letter] = ws.column_dimensions[col_letter].width
        
        # 행 높이
        for row_num in ws.row_dimensions:
            if ws.row_dimensions[row_num].height:
                sheet_info['row_heights'][row_num] = ws.row_dimensions[row_num].height
        
        # 병합된 셀
        for merged in ws.merged_cells.ranges:
            sheet_info['merged_cells'].append(str(merged))
        
        # 셀 분석 (샘플링: 처음 1000개 셀)
        cell_count = 0
        sample_cells = []
        
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value is None and not cell.data_type:
                    continue
                
                cell_count += 1
                if cell_count <= 100:  # 처음 100개 셀만 상세 분석
                    cell_info = analyze_cell_formatting(cell)
                    cell_info['coordinate'] = cell.coordinate
                    sample_cells.append(cell_info)
                    
                    # 함수 추출
                    if cell_info['formula']:
                        functions = extract_functions_from_formula(cell_info['formula'])
                        for func in functions:
                            analysis['all_functions'][func].append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'formula': cell_info['formula']
                            })
                    
                    # 포맷 통계
                    if cell_info['number_format']:
                        analysis['all_formats'][cell_info['number_format']] += 1
                
                # 데이터 타입 통계
                if cell.data_type:
                    sheet_info['data_types'][cell.data_type] += 1
                
                # 함수 추출 (모든 셀)
                if cell.data_type == 'f' and cell.value:
                    formula = cell.value
                    if hasattr(formula, 'text'):
                        formula_str = formula.text
                    else:
                        formula_str = str(formula) if formula else ''
                    sheet_info['formulas'].append({
                        'cell': cell.coordinate,
                        'formula': formula_str
                    })
        
        sheet_info['sample_cells'] = sample_cells[:20]  # 처음 20개만 저장
        sheet_info['total_cells_analyzed'] = cell_count
        
        analysis['sheets'].append(sheet_info)
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}, Formulas: {len(sheet_info['formulas'])}")
    
    return analysis


def generate_guide_document(analysis, output_path: str):
    """가이드 문서 생성 (Markdown)"""
    output_path = Path(output_path)
    
    md_content = []
    md_content.append("# Excel to Python Implementation Guide\n")
    md_content.append(f"**Source File:** `{analysis['file_name']}`\n")
    md_content.append(f"**Generated:** {Path(__file__).name}\n")
    md_content.append("\n---\n")
    
    # 개요
    md_content.append("## Overview\n")
    md_content.append(f"- **Total Sheets:** {len(analysis['sheets'])}\n")
    md_content.append(f"- **Total Functions Used:** {len(analysis['all_functions'])}\n")
    md_content.append(f"- **Total Number Formats:** {len(analysis['all_formats'])}\n")
    md_content.append("\n### Sheet List\n")
    for sheet in analysis['sheets']:
        md_content.append(f"- **{sheet['name']}**: {sheet['max_row']} rows × {sheet['max_column']} columns")
        if sheet['formulas']:
            md_content.append(f"  - Formulas: {len(sheet['formulas'])}")
        md_content.append("")
    
    md_content.append("\n---\n")
    
    # 각 시트 상세 분석
    for sheet in analysis['sheets']:
        md_content.append(f"## Sheet: {sheet['name']}\n")
        md_content.append(f"- **Dimensions:** {sheet['max_row']} rows × {sheet['max_column']} columns\n")
        
        # 열 너비
        if sheet['column_widths']:
            md_content.append("### Column Widths\n")
            md_content.append("| Column | Width |\n|--------|-------|\n")
            for col, width in sorted(sheet['column_widths'].items()):
                md_content.append(f"| {col} | {width:.2f} |\n")
            md_content.append("\n")
        
        # 병합된 셀
        if sheet['merged_cells']:
            md_content.append("### Merged Cells\n")
            for merged in sheet['merged_cells'][:20]:  # 최대 20개
                md_content.append(f"- `{merged}`\n")
            if len(sheet['merged_cells']) > 20:
                md_content.append(f"- ... and {len(sheet['merged_cells']) - 20} more\n")
            md_content.append("\n")
        
        # 함수 목록
        if sheet['formulas']:
            md_content.append("### Formulas\n")
            md_content.append(f"**Total:** {len(sheet['formulas'])}\n\n")
            md_content.append("| Cell | Formula |\n|------|---------|\n")
            for f in sheet['formulas'][:50]:  # 최대 50개
                formula = f['formula']
                if hasattr(formula, 'text'):
                    formula_str = formula.text
                else:
                    formula_str = str(formula) if formula else ''
                formula_display = formula_str.replace('|', '\\|')[:100]  # 최대 100자
                md_content.append(f"| {f['cell']} | `{formula_display}` |\n")
            if len(sheet['formulas']) > 50:
                md_content.append(f"| ... | ... ({len(sheet['formulas']) - 50} more) |\n")
            md_content.append("\n")
        
        # 샘플 셀 포맷
        if sheet['sample_cells']:
            md_content.append("### Sample Cell Formatting\n")
            md_content.append("| Cell | Value | Type | Formula | Font | Fill | Alignment |\n")
            md_content.append("|------|-------|------|---------|------|------|-----------|\n")
            for cell in sheet['sample_cells'][:10]:  # 최대 10개
                value = str(cell.get('value', ''))[:30].replace('|', '\\|')
                data_type = cell.get('data_type', '')
                formula = cell.get('formula', '')[:30].replace('|', '\\|') if cell.get('formula') else ''
                font_info = f"Bold:{cell['font']['bold']}" if cell.get('font') else ''
                fill_info = cell['fill']['fgColor'][:10] if cell.get('fill') and cell['fill'].get('fgColor') else ''
                align = cell['alignment']['horizontal'] if cell.get('alignment') else ''
                md_content.append(f"| {cell['coordinate']} | {value} | {data_type} | {formula} | {font_info} | {fill_info} | {align} |\n")
            md_content.append("\n")
        
        md_content.append("---\n")
    
    # 사용된 함수 목록
    if analysis['all_functions']:
        md_content.append("## Excel Functions Used\n")
        md_content.append("| Function | Count | Locations |\n|----------|-------|-----------|\n")
        for func, locations in sorted(analysis['all_functions'].items(), key=lambda x: len(x[1]), reverse=True):
            count = len(locations)
            sample_locs = ', '.join([f"{loc['sheet']}!{loc['cell']}" for loc in locations[:5]])
            if count > 5:
                sample_locs += f" ... ({count - 5} more)"
            md_content.append(f"| `{func}` | {count} | {sample_locs} |\n")
        md_content.append("\n")
    
    # 숫자 포맷
    if analysis['all_formats']:
        md_content.append("## Number Formats Used\n")
        md_content.append("| Format | Count |\n|--------|-------|\n")
        for fmt, count in sorted(analysis['all_formats'].items(), key=lambda x: x[1], reverse=True):
            md_content.append(f"| `{fmt}` | {count} |\n")
        md_content.append("\n")
    
    # Python 구현 가이드
    md_content.append("---\n")
    md_content.append("## Python Implementation Guide\n")
    md_content.append("\n### 1. Data Structure\n")
    md_content.append("```python\n")
    md_content.append("# Recommended data structure\n")
    md_content.append("from dataclasses import dataclass\n")
    md_content.append("from typing import Dict, List, Any\n\n")
    md_content.append("@dataclass\n")
    md_content.append("class SheetData:\n")
    md_content.append("    name: str\n")
    md_content.append("    rows: int\n")
    md_content.append("    cols: int\n")
    md_content.append("    data: List[List[Any]]\n")
    md_content.append("    formulas: Dict[str, str]  # cell_coordinate -> formula\n")
    md_content.append("    formats: Dict[str, Dict]  # cell_coordinate -> format_info\n\n")
    md_content.append("@dataclass\n")
    md_content.append("class ExcelWorkbook:\n")
    md_content.append("    sheets: Dict[str, SheetData]\n")
    md_content.append("```\n\n")
    
    md_content.append("### 2. Function Mapping\n")
    md_content.append("| Excel Function | Python Equivalent | Notes |\n")
    md_content.append("|----------------|-------------------|-------|\n")
    
    # 일반적인 함수 매핑
    function_mapping = {
        'SUM': 'sum()',
        'AVERAGE': 'statistics.mean()',
        'COUNT': 'len()',
        'IF': 'if/else or ternary operator',
        'VLOOKUP': 'dict lookup or pandas merge',
        'INDEX': 'list/dict indexing',
        'MATCH': 'list.index() or enumerate',
        'CONCATENATE': 'str.join() or f-strings',
        'TEXT': 'str.format() or f-strings',
        'DATE': 'datetime.date()',
        'TODAY': 'datetime.date.today()',
        'NOW': 'datetime.datetime.now()'
    }
    
    for excel_func in sorted(analysis['all_functions'].keys()):
        python_equiv = function_mapping.get(excel_func, 'TBD - needs implementation')
        md_content.append(f"| `{excel_func}` | `{python_equiv}` | Used in {len(analysis['all_functions'][excel_func])} cells |\n")
    
    md_content.append("\n### 3. Implementation Steps\n")
    md_content.append("\n1. **Load Data Structure**\n")
    md_content.append("   - Read all sheets\n")
    md_content.append("   - Extract cell values and formulas\n")
    md_content.append("   - Store formatting information\n\n")
    
    md_content.append("2. **Implement Functions**\n")
    md_content.append("   - Map Excel functions to Python equivalents\n")
    md_content.append("   - Handle cell references (e.g., A1, $B$2)\n")
    md_content.append("   - Implement calculation engine\n\n")
    
    md_content.append("3. **Apply Formatting**\n")
    md_content.append("   - Recreate cell styles (font, fill, alignment)\n")
    md_content.append("   - Apply number formats\n")
    md_content.append("   - Handle merged cells\n\n")
    
    md_content.append("4. **Output Generation**\n")
    md_content.append("   - Generate Excel file (openpyxl)\n")
    md_content.append("   - Or generate other formats (CSV, JSON, etc.)\n\n")
    
    # 저장
    output_path.write_text('\n'.join(md_content), encoding='utf-8')
    print(f"\nGuide document saved: {output_path}")


def main():
    """메인 함수"""
    if len(sys.argv) < 2:
        print("Usage: python analyze_excel_structure.py <excel_file.xlsx> [output.md]")
        print("\nExample:")
        print("  python analyze_excel_structure.py content-calendar.xlsx")
        print("  python analyze_excel_structure.py content-calendar.xlsx guide.md")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else Path(excel_path).with_suffix('.guide.md')
    
    try:
        analysis = analyze_excel_file(excel_path)
        generate_guide_document(analysis, output_path)
        
        # JSON도 저장 (상세 데이터)
        json_path = Path(output_path).with_suffix('.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(analysis, f, indent=2, ensure_ascii=False, default=str)
        print(f"Detailed analysis (JSON) saved: {json_path}")
        
        print(f"\nSuccess! Guide document: {output_path}")
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

