import datetime as dt
import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# === COLORS ===
COLORS = {
    "HEADER": "1F4E79",
    "SUBHEADER": "2E75B6",
    "MOBILIZATION": "8E7CC3",
    "DECK_PREP": "6FA8DC",
    "LOADOUT": "93C47D",
    "SEAFAST": "76A5AF",
    "SAIL": "A4C2F4",
    "AGI_UNLOAD": "F6B26B",
    "TURNING": "FFD966",
    "JACKDOWN": "E06666",
    "RETURN": "999999",
    "BUFFER": "D9D9D9",
    "MILESTONE": "FF0000",
    "SHAMAL": "FF9800",
    "INPUT": "FFFDE7",
    "FORMULA": "E3F2FD",
    # Phase mappings for TSV data
    "Governance": "8E7CC3",
    "Contract/Cost": "6FA8DC",
    "Port Condition": "93C47D",
    "Resources": "76A5AF",
    "Resources/Cost": "76A5AF",
    "Base Docs": "A4C2F4",
    "Engineering": "F6B26B",
    "HSE/Risk": "FFD966",
    "HSE": "FFD966",
    "Mandatory Docs": "E06666",
    "Review": "999999",
    "Submission": "D9D9D9",
    "Sign-off": "8E7CC3",
    "Sailing / Pilotage": "A4C2F4",
    "PTW": "FFD966",
    "Access / Gate Pass": "93C47D",
    "Verification/Approval": "6FA8DC",
    "Final Approval": "E06666",
    "Execution": "93C47D",
}

BORDER = Side(style="thin", color="A6A6A6")
def tb(): return Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

def parse_tsv(tsv_path):
    """TSV 파일을 읽어서 작업 목록 반환"""
    tasks = []
    with open(tsv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            # 날짜 파싱
            try:
                start_date = dt.datetime.strptime(row['Start_Date'], '%Y-%m-%d').date()
                due_date_str = row['Due_Date']
                if 'Fixed' in due_date_str:
                    # "Fixed (2026-01-23)" 형식 처리
                    due_date = dt.datetime.strptime(due_date_str.split('(')[1].split(')')[0], '%Y-%m-%d').date()
                else:
                    due_date = dt.datetime.strptime(due_date_str, '%Y-%m-%d').date()
                
                duration = int(float(row['Duration_days']))
                
                tasks.append({
                    'WBS': row['WBS'],
                    'Phase': row['Phase'],
                    'Site': row['Site'],
                    'Task': row['Task/Document'],
                    'Owner': row['Owner'],
                    'Approver': row['Approver'],
                    'Predecessor': row['Predecessor'],
                    'Start_Tag': row['Start_Tag'],
                    'Start_Date': start_date,
                    'Due_Tag': row['Due_Tag'],
                    'Due_Date': due_date,
                    'Duration_days': duration,
                    'Risk': row['Risk'],
                    'Evidence': row['Evidence'],
                })
            except Exception as e:
                print(f"Warning: Could not parse row {row.get('WBS', 'unknown')}: {e}")
                continue
    return tasks

def create_gantt_from_tsv(tsv_path):
    """TSV 파일로부터 Gantt 차트 Excel 생성"""
    tasks = parse_tsv(tsv_path)
    
    if not tasks:
        raise ValueError("No tasks found in TSV file")
    
    # 프로젝트 시작일 계산 (가장 이른 시작일)
    project_start = min(task['Start_Date'] for task in tasks)
    project_end = max(task['Due_Date'] for task in tasks)
    total_days = (project_end - project_start).days + 1
    
    wb = Workbook()
    
    # === CONTROL PANEL ===
    ws_ctrl = wb.active
    ws_ctrl.title = "Control_Panel"
    
    # Title
    ws_ctrl.merge_cells("A1:H1")
    ws_ctrl["A1"] = "🎛️ AGI TR 1-6 Transportation - Control Panel"
    ws_ctrl["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_ctrl["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    ws_ctrl["A1"].alignment = Alignment(horizontal="center")
    ws_ctrl.row_dimensions[1].height = 30
    
    ws_ctrl.merge_cells("A2:H2")
    ws_ctrl["A2"] = "📌 시작일(B4)을 변경하면 모든 일정이 자동 업데이트됩니다."
    ws_ctrl["A2"].fill = PatternFill("solid", fgColor="FFF9C4")
    
    # Input Section
    ws_ctrl["A4"] = "📅 프로젝트 시작일:"
    ws_ctrl["A4"].font = Font(bold=True, size=12)
    ws_ctrl["B4"] = project_start
    ws_ctrl["B4"].number_format = "YYYY-MM-DD"
    ws_ctrl["B4"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B4"].border = tb()
    ws_ctrl["B4"].font = Font(bold=True, size=12)
    
    ws_ctrl["A5"] = "🎯 목표 완료일:"
    ws_ctrl["A5"].font = Font(bold=True)
    ws_ctrl["B5"] = project_end
    ws_ctrl["B5"].number_format = "YYYY-MM-DD"
    ws_ctrl["B5"].fill = PatternFill("solid", fgColor=COLORS["INPUT"])
    ws_ctrl["B5"].border = tb()
    
    # Named Ranges
    wb.defined_names["PROJECT_START"] = DefinedName("PROJECT_START", attr_text="Control_Panel!$B$4")
    wb.defined_names["TARGET_END"] = DefinedName("TARGET_END", attr_text="Control_Panel!$B$5")
    
    # Summary Section
    ws_ctrl["A8"] = "📊 자동 계산 요약"
    ws_ctrl["A8"].font = Font(bold=True, size=12)
    
    summary_items = [
        ("A9", "예상 완료일:", "B9", "=MAX(Schedule_Data!G:G)"),
        ("A10", "총 기간 (일):", "B10", "=B9-B4+1"),
        ("A11", "목표 대비:", "B11", '=IF(B9<=B5,"✅ 목표 달성","❌ 지연")'),
        ("A12", "총 작업 수:", "B12", f"={len(tasks)}"),
    ]
    
    for lc, lt, vc, formula in summary_items:
        ws_ctrl[lc] = lt
        ws_ctrl[lc].font = Font(bold=True)
        ws_ctrl[vc] = formula
        ws_ctrl[vc].fill = PatternFill("solid", fgColor=COLORS["FORMULA"])
        ws_ctrl[vc].border = tb()
        if "MAX" in str(formula):
            ws_ctrl[vc].number_format = "YYYY-MM-DD"
    
    # Column widths
    ws_ctrl.column_dimensions["A"].width = 20
    ws_ctrl.column_dimensions["B"].width = 15
    
    # === SCHEDULE DATA ===
    ws_sched = wb.create_sheet("Schedule_Data")
    
    # Title rows
    ws_sched.merge_cells("A1:J1")
    ws_sched["A1"] = "AGI TR 1-6 Transportation Master Schedule"
    ws_sched["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_sched["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    ws_sched.merge_cells("A2:J2")
    ws_sched["A2"] = "Pre-Execution Phase | D-14 to D+1 | Preparation & Approval"
    ws_sched["A2"].font = Font(size=11, color="FFFFFF")
    ws_sched["A2"].fill = PatternFill("solid", fgColor=COLORS["SUBHEADER"])
    
    # Headers (Row 5)
    headers = ["ID", "WBS", "Task", "Phase", "Owner", "Start", "End", "Duration", "Risk", "Evidence"]
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(5, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Task rows
    for r, task in enumerate(tasks, 6):
        # Calculate offset from project start
        offset = (task['Start_Date'] - project_start).days
        
        ws_sched.cell(r, 1, value=task['WBS'])  # ID
        ws_sched.cell(r, 2, value=task['WBS'])  # WBS
        ws_sched.cell(r, 3, value=task['Task'])  # Task
        ws_sched.cell(r, 4, value=task['Phase'])  # Phase
        ws_sched.cell(r, 5, value=task['Owner'])  # Owner
        
        # Start = PROJECT_START + Offset
        ws_sched.cell(r, 6, value=f"=PROJECT_START+{offset}")
        ws_sched.cell(r, 6).number_format = "YYYY-MM-DD"
        
        # Duration
        ws_sched.cell(r, 8, value=task['Duration_days'])
        
        # End = Start + Duration
        ws_sched.cell(r, 7, value=f"=F{r}+H{r}")
        ws_sched.cell(r, 7).number_format = "YYYY-MM-DD"
        
        ws_sched.cell(r, 9, value=task['Risk'])  # Risk
        ws_sched.cell(r, 10, value=task['Evidence'])  # Evidence
        
        # Styling
        phase = task['Phase']
        pc = COLORS.get(phase, "FFFFFF")
        for c in range(1, 11):
            ws_sched.cell(r, c).border = tb()
        ws_sched.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        # Milestone 강조
        if phase == "Milestone":
            for c in range(1, 11):
                ws_sched.cell(r, c).font = Font(bold=True, color="FFFFFF")
                ws_sched.cell(r, c).fill = PatternFill("solid", fgColor=COLORS["MILESTONE"])
        # CRITICAL Risk 강조 (Milestone이 아닌 경우에만)
        elif task['Risk'] == "CRITICAL":
            for c in range(1, 11):
                ws_sched.cell(r, c).font = Font(bold=True, color="B71C1C")
    
    # Column widths
    col_widths = {"A":10, "B":6, "C":45, "D":18, "E":25, "F":12, "G":12, "H":10, "I":10, "J":40}
    for col, w in col_widths.items():
        ws_sched.column_dimensions[col].width = w
    ws_sched.freeze_panes = "A6"
    
    # === GANTT CHART ===
    ws_gantt = wb.create_sheet("Gantt_Chart")
    
    # Title rows
    ws_gantt.merge_cells(f"A1:{get_column_letter(7 + total_days)}1")
    ws_gantt["A1"] = "AGI TR 1-6 Transportation Master Gantt Chart"
    ws_gantt["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_gantt["A1"].fill = PatternFill("solid", fgColor=COLORS["HEADER"])
    
    # Column headers (Row 4)
    meta_headers = ["ID", "WBS", "Task", "Phase", "Start", "End", "Dur"]
    for c, h in enumerate(meta_headers, 1):
        cell = ws_gantt.cell(4, c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb()
    
    # Date columns
    date_col = 8
    for i in range(total_days):
        c = ws_gantt.cell(4, date_col + i, value=f"=PROJECT_START+{i}")
        c.number_format = "D"
        c.font = Font(bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["HEADER"])
        c.alignment = Alignment(horizontal="center")
        c.border = tb()
        ws_gantt.column_dimensions[get_column_letter(date_col + i)].width = 2.5
    
    # Task rows
    for r, task in enumerate(tasks, 5):
        offset = (task['Start_Date'] - project_start).days
        duration = task['Duration_days']
        
        ws_gantt.cell(r, 1, value=f"=Schedule_Data!A{r+1}")
        ws_gantt.cell(r, 2, value=f"=Schedule_Data!B{r+1}")
        ws_gantt.cell(r, 3, value=f"=Schedule_Data!C{r+1}")
        ws_gantt.cell(r, 4, value=f"=Schedule_Data!D{r+1}")
        
        start_cell = ws_gantt.cell(r, 5, value=f"=Schedule_Data!F{r+1}")
        start_cell.number_format = "MM/DD"
        
        end_cell = ws_gantt.cell(r, 6, value=f"=Schedule_Data!G{r+1}")
        end_cell.number_format = "MM/DD"
        
        ws_gantt.cell(r, 7, value=f"=Schedule_Data!H{r+1}")
        
        # Meta columns borders
        for c in range(1, 8):
            ws_gantt.cell(r, c).border = tb()
        
        # Phase color in column D
        phase = task['Phase']
        pc = COLORS.get(phase, "FFFFFF")
        ws_gantt.cell(r, 4).fill = PatternFill("solid", fgColor=pc)
        
        # Fill Gantt bars in date cells (작업 기간에 해당하는 날짜 셀에 색상 채우기)
        for i in range(total_days):
            cell = ws_gantt.cell(r, date_col + i)
            cell.border = tb()
            
            # 작업 기간 내의 날짜 셀에 색상 채우기
            if offset <= i < offset + duration:
                cell.fill = PatternFill("solid", fgColor=pc)
            
            # Milestone인 경우 시작일 셀에 별표 표시
            if phase == "Milestone" and i == offset:
                cell.value = "★"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=COLORS["MILESTONE"])
    
    # Column widths for meta columns
    ws_gantt.column_dimensions["A"].width = 10
    ws_gantt.column_dimensions["B"].width = 5
    ws_gantt.column_dimensions["C"].width = 35
    ws_gantt.column_dimensions["D"].width = 18
    ws_gantt.column_dimensions["E"].width = 7
    ws_gantt.column_dimensions["F"].width = 7
    ws_gantt.column_dimensions["G"].width = 4
    
    ws_gantt.freeze_panes = ws_gantt.cell(5, date_col)
    
    # === SUMMARY ===
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "AGI TR 1-6 Transportation - Project Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")
    
    summary_data = [
        ("Key Parameters", ""),
        ("Total Tasks", f"{len(tasks)}"),
        ("Project Start", "=PROJECT_START"),
        ("Target End", "=TARGET_END"),
        ("Project Complete", "=MAX(Schedule_Data!G:G)"),
        ("", ""),
        ("Phase Breakdown", ""),
    ]
    
    # Phase counts
    phase_counts = {}
    for task in tasks:
        phase = task['Phase']
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
    
    for phase, count in sorted(phase_counts.items()):
        summary_data.append((phase, str(count)))
    
    summary_data.extend([
        ("", ""),
        ("Risk Breakdown", ""),
    ])
    
    risk_counts = {}
    for task in tasks:
        risk = task['Risk']
        risk_counts[risk] = risk_counts.get(risk, 0) + 1
    
    for risk, count in sorted(risk_counts.items()):
        summary_data.append((risk, str(count)))
    
    for r, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(r, 1, value=label)
        ws_summary.cell(r, 1).font = Font(bold=True) if label and not value else Font()
        ws_summary.cell(r, 2, value=value)
        if "=" in str(value):
            ws_summary.cell(r, 2).number_format = "YYYY-MM-DD"
    
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30
    
    return wb

if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
    
    # TSV 파일 경로
    script_dir = os.path.dirname(os.path.abspath(__file__))
    tsv_path = os.path.join(script_dir, "Untitled-2.tsv")
    
    if not os.path.exists(tsv_path):
        print(f"Error: TSV file not found: {tsv_path}")
        sys.exit(1)
    
    print(f"Reading TSV file: {tsv_path}")
    wb = create_gantt_from_tsv(tsv_path)
    
    output_path = os.path.join(script_dir, "AGI_TR_1-6_Master_Gantt_from_TSV.xlsx")
    wb.save(output_path)
    print(f"[OK] Generated: {output_path}")